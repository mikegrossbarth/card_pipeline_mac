from __future__ import annotations

import csv
import difflib
import calendar
import html
import queue
import base64
import hashlib
import json
import mimetypes
import os
import re
import secrets
import shutil
import socket
import subprocess
import sys
import threading
import time
import tkinter as tk
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
APP_DEBUG_LOG = ROOT / "work" / "lucas-debug.log"
ENGINE_DIR = ROOT / "comp_engine"
if str(ENGINE_DIR) not in sys.path:
    sys.path.insert(0, str(ENGINE_DIR))

from bridge_server import (  # noqa: E402
    COMP_STRATEGY_AVERAGE,
    COMP_STRATEGY_HIGH,
    COMP_STRATEGY_LOW,
    COMP_STRATEGY_STALE_NEWEST,
    EXPECTED_CARDLADDER_EXTENSION_VERSION,
    EXPECTED_CARDLADDER_MANIFEST_VERSION,
    BridgeServer,
    BridgeState,
    comp_price,
    format_comps,
    parse_formatted_comps,
    row_has_comp_data,
)
from workbook_io import WorkbookRow  # noqa: E402
import assignment_engine  # noqa: E402
from assignment_engine import AssignmentEngine  # noqa: E402
from assignment_engine import CONFIG_PATH as ASSIGNMENT_CONFIG_PATH  # noqa: E402
from assignment_engine import gsheet_shortcut_url, is_google_keep_url, keep_note_cache_path, load_gsheet_shortcut, normalize_source_value, path_from_source_value, safe_filename  # noqa: E402
from assignment_config_ui import open_assignment_rules_dialog, open_people_rules_dialog, seller_terms_health_lines  # noqa: E402
from google_sheets_import import export_google_sheet_to_xlsx  # noqa: E402
from lucas_diagnostics import diagnostic_json, lucas_version_label, setup_doctor_results  # noqa: E402
from shared_state import atomic_write_json, local_identity, shared_lock  # noqa: E402

from intake_io import (  # noqa: E402
    append_company_sheet_rows,
    build_card_title,
    clear_received_in_workbooks,
    clean_part,
    company_week_start_for_time as intake_company_week_start_for_time,
    default_output_path,
    ensure_company_weekly_sheets,
    format_money,
    infer_grader,
    mark_received_in_workbooks,
    normalize_grader,
    read_company_profit_records,
    read_photo_export,
    read_simple_spreadsheet,
    remove_company_sheet_rows_for_source,
    scan_to_cert,
    summarize_workbook,
    working_sheet_path,
    write_working_sheet,
    workbook_sheet_names,
    write_pipeline_output,
)


PHOTO_APP_ROOT = ROOT / "photo_tool"
PHOTO_APP_DIR = PHOTO_APP_ROOT / "app"
if PHOTO_APP_DIR.exists() and str(PHOTO_APP_DIR) not in sys.path:
    sys.path.insert(0, str(PHOTO_APP_DIR))
try:
    from dotenv import load_dotenv
except Exception:
    load_dotenv = None
if load_dotenv:
    load_dotenv(ROOT / ".env", override=False)
    load_dotenv(PHOTO_APP_DIR / ".env", override=False)
try:
    from google import genai
    from google.genai import types as genai_types
    from multi_card_extraction import (
        ModelQuotaExceeded,
        ModelResponseParseError,
        TemporaryModelUnavailable,
        _verify_cert_only_sync,
        identify_cards_sync,
    )
except Exception:
    genai = None
    genai_types = None
    identify_cards_sync = None
    _verify_cert_only_sync = None
    TemporaryModelUnavailable = ModelQuotaExceeded = ModelResponseParseError = Exception
SETTINGS_PATH = Path(os.environ.get("LUCAS_SETTINGS_PATH") or ROOT / "lucas_settings.json").expanduser()
DEFAULT_CARD_PIPELINE_DIR = ROOT / "CARD_PIPELINE"
CARD_PIPELINE_DIR = Path(os.environ.get("LUCAS_PIPELINE_DIR") or DEFAULT_CARD_PIPELINE_DIR)
WORKING_SHEETS_DIR = Path(os.environ.get("LUCAS_WORKING_SHEETS_DIR") or CARD_PIPELINE_DIR / "WORKING SHEETS")
INCOMING_SHEETS_DIR = CARD_PIPELINE_DIR / "INCOMING SHEETS"
RECEIVED_SHEETS_DIR = CARD_PIPELINE_DIR / "RECEIVED SHEETS"
ARCHIVED_SHEETS_DIR = CARD_PIPELINE_DIR / "ARCHIVED SHEETS"
COMPANY_SHEETS_DIR = CARD_PIPELINE_DIR / "COMPANY SHEETS"
SHEET_MARKERS_PATH = CARD_PIPELINE_DIR / "sheet_markers.json"
WEEKLY_COMPANY_SHEETS_PATH = CARD_PIPELINE_DIR / "weekly_company_sheets.json"
PROFIT_LEDGER_PATH = CARD_PIPELINE_DIR / "profit_ledger.json"
INVENTORY_LEDGER_PATH = CARD_PIPELINE_DIR / "inventory_ledger.json"
INVENTORY_DELETED_TOMBSTONES_PATH = CARD_PIPELINE_DIR / "inventory_deleted_tombstones.json"
INVENTORY_PHOTOS_DIR = CARD_PIPELINE_DIR / "INVENTORY PHOTOS"
INVENTORY_PHOTO_STATE_PATH = CARD_PIPELINE_DIR / "inventory_photo_state.json"
INSTAGRAM_INVENTORY_STATE_PATH = CARD_PIPELINE_DIR / "instagram_inventory_state.json"
DELETED_ARCHIVE_DIR = CARD_PIPELINE_DIR / "DELETED ARCHIVE"
DELETED_INVENTORY_PHOTOS_DIR = DELETED_ARCHIVE_DIR / "INVENTORY PHOTOS"
DELETED_SHEETS_DIR = DELETED_ARCHIVE_DIR / "SHEETS"
ACTIVITY_LOG_PATH = CARD_PIPELINE_DIR / "activity_log.json"
MOBILE_ACTION_LOG_PATH = CARD_PIPELINE_DIR / "mobile_action_log.json"
UNASSIGNED_PLAYERS_PATH = CARD_PIPELINE_DIR / "unassigned_players.json"
PLAYER_OVERRIDES_PATH = CARD_PIPELINE_DIR / "assignment_player_overrides.json"
SELLER_TERMS_PATH = CARD_PIPELINE_DIR / "ASSIGNMENT RULES" / "seller_terms.csv"
PERFORMANCE_LOG_PATH = CARD_PIPELINE_DIR / "lucas_performance.log"
DELETED_ARCHIVE_RETENTION_DAYS = 14
MAX_INVENTORY_PHOTOS_PER_CARD = 4
INVENTORY_PHOTO_GROUP_WINDOW_SECONDS = 75
try:
    PHOTO_OCR_REQUEST_TIMEOUT_MS = int(os.environ.get("LUCAS_PHOTO_OCR_TIMEOUT_MS") or "120000")
except ValueError:
    PHOTO_OCR_REQUEST_TIMEOUT_MS = 120000


def instagram_inventory_photo_order(paths: list[Path]) -> list[Path]:
    def rank(path: Path) -> tuple[int, int]:
        name = path.name.lower()
        if re.search(r"\]-\[(?:0*1|front)\]-", name) or re.search(r"\bfront\b", name):
            return (0, paths.index(path))
        if re.search(r"\]-\[(?:0*2|back)\]-", name) or re.search(r"\bback\b", name):
            return (1, paths.index(path))
        return (2, paths.index(path))

    return sorted(paths, key=rank)


def instagram_ready_photo_urls(item: dict[str, object]) -> list[str]:
    raw_urls = item.get("photo_urls") if isinstance(item.get("photo_urls"), list) else [item.get("photo_url")]
    urls: list[str] = []
    seen: set[str] = set()
    for raw_url in raw_urls:
        url = str(raw_url or "").strip()
        if not url or url in seen:
            continue
        seen.add(url)
        urls.append(url)
    return urls
LUCAS_LOGO_PATH = ROOT / "assets" / "lucas.png"
MIKEYS_CARDS_LOGO_PATH = ROOT / "assets" / "mikeys_cards_logo.png"
CARDLADDER_EXTENSION_DIR = ROOT / "cardladder-autocomp" / "extension"
APP_TITLE = "L.U.C.A.S"
APP_SUBTITLE = "Lot Upload, Comping & Assignment System"

COMP_STRATEGY_DISPLAY = {
    "Average last 5": COMP_STRATEGY_AVERAGE,
    "Highest of last 5": COMP_STRATEGY_HIGH,
    "Lowest of last 5": COMP_STRATEGY_LOW,
    "Date weighted": COMP_STRATEGY_STALE_NEWEST,
}
COMP_SCOPE_EMPTY = "Empty Comps Only"
COMP_SCOPE_ALL = "Recomp All"
COMP_SOURCE_BOTH = "Card Ladder + CY"
COMP_SOURCE_CARD_LADDER = "Card Ladder"
COMP_SOURCE_CY = "CY"
COMP_CY_ENABLED = False
COMP_SOURCE_OPTIONS = (COMP_SOURCE_BOTH, COMP_SOURCE_CARD_LADDER, COMP_SOURCE_CY) if COMP_CY_ENABLED else (COMP_SOURCE_CARD_LADDER,)
COMP_LOT_VALUE_SOURCE_OPTIONS = ("Comps Average", "Card Ladder Value", "CY Estimate")
NO_COMPANY_TAKES_LABEL = "NOBODY TAKES"
PROFIT_PERIOD_OPTIONS = ("5 Days", "Week", "Last 30 Days", "Calendar Month", "Year", "YTD", "Total")
PROFIT_GRAPH_OPTIONS = ("Overall Profit", "Profit to Sales Ratio", "Daily Trend", "Profit by Company")
PROFIT_PLOT_OPTIONS = ("Overall", "By Sport")
DEFAULT_PROFIT_PERIOD = "Calendar Month"
DEFAULT_PROFIT_GRAPH = "Overall Profit"
DEFAULT_PROFIT_PLOT = "Overall"
COMPANY_RESET_WEEKDAYS = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")
DEFAULT_COMPANY_RESET_WEEKDAY = "Monday"
DEFAULT_COMPANY_RESET_TIME = "20:00"
PROFIT_SPORT_COLORS = {
    "Football": "#3b82f6",
    "Baseball": "#ef4444",
    "Basketball": "#f4b400",
    "Golf": "#22c55e",
    "Hockey": "#a855f7",
    "Soccer": "#14b8a6",
    "TCG": "#f97316",
    "Other": "#94a3b8",
}
EXPENSE_CATEGORY_OPTIONS = ("Travel", "Supplies", "Travel Meal", "Fees", "Shipping")
EXPENSE_LINK_OPTIONS = ("General", "Card", "Sheet")
INVENTORY_GRADER_OPTIONS = ("PSA", "BGS", "CGC", "SGC")
ASSIGNMENT_CATEGORY_OPTIONS = (
    "basketball",
    "football",
    "baseball",
    "soccer",
    "hockey",
    "pokemon",
    "one piece",
    "wwe",
    "f1",
    "marvel",
    "disney",
    "star wars",
    "ufc",
)
ASSIGNMENT_CATEGORY_WEB_SIGNALS = {
    "basketball": ("basketball", "nba", "wnba", "ncaa basketball", "point guard", "shooting guard", "small forward", "power forward", "center"),
    "football": ("football", "nfl", "quarterback", "running back", "wide receiver", "linebacker", "cornerback", "defensive end"),
    "baseball": ("baseball", "mlb", "pitcher", "catcher", "shortstop", "outfielder", "first baseman", "second baseman", "third baseman"),
    "soccer": ("soccer", "footballer", "futbol", "fifa", "uefa", "premier league", "la liga", "serie a", "bundesliga"),
    "hockey": ("hockey", "nhl", "goaltender", "defenceman", "defenseman", "left wing", "right wing"),
    "pokemon": ("pokemon", "pokémon", "trading card game", "tcg", "pokédex", "pokedex"),
    "one piece": ("one piece", "straw hat", "manga", "anime", "pirate crew"),
    "wwe": ("wwe", "wwf", "professional wrestler", "pro wrestler", "wrestling"),
    "f1": ("formula 1", "formula one", "f1 driver", "grand prix", "racing driver", "motorsport"),
    "marvel": ("marvel", "marvel comics", "marvel cinematic", "superhero", "supervillain"),
    "disney": ("disney", "pixar", "disney character", "animated character"),
    "star wars": ("star wars", "jedi", "sith", "galactic", "lucasfilm"),
    "ufc": ("ufc", "mma", "mixed martial artist", "mixed martial arts", "ultimate fighting championship"),
}

PERF_LOG_LOCK = threading.Lock()


def load_app_settings() -> dict[str, object]:
    if not SETTINGS_PATH.exists():
        return {}
    try:
        raw = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return raw if isinstance(raw, dict) else {}


def save_app_settings(settings: dict[str, object]) -> None:
    atomic_write_json(SETTINGS_PATH, settings)


def ensure_mobile_pin(settings: dict[str, object]) -> str:
    pin = re.sub(r"\D", "", str(settings.get("mobile_pin") or ""))
    if len(pin) >= 4:
        return pin
    pin = f"{secrets.randbelow(1_000_000):06d}"
    settings["mobile_pin"] = pin
    save_app_settings(settings)
    return pin


def clean_mobile_host(value: object) -> str:
    host = str(value or "").strip().strip("/")
    if not host or "/" in host or ":" in host:
        return ""
    return host if re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9.-]{0,251}", host) else ""


def macos_local_mobile_host() -> str:
    try:
        result = subprocess.run(
            ["scutil", "--get", "LocalHostName"],
            capture_output=True,
            text=True,
            timeout=1,
            check=False,
        )
    except Exception:
        result = None
    local_name = clean_mobile_host(result.stdout if result else "")
    if local_name:
        return f"{local_name.removesuffix('.local')}.local"
    hostname = clean_mobile_host(socket.gethostname())
    if hostname and not re.fullmatch(r"\d+(?:\.\d+){3}", hostname):
        return hostname if hostname.endswith(".local") else f"{hostname}.local"
    return ""


def lan_mobile_host() -> str:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as handle:
            handle.connect(("8.8.8.8", 80))
            return handle.getsockname()[0]
    except OSError:
        try:
            return socket.gethostbyname(socket.gethostname())
        except OSError:
            return "127.0.0.1"


def mobile_app_host(settings: dict[str, object] | None = None) -> str:
    settings = settings or {}
    configured = clean_mobile_host(os.environ.get("LUCAS_MOBILE_HOST") or settings.get("mobile_host"))
    if configured:
        return configured
    return macos_local_mobile_host() or lan_mobile_host()


def mobile_public_app_url(profile: str, settings: dict[str, object] | None = None) -> str:
    settings = settings or {}
    profile_key = "LUCAS_PERSONAL_MOBILE_PUBLIC_URL" if profile == "personal" else "LUCAS_TEAM_MOBILE_PUBLIC_URL"
    raw = str(
        os.environ.get(profile_key)
        or settings.get("mobile_public_url")
        or os.environ.get("LUCAS_MOBILE_PUBLIC_URL")
        or ""
    ).strip().rstrip("/")
    if not raw:
        return ""
    parsed = urllib.parse.urlparse(raw)
    if parsed.scheme != "https" or not parsed.netloc:
        return ""
    if re.search(r"/mobile/(?:team|personal)(?:/|$)", parsed.path):
        return raw
    if parsed.path.rstrip("/").endswith("/mobile"):
        return f"{raw}/{profile}"
    return f"{raw}/mobile/{profile}"


def is_personal_lucas_profile(settings: dict[str, object] | None = None, settings_path: Path | None = None) -> bool:
    settings = settings or {}
    path = settings_path or SETTINGS_PATH
    path_text = str(path).lower()
    pipeline_text = str(settings.get("pipeline_root") or CARD_PIPELINE_DIR).lower()
    folder_text = str(settings.get("working_sheets_dir") or "").lower()
    return any(
        marker in text
        for text in (path_text, pipeline_text, folder_text)
        for marker in ("lucas_personal", "personal lucas", "lucas_settings.michael")
    )


def mobile_bridge_port(settings: dict[str, object] | None = None, settings_path: Path | None = None) -> int:
    raw_port = str(os.environ.get("LUCAS_MOBILE_PORT") or (settings or {}).get("mobile_port") or "").strip()
    if raw_port:
        try:
            port = int(raw_port)
        except ValueError:
            port = 0
        if 1024 <= port <= 65535:
            return port
    return 8766 if is_personal_lucas_profile(settings, settings_path) else 8765


def make_photo_ocr_client(api_key: str):
    if genai is None:
        return None
    if genai_types is None or not hasattr(genai_types, "HttpOptions"):
        return genai.Client(api_key=api_key)
    try:
        return genai.Client(
            api_key=api_key,
            http_options=genai_types.HttpOptions(timeout=PHOTO_OCR_REQUEST_TIMEOUT_MS),
        )
    except TypeError:
        return genai.Client(api_key=api_key)


def app_debug_log(message: str) -> None:
    try:
        APP_DEBUG_LOG.parent.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with APP_DEBUG_LOG.open("a", encoding="utf-8") as handle:
            handle.write(f"[{stamp}] {message}\n")
    except Exception:
        pass


def performance_threshold_seconds() -> float:
    raw = os.environ.get("LUCAS_PERF_LOG_SECONDS", "0.25")
    try:
        return max(float(raw), 0.0)
    except (TypeError, ValueError):
        return 0.25


def record_performance_event(operation: str, started_at: float, details: str = "", force: bool = False) -> None:
    elapsed = time.perf_counter() - started_at
    if not force and elapsed < performance_threshold_seconds():
        return
    try:
        line = f"{datetime.now().isoformat(timespec='seconds')}\t{operation}\t{elapsed:.3f}s"
        if details:
            line = f"{line}\t{details}"
        with PERF_LOG_LOCK:
            PERFORMANCE_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
            with PERFORMANCE_LOG_PATH.open("a", encoding="utf-8") as handle:
                handle.write(line + "\n")
    except Exception:
        return


def is_google_sheet_url(value: object) -> bool:
    parsed = urllib.parse.urlparse(str(value or "").strip())
    return parsed.scheme in {"http", "https"} and parsed.netloc.lower().endswith("docs.google.com") and "/spreadsheets/" in parsed.path


def set_pipeline_root(path: Path, working_sheets_dir: Path | None = None) -> None:
    global CARD_PIPELINE_DIR, WORKING_SHEETS_DIR, INCOMING_SHEETS_DIR, RECEIVED_SHEETS_DIR, ARCHIVED_SHEETS_DIR, COMPANY_SHEETS_DIR, SHEET_MARKERS_PATH, WEEKLY_COMPANY_SHEETS_PATH, PROFIT_LEDGER_PATH, INVENTORY_LEDGER_PATH, INVENTORY_DELETED_TOMBSTONES_PATH, INVENTORY_PHOTOS_DIR, INVENTORY_PHOTO_STATE_PATH, INSTAGRAM_INVENTORY_STATE_PATH, DELETED_ARCHIVE_DIR, DELETED_INVENTORY_PHOTOS_DIR, DELETED_SHEETS_DIR, ACTIVITY_LOG_PATH, MOBILE_ACTION_LOG_PATH, UNASSIGNED_PLAYERS_PATH, PLAYER_OVERRIDES_PATH, SELLER_TERMS_PATH, PERFORMANCE_LOG_PATH
    CARD_PIPELINE_DIR = Path(path).expanduser()
    WORKING_SHEETS_DIR = Path(working_sheets_dir).expanduser() if working_sheets_dir else CARD_PIPELINE_DIR / "WORKING SHEETS"
    INCOMING_SHEETS_DIR = CARD_PIPELINE_DIR / "INCOMING SHEETS"
    RECEIVED_SHEETS_DIR = CARD_PIPELINE_DIR / "RECEIVED SHEETS"
    ARCHIVED_SHEETS_DIR = CARD_PIPELINE_DIR / "ARCHIVED SHEETS"
    COMPANY_SHEETS_DIR = CARD_PIPELINE_DIR / "COMPANY SHEETS"
    SHEET_MARKERS_PATH = CARD_PIPELINE_DIR / "sheet_markers.json"
    WEEKLY_COMPANY_SHEETS_PATH = CARD_PIPELINE_DIR / "weekly_company_sheets.json"
    PROFIT_LEDGER_PATH = CARD_PIPELINE_DIR / "profit_ledger.json"
    INVENTORY_LEDGER_PATH = CARD_PIPELINE_DIR / "inventory_ledger.json"
    INVENTORY_DELETED_TOMBSTONES_PATH = CARD_PIPELINE_DIR / "inventory_deleted_tombstones.json"
    INVENTORY_PHOTOS_DIR = CARD_PIPELINE_DIR / "INVENTORY PHOTOS"
    INVENTORY_PHOTO_STATE_PATH = CARD_PIPELINE_DIR / "inventory_photo_state.json"
    INSTAGRAM_INVENTORY_STATE_PATH = CARD_PIPELINE_DIR / "instagram_inventory_state.json"
    DELETED_ARCHIVE_DIR = CARD_PIPELINE_DIR / "DELETED ARCHIVE"
    DELETED_INVENTORY_PHOTOS_DIR = DELETED_ARCHIVE_DIR / "INVENTORY PHOTOS"
    DELETED_SHEETS_DIR = DELETED_ARCHIVE_DIR / "SHEETS"
    ACTIVITY_LOG_PATH = CARD_PIPELINE_DIR / "activity_log.json"
    MOBILE_ACTION_LOG_PATH = CARD_PIPELINE_DIR / "mobile_action_log.json"
    UNASSIGNED_PLAYERS_PATH = CARD_PIPELINE_DIR / "unassigned_players.json"
    PLAYER_OVERRIDES_PATH = CARD_PIPELINE_DIR / "assignment_player_overrides.json"
    SELLER_TERMS_PATH = CARD_PIPELINE_DIR / "ASSIGNMENT RULES" / "seller_terms.csv"
    PERFORMANCE_LOG_PATH = CARD_PIPELINE_DIR / "lucas_performance.log"


def set_pipeline_from_working_dir(path: Path) -> None:
    working_dir = normalize_working_dir_selection(Path(path).expanduser())
    set_pipeline_root(working_dir.parent, working_dir)


def normalize_working_dir_selection(path: Path) -> Path:
    child = path / "WORKING SHEETS"
    if path.name.upper() != "WORKING SHEETS" and child.exists() and child.is_dir():
        return child
    return path


def initialize_pipeline_root() -> None:
    settings = load_app_settings()
    configured_working = str(settings.get("working_sheets_dir") or os.environ.get("LUCAS_WORKING_SHEETS_DIR") or "").strip()
    if configured_working:
        set_pipeline_from_working_dir(Path(configured_working))
        return
    configured = str(settings.get("pipeline_root") or "").strip()
    if configured:
        set_pipeline_root(Path(configured))


initialize_pipeline_root()


def company_sheet_week_start_for_time(moment: datetime) -> datetime.date:
    return intake_company_week_start_for_time(moment)


def parse_company_reset_time(value: object) -> datetime.time:
    text = str(value or "").strip()
    if not text:
        text = DEFAULT_COMPANY_RESET_TIME
    for pattern in ("%H:%M", "%H%M", "%I:%M %p", "%I %p", "%I:%M%p", "%I%p"):
        try:
            return datetime.strptime(text.upper(), pattern).time().replace(second=0, microsecond=0)
        except ValueError:
            continue
    raise ValueError("Use a time like 20:00 or 8:00 PM.")


def company_sheet_week_start_for_schedule(moment: datetime, weekday: object = DEFAULT_COMPANY_RESET_WEEKDAY, reset_time: object = DEFAULT_COMPANY_RESET_TIME) -> datetime.date:
    weekday_text = str(weekday or DEFAULT_COMPANY_RESET_WEEKDAY).strip().title()
    if weekday_text not in COMPANY_RESET_WEEKDAYS:
        weekday_text = DEFAULT_COMPANY_RESET_WEEKDAY
    reset_weekday = COMPANY_RESET_WEEKDAYS.index(weekday_text)
    reset_clock = parse_company_reset_time(reset_time)
    days_since_reset_day = (moment.weekday() - reset_weekday) % 7
    current_reset_date = moment.date() - timedelta(days=days_since_reset_day)
    current_reset = datetime.combine(current_reset_date, reset_clock)
    if moment < current_reset:
        current_reset_date -= timedelta(days=7)
    return current_reset_date

DISPLAY_COLUMNS = (
    "excel_row",
    "source",
    "sheet_source",
    "cert_number",
    "grader",
    "category",
    "card_title",
    "purchase_price",
    "card_ladder_value",
    "card_ladder_comps_average",
    "cy_value",
    "cy_confidence",
    "best_company",
    "estimated_payout",
    "status",
)

INTAKE_COLUMNS = (
    "excel_row",
    "source",
    "cert_number",
    "grader",
    "category",
    "card_title",
    "purchase_price",
    "card_ladder_value",
    "card_ladder_comps_average",
    "cy_value",
    "cy_confidence",
    "status",
)

COMP_COLUMNS = (
    "excel_row",
    "source",
    "cert_number",
    "grader",
    "category",
    "card_title",
    "purchase_price",
    "card_ladder_value",
    "card_ladder_comps_average",
    "cy_value",
    "cy_confidence",
    "best_company",
    "estimated_payout",
    "received",
    "status",
    "sheet_source",
)

RECEIVE_COLUMNS = (
    "excel_row",
    "person",
    "item_id",
    "sheet_source",
    "cert_number",
    "grader",
    "category",
    "card_title",
    "purchase_price",
    "card_ladder_value",
    "card_ladder_comps_average",
    "cy_value",
    "cy_confidence",
    "best_company",
    "estimated_payout",
    "status",
    "company_pile",
    "source",
)

REVIEW_COLUMNS = DISPLAY_COLUMNS

INVENTORY_TABLE_COLUMNS = (
    "date",
    "type",
    "item_id",
    "person",
    "sport",
    "cert",
    "grader",
    "card",
    "purchase",
    "card_ladder",
    "comps",
    "cy_estimate",
    "cy_confidence",
    "company",
    "payout",
    "paid_with",
    "source",
    "status",
    "photos",
    "notes",
)

ADD_INTAKE_ROW_IID = "__add_intake_row__"
ADD_REVIEW_ROW_IID = "__add_review_row__"
ADD_COMP_ROW_IID = "__add_comp_row__"

EDITABLE_COLUMNS = {
    "source",
    "item_id",
    "cert_number",
    "grader",
    "category",
    "card_title",
    "purchase_price",
    "card_ladder_value",
    "card_ladder_comps_average",
    "cy_value",
    "cy_confidence",
}

HEADINGS = {
    "excel_row": "Row",
    "source": "Source",
    "person": "Person",
    "item_id": "Item ID",
    "sheet_source": "Sheet Source",
    "cert_number": "Cert #",
    "grader": "Company",
    "category": "Sport",
    "card_title": "Card",
    "purchase_price": "Purchase",
    "card_ladder_value": "Card Ladder",
    "card_ladder_comps_average": "Comps",
    "cy_value": "CY Estimate",
    "cy_confidence": "CY Confidence",
    "best_company": "Best Company",
    "estimated_payout": "Est. Payout",
    "received": "Received",
    "status": "Status",
    "company_pile": "Company Pile",
}

COLUMN_WIDTHS = {
    "excel_row": 52,
    "source": 130,
    "person": 145,
    "item_id": 150,
    "sheet_source": 150,
    "cert_number": 110,
    "grader": 86,
    "category": 95,
    "card_title": 390,
    "purchase_price": 90,
    "card_ladder_value": 100,
    "card_ladder_comps_average": 100,
    "cy_value": 100,
    "cy_confidence": 110,
    "best_company": 130,
    "estimated_payout": 100,
    "received": 90,
    "status": 160,
    "company_pile": 105,
}

INVENTORY_HEADINGS = {
    "date": "Date",
    "type": "Type",
    "item_id": "Item ID",
    "person": "Person",
    "sport": "Sport",
    "cert": "Cert",
    "grader": "Grader",
    "card": "Card",
    "purchase": "Purchase",
    "paid_with": "Paid With",
    "card_ladder": "Card Ladder",
    "comps": "Comps",
    "cy_estimate": "CY Estimate",
    "cy_confidence": "CY Confidence",
    "company": "Best Company",
    "payout": "Est. Payout",
    "source": "Source Sheet",
    "status": "Status",
    "photos": "Photos",
    "notes": "Notes",
}

INVENTORY_COLUMN_WIDTHS = {
    "date": 95,
    "type": 75,
    "item_id": 150,
    "person": 130,
    "sport": 95,
    "cert": 110,
    "grader": 80,
    "card": 320,
    "purchase": 100,
    "paid_with": 115,
    "card_ladder": 100,
    "comps": 100,
    "cy_estimate": 100,
    "cy_confidence": 110,
    "company": 140,
    "payout": 100,
    "source": 170,
    "status": 110,
    "photos": 80,
    "notes": 240,
}

INVENTORY_EDIT_COLUMN_FIELDS = {
    "date": "date_added",
    "person": "assigned_person",
    "sport": "sport",
    "cert": "cert_number",
    "grader": "grader",
    "card": "card_title",
    "purchase": "purchase_price",
    "paid_with": "paid_with",
    "card_ladder": "card_ladder_value",
    "comps": "card_ladder_comps_average",
    "cy_estimate": "cy_value",
    "cy_confidence": "cy_confidence",
    "company": "best_company",
    "payout": "estimated_payout",
    "source": "source_sheet",
    "notes": "notes",
}
INVENTORY_EDIT_MONEY_COLUMNS = {"purchase", "card_ladder", "comps", "cy_estimate", "payout"}

AUTO_INVENTORY_NOTES = {
    "backfilled from received sheets",
    "received without company pile",
    "moved from inventory",
}

BUTTON_TOOLTIPS = {
    "lucas settings": "Open LUCAS tools like Activity Log, System Health, Working Folder, and Mobile Help.",
    "instagram inventory sync": "Preview or run your personal Instagram inventory page sync.",
    "delete selected": "Remove the selected row or rows from this table. Right-click table rows for this action where available.",
    "clear rows": "Clear the current Create table rows from the screen.",
    "save as working sheet": "Save the Create rows as a new Working Sheet.",
    "load selected sheet": "Load the selected incoming or working sheet into Comp.",
    "refresh sheet list": "Reload the list of incoming and working sheets.",
    "save back to source sheet": "Write the current Comp table values back to the loaded source sheet.",
    "run all comps": "Run selected comp sources for the loaded rows.",
    "stop run": "Stop the active comp run after the current work finishes.",
    "clear comp rows": "Clear the Comp table from the screen without changing source sheets.",
    "add row": "Add a blank editable row to the loaded Comp sheet. Save back to persist it.",
    "mark received in sheets": "Mark scanned or loaded receive rows as received in their matching sheets.",
    "refresh incoming index": "Rebuild the lookup index from Incoming and Working sheets.",
    "load": "Load the selected sheet.",
    "refresh received sheets": "Reload the list of sheets in Received.",
    "company rules": "Open company assignment and payout rules.",
    "people rules": "Open person-specific rules, seller terms, and payout preferences.",
    "unassigned players": "Review players LUCAS could not categorize and teach their sport/category.",
    "copy details": "Copy the diagnostic details to the clipboard.",
    "close": "Close this window.",
    "incoming": "Show Incoming sheets on Home.",
    "working": "Show Working sheets on Home.",
    "received": "Show Received sheets on Home.",
    "edit markers": "Edit the selected Home sheet markers.",
    "refresh home view": "Reload Home lists, summaries, and sheet statuses.",
    "add card": "Add a card directly to inventory.",
    "export": "Export the current visible inventory rows.",
    "filters": "Open inventory filters for sport, grader, price, and date.",
    "settings": "Open inventory maintenance actions.",
    "bulk edit": "Turn multi-cell inventory editing on or off.",
    "clear filters": "Reset the inventory filter fields.",
    "apply filters": "Refresh inventory using the selected filters.",
    "refresh view": "Reload the current Profit view.",
    "recover sold ledger": "Repair missing sold-card profit rows by scanning company sheets.",
    "add expense": "Add an expense record to Profit.",
    "sold cards": "Show individual sold-card profit rows.",
    "sold sheets": "Group Profit by sold sheet.",
    "expenses": "Show expense rows.",
    "refresh": "Reload this view.",
    "cancel": "Close without saving changes.",
    "mark sold": "Move the inventory card to Profit as sold.",
    "save": "Save these changes.",
    "move": "Move the selected item.",
    "start recomp": "Start recomping the selected inventory value fields.",
    "delete": "Delete the selected item after confirmation.",
    "enter receive scanning mode": "Arm Receive so barcode scans add received rows.",
    "exit receive scanning mode": "Turn off Receive barcode scanning mode.",
    "add receive photos": "Choose receive photos to scan.",
    "scan receive photos": "Read selected receive photos and add detected cards.",
    "clear receive photos": "Remove selected receive photos from the scan queue.",
    "enter scanning station mode": "Arm Create so barcode scans add new rows.",
    "exit scanning station mode": "Turn off Create barcode scanning mode.",
    "browse": "Choose a spreadsheet file to load.",
    "load rows": "Load rows from the selected file.",
    "add photos": "Choose card photos to scan.",
    "add folder": "Add all card photos from a folder.",
    "scan photos": "Read selected photos and add detected cards.",
    "clear photos": "Remove selected photos from the scan queue.",
    "web search": "Search the selected player/card in a browser.",
    "save category": "Save this player/category so LUCAS can assign it next time.",
    "auto categorize all": "Try to categorize all unassigned players automatically.",
    "remove": "Remove the selected entry from this list.",
}


def inventory_display_notes(record: dict[str, object]) -> str:
    notes = str(record.get("notes") or "").strip()
    return "" if notes.lower() in AUTO_INVENTORY_NOTES else notes


def inventory_grader_filter_values(raw: object) -> set[str]:
    values: set[str] = set()
    for part in re.split(r"[,;/|]", str(raw or "")):
        text = part.strip().upper()
        if text:
            values.add(text)
    return values


class CardPipelineApp(tk.Tk):
    def __init__(self) -> None:
        perf_start = time.perf_counter()
        super().__init__()
        app_debug_log("app_start")
        self.title(f"{APP_TITLE} - {APP_SUBTITLE}")
        self.geometry("1420x820")
        self.minsize(760, 520)
        self.logo_image: tk.PhotoImage | None = None
        self._tab_scroll_canvases: dict[str, tk.Canvas] = {}
        self._tab_scroll_hosts: dict[str, tk.Widget] = {}
        self._tab_scroll_contents: list[tk.Widget] = []
        self._tab_scroll_bound_widgets: set[str] = set()
        self.tooltip_window: tk.Toplevel | None = None
        self.tooltip_after_id: str | None = None
        self.tooltip_widget: tk.Widget | None = None
        self._button_tooltip_classes_bound = False
        self.instagram_tunnel_process: subprocess.Popen | None = None
        self.instagram_tunnel_public_url = ""
        self.instagram_tunnel_log_path = ROOT / "work" / "instagram-cloudflared.log"
        self.instagram_auto_sync_running = False

        self.events: queue.Queue[str] = queue.Queue()
        self.intake_rows: list[WorkbookRow] = []
        self.intake_sources: dict[int, str] = {}
        self.intake_sheet_sources: dict[int, str] = {}
        self.row_sources: dict[int, str] = {}
        self.comp_sheet_sources: dict[int, str] = {}
        self.loaded_comp_sheet_label = ""
        self.review_rows: list[WorkbookRow] = []
        self.review_sources: dict[int, str] = {}
        self.review_sheet_sources: dict[int, str] = {}
        self.incoming_cert_index: dict[str, dict[str, object]] = {}
        self.comp_output_saved = True
        self.lucas_identity = local_identity(SETTINGS_PATH)
        self.app_settings = load_app_settings()
        self.mobile_pin = ensure_mobile_pin(self.app_settings)
        self.state = BridgeState()
        self.state.on_update = lambda: self.events.put("comp_refresh")
        self.state.mobile_pin_provider = lambda: self.mobile_pin
        self.state.mobile_inventory_search = self.mobile_inventory_search
        self.state.mobile_inventory_add = self.mobile_inventory_add
        self.state.mobile_inventory_mark_sold = self.mobile_inventory_mark_sold
        self.state.mobile_inventory_trade = self.mobile_inventory_trade
        self.state.mobile_card_identify = self.mobile_card_identify
        self.state.mobile_profit_summary = self.mobile_profit_summary
        self.state.mobile_expense_add = self.mobile_expense_add
        self.state.mobile_payouts = self.mobile_payouts
        self.state.mobile_queue_sync = self.mobile_queue_sync
        self.state.mobile_inventory_photo_resolver = self.mobile_inventory_photo_response
        self.state.instagram_media_resolver = self.instagram_inventory_media_response
        self.bridge = BridgeServer(self.state, port=mobile_bridge_port(self.app_settings, SETTINGS_PATH))
        self.bridge.start()
        self._refresh_keep_source_registry()
        app_debug_log(f"bridge_started started={self.bridge.started} port={self.bridge.port} error={self.bridge.error}")
        mobile_url = self._mobile_app_url()
        self.bridge_status_text = (
            f"Card Ladder bridge running at http://127.0.0.1:{self.bridge.port} | Mobile: {mobile_url} PIN {self.mobile_pin}"
            if self.bridge.started
            else f"Card Ladder bridge failed to start: {self.bridge.error}"
        )
        if self.bridge.started and self._personal_instagram_sync_enabled():
            self._ensure_instagram_background_tunnel()

        self.input_mode = tk.StringVar(value="Barcode Scanner")
        self.review_mode = tk.StringVar(value="Automatic Receive")
        self.review_input_mode = tk.StringVar(value="Barcode Scanner")
        self.comp_strategy_label = tk.StringVar(value="Average last 5")
        self.comp_low_outlier_pct_var = tk.StringVar(value="Off")
        self.comp_scope_label = tk.StringVar(value=COMP_SCOPE_EMPTY)
        self.comp_source_label = tk.StringVar(value=COMP_SOURCE_CARD_LADDER)
        self.working_sheet_title = tk.StringVar()
        self.create_network_mode_var = tk.BooleanVar(value=bool(self.app_settings.get("network_mode")))
        self.seller_terms_seller_var = tk.StringVar()
        self.seller_terms_sheet_type_var = tk.StringVar()
        self.selected_working_sheet = tk.StringVar()
        self.summary_var = tk.StringVar(value="Choose a create mode to begin.")
        self.status_var = tk.StringVar(value="Card Ladder bridge starting...")
        self.bridge_status_var = tk.StringVar(value=self.bridge_status_text)
        self.version_var = tk.StringVar(value=lucas_version_label("Mac"))
        self.pipeline_root_var = tk.StringVar(value=str(CARD_PIPELINE_DIR))

        self.scan_cert = tk.StringVar()
        self.scan_grader = tk.StringVar(value="PSA")
        self.scan_card = tk.StringVar()
        self.scan_status = tk.StringVar(value="Scanning station is off.")
        self.scan_entry: ttk.Entry | None = None
        self.cell_editor: ttk.Entry | ttk.Combobox | None = None
        self.cell_edit: tuple[ttk.Treeview, str, str] | None = None
        self.receive_cell_autocomplete_matches: dict[str, dict[str, object]] = {}
        self.column_widths_by_tree: dict[int, dict[str, int]] = {}
        self.scanning_station_active = False

        self.file_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.photo_paths: list[Path] = []
        self.photo_status = tk.StringVar(value="No photos selected.")
        self.photo_worker: threading.Thread | None = None
        self.photo_client = None
        self.review_scan_cert = tk.StringVar()
        self.review_scan_entry: ttk.Entry | None = None
        self.review_scanning_active = False
        self.review_status = tk.StringVar(value="Receive station is off.")
        self.assignment_progress_value = tk.DoubleVar(value=0)
        self.review_photo_paths: list[Path] = []
        self.review_photo_status = tk.StringVar(value="No receive photos selected.")
        self.review_photo_worker: threading.Thread | None = None
        self._load_player_overrides()
        assignment_start = time.perf_counter()
        self.assignment_engine = AssignmentEngine.load()
        record_performance_event("assignment.load.initial", assignment_start, f"companies={len(self.assignment_engine.companies)}")
        self.assignment_recommendation_job = 0
        self.assignment_recommendation_running = False
        self.assignment_recommendation_after_id: str | None = None
        self.assignment_recommendation_row_ids: set[int] | None = None
        self.pending_comp_assignment_row_ids: set[int] = set()
        self.assignment_config_status = tk.StringVar(value=self._assignment_config_status())
        self._ensure_company_sheet_folders()
        self._ensure_weekly_company_sheets_due()
        self.received_sheet_paths: dict[str, Path] = {}
        self.selected_received_sheet = tk.StringVar()
        self.incoming_sheet_paths: dict[str, Path] = {}
        self.working_sheet_paths: dict[str, Path] = {}
        self.comp_sheet_paths: dict[str, Path] = {}
        self.comp_sheet_stages: dict[str, str] = {}
        self.home_sheet_kind = tk.StringVar(value="Incoming")
        self.home_sheet_paths: dict[str, dict[str, Path]] = {"Incoming": {}, "Working": {}, "Received": {}}
        self.home_sheet_summaries: dict[str, dict[str, object]] = {}
        self.home_summary_cache: dict[str, dict[str, object]] = {}
        self.home_summary_cache_lock = threading.Lock()
        self.home_sheet_markers: dict[str, dict[str, object]] = self._load_sheet_markers()
        self.deleted_sheet_marker_keys: set[str] = set()
        self.home_selected_sheet_key = ""
        self.home_person_var = tk.StringVar()
        self.home_sheet_sort_var = tk.StringVar(value="Date Created")
        self.home_incoming_volume_sort_column = "sheet"
        self.home_incoming_volume_sort_descending = False
        self.payout_person_var = tk.StringVar()
        self.payout_status_var = tk.StringVar(value="No unpaid sheets loaded.")
        self.payout_summary_people: dict[str, str] = {}
        self.payout_detail_keys: dict[str, str] = {}
        self.inventory_status_var = tk.StringVar(value="No inventory loaded.")
        self.inventory_metric_var = tk.StringVar(value="")
        self.inventory_person_var = tk.StringVar()
        self.inventory_sport_var = tk.StringVar()
        self.inventory_grader_var = tk.StringVar()
        self.inventory_year_var = tk.StringVar()
        self.inventory_search_var = tk.StringVar()
        self.inventory_min_var = tk.StringVar()
        self.inventory_max_var = tk.StringVar()
        self.inventory_date_min_var = tk.StringVar()
        self.inventory_date_max_var = tk.StringVar()
        self.inventory_missing_title_var = tk.BooleanVar(value=False)
        self.inventory_missing_comps_var = tk.BooleanVar(value=False)
        self.inventory_missing_cl_var = tk.BooleanVar(value=False)
        self.inventory_missing_photos_var = tk.BooleanVar(value=False)
        self.inventory_bulk_edit_var = tk.BooleanVar(value=False)
        self.inventory_rows: list[dict[str, object]] = []
        self.filtered_inventory_rows: list[dict[str, object]] = []
        self.inventory_tree_records: dict[str, dict[str, object]] = {}
        self.inventory_recomp_context: dict[str, object] | None = None
        self.inventory_bulk_cell: tuple[str, str] | None = None
        self.inventory_cell_editor: ttk.Entry | None = None
        self.inventory_cell_edit: tuple[str, str] | None = None
        self.inventory_bulk_undo_stack: list[dict[str, object]] = []
        self.inventory_filter_after_id: str | None = None
        self.inventory_sort_column = "date"
        self.inventory_sort_descending = True
        self.inventory_photo_worker: threading.Thread | None = None
        self.inventory_photo_client = None
        self.inventory_photo_scan_after_id: str | None = None
        self.profit_status_var = tk.StringVar(value="No profit ledger loaded.")
        self.profit_metric_var = tk.StringVar(value="")
        self.profit_person_var = tk.StringVar()
        self.profit_period_var = tk.StringVar(value=DEFAULT_PROFIT_PERIOD)
        self.profit_graph_var = tk.StringVar(value=DEFAULT_PROFIT_GRAPH)
        self.profit_plot_var = tk.StringVar(value=DEFAULT_PROFIT_PLOT)
        self.profit_search_var = tk.StringVar()
        self.profit_chart_title_var = tk.StringVar(value=self._profit_chart_title())
        self.profit_view_mode = tk.StringVar(value="Sold Cards")
        self.profit_rows: list[dict[str, object]] = []
        self.filtered_profit_rows: list[dict[str, object]] = []
        self.profit_tree_records: dict[str, dict[str, object]] = {}
        self.profit_sport_cache: dict[str, str] = {}
        self.profit_sort_column = "date"
        self.profit_sort_descending = True

        self._install_button_tooltips()
        self._build_ui()
        self._show_mode()
        self.after_idle(self._refresh_tab_scroll_bindings)
        self.refresh_profit_tab()
        self._poll_events()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.status_var.set(self.bridge_status_text)
        self.after(100, self._start_startup_refresh)
        self.after(5 * 60 * 1000, self._weekly_company_sheet_timer)
        if self._personal_instagram_sync_enabled():
            self.after(90 * 1000, self._instagram_auto_sync_timer)
        record_performance_event("app.init", perf_start, f"pipeline={CARD_PIPELINE_DIR}", force=True)

    def _on_close(self) -> None:
        self._hide_tooltip()
        self.destroy()

    def _install_button_tooltips(self) -> None:
        if getattr(self, "_button_tooltip_classes_bound", False):
            return
        self._button_tooltip_classes_bound = True
        for class_name in ("TButton", "Button", "TCheckbutton", "Checkbutton"):
            self.bind_class(class_name, "<Enter>", self._schedule_tooltip_for_event, add="+")
            self.bind_class(class_name, "<Leave>", lambda _event: self._hide_tooltip(), add="+")
            self.bind_class(class_name, "<ButtonPress>", lambda _event: self._hide_tooltip(), add="+")

    def _set_tooltip(self, widget: tk.Widget, text: str | None = None) -> None:
        if text:
            setattr(widget, "_lucas_tooltip", text)
        widget.bind("<Enter>", self._schedule_tooltip_for_event, add="+")
        widget.bind("<Leave>", lambda _event: self._hide_tooltip(), add="+")
        widget.bind("<ButtonPress>", lambda _event: self._hide_tooltip(), add="+")

    def _schedule_tooltip_for_event(self, event: tk.Event) -> None:
        widget = event.widget
        text = self._button_tooltip_text(widget)
        if not text:
            return
        self._hide_tooltip()
        self.tooltip_widget = widget
        self.tooltip_after_id = self.after(550, lambda target=widget, message=text: self._show_tooltip(target, message))

    def _button_tooltip_text(self, widget: tk.Widget) -> str:
        direct = str(getattr(widget, "_lucas_tooltip", "") or "").strip()
        if direct:
            return direct
        try:
            label = str(widget.cget("text") or "").strip()
        except Exception:
            label = ""
        if not label:
            return ""
        key = re.sub(r"[^a-z0-9]+", " ", label.lower()).strip()
        mapped = BUTTON_TOOLTIPS.get(key)
        if mapped:
            return mapped
        return f"Click to {label[0].lower() + label[1:] if len(label) > 1 else label.lower()}."

    def _show_tooltip(self, widget: tk.Widget, text: str) -> None:
        if self.tooltip_widget is not widget or not str(text or "").strip():
            return
        try:
            if not widget.winfo_viewable():
                return
            x = widget.winfo_rootx() + min(widget.winfo_width(), 26)
            y = widget.winfo_rooty() + widget.winfo_height() + 8
        except tk.TclError:
            return
        self._hide_tooltip(cancel_after=False)
        tooltip = tk.Toplevel(self)
        tooltip.wm_overrideredirect(True)
        tooltip.configure(bg="#0f0f0f")
        tooltip.attributes("-topmost", True)
        label = tk.Label(
            tooltip,
            text=text,
            bg="#0f0f0f",
            fg="#f5f5f5",
            justify=tk.LEFT,
            wraplength=320,
            padx=10,
            pady=7,
            borderwidth=1,
            relief=tk.SOLID,
            font=("Segoe UI", 9),
        )
        label.pack()
        tooltip.wm_geometry(f"+{x}+{y}")
        self.tooltip_window = tooltip

    def _hide_tooltip(self, cancel_after: bool = True) -> None:
        if cancel_after and self.tooltip_after_id:
            try:
                self.after_cancel(self.tooltip_after_id)
            except tk.TclError:
                pass
        self.tooltip_after_id = None
        if self.tooltip_window is not None:
            try:
                self.tooltip_window.destroy()
            except tk.TclError:
                pass
        self.tooltip_window = None
        self.tooltip_widget = None

    def _scroll_canvas_pixels(self, canvas: tk.Canvas, orient: str, pixels: int) -> bool:
        if pixels == 0:
            return False
        bbox = canvas.bbox("all")
        if not bbox:
            return False
        if orient == "x":
            content_size = bbox[2] - bbox[0]
            viewport_size = canvas.winfo_width()
            first, last = canvas.xview()
            view = canvas.xview_moveto
        else:
            content_size = bbox[3] - bbox[1]
            viewport_size = canvas.winfo_height()
            first, last = canvas.yview()
            view = canvas.yview_moveto
        scrollable = content_size - viewport_size
        if scrollable <= 0 or (pixels < 0 and first <= 0) or (pixels > 0 and last >= 1):
            return False
        view(max(0.0, min(1.0, first + (pixels / scrollable))))
        return True

    def _wheel_pixels(self, event: tk.Event) -> int:
        delta = getattr(event, "delta", 0)
        if not delta:
            return 0
        if sys.platform == "darwin":
            scale = 4 if abs(delta) < 10 else 1
            return int(-delta * scale)
        return int(-delta / 120 * 72)

    def _current_tab_scroll_canvas(self) -> tk.Canvas | None:
        if not hasattr(self, "tabs"):
            return None
        try:
            return self._tab_scroll_canvases.get(str(self.tabs.select()))
        except tk.TclError:
            return None

    def _current_tab_scroll_host(self) -> tk.Widget | None:
        if not hasattr(self, "tabs"):
            return None
        try:
            return self._tab_scroll_hosts.get(str(self.tabs.select()))
        except tk.TclError:
            return None

    def _event_is_inside_active_tab_scroll_area(self, event: tk.Event) -> bool:
        host = self._current_tab_scroll_host()
        if host is None or not host.winfo_exists():
            return False
        x_root = getattr(event, "x_root", None)
        y_root = getattr(event, "y_root", None)
        if x_root is None or y_root is None:
            return False
        left = host.winfo_rootx()
        top = host.winfo_rooty()
        return left <= x_root < left + host.winfo_width() and top <= y_root < top + host.winfo_height()

    def _handle_tab_mousewheel(self, event: tk.Event) -> str | None:
        if str(event.widget.winfo_toplevel()) != str(self):
            return None
        if not self._event_is_inside_active_tab_scroll_area(event):
            return None
        canvas = self._current_tab_scroll_canvas()
        if canvas is None:
            return None
        orient = "x" if getattr(event, "state", 0) & 0x0001 else "y"
        return "break" if self._scroll_canvas_pixels(canvas, orient, self._wheel_pixels(event)) else None

    def _handle_tab_button4(self, event: tk.Event) -> str | None:
        if str(event.widget.winfo_toplevel()) != str(self):
            return None
        if not self._event_is_inside_active_tab_scroll_area(event):
            return None
        canvas = self._current_tab_scroll_canvas()
        return "break" if canvas is not None and self._scroll_canvas_pixels(canvas, "y", -72) else None

    def _handle_tab_button5(self, event: tk.Event) -> str | None:
        if str(event.widget.winfo_toplevel()) != str(self):
            return None
        if not self._event_is_inside_active_tab_scroll_area(event):
            return None
        canvas = self._current_tab_scroll_canvas()
        return "break" if canvas is not None and self._scroll_canvas_pixels(canvas, "y", 72) else None

    def _install_tab_scroll_bindings(self) -> None:
        self.bind_all("<MouseWheel>", self._handle_tab_mousewheel)
        self.bind_all("<Button-4>", self._handle_tab_button4)
        self.bind_all("<Button-5>", self._handle_tab_button5)

    def _should_skip_direct_tab_scroll(self, widget: tk.Widget) -> bool:
        widget_class = widget.winfo_class()
        return widget_class in {"Entry", "TEntry", "TCombobox", "Text", "Listbox", "Treeview"}

    def _bind_direct_tab_scroll(self, widget: tk.Widget) -> None:
        widget_id = str(widget)
        if widget_id not in self._tab_scroll_bound_widgets and not self._should_skip_direct_tab_scroll(widget):
            widget.bind("<MouseWheel>", self._handle_tab_mousewheel, add="+")
            widget.bind("<Button-4>", self._handle_tab_button4, add="+")
            widget.bind("<Button-5>", self._handle_tab_button5, add="+")
            self._tab_scroll_bound_widgets.add(widget_id)
        for child in widget.winfo_children():
            self._bind_direct_tab_scroll(child)

    def _refresh_tab_scroll_bindings(self) -> None:
        for content in self._tab_scroll_contents:
            if content.winfo_exists():
                self._bind_direct_tab_scroll(content)

    def _make_scrollable_tab(self, tab: ttk.Frame) -> ttk.Frame:
        host = ttk.Frame(tab, style="App.TFrame")
        host.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(
            host,
            bg=self.app_palette["bg"],
            borderwidth=0,
            highlightthickness=0,
            xscrollincrement=24,
            yscrollincrement=24,
        )
        y_scroll = ttk.Scrollbar(host, orient=tk.VERTICAL, command=canvas.yview)
        x_scroll = ttk.Scrollbar(host, orient=tk.HORIZONTAL, command=canvas.xview)
        canvas.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        host.columnconfigure(0, weight=1)
        host.rowconfigure(0, weight=1)
        scrollbars_visible = {"x": False, "y": False}

        content = ttk.Frame(canvas, style="App.TFrame")
        window_id = canvas.create_window((0, 0), window=content, anchor=tk.NW)
        sync_pending = False

        def sync_scroll_region() -> None:
            nonlocal sync_pending
            sync_pending = False
            content.update_idletasks()
            req_width = content.winfo_reqwidth()
            req_height = content.winfo_reqheight()
            canvas_width = max(canvas.winfo_width(), 1)
            canvas_height = max(canvas.winfo_height(), 1)
            needs_x = req_width > canvas_width
            needs_y = req_height > canvas_height
            if needs_y != scrollbars_visible["y"]:
                if needs_y:
                    y_scroll.grid(row=0, column=1, sticky="ns")
                else:
                    y_scroll.grid_remove()
                    canvas.yview_moveto(0)
                scrollbars_visible["y"] = needs_y
                canvas.after_idle(schedule_scroll_region_sync)
            if needs_x != scrollbars_visible["x"]:
                if needs_x:
                    x_scroll.grid(row=1, column=0, sticky="ew")
                else:
                    x_scroll.grid_remove()
                    canvas.xview_moveto(0)
                scrollbars_visible["x"] = needs_x
                canvas.after_idle(schedule_scroll_region_sync)
            width = max(req_width, canvas_width)
            height = max(req_height, canvas_height)
            canvas.itemconfigure(window_id, width=width, height=height)
            canvas.configure(scrollregion=(0, 0, width, height))

        def schedule_scroll_region_sync(_event: tk.Event | None = None) -> None:
            nonlocal sync_pending
            if sync_pending:
                return
            sync_pending = True
            canvas.after_idle(sync_scroll_region)

        content.bind("<Configure>", schedule_scroll_region_sync)
        canvas.bind("<Configure>", schedule_scroll_region_sync)
        self._tab_scroll_canvases[str(tab)] = canvas
        self._tab_scroll_hosts[str(tab)] = host
        self._tab_scroll_contents.append(content)
        return content

    def _build_ui(self) -> None:
        palette = {
            "bg": "#121212",
            "surface": "#181818",
            "panel": "#1f1f1f",
            "panel_high": "#242424",
            "field": "#2a2a2a",
            "border": "#333333",
            "muted": "#b3b3b3",
            "button": "#1ed760",
            "button_hover": "#1fdf64",
            "button_pressed": "#169c46",
            "soft_button": "#2a2a2a",
            "soft_button_hover": "#3a3a3a",
            "text": "#ffffff",
            "subtle_text": "#d9d9d9",
            "selection": "#1db954",
            "warning": "#5a4a14",
            "danger": "#5a1f1f",
        }
        self.app_palette = palette
        self.colors = palette
        self.configure(bg=palette["bg"])
        self.option_add("*TCombobox*Listbox.background", palette["field"])
        self.option_add("*TCombobox*Listbox.foreground", palette["text"])
        self.option_add("*TCombobox*Listbox.selectBackground", palette["selection"])
        self.option_add("*TCombobox*Listbox.selectForeground", "#000000")
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".", font=("Segoe UI", 10))
        style.configure("App.TFrame", background=palette["bg"])
        style.configure("Panel.TFrame", background=palette["panel"])
        style.configure("Header.TFrame", background=palette["surface"])
        style.configure("Header.TLabel", background=palette["surface"])
        style.configure("HeaderTitle.TLabel", background=palette["surface"], foreground=palette["text"], font=("Segoe UI Semibold", 22))
        style.configure("HeaderSub.TLabel", background=palette["surface"], foreground=palette["muted"])
        style.configure("BridgeBadge.TLabel", background=palette["panel_high"], foreground=palette["button"], font=("Segoe UI Semibold", 9), padding=(12, 7))
        style.configure("Panel.TLabel", background=palette["panel"], foreground=palette["text"])
        style.configure("Muted.TLabel", background=palette["panel"], foreground=palette["muted"])
        style.configure("AppTitle.TLabel", background=palette["bg"], foreground=palette["text"])
        style.configure("AppMuted.TLabel", background=palette["bg"], foreground=palette["muted"])
        style.configure("Status.TLabel", background=palette["bg"], foreground=palette["muted"])
        style.configure("Panel.TCheckbutton", background=palette["panel"], foreground=palette["text"])
        style.map(
            "Panel.TCheckbutton",
            background=[("active", palette["panel"])],
            foreground=[("active", palette["text"]), ("disabled", "#777777")],
        )
        style.configure(
            "ChromeTab.TButton",
            font=("Segoe UI Semibold", 9),
            padding=(12, 6),
            background=palette["soft_button"],
            foreground=palette["muted"],
            borderwidth=0,
            relief=tk.FLAT,
        )
        style.map(
            "ChromeTab.TButton",
            background=[("pressed", palette["border"]), ("active", palette["soft_button_hover"])],
            foreground=[("active", palette["text"])],
        )
        style.configure(
            "ChromeTabActive.TButton",
            font=("Segoe UI Semibold", 9),
            padding=(12, 6),
            background=palette["panel_high"],
            foreground=palette["text"],
            borderwidth=0,
            relief=tk.FLAT,
        )
        style.map(
            "ChromeTabActive.TButton",
            background=[("pressed", palette["panel_high"]), ("active", palette["panel_high"])],
            foreground=[("active", palette["text"])],
        )
        style.configure(
            "Primary.TButton",
            font=("Segoe UI Semibold", 10),
            padding=(18, 9),
            background=palette["button"],
            foreground="#000000",
            borderwidth=0,
            focusthickness=0,
            relief=tk.FLAT,
        )
        style.map(
            "Primary.TButton",
            background=[("pressed", palette["button_pressed"]), ("active", palette["button_hover"]), ("disabled", "#535353")],
            foreground=[("disabled", "#b3b3b3")],
            relief=[("pressed", tk.FLAT), ("!pressed", tk.FLAT)],
        )
        style.configure(
            "Soft.TButton",
            font=("Segoe UI Semibold", 10),
            padding=(16, 9),
            background=palette["button"],
            foreground="#000000",
            borderwidth=0,
            focusthickness=0,
            relief=tk.FLAT,
        )
        style.map(
            "Soft.TButton",
            background=[("pressed", palette["button_pressed"]), ("active", palette["button_hover"]), ("disabled", "#535353")],
            foreground=[("disabled", "#b3b3b3")],
            relief=[("pressed", tk.FLAT), ("!pressed", tk.FLAT)],
        )
        style.configure(
            "TButton",
            font=("Segoe UI Semibold", 10),
            padding=(16, 9),
            background=palette["button"],
            foreground="#000000",
            borderwidth=0,
            focusthickness=0,
            relief=tk.FLAT,
        )
        style.map(
            "TButton",
            background=[("pressed", palette["button_pressed"]), ("active", palette["button_hover"]), ("disabled", "#535353")],
            foreground=[("disabled", "#b3b3b3")],
            relief=[("pressed", tk.FLAT), ("!pressed", tk.FLAT)],
        )
        style.configure(
            "TEntry",
            fieldbackground=palette["field"],
            background=palette["field"],
            foreground=palette["text"],
            insertcolor=palette["text"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            padding=(8, 7),
        )
        style.map("TEntry", bordercolor=[("focus", palette["selection"])])
        style.configure(
            "TCombobox",
            fieldbackground=palette["field"],
            background=palette["field"],
            foreground=palette["text"],
            arrowcolor=palette["muted"],
            bordercolor=palette["border"],
            lightcolor=palette["border"],
            darkcolor=palette["border"],
            padding=(8, 6),
        )
        style.map(
            "TCombobox",
            fieldbackground=[("readonly", palette["field"])],
            foreground=[("readonly", palette["text"])],
            bordercolor=[("focus", palette["selection"])],
            arrowcolor=[("active", palette["text"])],
        )
        style.configure("TNotebook", background=palette["bg"], borderwidth=0, tabmargins=(0, 0, 0, 0))
        style.configure(
            "TNotebook.Tab",
            background=palette["bg"],
            foreground=palette["muted"],
            padding=(18, 10),
            borderwidth=0,
            font=("Segoe UI Semibold", 10),
        )
        style.map(
            "TNotebook.Tab",
            background=[("selected", palette["panel"]), ("active", palette["panel_high"])],
            foreground=[("selected", palette["text"]), ("active", palette["text"])],
        )
        style.configure("Vertical.TScrollbar", background=palette["field"], troughcolor=palette["panel"], bordercolor=palette["panel"], arrowcolor=palette["muted"])
        style.configure("Horizontal.TScrollbar", background=palette["field"], troughcolor=palette["panel"], bordercolor=palette["panel"], arrowcolor=palette["muted"])
        style.configure(
            "Assignment.Horizontal.TProgressbar",
            background="#16a34a",
            troughcolor="#ffffff",
            bordercolor="#d7dde3",
            lightcolor="#16a34a",
            darkcolor="#15803d",
        )
        style.configure("Treeview", rowheight=34, font=("Segoe UI", 10), background=palette["panel"], fieldbackground=palette["panel"], foreground=palette["subtle_text"], borderwidth=0)
        style.configure("Treeview.Heading", font=("Segoe UI Semibold", 9), background=palette["panel_high"], foreground=palette["muted"], padding=(10, 8), borderwidth=0)
        style.map("Treeview", background=[("selected", palette["selection"])], foreground=[("selected", "#000000")])

        header = ttk.Frame(self, style="Header.TFrame", padding=(18, 16))
        header.pack(fill=tk.X)
        header.columnconfigure(1, weight=1)
        logo_path = MIKEYS_CARDS_LOGO_PATH if self._is_personal_lucas() and MIKEYS_CARDS_LOGO_PATH.exists() else LUCAS_LOGO_PATH
        if logo_path.exists():
            try:
                self.logo_image = tk.PhotoImage(file=str(logo_path)).subsample(6, 6)
                self.iconphoto(False, self.logo_image)
                ttk.Label(header, image=self.logo_image, style="Header.TLabel").grid(row=0, column=0, rowspan=2, sticky="nw", padx=(0, 14))
            except tk.TclError:
                self.logo_image = None
        title_group = ttk.Frame(header, style="Header.TFrame")
        title_group.grid(row=0, column=1, sticky="w")
        ttk.Label(title_group, text=APP_TITLE, style="HeaderTitle.TLabel").pack(anchor=tk.W)
        ttk.Label(title_group, text=APP_SUBTITLE, style="HeaderSub.TLabel").pack(anchor=tk.W, pady=(3, 0))
        ttk.Label(header, textvariable=self.bridge_status_var, style="BridgeBadge.TLabel").grid(row=0, column=2, sticky="ne", padx=(16, 0))
        header_actions = ttk.Frame(header, style="Header.TFrame")
        header_actions.grid(row=1, column=1, columnspan=2, sticky="ew", pady=(12, 0))
        lucas_settings_button = self._make_colored_button(header_actions, "⚙ LUCAS Settings", lambda: self._show_lucas_settings_menu(lucas_settings_button), variant="primary")
        header_buttons = [lucas_settings_button]
        self._bind_responsive_button_row(header_actions, header_buttons, min_button_width=132)

        self.tabs = ttk.Notebook(self)
        self.tabs.pack(fill=tk.BOTH, expand=True, padx=18, pady=(16, 12))
        self.home_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=0)
        self.intake_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=0)
        self.comp_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=0)
        self.receive_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=0)
        self.review_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=0)
        self.payouts_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=0)
        self.inventory_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=0)
        self.profit_tab = ttk.Frame(self.tabs, style="App.TFrame", padding=0)
        self.tabs.add(self.home_tab, text="Home")
        self.tabs.add(self.intake_tab, text="Create")
        self.tabs.add(self.comp_tab, text="Comp")
        self.tabs.add(self.receive_tab, text="Receive")
        self.tabs.add(self.review_tab, text="Assignment")
        if not self._is_personal_lucas():
            self.tabs.add(self.payouts_tab, text="Payouts/Tabs")
        self.tabs.add(self.inventory_tab, text="Inventory")
        self.tabs.add(self.profit_tab, text="Profit")
        for tab_attr in (
            "home_tab",
            "intake_tab",
            "comp_tab",
            "receive_tab",
            "review_tab",
            "payouts_tab",
            "inventory_tab",
            "profit_tab",
        ):
            setattr(self, tab_attr, self._make_scrollable_tab(getattr(self, tab_attr)))
        self._install_tab_scroll_bindings()
        self.row_trees: list[ttk.Treeview] = []

        self._build_home_tab(palette)

        intake_controls = ttk.Frame(self.intake_tab, style="Panel.TFrame", padding=(16, 12))
        intake_controls.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(intake_controls, text="Input Mode", style="Panel.TLabel").grid(row=0, column=0, sticky="w")
        mode = ttk.Combobox(
            intake_controls,
            textvariable=self.input_mode,
            state="readonly",
            values=["Barcode Scanner", "Manual Entry", "Photo OCR", "Existing Spreadsheet"],
            width=22,
        )
        mode.grid(row=0, column=1, sticky="w", padx=(8, 16))
        mode.bind("<<ComboboxSelected>>", lambda _event: self._show_mode())
        ttk.Button(intake_controls, text="Clear Rows", command=self.clear_rows, style="Soft.TButton").grid(row=0, column=2, sticky="w", padx=(0, 8))
        ttk.Checkbutton(
            intake_controls,
            text="Network Mode",
            variable=self.create_network_mode_var,
            command=self._toggle_create_network_mode,
            style="Panel.TCheckbutton",
        ).grid(row=0, column=3, sticky="e", padx=(16, 0))
        intake_controls.columnconfigure(3, weight=1)
        self.network_seller_label = ttk.Label(intake_controls, text="Seller", style="Muted.TLabel")
        self.network_seller_label.grid(row=1, column=0, sticky="w", pady=(10, 0))
        self.seller_terms_seller_combo = ttk.Combobox(intake_controls, textvariable=self.seller_terms_seller_var, width=24)
        self.seller_terms_seller_combo.grid(row=1, column=1, sticky="w", padx=(8, 16), pady=(10, 0))
        self._bind_person_autocomplete(self.seller_terms_seller_combo, refresh_callback=self.apply_create_seller_terms, allow_blank=True)
        self.seller_terms_seller_combo.bind("<<ComboboxSelected>>", lambda _event: self.apply_create_seller_terms(), add="+")
        self.network_sheet_type_label = ttk.Label(intake_controls, text="Sheet Type", style="Muted.TLabel")
        self.network_sheet_type_label.grid(row=1, column=2, sticky="e", padx=(0, 6), pady=(10, 0))
        self.seller_terms_sheet_type_combo = ttk.Combobox(intake_controls, textvariable=self.seller_terms_sheet_type_var, width=18)
        self.seller_terms_sheet_type_combo.grid(row=1, column=3, sticky="w", padx=(0, 16), pady=(10, 0))
        self.seller_terms_sheet_type_combo.configure(postcommand=self._refresh_seller_terms_dropdowns)
        self.seller_terms_sheet_type_combo.bind("<<ComboboxSelected>>", lambda _event: self.apply_create_seller_terms(), add="+")
        ttk.Label(intake_controls, textvariable=self.summary_var, style="Muted.TLabel").grid(row=2, column=0, columnspan=5, sticky="w", pady=(10, 0))
        self._set_create_network_controls_visible(self._network_mode_enabled())

        self.mode_host = ttk.Frame(self.intake_tab, style="Panel.TFrame", padding=(16, 12))
        self.mode_host.pack(fill=tk.X, pady=(0, 10))
        self.intake_tree = self._build_table(self.intake_tab, editable=True, columns=INTAKE_COLUMNS)
        self._bind_context_menu(self.intake_tree, self._show_intake_context_menu)
        intake_save = ttk.Frame(self.intake_tab, style="Panel.TFrame", padding=(16, 12))
        intake_save.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(intake_save, text="Working Sheet Title", style="Panel.TLabel").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Entry(intake_save, textvariable=self.working_sheet_title, width=42).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(intake_save, text="Save as Working Sheet", command=self.save_working_sheet, style="Primary.TButton").pack(side=tk.LEFT)

        comp_body = ttk.Frame(self.comp_tab, style="App.TFrame")
        comp_body.pack(fill=tk.BOTH, expand=True)
        comp_body.columnconfigure(0, weight=0)
        comp_body.columnconfigure(1, weight=1)
        comp_body.rowconfigure(0, weight=1)
        sheet_panel = ttk.Frame(comp_body, style="Panel.TFrame", padding=(12, 12))
        sheet_panel.grid(row=0, column=0, sticky="ns", padx=(0, 10))
        sheet_panel.rowconfigure(1, weight=1)
        sheet_panel.columnconfigure(0, weight=1)
        ttk.Label(sheet_panel, text="Active Sheets", style="Panel.TLabel").grid(row=0, column=0, sticky="w")
        self.working_sheet_list = tk.Listbox(
            sheet_panel,
            width=34,
            height=10,
            activestyle="none",
            exportselection=False,
            bg=palette["panel"],
            fg=palette["subtle_text"],
            selectbackground=palette["selection"],
            selectforeground="#000000",
            highlightthickness=1,
            highlightbackground=palette["border"],
            highlightcolor=palette["selection"],
            relief=tk.FLAT,
            borderwidth=0,
            font=("Segoe UI", 10),
        )
        self.working_sheet_list.grid(row=1, column=0, sticky="nsew", pady=(8, 8))
        self.working_sheet_list.bind("<Double-Button-1>", lambda _event: self.load_selected_working_sheet())
        ttk.Button(sheet_panel, text="Load Selected Sheet", command=self.load_selected_working_sheet, style="Primary.TButton").grid(row=2, column=0, sticky="ew", pady=(0, 8))
        ttk.Button(sheet_panel, text="Refresh Sheet List", command=self.refresh_pipeline, style="Soft.TButton").grid(row=3, column=0, sticky="ew")
        comp_main = ttk.Frame(comp_body, style="App.TFrame")
        comp_main.grid(row=0, column=1, sticky="nsew")
        comp_main.rowconfigure(0, weight=1)
        comp_main.columnconfigure(0, weight=1)
        self.comp_tree = self._build_table(comp_main, editable=True, columns=COMP_COLUMNS)
        self._bind_context_menu(self.comp_tree, self._show_comp_context_menu)
        comp_controls = ttk.Frame(comp_main, style="Panel.TFrame", padding=(16, 12))
        comp_controls.pack(fill=tk.X, pady=(10, 0))
        comp_actions = ttk.Frame(comp_controls, style="Panel.TFrame")
        comp_actions.pack(fill=tk.X)
        comp_options = ttk.Frame(comp_controls, style="Panel.TFrame")
        comp_options.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(comp_actions, text="Save Back to Source Sheet", command=self.save_comp_to_source_sheet, style="Soft.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(comp_actions, text="Run All Comps", command=self.run_all_comps, style="Primary.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(comp_actions, text="Stop Run", command=self.stop_comp_run, style="Soft.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(comp_actions, text="Clear Comp Rows", command=self.clear_comp_rows, style="Soft.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(comp_actions, text="Lot Price Fill", command=self.open_lot_purchase_fill_popup, style="Soft.TButton").pack(side=tk.LEFT, padx=(8, 0))
        self.comp_scope_combo = ttk.Combobox(
            comp_options,
            textvariable=self.comp_scope_label,
            state="readonly",
            values=(COMP_SCOPE_EMPTY, COMP_SCOPE_ALL),
            width=17,
        )
        self.comp_scope_combo.pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Label(comp_options, text="Run Scope", style="Panel.TLabel").pack(side=tk.RIGHT)
        self.comp_source_combo = ttk.Combobox(
            comp_options,
            textvariable=self.comp_source_label,
            state="readonly",
            values=COMP_SOURCE_OPTIONS,
            width=18,
        )
        self.comp_source_combo.pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Label(comp_options, text="Run", style="Panel.TLabel").pack(side=tk.RIGHT)
        self.comp_method_combo = ttk.Combobox(
            comp_options,
            textvariable=self.comp_strategy_label,
            state="readonly",
            values=list(COMP_STRATEGY_DISPLAY.keys()),
            width=20,
        )
        self.comp_method_combo.pack(side=tk.RIGHT, padx=(8, 0))
        self.comp_method_combo.bind("<<ComboboxSelected>>", self.recalculate_comp_method)
        ttk.Label(comp_options, text="Comp Method", style="Panel.TLabel").pack(side=tk.RIGHT)
        self.comp_low_outlier_combo = ttk.Combobox(
            comp_options,
            textvariable=self.comp_low_outlier_pct_var,
            state="readonly",
            values=("Off", "50%", "75%", "90%"),
            width=7,
        )
        self.comp_low_outlier_combo.pack(side=tk.RIGHT, padx=(8, 0))
        self.comp_low_outlier_combo.bind("<<ComboboxSelected>>", self.recalculate_comp_method)
        ttk.Label(comp_options, text="Low Comp % Avg", style="Panel.TLabel").pack(side=tk.RIGHT)

        receive_controls = ttk.Frame(self.receive_tab, style="Panel.TFrame", padding=(16, 12))
        receive_controls.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(receive_controls, text="Receive Mode", style="Panel.TLabel").grid(row=0, column=0, sticky="w")
        review_mode = ttk.Combobox(
            receive_controls,
            textvariable=self.review_mode,
            state="readonly",
            values=["Automatic Receive", "Manual Receive"],
            width=20,
        )
        review_mode.grid(row=0, column=1, sticky="w", padx=(8, 16))
        review_mode.bind("<<ComboboxSelected>>", lambda _event: self._show_review_mode())
        ttk.Label(receive_controls, textvariable=self.review_status, style="Muted.TLabel").grid(row=1, column=0, columnspan=5, sticky="w", pady=(10, 0))
        receive_controls.columnconfigure(4, weight=1)

        self.review_mode_host = ttk.Frame(self.receive_tab, style="Panel.TFrame", padding=(16, 12))
        self.review_mode_host.pack(fill=tk.X, pady=(0, 10))
        self.receive_tree = self._build_table(self.receive_tab, editable=True, columns=self._personal_person_last_columns(RECEIVE_COLUMNS))
        self._bind_context_menu(self.receive_tree, self._show_receive_context_menu)
        receive_bottom = ttk.Frame(self.receive_tab, style="Panel.TFrame", padding=(16, 12))
        receive_bottom.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(receive_bottom, text="Mark Received in Sheets", command=self.mark_review_received_in_sheets, style="Primary.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(receive_bottom, text="Refresh Incoming Index", command=self.refresh_incoming_index, style="Primary.TButton").pack(side=tk.RIGHT, padx=(8, 0))

        review_controls = ttk.Frame(self.review_tab, style="Panel.TFrame", padding=(16, 12))
        review_controls.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(review_controls, text="Received Sheet", style="Panel.TLabel").grid(row=0, column=0, sticky="w")
        self.received_sheet_combo = ttk.Combobox(review_controls, textvariable=self.selected_received_sheet, state="readonly", width=32)
        self.received_sheet_combo.grid(row=0, column=1, sticky="ew", padx=(8, 8))
        ttk.Button(review_controls, text="Load", command=self.load_selected_received_sheet_for_review, style="Primary.TButton").grid(row=0, column=2, sticky="w", padx=(0, 8))
        ttk.Button(review_controls, text="Refresh Received Sheets", command=self.refresh_received_sheets, style="Soft.TButton").grid(row=0, column=3, sticky="w")
        ttk.Button(review_controls, text="Company Rules", command=self.open_assignment_rules, style="Soft.TButton").grid(row=0, column=4, sticky="w", padx=(8, 0))
        ttk.Button(review_controls, text="People Rules", command=self.open_people_rules, style="Soft.TButton").grid(row=0, column=5, sticky="w", padx=(8, 0))
        ttk.Button(review_controls, text="Unassigned Players", command=self.open_unassigned_players_dialog, style="Soft.TButton").grid(row=0, column=7, sticky="w", padx=(8, 0))
        review_controls.columnconfigure(1, weight=1)
        ttk.Label(review_controls, textvariable=self.review_status, style="Muted.TLabel").grid(row=1, column=0, columnspan=8, sticky="w", pady=(10, 0))
        ttk.Label(review_controls, textvariable=self.assignment_config_status, style="Muted.TLabel").grid(row=2, column=0, columnspan=8, sticky="w", pady=(4, 0))
        self.assignment_progress = ttk.Progressbar(
            review_controls,
            style="Assignment.Horizontal.TProgressbar",
            variable=self.assignment_progress_value,
            maximum=100,
            mode="determinate",
        )
        self.assignment_progress.grid(row=3, column=0, columnspan=8, sticky="ew", pady=(8, 0))
        self.review_tree = self._build_table(self.review_tab, editable=True, columns=REVIEW_COLUMNS)
        self._bind_context_menu(self.review_tree, self._show_receive_context_menu)
        self._show_review_mode()
        self._build_payouts_tab()
        self._build_inventory_tab()
        self._build_profit_tab()

        bottom = ttk.Frame(self, style="App.TFrame", padding=(16, 0, 16, 14))
        bottom.pack(fill=tk.X)
        ttk.Label(bottom, textvariable=self.status_var, style="Status.TLabel").pack(side=tk.LEFT)
        ttk.Label(bottom, textvariable=self.version_var, style="Muted.TLabel").pack(side=tk.RIGHT)

    def open_setup_doctor(self) -> None:
        rows = self._setup_doctor_results()
        dialog = tk.Toplevel(self)
        dialog.title("LUCAS System Health")
        dialog.geometry("840x520")
        dialog.transient(self)
        dialog.configure(bg="#121212")
        frame = ttk.Frame(dialog, style="App.TFrame", padding=16)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="System Health", style="HeaderTitle.TLabel").pack(anchor=tk.W, pady=(0, 10))
        tree = ttk.Treeview(frame, columns=("status", "detail"), show="tree headings", height=14)
        tree.heading("#0", text="Check", anchor=tk.W)
        tree.heading("status", text="Status", anchor=tk.W)
        tree.heading("detail", text="Detail", anchor=tk.W)
        tree.column("#0", width=220, stretch=False)
        tree.column("status", width=130, stretch=False)
        tree.column("detail", width=460, stretch=True)
        tree.pack(fill=tk.BOTH, expand=True)
        for row in rows:
            tree.insert("", tk.END, text=row["name"], values=(row["status"], row["detail"]))
        actions = ttk.Frame(frame, style="App.TFrame")
        actions.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(actions, text="Copy Details", command=lambda: self._copy_setup_doctor_details(rows), style="Soft.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(actions, text="Close", command=dialog.destroy, style="Soft.TButton").pack(side=tk.RIGHT)
        self.status_var.set("System health check complete.")

    def _setup_doctor_results(self) -> list[dict[str, str]]:
        snapshot = self.state.snapshot() if hasattr(self, "state") else {}
        rows = setup_doctor_results(CARD_PIPELINE_DIR, snapshot, "Mac")
        rows.extend(self._operational_health_rows())
        return rows

    def _health_row(self, name: str, ok: bool, detail: str) -> dict[str, str]:
        return {"name": name, "status": "OK" if ok else "Needs attention", "detail": detail}

    def _stale_lock_files(self, minutes: int = 30) -> list[Path]:
        lock_dir = CARD_PIPELINE_DIR / ".locks"
        if not lock_dir.exists():
            return []
        cutoff = time.time() - (minutes * 60)
        stale: list[Path] = []
        for path in lock_dir.glob("*.lock"):
            try:
                if path.stat().st_mtime < cutoff:
                    stale.append(path)
            except OSError:
                continue
        return sorted(stale, key=lambda item: item.name.lower())

    def _operational_health_rows(self) -> list[dict[str, str]]:
        conflict_files = self._shared_conflict_files() if CARD_PIPELINE_DIR.exists() else []
        stale_locks = self._stale_lock_files()
        assignment_companies = [company for company in getattr(self.assignment_engine, "companies", []) if getattr(company, "name", "")]
        seller_lines = seller_terms_health_lines(SELLER_TERMS_PATH, self._assignment_company_health_payload())
        seller_ok = "error(s)" in seller_lines[0] and "0 error(s)" in seller_lines[0]
        activity_count = len(self._load_activity_log())
        profit_count = len(self._load_profit_ledger())
        inventory_count = len(self._load_inventory_ledger())
        return [
            self._health_row("Shared conflict files", not conflict_files, ", ".join(path.name for path in conflict_files[:5]) or "none"),
            self._health_row("Stale shared locks", not stale_locks, ", ".join(path.name for path in stale_locks[:5]) or "none"),
            self._health_row("Assignment companies loaded", bool(assignment_companies), f"{len(assignment_companies)} active/configured company object(s)"),
            self._health_row("Seller terms health", seller_ok, seller_lines[0] if seller_lines else "not checked"),
            self._health_row("Inventory ledger", INVENTORY_LEDGER_PATH.exists(), f"{inventory_count} active/raw ledger item(s) | {INVENTORY_LEDGER_PATH}"),
            self._health_row("Profit ledger", PROFIT_LEDGER_PATH.exists(), f"{profit_count} ledger row(s) | {PROFIT_LEDGER_PATH}"),
            self._health_row("Activity log", True, f"{activity_count} recent operation record(s) | {ACTIVITY_LOG_PATH}"),
            self._health_row("Mobile queue log", MOBILE_ACTION_LOG_PATH.exists(), f"{MOBILE_ACTION_LOG_PATH}"),
        ]

    def _assignment_company_health_payload(self) -> list[dict[str, object]]:
        try:
            raw = json.loads(ASSIGNMENT_CONFIG_PATH.read_text(encoding="utf-8")) if ASSIGNMENT_CONFIG_PATH.exists() else {}
        except Exception:
            raw = {}
        companies = raw.get("companies", raw) if isinstance(raw, dict) else raw
        return [company for company in companies if isinstance(company, dict)] if isinstance(companies, list) else []

    def _copy_setup_doctor_details(self, rows: list[dict[str, str]]) -> None:
        payload = {"version": self.version_var.get(), "checks": rows}
        self.clipboard_clear()
        self.clipboard_append(diagnostic_json(payload))
        self.status_var.set("Copied system health details.")

    def _build_table(self, parent: ttk.Frame, editable: bool = False, columns: tuple[str, ...] = DISPLAY_COLUMNS) -> ttk.Treeview:
        content = ttk.Frame(parent, style="Panel.TFrame", padding=(1, 1))
        content.configure(width=900, height=260)
        content.grid_propagate(False)
        content.pack(fill=tk.BOTH, expand=True)
        tree = ttk.Treeview(content, columns=columns, show="headings", selectmode="extended")
        setattr(tree, "_display_columns", columns)
        for col in columns:
            tree.heading(col, text=HEADINGS[col], anchor=tk.W)
            tree.column(col, width=COLUMN_WIDTHS[col], minwidth=45, stretch=False)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(content, orient=tk.VERTICAL, command=tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(content, orient=tk.HORIZONTAL, command=tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.tag_configure("duplicate_cert", background="#4a3d12", foreground="#fff3b0")
        tree.tag_configure("no_sheet_found", background="#4a1717", foreground="#ffd1d1")
        tree.tag_configure("add_review_row", background="#242424", foreground="#1ed760")
        if editable:
            tree.bind("<Double-1>", self._begin_cell_edit)
            tree.bind("<Button-1>", self._handle_table_click, add="+")
            tree.bind("<Delete>", self._delete_selected_table_rows)
        tree.bind("<ButtonRelease-1>", lambda _event, target=tree: self._remember_column_widths(target), add="+")
        content.columnconfigure(0, weight=1)
        content.rowconfigure(0, weight=1)
        setattr(tree, "_table_frame", content)
        self.row_trees.append(tree)
        self.column_widths_by_tree[id(tree)] = {col: COLUMN_WIDTHS[col] for col in columns}
        return tree

    def _build_home_tab(self, palette: dict[str, str]) -> None:
        body = ttk.Frame(self.home_tab, style="App.TFrame")
        body.pack(fill=tk.BOTH, expand=True)

        sheet_panel = ttk.Frame(body, style="Panel.TFrame", padding=(12, 12))
        sheet_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        sheet_panel.configure(width=360)
        sheet_panel.pack_propagate(False)
        person_row = ttk.Frame(sheet_panel, style="Panel.TFrame")
        if not self._is_personal_lucas():
            person_row.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(person_row, text="Person", style="Muted.TLabel").pack(side=tk.LEFT)
        self.home_person_combo = ttk.Combobox(person_row, textvariable=self.home_person_var, width=24)
        self.home_person_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))
        self._bind_person_autocomplete(self.home_person_combo, refresh_callback=self._on_home_person_filter_changed, allow_blank=True)
        self.home_person_combo.bind("<<ComboboxSelected>>", lambda _event: self._on_home_person_filter_changed(), add="+")
        toggle_row = tk.Frame(sheet_panel, bg=palette["panel"])
        toggle_row.pack(fill=tk.X, pady=(0, 8))
        self.home_tab_palette = palette
        self.home_incoming_tab = self._build_home_tab_button(toggle_row, "Incoming", lambda: self._set_home_sheet_kind("Incoming"))
        self.home_working_tab = self._build_home_tab_button(toggle_row, "Working", lambda: self._set_home_sheet_kind("Working"))
        self.home_received_tab = self._build_home_tab_button(toggle_row, "Received", lambda: self._set_home_sheet_kind("Received"))
        self.home_edit_markers_tab = self._build_home_tab_button(toggle_row, "Edit Markers", self.open_sheet_marker_editor)
        self._bind_responsive_button_row(
            toggle_row,
            [self.home_incoming_tab, self.home_working_tab, self.home_received_tab, self.home_edit_markers_tab],
            min_button_width=70,
            uniform_columns=True,
        )
        sort_row = ttk.Frame(sheet_panel, style="Panel.TFrame")
        sort_row.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(sort_row, text="Sort", style="Muted.TLabel").pack(side=tk.LEFT)
        self.home_sheet_sort_combo = ttk.Combobox(
            sort_row,
            textvariable=self.home_sheet_sort_var,
            values=["Date Created", "Name"],
            width=16,
            state="readonly",
        )
        self.home_sheet_sort_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(8, 0))
        self.home_sheet_sort_combo.bind("<<ComboboxSelected>>", lambda _event: self._refresh_home_sheet_list(), add="+")
        self.home_sheet_list = tk.Listbox(
            sheet_panel,
            width=1,
            height=10,
            activestyle="none",
            exportselection=False,
            bg=palette["panel"],
            fg=palette["subtle_text"],
            selectbackground=palette["selection"],
            selectforeground="#000000",
            highlightthickness=1,
            highlightbackground=palette["border"],
            highlightcolor=palette["selection"],
            relief=tk.FLAT,
            borderwidth=0,
            font=("Segoe UI", 10),
        )
        self.home_sheet_list.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        self.home_sheet_list.bind("<<ListboxSelect>>", lambda _event: self._load_home_selected_marker())
        self._bind_context_menu(self.home_sheet_list, self._show_home_sheet_context_menu)
        self._make_colored_button(sheet_panel, "Refresh Home View", self.refresh_home, variant="primary").pack(fill=tk.X)

        right = ttk.Frame(body, style="App.TFrame")
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        metrics = tk.PanedWindow(
            right,
            orient=tk.VERTICAL,
            sashwidth=6,
            sashrelief=tk.RAISED,
            bg=palette["bg"],
            bd=0,
            showhandle=False,
        )
        metrics.pack(fill=tk.BOTH, expand=True)
        self.home_metrics_pane = metrics
        volume_panel = ttk.Frame(metrics, style="Panel.TFrame", padding=(12, 12))
        ttk.Label(volume_panel, text="Incoming Volume by Sheet", style="Panel.TLabel").pack(anchor=tk.W)
        incoming_volume_headings = {"sheet": "Sheet", "person": "Person", "cards": "Cards", "received": "Received", "volume": "Price Volume", "status": "Status"}
        self.incoming_volume_tree = self._build_home_tree(
            volume_panel,
            columns=("sheet", "person", "cards", "received", "volume", "status"),
            headings=incoming_volume_headings,
            widths={"sheet": 320, "person": 130, "cards": 80, "received": 95, "volume": 130, "status": 150},
            height=9,
        )
        self.incoming_volume_headings = incoming_volume_headings
        self._configure_sortable_tree_headings(self.incoming_volume_tree, incoming_volume_headings, "home_incoming_volume")
        self.incoming_volume_tree.tag_configure("total_divider", background="#1f1f1f", foreground="#ffffff", font=("Segoe UI Semibold", 10))
        self.incoming_volume_tree.tag_configure("total_row", background="#242424", foreground="#ffffff", font=("Segoe UI Semibold", 10))

        partial_panel = ttk.Frame(metrics, style="Panel.TFrame", padding=(12, 12))
        ttk.Label(partial_panel, text="Partially Received Incoming Sheets", style="Panel.TLabel").pack(anchor=tk.W)
        self.partial_received_tree = self._build_home_tree(
            partial_panel,
            columns=("sheet", "progress", "volume", "person", "tracking", "all_received"),
            headings={"sheet": "Sheet", "progress": "Received", "volume": "Price Volume", "person": "Person", "tracking": "Tracking", "all_received": "All Received"},
            widths={"sheet": 280, "progress": 100, "volume": 130, "person": 130, "tracking": 180, "all_received": 110},
            height=8,
        )
        self.partial_received_tree.tag_configure("partial_sheet", background="#4a3d12", foreground="#fff3b0")
        metrics.add(volume_panel, minsize=150, stretch="always")
        metrics.add(partial_panel, minsize=150, stretch="always")

    def _build_home_tree(
        self,
        parent: ttk.Frame,
        columns: tuple[str, ...],
        headings: dict[str, str],
        widths: dict[str, int],
        height: int,
        scrollbars: bool = False,
        max_height: int = 180,
    ) -> ttk.Treeview:
        container = ttk.Frame(parent, style="Panel.TFrame")
        table_height = max(120, min(max_height, height * 24 + 42))
        container.configure(width=560, height=table_height)
        container.grid_propagate(False)
        container.pack_propagate(False)
        tree = ttk.Treeview(container, columns=columns, show="headings", selectmode="browse", height=height)
        for col in columns:
            tree.heading(col, text=headings[col], anchor=tk.W)
            tree.column(col, width=widths[col], minwidth=60, stretch=col == "sheet", anchor=tk.W)
        container.pack(fill=tk.BOTH, expand=True, pady=(8, 0))
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(container, orient=tk.VERTICAL, command=tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(container, orient=tk.HORIZONTAL, command=tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)
        return tree

    def _personal_person_last_columns(self, columns: tuple[str, ...]) -> tuple[str, ...]:
        if not getattr(self, "_is_personal_lucas", lambda: False)() or "person" not in columns:
            return columns
        return tuple(column for column in columns if column != "person") + ("person",)

    def _configure_sortable_tree_headings(self, tree: ttk.Treeview, headings: dict[str, str], table: str) -> None:
        if table == "inventory":
            sort_column = self.inventory_sort_column
            descending = self.inventory_sort_descending
        elif table == "home_incoming_volume":
            sort_column = self.home_incoming_volume_sort_column
            descending = self.home_incoming_volume_sort_descending
        else:
            sort_column = self.profit_sort_column
            descending = self.profit_sort_descending
        for column in tree["columns"]:
            label = headings.get(column, column)
            if column == sort_column:
                label = f"{label} {'v' if descending else '^'}"
            if table == "inventory":
                command = lambda col=column: self._sort_inventory_by_column(col)
            elif table == "home_incoming_volume":
                command = lambda col=column: self._sort_home_incoming_volume_by_column(col)
            else:
                command = lambda col=column: self._sort_profit_by_column(col)
            tree.heading(column, text=label, anchor=tk.W, command=command)

    def _sort_inventory_by_column(self, column: str) -> None:
        if column == self.inventory_sort_column:
            self.inventory_sort_descending = not self.inventory_sort_descending
        else:
            self.inventory_sort_column = column
            self.inventory_sort_descending = False
        self.refresh_inventory_tab()

    def _sort_home_incoming_volume_by_column(self, column: str) -> None:
        if column == self.home_incoming_volume_sort_column:
            self.home_incoming_volume_sort_descending = not self.home_incoming_volume_sort_descending
        else:
            self.home_incoming_volume_sort_column = column
            self.home_incoming_volume_sort_descending = column in {"cards", "received", "volume"}
        self._refresh_home_metrics()

    def _sort_profit_by_column(self, column: str) -> None:
        if column == self.profit_sort_column:
            self.profit_sort_descending = not self.profit_sort_descending
        else:
            self.profit_sort_column = column
            self.profit_sort_descending = False
        self.refresh_profit_tab()

    def _profit_added_sort_value(self, record: dict[str, object]) -> str:
        added_at = str(record.get("ledger_added_at") or record.get("created_at") or record.get("recorded_at") or "").strip()
        if added_at:
            return added_at
        expense_id = str(record.get("expense_id") or "").strip()
        if re.fullmatch(r"\d{14,20}", expense_id):
            return (
                f"{expense_id[0:4]}-{expense_id[4:6]}-{expense_id[6:8]}"
                f"T{expense_id[8:10]}:{expense_id[10:12]}:{expense_id[12:14]}.{expense_id[14:]}"
            )
        date_text = str(record.get("date_added") or "").strip()[:10]
        return f"{date_text}T00:00:00" if self._profit_record_date(date_text) is not None else ""

    def _record_sort_value(self, record: dict[str, object], column: str, table: str, mode: str = "") -> tuple[bool, object]:
        money_columns = {"purchase", "sale", "profit", "amount", "payout", "card_ladder", "comps", "cy_estimate", "volume"}
        int_columns = {"cards", "received"}
        if table == "home_incoming_volume":
            raw = record.get(column)
        elif table == "inventory":
            field_map = {
                "date": "date_added",
                "type": "item_type",
                "item_id": "item_id",
                "person": "assigned_person",
                "sport": "sport",
                "cert": "cert_number",
                "grader": "grader",
                "card": "card_title",
                "purchase": "purchase_price",
                "card_ladder": "card_ladder_value",
                "comps": "card_ladder_comps_average",
                "cy_estimate": "cy_value",
                "cy_confidence": "cy_confidence",
                "company": "best_company",
                "payout": "estimated_payout",
                "source": "source_sheet",
                "status": "status",
                "photos": "photo_count",
                "notes": "notes",
            }
            raw = inventory_display_notes(record) if column == "notes" else record.get(field_map.get(column, column))
        elif table == "profit_sheet":
            raw = record.get(column)
        else:
            field_map = {
                "date": "date_added",
                "person": "assigned_person",
                "company": "company",
                "card": "card_title",
                "cert": "cert_number",
                "purchase": "purchase_price",
                "sale": "sale_price",
                "profit": "profit",
                "sheet": "source_sheet",
                "type": "expense_type",
                "amount": "expense_amount",
                "related": "source_sheet",
                "notes": "notes",
            }
            if column == "cert":
                raw = record.get("cert_number") or record.get("item_id")
            elif column == "sheet":
                raw = record.get("weekly_sheet_name") or record.get("source_sheet")
            elif column == "related":
                raw = self._expense_related_label(record)
            else:
                raw = record.get(field_map.get(column, column))
            if column == "date":
                text = str(raw or "").strip()
                return not bool(text), (text.casefold(), self._profit_added_sort_value(record))
        if column in money_columns:
            value = self._money_value(raw)
            return value is None, value if value is not None else 0.0
        if column in int_columns:
            try:
                return False, int(raw or 0)
            except (TypeError, ValueError):
                return True, 0
        text = str(raw or "").strip()
        return not bool(text), text.casefold()

    def _sorted_records(self, rows: list[dict[str, object]], column: str, descending: bool, table: str, mode: str = "") -> list[dict[str, object]]:
        keyed = [(self._record_sort_value(record, column, table, mode), index, record) for index, record in enumerate(rows)]
        present = [item for item in keyed if not item[0][0]]
        missing = [item for item in keyed if item[0][0]]
        present.sort(key=lambda item: (item[0][1], item[1]), reverse=descending)
        return [item[2] for item in present + missing]

    def _build_payouts_tab(self) -> None:
        controls = ttk.Frame(self.payouts_tab, style="Panel.TFrame", padding=(16, 12))
        controls.pack(fill=tk.X, pady=(0, 10))
        payout_person_label = ttk.Label(controls, text="Filter by Assigned Person", style="Panel.TLabel")
        if not self._is_personal_lucas():
            payout_person_label.grid(row=0, column=0, sticky="w")
        self.payout_person_combo = ttk.Combobox(controls, textvariable=self.payout_person_var, width=30)
        if not self._is_personal_lucas():
            self.payout_person_combo.grid(row=0, column=1, sticky="w", padx=(8, 10))
        self._bind_person_autocomplete(self.payout_person_combo, refresh_callback=self.refresh_payouts_tab, allow_blank=True)
        self.payout_person_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_payouts_tab(), add="+")
        controls.columnconfigure(2, weight=1)
        ttk.Label(controls, textvariable=self.payout_status_var, style="Muted.TLabel").grid(row=1, column=0, columnspan=3, sticky="w", pady=(10, 0))

        payout_split = tk.PanedWindow(
            self.payouts_tab,
            orient=tk.HORIZONTAL,
            bg=self.app_palette["border"],
            bd=0,
            sashwidth=8,
            sashrelief=tk.RAISED,
            showhandle=True,
            handlesize=28,
            opaqueresize=True,
        )
        payout_split.pack(fill=tk.BOTH, expand=True)
        summary_panel = ttk.Frame(payout_split, style="Panel.TFrame", padding=(12, 12))
        payout_split.add(summary_panel, minsize=320)
        ttk.Label(summary_panel, text="Active Balances", style="Panel.TLabel").pack(anchor=tk.W)
        self.payout_summary_tree = self._build_home_tree(
            summary_panel,
            columns=("person", "sheets", "cards", "expenses", "total_net_profit", "unpaid_net_profit", "balance"),
            headings={
                "person": "Person",
                "sheets": "Sheets",
                "cards": "Cards",
                "expenses": "Expenses",
                "total_net_profit": "Total Net Profit",
                "unpaid_net_profit": "Unpaid Net Profit",
                "balance": "Balance Owed",
            },
            widths={"person": 180, "sheets": 65, "cards": 65, "expenses": 100, "total_net_profit": 120, "unpaid_net_profit": 125, "balance": 120},
            height=18,
        )
        self.payout_summary_tree.tag_configure("total_divider", background="#1f1f1f", foreground="#ffffff", font=("Segoe UI Semibold", 10))
        self.payout_summary_tree.tag_configure("total_row", background="#242424", foreground="#ffffff", font=("Segoe UI Semibold", 10))
        self.payout_summary_tree.bind("<ButtonRelease-1>", self.mark_payout_person_paid)
        self.payout_summary_tree.bind("<Button-3>", self.open_payout_history_menu)
        self.payout_summary_tree.bind("<Control-Button-1>", self.open_payout_history_menu)

        detail_panel = ttk.Frame(payout_split, style="Panel.TFrame", padding=(12, 12))
        payout_split.add(detail_panel, minsize=360)
        ttk.Label(detail_panel, text="Payment Sheets", style="Panel.TLabel").pack(anchor=tk.W)
        self.payout_detail_tree = self._build_home_tree(
            detail_panel,
            columns=("sheet", "stage", "person", "cards", "received", "volume", "status"),
            headings={"sheet": "Sheet", "stage": "Stage", "person": "Person", "cards": "Cards", "received": "Received", "volume": "Balance", "status": "Status"},
            widths={"sheet": 280, "stage": 90, "person": 150, "cards": 80, "received": 95, "volume": 130, "status": 140},
            height=18,
        )
        self.payout_detail_tree.configure(selectmode="extended")
        self.payout_detail_tree.bind("<ButtonRelease-1>", self.open_payout_marker_editor)

    def _build_inventory_tab(self) -> None:
        controls = ttk.Frame(self.inventory_tab, style="Panel.TFrame", padding=(16, 12))
        controls.pack(fill=tk.X, pady=(0, 10))
        controls.columnconfigure(9, weight=1)
        ttk.Label(controls, text="Inventory", style="Panel.TLabel", font=("Segoe UI Semibold", 13)).grid(row=0, column=0, sticky="w")
        ttk.Label(controls, textvariable=self.inventory_metric_var, style="Panel.TLabel").grid(row=0, column=1, columnspan=8, sticky="e", padx=(18, 0))
        inventory_person_label = ttk.Label(controls, text="Person", style="Muted.TLabel")
        if not self._is_personal_lucas():
            inventory_person_label.grid(row=1, column=0, sticky="w", pady=(10, 0))
        self.inventory_person_combo = ttk.Combobox(controls, textvariable=self.inventory_person_var, width=22)
        if not self._is_personal_lucas():
            self.inventory_person_combo.grid(row=1, column=1, sticky="w", padx=(8, 14), pady=(10, 0))
        self._bind_person_autocomplete(self.inventory_person_combo, refresh_callback=self.refresh_inventory_tab, allow_blank=True)
        self.inventory_person_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_inventory_tab(), add="+")
        search_label_column = 0 if self._is_personal_lucas() else 2
        search_entry_column = 1 if self._is_personal_lucas() else 3
        ttk.Label(controls, text="Search ID/Cert/Card", style="Muted.TLabel").grid(row=1, column=search_label_column, sticky="e", padx=(0, 6), pady=(10, 0))
        ttk.Entry(controls, textvariable=self.inventory_search_var, width=42).grid(row=1, column=search_entry_column, columnspan=4, sticky="w", pady=(10, 0))
        action_row = ttk.Frame(controls, style="Panel.TFrame")
        action_row.grid(row=2, column=0, columnspan=10, sticky="w", pady=(10, 0))
        ttk.Button(action_row, text="Add Card", command=self.add_raw_inventory_card, style="Primary.TButton").pack(side=tk.LEFT)
        ttk.Button(action_row, text="Export", command=self.export_inventory, style="Primary.TButton").pack(side=tk.LEFT, padx=(8, 0))
        self._make_inventory_toolbar_icon_button(
            action_row,
            "filter",
            "Open inventory filters for sport, grader, card year, price, date, missing values, descriptions, and photos.",
            self.open_inventory_filters_popup,
        ).pack(side=tk.LEFT, padx=(8, 0))
        settings_button = self._make_inventory_toolbar_icon_button(
            action_row,
            "gear",
            "Open inventory settings and maintenance actions.",
            lambda: self._show_inventory_settings_menu(settings_button),
        )
        settings_button.pack(side=tk.LEFT, padx=(8, 0))
        self.inventory_bulk_toggle = tk.Checkbutton(
            action_row,
            text="Bulk Edit",
            variable=self.inventory_bulk_edit_var,
            command=self._toggle_inventory_bulk_edit,
            indicatoron=False,
            cursor="hand2",
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=0,
            padx=14,
            pady=7,
            font=("Segoe UI Semibold", 9),
        )
        self.inventory_bulk_toggle.pack(side=tk.LEFT, padx=(14, 0))
        self._style_inventory_bulk_toggle()
        ttk.Label(controls, textvariable=self.inventory_status_var, style="Muted.TLabel").grid(row=3, column=0, columnspan=10, sticky="w", pady=(8, 0))
        for var in (self.inventory_sport_var, self.inventory_grader_var, self.inventory_year_var, self.inventory_search_var, self.inventory_min_var, self.inventory_max_var, self.inventory_date_min_var, self.inventory_date_max_var, self.inventory_missing_title_var, self.inventory_missing_comps_var, self.inventory_missing_cl_var, self.inventory_missing_photos_var):
            var.trace_add("write", lambda *_args: self._schedule_inventory_filter_refresh())

        self.inventory_tree = self._build_home_tree(
            self.inventory_tab,
            columns=self._personal_person_last_columns(INVENTORY_TABLE_COLUMNS),
            headings=INVENTORY_HEADINGS,
            widths=INVENTORY_COLUMN_WIDTHS,
            height=22,
            scrollbars=True,
        )
        self.inventory_tree.configure(selectmode="extended")
        self._configure_sortable_tree_headings(self.inventory_tree, INVENTORY_HEADINGS, "inventory")
        self._bind_context_menu(self.inventory_tree, self._show_inventory_context_menu)
        self.inventory_tree.bind("<Button-1>", self._inventory_bulk_click, add="+")
        self.inventory_tree.bind("<Double-1>", self._begin_inventory_bulk_edit, add="+")
        self.inventory_tree.bind("<Return>", self._begin_inventory_bulk_edit, add="+")
        self.inventory_tree.bind("<F2>", self._begin_inventory_bulk_edit, add="+")
        self.inventory_tree.bind("<Up>", lambda event: self._move_inventory_bulk_cell(-1, 0), add="+")
        self.inventory_tree.bind("<Down>", lambda event: self._move_inventory_bulk_cell(1, 0), add="+")
        self.inventory_tree.bind("<Left>", lambda event: self._move_inventory_bulk_cell(0, -1), add="+")
        self.inventory_tree.bind("<Right>", lambda event: self._move_inventory_bulk_cell(0, 1), add="+")
        self.inventory_tree.bind("<Control-z>", self._undo_inventory_bulk_edit, add="+")
        self.inventory_tree.bind("<Control-Z>", self._undo_inventory_bulk_edit, add="+")
        self.inventory_tree.bind("<Command-z>", self._undo_inventory_bulk_edit, add="+")
        self.inventory_tree.bind("<Command-Z>", self._undo_inventory_bulk_edit, add="+")
        self.refresh_inventory_tab()

    def open_inventory_filters_popup(self) -> None:
        popup = tk.Toplevel(self)
        popup.title("Inventory Filters")
        popup.configure(bg=self.colors["bg"])
        popup.transient(self)
        popup.geometry("600x500")
        popup.minsize(560, 455)
        frame = ttk.Frame(popup, style="App.TFrame", padding=18)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Inventory Filters", style="AppTitle.TLabel", font=("Segoe UI Semibold", 13)).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 12))

        ttk.Label(frame, text="Sport", style="AppMuted.TLabel").grid(row=1, column=0, sticky="w", pady=(0, 10))
        sport_frame = ttk.Frame(frame, style="App.TFrame")
        sport_frame.grid(row=1, column=1, columnspan=3, sticky="ew", pady=(0, 10))
        selected_sports = self._inventory_sport_filter_values()
        sport_vars: dict[str, tk.BooleanVar] = {}

        def sync_sport_filter() -> None:
            values = [sport for sport in ASSIGNMENT_CATEGORY_OPTIONS if sport_vars[sport].get()]
            self.inventory_sport_var.set(", ".join(values))

        for index, sport in enumerate(ASSIGNMENT_CATEGORY_OPTIONS):
            sport_key = (assignment_engine.canonical_sport_label(sport) or sport).strip().lower()
            var = tk.BooleanVar(value=sport_key in selected_sports)
            sport_vars[sport] = var
            ttk.Checkbutton(
                sport_frame,
                text=sport.title(),
                variable=var,
                command=sync_sport_filter,
                style="Panel.TCheckbutton",
            ).grid(row=index // 3, column=index % 3, sticky="w", padx=(0, 14), pady=(0, 4))

        ttk.Label(frame, text="Grader", style="AppMuted.TLabel").grid(row=2, column=0, sticky="w", pady=(0, 10))
        grader_frame = ttk.Frame(frame, style="App.TFrame")
        grader_frame.grid(row=2, column=1, columnspan=3, sticky="ew", pady=(0, 10))
        selected_graders = inventory_grader_filter_values(self.inventory_grader_var.get())
        grader_vars: dict[str, tk.BooleanVar] = {}

        def sync_grader_filter() -> None:
            values = [grader for grader in INVENTORY_GRADER_OPTIONS if grader_vars[grader].get()]
            self.inventory_grader_var.set(", ".join(values))

        for index, grader in enumerate(INVENTORY_GRADER_OPTIONS):
            var = tk.BooleanVar(value=grader in selected_graders)
            grader_vars[grader] = var
            ttk.Checkbutton(
                grader_frame,
                text=grader,
                variable=var,
                command=sync_grader_filter,
                style="Panel.TCheckbutton",
            ).grid(row=0, column=index, sticky="w", padx=(0, 14), pady=(0, 4))

        ttk.Label(frame, text="Card Year", style="AppMuted.TLabel").grid(row=3, column=0, sticky="w", pady=(0, 10))
        ttk.Entry(frame, textvariable=self.inventory_year_var, width=12).grid(row=3, column=1, sticky="w", pady=(0, 10))

        ttk.Label(frame, text="Price", style="AppMuted.TLabel").grid(row=4, column=0, sticky="w", pady=(0, 10))
        ttk.Entry(frame, textvariable=self.inventory_min_var, width=12).grid(row=4, column=1, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="to", style="AppMuted.TLabel").grid(row=4, column=2, sticky="w", padx=(8, 8), pady=(0, 10))
        ttk.Entry(frame, textvariable=self.inventory_max_var, width=12).grid(row=4, column=3, sticky="w", pady=(0, 10))

        ttk.Label(frame, text="Date Added", style="AppMuted.TLabel").grid(row=5, column=0, sticky="w", pady=(0, 10))
        self._inventory_date_picker(frame, self.inventory_date_min_var).grid(row=5, column=1, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="to", style="AppMuted.TLabel").grid(row=5, column=2, sticky="w", padx=(8, 8), pady=(0, 10))
        self._inventory_date_picker(frame, self.inventory_date_max_var).grid(row=5, column=3, sticky="w", pady=(0, 10))

        ttk.Checkbutton(
            frame,
            text="Missing Card Description Only",
            variable=self.inventory_missing_title_var,
            style="Panel.TCheckbutton",
        ).grid(row=6, column=0, columnspan=4, sticky="w", pady=(0, 10))

        ttk.Checkbutton(
            frame,
            text="Missing Comps Only",
            variable=self.inventory_missing_comps_var,
            style="Panel.TCheckbutton",
        ).grid(row=7, column=0, columnspan=4, sticky="w", pady=(0, 10))

        ttk.Checkbutton(
            frame,
            text="Missing CL Value Only",
            variable=self.inventory_missing_cl_var,
            style="Panel.TCheckbutton",
        ).grid(row=8, column=0, columnspan=4, sticky="w", pady=(0, 10))

        ttk.Checkbutton(
            frame,
            text="Missing Photos Only",
            variable=self.inventory_missing_photos_var,
            style="Panel.TCheckbutton",
        ).grid(row=9, column=0, columnspan=4, sticky="w", pady=(0, 10))

        actions = ttk.Frame(frame, style="App.TFrame")
        actions.grid(row=10, column=0, columnspan=4, sticky="ew", pady=(16, 0))
        actions.columnconfigure(1, weight=1)
        ttk.Button(actions, text="Clear Filters", command=self.clear_inventory_filters, style="Soft.TButton").grid(row=0, column=0, sticky="w")
        ttk.Button(actions, text="Close", command=popup.destroy, style="Soft.TButton").grid(row=0, column=2, sticky="e", padx=(0, 8))
        ttk.Button(actions, text="Apply Filters", command=self.refresh_inventory_tab, style="Primary.TButton").grid(row=0, column=3, sticky="e")
        frame.columnconfigure(3, weight=1)

    def _inventory_date_picker(self, parent, variable: tk.StringVar) -> ttk.Frame:
        picker = ttk.Frame(parent, style="App.TFrame")
        entry = ttk.Entry(picker, textvariable=variable, width=12, state="readonly")
        entry.grid(row=0, column=0, sticky="w")
        ttk.Button(
            picker,
            text="Pick",
            command=lambda: self._open_inventory_date_calendar(variable),
            style="Soft.TButton",
            width=6,
        ).grid(row=0, column=1, sticky="w", padx=(6, 0))
        return picker

    def _open_inventory_date_calendar(self, variable: tk.StringVar) -> None:
        selected_date = self._profit_record_date(variable.get())
        current = selected_date or datetime.now().date()
        month_state = {"year": current.year, "month": current.month}

        popup = tk.Toplevel(self)
        popup.title("Select Date Added")
        popup.configure(bg=self.colors["bg"])
        popup.transient(self)
        popup.grab_set()

        frame = ttk.Frame(popup, style="App.TFrame", padding=14)
        frame.pack(fill=tk.BOTH, expand=True)
        header = ttk.Frame(frame, style="App.TFrame")
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.columnconfigure(1, weight=1)
        month_label = tk.StringVar()
        days_frame = ttk.Frame(frame, style="App.TFrame")
        days_frame.grid(row=1, column=0, sticky="nsew")

        def set_month(delta: int) -> None:
            month = month_state["month"] + delta
            year = month_state["year"]
            if month < 1:
                month = 12
                year -= 1
            elif month > 12:
                month = 1
                year += 1
            month_state["year"] = year
            month_state["month"] = month
            render_days()

        def choose(year: int, month: int, day: int) -> None:
            variable.set(f"{year:04d}-{month:02d}-{day:02d}")
            popup.destroy()

        def render_days() -> None:
            for child in days_frame.winfo_children():
                child.destroy()
            year = month_state["year"]
            month = month_state["month"]
            month_label.set(f"{calendar.month_name[month]} {year}")
            for index, name in enumerate(("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")):
                ttk.Label(days_frame, text=name, style="AppMuted.TLabel", anchor="center").grid(row=0, column=index, sticky="ew", padx=2, pady=(0, 4))
                days_frame.columnconfigure(index, weight=1)
            selected = self._profit_record_date(variable.get())
            today = datetime.now().date()
            for week_index, week in enumerate(calendar.Calendar(firstweekday=0).monthdayscalendar(year, month), start=1):
                for day_index, day in enumerate(week):
                    if not day:
                        ttk.Label(days_frame, text="", style="AppMuted.TLabel", width=4).grid(row=week_index, column=day_index, padx=2, pady=2)
                        continue
                    button_style = "Primary.TButton" if selected and selected.year == year and selected.month == month and selected.day == day else "Soft.TButton"
                    if today.year == year and today.month == month and today.day == day and not (selected and selected == today):
                        button_style = "Primary.TButton"
                    ttk.Button(
                        days_frame,
                        text=str(day),
                        command=lambda d=day: choose(year, month, d),
                        style=button_style,
                        width=4,
                    ).grid(row=week_index, column=day_index, padx=2, pady=2, sticky="ew")

        ttk.Button(header, text="<", command=lambda: set_month(-1), style="Soft.TButton", width=4).grid(row=0, column=0, sticky="w")
        ttk.Label(header, textvariable=month_label, style="Panel.TLabel", anchor="center").grid(row=0, column=1, sticky="ew", padx=8)
        ttk.Button(header, text=">", command=lambda: set_month(1), style="Soft.TButton", width=4).grid(row=0, column=2, sticky="e")

        actions = ttk.Frame(frame, style="App.TFrame")
        actions.grid(row=2, column=0, sticky="ew", pady=(12, 0))
        actions.columnconfigure(1, weight=1)
        ttk.Button(actions, text="Clear", command=lambda: (variable.set(""), popup.destroy()), style="Soft.TButton").grid(row=0, column=0, sticky="w")
        ttk.Button(actions, text="Today", command=lambda: choose(datetime.now().year, datetime.now().month, datetime.now().day), style="Soft.TButton").grid(row=0, column=2, sticky="e", padx=(0, 8))
        ttk.Button(actions, text="Cancel", command=popup.destroy, style="Soft.TButton").grid(row=0, column=3, sticky="e")
        render_days()

    def clear_inventory_filters(self) -> None:
        for var in (self.inventory_sport_var, self.inventory_grader_var, self.inventory_year_var, self.inventory_min_var, self.inventory_max_var, self.inventory_date_min_var, self.inventory_date_max_var):
            var.set("")
        self.inventory_missing_title_var.set(False)
        self.inventory_missing_comps_var.set(False)
        self.inventory_missing_cl_var.set(False)
        self.inventory_missing_photos_var.set(False)
        self.refresh_inventory_tab()

    def _show_inventory_settings_menu(self, anchor: tk.Widget) -> None:
        menu = tk.Menu(self, tearoff=False, bg="#1f1f1f", fg="#ffffff", activebackground="#1ed760", activeforeground="#000000")
        menu.add_command(label="Sync Received to Inventory", command=lambda: self.refresh_inventory_tab(reconcile=True, enrich=True, filtered_only=True))
        menu.add_command(label="Update Best Company/Payouts", command=self.update_inventory_payouts)
        menu.add_command(label="Recomp Visible Cards", command=self.open_inventory_recomp_popup)
        menu.add_separator()
        menu.add_command(label="Photo Folder", command=self.choose_inventory_photo_folder)
        menu.add_command(label="Scan Photos", command=lambda: self.scan_inventory_photos(manual=True))
        menu.add_separator()
        menu.add_command(label="Import Mobile Queue", command=self.import_mobile_queue_file)
        try:
            menu.tk_popup(anchor.winfo_rootx(), anchor.winfo_rooty() + anchor.winfo_height())
        finally:
            menu.grab_release()

    def _build_profit_tab(self) -> None:
        controls = ttk.Frame(self.profit_tab, style="Panel.TFrame", padding=(14, 8))
        controls.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(controls, text="Profit", style="Panel.TLabel", font=("Segoe UI Semibold", 13)).grid(row=0, column=0, sticky="w")
        profit_person_label = ttk.Label(controls, text="Person", style="Muted.TLabel")
        if not self._is_personal_lucas():
            profit_person_label.grid(row=0, column=1, sticky="e", padx=(18, 6))
        self.profit_person_combo = ttk.Combobox(controls, textvariable=self.profit_person_var, width=28)
        if not self._is_personal_lucas():
            self.profit_person_combo.grid(row=0, column=2, sticky="w")
        self._bind_person_autocomplete(self.profit_person_combo, refresh_callback=self.refresh_profit_tab, allow_blank=True)
        self.profit_person_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_profit_tab(), add="+")
        period_label_column = 1 if self._is_personal_lucas() else 3
        period_combo_column = 2 if self._is_personal_lucas() else 4
        ttk.Label(controls, text="Period", style="Muted.TLabel").grid(row=0, column=period_label_column, sticky="e", padx=(18, 6))
        self.profit_period_combo = ttk.Combobox(
            controls,
            textvariable=self.profit_period_var,
            values=PROFIT_PERIOD_OPTIONS,
            width=10,
            state="readonly",
        )
        self.profit_period_combo.grid(row=0, column=period_combo_column, sticky="w")
        self.profit_period_combo.bind("<<ComboboxSelected>>", lambda _event: self.refresh_profit_tab(), add="+")
        ttk.Label(controls, text="Metric", style="Muted.TLabel").grid(row=0, column=5, sticky="e", padx=(18, 6))
        self.profit_graph_combo = ttk.Combobox(
            controls,
            textvariable=self.profit_graph_var,
            values=PROFIT_GRAPH_OPTIONS,
            width=20,
            state="readonly",
        )
        self.profit_graph_combo.grid(row=0, column=6, sticky="w")
        self.profit_graph_combo.bind("<<ComboboxSelected>>", lambda _event: self._draw_profit_chart(), add="+")
        ttk.Label(controls, text="Plot", style="Muted.TLabel").grid(row=0, column=7, sticky="e", padx=(18, 6))
        self.profit_plot_combo = ttk.Combobox(
            controls,
            textvariable=self.profit_plot_var,
            values=PROFIT_PLOT_OPTIONS,
            width=10,
            state="readonly",
        )
        self.profit_plot_combo.grid(row=0, column=8, sticky="w")
        self.profit_plot_combo.bind("<<ComboboxSelected>>", lambda _event: self._draw_profit_chart(), add="+")
        ttk.Button(controls, text="Refresh View", command=self.refresh_profit_tab, style="Soft.TButton").grid(row=0, column=9, sticky="w", padx=(10, 0))
        ttk.Button(controls, text="Add Expense", command=self.open_add_expense_popup, style="Soft.TButton").grid(row=0, column=10, sticky="w", padx=(8, 0))
        controls.columnconfigure(11, weight=1)
        self.profit_search_var.trace_add("write", lambda *_args: self.refresh_profit_tab())
        ttk.Label(controls, textvariable=self.profit_metric_var, style="Panel.TLabel").grid(row=1, column=0, columnspan=12, sticky="w", pady=(8, 0))
        ttk.Label(controls, textvariable=self.profit_status_var, style="Muted.TLabel").grid(row=2, column=0, columnspan=12, sticky="w", pady=(4, 0))

        profit_split = tk.PanedWindow(
            self.profit_tab,
            orient=tk.VERTICAL,
            bg=self.app_palette["border"],
            bd=0,
            sashwidth=8,
            sashrelief=tk.RAISED,
            showhandle=True,
            handlesize=28,
            opaqueresize=True,
        )
        profit_split.pack(fill=tk.BOTH, expand=True)

        chart_panel = ttk.Frame(profit_split, style="Panel.TFrame", padding=(10, 8))
        ttk.Label(chart_panel, textvariable=self.profit_chart_title_var, style="Panel.TLabel").pack(anchor=tk.W)
        self.profit_chart_canvas = tk.Canvas(
            chart_panel,
            height=210,
            bg="#1f1f1f",
            highlightthickness=1,
            highlightbackground="#333333",
        )
        self.profit_chart_canvas.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        self.profit_chart_canvas.bind("<Configure>", lambda _event: self._draw_profit_chart())
        self.profit_chart_canvas.bind("<Motion>", self._show_profit_chart_tooltip)
        self.profit_chart_canvas.bind("<Leave>", lambda _event: self._hide_profit_chart_tooltip())
        profit_split.add(chart_panel, minsize=120)

        ledger_panel = ttk.Frame(profit_split, style="Panel.TFrame", padding=(12, 10))
        profit_split.add(ledger_panel, minsize=220)
        toolbar = ttk.Frame(ledger_panel, style="Panel.TFrame")
        toolbar.pack(fill=tk.X, pady=(0, 8))
        self.profit_cards_button = ttk.Button(toolbar, text="Sold Cards", command=lambda: self._set_profit_view_mode("Sold Cards"), style="Soft.TButton")
        self.profit_cards_button.pack(side=tk.LEFT)
        self.profit_sheets_button = ttk.Button(toolbar, text="Sold Sheets", command=lambda: self._set_profit_view_mode("Sold Sheets"), style="Soft.TButton")
        self.profit_sheets_button.pack(side=tk.LEFT, padx=(8, 0))
        self.profit_expenses_button = ttk.Button(toolbar, text="Expenses", command=lambda: self._set_profit_view_mode("Expenses"), style="Soft.TButton")
        self.profit_expenses_button.pack(side=tk.LEFT, padx=(8, 0))
        ttk.Label(toolbar, text="Search", style="Muted.TLabel").pack(side=tk.LEFT, padx=(18, 6))
        ttk.Entry(toolbar, textvariable=self.profit_search_var, width=36).pack(side=tk.LEFT)
        self.profit_table_title_var = tk.StringVar(value="Sold Cards")
        ttk.Label(toolbar, textvariable=self.profit_table_title_var, style="Panel.TLabel").pack(side=tk.RIGHT)
        self.profit_tree = self._build_home_tree(
            ledger_panel,
            columns=("date", "company", "card", "cert", "purchase", "sale", "profit", "sheet"),
            headings={
                "date": "Date",
                "company": "Company",
                "card": "Card",
                "cert": "Cert",
                "purchase": "Purchase",
                "sale": "Sale Price",
                "profit": "Profit",
                "sheet": "Company Sheet",
            },
            widths={"date": 95, "company": 150, "card": 440, "cert": 110, "purchase": 105, "sale": 105, "profit": 105, "sheet": 220},
            height=28,
            max_height=520,
        )
        self.profit_tree.tag_configure("profit_positive", foreground="#d7fbe8")
        self.profit_tree.tag_configure("profit_negative", foreground="#ffd1d1")
        self.profit_tree.tag_configure("total_row", background="#242424", foreground="#ffffff", font=("Segoe UI Semibold", 10))
        self._configure_sortable_tree_headings(self.profit_tree, {
            "date": "Date",
            "person": "Person",
            "company": "Company",
            "card": "Card",
            "cert": "Cert / Item ID",
            "purchase": "Purchase",
            "sale": "Sale Price",
            "profit": "Profit",
            "sheet": "Company Sheet",
        }, "profit")
        self._bind_context_menu(self.profit_tree, self._show_profit_context_menu)

    def _load_profit_ledger(self) -> list[dict[str, object]]:
        if not PROFIT_LEDGER_PATH.exists():
            return []
        try:
            raw = json.loads(PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
        except Exception:
            return []
        if not isinstance(raw, list):
            return []
        return [item for item in raw if isinstance(item, dict)]

    def _save_profit_ledger(self, rows: list[dict[str, object]]) -> None:
        atomic_write_json(PROFIT_LEDGER_PATH, rows)

    def _load_inventory_ledger(self) -> list[dict[str, object]]:
        if not INVENTORY_LEDGER_PATH.exists():
            return []
        try:
            raw = json.loads(INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))
        except Exception:
            return []
        entries = raw.get("items", raw) if isinstance(raw, dict) else raw
        return [item for item in entries if isinstance(item, dict)] if isinstance(entries, list) else []

    def _save_inventory_ledger(self, rows: list[dict[str, object]]) -> None:
        INVENTORY_LEDGER_PATH.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_json(INVENTORY_LEDGER_PATH, {"items": rows})

    def _load_inventory_deleted_tombstones(self) -> list[dict[str, object]]:
        if not INVENTORY_DELETED_TOMBSTONES_PATH.exists():
            return []
        try:
            raw = json.loads(INVENTORY_DELETED_TOMBSTONES_PATH.read_text(encoding="utf-8"))
        except Exception:
            return []
        entries = raw.get("items", raw) if isinstance(raw, dict) else raw
        return [item for item in entries if isinstance(item, dict)] if isinstance(entries, list) else []

    def _save_inventory_deleted_tombstones(self, rows: list[dict[str, object]]) -> None:
        INVENTORY_DELETED_TOMBSTONES_PATH.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_json(INVENTORY_DELETED_TOMBSTONES_PATH, {"items": rows[-5000:]})

    def _inventory_deleted_source_cert_keys(self) -> set[tuple[str, str]]:
        keys: set[tuple[str, str]] = set()
        for record in self._load_inventory_deleted_tombstones():
            source_sheet = Path(str(record.get("source_sheet") or "")).name.strip().lower()
            cert = scan_to_cert(record.get("cert_number"))
            if source_sheet and cert:
                keys.add((source_sheet, cert))
        return keys

    def _record_inventory_deleted_tombstones(self, records: list[dict[str, object]], reason: str = "inventory_delete") -> int:
        existing = self._load_inventory_deleted_tombstones()
        existing_keys = {
            (
                Path(str(record.get("source_sheet") or "")).name.strip().lower(),
                scan_to_cert(record.get("cert_number")),
            )
            for record in existing
        }
        added = 0
        deleted_at = datetime.now().isoformat(timespec="seconds")
        for raw_record in records:
            record = self._normalize_inventory_record(raw_record)
            source_sheet = Path(str(record.get("source_sheet") or "")).name.strip()
            cert = scan_to_cert(record.get("cert_number"))
            if not source_sheet or not cert:
                continue
            key = (source_sheet.lower(), cert)
            if key in existing_keys:
                continue
            existing.append(
                {
                    "source_sheet": source_sheet,
                    "cert_number": cert,
                    "inventory_key": str(record.get("inventory_key") or ""),
                    "card_title": str(record.get("card_title") or ""),
                    "assigned_person": str(record.get("assigned_person") or ""),
                    "deleted_at": deleted_at,
                    "reason": reason,
                }
            )
            existing_keys.add(key)
            added += 1
        if added:
            self._save_inventory_deleted_tombstones(existing)
        return added

    def _load_activity_log(self) -> list[dict[str, object]]:
        if not ACTIVITY_LOG_PATH.exists():
            return []
        try:
            raw = json.loads(ACTIVITY_LOG_PATH.read_text(encoding="utf-8"))
        except Exception:
            return []
        entries = raw.get("entries", raw) if isinstance(raw, dict) else raw
        return [entry for entry in entries if isinstance(entry, dict)] if isinstance(entries, list) else []

    def _save_activity_log(self, entries: list[dict[str, object]]) -> None:
        ACTIVITY_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_json(ACTIVITY_LOG_PATH, {"entries": entries[-300:]})

    def _append_activity(self, action: str, summary: str, details: dict[str, object] | None = None) -> None:
        entry = {
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "action": str(action or "").strip() or "Activity",
            "summary": str(summary or "").strip(),
            "user": self.lucas_identity.get("display_name") or "",
            "machine": self.lucas_identity.get("machine") or "",
            "details": details or {},
        }
        try:
            with shared_lock(CARD_PIPELINE_DIR, "activity-log", self.lucas_identity, timeout=8):
                entries = self._load_activity_log()
                entries.append(entry)
                self._save_activity_log(entries)
        except Exception:
            record_performance_event("activity.log_failed", time.perf_counter(), f"action={entry['action']}", force=True)

    def open_activity_log(self) -> None:
        entries = list(reversed(self._load_activity_log()[-100:]))
        popup = tk.Toplevel(self)
        popup.title("LUCAS Activity Log")
        popup.geometry("980x520")
        popup.transient(self)
        popup.configure(bg="#121212")
        frame = ttk.Frame(popup, style="App.TFrame", padding=14)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Activity Log", style="HeaderTitle.TLabel").pack(anchor=tk.W, pady=(0, 10))
        tree = ttk.Treeview(frame, columns=("time", "action", "summary", "user"), show="headings", height=16)
        for column, label, width in (("time", "Time", 150), ("action", "Action", 150), ("summary", "Summary", 500), ("user", "User", 130)):
            tree.heading(column, text=label, anchor=tk.W)
            tree.column(column, width=width, anchor=tk.W, stretch=column == "summary")
        tree.pack(fill=tk.BOTH, expand=True)
        for entry in entries:
            tree.insert(
                "",
                tk.END,
                values=(entry.get("timestamp") or "", entry.get("action") or "", entry.get("summary") or "", entry.get("user") or ""),
            )
        actions = ttk.Frame(frame, style="App.TFrame")
        actions.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(actions, text="Refresh", command=lambda: (popup.destroy(), self.open_activity_log()), style="Soft.TButton").pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(actions, text="Close", command=popup.destroy, style="Soft.TButton").pack(side=tk.RIGHT)

    def _show_error_with_copy(self, title: str, message: str, details: dict[str, object] | str | None = None) -> None:
        if isinstance(details, str):
            detail_text = details
        elif details:
            try:
                detail_text = json.dumps(details, indent=2, sort_keys=True, default=str)
            except TypeError:
                detail_text = str(details)
        else:
            detail_text = ""
        copy_text = "\n\n".join(part for part in (message, detail_text) if part)
        popup = tk.Toplevel(self)
        popup.title(title)
        popup.configure(bg=self.colors["bg"])
        popup.transient(self)
        popup.geometry("720x420")
        frame = ttk.Frame(popup, style="App.TFrame", padding=16)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=title, style="Panel.TLabel", font=("Segoe UI Semibold", 13)).pack(anchor=tk.W, pady=(0, 8))
        ttk.Label(frame, text=message, style="Muted.TLabel", wraplength=660).pack(anchor=tk.W, pady=(0, 10))
        if detail_text:
            text = tk.Text(frame, bg="#111111", fg="#f5f5f5", insertbackground="#ffffff", relief=tk.FLAT, wrap=tk.WORD, height=12)
            text.pack(fill=tk.BOTH, expand=True)
            text.insert("1.0", detail_text)
            text.configure(state=tk.DISABLED)
        actions = ttk.Frame(frame, style="App.TFrame")
        actions.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(actions, text="Copy Details", command=lambda: self._copy_inventory_text(copy_text, "error details"), style="Soft.TButton").pack(side=tk.LEFT)
        ttk.Button(actions, text="Close", command=popup.destroy, style="Primary.TButton").pack(side=tk.RIGHT)

    def _inventory_record_key(self, record: dict[str, object]) -> str:
        item_id = str(record.get("item_id") or "").strip()
        if item_id:
            return item_id.lower()
        return "|".join(
            str(record.get(field) or "").strip().lower()
            for field in ("cert_number", "source_sheet", "assigned_person")
        )

    def _raw_item_id_namespace(self) -> str:
        return "MIKEY" if self._is_personal_lucas() else "TEAM"

    def _next_raw_item_id(self, existing_records: list[dict[str, object]] | None = None) -> str:
        today = datetime.now().strftime("%Y%m%d")
        prefix = f"RAW-{self._raw_item_id_namespace()}-{today}-"
        records = existing_records if existing_records is not None else self._load_inventory_ledger()
        max_sequence = 0
        for record in records:
            item_id = str(record.get("item_id") or "").strip().upper()
            if not item_id.startswith(prefix):
                continue
            suffix = item_id[len(prefix):]
            if suffix.isdigit():
                max_sequence = max(max_sequence, int(suffix))
        return f"{prefix}{max_sequence + 1:04d}"

    def _live_sheet_raw_item_records(self) -> list[dict[str, object]]:
        records: list[dict[str, object]] = []
        for folder in (INCOMING_SHEETS_DIR, WORKING_SHEETS_DIR, RECEIVED_SHEETS_DIR):
            try:
                paths = sorted(folder.glob("*.xlsx"), key=lambda path: path.name.lower())
            except Exception:
                continue
            for path in paths:
                try:
                    rows = read_simple_spreadsheet(path)
                except Exception:
                    continue
                for row in rows:
                    item_id = str(row.get("item_id") or "").strip()
                    if item_id.upper().startswith("RAW-"):
                        records.append({"item_id": item_id})
        return records

    def _ensure_raw_item_ids_for_rows(self, rows: list[WorkbookRow]) -> int:
        existing_records = list(self._load_inventory_ledger())
        if hasattr(self, "_live_sheet_raw_item_records"):
            existing_records.extend(self._live_sheet_raw_item_records())
        for row in rows:
            item_id = str(getattr(row, "item_id", "") or "").strip()
            if item_id:
                existing_records.append({"item_id": item_id})
        added = 0
        for row in rows:
            cert = str(getattr(row, "cert_number", "") or "").strip()
            item_id = str(getattr(row, "item_id", "") or "").strip()
            has_row_data = any(
                str(value or "").strip()
                for value in (
                    getattr(row, "card_title", ""),
                    getattr(row, "grader", ""),
                    getattr(row, "category", ""),
                    getattr(row, "existing_value", ""),
                    getattr(row, "card_ladder_value", ""),
                    getattr(row, "card_ladder_comps_average", ""),
                    getattr(row, "cy_value", ""),
                )
            )
            if cert or item_id or not has_row_data:
                continue
            item_id = self._next_raw_item_id(existing_records)
            setattr(row, "item_id", item_id)
            existing_records.append({"item_id": item_id})
            added += 1
        return added

    def _workbook_header_lookup(self, sheet) -> dict[str, int]:
        headers: dict[str, int] = {}
        for cell in sheet[1]:
            value = re.sub(r"[^a-z0-9]+", "", str(cell.value or "").strip().lower())
            if value:
                headers[value] = int(cell.column)
        return headers

    def _sheet_cell_by_any_header(self, sheet, row_index: int, headers: dict[str, int], names: tuple[str, ...], fallback_col: int | None = None):
        for name in names:
            col = headers.get(re.sub(r"[^a-z0-9]+", "", name.strip().lower()))
            if col:
                return sheet.cell(row_index, col).value
        if fallback_col:
            return sheet.cell(row_index, fallback_col).value
        return None

    def _ensure_workbook_item_id_column(self, sheet, headers: dict[str, int]) -> tuple[dict[str, int], int]:
        for name in ("itemid", "rawitemid", "inventoryitemid", "inventoryid"):
            if name in headers:
                return headers, headers[name]
        sheet.insert_cols(1)
        sheet.cell(1, 1).value = "Item ID"
        headers = self._workbook_header_lookup(sheet)
        return headers, 1

    def _ensure_raw_item_ids_in_sheet_paths(self, paths: list[Path]) -> dict[str, object]:
        existing_records = list(self._load_inventory_ledger())
        existing_records.extend(self._live_sheet_raw_item_records())
        result: dict[str, object] = {"files_updated": 0, "ids_added": 0, "errors": []}
        changed_paths: set[Path] = set()
        seen_raw_ids: set[str] = set()
        for path in paths:
            try:
                workbook = load_workbook(path)
            except Exception as error:
                result["errors"].append(f"{path.name}: {error}")  # type: ignore[index]
                continue
            changed = False
            try:
                for sheet in workbook.worksheets:
                    if sheet.max_row < 2:
                        continue
                    headers = self._workbook_header_lookup(sheet)
                    headers, item_id_col = self._ensure_workbook_item_id_column(sheet, headers)
                    cert_col = next((headers.get(name) for name in ("certificationnumber", "certnumber", "cert", "certification")), None)
                    card_col = next((headers.get(name) for name in ("carddescription", "card", "cardtitle", "title", "description")), None)
                    grader_col = next((headers.get(name) for name in ("company", "grader", "gradingcompany")), None)
                    sport_col = next((headers.get(name) for name in ("sport", "category")), None)
                    purchase_col = next((headers.get(name) for name in ("purchaseprice", "purchase", "cost", "buyprice")), None)
                    for row_index in range(2, sheet.max_row + 1):
                        item_id = str(sheet.cell(row_index, item_id_col).value or "").strip()
                        cert = scan_to_cert(sheet.cell(row_index, cert_col).value if cert_col else "")
                        card = str((sheet.cell(row_index, card_col).value if card_col else "") or "").strip()
                        if item_id:
                            item_key = item_id.upper()
                            if item_key.startswith(f"RAW-{self._raw_item_id_namespace()}-") and not cert and not card:
                                sheet.cell(row_index, item_id_col).value = None
                                changed = True
                                continue
                            if item_key.startswith(f"RAW-{self._raw_item_id_namespace()}-") and item_key in seen_raw_ids:
                                new_item_id = self._next_raw_item_id(existing_records)
                                sheet.cell(row_index, item_id_col).value = new_item_id
                                existing_records.append({"item_id": new_item_id})
                                seen_raw_ids.add(new_item_id.upper())
                                result["ids_added"] = int(result["ids_added"]) + 1
                                changed = True
                                continue
                            seen_raw_ids.add(item_key)
                            existing_records.append({"item_id": item_id})
                            continue
                        if cert:
                            continue
                        grader = str((sheet.cell(row_index, grader_col).value if grader_col else "") or "").strip()
                        sport = str((sheet.cell(row_index, sport_col).value if sport_col else "") or "").strip()
                        purchase = sheet.cell(row_index, purchase_col).value if purchase_col else None
                        if not card:
                            continue
                        new_item_id = self._next_raw_item_id(existing_records)
                        sheet.cell(row_index, item_id_col).value = new_item_id
                        existing_records.append({"item_id": new_item_id})
                        result["ids_added"] = int(result["ids_added"]) + 1
                        changed = True
                if changed:
                    workbook.save(path)
                    changed_paths.add(path)
            except Exception as error:
                result["errors"].append(f"{path.name}: {error}")  # type: ignore[index]
            finally:
                workbook.close()
        result["files_updated"] = len(changed_paths)
        return result

    def ensure_raw_item_ids_in_stage_sheets(self, include_received: bool = False) -> dict[str, object]:
        folders = [INCOMING_SHEETS_DIR, WORKING_SHEETS_DIR]
        if include_received:
            folders.append(RECEIVED_SHEETS_DIR)
        paths: list[Path] = []
        for folder in folders:
            try:
                folder.mkdir(parents=True, exist_ok=True)
                paths.extend(sorted(folder.glob("*.xlsx"), key=lambda path: path.name.lower()))
            except Exception:
                continue
        return self._ensure_raw_item_ids_in_sheet_paths(paths)

    def _normalize_inventory_record(self, record: dict[str, object]) -> dict[str, object]:
        normalized = dict(record)
        normalized["date_added"] = str(normalized.get("date_added") or datetime.now().strftime("%Y-%m-%d"))[:10]
        item_type = str(normalized.get("item_type") or normalized.get("type") or "").strip().title()
        if item_type not in {"Raw", "Graded"}:
            item_type = "Raw" if str(normalized.get("item_id") or "").strip().upper().startswith("RAW-") else "Graded"
        normalized["item_type"] = item_type
        normalized["item_id"] = str(normalized.get("item_id") or "").strip()
        owner_for_profile = getattr(self, "_owner_for_profile", lambda person="": str(person or "").strip() or "Unassigned")
        normalized["assigned_person"] = owner_for_profile(normalized.get("assigned_person") or normalized.get("person"))
        sport = str(normalized.get("sport") or normalized.get("category") or "").strip()
        normalized["sport"] = assignment_engine.canonical_sport_label(sport) or sport
        normalized["cert_number"] = str(normalized.get("cert_number") or "").strip()
        normalized["grader"] = str(normalized.get("grader") or "").strip()
        normalized["card_title"] = str(normalized.get("card_title") or "").strip()
        if not normalized["sport"] and normalized["card_title"]:
            normalized["sport"] = CardPipelineApp._inventory_sport_from_value(self, "", normalized["card_title"])
        normalized["purchase_price"] = self._money_value(normalized.get("purchase_price"))
        normalized["paid_with"] = str(normalized.get("paid_with") or normalized.get("payment_method") or "").strip()
        normalized["card_ladder_value"] = self._money_value(normalized.get("card_ladder_value"))
        normalized["card_ladder_comps_average"] = self._money_value(normalized.get("card_ladder_comps_average") or normalized.get("comps"))
        normalized["cy_value"] = self._money_value(normalized.get("cy_value") or normalized.get("cy_estimate"))
        cy_confidence = normalized.get("cy_confidence")
        normalized["cy_confidence"] = cy_confidence.strip() if isinstance(cy_confidence, str) else "" if cy_confidence is None else cy_confidence
        normalized["inventory_value"] = self._money_value(normalized.get("inventory_value") or normalized.get("value") or normalized.get("sale_price") or normalized.get("estimated_payout"))
        normalized["best_company"] = str(normalized.get("best_company") or normalized.get("company") or "").strip()
        normalized["estimated_payout"] = self._money_value(normalized.get("estimated_payout") or normalized.get("payout"))
        normalized["source_sheet"] = str(normalized.get("source_sheet") or "").strip()
        normalized["source"] = str(normalized.get("source") or "").strip()
        normalized["status"] = str(normalized.get("status") or "Active").strip() or "Active"
        photo_paths = normalized.get("photo_paths") or normalized.get("photos") or []
        if isinstance(photo_paths, str):
            photo_paths = [part.strip() for part in re.split(r"[;\n]", photo_paths) if part.strip()]
        elif not isinstance(photo_paths, list):
            photo_paths = []
        normalized["photo_paths"] = [str(path).strip() for path in photo_paths if str(path or "").strip()]
        normalized["photo_count"] = len(normalized["photo_paths"])
        normalized["notes"] = str(normalized.get("notes") or "").strip()
        normalized["inventory_key"] = str(normalized.get("inventory_key") or self._inventory_record_key(normalized))
        return normalized

    def _inventory_sport_from_value(self, sport_value: object = "", card_title: object = "") -> str:
        sport = str(sport_value or "").strip()
        if sport:
            return assignment_engine.canonical_sport_label(sport) or sport
        title = str(card_title or "").strip()
        if not title:
            return ""
        return str(assignment_engine.parse_card_for_matching(title).get("sport") or "")

    def _inventory_record_from_row(self, row: WorkbookRow, person: str, source_sheet: str = "", source: str = "", status: str = "Active", notes: str = "") -> dict[str, object]:
        card_title = str(row.card_title or "")
        sport = CardPipelineApp._inventory_sport_from_value(self, getattr(row, "category", ""), card_title)
        cert = str(row.cert_number or "").strip()
        return self._normalize_inventory_record(
            {
                "date_added": datetime.now().strftime("%Y-%m-%d"),
                "item_type": "Graded" if cert else "Raw",
                "item_id": "" if cert else str(getattr(row, "item_id", "") or "").strip(),
                "assigned_person": person or "Unassigned",
                "sport": sport,
                "cert_number": cert,
                "grader": row.grader,
                "card_title": row.card_title,
                "purchase_price": row.existing_value,
                "card_ladder_value": row.card_ladder_value,
                "card_ladder_comps_average": row.card_ladder_comps_average,
                "cy_value": row.cy_value,
                "cy_confidence": getattr(row, "cy_confidence", None),
                "inventory_value": row.card_ladder_comps_average or row.card_ladder_value or row.cy_value,
                "best_company": row.best_company,
                "estimated_payout": row.estimated_payout,
                "source_sheet": source_sheet,
                "source": source,
                "status": status,
                "notes": notes if str(notes or "").strip().lower() not in AUTO_INVENTORY_NOTES else "",
            }
        )

    def add_inventory_records(self, records: list[dict[str, object]], refresh: bool = True) -> int:
        if not records:
            return 0
        ledger = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        by_key = {str(record.get("inventory_key") or ""): record for record in ledger}
        added = 0
        for record in records:
            if not str(record.get("cert_number") or "").strip() and not str(record.get("item_id") or "").strip():
                record = dict(record)
                record["item_type"] = "Raw"
                record["item_id"] = self._next_raw_item_id([*ledger, *by_key.values()])
            normalized = self._normalize_inventory_record(record)
            normalized = self._enrich_inventory_record_assignment(normalized)
            key = str(normalized.get("inventory_key") or "")
            if not key:
                continue
            if key not in by_key:
                ledger.append(normalized)
                by_key[key] = normalized
                added += 1
            else:
                existing = by_key[key]
                existing.update(normalized)
                existing["status"] = "Active"
        self._save_inventory_ledger(ledger)
        if refresh:
            self.refresh_inventory_tab()
        return added

    def _mobile_app_url(self) -> str:
        profile = "personal" if self._is_personal_lucas() else "team"
        public_url = mobile_public_app_url(profile, getattr(self, "app_settings", {}))
        if public_url:
            return public_url
        host = mobile_app_host(getattr(self, "app_settings", {}))
        return f"http://{host}:{self.bridge.port}/mobile/{profile}"

    def _mobile_local_app_url(self) -> str:
        host = mobile_app_host(getattr(self, "app_settings", {}))
        profile = "personal" if self._is_personal_lucas() else "team"
        return f"http://{host}:{self.bridge.port}/mobile/{profile}"

    def open_mobile_connection_helper(self) -> None:
        url = self._mobile_app_url()
        local_url = self._mobile_local_app_url()
        profile = "personal" if self._is_personal_lucas() else "team"
        public_url = mobile_public_app_url(profile, getattr(self, "app_settings", {}))
        details = "\n".join(
            [
                f"Mobile URL: {url}",
                f"Local Wi-Fi URL: {local_url}",
                f"PIN: {self.mobile_pin}",
                "",
                "Live reads and syncing need a reachable desktop LUCAS bridge.",
                "For LTE/offline Home Screen use, install from a stable HTTPS public URL set in LUCAS_MOBILE_PUBLIC_URL.",
                "A local http:// Mac URL may white-screen on LTE because iPhone cannot reach that origin or reliably boot its cache.",
                "" if public_url else f"No HTTPS public mobile URL is configured for {profile} LUCAS.",
                "Open the Mobile URL in Safari on the iPhone, then enter the PIN.",
                "Offline adds, expenses, and cached-card sales stay in the phone Sync queue until desktop LUCAS is reachable.",
            ]
        )
        popup = tk.Toplevel(self)
        popup.title("Mobile Connection")
        popup.configure(bg=self.colors["bg"])
        popup.transient(self)
        popup.geometry("640x300")
        frame = ttk.Frame(popup, style="App.TFrame", padding=18)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Mobile Connection", style="Panel.TLabel", font=("Segoe UI Semibold", 13)).pack(anchor=tk.W, pady=(0, 10))
        ttk.Label(frame, text=details, style="Muted.TLabel", wraplength=580, justify=tk.LEFT).pack(anchor=tk.W, fill=tk.X)
        actions = ttk.Frame(frame, style="App.TFrame")
        actions.pack(fill=tk.X, pady=(16, 0))
        ttk.Button(actions, text="Copy Details", command=lambda: self._copy_inventory_text(details, "mobile connection details"), style="Soft.TButton").pack(side=tk.LEFT)
        ttk.Button(actions, text="Close", command=popup.destroy, style="Primary.TButton").pack(side=tk.RIGHT)

    def _mobile_inventory_payload_record(self, payload: dict) -> dict[str, object]:
        cert = scan_to_cert(payload.get("cert_number") or payload.get("cert") or payload.get("barcode") or "")
        card_title = str(payload.get("card_title") or payload.get("card") or "").strip()
        grader = normalize_grader(payload.get("grader") or "") or infer_grader(card_title)
        source = str(payload.get("source") or payload.get("seller") or "Mobile").strip() or "Mobile"
        person = str(payload.get("assigned_person") or payload.get("person") or "").strip()
        is_personal = getattr(self, "_is_personal_lucas", lambda: False)
        default_person = getattr(self, "_personal_default_person", lambda: "Mikey")
        if is_personal() and not person:
            person = default_person()
        elif not is_personal():
            person = self._canonical_person_choice(person) or ""
        sport = str(payload.get("sport") or "").strip()
        if not sport and card_title:
            sport = str(assignment_engine.parse_card_for_matching(card_title).get("sport") or "")
        notes = str(payload.get("notes") or "").strip()
        item_type = "Graded" if cert and grader else "Raw"
        if item_type == "Raw" and cert:
            notes = "\n".join(part for part in [notes, f"Mobile entered cert/item: {cert}"] if part).strip()
            cert = ""
        card_title = self._inventory_title_with_grader(card_title, grader)
        return self._normalize_inventory_record(
            {
                "date_added": datetime.now().strftime("%Y-%m-%d"),
                "assigned_person": person,
                "sport": sport,
                "cert_number": cert,
                "grader": grader,
                "card_title": card_title,
                "item_type": item_type,
                "purchase_price": payload.get("purchase_price") or payload.get("purchase") or payload.get("price_paid"),
                "inventory_value": payload.get("inventory_value") or payload.get("value"),
                "source_sheet": str(payload.get("source_sheet") or "Mobile Inventory").strip() or "Mobile Inventory",
                "source": source,
                "status": "Active",
                "notes": notes,
            }
        )

    def _inventory_title_with_grader(self, card_title: object, grader: object) -> str:
        title = re.sub(r"\s+", " ", str(card_title or "")).strip()
        normalized_grader = normalize_grader(grader or "")
        if not title or not normalized_grader:
            return title
        if re.search(rf"\b{re.escape(normalized_grader)}\b", title, flags=re.IGNORECASE):
            return title
        grade_match = re.search(r"\b(10|9(?:\.5)?|8(?:\.5)?|7(?:\.5)?|6(?:\.5)?|5(?:\.5)?|4(?:\.5)?|3(?:\.5)?|2(?:\.5)?|1(?:\.5)?)\s*$", title)
        if grade_match:
            prefix = title[: grade_match.start()].rstrip()
            return f"{prefix} {normalized_grader} {grade_match.group(1)}".strip()
        return f"{title} {normalized_grader}".strip()

    def _mobile_inventory_json_record(self, record: dict[str, object]) -> dict[str, object]:
        normalized = self._normalize_inventory_record(record)
        photo_items = getattr(self, "_mobile_inventory_photo_items", None)
        photos = photo_items(normalized) if callable(photo_items) else []
        return {
            "inventory_key": normalized.get("inventory_key"),
            "item_type": normalized.get("item_type"),
            "item_id": normalized.get("item_id"),
            "date_added": normalized.get("date_added"),
            "assigned_person": normalized.get("assigned_person"),
            "sport": normalized.get("sport"),
            "cert_number": normalized.get("cert_number"),
            "grader": normalized.get("grader"),
            "card_title": normalized.get("card_title"),
            "purchase_price": normalized.get("purchase_price"),
            "purchase_price_display": format_money(self._money_value(normalized.get("purchase_price"))),
            "inventory_value": normalized.get("inventory_value"),
            "inventory_value_display": format_money(self._money_value(normalized.get("inventory_value"))),
            "best_company": normalized.get("best_company"),
            "estimated_payout": normalized.get("estimated_payout"),
            "estimated_payout_display": format_money(self._money_value(normalized.get("estimated_payout"))),
            "source_sheet": normalized.get("source_sheet"),
            "source": normalized.get("source"),
            "status": normalized.get("status"),
            "notes": normalized.get("notes"),
            "photo_count": len(photos),
            "photos": photos,
        }

    def _mobile_inventory_photo_items(self, record: dict[str, object]) -> list[dict[str, object]]:
        bridge_state = getattr(self, "state", None)
        mobile_path = getattr(bridge_state, "mobile_inventory_photo_path", None)
        items: list[dict[str, object]] = []
        for value in list(record.get("photo_paths") or [])[:MAX_INVENTORY_PHOTOS_PER_CARD]:
            path = self._safe_inventory_photo_path(value)
            if path is None or not path.is_file():
                continue
            photo_id = self._inventory_photo_encoded_id(path)
            url_path = mobile_path(photo_id, path.name) if callable(mobile_path) else f"/mobile/api/inventory/photo/{photo_id}/{path.name}"
            items.append(
                {
                    "id": photo_id,
                    "name": path.name,
                    "url": f"{url_path}?pin={urllib.parse.quote(str(self.mobile_pin or ''))}",
                }
            )
        return items

    def mobile_inventory_search(self, payload: dict) -> dict:
        query = str(payload.get("query") or payload.get("q") or "").strip().lower()
        cert_query = scan_to_cert(query)
        person = str(payload.get("person") or "").strip().lower()
        sport_filters = self._mobile_inventory_sport_filters(payload)
        include_sold = bool(payload.get("include_sold"))
        try:
            limit = int(payload.get("limit") or 75)
        except (TypeError, ValueError):
            limit = 75
        limit = max(1, min(limit, 1000))
        rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        results: list[dict[str, object]] = []
        for record in rows:
            status = str(record.get("status") or "").lower()
            if status != "active" and not include_sold:
                continue
            if person and person not in str(record.get("assigned_person") or "Unassigned").lower():
                continue
            if sport_filters:
                record_sport_text = str(record.get("sport") or "").strip().lower()
                record_sport = (assignment_engine.canonical_sport_label(record_sport_text) or record_sport_text).strip().lower()
                if record_sport not in sport_filters and not any(sport in record_sport_text for sport in sport_filters):
                    continue
            if query:
                record_cert = str(record.get("cert_number") or "").strip().lower()
                record_cert_digits = scan_to_cert(record_cert)
                record_title = str(record.get("card_title") or "").strip().lower()
                cert_matches = bool(
                    (cert_query and cert_query in record_cert_digits)
                    or (query and query in record_cert)
                )
                title_matches = query in record_title
                if not (cert_matches or title_matches):
                    continue
            results.append(self._mobile_inventory_json_record(record))
            if len(results) >= limit:
                break
        return {"ok": True, "count": len(results), "items": results, "people": self._known_people()}

    def _mobile_inventory_sport_filters(self, payload: dict) -> set[str]:
        raw = payload.get("sport") or payload.get("category") or ""
        if isinstance(raw, list):
            parts = [str(value or "") for value in raw]
        else:
            parts = re.split(r"[,;/|]", str(raw or ""))
        values: set[str] = set()
        for part in parts:
            text = part.strip().lower()
            if not text:
                continue
            values.add((assignment_engine.canonical_sport_label(text) or text).strip().lower())
        return values

    def mobile_inventory_add(self, payload: dict) -> dict:
        if not getattr(self, "_is_personal_lucas", lambda: False)():
            raw_person = str(payload.get("assigned_person") or payload.get("person") or "").strip()
            if self._canonical_person_choice(raw_person) is None:
                return {"ok": False, "error": "Choose an existing person from People Rules."}
        record = self._mobile_inventory_payload_record(payload)
        if not record.get("cert_number") and not record.get("card_title"):
            return {"ok": False, "error": "Enter or scan a cert number, or enter a card title."}
        cert = scan_to_cert(record.get("cert_number"))
        update_existing = bool(payload.get("update_existing"))
        with shared_lock(CARD_PIPELINE_DIR, "mobile-inventory", self.lucas_identity):
            ledger = [self._normalize_inventory_record(item) for item in self._load_inventory_ledger()]
            if not cert and not str(record.get("item_id") or "").strip():
                record["item_type"] = "Raw"
                record["item_id"] = self._next_raw_item_id(ledger)
                record["source_sheet"] = "Raw Inventory"
                record["source"] = record.get("source") or "Mobile Raw Card"
                record.pop("inventory_key", None)
                record = self._normalize_inventory_record(record)
            existing_index = next(
                (
                    index
                    for index, item in enumerate(ledger)
                    if cert and scan_to_cert(item.get("cert_number")) == cert and str(item.get("status") or "").lower() == "active"
                ),
                None,
            )
            if existing_index is not None and not update_existing:
                return {
                    "ok": False,
                    "duplicate": True,
                    "error": "That cert is already active in inventory.",
                    "record": self._mobile_inventory_json_record(ledger[existing_index]),
                }
            if existing_index is not None:
                existing = ledger[existing_index]
                for key, value in record.items():
                    if key in {"inventory_key", "date_added"}:
                        continue
                    if value not in ("", None):
                        existing[key] = value
                existing["status"] = "Active"
                ledger[existing_index] = self._enrich_inventory_record_assignment(self._normalize_inventory_record(existing), force=True)
                self._save_inventory_ledger(ledger)
                saved = ledger[existing_index]
                action = "updated"
            else:
                saved = self._enrich_inventory_record_assignment(record)
                ledger.append(saved)
                self._save_inventory_ledger(ledger)
                action = "added"
        self.events.put(("inventory_refresh", f"Mobile inventory {action}: {saved.get('cert_number') or saved.get('card_title') or 'card'}"))
        self._append_activity("Mobile Inventory", f"Mobile inventory {action}: {saved.get('cert_number') or saved.get('card_title') or 'card'}.", {"action": action, "inventory_key": saved.get("inventory_key")})
        return {"ok": True, "action": action, "record": self._mobile_inventory_json_record(saved)}

    def mobile_inventory_mark_sold(self, payload: dict) -> dict:
        inventory_key = str(payload.get("inventory_key") or payload.get("key") or "").strip()
        has_fallback_identifier = any(
            str(payload.get(name) or "").strip()
            for name in ("cert_number", "cert", "item_id", "card_title", "card")
        )
        if not inventory_key and not has_fallback_identifier:
            return {"ok": False, "error": "Choose an inventory card to mark sold."}
        sale_price = self._money_value(payload.get("sale_price") or payload.get("amount") or payload.get("price"))
        if sale_price is None or sale_price < 0:
            return {"ok": False, "error": "Enter a valid sale price."}
        sale_date = str(payload.get("sale_date") or payload.get("date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
        if self._profit_record_date(sale_date) is None:
            return {"ok": False, "error": "Enter the sale date as YYYY-MM-DD."}
        sale_date = self._mobile_local_calendar_date(sale_date)
        sale_method = str(payload.get("sale_method") or payload.get("method") or "").strip()
        company = str(payload.get("company") or payload.get("buyer") or "").strip()
        with shared_lock(CARD_PIPELINE_DIR, "mobile-inventory-sold", self.lucas_identity):
            ledger = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
            record = next((item for item in ledger if str(item.get("inventory_key") or "") == inventory_key), None)
            if record is None:
                record = self._mobile_inventory_sale_match(ledger, payload)
            if record is None:
                already_applied = self._mobile_sold_profit_match(payload, float(sale_price), sale_date)
                if already_applied is not None:
                    return self._mobile_sold_already_applied_result(already_applied, sale_date, sale_method, company, float(sale_price))
                return {"ok": False, "error": "That inventory card was not found."}
            if str(record.get("status") or "").lower() != "active":
                already_payload = dict(payload)
                already_payload.setdefault("cert_number", record.get("cert_number"))
                already_payload.setdefault("item_id", record.get("item_id"))
                already_payload.setdefault("card_title", record.get("card_title"))
                already_applied = self._mobile_sold_profit_match(already_payload, float(sale_price), sale_date)
                if already_applied is not None:
                    return self._mobile_sold_already_applied_result(already_applied, sale_date, sale_method, company, float(sale_price))
                return {"ok": False, "error": "Only active inventory cards can be marked sold."}
            sold_inventory_key = str(record.get("inventory_key") or inventory_key)
            profit_record = self._inventory_sale_profit_record(record, company, float(sale_price), sale_date=sale_date, sale_method=sale_method)
            added = self._append_profit_records([profit_record])
            changed = self._mark_inventory_record_sold(sold_inventory_key, company or "General Sold", float(sale_price))
        if not (added or changed):
            return {"ok": False, "error": "That sale already exists."}
        title = record.get("cert_number") or record.get("card_title") or "card"
        self.events.put(("inventory_refresh", f"Mobile marked sold: {title} for {format_money(sale_price)}."))
        self.events.put(("profit_refresh", f"Mobile marked sold: {title} for {format_money(sale_price)}."))
        self._append_activity("Mobile Sold", f"Mobile marked sold: {title} for {format_money(sale_price)}.", {"inventory_key": sold_inventory_key, "company": company or "General Sold", "sale_price": sale_price})
        return {
            "ok": True,
            "record": self._mobile_inventory_json_record(record),
            "sale": {
                "date": sale_date[:10],
                "method": sale_method,
                "company": company or "General Sold",
                "sale_price": round(float(sale_price), 2),
                "sale_price_display": format_money(sale_price),
                "profit": profit_record.get("profit"),
                "profit_display": format_money(profit_record.get("profit")),
            },
            "people": self._known_people(),
        }

    def _mobile_trade_basis(self, record: dict[str, object]) -> float:
        for key in ("purchase_price", "inventory_value", "estimated_payout"):
            value = self._money_value(record.get(key))
            if value is not None and value > 0:
                return round(float(value), 2)
        return 0.0

    def _mobile_trade_allocations(
        self,
        outgoing_records: list[dict[str, object]],
        incoming_payloads: list[dict[str, object]],
        cash_paid: object = "",
        cash_received: object = "",
    ) -> dict[str, object]:
        paid = self._money_value(cash_paid) or 0.0
        received = self._money_value(cash_received) or 0.0
        outgoing_basis = round(sum(self._mobile_trade_basis(record) for record in outgoing_records), 2)
        incoming_values = [self._money_value(item.get("trade_value") or item.get("inventory_value") or item.get("value")) or 0.0 for item in incoming_payloads]
        total_cost = round(max(0.0, outgoing_basis + float(paid) - float(received)), 2)
        total_value = round(sum(max(0.0, value) for value in incoming_values), 2)
        allocations: list[float] = []
        if incoming_payloads:
            if total_value > 0:
                remaining = total_cost
                for index, value in enumerate(incoming_values):
                    if index == len(incoming_values) - 1:
                        amount = remaining
                    else:
                        amount = round(total_cost * (max(0.0, value) / total_value), 2)
                        remaining = round(remaining - amount, 2)
                    allocations.append(round(max(0.0, amount), 2))
            else:
                split = round(total_cost / len(incoming_payloads), 2)
                remaining = total_cost
                for index in range(len(incoming_payloads)):
                    if index == len(incoming_payloads) - 1:
                        amount = remaining
                    else:
                        amount = split
                        remaining = round(remaining - amount, 2)
                    allocations.append(round(max(0.0, amount), 2))
        return {
            "outgoing_basis": outgoing_basis,
            "cash_paid": round(float(paid), 2),
            "cash_received": round(float(received), 2),
            "incoming_value": total_value,
            "outgoing_side": round(outgoing_basis + float(paid), 2),
            "incoming_side": round(total_value + float(received), 2),
            "difference": round((outgoing_basis + float(paid)) - (total_value + float(received)), 2),
            "total_cost": total_cost,
            "allocations": allocations,
        }

    def mobile_inventory_trade(self, payload: dict) -> dict:
        raw_outgoing = payload.get("outgoing")
        raw_incoming = payload.get("incoming")
        outgoing_payloads = [item for item in raw_outgoing if isinstance(item, dict)] if isinstance(raw_outgoing, list) else []
        incoming_payloads = [item for item in raw_incoming if isinstance(item, dict)] if isinstance(raw_incoming, list) else []
        if not outgoing_payloads and not incoming_payloads:
            return {"ok": False, "error": "Choose outgoing inventory or enter incoming trade cards."}
        if incoming_payloads and not any(str(item.get("cert_number") or item.get("cert") or item.get("card_title") or item.get("card") or "").strip() for item in incoming_payloads):
            return {"ok": False, "error": "Enter at least one incoming card."}
        for item in incoming_payloads:
            value = self._money_value(item.get("trade_value") or item.get("inventory_value") or item.get("value"))
            if value is None or value <= 0:
                return {"ok": False, "error": "Every incoming trade card needs a trade value."}
        if not getattr(self, "_is_personal_lucas", lambda: False)():
            raw_person = str(payload.get("assigned_person") or payload.get("person") or "").strip()
            if self._canonical_person_choice(raw_person) is None:
                return {"ok": False, "error": "Choose an existing person from People Rules."}
        trade_date = str(payload.get("trade_date") or payload.get("date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
        if self._profit_record_date(trade_date) is None:
            return {"ok": False, "error": "Enter the trade date as YYYY-MM-DD."}
        trade_date = self._mobile_local_calendar_date(trade_date)
        trade_partner = "Trade"
        trade_notes = str(payload.get("notes") or "").strip()
        with shared_lock(CARD_PIPELINE_DIR, "mobile-inventory-trade", self.lucas_identity):
            ledger = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
            outgoing_records: list[dict[str, object]] = []
            for item in outgoing_payloads:
                record = next((row for row in ledger if str(row.get("inventory_key") or "") == str(item.get("inventory_key") or item.get("key") or "").strip()), None)
                if record is None:
                    record = self._mobile_inventory_sale_match(ledger, item)
                if record is None:
                    return {"ok": False, "error": f"Outgoing card not found: {item.get('card_title') or item.get('cert_number') or item.get('inventory_key') or 'unknown'}"}
                if str(record.get("status") or "").lower() != "active":
                    return {"ok": False, "error": f"Outgoing card is not active: {record.get('card_title') or record.get('cert_number') or 'card'}"}
                if str(record.get("inventory_key") or "") not in {str(existing.get("inventory_key") or "") for existing in outgoing_records}:
                    outgoing_records.append(record)
            allocation = self._mobile_trade_allocations(outgoing_records, incoming_payloads, payload.get("cash_paid"), payload.get("cash_received"))
            if abs(float(allocation.get("difference") or 0.0)) > 0.01:
                return {
                    "ok": False,
                    "error": f"Trade is not balanced. Difference: {format_money(abs(float(allocation.get('difference') or 0.0)))}.",
                    "trade": allocation,
                }
            sale_records = [
                self._inventory_sale_profit_record(
                    record,
                    trade_partner,
                    self._mobile_trade_basis(record),
                    sale_date=trade_date,
                    sale_method="Trade",
                )
                for record in outgoing_records
            ]
            sold_keys = {str(record.get("inventory_key") or "") for record in outgoing_records}
            added_records: list[dict[str, object]] = []
            for index, incoming in enumerate(incoming_payloads):
                incoming_payload = dict(incoming)
                incoming_payload["purchase_price"] = allocation["allocations"][index] if index < len(allocation["allocations"]) else 0.0
                incoming_payload.setdefault("inventory_value", incoming.get("trade_value") or incoming.get("value") or "")
                incoming_payload.setdefault("assigned_person", payload.get("assigned_person") or payload.get("person") or "")
                incoming_payload.setdefault("source", trade_partner)
                incoming_payload.setdefault("source_sheet", "Mobile Trade")
                notes = "\n".join(part for part in [str(incoming_payload.get("notes") or "").strip(), trade_notes, "Added from mobile trade."] if part)
                incoming_payload["notes"] = notes
                record = self._mobile_inventory_payload_record(incoming_payload)
                if not record.get("cert_number") and not record.get("card_title"):
                    return {"ok": False, "error": "Incoming trade cards need a title or cert."}
                cert = scan_to_cert(record.get("cert_number"))
                if cert and any(scan_to_cert(row.get("cert_number")) == cert and str(row.get("status") or "").lower() == "active" and str(row.get("inventory_key") or "") not in sold_keys for row in ledger + added_records):
                    return {"ok": False, "error": f"Incoming cert is already active in inventory: {cert}"}
                if not record.get("cert_number") and not str(record.get("item_id") or "").strip():
                    record["item_type"] = "Raw"
                    record["item_id"] = self._next_raw_item_id(ledger + added_records)
                    record["source_sheet"] = "Raw Inventory"
                    record["source"] = record.get("source") or trade_partner
                    record.pop("inventory_key", None)
                    record = self._normalize_inventory_record(record)
                added_records.append(self._enrich_inventory_record_assignment(record))
            if sale_records:
                self._append_profit_records(sale_records)
            kept = [record for record in ledger if str(record.get("inventory_key") or "") not in sold_keys]
            kept.extend(added_records)
            self._save_inventory_ledger(kept)
            cleanup = getattr(self, "_delete_inventory_photo_files_for_removed_records", None)
            if callable(cleanup) and outgoing_records:
                cleanup(outgoing_records, kept)
        self.events.put(("inventory_refresh", f"Mobile trade saved: {len(outgoing_records)} outgoing, {len(added_records)} incoming."))
        self.events.put(("profit_refresh", f"Mobile trade saved: {len(outgoing_records)} outgoing, {len(added_records)} incoming."))
        self._append_activity(
            "Mobile Trade",
            f"Mobile trade saved: {len(outgoing_records)} outgoing, {len(added_records)} incoming, basis {format_money(allocation['total_cost'])}.",
            {"outgoing": len(outgoing_records), "incoming": len(added_records), "trade_partner": trade_partner, "total_cost": allocation["total_cost"]},
        )
        return {
            "ok": True,
            "trade": {
                **allocation,
                "total_cost_display": format_money(allocation["total_cost"]),
                "difference_display": format_money(abs(float(allocation.get("difference") or 0.0))),
                "outgoing_count": len(outgoing_records),
                "incoming_count": len(added_records),
                "trade_partner": trade_partner,
            },
            "outgoing": [self._mobile_inventory_json_record(record) for record in outgoing_records],
            "records": [self._mobile_inventory_json_record(record) for record in added_records],
            "people": self._known_people(),
        }

    def _mobile_inventory_sale_match(self, ledger: list[dict[str, object]], payload: dict) -> dict[str, object] | None:
        active = [record for record in ledger if str(record.get("status") or "").lower() == "active"]
        cert = scan_to_cert(payload.get("cert_number") or payload.get("cert") or payload.get("barcode") or "")
        if cert:
            matches = [record for record in active if scan_to_cert(record.get("cert_number")) == cert]
            if len(matches) == 1:
                return matches[0]
        item_id = str(payload.get("item_id") or "").strip().lower()
        if item_id:
            matches = [record for record in active if str(record.get("item_id") or "").strip().lower() == item_id]
            if len(matches) == 1:
                return matches[0]
        title_key = self._mobile_inventory_title_key(payload.get("card_title") or payload.get("card") or "")
        if title_key:
            matches = [record for record in active if self._mobile_inventory_title_key(record.get("card_title")) == title_key]
            if len(matches) == 1:
                return matches[0]
        return None

    def _mobile_inventory_title_key(self, value: object) -> str:
        return re.sub(r"\s+", " ", str(value or "").strip()).casefold()

    def _mobile_sold_profit_match(self, payload: dict, sale_price: float, sale_date: str) -> dict[str, object] | None:
        inventory_key = str(payload.get("inventory_key") or payload.get("key") or "").strip().lower()
        key_parts = [part.strip() for part in inventory_key.split("|") if part.strip()]
        cert = scan_to_cert(payload.get("cert_number") or payload.get("cert") or payload.get("barcode") or "")
        if not cert and key_parts:
            cert = scan_to_cert(key_parts[0])
        item_id = str(payload.get("item_id") or "").strip().lower()
        if not item_id and inventory_key.startswith("raw-"):
            item_id = inventory_key
        title_key = self._mobile_inventory_title_key(payload.get("card_title") or payload.get("card") or "")
        expected_date = self._profit_record_date(sale_date)
        matches: list[dict[str, object]] = []
        for raw_record in self._load_profit_ledger():
            record = self._normalize_profit_record(raw_record)
            if str(record.get("status") or "") != "Sold from inventory":
                continue
            if self._money_value(record.get("sale_price")) != round(float(sale_price), 2):
                continue
            sold_date = self._profit_record_date(record.get("date_added"))
            if expected_date is not None and sold_date is not None and abs((sold_date - expected_date).days) > 1:
                continue
            if cert and scan_to_cert(record.get("cert_number")) == cert:
                matches.append(record)
                continue
            if item_id and str(record.get("item_id") or "").strip().lower() == item_id:
                matches.append(record)
                continue
            if title_key and self._mobile_inventory_title_key(record.get("card_title")) == title_key:
                matches.append(record)
        return matches[0] if len(matches) == 1 else None

    def _mobile_sold_already_applied_result(
        self,
        record: dict[str, object],
        sale_date: str,
        sale_method: str,
        company: str,
        sale_price: float,
    ) -> dict[str, object]:
        return {
            "ok": True,
            "already_applied": True,
            "record": self._mobile_inventory_json_record(record),
            "sale": {
                "date": str(record.get("date_added") or sale_date)[:10],
                "method": str(record.get("sale_method") or sale_method or ""),
                "company": str(record.get("company") or company or "General Sold"),
                "sale_price": round(float(record.get("sale_price") or sale_price), 2),
                "sale_price_display": format_money(record.get("sale_price") or sale_price),
                "profit": record.get("profit"),
                "profit_display": format_money(record.get("profit")),
            },
            "people": self._known_people(),
        }

    def _mobile_local_calendar_date(self, value: object) -> str:
        return self._profit_local_calendar_date(value)

    def _profit_local_calendar_date(self, value: object, ledger_added_at: object = "") -> str:
        date_text = str(value or "").strip()[:10]
        parsed = self._profit_record_date(date_text)
        ledger_date = self._profit_record_date(str(ledger_added_at or "").strip()[:10])
        if parsed is not None and ledger_date is not None and parsed == ledger_date + timedelta(days=1):
            return ledger_date.isoformat()
        today = date.today()
        if parsed == today + timedelta(days=1):
            return today.isoformat()
        return date_text

    def _mobile_profit_rows(self, person: str = "", period: str = "Total") -> list[dict[str, object]]:
        needle = person.strip().lower()
        period_start, period_end = self._profit_period_bounds(period)
        rows = self._enrich_profit_records_with_people(self._load_profit_ledger())
        filtered: list[dict[str, object]] = []
        for record in rows:
            if needle and needle not in str(record.get("assigned_person") or "Unassigned").lower():
                continue
            if period_start is not None:
                sold_date = self._profit_record_date(record.get("date_added"))
                if sold_date is None or sold_date < period_start or sold_date > period_end:
                    continue
            filtered.append(record)
        return sorted(
            filtered,
            key=lambda item: (
                self._profit_added_sort_value(item),
                str(item.get("date_added") or ""),
                str(item.get("company") or ""),
                str(item.get("card_title") or ""),
            ),
            reverse=True,
        )

    def _mobile_profit_chart_series(self, rows: list[dict[str, object]], period: str, graph: str) -> tuple[list[str], list[float]]:
        daily: dict[str, float] = {}
        for record in rows:
            profit = self._money_value(record.get("profit"))
            sold_date = self._profit_record_date(record.get("date_added"))
            if profit is None or sold_date is None:
                continue
            day = sold_date.isoformat()
            daily[day] = daily.get(day, 0.0) + float(profit)
        period_start, period_end = self._profit_period_bounds(period)
        if period_start is not None:
            cursor = period_start
            while cursor <= period_end:
                daily.setdefault(cursor.isoformat(), 0.0)
                cursor += timedelta(days=1)
        days = sorted(daily)
        values = [daily[day] for day in days]
        if graph == "Overall Profit":
            running = 0.0
            cumulative: list[float] = []
            for value in values:
                running += value
                cumulative.append(round(running, 2))
            values = cumulative
        return days, [round(value, 2) for value in values]

    def mobile_profit_summary(self, payload: dict) -> dict:
        period = self._canonical_profit_period(str(payload.get("period") or "Total").strip())
        if period not in PROFIT_PERIOD_OPTIONS:
            period = "Total"
        graph = str(payload.get("graph") or "Daily Trend").strip()
        if graph not in PROFIT_GRAPH_OPTIONS:
            graph = "Daily Trend"
        rows = self._mobile_profit_rows(str(payload.get("person") or ""), period)
        total_purchase = 0.0
        total_sale = 0.0
        gross_profit = 0.0
        expenses = 0.0
        net_profit = 0.0
        complete_count = 0
        recent: list[dict[str, object]] = []
        for record in rows:
            is_expense = str(record.get("record_type") or "").strip().lower() == "expense"
            purchase = self._money_value(record.get("purchase_price"))
            sale = self._money_value(record.get("sale_price"))
            profit = self._money_value(record.get("profit"))
            if purchase is not None:
                total_purchase += purchase
            if sale is not None:
                total_sale += sale
            if profit is not None:
                net_profit += profit
                if is_expense:
                    expenses += abs(profit)
                else:
                    gross_profit += profit
                complete_count += 1
            if len(recent) < 25:
                recent.append(
                    {
                        "date": record.get("date_added") or "",
                        "person": record.get("assigned_person") or "Unassigned",
                        "type": "Expense" if is_expense else "Sale",
                        "title": record.get("card_title") or record.get("company") or "",
                        "company": record.get("company") or "",
                        "profit": round(profit or 0.0, 2) if profit is not None else None,
                        "profit_display": format_money(profit),
                    }
                )
        labels, values = self._mobile_profit_chart_series(rows, period, graph)
        return {
            "ok": True,
            "people": self._known_people(),
            "periods": list(PROFIT_PERIOD_OPTIONS),
            "graphs": list(PROFIT_GRAPH_OPTIONS),
            "totals": {
                "purchase": round(total_purchase, 2),
                "sale": round(total_sale, 2),
                "gross_profit": round(gross_profit, 2),
                "expenses": round(expenses, 2),
                "net_profit": round(net_profit, 2),
                "complete_count": complete_count,
                "row_count": len(rows),
            },
            "chart": {"labels": labels, "values": values},
            "recent": recent,
        }

    def mobile_expense_add(self, payload: dict) -> dict:
        person = str(payload.get("person") or payload.get("assigned_person") or "").strip()
        is_personal = getattr(self, "_is_personal_lucas", lambda: False)
        default_person = getattr(self, "_personal_default_person", lambda: "Mikey")
        if not person and is_personal():
            person = default_person()
        elif not is_personal():
            person = self._canonical_person_choice(person) or ""
        if not person:
            return {"ok": False, "error": "Choose an existing person from People Rules."}
        expense_date = str(payload.get("date") or payload.get("date_added") or "").strip() or datetime.now().strftime("%Y-%m-%d")
        if self._profit_record_date(expense_date) is None:
            return {"ok": False, "error": "Enter the expense date as YYYY-MM-DD."}
        expense_date = self._mobile_local_calendar_date(expense_date)
        amount = self._money_value(payload.get("amount") or payload.get("expense_amount"))
        if amount is None or amount <= 0:
            return {"ok": False, "error": "Enter an expense amount greater than zero."}
        expense_type = str(payload.get("expense_type") or payload.get("type") or "").strip()
        if expense_type not in EXPENSE_CATEGORY_OPTIONS:
            expense_type = "Fees"
        related_type = str(payload.get("related_type") or payload.get("tie_to") or "").strip()
        if related_type not in EXPENSE_LINK_OPTIONS:
            related_type = "General"
        related_sheet = str(payload.get("source_sheet") or payload.get("sheet") or "").strip()
        related_cert = str(payload.get("cert_number") or payload.get("cert") or "").strip()
        if related_type == "Sheet" and not related_sheet:
            return {"ok": False, "error": "Enter the sold sheet this expense belongs to."}
        if related_type == "Card" and not (related_sheet or related_cert):
            return {"ok": False, "error": "Enter a cert number or sold sheet for the card expense."}
        record = {
            "record_type": "expense",
            "expense_id": datetime.now().strftime("%Y%m%d%H%M%S%f"),
            "date_added": expense_date[:10],
            "assigned_person": person,
            "expense_type": expense_type,
            "expense_amount": amount,
            "related_type": related_type,
            "source_sheet": related_sheet,
            "cert_number": related_cert,
            "notes": str(payload.get("notes") or "").strip(),
        }
        added = self._append_profit_records([record])
        if not added:
            return {"ok": False, "error": "That expense already exists in the profit ledger."}
        self.events.put(("profit_refresh", f"Added {expense_type} expense for {person}: {format_money(amount)}."))
        return {"ok": True, "record": self._normalize_profit_record(record), "people": self._known_people()}

    def _load_mobile_action_log(self) -> dict[str, dict[str, object]]:
        if not MOBILE_ACTION_LOG_PATH.exists():
            return {}
        try:
            raw = json.loads(MOBILE_ACTION_LOG_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
        if isinstance(raw, dict) and isinstance(raw.get("applied"), dict):
            return {str(key): value for key, value in raw["applied"].items() if isinstance(value, dict)}
        if isinstance(raw, dict) and isinstance(raw.get("applied"), list):
            return {str(item): {"applied_at": ""} for item in raw["applied"]}
        return {}

    def _save_mobile_action_log(self, applied: dict[str, dict[str, object]]) -> None:
        MOBILE_ACTION_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        items = sorted(
            applied.items(),
            key=lambda pair: str(pair[1].get("applied_at") or ""),
            reverse=True,
        )[:2000]
        atomic_write_json(
            MOBILE_ACTION_LOG_PATH,
            {
                "version": 1,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "applied": dict(items),
            },
        )

    def _apply_mobile_queue_action(self, action: dict[str, object]) -> dict:
        action_type = str(action.get("type") or action.get("action") or "").strip().lower()
        payload = action.get("payload") if isinstance(action.get("payload"), dict) else {}
        if action_type in {"inventory.add", "inventory_add", "add_inventory"}:
            return self.mobile_inventory_add(dict(payload))
        if action_type in {"inventory.sold", "inventory.mark_sold", "inventory_sold", "mark_sold"}:
            return self.mobile_inventory_mark_sold(dict(payload))
        if action_type in {"inventory.trade", "inventory_trade", "trade_inventory"}:
            return self.mobile_inventory_trade(dict(payload))
        if action_type in {"expense.add", "expense_add", "add_expense"}:
            return self.mobile_expense_add(dict(payload))
        return {"ok": False, "error": f"Unsupported mobile queue action type: {action_type or 'blank'}."}

    def mobile_queue_sync(self, payload: dict) -> dict:
        raw_actions = payload.get("actions")
        if not isinstance(raw_actions, list):
            return {"ok": False, "error": "Mobile queue payload must include an actions list."}
        applied = self._load_mobile_action_log()
        results: list[dict[str, object]] = []
        applied_count = 0
        skipped_count = 0
        failed_count = 0
        changed = False
        for index, raw_action in enumerate(raw_actions, start=1):
            if not isinstance(raw_action, dict):
                failed_count += 1
                results.append({"ok": False, "index": index, "error": "Queue action was not an object."})
                continue
            action_id = str(raw_action.get("id") or raw_action.get("action_id") or "").strip()
            action_type = str(raw_action.get("type") or raw_action.get("action") or "").strip()
            if not action_id:
                failed_count += 1
                results.append({"ok": False, "index": index, "type": action_type, "error": "Queue action is missing an id."})
                continue
            if action_id in applied:
                skipped_count += 1
                results.append({"ok": True, "id": action_id, "type": action_type, "status": "already_applied"})
                continue
            try:
                result = self._apply_mobile_queue_action(raw_action)
            except Exception as error:
                result = {"ok": False, "error": str(error)}
            if result.get("ok"):
                applied_count += 1
                changed = True
                applied[action_id] = {
                    "type": action_type,
                    "applied_at": datetime.now(timezone.utc).isoformat(),
                    "client_id": str(payload.get("client_id") or raw_action.get("client_id") or ""),
                }
                results.append({"ok": True, "id": action_id, "type": action_type, "status": "applied", "result": result})
            else:
                failed_count += 1
                results.append({"ok": False, "id": action_id, "type": action_type, "status": "failed", "error": result.get("error") or "Action failed.", "result": result})
        if changed:
            self._save_mobile_action_log(applied)
            self.events.put(("inventory_refresh", f"Applied {applied_count} mobile queued action(s)."))
            self.events.put(("profit_refresh", f"Applied {applied_count} mobile queued action(s)."))
        return {
            "ok": failed_count == 0,
            "applied": applied_count,
            "skipped": skipped_count,
            "failed": failed_count,
            "results": results,
            "people": self._known_people(),
        }

    def import_mobile_queue_file(self) -> None:
        path_text = filedialog.askopenfilename(
            title="Import Mobile Queue",
            filetypes=[("LUCAS mobile queue", "*.json"), ("All files", "*.*")],
        )
        if not path_text:
            return
        path = Path(path_text)
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except Exception as error:
            messagebox.showerror("Import Mobile Queue", f"Could not read queue file: {error}")
            return
        payload = raw if isinstance(raw, dict) else {"actions": raw}
        result = self.mobile_queue_sync(payload)
        applied = int(result.get("applied") or 0)
        skipped = int(result.get("skipped") or 0)
        failed = int(result.get("failed") or 0)
        self.refresh_inventory_tab()
        self.refresh_profit_tab()
        self.inventory_status_var.set(f"Imported mobile queue: applied {applied}, skipped {skipped}, failed {failed}.")
        if failed:
            errors = [str(item.get("error") or item.get("status") or "Unknown error") for item in result.get("results", []) if isinstance(item, dict) and not item.get("ok")]
            messagebox.showwarning("Import Mobile Queue", "\n".join([f"Applied: {applied}", f"Skipped: {skipped}", f"Failed: {failed}", "", *errors[:8]]))
        else:
            messagebox.showinfo("Import Mobile Queue", f"Applied {applied} queued mobile action(s). Skipped {skipped} already-applied action(s).")

    def mobile_payouts(self, payload: dict) -> dict:
        needle = str(payload.get("person") or "").strip().lower()
        balances: dict[str, dict[str, float | int]] = {}
        details: list[dict[str, object]] = []
        for item in self._payout_sheet_items():
            person = str(item.get("person") or "Unassigned")
            if needle and needle not in person.lower():
                continue
            if not item.get("paid"):
                balance = balances.setdefault(person, {"sheets": 0, "cards": 0, "balance": 0.0})
                balance["sheets"] = int(balance["sheets"]) + 1
                balance["cards"] = int(balance["cards"]) + int(item.get("row_count") or 0)
                balance["balance"] = float(balance["balance"]) + float(item.get("payout_balance") or 0.0)
            details.append(
                {
                    "name": item.get("name") or "",
                    "stage": item.get("stage") or "",
                    "person": person,
                    "row_count": int(item.get("row_count") or 0),
                    "received_count": int(item.get("received_count") or 0),
                    "payout_balance": round(float(item.get("payout_balance") or 0.0), 2),
                    "payout_balance_display": format_money(float(item.get("payout_balance") or 0.0)),
                    "status": item.get("status") or "",
                    "paid": bool(item.get("paid")),
                }
            )
        summary = [
            {
                "person": person,
                "sheets": int(values["sheets"]),
                "cards": int(values["cards"]),
                "balance": round(float(values["balance"]), 2),
                "balance_display": format_money(float(values["balance"])),
            }
            for person, values in sorted(balances.items(), key=lambda pair: (-float(pair[1]["balance"]), pair[0].lower()))
        ]
        total_balance = sum(item["balance"] for item in summary)
        return {
            "ok": True,
            "people": self._known_people(),
            "summary": summary,
            "details": details,
            "totals": {
                "balance": round(total_balance, 2),
                "balance_display": format_money(total_balance),
                "sheets": sum(int(item["sheets"]) for item in summary),
                "cards": sum(int(item["cards"]) for item in summary),
            },
        }

    def _mobile_image_parts(self, image: str) -> tuple[str, str, bytes]:
        match = re.match(r"^data:([^;]+);base64,(.*)$", image, re.S)
        if match:
            mime_type = match.group(1) or "image/jpeg"
            image_b64 = match.group(2)
        else:
            mime_type = "image/jpeg"
            image_b64 = image
        return mime_type, image_b64, base64.b64decode(image_b64)

    def _parse_mobile_quick_card_response(self, raw: str) -> dict:
        text = str(raw or "").strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.I | re.S).strip()
        try:
            parsed = json.loads(text)
        except Exception:
            match = re.search(r"\{.*\}", text, re.S)
            if not match:
                return {}
            try:
                parsed = json.loads(match.group(0))
            except Exception:
                return {}
        return parsed if isinstance(parsed, dict) else {}

    def _mobile_quick_card_to_row(self, parsed: dict) -> dict[str, object]:
        grader = normalize_grader(parsed.get("grading_company") or parsed.get("grader") or "")
        title = build_card_title(
            {
                "description": "",
                "year": parsed.get("year"),
                "set": parsed.get("set"),
                "player": parsed.get("player") or parsed.get("subject"),
                "card_number": parsed.get("card_number"),
                "parallel": parsed.get("parallel"),
                "subset": parsed.get("subset") or parsed.get("attributes"),
                "grader": grader,
                "grade": parsed.get("grade"),
            }
        )
        label_text = str(parsed.get("label_text") or "").strip()
        if not title:
            title = str(parsed.get("card_title") or parsed.get("title") or "").strip()
        notes = clean_part("; ".join(part for part in ("Mobile quick scan", label_text[:180]) if part))
        return {
            "cert_number": scan_to_cert(parsed.get("cert_number")),
            "grader": grader or infer_grader(title),
            "card_title": title,
            "purchase_price": None,
            "source": "Mobile Photo",
            "notes": notes,
        }

    def _mobile_single_card_quick_read(self, client, mime_type: str, image_bytes: bytes) -> dict[str, object] | None:
        if genai_types is None:
            return None
        prompt = (
            "Read this single trading card or graded slab photo for inventory entry. "
            "Assume the user is photographing one card/slab. Extract only visible facts; do not guess. "
            "Return JSON only with keys: grading_company, cert_number, player, year, set, card_number, "
            "parallel, subset, attributes, grade, card_title, label_text, confidence. "
            "Normalize cert_number to digits only when possible. If a field is unreadable, use an empty string. "
            "For BGS/Beckett, grade must be the overall slab grade, not a subgrade."
        )
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[
                prompt,
                genai_types.Part.from_bytes(data=image_bytes, mime_type=mime_type),
            ],
            config=genai_types.GenerateContentConfig(
                thinking_config=genai_types.ThinkingConfig(thinking_budget=0),
                max_output_tokens=700,
                response_mime_type="application/json",
                temperature=0,
            ),
        )
        parsed = self._parse_mobile_quick_card_response(response.text or "")
        row = self._mobile_quick_card_to_row(parsed)
        if row.get("cert_number") or row.get("card_title") or row.get("grader"):
            return row
        return None

    def mobile_card_identify(self, payload: dict) -> dict:
        image = str(payload.get("image") or "").strip()
        if not image:
            return {"ok": False, "error": "Take or choose a card photo first."}
        if genai is None or identify_cards_sync is None:
            return {"ok": False, "error": "Photo OCR dependencies are not available."}
        if hasattr(self, "_load_photo_env"):
            self._load_photo_env()
        api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
        if not api_key:
            return {"ok": False, "error": "Missing GOOGLE_API_KEY for photo card search."}
        try:
            mime_type, image_b64, image_bytes = self._mobile_image_parts(image)
        except Exception as error:
            return {"ok": False, "error": f"Could not read that photo: {error}"}
        if len(image_bytes) > 8 * 1024 * 1024:
            return {"ok": False, "error": "That photo is too large for mobile OCR. Retake it a little closer or choose a smaller image."}
        try:
            client = make_photo_ocr_client(api_key)
            row = self._mobile_single_card_quick_read(client, mime_type, image_bytes)
            if row is None:
                cards = identify_cards_sync(client, image_b64)
                self._inventory_photo_rescue_single_bgs_cert(cards, image_b64, client=client)
                rows = [
                    self._photo_card_to_row(Path("mobile-photo.jpg"), card)
                    for card in cards
                    if self._photo_card_has_inventory(card)
                ]
                if not rows:
                    return {"ok": False, "error": "No card was found in that photo."}
                row = rows[0]
                cards_found = len(rows)
                mode = "fallback"
            else:
                cards_found = 1
                mode = "quick"
        except (TemporaryModelUnavailable, ModelQuotaExceeded, ModelResponseParseError) as error:
            return {"ok": False, "error": str(error)}
        except Exception as error:
            return {"ok": False, "error": f"Photo search failed: {error}"}
        query = scan_to_cert(row.get("cert_number")) or str(row.get("card_title") or "").strip()
        return {
            "ok": True,
            "query": query,
            "card": {
                "cert_number": row.get("cert_number"),
                "grader": row.get("grader"),
                "card_title": row.get("card_title"),
                "notes": row.get("notes"),
            },
            "cards_found": cards_found,
            "mode": mode,
        }

    def _retarget_inventory_rows_for_source(self, source_sheet_name: str, assigned_person: str) -> int:
        source_name = Path(str(source_sheet_name or "")).name.strip().lower()
        if not source_name:
            return 0
        new_person = str(assigned_person or "").strip() or "Unassigned"
        changed = 0
        merged: dict[str, dict[str, object]] = {}
        for record in [self._normalize_inventory_record(item) for item in self._load_inventory_ledger()]:
            if Path(str(record.get("source_sheet") or "")).name.strip().lower() == source_name:
                if str(record.get("assigned_person") or "").strip() != new_person:
                    changed += 1
                record["assigned_person"] = new_person
                record["best_company"] = ""
                record["estimated_payout"] = None
                record.pop("inventory_key", None)
                record = self._normalize_inventory_record(record)
                record = self._enrich_inventory_record_assignment(record)
            key = str(record.get("inventory_key") or "")
            if not key:
                continue
            existing = merged.get(key)
            if existing is None:
                merged[key] = record
                continue
            status = "Active" if "active" in {str(existing.get("status") or "").lower(), str(record.get("status") or "").lower()} else str(record.get("status") or existing.get("status") or "")
            existing.update(record)
            existing["status"] = status or "Active"
        if changed:
            self._save_inventory_ledger(list(merged.values()))
        return changed

    def _retarget_profit_rows_for_source(self, source_sheet_name: str, assigned_person: str) -> int:
        source_name = Path(str(source_sheet_name or "")).name.strip().lower()
        if not source_name:
            return 0
        new_person = str(assigned_person or "").strip() or "Unassigned"
        changed = 0
        updated: list[dict[str, object]] = []
        for record in [self._normalize_profit_record(item) for item in self._load_profit_ledger()]:
            source = Path(str(record.get("source_sheet") or "")).name.strip().lower()
            original_source = Path(str(record.get("original_source_sheet") or "")).name.strip().lower()
            if source == source_name or original_source == source_name:
                if str(record.get("assigned_person") or "").strip() != new_person:
                    changed += 1
                record["assigned_person"] = new_person
                record = self._normalize_profit_record(record)
            updated.append(record)
        if changed:
            self._save_profit_ledger(updated)
        return changed

    def _received_certs_in_workbook(self, path: Path) -> set[str]:
        certs: set[str] = set()
        if not path.exists():
            return certs
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                headers = {
                    re.sub(r"[^a-z0-9]", "", str(cell.value or "").strip().lower()): index
                    for index, cell in enumerate(sheet[1], start=1)
                    if str(cell.value or "").strip()
                }
                received_col = headers.get("received")
                cert_col = headers.get("certificationnumber") or headers.get("certnumber") or headers.get("cert")
                if not received_col or not cert_col:
                    continue
                for row_index in range(2, sheet.max_row + 1):
                    received_text = str(sheet.cell(row_index, received_col).value or "").strip().upper()
                    if received_text not in {"X", "Y", "YES", "TRUE", "1", "RECEIVED"}:
                        continue
                    cert = scan_to_cert(sheet.cell(row_index, cert_col).value)
                    if cert:
                        certs.add(cert)
        finally:
            workbook.close()
        return certs

    def _company_sheet_source_cert_keys(self) -> set[tuple[str, str]]:
        keys: set[tuple[str, str]] = set()
        for record in read_company_profit_records(COMPANY_SHEETS_DIR):
            source_sheet = Path(str(record.get("source_sheet") or "")).name.strip().lower()
            cert = scan_to_cert(record.get("cert_number"))
            if source_sheet and cert:
                keys.add((source_sheet, cert))
        return keys

    def _received_inventory_accounted_source_cert_keys(self) -> set[tuple[str, str]]:
        keys: set[tuple[str, str]] = set()
        for record in [self._normalize_inventory_record(row) for row in self._load_inventory_ledger()]:
            source_sheet = Path(str(record.get("source_sheet") or "")).name.strip().lower()
            cert = scan_to_cert(record.get("cert_number"))
            if source_sheet and cert:
                keys.add((source_sheet, cert))
        for record in [self._normalize_profit_record(row) for row in self._load_profit_ledger()]:
            cert = scan_to_cert(record.get("cert_number"))
            for source_value in (record.get("source_sheet"), record.get("original_source_sheet")):
                source_sheet = Path(str(source_value or "")).name.strip().lower()
                if source_sheet and cert:
                    keys.add((source_sheet, cert))
        keys.update(self._inventory_deleted_source_cert_keys())
        return keys

    def _received_inventory_candidate_records_for_sheet(
        self,
        stage: str,
        path: Path,
        person: str,
        company_keys: set[tuple[str, str]] | None = None,
        accounted_keys: set[tuple[str, str]] | None = None,
    ) -> list[dict[str, object]]:
        assigned_person = str(person or "").strip()
        if not assigned_person or not path.exists():
            return []
        received_certs = None if stage == "Received" else self._received_certs_in_workbook(path)
        if received_certs == set():
            return []
        try:
            rows = read_simple_spreadsheet(path)
        except Exception:
            return []
        company_keys = company_keys if company_keys is not None else self._company_sheet_source_cert_keys()
        accounted_loader = getattr(self, "_received_inventory_accounted_source_cert_keys", None)
        accounted_keys = accounted_keys if accounted_keys is not None else accounted_loader() if callable(accounted_loader) else set()
        candidates: list[dict[str, object]] = []
        for row in rows:
            cert = scan_to_cert(row.get("cert_number"))
            if not cert:
                continue
            if received_certs is not None and cert not in received_certs:
                continue
            if (path.name.lower(), cert) in company_keys:
                continue
            if (path.name.lower(), cert) in accounted_keys:
                continue
            card_title = str(row.get("card_title") or "")
            sport = CardPipelineApp._inventory_sport_from_value(self, row.get("sport") or row.get("category"), card_title)
            candidates.append(
                self._normalize_inventory_record(
                    {
                        "date_added": datetime.now().strftime("%Y-%m-%d"),
                        "item_type": "Graded",
                        "item_id": "",
                        "assigned_person": assigned_person,
                        "sport": sport,
                        "cert_number": cert,
                        "grader": row.get("grader") or "",
                        "card_title": card_title,
                        "purchase_price": row.get("purchase_price"),
                        "card_ladder_value": row.get("card_ladder_value"),
                        "card_ladder_comps_average": row.get("card_ladder_comps_average"),
                        "cy_value": row.get("cy_value"),
                        "cy_confidence": row.get("cy_confidence"),
                        "inventory_value": row.get("card_ladder_comps_average") or row.get("card_ladder_value") or row.get("cy_value"),
                        "best_company": row.get("best_company") or "",
                        "estimated_payout": row.get("estimated_payout"),
                        "source_sheet": path.name,
                        "source": row.get("source") or "",
                        "status": "Active",
                        "notes": "",
                    }
                )
            )
        return candidates

    def _received_inventory_candidate_records(self) -> list[dict[str, object]]:
        company_keys = self._company_sheet_source_cert_keys()
        accounted_keys = self._received_inventory_accounted_source_cert_keys()
        candidates: list[dict[str, object]] = []
        default_person = self._personal_default_person() if self._is_personal_lucas() else "Unassigned"
        for stage, directory in (("Received", RECEIVED_SHEETS_DIR), ("Incoming", INCOMING_SHEETS_DIR), ("Working", WORKING_SHEETS_DIR)):
            if not directory.exists():
                continue
            for path in sorted(directory.glob("*.xlsx"), key=lambda item: item.name.lower()):
                marker = self.home_sheet_markers.get(self._home_sheet_key(stage, path.name), {})
                person = str(marker.get("assigned_person") or "").strip() or default_person
                if not person:
                    continue
                candidates.extend(self._received_inventory_candidate_records_for_sheet(stage, path, person, company_keys, accounted_keys))
        return candidates

    def _sync_received_inventory_to_ledger(self, filtered_only: bool = False) -> tuple[int, int]:
        records = self._received_inventory_candidate_records()
        if filtered_only:
            records = self._filtered_inventory_records([self._normalize_inventory_record(record) for record in records])
        added = self.add_inventory_records(records, refresh=False)
        return added, len(records)

    def _sync_received_sheet_inventory_to_ledger(self, stage: str, path: Path, person: str) -> tuple[int, int]:
        records = self._received_inventory_candidate_records_for_sheet(stage, path, person)
        added = self.add_inventory_records(records, refresh=False)
        return added, len(records)

    def _inventory_workbook_row(self, record: dict[str, object], excel_row: int) -> WorkbookRow:
        inventory_value = self._money_value(record.get("inventory_value"))
        card_ladder_value = self._money_value(record.get("card_ladder_value"))
        comps_average = self._money_value(record.get("card_ladder_comps_average"))
        cy_value = self._money_value(record.get("cy_value"))
        if card_ladder_value is None and comps_average is None and cy_value is None:
            card_ladder_value = inventory_value
            comps_average = inventory_value
        return WorkbookRow(
            excel_row=excel_row,
            cert_number=str(record.get("cert_number") or ""),
            grader=str(record.get("grader") or ""),
            card_title=str(record.get("card_title") or ""),
            item_id=str(record.get("item_id") or ""),
            category=str(record.get("sport") or ""),
            existing_value=self._money_value(record.get("purchase_price")),
            card_ladder_value=card_ladder_value,
            card_ladder_comps_average=comps_average,
            cy_value=cy_value,
            cy_confidence=record.get("cy_confidence"),
            best_company=str(record.get("best_company") or ""),
            estimated_payout=self._money_value(record.get("estimated_payout")),
            company_pile=True,
            status="Moved from inventory",
            notes=str(record.get("notes") or ""),
        )

    def _inventory_source_sheet_path(self, source_sheet: str) -> Path | None:
        name = Path(str(source_sheet or "")).name
        if not name:
            return None
        for directory in (RECEIVED_SHEETS_DIR, WORKING_SHEETS_DIR, INCOMING_SHEETS_DIR):
            path = directory / name
            if path.exists():
                return path
        return None

    def _inventory_source_rows_by_cert(self, path: Path) -> dict[str, dict[str, object]]:
        cache = getattr(self, "_inventory_source_rows_cache", None)
        if not isinstance(cache, dict):
            cache = {}
            self._inventory_source_rows_cache = cache
        cache_key = str(path.resolve()) if path.exists() else str(path)
        if cache_key in cache:
            return cache[cache_key]
        try:
            rows = read_simple_spreadsheet(path)
        except Exception:
            cache[cache_key] = {}
            return cache[cache_key]
        by_cert: dict[str, dict[str, object]] = {}
        for row in rows:
            cert = scan_to_cert(row.get("cert_number"))
            if cert and cert not in by_cert:
                by_cert[cert] = row
        cache[cache_key] = by_cert
        return by_cert

    def _hydrate_inventory_record_source_values(self, record: dict[str, object]) -> dict[str, object]:
        normalized = self._normalize_inventory_record(record)
        if (
            normalized.get("card_ladder_value") is not None
            and normalized.get("card_ladder_comps_average") is not None
            and normalized.get("cy_value") is not None
            and str(normalized.get("cy_confidence") or "").strip()
        ):
            return normalized
        cert = scan_to_cert(normalized.get("cert_number"))
        path = self._inventory_source_sheet_path(str(normalized.get("source_sheet") or ""))
        if not cert or path is None:
            return normalized
        source_reader = getattr(self, "_inventory_source_rows_by_cert", None)
        if callable(source_reader):
            row = source_reader(path).get(cert)
        else:
            try:
                row = next((candidate for candidate in read_simple_spreadsheet(path) if scan_to_cert(candidate.get("cert_number")) == cert), None)
            except Exception:
                row = None
        if not row:
            return normalized
        if not str(normalized.get("sport") or "").strip():
            normalized["sport"] = CardPipelineApp._inventory_sport_from_value(self, row.get("sport") or row.get("category"), row.get("card_title"))
        for source_field in ("card_ladder_value", "card_ladder_comps_average", "cy_value"):
            if normalized.get(source_field) is None:
                normalized[source_field] = self._money_value(row.get(source_field))
        if not str(normalized.get("cy_confidence") or "").strip() and row.get("cy_confidence") is not None:
            normalized["cy_confidence"] = row.get("cy_confidence")
        return self._normalize_inventory_record(normalized)

    def _enrich_inventory_record_assignment(self, record: dict[str, object], force: bool = False) -> dict[str, object]:
        hydrator = getattr(self, "_hydrate_inventory_record_source_values", None)
        normalized = hydrator(record) if callable(hydrator) else self._normalize_inventory_record(record)
        current_best = str(normalized.get("best_company") or "").strip()
        if (
            not force
            and current_best
            and current_best.upper() != NO_COMPANY_TAKES_LABEL.upper()
            and normalized.get("estimated_payout") is not None
        ):
            return normalized
        try:
            row = self._inventory_workbook_row(normalized, 1)
            if not str(normalized.get("sport") or "").strip():
                parsed = assignment_engine.parse_card_for_matching(str(normalized.get("card_title") or ""))
                if parsed.get("sport"):
                    normalized["sport"] = parsed.get("sport")
                    row.category = str(parsed.get("sport") or "")
            recommendation = self.assignment_engine.recommend(row, person=str(normalized.get("assigned_person") or ""))
        except Exception:
            return normalized
        if recommendation.payout is None:
            normalized["best_company"] = NO_COMPANY_TAKES_LABEL
            normalized["estimated_payout"] = None
            normalized["inventory_value"] = getattr(recommendation, "source_value", None) or assignment_engine.assignment_value(row)
            return normalized
        normalized["best_company"] = recommendation.company
        normalized["estimated_payout"] = recommendation.payout
        normalized["inventory_value"] = getattr(recommendation, "source_value", None) or assignment_engine.assignment_value(row)
        return normalized

    def _raw_inventory_card_dialog(self) -> dict[str, object] | None:
        personal_inventory = self._is_personal_lucas()
        person_var = tk.StringVar(value=self._personal_default_person() if personal_inventory else (self.inventory_person_var.get().strip() if hasattr(self, "inventory_person_var") else ""))
        cert_var = tk.StringVar()
        grader_var = tk.StringVar()
        title_var = tk.StringVar()
        purchase_var = tk.StringVar()
        paid_with_var = tk.StringVar()
        card_ladder_var = tk.StringVar()
        comps_var = tk.StringVar()
        cy_var = tk.StringVar()
        cy_confidence_var = tk.StringVar()
        best_company_var = tk.StringVar()
        payout_var = tk.StringVar()
        notes_var = tk.StringVar()
        result: dict[str, object] = {}

        popup = tk.Toplevel(self)
        popup.title("Add Card")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Add Card", style="Panel.TLabel", font=("Segoe UI Semibold", 12)).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 2))

        fields = [
            ("Cert", cert_var, 24),
            ("Grader", grader_var, 18),
            ("Card description", title_var, 52),
            ("Purchase", purchase_var, 18),
            ("Paid With", paid_with_var, 28),
            ("Card Ladder", card_ladder_var, 18),
            ("Comps", comps_var, 18),
            ("CY Estimate", cy_var, 18),
            ("CY Confidence", cy_confidence_var, 18),
            ("Best Company", best_company_var, 28),
            ("Est. Payout", payout_var, 18),
            ("Notes", notes_var, 52),
        ]
        if not personal_inventory:
            fields.insert(0, ("Person", person_var, 28))
        for index, (label, var, width) in enumerate(fields):
            row = 1 + index
            ttk.Label(frame, text=label, style="Panel.TLabel").grid(row=row, column=0, sticky="w", padx=(0, 10), pady=(0, 8))
            if label == "Person":
                person_combo = ttk.Combobox(frame, textvariable=var, width=width)
                person_combo.grid(row=row, column=1, columnspan=3, sticky="ew", pady=(0, 8))
                self._bind_person_autocomplete(person_combo)
            else:
                ttk.Entry(frame, textvariable=var, width=width).grid(row=row, column=1, columnspan=3, sticky="ew", pady=(0, 8))

        required_text = "Card description and purchase are required." if personal_inventory else "Person, card description, and purchase are required. Cert and grader are optional."
        status_var = tk.StringVar(value=required_text)
        status_row = 1 + len(fields)
        ttk.Label(frame, textvariable=status_var, style="Muted.TLabel").grid(row=status_row, column=0, columnspan=4, sticky="w", pady=(4, 14))

        def optional_money(var: tk.StringVar, label: str) -> float | None:
            text = var.get().strip()
            if not text:
                return None
            value = self._money_value(text)
            if value is None or value < 0:
                raise ValueError(label)
            return float(value)

        def show_validation(message: str) -> None:
            status_var.set(message)
            popup.bell()
            messagebox.showinfo("Add Card", message, parent=popup)

        def submit() -> None:
            person = person_var.get().strip()
            title = title_var.get().strip()
            purchase = self._money_value(purchase_var.get())
            if not person and personal_inventory:
                person = self._personal_default_person()
            person_choice = self._canonical_person_choice(person)
            if person_choice is None:
                show_validation("Choose an existing person.")
                return
            person = person_choice
            if not title:
                show_validation("Enter a card description.")
                return
            if purchase is None or purchase < 0:
                show_validation("Enter a valid purchase price.")
                return
            try:
                card_ladder = optional_money(card_ladder_var, "Card Ladder")
                comps = optional_money(comps_var, "Comps")
                cy_value = optional_money(cy_var, "CY Estimate")
                payout = optional_money(payout_var, "Est. Payout")
            except ValueError as error:
                show_validation(f"Enter a valid value for {error}.")
                return
            result.update(
                {
                    "assigned_person": person,
                    "cert_number": scan_to_cert(cert_var.get()),
                    "grader": grader_var.get().strip(),
                    "card_title": title,
                    "purchase_price": float(purchase),
                    "paid_with": paid_with_var.get().strip(),
                    "card_ladder_value": card_ladder,
                    "card_ladder_comps_average": comps,
                    "cy_value": cy_value,
                    "cy_confidence": cy_confidence_var.get().strip(),
                    "best_company": best_company_var.get().strip(),
                    "estimated_payout": payout,
                    "notes": notes_var.get().strip(),
                }
            )
            popup.destroy()

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=status_row + 1, column=0, columnspan=4, sticky="e")
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="Add Card", command=submit, style="Primary.TButton").pack(side=tk.LEFT)
        frame.columnconfigure(1, weight=1)
        popup.bind("<Return>", lambda _event: submit())
        popup.bind("<Escape>", lambda _event: popup.destroy())
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")
        self.wait_window(popup)
        return result or None

    def add_raw_inventory_card(self) -> None:
        values = self._raw_inventory_card_dialog()
        if values is None:
            return
        try:
            with shared_lock(CARD_PIPELINE_DIR, "inventory-raw-add", self.lucas_identity):
                existing = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
                cert = scan_to_cert(values.get("cert_number"))
                item_id = "" if cert else self._next_raw_item_id(existing)
                record = self._normalize_inventory_record(
                    {
                        **values,
                        "date_added": datetime.now().strftime("%Y-%m-%d"),
                        "item_type": "Graded" if cert else "Raw",
                        "item_id": item_id,
                        "sport": CardPipelineApp._inventory_sport_from_value(self, "", values.get("card_title")),
                        "cert_number": cert,
                        "grader": str(values.get("grader") or "").strip(),
                        "source_sheet": "Manual Inventory" if cert else "Raw Inventory",
                        "source": "Manual Card",
                        "status": "Active",
                    }
                )
                record = self._enrich_inventory_record_assignment(record)
                existing.append(record)
                self._save_inventory_ledger(existing)
            self.refresh_inventory_tab()
            card_id = record.get("cert_number") or record.get("item_id") or "manual card"
            message = f"Added inventory card {card_id}."
            if hasattr(self, "inventory_status_var"):
                self.inventory_status_var.set(message)
            if hasattr(self, "status_var"):
                self.status_var.set(message)
            self._append_activity("Inventory Add", message, {"item_id": record.get("item_id"), "cert_number": record.get("cert_number"), "person": record.get("assigned_person"), "card": record.get("card_title")})
        except Exception as error:
            message = f"Add Card failed: {error}"
            if hasattr(self, "inventory_status_var"):
                self.inventory_status_var.set(message)
            if hasattr(self, "status_var"):
                self.status_var.set(message)
            try:
                messagebox.showerror("Add Card failed", message)
            except Exception:
                pass

    def _is_personal_lucas(self) -> bool:
        return is_personal_lucas_profile(getattr(self, "app_settings", {}), SETTINGS_PATH)

    def _personal_default_person(self) -> str:
        return "Mikey"

    def _owner_for_profile(self, person: object = "") -> str:
        if self._is_personal_lucas():
            return self._personal_default_person()
        return str(person or "").strip() or "Unassigned"

    def _personal_instagram_sync_enabled(self) -> bool:
        return str(os.environ.get("LUCAS_ENABLE_PERSONAL_INSTAGRAM_SYNC") or "").strip().lower() in {"1", "true", "yes", "on"}

    def _instagram_background_tunnel_enabled(self) -> bool:
        value = str(os.environ.get("LUCAS_INSTAGRAM_BACKGROUND_TUNNEL") or "").strip().lower()
        if value in {"0", "false", "no", "off"}:
            return False
        return self._personal_instagram_sync_enabled()

    def _cloudflared_executable(self) -> str:
        candidates = [
            os.environ.get("LUCAS_CLOUDFLARED_PATH"),
            shutil.which("cloudflared"),
            "/opt/homebrew/bin/cloudflared",
            "/usr/local/bin/cloudflared",
        ]
        for candidate in candidates:
            path = str(candidate or "").strip()
            if path and Path(path).exists():
                return path
        return ""

    def _ensure_instagram_background_tunnel(self) -> str:
        if self.instagram_tunnel_public_url:
            return self.instagram_tunnel_public_url
        process = self.instagram_tunnel_process
        if process is not None and process.poll() is None:
            return ""
        if not self._instagram_background_tunnel_enabled() or not getattr(self, "bridge", None) or not self.bridge.started:
            return ""
        cloudflared = self._cloudflared_executable()
        if not cloudflared:
            return ""
        target = f"http://127.0.0.1:{self.bridge.port}"
        try:
            self.instagram_tunnel_log_path.parent.mkdir(parents=True, exist_ok=True)
            self.instagram_tunnel_process = subprocess.Popen(
                [cloudflared, "tunnel", "--url", target, "--no-autoupdate"],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
            )
        except Exception as error:
            app_debug_log(f"instagram_tunnel_start_failed error={error}")
            return ""

        threading.Thread(target=self._watch_instagram_background_tunnel, daemon=True).start()
        return ""

    def _watch_instagram_background_tunnel(self) -> None:
        process = self.instagram_tunnel_process
        if process is None or process.stdout is None:
            return
        url_pattern = re.compile(r"https://[a-z0-9-]+\.trycloudflare\.com", re.I)
        try:
            with self.instagram_tunnel_log_path.open("a", encoding="utf-8") as log:
                log.write(f"\n--- LUCAS Instagram tunnel started {datetime.now().isoformat(timespec='seconds')} ---\n")
                for line in process.stdout:
                    log.write(line)
                    log.flush()
                    match = url_pattern.search(line)
                    if match:
                        url = match.group(0).rstrip("/")
                        if url != self.instagram_tunnel_public_url:
                            self.instagram_tunnel_public_url = url
                            os.environ["LUCAS_INSTAGRAM_PUBLIC_BRIDGE_URL"] = url
                            self.events.put(("status", f"Instagram photo bridge running in background: {url}"))
        except Exception as error:
            app_debug_log(f"instagram_tunnel_watch_failed error={error}")

    def _stop_instagram_background_tunnel(self) -> None:
        process = self.instagram_tunnel_process
        self.instagram_tunnel_process = None
        if process is None or process.poll() is not None:
            return
        try:
            process.terminate()
            process.wait(timeout=5)
        except Exception:
            try:
                process.kill()
            except Exception:
                pass

    def _instagram_env_config(self) -> dict[str, str]:
        if load_dotenv:
            load_dotenv(ROOT / ".env", override=False)
        background_enabled = self._instagram_background_tunnel_enabled() if hasattr(self, "_instagram_background_tunnel_enabled") else False
        background_url = self._ensure_instagram_background_tunnel() if background_enabled and hasattr(self, "_ensure_instagram_background_tunnel") else ""
        manual_bridge_url = str(os.environ.get("LUCAS_INSTAGRAM_PUBLIC_BRIDGE_URL") or os.environ.get("LUCAS_PUBLIC_BRIDGE_URL") or "").strip().rstrip("/")
        return {
            "user_id": str(os.environ.get("LUCAS_INSTAGRAM_USER_ID") or os.environ.get("LUCAS_INSTAGRAM_ACCOUNT_ID") or "").strip(),
            "access_token": str(os.environ.get("LUCAS_INSTAGRAM_ACCESS_TOKEN") or "").strip(),
            "public_photo_base_url": str(os.environ.get("LUCAS_INSTAGRAM_PUBLIC_PHOTO_BASE_URL") or "").strip().rstrip("/"),
            "public_bridge_url": str(background_url or ("" if background_enabled else manual_bridge_url)).strip().rstrip("/"),
            "daily_post_limit": str(os.environ.get("LUCAS_INSTAGRAM_DAILY_POST_LIMIT") or "75").strip(),
        }

    def _instagram_daily_post_limit(self, config: dict[str, object] | None = None) -> int:
        value = ""
        if isinstance(config, dict):
            value = str(config.get("daily_post_limit") or "").strip()
        if not value:
            value = str(os.environ.get("LUCAS_INSTAGRAM_DAILY_POST_LIMIT") or "75").strip()
        try:
            limit = int(value)
        except (TypeError, ValueError):
            limit = 75
        return max(1, min(limit, 100))

    def _instagram_meta_publish_limit(self) -> int:
        value = str(os.environ.get("LUCAS_INSTAGRAM_META_POST_LIMIT") or "50").strip()
        try:
            limit = int(value)
        except (TypeError, ValueError):
            limit = 50
        return max(1, limit)

    def _instagram_content_publishing_limit(self, config: dict[str, object] | None = None) -> dict[str, object]:
        config = config if isinstance(config, dict) else self._instagram_env_config()
        meta_limit = self._instagram_meta_publish_limit()
        user_id = str(config.get("user_id") or "").strip()
        token = str(config.get("access_token") or "").strip()
        if not user_id or not token:
            return {"limit": meta_limit, "quota_usage": None, "remaining": None, "error": "Missing Instagram token or user id."}
        try:
            response = self._instagram_api_json(f"{user_id}/content_publishing_limit", {"access_token": token})
            data = response.get("data") if isinstance(response, dict) else []
            first = data[0] if isinstance(data, list) and data and isinstance(data[0], dict) else {}
            usage = int(first.get("quota_usage") or 0)
            return {"limit": meta_limit, "quota_usage": usage, "remaining": max(0, meta_limit - usage), "error": ""}
        except Exception as error:
            return {"limit": meta_limit, "quota_usage": None, "remaining": None, "error": str(error)}

    def _load_instagram_inventory_state(self) -> dict[str, object]:
        if not INSTAGRAM_INVENTORY_STATE_PATH.exists():
            return {"version": 1, "posts": {}}
        try:
            raw = json.loads(INSTAGRAM_INVENTORY_STATE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {"version": 1, "posts": {}}
        if not isinstance(raw, dict):
            return {"version": 1, "posts": {}}
        if not isinstance(raw.get("posts"), dict):
            raw["posts"] = {}
        raw.setdefault("version", 1)
        return raw

    def _save_instagram_inventory_state(self, state: dict[str, object]) -> None:
        INSTAGRAM_INVENTORY_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_json(INSTAGRAM_INVENTORY_STATE_PATH, state)

    def _instagram_profile_url(self) -> str:
        if load_dotenv:
            load_dotenv(ROOT / ".env", override=False)
        profile_url = str(os.environ.get("LUCAS_INSTAGRAM_PROFILE_URL") or "").strip()
        if profile_url:
            return profile_url
        username = str(os.environ.get("LUCAS_INSTAGRAM_USERNAME") or "").strip().strip("@")
        if username:
            return f"https://www.instagram.com/{username}/"
        return "https://www.instagram.com/"

    def _instagram_inventory_identity(self, record: dict[str, object] | None) -> str:
        if not isinstance(record, dict):
            return ""
        cert = scan_to_cert(record.get("cert_number"))
        if cert and len(cert) >= 5:
            return f"cert:{cert}"
        item_id = str(record.get("item_id") or "").strip().lower()
        if item_id:
            return f"item:{re.sub(r'[^a-z0-9]+', '', item_id)}"
        title = str(record.get("card_title") or record.get("caption") or "").strip().lower()
        title = re.sub(r"[^a-z0-9]+", " ", title)
        title = re.sub(r"\s+", " ", title).strip()
        return f"title:{title}" if title else ""

    def _instagram_post_entry_identity(self, entry: dict[str, object]) -> str:
        explicit = str(entry.get("inventory_identity") or "").strip()
        if explicit:
            return explicit
        return self._instagram_inventory_identity(
            {
                "cert_number": entry.get("cert_number"),
                "item_id": entry.get("item_id"),
                "card_title": entry.get("card_title") or entry.get("caption"),
            }
        )

    def _instagram_record_duplicate_post(self, state: dict[str, object], record: dict[str, object], post: dict[str, object], method: str, score: float) -> None:
        media_id = str(post.get("id") or post.get("media_id") or "").strip()
        if not media_id:
            return
        duplicates = state.setdefault("duplicate_posts", [])
        if not isinstance(duplicates, list):
            duplicates = []
            state["duplicate_posts"] = duplicates
        if any(isinstance(item, dict) and str(item.get("media_id") or "").strip() == media_id for item in duplicates):
            return
        duplicates.append(
            {
                "inventory_key": str(record.get("inventory_key") or ""),
                "inventory_identity": self._instagram_inventory_identity(record),
                "media_id": media_id,
                "caption": str(post.get("caption") or record.get("card_title") or "").strip(),
                "permalink": str(post.get("permalink") or ""),
                "photo_url": str(post.get("media_url") or ""),
                "detected_at": datetime.now().isoformat(timespec="seconds"),
                "reason": "duplicate_inventory_post",
                "matched_by": method,
                "match_score": round(float(score), 3),
            }
        )
        if len(duplicates) > 500:
            del duplicates[:-500]

    def _instagram_auto_sync_due(self, now: datetime | None = None) -> bool:
        if not self._personal_instagram_sync_enabled():
            return False
        if getattr(self, "instagram_auto_sync_running", False):
            return False
        state = self._load_instagram_inventory_state()
        today = (now or datetime.now()).date().isoformat()
        return str(state.get("last_auto_sync_date") or "") != today

    def _mark_instagram_auto_sync_completed(self, day: str, summary: str) -> None:
        state = self._load_instagram_inventory_state()
        state["last_auto_sync_date"] = day
        state["last_auto_sync_at"] = datetime.now().isoformat(timespec="seconds")
        state["last_auto_sync_summary"] = summary
        self._save_instagram_inventory_state(state)

    def _instagram_auto_sync_timer(self) -> None:
        try:
            self._run_instagram_auto_sync_if_due()
        finally:
            if self._personal_instagram_sync_enabled():
                self.after(60 * 60 * 1000, self._instagram_auto_sync_timer)

    def _run_instagram_auto_sync_if_due(self) -> bool:
        if not self._instagram_auto_sync_due():
            return False
        config = self._instagram_env_config()
        if not str(config.get("user_id") or "").strip() or not str(config.get("access_token") or "").strip():
            self.events.put(("status", "Instagram daily sync skipped: missing Instagram token or user id."))
            return False
        plan = self._instagram_inventory_plan()
        daily_limit = self._instagram_daily_post_limit(config)
        quota = self._instagram_content_publishing_limit(config)
        meta_remaining = quota.get("remaining") if isinstance(quota.get("remaining"), int) else None
        ready_posts = [item for item in plan.get("to_post") or [] if isinstance(item, dict) and instagram_ready_photo_urls(item)]
        if meta_remaining is None and ready_posts:
            self.events.put(("status", f"Instagram daily sync skipped: could not verify Meta publishing quota: {quota.get('error') or 'unknown error'}"))
            return False
        post_limit = min(daily_limit, meta_remaining if meta_remaining is not None else daily_limit)
        ready_posts = ready_posts[:post_limit]
        removable = [item for item in plan.get("to_remove") or [] if isinstance(item, dict) and str(item.get("media_id") or "").strip()]
        today = datetime.now().date().isoformat()
        if not ready_posts and not removable:
            summary = "meta_quota_full" if meta_remaining == 0 else "no_changes"
            self._mark_instagram_auto_sync_completed(today, summary)
            self.events.put(("status", "Instagram daily sync checked: Meta publishing quota is full." if meta_remaining == 0 else "Instagram daily sync checked: no inventory changes to post or remove."))
            return False
        auto_plan = {**plan, "to_post": ready_posts, "to_remove": removable, "missing_public_urls": []}
        self.instagram_auto_sync_running = True
        worker = threading.Thread(target=self._instagram_auto_sync_worker, args=(auto_plan, today), daemon=True)
        worker.start()
        return True

    def _instagram_auto_sync_worker(self, plan: dict[str, object], day: str) -> None:
        try:
            self._instagram_inventory_sync_worker(plan)
            limit = self._instagram_daily_post_limit(plan.get("config") if isinstance(plan.get("config"), dict) else None)
            quota = self._instagram_content_publishing_limit(plan.get("config") if isinstance(plan.get("config"), dict) else None)
            remaining = quota.get("remaining")
            summary = f"posted={len(plan.get('to_post') or [])} remove_candidates={len(plan.get('to_remove') or [])} daily_limit={limit} meta_remaining={remaining if remaining is not None else 'unknown'}"
            self._mark_instagram_auto_sync_completed(day, summary)
        finally:
            self.instagram_auto_sync_running = False

    def _instagram_limited_manual_sync_plan(self, plan: dict[str, object], meta_remaining: int | None = None) -> tuple[dict[str, object], int, int]:
        limit = self._instagram_daily_post_limit(plan.get("config") if isinstance(plan.get("config"), dict) else None)
        if meta_remaining is not None:
            limit = min(limit, max(0, meta_remaining))
        ready_posts = [
            item
            for item in plan.get("to_post") or []
            if isinstance(item, dict) and instagram_ready_photo_urls(item)
        ]
        limited_posts = ready_posts[:limit]
        limited_plan = {
            **plan,
            "to_post": limited_posts,
            "missing_public_urls": [],
        }
        return limited_plan, len(ready_posts), limit

    def _inventory_photo_encoded_id(self, path: Path) -> str:
        try:
            storage_value = self._inventory_photo_storage_value(path)
        except Exception:
            storage_value = str(path)
        raw = storage_value.encode("utf-8")
        return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")

    def _instagram_inventory_photo_id(self, path: Path) -> str:
        return self._inventory_photo_encoded_id(path)

    def _instagram_inventory_photo_url(self, path: Path, config: dict[str, str]) -> str:
        bridge_url = config.get("public_bridge_url", "").strip().rstrip("/")
        bridge_state = getattr(self, "state", None)
        if bridge_url and bridge_state is not None and hasattr(bridge_state, "instagram_media_path"):
            photo_id = self._instagram_inventory_photo_id(path)
            media_path = bridge_state.instagram_media_path(photo_id, path.name)
            return f"{bridge_url}{media_path}"
        base_url = config.get("public_photo_base_url", "").strip().rstrip("/")
        if not base_url:
            return ""
        try:
            relative = self._inventory_photo_relative_path(path)
        except Exception:
            relative = None
        if relative is None:
            relative = Path(path.name)
        return f"{base_url}/{urllib.parse.quote(relative.as_posix())}"

    def _instagram_inventory_photo_is_postable(self, path: Path) -> bool:
        if not path.exists() or not path.is_file():
            return False
        try:
            from PIL import Image

            with Image.open(path) as image:
                extrema = image.convert("L").getextrema()
        except Exception:
            return False
        if not extrema:
            return False
        return int(extrema[1]) - int(extrema[0]) > 8

    def _instagram_post_photo_id(self, post_entry: dict[str, object]) -> str:
        photo_id = str(post_entry.get("photo_id") or "").strip()
        if photo_id:
            return photo_id
        photo_url = str(post_entry.get("photo_url") or "").strip()
        parsed = urllib.parse.urlparse(photo_url)
        match = re.match(r"^/instagram/media/[^/]+/([^/]+)(?:/|$)", parsed.path)
        return urllib.parse.unquote(match.group(1)) if match else ""

    def _instagram_cover_photo_path(self, paths: list[Path]) -> Path | None:
        if not paths:
            return None
        return instagram_inventory_photo_order(paths)[0]

    def instagram_inventory_media_response(self, photo_id: str) -> tuple[bytes, str] | None:
        if not self._personal_instagram_sync_enabled():
            return None
        return self._inventory_photo_media_response(photo_id)

    def mobile_inventory_photo_response(self, photo_id: str) -> tuple[bytes, str] | None:
        return self._inventory_photo_media_response(photo_id)

    def _inventory_photo_media_response(self, photo_id: str) -> tuple[bytes, str] | None:
        padding = "=" * (-len(str(photo_id or "")) % 4)
        try:
            path_value = base64.urlsafe_b64decode(f"{photo_id}{padding}".encode("ascii")).decode("utf-8")
        except Exception:
            return None
        path = self._safe_inventory_photo_path(path_value)
        if path is None or not path.is_file():
            return None
        content_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        if not content_type.startswith("image/"):
            return None
        try:
            return path.read_bytes(), content_type
        except OSError:
            return None

    def _instagram_inventory_active_records(self) -> list[dict[str, object]]:
        rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        active = []
        for record in rows:
            if str(record.get("status") or "Active").strip().lower() != "active":
                continue
            if not str(record.get("card_title") or "").strip():
                continue
            active.append(record)
        return active

    def _instagram_active_identity_map(self, active_records: list[dict[str, object]]) -> dict[str, dict[str, object]]:
        active_by_identity: dict[str, dict[str, object]] = {}
        for record in active_records:
            identity = self._instagram_inventory_identity(record)
            if identity and identity not in active_by_identity:
                active_by_identity[identity] = record
        return active_by_identity

    def _repair_instagram_state_for_active_inventory(self, state: dict[str, object], active_records: list[dict[str, object]]) -> int:
        posts = state.get("posts") if isinstance(state.get("posts"), dict) else {}
        if not isinstance(posts, dict):
            return 0
        active_by_key = {str(record.get("inventory_key") or ""): record for record in active_records if str(record.get("inventory_key") or "")}
        active_by_identity = self._instagram_active_identity_map(active_records)
        active_by_title: dict[str, dict[str, object] | None] = {}
        for record in active_records:
            title_identity = self._instagram_inventory_identity({"card_title": record.get("card_title")})
            if not title_identity:
                continue
            if title_identity in active_by_title:
                active_by_title[title_identity] = None
            else:
                active_by_title[title_identity] = record
        repaired = 0
        now = datetime.now().isoformat(timespec="seconds")
        for old_key, entry in list(posts.items()):
            if not isinstance(entry, dict):
                continue
            status = str(entry.get("status") or "").strip().lower()
            if status == "posted":
                matched_by = str(entry.get("matched_by") or "").strip().lower()
                caption_title = self._instagram_inventory_identity({"card_title": entry.get("caption")})
                exact_title_match = active_by_title.get(caption_title)
                old_record = active_by_key.get(str(old_key))
                old_record_title = self._instagram_inventory_identity(
                    {"card_title": old_record.get("card_title") if isinstance(old_record, dict) else ""}
                )
                if (
                    matched_by == "title_tokens"
                    and caption_title
                    and exact_title_match is not None
                    and str(exact_title_match.get("inventory_key") or "") != str(old_key)
                    and old_record_title != caption_title
                ):
                    new_key = str(exact_title_match.get("inventory_key") or old_key)
                    repaired_entry = dict(entry)
                    repaired_entry.update(
                        {
                            "status": "posted",
                            "inventory_identity": self._instagram_inventory_identity(exact_title_match),
                            "card_title": str(exact_title_match.get("card_title") or repaired_entry.get("caption") or "").strip(),
                            "cert_number": scan_to_cert(exact_title_match.get("cert_number")),
                            "item_id": str(exact_title_match.get("item_id") or "").strip(),
                            "repaired_at": now,
                            "repair_reason": "exact_caption_replaced_fuzzy_title_mapping",
                            "previous_status": status,
                            "previous_inventory_key": str(old_key),
                            "previous_card_title": str(entry.get("card_title") or ""),
                        }
                    )
                    posts[new_key] = repaired_entry
                    posts.pop(old_key, None)
                    repaired += 1
                continue
            if status not in {"delete_review_needed", "post_error"}:
                continue
            media_id = str(entry.get("media_id") or "").strip()
            permalink = str(entry.get("permalink") or "").strip()
            if not media_id or not permalink:
                continue
            identity = self._instagram_post_entry_identity(entry)
            record = active_by_key.get(str(old_key)) or active_by_identity.get(identity)
            if record is None:
                title_identity = self._instagram_inventory_identity({"card_title": entry.get("card_title") or entry.get("caption")})
                title_match = active_by_title.get(title_identity)
                if title_match is not None:
                    record = title_match
            if record is None:
                continue
            new_key = str(record.get("inventory_key") or old_key)
            repaired_entry = dict(entry)
            repaired_entry.update(
                {
                    "status": "posted",
                    "inventory_identity": self._instagram_inventory_identity(record),
                    "card_title": str(record.get("card_title") or repaired_entry.get("card_title") or repaired_entry.get("caption") or "").strip(),
                    "cert_number": scan_to_cert(record.get("cert_number")) if isinstance(record, dict) else str(repaired_entry.get("cert_number") or ""),
                    "item_id": str(record.get("item_id") or "").strip() if isinstance(record, dict) else str(repaired_entry.get("item_id") or ""),
                    "repaired_at": now,
                    "repair_reason": "active_inventory_mapping_restored",
                    "previous_status": status,
                    "previous_inventory_key": str(old_key),
                }
            )
            repaired_entry.pop("delete_queued_at", None)
            posts[new_key] = repaired_entry
            if new_key != old_key:
                posts.pop(old_key, None)
            repaired += 1
        if repaired:
            state["posts"] = posts
            self._save_instagram_inventory_state(state)
            self._append_activity(
                "Instagram Inventory Repair",
                f"Restored {repaired} active Instagram inventory post mapping(s).",
                {"repaired": repaired},
            )
        return repaired

    def _instagram_inventory_plan(self) -> dict[str, object]:
        state = self._load_instagram_inventory_state()
        posts = state.get("posts") if isinstance(state.get("posts"), dict) else {}
        config = self._instagram_env_config()
        active_records = self._instagram_inventory_active_records()
        repair_state = getattr(self, "_repair_instagram_state_for_active_inventory", None)
        if callable(repair_state) and repair_state(state, active_records):
            posts = state.get("posts") if isinstance(state.get("posts"), dict) else {}
        active_by_key = {str(record.get("inventory_key") or ""): record for record in active_records if str(record.get("inventory_key") or "")}
        active_by_identity = self._instagram_active_identity_map(active_records)
        to_post: list[dict[str, object]] = []
        already_posted: list[dict[str, object]] = []
        missing_photos: list[dict[str, object]] = []
        missing_public_urls: list[dict[str, object]] = []
        cover_replacements: list[dict[str, object]] = []
        posted_identities: set[str] = set()

        for key, post_entry in posts.items():
            if not isinstance(post_entry, dict):
                continue
            if str(post_entry.get("status") or "").strip().lower() != "posted":
                continue
            identity = self._instagram_post_entry_identity(post_entry)
            if not identity and str(key) in active_by_key:
                identity = self._instagram_inventory_identity(active_by_key[str(key)])
            if identity:
                posted_identities.add(identity)

        for key, record in active_by_key.items():
            post_entry = posts.get(key) if isinstance(posts.get(key), dict) else {}
            if str(post_entry.get("media_id") or "").strip() and str(post_entry.get("status") or "").strip().lower() == "posted":
                paths = self._inventory_photo_paths_for_record(record)
                cover_photo = self._instagram_cover_photo_path(paths)
                expected_photo_id = self._instagram_inventory_photo_id(cover_photo) if cover_photo is not None else ""
                posted_photo_id = self._instagram_post_photo_id(post_entry)
                if expected_photo_id and posted_photo_id and expected_photo_id != posted_photo_id:
                    cover_replacements.append(
                        {
                            "inventory_key": key,
                            **post_entry,
                            "reason": "cover_photo_changed",
                            "expected_photo_path": str(cover_photo),
                            "expected_photo_id": expected_photo_id,
                        }
                    )
                already_posted.append(record)
                continue
            identity = self._instagram_inventory_identity(record)
            if identity and identity in posted_identities:
                already_posted.append(record)
                continue
            paths = instagram_inventory_photo_order(self._inventory_photo_paths_for_record(record))
            if hasattr(self, "_instagram_inventory_photo_is_postable"):
                paths = [path for path in paths if self._instagram_inventory_photo_is_postable(path)]
            if not paths:
                missing_photos.append(record)
                continue
            cover_photo = paths[0]
            if cover_photo is None:
                missing_photos.append(record)
                continue
            photo_urls = [self._instagram_inventory_photo_url(path, config) for path in paths[:10]]
            photo_url = photo_urls[0] if photo_urls else ""
            item = {
                "inventory_key": key,
                "record": record,
                "photo_path": cover_photo,
                "photo_paths": paths,
                "photo_count": len(paths),
                "photo_url": photo_url,
                "photo_urls": photo_urls,
                "caption": str(record.get("card_title") or "").strip(),
            }
            if not photo_url or any(not url for url in photo_urls):
                missing_public_urls.append(item)
            to_post.append(item)

        to_remove: list[dict[str, object]] = list(cover_replacements)
        duplicate_posts = state.get("duplicate_posts") if isinstance(state.get("duplicate_posts"), list) else []
        duplicate_media_ids: set[str] = set()
        for duplicate in duplicate_posts:
            if not isinstance(duplicate, dict):
                continue
            media_id = str(duplicate.get("media_id") or "").strip()
            if not media_id or media_id in duplicate_media_ids:
                continue
            duplicate_media_ids.add(media_id)
            to_remove.append({"inventory_key": str(duplicate.get("inventory_key") or ""), **duplicate})
        for key, post_entry in posts.items():
            if not isinstance(post_entry, dict):
                continue
            if key in active_by_key:
                continue
            identity = self._instagram_post_entry_identity(post_entry)
            if identity and identity in active_by_identity:
                continue
            media_id = str(post_entry.get("media_id") or "").strip()
            if media_id and media_id not in duplicate_media_ids:
                to_remove.append({"inventory_key": key, **post_entry})

        return {
            "config": config,
            "active_count": len(active_records),
            "posted_count": len(already_posted),
            "to_post": to_post,
            "to_remove": to_remove,
            "missing_photos": missing_photos,
            "missing_public_urls": missing_public_urls,
        }

    def _instagram_existing_media_posts(self, config: dict[str, str], limit: int = 500) -> list[dict[str, object]]:
        user_id = str(config.get("user_id") or "").strip()
        if not user_id:
            raise RuntimeError("Missing Instagram user ID.")
        posts: list[dict[str, object]] = []
        after = ""
        while len(posts) < limit:
            params: dict[str, object] = {
                "fields": "id,caption,media_type,media_url,permalink,timestamp",
                "limit": min(100, limit - len(posts)),
            }
            if after:
                params["after"] = after
            response = self._instagram_api_json(f"{user_id}/media", params)
            data = response.get("data") if isinstance(response.get("data"), list) else []
            posts.extend(item for item in data if isinstance(item, dict))
            paging = response.get("paging") if isinstance(response.get("paging"), dict) else {}
            cursors = paging.get("cursors") if isinstance(paging.get("cursors"), dict) else {}
            next_after = str(cursors.get("after") or "").strip()
            if not next_after or next_after == after or not data:
                break
            after = next_after
        return posts

    def _instagram_match_text_tokens(self, value: object) -> set[str]:
        text = str(value or "").lower()
        tokens = set(re.findall(r"[a-z0-9]+", text))
        stop_words = {
            "the", "and", "with", "for", "card", "cards", "auto", "rc", "rookie", "psa", "bgs",
            "cgc", "sgc", "gem", "mint", "auto", "autograph", "number", "serial", "refractor",
        }
        return {token for token in tokens if len(token) >= 3 and token not in stop_words}

    def _instagram_record_duplicate_media(
        self,
        state: dict[str, object],
        record: dict[str, object],
        post: dict[str, object],
        method: str,
        score: float,
        reason: str = "duplicate_inventory_post",
        keep_media_id: str = "",
        keep_post: dict[str, object] | None = None,
    ) -> bool:
        media_id = str(post.get("id") or post.get("media_id") or "").strip()
        if not media_id:
            return False
        duplicates = state.setdefault("duplicate_posts", [])
        if not isinstance(duplicates, list):
            duplicates = []
            state["duplicate_posts"] = duplicates
        if any(isinstance(item, dict) and str(item.get("media_id") or "").strip() == media_id for item in duplicates):
            return False
        duplicate = {
            "inventory_key": str(record.get("inventory_key") or ""),
            "inventory_identity": self._instagram_inventory_identity(record),
            "media_id": media_id,
            "caption": str(post.get("caption") or record.get("card_title") or "").strip(),
            "permalink": str(post.get("permalink") or ""),
            "photo_url": str(post.get("media_url") or ""),
            "detected_at": datetime.now().isoformat(timespec="seconds"),
            "reason": reason,
            "matched_by": method,
            "match_score": round(float(score), 3),
            "keep_media_id": keep_media_id,
        }
        if isinstance(keep_post, dict):
            duplicate.update(
                {
                    "keep_caption": str(keep_post.get("caption") or record.get("card_title") or "").strip(),
                    "keep_permalink": str(keep_post.get("permalink") or ""),
                    "keep_photo_url": str(keep_post.get("media_url") or ""),
                    "keep_timestamp": str(keep_post.get("timestamp") or ""),
                }
            )
        duplicates.append(duplicate)
        if len(duplicates) > 500:
            del duplicates[:-500]
        return True

    def _instagram_queue_duplicate_inventory_media(
        self,
        state: dict[str, object],
        media_items: list[dict[str, object]],
        active_records: list[dict[str, object]],
    ) -> int:
        posts = state.get("posts") if isinstance(state.get("posts"), dict) else {}
        if not isinstance(state.get("posts"), dict):
            state["posts"] = posts
        known_posted_media_ids = {
            str(entry.get("media_id") or "").strip()
            for entry in posts.values()
            if isinstance(entry, dict)
            and str(entry.get("status") or "").strip().lower() == "posted"
            and str(entry.get("media_id") or "").strip()
        }
        matches_by_key: dict[str, list[tuple[dict[str, object], dict[str, object], str, float]]] = {}
        for post in media_items:
            if not isinstance(post, dict):
                continue
            media_id = str(post.get("id") or post.get("media_id") or "").strip()
            if not media_id:
                continue
            record, method, score = self._instagram_find_inventory_match_for_post(post, active_records, set())
            if not record:
                continue
            key = str(record.get("inventory_key") or "").strip()
            if key:
                matches_by_key.setdefault(key, []).append((post, record, method, score))
        queued = 0
        for key, matches in matches_by_key.items():
            unique: list[tuple[dict[str, object], dict[str, object], str, float]] = []
            seen_ids: set[str] = set()
            for post, record, method, score in matches:
                media_id = str(post.get("id") or post.get("media_id") or "").strip()
                if media_id and media_id not in seen_ids:
                    unique.append((post, record, method, score))
                    seen_ids.add(media_id)
            if len(unique) < 2:
                continue
            known = [item for item in unique if str(item[0].get("id") or item[0].get("media_id") or "").strip() in known_posted_media_ids]
            keeper = known[0] if known else sorted(unique, key=lambda item: str(item[0].get("timestamp") or ""))[0]
            keep_post, keep_record, keep_method, keep_score = keeper
            keep_media_id = str(keep_post.get("id") or keep_post.get("media_id") or "").strip()
            existing = posts.get(key) if isinstance(posts.get(key), dict) else {}
            existing_status = str(existing.get("status") or "").strip().lower() if isinstance(existing, dict) else ""
            existing_media_id = str(existing.get("media_id") or "").strip() if isinstance(existing, dict) else ""
            if keep_media_id and (existing_status != "posted" or existing_media_id != keep_media_id):
                posts[key] = {
                    "status": "posted",
                    "media_id": keep_media_id,
                    "caption": str(keep_post.get("caption") or keep_record.get("card_title") or "").strip(),
                    "photo_url": str(keep_post.get("media_url") or ""),
                    "permalink": str(keep_post.get("permalink") or ""),
                    "posted_at": str(keep_post.get("timestamp") or datetime.now().isoformat(timespec="seconds")),
                    "imported_at": datetime.now().isoformat(timespec="seconds"),
                    "inventory_identity": self._instagram_inventory_identity(keep_record),
                    "card_title": str(keep_record.get("card_title") or ""),
                    "cert_number": scan_to_cert(keep_record.get("cert_number")),
                    "item_id": str(keep_record.get("item_id") or ""),
                    "matched_by": keep_method or "inventory_match",
                    "match_score": round(float(keep_score), 3),
                    "imported_from": "instagram_duplicate_keeper",
                }
                known_posted_media_ids.add(keep_media_id)
            for post, record, method, score in unique:
                media_id = str(post.get("id") or post.get("media_id") or "").strip()
                if not media_id or media_id == keep_media_id:
                    continue
                if self._instagram_record_duplicate_media(
                    state,
                    record,
                    post,
                    method or "inventory_match",
                    score,
                    reason="duplicate_live_inventory_post",
                    keep_media_id=keep_media_id,
                    keep_post=keep_post,
                ):
                    queued += 1
        return queued

    def _instagram_find_inventory_match_for_post(
        self,
        post: dict[str, object],
        active_records: list[dict[str, object]],
        used_keys: set[str],
    ) -> tuple[dict[str, object] | None, str, float]:
        text = " ".join(str(post.get(field) or "") for field in ("caption", "ocr_text", "permalink"))
        compact_text = re.sub(r"\D+", "", text)
        lowered_text = text.lower()
        normalized_post_text = re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", text.lower())).strip()
        best_record: dict[str, object] | None = None
        best_method = ""
        best_score = 0.0

        for record in active_records:
            key = str(record.get("inventory_key") or "")
            if not key or key in used_keys:
                continue
            cert = scan_to_cert(record.get("cert_number"))
            item_id = str(record.get("item_id") or "").strip().lower()
            if cert and len(cert) >= 5 and cert in compact_text:
                return record, "cert_number", 1.0
            if item_id and item_id in lowered_text:
                return record, "item_id", 1.0

            title = str(record.get("card_title") or "").strip()
            normalized_title = re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", title.lower())).strip()
            if normalized_title and normalized_title in normalized_post_text:
                score = 0.97 + min(0.02, len(normalized_title) / 10000)
                if score > best_score:
                    best_record = record
                    best_method = "title_exact"
                    best_score = score

        return best_record, best_method, best_score

    def _instagram_ocr_post_text(self, post: dict[str, object], client=None) -> str:
        media_type = str(post.get("media_type") or "").upper()
        media_url = str(post.get("media_url") or "").strip()
        if not media_url or media_type == "VIDEO":
            return ""
        if genai is None or genai_types is None:
            return ""
        self._load_photo_env()
        api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
        if not api_key:
            return ""
        try:
            with urllib.request.urlopen(media_url, timeout=45) as response:
                image_bytes = response.read()
                content_type = response.headers.get("content-type") or "image/jpeg"
        except Exception:
            return ""
        if not content_type.startswith("image/"):
            content_type = "image/jpeg"
        try:
            ocr_client = client or make_photo_ocr_client(api_key)
            prompt = (
                "Read this trading card inventory Instagram post image. Extract visible card-identifying text only. "
                "Prioritize grading company, cert number, player, year, set, parallel, card number, grade, and any slab label text. "
                "Return plain text only. Do not describe the background or invent missing details."
            )
            response = ocr_client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[
                    prompt,
                    genai_types.Part.from_bytes(data=image_bytes, mime_type=content_type),
                ],
                config=genai_types.GenerateContentConfig(
                    thinking_config=genai_types.ThinkingConfig(thinking_budget=0),
                    max_output_tokens=450,
                    temperature=0,
                ),
            )
            return str(response.text or "").strip()
        except Exception:
            return ""

    def _instagram_import_existing_posts(self, limit: int = 500, use_ocr: bool = True) -> dict[str, object]:
        state = self._load_instagram_inventory_state()
        posts = state.setdefault("posts", {})
        if not isinstance(posts, dict):
            posts = {}
            state["posts"] = posts
        config = self._instagram_env_config()
        active_records = self._instagram_inventory_active_records()
        known_media_ids = {
            str(entry.get("media_id") or "").strip()
            for entry in posts.values()
            if isinstance(entry, dict) and str(entry.get("media_id") or "").strip()
        }
        used_keys = {
            str(key)
            for key, entry in posts.items()
            if isinstance(entry, dict) and str(entry.get("status") or "").strip().lower() == "posted"
        }
        media_items = self._instagram_existing_media_posts(config, limit=limit)
        imported = 0
        already_known = 0
        queue_duplicate_media = getattr(self, "_instagram_queue_duplicate_inventory_media", None)
        duplicates_found = queue_duplicate_media(state, media_items, active_records) if callable(queue_duplicate_media) else 0
        queued_duplicate_media_ids = {
            str(item.get("media_id") or "").strip()
            for item in state.get("duplicate_posts", [])
            if isinstance(item, dict) and str(item.get("media_id") or "").strip()
        }
        known_media_ids.update(queued_duplicate_media_ids)
        known_media_ids.update(
            str(entry.get("media_id") or "").strip()
            for entry in posts.values()
            if isinstance(entry, dict) and str(entry.get("media_id") or "").strip()
        )
        used_keys.update(
            str(key)
            for key, entry in posts.items()
            if isinstance(entry, dict) and str(entry.get("status") or "").strip().lower() == "posted"
        )
        ocr_attempted = 0
        ocr_matched = 0
        unmatched: list[dict[str, object]] = []
        ocr_client = None

        for post in media_items:
            media_id = str(post.get("id") or "").strip()
            if not media_id:
                continue
            if media_id in known_media_ids:
                already_known += 1
                continue
            record, method, score = self._instagram_find_inventory_match_for_post(post, active_records, used_keys)
            duplicate_record: dict[str, object] | None = None
            duplicate_method = ""
            duplicate_score = 0.0
            if not record:
                candidate, candidate_method, candidate_score = self._instagram_find_inventory_match_for_post(post, active_records, set())
                candidate_key = str(candidate.get("inventory_key") or "") if isinstance(candidate, dict) else ""
                if candidate_key and candidate_key in used_keys:
                    duplicate_record = candidate
                    duplicate_method = candidate_method
                    duplicate_score = candidate_score
            if not record and use_ocr:
                if ocr_client is None and genai is not None and genai_types is not None:
                    if hasattr(self, "_load_photo_env"):
                        self._load_photo_env()
                    api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
                    if api_key:
                        try:
                            ocr_client = make_photo_ocr_client(api_key)
                        except Exception:
                            ocr_client = False
                if ocr_client:
                    ocr_attempted += 1
                    events = getattr(self, "events", None)
                    if events is not None:
                        events.put(("status", f"OCR reading Instagram post {ocr_attempted}: {str(post.get('permalink') or media_id)}"))
                    ocr_text = self._instagram_ocr_post_text(post, client=ocr_client)
                    if ocr_text:
                        post = {**post, "ocr_text": ocr_text}
                        record, method, score = self._instagram_find_inventory_match_for_post(post, active_records, used_keys)
                        if record:
                            method = f"ocr_{method}"
                            ocr_matched += 1
                        else:
                            candidate, candidate_method, candidate_score = self._instagram_find_inventory_match_for_post(post, active_records, set())
                            candidate_key = str(candidate.get("inventory_key") or "") if isinstance(candidate, dict) else ""
                            if candidate_key and candidate_key in used_keys:
                                duplicate_record = candidate
                                duplicate_method = f"ocr_{candidate_method}"
                                duplicate_score = candidate_score
            if not record and duplicate_record:
                self._instagram_record_duplicate_post(state, duplicate_record, post, duplicate_method, duplicate_score)
                known_media_ids.add(media_id)
                duplicates_found += 1
                continue
            if not record:
                unmatched.append(
                    {
                        "media_id": media_id,
                        "caption": str(post.get("caption") or "")[:240],
                        "ocr_text": str(post.get("ocr_text") or "")[:240],
                        "permalink": str(post.get("permalink") or ""),
                    }
                )
                continue
            key = str(record.get("inventory_key") or "")
            caption = str(post.get("caption") or record.get("card_title") or "").strip()
            posts[key] = {
                "status": "posted",
                "media_id": media_id,
                "caption": caption,
                "inventory_identity": self._instagram_inventory_identity(record),
                "card_title": str(record.get("card_title") or "").strip(),
                "cert_number": scan_to_cert(record.get("cert_number")),
                "item_id": str(record.get("item_id") or "").strip(),
                "photo_url": str(post.get("media_url") or ""),
                "permalink": str(post.get("permalink") or ""),
                "posted_at": str(post.get("timestamp") or ""),
                "imported_at": datetime.now().isoformat(timespec="seconds"),
                "imported_from": "instagram_existing_posts",
                "matched_by": method,
                "match_score": round(float(score), 3),
            }
            known_media_ids.add(media_id)
            used_keys.add(key)
            imported += 1

        self._save_instagram_inventory_state(state)
        self._append_activity(
            "Instagram Inventory Import",
            f"Imported {imported} existing Instagram post(s); {len(unmatched)} unmatched.",
            {
                "imported": imported,
                "already_known": already_known,
                "duplicates_found": duplicates_found,
                "ocr_attempted": ocr_attempted,
                "ocr_matched": ocr_matched,
                "unmatched": unmatched[:12],
            },
        )
        return {
            "imported": imported,
            "already_known": already_known,
            "duplicates_found": duplicates_found,
            "ocr_attempted": ocr_attempted,
            "ocr_matched": ocr_matched,
            "unmatched": unmatched,
            "total_seen": len(media_items),
        }

    def open_instagram_inventory_sync(self) -> None:
        if not self._personal_instagram_sync_enabled():
            messagebox.showinfo("Instagram Sync", "Personal Instagram sync is disabled on this LUCAS install.")
            return
        plan = self._instagram_inventory_plan()
        popup = tk.Toplevel(self)
        popup.title("Instagram Inventory Sync")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.geometry("980x660")
        popup.minsize(900, 600)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(14, 12))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Instagram Inventory Sync", style="Panel.TLabel", font=("Segoe UI Semibold", 14)).pack(anchor="w")

        summary_var = tk.StringVar()
        ttk.Label(frame, textvariable=summary_var, style="Muted.TLabel").pack(anchor="w", pady=(6, 10))

        columns = ("action", "card", "id", "photo", "detail")
        table_frame = ttk.Frame(frame, style="Panel.TFrame")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=14)
        headings = {
            "action": "Action",
            "card": "Card",
            "id": "Cert / Item",
            "photo": "Photo",
            "detail": "Detail",
        }
        widths = {"action": 120, "card": 320, "id": 120, "photo": 180, "detail": 220}
        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=widths[column], anchor="w", stretch=column in {"card", "detail"})
        y_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        post_items: dict[str, dict[str, object]] = {}
        remove_items: dict[str, dict[str, object]] = {}

        def reload_plan() -> None:
            nonlocal plan
            plan = self._instagram_inventory_plan()
            post_items.clear()
            remove_items.clear()
            tree.delete(*tree.get_children())
            for index, item in enumerate(plan["to_post"]):
                record = item["record"]
                iid = f"post:{index}"
                tree.insert(
                    "",
                    tk.END,
                    iid=iid,
                    values=(
                        "Post",
                        record.get("card_title") or "",
                        record.get("cert_number") or record.get("item_id") or "",
                        Path(str(item["photo_path"])).name,
                        "Needs public photo URL" if not item.get("photo_url") else "Ready",
                    ),
                )
                post_items[iid] = item
            for index, item in enumerate(plan["to_remove"]):
                detail = "No longer active in inventory"
                if str(item.get("status") or "").strip().lower() == "delete_review_needed":
                    detail = "Manual delete needed; queued for review"
                elif str(item.get("reason") or "").strip().lower() == "cover_photo_changed":
                    detail = "Manual delete needed; cover photo changed"
                else:
                    detail = "Manual delete needed"
                iid = f"remove:{index}"
                tree.insert("", tk.END, iid=iid, values=("Remove", item.get("caption") or item.get("title") or "", item.get("media_id") or "", "", detail))
                remove_items[iid] = item
            for record in plan["missing_photos"]:
                tree.insert(
                    "",
                    tk.END,
                    values=(
                        "Missing Photo",
                        record.get("card_title") or "",
                        record.get("cert_number") or record.get("item_id") or "",
                        "",
                        "Active inventory row has no attached photo",
                    ),
                )
            summary_var.set(
                f"Active inventory: {plan['active_count']} | Posted: {plan['posted_count']} | "
                f"To post: {len(plan['to_post'])} | To remove: {len(plan['to_remove'])} | "
                f"Missing photos: {len(plan['missing_photos'])} | Daily auto cap: {self._instagram_daily_post_limit(plan.get('config') if isinstance(plan.get('config'), dict) else None)}"
            )

        def post_one_test_item() -> None:
            selected = [iid for iid in tree.selection() if iid in post_items]
            item = post_items[selected[0]] if selected else next((candidate for candidate in plan.get("to_post") or [] if isinstance(candidate, dict) and candidate.get("photo_url")), None)
            if not item:
                messagebox.showinfo("Instagram Sync", "No ready item with a public photo URL is available to test.")
                return
            if not item.get("photo_url"):
                messagebox.showinfo("Instagram Sync", "Choose a row marked Ready before posting a test item.")
                return
            test_plan = {**plan, "to_post": [item], "to_remove": [], "missing_public_urls": []}
            record = item.get("record") if isinstance(item.get("record"), dict) else {}
            title = str(record.get("card_title") or item.get("caption") or "selected card")
            if not messagebox.askyesno("Instagram Sync", f"Post one test item to Instagram?\n\n{title}"):
                return
            worker = threading.Thread(target=self._instagram_inventory_sync_worker, args=(test_plan,), daemon=True)
            worker.start()
            popup.after(1000, reload_plan)

        def selected_remove_items() -> list[dict[str, object]]:
            return [remove_items[iid] for iid in tree.selection() if iid in remove_items]

        def open_manual_delete_target() -> None:
            selected = selected_remove_items()
            if not selected:
                messagebox.showinfo("Instagram Sync", "Select a Remove row first. Ready rows are planned posts and do not have an existing Instagram target.")
                return
            item = selected[0]
            url = str(item.get("permalink") or "").strip() or self._instagram_profile_url()
            if not url:
                messagebox.showinfo("Instagram Sync", "No Instagram URL is configured. Open the inventory profile manually.")
                return
            webbrowser.open(url)

        def copy_manual_delete_title() -> None:
            selected = selected_remove_items()
            if not selected:
                messagebox.showinfo("Instagram Sync", "Select a Remove row first.")
                return
            item = selected[0]
            text = str(item.get("caption") or item.get("title") or item.get("card_title") or item.get("media_id") or "").strip()
            self.clipboard_clear()
            self.clipboard_append(text)
            self.status_var.set("Copied Instagram delete queue title.")

        def mark_manual_deleted() -> None:
            selected = selected_remove_items()
            if not selected:
                messagebox.showinfo("Instagram Sync", "Select one or more Remove rows to mark manually deleted.")
                return
            if not messagebox.askyesno("Instagram Sync", f"Mark {len(selected)} Instagram post(s) manually deleted in LUCAS?\n\nOnly do this after you deleted them on Instagram."):
                return
            marked = self._mark_instagram_posts_manually_deleted(selected)
            reload_plan()
            skipped = max(0, len(selected) - marked)
            if skipped:
                messagebox.showwarning(
                    "Instagram Sync",
                    f"Marked {marked} post(s) manually deleted.\n\nSkipped {skipped} active inventory post(s) because LUCAS could not find a separate keeper post.",
                )
            else:
                messagebox.showinfo("Instagram Sync", f"Marked {marked} post(s) manually deleted.")

        def import_existing_posts() -> None:
            config = self._instagram_env_config()
            if not str(config.get("user_id") or "").strip() or not str(config.get("access_token") or "").strip():
                messagebox.showerror("Instagram Import", "Missing Instagram user ID or access token in .env.")
                return
            if not messagebox.askyesno("Instagram Import", "Scan existing Instagram posts and match them to active LUCAS inventory?\n\nBlank-caption posts will be OCR-read from the image, so this can take a few minutes. LUCAS will only record confident matches and leave unclear posts alone."):
                return

            def worker() -> None:
                try:
                    result = self._instagram_import_existing_posts()
                    message = (
                        f"Imported: {result['imported']}\n"
                        f"Already known: {result['already_known']}\n"
                        f"Duplicates queued for removal: {result['duplicates_found']}\n"
                        f"OCR attempted: {result['ocr_attempted']}\n"
                        f"OCR matched: {result['ocr_matched']}\n"
                        f"Unmatched: {len(result['unmatched'])}\n"
                        f"Posts scanned: {result['total_seen']}"
                    )
                    self.events.put(("status", f"Instagram import matched {result['imported']} existing post(s); {len(result['unmatched'])} unmatched."))
                    popup.after(0, lambda: (reload_plan(), messagebox.showinfo("Instagram Import", message)))
                except Exception as error:
                    error_text = str(error)
                    popup.after(0, lambda: messagebox.showerror("Instagram Import", error_text))

            threading.Thread(target=worker, daemon=True).start()

        actions = ttk.Frame(frame, style="Panel.TFrame")
        actions.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(actions, text="Refresh", command=reload_plan, style="Soft.TButton").pack(side=tk.LEFT)
        ttk.Button(actions, text="Import Posts", command=import_existing_posts, style="Primary.TButton").pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(actions, text="Test Post", command=post_one_test_item, style="Primary.TButton").pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(actions, text="Live Sync", command=lambda: self._run_instagram_inventory_sync(plan, popup, reload_plan), style="Primary.TButton").pack(side=tk.LEFT, padx=(8, 0))
        delete_actions = ttk.Frame(frame, style="Panel.TFrame")
        delete_actions.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(delete_actions, text="Open Remove Target", command=open_manual_delete_target, style="Soft.TButton").pack(side=tk.LEFT)
        ttk.Button(delete_actions, text="Copy Title", command=copy_manual_delete_title, style="Soft.TButton").pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(delete_actions, text="Mark Deleted", command=mark_manual_deleted, style="Primary.TButton").pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(delete_actions, text="Close", command=popup.destroy, style="Soft.TButton").pack(side=tk.RIGHT)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        reload_plan()

    def _instagram_api_json(self, endpoint: str, params: dict[str, object] | None = None, method: str = "GET") -> dict[str, object]:
        params = dict(params or {})
        token = str(params.pop("access_token", "") or self._instagram_env_config().get("access_token") or "").strip()
        if token:
            params["access_token"] = token
        data = urllib.parse.urlencode(params).encode("utf-8")
        url = f"https://graph.instagram.com/v21.0/{endpoint.lstrip('/')}"
        request: urllib.request.Request
        if method.upper() == "GET":
            request = urllib.request.Request(url + ("?" + data.decode("utf-8") if data else ""))
        else:
            request = urllib.request.Request(url, data=data, method=method.upper())
        try:
            with urllib.request.urlopen(request, timeout=45) as response:
                raw = response.read().decode("utf-8")
        except urllib.error.HTTPError as error:
            body = error.read().decode("utf-8", errors="replace")
            raise RuntimeError(body or str(error)) from error
        return json.loads(raw) if raw else {}

    def _run_instagram_inventory_sync(self, plan: dict[str, object], popup: tk.Toplevel, refresh_callback) -> None:
        config = plan.get("config") if isinstance(plan.get("config"), dict) else self._instagram_env_config()
        if not str(config.get("user_id") or "").strip() or not str(config.get("access_token") or "").strip():
            messagebox.showerror("Instagram Sync", "Missing Instagram user ID or access token in .env.")
            return
        ready_candidates = [item for item in plan.get("to_post") or [] if isinstance(item, dict) and instagram_ready_photo_urls(item)]
        quota = self._instagram_content_publishing_limit(config)
        meta_remaining = quota.get("remaining") if isinstance(quota.get("remaining"), int) else None
        if ready_candidates and meta_remaining is None:
            messagebox.showerror(
                "Instagram Sync",
                "Could not verify Meta publishing quota, so LUCAS will not post new cards.\n\n"
                f"{quota.get('error') or 'Unknown quota check error.'}",
            )
            return
        sync_plan, ready_total, post_limit = self._instagram_limited_manual_sync_plan(plan, meta_remaining)
        if not sync_plan.get("to_post") and not sync_plan.get("to_remove"):
            if ready_candidates and meta_remaining == 0:
                messagebox.showinfo(
                    "Instagram Sync",
                    f"Meta publishing quota is full.\n\nQuota usage: {quota.get('quota_usage')}/{quota.get('limit')}\nTry again after the 24-hour rolling window frees up.",
                )
                return
            missing_public = plan.get("missing_public_urls") or []
            if missing_public:
                messagebox.showinfo(
                    "Instagram Sync",
                    "Preview is ready, but live posting needs public photo URLs.\n\n"
                    "Add LUCAS_INSTAGRAM_PUBLIC_BRIDGE_URL or LUCAS_INSTAGRAM_PUBLIC_PHOTO_BASE_URL to .env after we set up a hosted photo URL source.",
                )
                return
            messagebox.showinfo(
                "Instagram Sync",
                "No ready Instagram inventory changes are available to sync.",
            )
            return
        post_count = len(sync_plan.get("to_post") or [])
        remove_count = len(sync_plan.get("to_remove") or [])
        skipped_count = max(0, ready_total - post_count)
        message = (
            "Run live Instagram sync?\n\n"
            f"Post {post_count} of {ready_total} ready card(s)."
            f"\nQueue {remove_count} old post(s) for manual deletion review."
            f"\n\nManual live posting is capped at {post_limit} card(s) for this run."
        )
        if meta_remaining is not None:
            message += f"\nMeta quota usage: {quota.get('quota_usage')}/{quota.get('limit')} used; {meta_remaining} remaining."
        if skipped_count:
            message += f"\n{skipped_count} ready card(s) will remain for the next batch."
        if not messagebox.askyesno("Instagram Sync", message):
            return
        worker = threading.Thread(target=self._instagram_inventory_sync_worker, args=(sync_plan,), daemon=True)
        worker.start()
        popup.after(1000, refresh_callback)

    def _instagram_publish_media_with_retry(self, user_id: str, creation_id: str, caption: str) -> dict[str, object]:
        last_error = ""
        for attempt in range(1, 7):
            try:
                return self._instagram_api_json(
                    f"{user_id}/media_publish",
                    {"creation_id": creation_id},
                    method="POST",
                )
            except RuntimeError as error:
                last_error = str(error)
                lowered = last_error.lower()
                if "media is not ready" not in lowered and "media id is not available" not in lowered and "cannot publish" not in lowered:
                    raise
                if attempt >= 6:
                    break
                self.events.put(("status", f"Instagram media not ready yet for {caption[:60]}; retrying publish ({attempt}/6)."))
                time.sleep(5)
        raise RuntimeError(last_error or f"Instagram media was not ready for publishing: {caption}")

    def _instagram_delete_media_post(self, media_id: str) -> dict[str, object]:
        media_id = str(media_id or "").strip()
        if not media_id:
            raise RuntimeError("Missing Instagram media id.")
        response = self._instagram_api_json(media_id, {}, method="DELETE")
        if isinstance(response, dict) and response.get("success") is False:
            raise RuntimeError(json.dumps(response))
        return response if isinstance(response, dict) else {}

    def _record_instagram_removed_post(self, state: dict[str, object], item: dict[str, object]) -> None:
        removed_posts = state.setdefault("removed_posts", [])
        if not isinstance(removed_posts, list):
            removed_posts = []
            state["removed_posts"] = removed_posts
        removed_posts.append(
            {
                "inventory_key": str(item.get("inventory_key") or ""),
                "media_id": str(item.get("media_id") or ""),
                "caption": str(item.get("caption") or item.get("title") or ""),
                "permalink": str(item.get("permalink") or ""),
                "removed_at": datetime.now().isoformat(timespec="seconds"),
                "reason": str(item.get("reason") or "inventory_not_active"),
            }
        )
        if len(removed_posts) > 250:
            del removed_posts[:-250]

    def _mark_instagram_posts_manually_deleted(self, items: list[dict[str, object]]) -> int:
        if not items:
            return 0
        state = self._load_instagram_inventory_state()
        posts = state.setdefault("posts", {})
        if not isinstance(posts, dict):
            posts = {}
            state["posts"] = posts
        duplicates = state.get("duplicate_posts")
        if not isinstance(duplicates, list):
            duplicates = []
        active_keys = {
            str(record.get("inventory_key") or "").strip()
            for record in self._instagram_inventory_active_records()
            if str(record.get("inventory_key") or "").strip()
        }
        marked = 0
        for raw_item in items:
            if not isinstance(raw_item, dict):
                continue
            key = str(raw_item.get("inventory_key") or "").strip()
            media_id = str(raw_item.get("media_id") or "").strip()
            if not key and not media_id:
                continue
            tracked = posts.get(key) if key and isinstance(posts.get(key), dict) else {}
            item = {**raw_item}
            if isinstance(tracked, dict):
                item = {**tracked, **item}
            tracked_media_id = str(tracked.get("media_id") or "").strip() if isinstance(tracked, dict) else ""
            keep_media_id = str(raw_item.get("keep_media_id") or "").strip()
            source_reason = str(raw_item.get("reason") or "").strip().lower()
            has_separate_keeper = bool(keep_media_id and media_id and media_id != keep_media_id)
            has_tracked_keeper = bool(tracked_media_id and media_id and media_id != tracked_media_id)
            is_cover_replacement = source_reason == "cover_photo_changed"
            if key in active_keys and not has_separate_keeper and not has_tracked_keeper and not is_cover_replacement:
                continue
            item["original_reason"] = str(raw_item.get("reason") or item.get("reason") or "")
            item["reason"] = "manual_delete_confirmed"
            self._record_instagram_removed_post(state, item)
            if key:
                if keep_media_id and media_id and media_id != keep_media_id:
                    if not tracked_media_id or tracked_media_id == media_id:
                        posts[key] = {
                            "status": "posted",
                            "media_id": keep_media_id,
                            "caption": str(raw_item.get("keep_caption") or raw_item.get("caption") or raw_item.get("title") or "").strip(),
                            "photo_url": str(raw_item.get("keep_photo_url") or ""),
                            "permalink": str(raw_item.get("keep_permalink") or ""),
                            "posted_at": str(raw_item.get("keep_timestamp") or datetime.now().isoformat(timespec="seconds")),
                            "imported_at": datetime.now().isoformat(timespec="seconds"),
                            "inventory_identity": str(raw_item.get("inventory_identity") or ""),
                            "card_title": str(raw_item.get("title") or raw_item.get("caption") or ""),
                            "matched_by": str(raw_item.get("matched_by") or "duplicate_keeper"),
                            "match_score": raw_item.get("match_score", ""),
                            "imported_from": "manual_delete_duplicate_keeper",
                        }
                elif not media_id or not tracked_media_id or media_id == tracked_media_id:
                    posts.pop(key, None)
            if media_id:
                duplicates = [
                    duplicate
                    for duplicate in duplicates
                    if not isinstance(duplicate, dict) or str(duplicate.get("media_id") or "").strip() != media_id
                ]
            marked += 1
        state["duplicate_posts"] = duplicates
        self._save_instagram_inventory_state(state)
        self._append_activity("Instagram Manual Delete", f"Marked {marked} Instagram inventory post(s) manually deleted.", {"marked": marked})
        return marked

    def _instagram_inventory_sync_worker(self, plan: dict[str, object]) -> None:
        started = time.perf_counter()
        try:
            import_existing = getattr(self, "_instagram_import_existing_posts", None)
            if callable(import_existing):
                import_existing(limit=500, use_ocr=False)
        except Exception as error:
            events = getattr(self, "events", None)
            if events is not None:
                events.put(("status", f"Instagram pre-sync duplicate scan skipped: {str(error)[:120]}"))
        state = self._load_instagram_inventory_state()
        posts = state.setdefault("posts", {})
        current_active_records = self._instagram_inventory_active_records()
        active_by_key = {str(record.get("inventory_key") or ""): record for record in current_active_records if str(record.get("inventory_key") or "")}
        active_by_identity = self._instagram_active_identity_map(current_active_records)
        posted = 0
        removed = 0
        queued_removals = 0
        errors: list[str] = []
        try:
            for item in plan.get("to_post") or []:
                if not isinstance(item, dict):
                    continue
                key = str(item.get("inventory_key") or "")
                caption = str(item.get("caption") or "").strip()
                photo_url = str(item.get("photo_url") or "").strip()
                if not key or not caption or not photo_url:
                    continue
                record = item.get("record") if isinstance(item.get("record"), dict) else {}
                current_record = active_by_key.get(key)
                if not current_record:
                    continue
                identity = self._instagram_inventory_identity(current_record)
                if identity and self._instagram_inventory_identity(record) and identity != self._instagram_inventory_identity(record):
                    continue
                caption = str(current_record.get("card_title") or caption).strip()
                existing_entry = posts.get(key) if isinstance(posts.get(key), dict) else {}
                if str(existing_entry.get("media_id") or "").strip() and str(existing_entry.get("status") or "").strip().lower() == "posted":
                    continue
                already_posted_identity = False
                if identity:
                    for existing_key, existing_post in posts.items():
                        if not isinstance(existing_post, dict) or existing_key == key:
                            continue
                        if str(existing_post.get("status") or "").strip().lower() != "posted":
                            continue
                        if self._instagram_post_entry_identity(existing_post) == identity:
                            already_posted_identity = True
                            break
                if already_posted_identity:
                    continue
                try:
                    photo_urls = instagram_ready_photo_urls(item)
                    if not photo_urls:
                        photo_urls = [photo_url]
                    photo_paths = item.get("photo_paths") if isinstance(item.get("photo_paths"), list) else []
                    if not photo_paths:
                        photo_paths = [item.get("photo_path")]
                    photo_ids = [
                        self._instagram_inventory_photo_id(Path(str(path or "")))
                        for path in photo_paths[: len(photo_urls)]
                    ]
                    child_creation_ids: list[str] = []
                    media_type = "IMAGE"
                    if len(photo_urls) > 1:
                        for url in photo_urls[:10]:
                            child_response = self._instagram_api_json(
                                f"{plan['config']['user_id']}/media",
                                {"image_url": url, "is_carousel_item": "true"},
                                method="POST",
                            )
                            child_id = str(child_response.get("id") or "").strip()
                            if not child_id:
                                raise RuntimeError(f"Instagram did not return a carousel child id for {caption}.")
                            child_creation_ids.append(child_id)
                        create_response = self._instagram_api_json(
                            f"{plan['config']['user_id']}/media",
                            {"media_type": "CAROUSEL", "children": ",".join(child_creation_ids), "caption": caption},
                            method="POST",
                        )
                        media_type = "CAROUSEL"
                    else:
                        create_response = self._instagram_api_json(
                            f"{plan['config']['user_id']}/media",
                            {"image_url": photo_urls[0], "caption": caption},
                            method="POST",
                        )
                    creation_id = str(create_response.get("id") or "").strip()
                    if not creation_id:
                        raise RuntimeError(f"Instagram did not return a creation id for {caption}.")
                    publish_response = self._instagram_publish_media_with_retry(str(plan["config"]["user_id"]), creation_id, caption)
                    media_id = str(publish_response.get("id") or "").strip()
                    permalink = ""
                    if media_id:
                        try:
                            media_details = self._instagram_api_json(media_id, {"fields": "permalink"})
                            permalink = str(media_details.get("permalink") or "").strip()
                        except Exception:
                            permalink = ""
                    posts[key] = {
                        "status": "posted",
                        "media_id": media_id,
                        "creation_id": creation_id,
                        "caption": caption,
                        "inventory_identity": identity,
                        "card_title": str(current_record.get("card_title") or "").strip(),
                        "cert_number": scan_to_cert(current_record.get("cert_number")) if isinstance(current_record, dict) else "",
                        "item_id": str(current_record.get("item_id") or "").strip() if isinstance(current_record, dict) else "",
                        "photo_id": photo_ids[0] if photo_ids else self._instagram_inventory_photo_id(Path(str(item.get("photo_path") or ""))),
                        "photo_ids": photo_ids,
                        "photo_url": photo_urls[0],
                        "photo_urls": photo_urls,
                        "media_type": media_type,
                        "child_creation_ids": child_creation_ids,
                        "permalink": permalink,
                        "posted_at": datetime.now().isoformat(timespec="seconds"),
                    }
                    posted += 1
                except Exception as error:
                    error_text = str(error)
                    errors.append(f"{caption or key}: {error_text}")
                    posts[key] = {
                        "status": "post_error",
                        "caption": caption,
                        "inventory_identity": identity,
                        "card_title": str(current_record.get("card_title") or "").strip(),
                        "cert_number": scan_to_cert(current_record.get("cert_number")) if isinstance(current_record, dict) else "",
                        "item_id": str(current_record.get("item_id") or "").strip() if isinstance(current_record, dict) else "",
                        "photo_url": photo_url,
                        "photo_urls": item.get("photo_urls") if isinstance(item.get("photo_urls"), list) else [photo_url],
                        "post_error": error_text[:500],
                        "post_error_at": datetime.now().isoformat(timespec="seconds"),
                    }
                    self.events.put(("status", f"Instagram skipped {caption[:60] or key}: {error_text[:120]}"))
                self._save_instagram_inventory_state(state)

            for item in plan.get("to_remove") or []:
                if not isinstance(item, dict):
                    continue
                key = str(item.get("inventory_key") or "")
                media_id = str(item.get("media_id") or "").strip()
                if not key or not media_id:
                    continue
                item_identity = str(item.get("inventory_identity") or "").strip() or self._instagram_post_entry_identity(item)
                is_duplicate_removal = str(item.get("reason") or "").strip().lower() == "duplicate_inventory_post"
                tracked_entry = posts.get(key) if isinstance(posts.get(key), dict) else {}
                tracked_media_id = str(tracked_entry.get("media_id") or "").strip()
                if not is_duplicate_removal:
                    if key in active_by_key:
                        continue
                    if item_identity and item_identity in active_by_identity:
                        continue
                if tracked_media_id == media_id:
                    entry = posts.get(key) if isinstance(posts.get(key), dict) else {}
                    entry["status"] = "delete_review_needed"
                    entry.pop("delete_error", None)
                    entry["delete_queued_at"] = datetime.now().isoformat(timespec="seconds")
                    posts[key] = entry
                else:
                    duplicates = state.get("duplicate_posts")
                    if isinstance(duplicates, list):
                        for duplicate in duplicates:
                            if isinstance(duplicate, dict) and str(duplicate.get("media_id") or "").strip() == media_id:
                                duplicate["status"] = "delete_review_needed"
                                duplicate.pop("delete_error", None)
                                duplicate["delete_queued_at"] = datetime.now().isoformat(timespec="seconds")
                queued_removals += 1
                self._save_instagram_inventory_state(state)
        except Exception as error:
            errors.append(str(error))
        self._append_activity(
            "Instagram Inventory Sync",
            f"Instagram inventory sync posted {posted}, removed {removed}, queued {queued_removals}, errors {len(errors)}.",
            {"posted": posted, "removed": removed, "queued_removals": queued_removals, "errors": errors[:3]},
        )
        self.events.put(("status", f"Instagram sync posted {posted}, removed {removed}, queued {queued_removals}, errors {len(errors)}."))
        record_performance_event("instagram.inventory_sync", started, f"posted={posted} removed={removed} queued={queued_removals} errors={len(errors)}", force=True)

    def _mark_inventory_records_moved_to_company(self, moved_keys: set[str]) -> None:
        if not moved_keys:
            return
        ledger = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        removed = [record for record in ledger if str(record.get("inventory_key") or "") in moved_keys]
        kept = [record for record in ledger if str(record.get("inventory_key") or "") not in moved_keys]
        if len(kept) != len(ledger):
            self._save_inventory_ledger(kept)
            cleanup = getattr(self, "_delete_inventory_photo_files_for_removed_records", None)
            if callable(cleanup):
                cleanup(removed, kept)

    def _safe_inventory_photo_path(self, path_value: object) -> Path | None:
        safe_candidates = getattr(self, "_inventory_photo_safe_candidates", None)
        if callable(safe_candidates):
            candidates = safe_candidates(path_value)
        else:
            path = Path(str(path_value or "")).expanduser()
            if not path.is_absolute():
                path = self._inventory_photo_source_folder() / path
            candidates = [path]
        for path in candidates:
            if path.exists():
                return path
        return None

    def _inventory_photo_safe_candidates(self, path_value: object) -> list[Path]:
        candidate_source = getattr(self, "_inventory_photo_path_candidates", None)
        candidates = candidate_source(path_value) if callable(candidate_source) else []
        if not candidates:
            path = Path(str(path_value or "")).expanduser()
            if not path.is_absolute():
                path = self._inventory_photo_source_folder() / path
            candidates = [path]
        roots: list[Path] = []
        root_sources = [self._inventory_photo_source_folder()]
        shared_folder = getattr(self, "_inventory_photo_shared_folder", None)
        if callable(shared_folder):
            root_sources.insert(0, shared_folder())
        for root in root_sources:
            try:
                roots.append(root.resolve())
            except Exception:
                continue
        safe: list[Path] = []
        seen: set[str] = set()
        for path in candidates:
            try:
                resolved = path.resolve()
            except Exception:
                continue
            if not any(resolved == root or root in resolved.parents for root in roots):
                continue
            key = str(resolved).lower()
            if key in seen:
                continue
            seen.add(key)
            safe.append(resolved)
        return safe

    def _deleted_archive_metadata_path(self, path: Path) -> Path:
        return path.with_name(f"{path.name}.archive.json")

    def _unique_deleted_archive_path(self, archive_root: Path, source: Path, archived_at: datetime) -> Path:
        folder = archive_root / archived_at.strftime("%Y-%m-%d")
        folder.mkdir(parents=True, exist_ok=True)
        destination = folder / source.name
        if not destination.exists() and not self._deleted_archive_metadata_path(destination).exists():
            return destination
        for index in range(2, 1000):
            candidate = folder / f"{source.stem}-{index}{source.suffix}"
            if not candidate.exists() and not self._deleted_archive_metadata_path(candidate).exists():
                return candidate
        return folder / f"{source.stem}-{archived_at.strftime('%H%M%S')}-{time.time_ns()}{source.suffix}"

    def _archive_deleted_file(self, source: Path, archive_root: Path, reason: str, details: dict[str, object] | None = None) -> Path:
        self._purge_expired_deleted_archive()
        archived_at = datetime.now()
        expires_at = archived_at + timedelta(days=DELETED_ARCHIVE_RETENTION_DAYS)
        destination = self._unique_deleted_archive_path(archive_root, source, archived_at)
        shutil.move(str(source), str(destination))
        atomic_write_json(
            self._deleted_archive_metadata_path(destination),
            {
                "original_path": str(source),
                "archive_path": str(destination),
                "reason": reason,
                "archived_at": archived_at.isoformat(timespec="seconds"),
                "expires_at": expires_at.isoformat(timespec="seconds"),
                "retention_days": DELETED_ARCHIVE_RETENTION_DAYS,
                "details": details or {},
            },
        )
        return destination

    def _purge_expired_deleted_archive(self) -> int:
        archive_root = DELETED_ARCHIVE_DIR
        if not archive_root.exists():
            return 0
        now = datetime.now()
        purged = 0
        for metadata_path in archive_root.rglob("*.archive.json"):
            try:
                metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            expires_raw = str(metadata.get("expires_at") or "").strip()
            try:
                expires_at = datetime.fromisoformat(expires_raw)
            except ValueError:
                continue
            if expires_at > now:
                continue
            archive_path = Path(str(metadata.get("archive_path") or "")).expanduser()
            try:
                resolved_archive = archive_path.resolve()
                resolved_root = archive_root.resolve()
            except Exception:
                continue
            if resolved_archive == resolved_root or resolved_root not in resolved_archive.parents:
                continue
            try:
                if resolved_archive.exists():
                    resolved_archive.unlink()
                    purged += 1
            except OSError:
                continue
            try:
                metadata_path.unlink()
            except OSError:
                pass
        for folder in sorted((path for path in archive_root.rglob("*") if path.is_dir()), key=lambda path: len(path.parts), reverse=True):
            try:
                folder.rmdir()
            except OSError:
                pass
        return purged

    def _delete_inventory_photo_files_for_removed_records(
        self,
        removed_records: list[dict[str, object]],
        remaining_records: list[dict[str, object]] | None = None,
    ) -> int:
        if not removed_records:
            return 0
        remaining = remaining_records if remaining_records is not None else [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        safe_candidates = getattr(self, "_inventory_photo_safe_candidates", None)
        def photo_safe_candidates(value: object) -> list[Path]:
            if callable(safe_candidates):
                return safe_candidates(value)
            path = self._safe_inventory_photo_path(value)
            return [path] if path else []

        still_used: set[str] = set()
        for record in remaining:
            for path in (record.get("photo_paths") or []):
                still_used.add(str(path))
                for candidate in photo_safe_candidates(path):
                    still_used.add(str(candidate))
        archived_paths: list[str] = []
        archived_by_original: dict[str, str] = {}
        for record in removed_records:
            for path_value in record.get("photo_paths") or []:
                if str(path_value) in still_used:
                    continue
                for path in photo_safe_candidates(path_value):
                    if str(path) in still_used or not path.exists():
                        continue
                    try:
                        archive_path = self._archive_deleted_file(
                            path,
                            DELETED_INVENTORY_PHOTOS_DIR,
                            "inventory_photo_removed",
                            {
                                "inventory_key": record.get("inventory_key") or "",
                                "cert_number": record.get("cert_number") or "",
                                "card_title": record.get("card_title") or "",
                                "source_sheet": record.get("source_sheet") or "",
                            },
                        )
                        archived_paths.append(str(archive_path))
                        archived_by_original[str(path)] = str(archive_path)
                    except Exception:
                        continue
        if archived_paths:
            state = self._load_inventory_photo_state()
            photos = state.setdefault("photos", {})
            archive_expires_at = (datetime.now() + timedelta(days=DELETED_ARCHIVE_RETENTION_DAYS)).isoformat(timespec="seconds")
            for record in photos.values():
                if not isinstance(record, dict):
                    continue
                record_path = record.get("path")
                try:
                    resolved = str(Path(str(record_path or "")).resolve())
                except Exception:
                    resolved = str(record_path or "")
                archive_path = archived_by_original.get(resolved)
                if archive_path:
                    record["status"] = "archived_from_album"
                    record["archived_at"] = datetime.now().isoformat(timespec="seconds")
                    record["archive_expires_at"] = archive_expires_at
                    record["archive_path"] = archive_path
            self._save_inventory_photo_state(state)
            self._append_activity("Inventory Photo Archive", f"Archived {len(archived_paths)} inventory photo file(s) for {DELETED_ARCHIVE_RETENTION_DAYS} days.", {"paths": archived_paths[:20]})
        return len(archived_paths)

    def _restore_inventory_photo_files_for_records(self, records: list[dict[str, object]]) -> int:
        if not records or not DELETED_INVENTORY_PHOTOS_DIR.exists():
            return 0
        metadata_for_records: list[dict[str, object]] = []
        metadata_by_original: dict[str, dict[str, object]] = {}
        for metadata_path in DELETED_INVENTORY_PHOTOS_DIR.rglob("*.archive.json"):
            try:
                metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            original = str(metadata.get("original_path") or "").strip()
            archive = str(metadata.get("archive_path") or "").strip()
            if original and archive:
                metadata_for_records.append(metadata)
                metadata_by_original[original] = metadata
                try:
                    metadata_by_original[str(Path(original).expanduser().resolve())] = metadata
                except Exception:
                    pass
        restored = 0
        restored_paths: list[str] = []
        for record in records:
            photo_values = [str(path) for path in (record.get("photo_paths") or []) if str(path or "").strip()]
            if not photo_values:
                cert = scan_to_cert(record.get("cert_number"))
                inventory_key = str(record.get("inventory_key") or "").strip().lower()
                source_sheet = Path(str(record.get("source_sheet") or "")).name.strip().lower()
                title = str(record.get("card_title") or "").strip().lower()
                recovered: list[str] = []
                for metadata in metadata_for_records:
                    details = metadata.get("details") if isinstance(metadata.get("details"), dict) else {}
                    detail_cert = scan_to_cert(details.get("cert_number"))
                    detail_key = str(details.get("inventory_key") or "").strip().lower()
                    detail_source = Path(str(details.get("source_sheet") or "")).name.strip().lower()
                    detail_title = str(details.get("card_title") or "").strip().lower()
                    matched = bool(cert and detail_cert and cert == detail_cert)
                    matched = matched or bool(inventory_key and detail_key and inventory_key == detail_key)
                    matched = matched or bool(source_sheet and detail_source and title and detail_title and source_sheet == detail_source and title == detail_title)
                    if matched:
                        original = str(metadata.get("original_path") or "").strip()
                        if original and original not in recovered:
                            recovered.append(original)
                if recovered:
                    record["photo_paths"] = recovered[:MAX_INVENTORY_PHOTOS_PER_CARD]
                    record["photo_count"] = len(record["photo_paths"])
                    photo_values = list(record["photo_paths"])
            for path_value in photo_values:
                candidates = self._inventory_photo_path_candidates(path_value)
                if any(candidate.exists() for candidate in candidates):
                    continue
                metadata = None
                for candidate in candidates:
                    metadata = metadata_by_original.get(str(candidate))
                    if metadata is not None:
                        break
                    try:
                        metadata = metadata_by_original.get(str(candidate.resolve()))
                    except Exception:
                        metadata = None
                    if metadata is not None:
                        break
                if metadata is None:
                    continue
                original_path = Path(str(metadata.get("original_path") or "")).expanduser()
                archive_path = Path(str(metadata.get("archive_path") or "")).expanduser()
                if not archive_path.exists() or original_path.exists():
                    continue
                try:
                    original_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(archive_path), str(original_path))
                    metadata_path = self._deleted_archive_metadata_path(archive_path)
                    if metadata_path.exists():
                        metadata_path.unlink()
                    restored += 1
                    restored_paths.append(str(original_path))
                except Exception:
                    continue
        if restored:
            self._append_activity("Inventory Photo Restore", f"Restored {restored} archived inventory photo file(s).", {"paths": restored_paths[:20]})
        return restored

    def _mark_inventory_record_sold(self, inventory_key: str, company: str, sale_price: float) -> int:
        if not inventory_key:
            return 0
        ledger = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        removed = [record for record in ledger if str(record.get("inventory_key") or "") == inventory_key]
        kept = [record for record in ledger if str(record.get("inventory_key") or "") != inventory_key]
        changed = len(ledger) - len(kept)
        if changed:
            self._save_inventory_ledger(kept)
            cleanup = getattr(self, "_delete_inventory_photo_files_for_removed_records", None)
            if callable(cleanup):
                cleanup(removed, kept)
        return changed

    def _general_sold_sheet_name(self, person: str) -> str:
        owner_for_profile = getattr(self, "_owner_for_profile", lambda value="": str(value or "").strip() or "Unassigned")
        person_name = owner_for_profile(person)
        return f"{person_name} General Sold"

    def _inventory_sale_profit_record(
        self,
        record: dict[str, object],
        company: str,
        sale_price: float,
        sale_date: str | None = None,
        sale_method: str = "",
    ) -> dict[str, object]:
        normalized = self._normalize_inventory_record(record)
        owner_for_profile = getattr(self, "_owner_for_profile", lambda person="": str(person or "").strip() or "Unassigned")
        assigned_person = owner_for_profile(normalized.get("assigned_person"))
        company_name = str(company or "").strip()
        source_sheet = normalized.get("source_sheet") or "Inventory"
        original_source_sheet = source_sheet
        if not company_name:
            company_name = "General Sold"
            source_sheet = self._general_sold_sheet_name(assigned_person)
        sold_date = str(sale_date or "").strip()[:10]
        if CardPipelineApp._profit_record_date(self, sold_date) is None:
            sold_date = datetime.now().strftime("%Y-%m-%d")
        notes = str(normalized.get("notes") or "").strip()
        method = str(sale_method or "").strip()
        if method:
            notes = clean_part("; ".join(part for part in (notes, f"Sale method: {method}") if part))
        return self._normalize_profit_record(
            {
                "date_added": sold_date,
                "company": company_name,
                "weekly_sheet_name": "Inventory Sale",
                "source_sheet": source_sheet,
                "original_source_sheet": original_source_sheet,
                "source": normalized.get("source") or "Inventory",
                "item_type": normalized.get("item_type") or "",
                "item_id": normalized.get("item_id") or "",
                "cert_number": normalized.get("cert_number") or "",
                "grader": normalized.get("grader") or "",
                "card_title": normalized.get("card_title") or "",
                "purchase_price": normalized.get("purchase_price"),
                "sale_price": sale_price,
                "sale_method": method,
                "assigned_person": assigned_person,
                "sport": normalized.get("sport") or "",
                "photo_paths": list(normalized.get("photo_paths") or []),
                "status": "Sold from inventory",
                "notes": notes,
            }
        )

    def _inventory_sale_expense_record(
        self,
        sale_record: dict[str, object],
        expense_type: str,
        expense_amount: float,
        notes: str = "",
    ) -> dict[str, object]:
        expense_type = str(expense_type or "").strip()
        if expense_type not in EXPENSE_CATEGORY_OPTIONS:
            expense_type = "Fees"
        return self._normalize_profit_record(
            {
                "record_type": "expense",
                "expense_id": datetime.now().strftime("%Y%m%d%H%M%S%f"),
                "date_added": str(sale_record.get("date_added") or datetime.now().strftime("%Y-%m-%d"))[:10],
                "assigned_person": sale_record.get("assigned_person") or "Unassigned",
                "expense_type": expense_type,
                "expense_amount": expense_amount,
                "related_type": "Card",
                "source_sheet": sale_record.get("source_sheet") or "",
                "item_id": sale_record.get("item_id") or "",
                "cert_number": sale_record.get("cert_number") or "",
                "notes": str(notes or "").strip(),
            }
        )

    def mark_inventory_record_sold(
        self,
        record: dict[str, object],
        company: str,
        sale_price: float,
        sale_date: str | None = None,
        sale_method: str = "",
        expense_type: str = "",
        expense_amount: float | None = None,
        expense_notes: str = "",
    ) -> bool:
        normalized = self._normalize_inventory_record(record)
        if str(normalized.get("status") or "").lower() != "active":
            return False
        company = str(company or "").strip()
        sold_company = company or "General Sold"
        profit_record = self._inventory_sale_profit_record(normalized, company, sale_price, sale_date=sale_date, sale_method=sale_method)
        profit_records = [profit_record]
        if expense_amount is not None and expense_amount > 0:
            profit_records.append(self._inventory_sale_expense_record(profit_record, expense_type, expense_amount, expense_notes))
        added = self.record_profit_sales(profit_records)
        changed = self._mark_inventory_record_sold(str(normalized.get("inventory_key") or ""), sold_company, sale_price)
        if added or changed:
            self._append_activity(
                "Inventory Sold",
                f"Marked inventory card sold to {sold_company} for {format_money(sale_price)}.",
                {"inventory_key": normalized.get("inventory_key"), "company": sold_company, "sale_price": sale_price, "card": normalized.get("card_title")},
            )
        return bool(added or changed)

    def _inventory_sale_dialog(self, record: dict[str, object]) -> dict[str, object] | None:
        normalized = self._normalize_inventory_record(record)
        default_sale = self._money_value(normalized.get("estimated_payout")) or self._money_value(normalized.get("inventory_value")) or 0.0
        company_var = tk.StringVar(value="")
        sale_var = tk.StringVar(value=f"{default_sale:.2f}" if default_sale else "")
        expense_type_var = tk.StringVar(value="Shipping")
        expense_amount_var = tk.StringVar()
        expense_notes_var = tk.StringVar()
        result: dict[str, object] = {}

        popup = tk.Toplevel(self)
        popup.title("Mark Inventory Sold")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Mark Sold", style="Panel.TLabel", font=("Segoe UI Semibold", 12)).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 2))
        ttk.Label(frame, text=str(normalized.get("card_title") or normalized.get("cert_number") or "Inventory card"), style="Muted.TLabel", wraplength=420).grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 14))
        ttk.Label(frame, text="Company / buyer", style="Panel.TLabel").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        ttk.Entry(frame, textvariable=company_var, width=34).grid(row=2, column=1, sticky="ew", pady=(0, 10))
        ttk.Label(frame, text="Sale price", style="Panel.TLabel").grid(row=3, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        sale_entry = ttk.Entry(frame, textvariable=sale_var, width=34)
        sale_entry.grid(row=3, column=1, sticky="ew", pady=(0, 10))
        ttk.Label(frame, text="Expense Type", style="Panel.TLabel").grid(row=4, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        ttk.Combobox(frame, textvariable=expense_type_var, values=EXPENSE_CATEGORY_OPTIONS, width=18, state="readonly").grid(row=4, column=1, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="Expense Amount", style="Panel.TLabel").grid(row=5, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        expense_entry = ttk.Entry(frame, textvariable=expense_amount_var, width=18)
        expense_entry.grid(row=5, column=1, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="Expense Notes", style="Panel.TLabel").grid(row=6, column=0, sticky="w", padx=(0, 10), pady=(0, 14))
        ttk.Entry(frame, textvariable=expense_notes_var, width=34).grid(row=6, column=1, sticky="ew", pady=(0, 14))
        status_var = tk.StringVar(value="Leave company / buyer blank for that person's General Sold sheet.")
        ttk.Label(frame, textvariable=status_var, style="Muted.TLabel").grid(row=7, column=0, columnspan=2, sticky="w", pady=(0, 14))

        def submit() -> None:
            sale_price = self._money_value(sale_var.get())
            if sale_price is None or sale_price < 0:
                status_var.set("Enter a valid sale price.")
                sale_entry.focus_set()
                return
            expense_amount = self._money_value(expense_amount_var.get())
            if expense_amount_var.get().strip() and (expense_amount is None or expense_amount <= 0):
                status_var.set("Enter a valid expense amount or leave it blank.")
                expense_entry.focus_set()
                return
            result["company"] = company_var.get().strip()
            result["sale_price"] = float(sale_price)
            result["expense_type"] = expense_type_var.get().strip()
            result["expense_amount"] = float(expense_amount) if expense_amount is not None else None
            result["expense_notes"] = expense_notes_var.get().strip()
            popup.destroy()

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=8, column=0, columnspan=2, sticky="e")
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="Mark Sold", command=submit, style="Primary.TButton").pack(side=tk.LEFT)
        frame.columnconfigure(1, weight=1)
        popup.bind("<Return>", lambda _event: submit())
        popup.bind("<Escape>", lambda _event: popup.destroy())
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")
        sale_entry.focus_set()
        self.wait_window(popup)
        if "sale_price" not in result:
            return None
        return result

    def mark_selected_inventory_sold(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        selected = list(self.inventory_tree.selection())
        records = [self.inventory_tree_records.get(iid) for iid in selected if self.inventory_tree_records.get(iid)]
        if len(records) != 1:
            messagebox.showinfo("Choose one card", "Select one active inventory card to mark sold.")
            return
        record = self._normalize_inventory_record(records[0])
        if str(record.get("status") or "").lower() != "active":
            messagebox.showinfo("Active card required", "Only active inventory cards can be marked sold.")
            return
        sale = self._inventory_sale_dialog(record)
        if sale is None:
            return
        sale_price = float(sale["sale_price"])
        if self.mark_inventory_record_sold(
            record,
            str(sale.get("company") or ""),
            sale_price,
            expense_type=str(sale.get("expense_type") or ""),
            expense_amount=sale.get("expense_amount") if sale.get("expense_amount") is not None else None,
            expense_notes=str(sale.get("expense_notes") or ""),
        ):
            self.refresh_inventory_tab()
            self.refresh_profit_tab()
            self.status_var.set(f"Marked inventory card sold: {record.get('cert_number') or record.get('card_title') or 'card'} for {format_money(sale_price)}.")

    def _inventory_edit_row_dialog(self, record: dict[str, object]) -> dict[str, object] | None:
        normalized = self._normalize_inventory_record(record)
        popup = tk.Toplevel(self)
        popup.title("Edit Inventory Row")
        popup.geometry("980x640")
        popup.minsize(760, 420)
        popup.transient(self)
        popup.grab_set()
        popup.configure(bg="#121212")
        result: dict[str, object] = {}
        fields = [
            ("date_added", "Date"),
            ("item_type", "Type"),
            ("item_id", "Item ID"),
            ("assigned_person", "Person"),
            ("sport", "Sport"),
            ("cert_number", "Cert"),
            ("grader", "Grader"),
            ("card_title", "Card"),
            ("purchase_price", "Purchase"),
            ("paid_with", "Paid With"),
            ("card_ladder_value", "Card Ladder"),
            ("card_ladder_comps_average", "Comps"),
            ("cy_value", "CY Estimate"),
            ("cy_confidence", "CY Confidence"),
            ("best_company", "Best Company"),
            ("estimated_payout", "Est. Payout"),
            ("source_sheet", "Source Sheet"),
        ]
        money_fields = {"purchase_price", "card_ladder_value", "card_ladder_comps_average", "cy_value", "estimated_payout"}
        vars_by_field: dict[str, tk.StringVar] = {}

        frame = self._scrollable_popup_frame(popup, style_name="Panel.TFrame", bg="#121212", padding=(18, 16))
        ttk.Label(frame, text="Edit Inventory Row", style="Panel.TLabel", font=("Segoe UI Semibold", 12)).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 2))
        ttk.Label(frame, text=str(normalized.get("status") or "Active"), style="Muted.TLabel").grid(row=1, column=0, columnspan=4, sticky="w", pady=(0, 14))
        for index, (field, label) in enumerate(fields):
            row = 2 + index // 2
            col = 0 if index % 2 == 0 else 2
            value = normalized.get(field)
            text = "" if value is None else f"{value:.2f}" if field in money_fields and isinstance(value, (int, float)) else str(value)
            var = tk.StringVar(value=text)
            vars_by_field[field] = var
            ttk.Label(frame, text=label, style="Panel.TLabel").grid(row=row, column=col, sticky="w", padx=(0, 8), pady=(0, 8))
            width = 46 if field == "card_title" else 24
            if field == "assigned_person":
                person_combo = ttk.Combobox(frame, textvariable=var, width=width)
                person_combo.grid(row=row, column=col + 1, sticky="ew", padx=(0, 14), pady=(0, 8))
                self._bind_person_autocomplete(person_combo)
            else:
                ttk.Entry(frame, textvariable=var, width=width).grid(row=row, column=col + 1, sticky="ew", padx=(0, 14), pady=(0, 8))
        status_var = tk.StringVar(value="Status changes use Mark Sold, Move, or Delete.")
        status_row = 2 + (len(fields) + 1) // 2
        ttk.Label(frame, textvariable=status_var, style="Muted.TLabel").grid(row=status_row, column=0, columnspan=4, sticky="w", pady=(4, 14))

        def submit() -> None:
            updates: dict[str, object] = {}
            for field, _label in fields:
                raw = vars_by_field[field].get().strip()
                if field in {"item_type", "item_id"}:
                    updates[field] = normalized.get(field) or raw
                    continue
                if field == "assigned_person":
                    person_choice = self._canonical_person_choice(raw)
                    if person_choice is None:
                        status_var.set("Choose an existing person.")
                        return
                    updates[field] = person_choice
                    continue
                if field in money_fields:
                    if not raw:
                        updates[field] = None
                        continue
                    value = self._money_value(raw)
                    if value is None or value < 0:
                        status_var.set(f"Enter a valid value for {field.replace('_', ' ')}.")
                        return
                    updates[field] = float(value)
                else:
                    updates[field] = raw
            result.update(updates)
            popup.destroy()

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=status_row + 1, column=0, columnspan=4, sticky="e")
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="Save", command=submit, style="Primary.TButton").pack(side=tk.LEFT)
        frame.columnconfigure(1, weight=1)
        frame.columnconfigure(3, weight=1)
        popup.bind("<Return>", lambda _event: submit())
        popup.bind("<Escape>", lambda _event: popup.destroy())
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")
        self.wait_window(popup)
        return result or None

    def _scrollable_popup_frame(self, popup: tk.Toplevel, style_name: str = "Panel.TFrame", bg: str = "#121212", padding: tuple[int, int] = (18, 16)) -> ttk.Frame:
        outer = ttk.Frame(popup, style=style_name)
        outer.pack(fill=tk.BOTH, expand=True)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)
        canvas = tk.Canvas(outer, bg=bg, highlightthickness=0, borderwidth=0)
        canvas.grid(row=0, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(outer, orient=tk.HORIZONTAL, command=canvas.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        canvas.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        frame = ttk.Frame(canvas, style=style_name, padding=padding)
        window_id = canvas.create_window((0, 0), window=frame, anchor="nw")

        def update_scrollregion(_event: tk.Event | None = None) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        def resize_window(event: tk.Event) -> None:
            canvas.itemconfigure(window_id, width=max(event.width, frame.winfo_reqwidth()))
            update_scrollregion()

        def on_mousewheel(event: tk.Event) -> str:
            delta = int(getattr(event, "delta", 0) or 0)
            if delta:
                canvas.yview_scroll(int(-1 * (delta / 120)), "units")
            return "break"

        def on_shift_mousewheel(event: tk.Event) -> str:
            delta = int(getattr(event, "delta", 0) or 0)
            if delta:
                canvas.xview_scroll(int(-1 * (delta / 120)), "units")
            return "break"

        frame.bind("<Configure>", update_scrollregion)
        canvas.bind("<Configure>", resize_window)
        for widget in (outer, canvas, frame):
            widget.bind("<MouseWheel>", on_mousewheel, add="+")
            widget.bind("<Shift-MouseWheel>", on_shift_mousewheel, add="+")
        return frame

    def edit_selected_inventory_row(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        selected = list(self.inventory_tree.selection())
        records = [self.inventory_tree_records.get(iid) for iid in selected if self.inventory_tree_records.get(iid)]
        if len(records) != 1:
            messagebox.showinfo("Choose one card", "Select one active inventory row to edit.")
            return
        record = self._normalize_inventory_record(records[0])
        if str(record.get("status") or "").lower() != "active":
            messagebox.showinfo("Active card required", "Only active inventory rows can be edited here.")
            return
        updates = self._inventory_edit_row_dialog(record)
        if not updates:
            return
        key = str(record.get("inventory_key") or "")
        updated = self._update_inventory_record_by_key(key, updates)
        self.refresh_inventory_tab()
        self.status_var.set(f"Edited {updated} inventory row(s).")

    def _update_inventory_record_by_key(self, key: str, updates: dict[str, object]) -> int:
        if not key:
            return 0
        with shared_lock(CARD_PIPELINE_DIR, "inventory-row-edit", self.lucas_identity):
            rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
            updated = 0
            for index, record in enumerate(rows):
                if str(record.get("inventory_key") or "") != key or str(record.get("status") or "").lower() != "active":
                    continue
                merged = dict(record)
                merged.update(updates)
                merged.pop("inventory_key", None)
                rows[index] = self._normalize_inventory_record(merged)
                updated += 1
                break
            if updated:
                self._save_inventory_ledger(rows)
        return updated

    def _replace_inventory_record_by_key(self, key: str, replacement: dict[str, object]) -> dict[str, object] | None:
        if not key:
            return None
        with shared_lock(CARD_PIPELINE_DIR, "inventory-row-edit", self.lucas_identity):
            rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
            for index, record in enumerate(rows):
                if str(record.get("inventory_key") or "") != key:
                    continue
                normalized = self._normalize_inventory_record(replacement)
                rows[index] = normalized
                self._save_inventory_ledger(rows)
                return normalized
        return None

    def _toggle_inventory_bulk_edit(self) -> None:
        self._destroy_inventory_cell_editor()
        self._style_inventory_bulk_toggle()
        if not self.inventory_bulk_edit_var.get():
            self.inventory_bulk_cell = None
            self.inventory_status_var.set("Bulk edit off.")
            return
        rows = self.inventory_tree.get_children() if hasattr(self, "inventory_tree") else ()
        if rows:
            selected = self.inventory_tree.selection()
            iid = selected[0] if selected else rows[0]
            self._set_inventory_bulk_cell(iid, self._inventory_editable_columns()[0])
        self.inventory_tree.focus_set()
        self.inventory_status_var.set("Bulk edit on. Use arrows to move, Enter/F2 to edit, Esc to cancel a cell.")

    def _style_inventory_bulk_toggle(self) -> None:
        toggle = getattr(self, "inventory_bulk_toggle", None)
        if toggle is None:
            return
        active = bool(self.inventory_bulk_edit_var.get())
        if active:
            toggle.configure(
                text="Bulk Edit ON",
                bg="#1ed760",
                fg="#000000",
                activebackground="#1fdf64",
                activeforeground="#000000",
                selectcolor="#1ed760",
            )
        else:
            toggle.configure(
                text="Bulk Edit",
                bg="#2a2a2a",
                fg="#ffffff",
                activebackground="#333333",
                activeforeground="#ffffff",
                selectcolor="#2a2a2a",
            )

    def _inventory_editable_columns(self) -> list[str]:
        return [column for column in self._personal_person_last_columns(INVENTORY_TABLE_COLUMNS) if column in INVENTORY_EDIT_COLUMN_FIELDS]

    def _inventory_bulk_click(self, event: tk.Event) -> None:
        if not getattr(self, "inventory_bulk_edit_var", None) or not self.inventory_bulk_edit_var.get():
            return
        column_id = self.inventory_tree.identify_column(event.x)
        row_id = self.inventory_tree.identify_row(event.y)
        if not row_id or not column_id:
            return
        try:
            column_index = int(str(column_id).replace("#", "")) - 1
        except ValueError:
            return
        columns = list(self.inventory_tree["columns"])
        if column_index < 0 or column_index >= len(columns):
            return
        column = columns[column_index]
        if column in INVENTORY_EDIT_COLUMN_FIELDS:
            self._set_inventory_bulk_cell(row_id, column)

    def _set_inventory_bulk_cell(self, iid: str, column: str) -> None:
        if not iid or column not in INVENTORY_EDIT_COLUMN_FIELDS:
            return
        self.inventory_bulk_cell = (iid, column)
        self.inventory_tree.selection_set(iid)
        self.inventory_tree.focus(iid)
        self.inventory_tree.see(iid)
        heading = INVENTORY_HEADINGS.get(column, column)
        self.inventory_status_var.set(f"Bulk edit cell: {heading}. Press Enter or F2 to edit.")

    def _move_inventory_bulk_cell(self, row_delta: int, column_delta: int, reopen: bool = False) -> str | None:
        if not getattr(self, "inventory_bulk_edit_var", None) or not self.inventory_bulk_edit_var.get():
            return None
        if self.inventory_cell_editor is not None:
            return self._commit_inventory_bulk_edit(row_delta, column_delta, reopen=True)
        rows = list(self.inventory_tree.get_children())
        columns = self._inventory_editable_columns()
        if not rows or not columns:
            return "break"
        iid, column = self.inventory_bulk_cell or (self.inventory_tree.focus() or rows[0], columns[0])
        if iid not in rows:
            iid = rows[0]
        if column not in columns:
            column = columns[0]
        row_index = max(0, min(len(rows) - 1, rows.index(iid) + row_delta))
        col_index = max(0, min(len(columns) - 1, columns.index(column) + column_delta))
        self._set_inventory_bulk_cell(rows[row_index], columns[col_index])
        if reopen:
            self._begin_inventory_bulk_edit()
        return "break"

    def _begin_inventory_bulk_edit(self, event: tk.Event | None = None) -> str | None:
        if not getattr(self, "inventory_bulk_edit_var", None) or not self.inventory_bulk_edit_var.get():
            return None
        if event is not None and getattr(event, "x", None) is not None:
            self._inventory_bulk_click(event)
        rows = list(self.inventory_tree.get_children())
        columns = self._inventory_editable_columns()
        if not rows or not columns:
            return "break"
        iid, column = self.inventory_bulk_cell or (self.inventory_tree.focus() or rows[0], columns[0])
        if iid not in rows:
            iid = rows[0]
        if column not in columns:
            column = columns[0]
        bbox = self.inventory_tree.bbox(iid, column)
        if not bbox:
            self.inventory_tree.see(iid)
            bbox = self.inventory_tree.bbox(iid, column)
        if not bbox:
            return "break"
        self._destroy_inventory_cell_editor()
        x, y, width, height = bbox
        editor = ttk.Entry(self.inventory_tree)
        editor.insert(0, self.inventory_tree.set(iid, column))
        editor.select_range(0, tk.END)
        editor.place(x=x, y=y, width=width, height=height)
        editor.focus_set()
        self.inventory_cell_editor = editor
        self.inventory_cell_edit = (iid, column)
        editor.bind("<Return>", lambda _event: self._commit_inventory_bulk_edit(1, 0, reopen=True))
        editor.bind("<KP_Enter>", lambda _event: self._commit_inventory_bulk_edit(1, 0, reopen=True))
        editor.bind("<Tab>", lambda _event: self._commit_inventory_bulk_edit(0, 1, reopen=True))
        editor.bind("<Shift-Tab>", lambda _event: self._commit_inventory_bulk_edit(0, -1, reopen=True))
        editor.bind("<Up>", lambda _event: self._commit_inventory_bulk_edit(-1, 0, reopen=True))
        editor.bind("<Down>", lambda _event: self._commit_inventory_bulk_edit(1, 0, reopen=True))
        editor.bind("<Left>", lambda _event: self._commit_inventory_bulk_edit(0, -1, reopen=True))
        editor.bind("<Right>", lambda _event: self._commit_inventory_bulk_edit(0, 1, reopen=True))
        editor.bind("<Control-z>", self._undo_inventory_bulk_edit)
        editor.bind("<Control-Z>", self._undo_inventory_bulk_edit)
        editor.bind("<Command-z>", self._undo_inventory_bulk_edit)
        editor.bind("<Command-Z>", self._undo_inventory_bulk_edit)
        editor.bind("<Escape>", lambda _event: self._cancel_inventory_bulk_edit())
        editor.bind("<FocusOut>", lambda _event: self._commit_inventory_bulk_edit(0, 0, reopen=False))
        return "break"

    def _commit_inventory_bulk_edit(self, row_delta: int = 0, column_delta: int = 0, reopen: bool = False) -> str:
        if self.inventory_cell_editor is None or self.inventory_cell_edit is None:
            self._move_inventory_bulk_cell(row_delta, column_delta, reopen=reopen)
            return "break"
        iid, column = self.inventory_cell_edit
        raw = self.inventory_cell_editor.get().strip()
        self._destroy_inventory_cell_editor()
        record = self.inventory_tree_records.get(iid)
        if not record:
            return "break"
        updates = self._inventory_bulk_updates_for_cell(column, raw)
        if updates is None:
            self._set_inventory_bulk_cell(iid, column)
            self._begin_inventory_bulk_edit()
            return "break"
        key = str(record.get("inventory_key") or "")
        before = dict(record)
        updated = self._update_inventory_record_by_key(key, updates)
        if updated:
            merged = dict(record)
            merged.update(updates)
            merged.pop("inventory_key", None)
            normalized = self._normalize_inventory_record(merged)
            self.inventory_bulk_undo_stack.append(
                {
                    "iid": iid,
                    "column": column,
                    "before": before,
                    "after_key": str(normalized.get("inventory_key") or key),
                }
            )
            self.inventory_tree_records[iid] = normalized
            self._refresh_inventory_tree_row(iid, normalized)
            self.status_var.set("Inventory cell saved.")
        self.inventory_bulk_cell = (iid, column)
        self._move_inventory_bulk_cell(row_delta, column_delta, reopen=reopen)
        return "break"

    def _inventory_bulk_updates_for_cell(self, column: str, raw: str) -> dict[str, object] | None:
        field = INVENTORY_EDIT_COLUMN_FIELDS.get(column)
        if not field:
            return {}
        if column in INVENTORY_EDIT_MONEY_COLUMNS:
            if not raw:
                return {field: None}
            value = self._money_value(raw)
            if value is None or value < 0:
                self.inventory_status_var.set(f"Enter a valid number for {INVENTORY_HEADINGS.get(column, column)}.")
                return None
            return {field: float(value)}
        return {field: raw}

    def _refresh_inventory_tree_row(self, iid: str, record: dict[str, object]) -> None:
        values = {
            "date": record.get("date_added") or "",
            "type": record.get("item_type") or "",
            "item_id": record.get("item_id") or "",
            "person": record.get("assigned_person") or "Unassigned",
            "sport": record.get("sport") or "",
            "cert": record.get("cert_number") or "",
            "grader": record.get("grader") or "",
            "card": record.get("card_title") or "",
            "purchase": format_money(record.get("purchase_price")),
            "card_ladder": format_money(record.get("card_ladder_value")),
            "comps": format_money(record.get("card_ladder_comps_average")),
            "cy_estimate": format_money(record.get("cy_value")),
            "cy_confidence": record.get("cy_confidence") if record.get("cy_confidence") is not None else "",
            "company": record.get("best_company") or "",
            "payout": format_money(record.get("estimated_payout")),
            "source": record.get("source_sheet") or "",
            "status": record.get("status") or "",
            "photos": str(len(record.get("photo_paths") or [])),
            "notes": inventory_display_notes(record),
        }
        for column, value in values.items():
            if column in self.inventory_tree["columns"]:
                self.inventory_tree.set(iid, column, value)

    def _cancel_inventory_bulk_edit(self) -> str:
        self._destroy_inventory_cell_editor()
        return "break"

    def _undo_inventory_bulk_edit(self, event: tk.Event | None = None) -> str:
        if not getattr(self, "inventory_bulk_edit_var", None) or not self.inventory_bulk_edit_var.get():
            return "break"
        self._destroy_inventory_cell_editor()
        if not self.inventory_bulk_undo_stack:
            self.inventory_status_var.set("Nothing to undo.")
            return "break"
        action = self.inventory_bulk_undo_stack.pop()
        before = dict(action.get("before") or {})
        after_key = str(action.get("after_key") or before.get("inventory_key") or "")
        restored = self._replace_inventory_record_by_key(after_key, before)
        if not restored:
            self.inventory_status_var.set("Could not undo; inventory row was not found.")
            return "break"
        iid = str(action.get("iid") or "")
        if iid and hasattr(self, "inventory_tree") and iid in self.inventory_tree.get_children():
            self.inventory_tree_records[iid] = restored
            self._refresh_inventory_tree_row(iid, restored)
            column = str(action.get("column") or "")
            if column in INVENTORY_EDIT_COLUMN_FIELDS:
                self._set_inventory_bulk_cell(iid, column)
        else:
            self.refresh_inventory_tab()
        self.inventory_status_var.set("Undid last bulk edit.")
        self.status_var.set("Inventory bulk edit undone.")
        return "break"

    def _destroy_inventory_cell_editor(self) -> None:
        editor = self.inventory_cell_editor
        self.inventory_cell_editor = None
        self.inventory_cell_edit = None
        if editor is not None:
            try:
                editor.destroy()
            except tk.TclError:
                pass

    def move_selected_inventory_to_company_sheets(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        selected = list(self.inventory_tree.selection())
        records = [self.inventory_tree_records.get(iid) for iid in selected if self.inventory_tree_records.get(iid)]
        movable_records = [record for record in records if self._inventory_record_can_move_to_company_sheet(record)]
        if not movable_records:
            messagebox.showinfo("Choose inventory", "Select one or more active inventory rows with an assignable best company.")
            return
        confirmed = messagebox.askyesno(
            "Move inventory card(s)?",
            f"Move {len(movable_records)} active inventory card(s) to company sheets?",
        )
        if not confirmed:
            return
        self._move_inventory_records_to_company_sheets(movable_records)

    def move_selected_inventory_to_specific_company_sheet(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        selected = list(self.inventory_tree.selection())
        records = [self.inventory_tree_records.get(iid) for iid in selected if self.inventory_tree_records.get(iid)]
        movable_records = [record for record in records if self._inventory_record_can_move_to_company_sheet(record)]
        if not movable_records:
            messagebox.showinfo("Choose inventory", "Select one or more active inventory rows with an assignable best company.")
            return
        default_company = str(movable_records[0].get("best_company") or "")
        company = self._choose_inventory_company(default_company)
        if not company:
            return
        confirmed = messagebox.askyesno(
            "Move inventory card(s)?",
            f"Move {len(movable_records)} active inventory card(s) to {company}?",
        )
        if not confirmed:
            return
        self._move_inventory_records_to_company_sheets(movable_records, company_override=company)

    def _choose_inventory_company(self, default_company: str = "") -> str:
        companies = [
            company.name
            for company in self.assignment_engine.companies
            if str(company.name or "").strip() and str(company.name or "").strip().upper() != NO_COMPANY_TAKES_LABEL
        ]
        if not companies:
            messagebox.showinfo("No companies", "No active assignment companies are loaded.")
            return ""
        popup = tk.Toplevel(self)
        popup.title("Move to Company Sheet")
        popup.transient(self)
        popup.grab_set()
        popup.configure(bg="#121212")
        result = tk.StringVar(value="")
        company_var = tk.StringVar(value=default_company if default_company in companies else companies[0])
        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Move to Company Sheet", style="Panel.TLabel", font=("Segoe UI Semibold", 12)).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 12))
        ttk.Label(frame, text="Company", style="Panel.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(0, 12))
        company_combo = ttk.Combobox(frame, textvariable=company_var, values=companies, width=34)
        company_combo.grid(row=1, column=1, sticky="ew", pady=(0, 12))

        def submit() -> None:
            company = company_var.get().strip()
            if not company:
                return
            result.set(company)
            popup.destroy()

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=2, column=0, columnspan=2, sticky="e")
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="Move", command=submit, style="Primary.TButton").pack(side=tk.LEFT)
        frame.columnconfigure(1, weight=1)
        popup.bind("<Return>", lambda _event: submit())
        popup.bind("<Escape>", lambda _event: popup.destroy())
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")
        company_combo.focus_set()
        self.wait_window(popup)
        return result.get().strip()

    def _specific_company_decision(self, row: WorkbookRow, person: str, company_name: str) -> assignment_engine.AssignmentDecision | None:
        target = clean_part(company_name).casefold()
        if not target:
            return None
        for decision in self.assignment_engine.evaluate(row, person=person):
            if clean_part(decision.company).casefold() == target:
                return decision
        return None

    def _move_inventory_records_to_company_sheets(self, movable_records: list[dict[str, object]], company_override: str = "") -> None:
        rows: list[WorkbookRow] = []
        source_lookup: dict[int, str] = {}
        sheet_source_lookup: dict[int, str] = {}
        people_by_cert: dict[str, str] = {}
        keys_by_cert: dict[str, str] = {}
        photos_by_cert: dict[str, list[str]] = {}
        unassigned = 0
        for index, record in enumerate(movable_records, start=1):
            row = self._inventory_workbook_row(record, index)
            person = str(record.get("assigned_person") or "")
            if company_override:
                decision = self._specific_company_decision(row, person, company_override)
                if not decision or not decision.accepted or decision.payout is None:
                    unassigned += 1
                    continue
                row.best_company = decision.company
                row.estimated_payout = round(decision.payout, 2)
            else:
                stored_company = str(record.get("best_company") or "").strip()
                stored_payout = self._money_value(record.get("estimated_payout"))
                if stored_company and stored_company.upper() != NO_COMPANY_TAKES_LABEL and stored_payout is not None:
                    row.best_company = stored_company
                    row.estimated_payout = stored_payout
                else:
                    recommendation = self.assignment_engine.recommend(row, person=person)
                    if recommendation.payout is None:
                        unassigned += 1
                        continue
                    row.best_company = recommendation.company
                    row.estimated_payout = recommendation.payout
            if not row.best_company or row.best_company.upper() == NO_COMPANY_TAKES_LABEL:
                unassigned += 1
                continue
            rows.append(row)
            source_lookup[index] = str(record.get("source") or "Inventory")
            sheet_source_lookup[index] = str(record.get("source_sheet") or "Inventory")
            cert = scan_to_cert(row.cert_number)
            if cert:
                people_by_cert[cert] = str(record.get("assigned_person") or "")
                keys_by_cert[cert] = str(record.get("inventory_key") or "")
                photos_by_cert[cert] = [str(path) for path in (record.get("photo_paths") or []) if str(path or "").strip()]
        if not rows:
            messagebox.showinfo("No company match", "No selected inventory cards matched an assignable company.")
            return
        with shared_lock(CARD_PIPELINE_DIR, "inventory-company-sheets", self.lucas_identity):
            company_result = append_company_sheet_rows(
                COMPANY_SHEETS_DIR,
                rows,
                source_lookup,
                sheet_source_lookup,
                sheet_name_lookup=self._company_sheet_name_lookup_for_rows(rows),
            )
            added_records = list(company_result.get("added_records") or [])
            moved_keys: set[str] = set()
            for record in added_records:
                cert = scan_to_cert(record.get("cert_number"))
                record["assigned_person"] = people_by_cert.get(cert, "")
                if cert and not record.get("photo_paths"):
                    record["photo_paths"] = list(photos_by_cert.get(cert, []))
                    record["photo_count"] = len(record["photo_paths"])
                key = keys_by_cert.get(cert, "")
                if key:
                    moved_keys.add(key)
            if added_records:
                self.record_profit_sales(added_records)
            self._mark_inventory_records_moved_to_company(moved_keys)
        self.refresh_inventory_tab()
        self.refresh_profit_tab()
        added = int(company_result.get("rows_added") or 0)
        errors = company_result.get("errors") or []
        if company_override:
            suffix = f" {unassigned} card(s) were not accepted by {company_override}." if unassigned else ""
        else:
            suffix = f" {unassigned} card(s) had no assignable company." if unassigned else ""
        self.status_var.set(f"Moved {added} inventory card(s) to company sheets.{suffix}")
        self._append_activity("Inventory Move", f"Moved {added} inventory card(s) to company sheets.{suffix}", {"rows_added": added, "unassigned": unassigned, "company_override": company_override})
        if errors:
            messagebox.showwarning("Inventory move completed with warnings", "\n".join([f"Moved rows: {added}", *errors[:8]]))

    def _inventory_record_can_move_to_company_sheet(self, record: dict[str, object] | None) -> bool:
        if not record or str(record.get("status") or "").lower() != "active":
            return False
        if str(record.get("item_type") or "").strip().lower() == "raw":
            return False
        best_company = str(record.get("best_company") or "").strip()
        return bool(best_company) and best_company.upper() != NO_COMPANY_TAKES_LABEL

    def _tree_cell_text(self, tree: ttk.Treeview, row_id: str, column_id: str) -> str:
        if not row_id or not column_id:
            return ""
        values = tree.item(row_id, "values") or ()
        try:
            index = int(str(column_id).lstrip("#")) - 1
        except ValueError:
            return ""
        if index < 0 or index >= len(values):
            return ""
        return str(values[index] or "")

    def _inventory_tree_cell_text(self, row_id: str, column_id: str) -> str:
        return self._tree_cell_text(self.inventory_tree, row_id, column_id)

    def _tree_row_text(self, tree: ttk.Treeview, row_id: str) -> str:
        if not row_id:
            return ""
        values = tree.item(row_id, "values") or ()
        return "\t".join(str(value or "") for value in values)

    def _inventory_tree_row_text(self, row_id: str) -> str:
        return self._tree_row_text(self.inventory_tree, row_id)

    def _copy_inventory_text(self, text: str, label: str = "inventory value") -> None:
        clipboard_text = str(text or "")
        for _attempt in range(2):
            self.clipboard_clear()
            try:
                self.update()
            except tk.TclError:
                pass
            self.clipboard_append(clipboard_text)
            try:
                self.update()
            except tk.TclError:
                pass
            try:
                if self.clipboard_get() == clipboard_text:
                    break
            except tk.TclError:
                break
        if hasattr(self, "status_var"):
            self.status_var.set(f"Copied {label}.")

    def copy_inventory_cell_value(self, row_id: str, column_id: str) -> None:
        self._copy_inventory_text(self._inventory_tree_cell_text(row_id, column_id), "inventory cell")

    def copy_inventory_row_values(self, row_id: str) -> None:
        self._copy_inventory_text(self._inventory_tree_row_text(row_id), "inventory row")

    def copy_tree_cell_value(self, tree: ttk.Treeview, row_id: str, column_id: str, label: str = "cell") -> None:
        self._copy_inventory_text(self._tree_cell_text(tree, row_id, column_id), label)

    def copy_tree_row_values(self, tree: ttk.Treeview, row_id: str, label: str = "row") -> None:
        self._copy_inventory_text(self._tree_row_text(tree, row_id), label)

    def _assignment_explanation_for_record(self, record: dict[str, object]) -> str:
        normalized = self._normalize_inventory_record(record)
        row = self._inventory_workbook_row(normalized, 1)
        person = str(normalized.get("assigned_person") or "").strip()
        recommendation = self.assignment_engine.recommend(row, person=person)
        decisions = self.assignment_engine.evaluate(row, person=person)
        lines = [
            str(normalized.get("card_title") or normalized.get("cert_number") or "Inventory card"),
            "",
            f"Person: {person or 'Unassigned'}",
            f"Sport: {normalized.get('sport') or 'blank'}",
            f"Cert: {normalized.get('cert_number') or 'blank'}",
            f"Grader: {normalized.get('grader') or 'blank'}",
            f"CL value: {format_money(self._money_value(normalized.get('card_ladder_value')) or 0.0)}",
            f"Comps: {format_money(self._money_value(normalized.get('card_ladder_comps_average')) or 0.0)}",
            f"CY estimate: {format_money(self._money_value(normalized.get('cy_value')) or 0.0)}",
            f"CY confidence: {normalized.get('cy_confidence') or 'blank'}",
            "",
            f"Recommended: {recommendation.company or NO_COMPANY_TAKES_LABEL} | {format_money(recommendation.payout) if recommendation.payout is not None else 'no payout'}",
            "",
            "Rule decisions:",
        ]
        if not decisions:
            lines.append("No assignment companies are loaded.")
        for decision in decisions:
            status = "TAKES" if decision.accepted and decision.payout is not None else "NO"
            payout = format_money(decision.payout) if decision.payout is not None else "no payout"
            source_value = format_money(decision.source_value) if decision.source_value is not None else "no source value"
            reason = self._assignment_decision_detail(decision)
            lines.append(f"- {decision.company}: {status} | {source_value} | {payout} | {reason}")
        return "\n".join(lines)

    def _assignment_explanation_for_workbook_row(self, row: WorkbookRow, label: str = "Workflow row") -> str:
        person = self._assignment_person_for_row(row)
        recommendation = self.assignment_engine.recommend(row, person=person)
        decisions = self.assignment_engine.evaluate(row, person=person)
        lines = [
            row.card_title or row.cert_number or row.item_id or label,
            "",
            f"Person: {person or 'Unassigned'}",
            f"Sport: {row.category or 'blank'}",
            f"Cert/Item ID: {row.cert_number or row.item_id or 'blank'}",
            f"Grader: {row.grader or 'blank'}",
            f"Purchase: {format_money(self._money_value(row.existing_value) or 0.0)}",
            f"CL value: {format_money(self._money_value(row.card_ladder_value) or 0.0)}",
            f"Comps: {format_money(self._money_value(row.card_ladder_comps_average) or 0.0)}",
            f"CY estimate: {format_money(self._money_value(row.cy_value) or 0.0)}",
            f"CY confidence: {row.cy_confidence or 'blank'}",
            f"Current Best Company: {row.best_company or 'blank'}",
            f"Current Est. Payout: {format_money(row.estimated_payout) if row.estimated_payout is not None else 'blank'}",
            "",
            f"Recommended: {recommendation.company or NO_COMPANY_TAKES_LABEL} | {format_money(recommendation.payout) if recommendation.payout is not None else 'no payout'}",
            "",
            "Rule decisions:",
        ]
        if not decisions:
            lines.append("No assignment companies are loaded.")
        for decision in decisions:
            status = "TAKES" if decision.accepted and decision.payout is not None else "NO"
            payout = format_money(decision.payout) if decision.payout is not None else "no payout"
            source_value = format_money(decision.source_value) if decision.source_value is not None else "no source value"
            reason = self._assignment_decision_detail(decision)
            lines.append(f"- {decision.company}: {status} | {source_value} | {payout} | {reason}")
        return "\n".join(lines)

    def _show_assignment_explanation_popup(self, explanation: str) -> None:
        popup = tk.Toplevel(self)
        popup.title("Assignment Explanation")
        popup.configure(bg=self.colors["bg"])
        popup.transient(self)
        popup.geometry("760x520")
        frame = ttk.Frame(popup, style="App.TFrame", padding=18)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Assignment Explanation", style="Panel.TLabel", font=("Segoe UI Semibold", 13)).pack(anchor=tk.W, pady=(0, 10))
        text = tk.Text(frame, bg="#111111", fg="#f5f5f5", insertbackground="#ffffff", relief=tk.FLAT, wrap=tk.WORD, height=22)
        text.pack(fill=tk.BOTH, expand=True)
        text.insert("1.0", explanation)
        text.configure(state=tk.DISABLED)
        actions = ttk.Frame(frame, style="App.TFrame")
        actions.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(actions, text="Copy Details", command=lambda: self._copy_inventory_text(explanation, "assignment explanation"), style="Soft.TButton").pack(side=tk.LEFT)
        ttk.Button(actions, text="Close", command=popup.destroy, style="Primary.TButton").pack(side=tk.RIGHT)

    def _workbook_row_for_tree_iid(self, tree: ttk.Treeview, row_id: str) -> WorkbookRow | None:
        try:
            excel_row = int(row_id)
        except (TypeError, ValueError):
            return None
        rows = self.state.rows if hasattr(self, "comp_tree") and tree is self.comp_tree else self.review_rows
        return next((row for row in rows if row.excel_row == excel_row), None)

    def explain_selected_workflow_assignment(self, tree: ttk.Treeview) -> None:
        selected = tree.selection() if tree is not None else ()
        if not selected:
            messagebox.showinfo("Explain Assignment", "Select one row first.")
            return
        row = self._workbook_row_for_tree_iid(tree, selected[0])
        if not row:
            messagebox.showinfo("Explain Assignment", "Could not find that row.")
            return
        label = "Comp row" if hasattr(self, "comp_tree") and tree is self.comp_tree else "Receive row"
        self._show_assignment_explanation_popup(self._assignment_explanation_for_workbook_row(row, label))

    def _assignment_decision_detail(self, decision: assignment_engine.AssignmentDecision) -> str:
        if decision.accepted and decision.payout is not None:
            details: list[str] = []
            rate = getattr(decision, "payout_rate", None)
            if rate is not None:
                percent = rate * 100
                percent_text = f"{percent:.2f}".rstrip("0").rstrip(".")
                details.append(f"{percent_text}%")
            category = str(getattr(decision, "payout_category", "") or "").strip()
            details.append(f"category: {category or 'default/all'}")
            min_price = getattr(decision, "payout_min_price", 0)
            max_price = getattr(decision, "payout_max_price", None)
            if min_price or max_price is not None:
                low = format_money(float(min_price or 0))
                high = format_money(float(max_price)) if max_price is not None else "no max"
                details.append(f"tier: {low} to {high}")
            return " | ".join(details)
        return decision.reason or "no reason returned"

    def explain_selected_inventory_assignment(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        selected = self.inventory_tree.selection()
        if not selected:
            messagebox.showinfo("Explain Assignment", "Select one inventory row first.")
            return
        record = self.inventory_tree_records.get(selected[0])
        if not record:
            messagebox.showinfo("Explain Assignment", "Could not find that inventory row.")
            return
        explanation = self._assignment_explanation_for_record(record)
        self._show_assignment_explanation_popup(explanation)

    def _bind_context_menu(self, widget: tk.Widget, callback) -> None:
        for sequence in ("<Button-3>", "<Button-2>", "<Control-Button-1>", "<Command-Button-1>"):
            widget.bind(sequence, callback, add="+")

    def _show_comp_context_menu(self, event: tk.Event) -> str:
        if not hasattr(self, "comp_tree"):
            return "break"
        row_id = self.comp_tree.identify_row(event.y)
        if not row_id or row_id == ADD_COMP_ROW_IID:
            return "break"
        column_id = self.comp_tree.identify_column(event.x)
        if row_id not in self.comp_tree.selection():
            self.comp_tree.selection_set(row_id)
            self.comp_tree.focus(row_id)
        menu = tk.Menu(self, tearoff=False, bg="#1f1f1f", fg="#ffffff", activebackground="#1ed760", activeforeground="#000000")
        menu.add_command(label="Copy Cell", command=lambda row=row_id, column=column_id: self.copy_tree_cell_value(self.comp_tree, row, column, "comp cell"))
        menu.add_command(label="Copy Row", command=lambda row=row_id: self.copy_tree_row_values(self.comp_tree, row, "comp row"))
        menu.add_separator()
        menu.add_command(label="Explain Assignment", command=lambda target=self.comp_tree: self.explain_selected_workflow_assignment(target))
        menu.add_command(label="Delete Selected", command=self.delete_selected_comp_rows)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        return "break"

    def _show_intake_context_menu(self, event: tk.Event) -> str:
        if not hasattr(self, "intake_tree"):
            return "break"
        row_id = self.intake_tree.identify_row(event.y)
        if not row_id or row_id == ADD_INTAKE_ROW_IID:
            return "break"
        column_id = self.intake_tree.identify_column(event.x)
        if row_id not in self.intake_tree.selection():
            self.intake_tree.selection_set(row_id)
            self.intake_tree.focus(row_id)
        menu = tk.Menu(self, tearoff=False, bg="#1f1f1f", fg="#ffffff", activebackground="#1ed760", activeforeground="#000000")
        menu.add_command(label="Copy Cell", command=lambda row=row_id, column=column_id: self.copy_tree_cell_value(self.intake_tree, row, column, "create cell"))
        menu.add_command(label="Copy Row", command=lambda row=row_id: self.copy_tree_row_values(self.intake_tree, row, "create row"))
        menu.add_separator()
        menu.add_command(label="Delete Selected", command=self.delete_selected_intake_rows)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        return "break"

    def _show_receive_context_menu(self, event: tk.Event) -> str:
        tree = event.widget
        if not self._is_review_row_tree(tree):
            return "break"
        row_id = tree.identify_row(event.y)
        if not row_id or row_id == ADD_REVIEW_ROW_IID:
            return "break"
        column_id = tree.identify_column(event.x)
        if row_id not in tree.selection():
            tree.selection_set(row_id)
            tree.focus(row_id)
        label_prefix = "receive" if self._is_receive_tree(tree) else "assignment"
        menu = tk.Menu(self, tearoff=False, bg="#1f1f1f", fg="#ffffff", activebackground="#1ed760", activeforeground="#000000")
        menu.add_command(label="Copy Cell", command=lambda row=row_id, column=column_id: self.copy_tree_cell_value(tree, row, column, f"{label_prefix} cell"))
        menu.add_command(label="Copy Row", command=lambda row=row_id: self.copy_tree_row_values(tree, row, f"{label_prefix} row"))
        menu.add_separator()
        menu.add_command(label="Explain Assignment", command=lambda target=tree: self.explain_selected_workflow_assignment(target))
        menu.add_command(label="Delete Selected", command=self.delete_selected_review_rows)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        return "break"

    def _show_inventory_context_menu(self, event) -> str:
        if not hasattr(self, "inventory_tree"):
            return "break"
        row_id = self.inventory_tree.identify_row(event.y)
        if not row_id:
            return "break"
        column_id = self.inventory_tree.identify_column(event.x)
        if row_id not in self.inventory_tree.selection():
            self.inventory_tree.selection_set(row_id)
            self.inventory_tree.focus(row_id)
        records = [self.inventory_tree_records.get(iid) for iid in self.inventory_tree.selection()]
        active_records = [record for record in records if record and str(record.get("status") or "").lower() == "active"]
        menu = tk.Menu(self, tearoff=False, bg="#1f1f1f", fg="#ffffff", activebackground="#1ed760", activeforeground="#000000")
        menu.add_command(label="Copy Cell", command=lambda row=row_id, column=column_id: self.copy_inventory_cell_value(row, column))
        menu.add_command(label="Copy Row", command=lambda row=row_id: self.copy_inventory_row_values(row))
        if active_records:
            menu.add_separator()
            menu.add_command(label="Edit Row", command=self.edit_selected_inventory_row)
            menu.add_command(label="Explain Assignment", command=self.explain_selected_inventory_assignment)
        if len(active_records) == 1 and len(records) == 1:
            menu.add_command(label="Attach Photo...", command=self.attach_photo_to_selected_inventory_row)
        if len(records) == 1 and self._inventory_photo_paths_for_record(records[0]):
            menu.add_separator()
            menu.add_command(label="Open Photo", command=self.open_selected_inventory_photo)
            menu.add_command(label="Export Copy to Desktop", command=self.export_selected_inventory_photos_to_desktop)
            menu.add_command(label="Open Photo Folder", command=self.open_selected_inventory_photo_folder)
            menu.add_command(label="Detach Photo...", command=self.detach_photo_from_selected_inventory_row)
        if len(active_records) == 1 and len(records) == 1:
            menu.add_command(label="Mark Sold", command=self.mark_selected_inventory_sold)
        if records and all(self._inventory_record_can_move_to_company_sheet(record) for record in records):
            if len(active_records) == 1 and len(records) == 1:
                menu.add_separator()
            menu.add_command(label="Move to Best Company Sheet", command=self.move_selected_inventory_to_company_sheets)
            menu.add_command(label="Move to Specific Company Sheet...", command=self.move_selected_inventory_to_specific_company_sheet)
        if records:
            menu.add_separator()
            menu.add_command(label="Delete from Inventory", command=self.delete_selected_inventory_records)
        if menu.index("end") is None:
            return "break"
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        return "break"

    def _inventory_photo_paths_for_record(self, record: dict[str, object] | None) -> list[Path]:
        if not record:
            return []
        paths: list[Path] = []
        seen: set[str] = set()
        for value in record.get("photo_paths") or []:
            path = self._resolve_inventory_photo_path(value)
            if not path.exists():
                continue
            try:
                key = self._inventory_photo_file_hash(path)
            except Exception:
                try:
                    key = str(path.resolve())
                except Exception:
                    key = str(path)
            if key in seen:
                continue
            seen.add(key)
            paths.append(path)
        return paths

    def _open_local_path(self, path: Path) -> bool:
        try:
            if sys.platform == "win32":
                os.startfile(str(path))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(path)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                subprocess.Popen(["xdg-open", str(path)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            return True
        except Exception as error:
            self.status_var.set(f"Could not open {path}: {error}")
            return False

    def open_selected_inventory_photo(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        records = [self.inventory_tree_records.get(iid) for iid in self.inventory_tree.selection()]
        record = next((item for item in records if item), None)
        paths = self._inventory_photo_paths_for_record(record)
        if not paths:
            messagebox.showinfo("No photo", "This inventory row does not have an existing linked photo.")
            return
        if len(paths) == 1:
            self._open_local_path(paths[0])
            return
        path = self._choose_inventory_photo_to_open(paths)
        if path:
            self._open_local_path(path)

    def open_selected_inventory_photo_folder(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        records = [self.inventory_tree_records.get(iid) for iid in self.inventory_tree.selection()]
        record = next((item for item in records if item), None)
        paths = self._inventory_photo_paths_for_record(record)
        if not paths:
            messagebox.showinfo("No photo", "This inventory row does not have an existing linked photo.")
            return
        self._open_local_path(paths[0].parent)

    def _desktop_photo_export_destination(self, source_path: Path, record: dict[str, object], index: int = 0, total: int = 1) -> Path:
        desktop = Path.home() / "Desktop"
        title = str(record.get("card_title") or record.get("cert_number") or record.get("inventory_key") or source_path.stem).strip()
        title_slug = re.sub(r"-+", "-", re.sub(r"[^A-Za-z0-9._-]+", "-", title)).strip("-")[:80] or "inventory-photo"
        source_slug = re.sub(r"-+", "-", re.sub(r"[^A-Za-z0-9._-]+", "-", source_path.stem)).strip("-") or "photo"
        suffix = source_path.suffix or ".jpg"
        part = f"-photo-{index + 1}" if total > 1 else ""
        destination = desktop / f"{title_slug}{part}-{source_slug}{suffix}"
        counter = 2
        while destination.exists():
            destination = desktop / f"{title_slug}{part}-{source_slug}-{counter}{suffix}"
            counter += 1
        return destination

    def _export_inventory_photos_to_desktop(self, record: dict[str, object], paths: list[Path]) -> list[Path]:
        if not paths:
            return []
        desktop = Path.home() / "Desktop"
        desktop.mkdir(parents=True, exist_ok=True)
        exported: list[Path] = []
        total = len(paths)
        for index, path in enumerate(paths):
            if not path.exists() or not path.is_file():
                continue
            destination = self._desktop_photo_export_destination(path, record, index, total)
            shutil.copy2(path, destination)
            exported.append(destination)
        return exported

    def export_selected_inventory_photos_to_desktop(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        records = [self.inventory_tree_records.get(iid) for iid in self.inventory_tree.selection()]
        record = next((item for item in records if item), None)
        paths = self._inventory_photo_paths_for_record(record)
        if not record or not paths:
            messagebox.showinfo("No photo", "This inventory row does not have an existing linked photo.")
            return
        try:
            exported = self._export_inventory_photos_to_desktop(record, paths)
        except Exception as error:
            messagebox.showinfo("Export failed", f"Could not export photo copy to Desktop:\n{error}")
            self.status_var.set(f"Photo export failed: {error}")
            return
        if not exported:
            messagebox.showinfo("Export failed", "No existing photo files could be exported.")
            return
        self.status_var.set(f"Exported {len(exported)} inventory photo copy/copies to Desktop.")
        if len(exported) == 1:
            messagebox.showinfo("Export complete", f"Copied photo to Desktop:\n\n{exported[0].name}")
        else:
            messagebox.showinfo("Export complete", f"Copied {len(exported)} photos to Desktop.")

    def _choose_inventory_photo_to_open(self, paths: list[Path]) -> Path | None:
        popup = tk.Toplevel(self)
        popup.title("Open Inventory Photo")
        popup.configure(bg="#121212")
        popup.geometry("760x420")
        popup.transient(self)
        popup.grab_set()
        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(16, 14))
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)
        ttk.Label(frame, text="Open Photo", style="Section.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
        tree = ttk.Treeview(frame, columns=("file", "folder"), show="headings", selectmode="browse")
        tree.heading("file", text="File")
        tree.heading("folder", text="Folder")
        tree.column("file", width=260, anchor="w")
        tree.column("folder", width=420, anchor="w", stretch=True)
        tree.grid(row=1, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        y_scroll.grid(row=1, column=1, sticky="ns")
        tree.configure(yscrollcommand=y_scroll.set)
        path_by_iid: dict[str, Path] = {}
        for index, path in enumerate(paths):
            iid = f"photo-{index}"
            path_by_iid[iid] = path
            tree.insert("", tk.END, iid=iid, values=(path.name, str(path.parent)))
        result: list[Path] = []

        def choose() -> None:
            selected = tree.selection()
            if not selected:
                messagebox.showinfo("Choose photo", "Select a photo to open.")
                return
            result[:] = [path_by_iid[selected[0]]]
            popup.destroy()

        def close() -> None:
            result.clear()
            popup.destroy()

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(12, 0))
        buttons.columnconfigure(0, weight=1)
        ttk.Button(buttons, text="Cancel", command=close, style="Soft.TButton").grid(row=0, column=1, padx=(0, 8))
        ttk.Button(buttons, text="Open Selected", command=choose, style="Primary.TButton").grid(row=0, column=2)
        tree.bind("<Double-1>", lambda _event: choose())
        popup.protocol("WM_DELETE_WINDOW", close)
        popup.update_idletasks()
        x = self.winfo_rootx() + max(70, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(70, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")
        self.wait_window(popup)
        return result[0] if result else None

    def _copy_inventory_photo_attachment(self, source_path: Path, record: dict[str, object]) -> Path:
        source_path = source_path.expanduser()
        if not source_path.exists() or not source_path.is_file():
            raise FileNotFoundError(str(source_path))
        shared = self._inventory_photo_shared_folder()
        shared.mkdir(parents=True, exist_ok=True)
        try:
            source_resolved = source_path.resolve()
            shared_resolved = shared.resolve()
            if source_resolved == shared_resolved or shared_resolved in source_resolved.parents:
                return source_resolved
        except Exception:
            pass

        identifier = str(record.get("item_id") or record.get("cert_number") or record.get("inventory_key") or "inventory").strip()
        identifier = re.sub(r"[^A-Za-z0-9._-]+", "-", identifier).strip("-") or "inventory"
        stem = re.sub(r"[^A-Za-z0-9._-]+", "-", source_path.stem).strip("-") or "photo"
        suffix = source_path.suffix.lower() or ".jpg"
        destination_dir = shared / "manual-attachments"
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = destination_dir / f"{identifier}-{stem}{suffix}"
        counter = 2
        while destination.exists():
            try:
                if destination.stat().st_size == source_path.stat().st_size:
                    return destination
            except Exception:
                pass
            destination = destination_dir / f"{identifier}-{stem}-{counter}{suffix}"
            counter += 1
        shutil.copy2(source_path, destination)
        return destination

    def _inventory_photo_used_path_keys(self, rows: list[dict[str, object]] | None = None) -> set[str]:
        rows = rows if rows is not None else [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        used: set[str] = set()
        for record in rows:
            for value in record.get("photo_paths") or []:
                text = str(value or "").strip()
                if not text:
                    continue
                used.add(text)
                for candidate in self._inventory_photo_path_candidates(text):
                    used.add(str(candidate))
                    try:
                        used.add(str(candidate.resolve()))
                    except Exception:
                        pass
        return used

    def _inventory_photo_used_hashes(self, rows: list[dict[str, object]] | None = None) -> set[str]:
        rows = rows if rows is not None else [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        used_hashes: set[str] = set()
        seen_paths: set[str] = set()
        for record in rows:
            for value in record.get("photo_paths") or []:
                for candidate in self._inventory_photo_path_candidates(value):
                    if not candidate.exists() or not candidate.is_file():
                        continue
                    try:
                        key = str(candidate.resolve())
                    except Exception:
                        key = str(candidate)
                    if key in seen_paths:
                        continue
                    seen_paths.add(key)
                    try:
                        used_hashes.add(self._inventory_photo_file_hash(candidate))
                    except Exception:
                        continue
        return used_hashes

    def _inventory_photo_state_used_keys(self) -> tuple[set[str], set[str], set[str]]:
        state = self._load_inventory_photo_state()
        photos = state.get("photos") if isinstance(state, dict) else {}
        if not isinstance(photos, dict):
            return set(), set(), set()
        used_names: set[str] = set()
        used_paths: set[str] = set()
        used_hashes: set[str] = set()
        used_statuses = {"linked", "missing_from_album", "archived_from_album"}
        for sha, record in photos.items():
            if not isinstance(record, dict):
                continue
            status = str(record.get("status") or "").strip()
            linked_keys = [str(key).strip() for key in (record.get("linked_keys") or []) if str(key).strip()]
            if status not in used_statuses and not linked_keys:
                continue
            sha_text = str(sha or "").strip()
            if sha_text:
                used_hashes.add(sha_text)
            for value in (
                record.get("path"),
                record.get("relative_path"),
                record.get("filename"),
                record.get("archived_path"),
                record.get("original_path"),
            ):
                text = str(value or "").strip()
                if not text:
                    continue
                used_paths.add(text)
                try:
                    path = Path(text).expanduser()
                    used_names.add(path.name)
                    used_paths.add(str(path))
                    if path.exists() and path.is_file():
                        used_hashes.add(self._inventory_photo_file_hash(path))
                    try:
                        used_paths.add(str(path.resolve()))
                    except Exception:
                        pass
                except Exception:
                    used_names.add(Path(text).name)
        return used_names, used_paths, used_hashes

    def _sold_inventory_cert_numbers(self) -> set[str]:
        certs: set[str] = set()
        for record in self._load_profit_ledger():
            normalized = self._normalize_profit_record(record)
            cert = scan_to_cert(normalized.get("cert_number"))
            if cert:
                certs.add(cert)
        return certs

    def _sold_inventory_photo_used_keys(self) -> tuple[set[str], set[str]]:
        sold_rows = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
        return self._inventory_photo_used_path_keys(sold_rows), self._inventory_photo_used_hashes(sold_rows)

    def _inventory_photo_image_matches_sold_photo(
        self,
        image: dict[str, object],
        sold_photo_paths: set[str],
        sold_photo_hashes: set[str],
    ) -> bool:
        sha = str(image.get("sha256") or "").strip()
        if sha and sha in sold_photo_hashes:
            return True
        image_path = Path(str(image.get("path") or "")).expanduser()
        keys = {
            str(image_path),
            str(image.get("relative_path") or "").strip(),
            str(image.get("filename") or "").strip(),
        }
        storage = self._inventory_photo_storage_value(image_path)
        if storage:
            keys.add(storage)
        try:
            keys.add(str(image_path.resolve()))
        except Exception:
            pass
        return bool({key for key in keys if key} & sold_photo_paths)

    def _inventory_photo_state_matches_sold_cert(self, existing: dict[str, object], sold_certs: set[str] | None = None) -> bool:
        if not existing:
            return False
        sold_certs = sold_certs if sold_certs is not None else self._sold_inventory_cert_numbers()
        if not sold_certs:
            return False
        photo_certs = {scan_to_cert(cert) for cert in (existing.get("certs") or []) if scan_to_cert(cert)}
        return bool(photo_certs & sold_certs)

    def _inventory_unattached_photo_paths(self) -> list[Path]:
        used = self._inventory_photo_used_path_keys()
        used_hashes = self._inventory_photo_used_hashes()
        state_used_names, state_used_paths, state_used_hashes = self._inventory_photo_state_used_keys()
        sold_certs = self._sold_inventory_cert_numbers()
        sold_photo_source = getattr(self, "_sold_inventory_photo_used_keys", None)
        sold_photo_paths, sold_photo_hashes = sold_photo_source() if callable(sold_photo_source) else (set(), set())
        state = self._load_inventory_photo_state()
        photos = state.get("photos") if isinstance(state, dict) else {}
        photos = photos if isinstance(photos, dict) else {}
        paths: list[Path] = []
        seen: set[str] = set()
        for folder in (self._inventory_photo_shared_folder(), self._inventory_photo_source_folder()):
            for path in self._inventory_photo_paths(folder):
                if path.name in state_used_names:
                    continue
                keys = {str(path)}
                try:
                    keys.add(str(path.resolve()))
                except Exception:
                    pass
                storage = self._inventory_photo_storage_value(path)
                if storage:
                    keys.add(storage)
                if keys & used or keys & state_used_paths or keys & sold_photo_paths:
                    continue
                unique_key = ""
                try:
                    unique_key = self._inventory_photo_file_hash(path)
                except Exception:
                    unique_key = ""
                if unique_key and (unique_key in used_hashes or unique_key in state_used_hashes or unique_key in sold_photo_hashes):
                    continue
                existing_state = photos.get(unique_key) if unique_key and isinstance(photos.get(unique_key), dict) else {}
                if existing_state and self._inventory_photo_state_matches_sold_cert(existing_state, sold_certs):
                    continue
                if not unique_key:
                    unique_key = next(iter(keys))
                    try:
                        unique_key = str(path.resolve())
                    except Exception:
                        pass
                try:
                    resolved_key = str(path.resolve())
                except Exception:
                    resolved_key = str(path)
                if unique_key in seen or resolved_key in seen:
                    continue
                seen.add(unique_key)
                seen.add(resolved_key)
                paths.append(path)
        return sorted(paths, key=lambda item: str(item).lower())

    def _inventory_photo_preview_image(self, path: Path, size: tuple[int, int] = (260, 340)) -> tk.PhotoImage | None:
        try:
            if path.suffix.lower() in {".heic", ".heif"}:
                try:
                    import pillow_heif

                    pillow_heif.register_heif_opener()
                except Exception:
                    return None
            from PIL import Image, ImageOps, ImageTk

            with Image.open(path) as opened:
                image = ImageOps.exif_transpose(opened)
                if image.mode not in {"RGB", "RGBA"}:
                    image = image.convert("RGB")
                image.thumbnail(size, Image.LANCZOS)
                return ImageTk.PhotoImage(image)
        except Exception:
            return None

    def _choose_unattached_inventory_photos(self, max_count: int = MAX_INVENTORY_PHOTOS_PER_CARD) -> list[Path]:
        available = self._inventory_unattached_photo_paths()
        if not available:
            messagebox.showinfo("No unattached photos", "No unattached inventory photos were found in the configured photo bucket.")
            return []
        popup = tk.Toplevel(self)
        popup.title("Attach Unattached Inventory Photos")
        popup.configure(bg="#121212")
        popup.geometry("980x560")
        popup.transient(self)
        popup.grab_set()
        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(16, 14))
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=0)
        frame.rowconfigure(2, weight=1)
        title_text = "Unattached Photos" if max_count >= MAX_INVENTORY_PHOTOS_PER_CARD else f"Unattached Photos ({max_count} slot(s) left)"
        ttk.Label(frame, text=title_text, style="Section.TLabel").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))
        search_var = tk.StringVar()
        search = ttk.Entry(frame, textvariable=search_var)
        search.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        columns = ("filename", "folder", "modified")
        tree_frame = ttk.Frame(frame, style="Panel.TFrame")
        tree_frame.grid(row=2, column=0, sticky="nsew")
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="extended")
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        headings = {"filename": "File", "folder": "Folder", "modified": "Modified"}
        widths = {"filename": 360, "folder": 420, "modified": 150}
        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=widths[column], anchor="w", stretch=column != "modified")
        preview_frame = ttk.Frame(frame, style="Panel.TFrame", padding=(12, 0, 0, 0))
        preview_frame.grid(row=2, column=1, sticky="ns")
        ttk.Label(preview_frame, text="Photo Preview", style="Muted.TLabel").pack(anchor="w", pady=(0, 6))
        preview_label = ttk.Label(preview_frame, style="Panel.TLabel")
        preview_label.pack(anchor="n")
        preview_detail = ttk.Label(preview_frame, text="Select a photo.", style="Muted.TLabel", wraplength=260, justify=tk.LEFT)
        preview_detail.pack(anchor="w", fill=tk.X, pady=(8, 0))
        preview_image_ref: dict[str, tk.PhotoImage | None] = {"image": None}
        path_by_iid: dict[str, Path] = {}
        selected_paths: list[Path] = []

        def row_text(path: Path) -> str:
            try:
                rel = path.relative_to(self._inventory_photo_picker_initial_dir()).as_posix()
            except Exception:
                rel = str(path)
            return f"{path.name} {path.parent} {rel}".lower()

        searchable = [(path, row_text(path)) for path in available]

        def render() -> None:
            query = search_var.get().strip().lower()
            tree.delete(*tree.get_children())
            path_by_iid.clear()
            for index, (path, text) in enumerate(searchable):
                if query and query not in text:
                    continue
                try:
                    modified = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %I:%M %p")
                except Exception:
                    modified = ""
                try:
                    folder = path.parent.relative_to(self._inventory_photo_picker_initial_dir()).as_posix() or "."
                except Exception:
                    folder = str(path.parent)
                iid = f"photo-{index}"
                path_by_iid[iid] = path
                tree.insert("", tk.END, iid=iid, values=(path.name, folder, modified))
            first = tree.get_children()
            if first:
                tree.selection_set(first[0])
                tree.focus(first[0])
                update_preview()
            else:
                update_preview()

        def update_preview(_event: tk.Event | None = None) -> None:
            selected = [iid for iid in tree.selection() if iid in path_by_iid]
            path = path_by_iid[selected[0]] if selected else None
            if path is None:
                preview_image_ref["image"] = None
                preview_label.configure(image="")
                preview_detail.configure(text="No matching photos.")
                return
            image = self._inventory_photo_preview_image(path)
            preview_image_ref["image"] = image
            if image is None:
                preview_label.configure(image="")
                preview_detail.configure(text=f"Preview unavailable\n{path.name}")
                return
            preview_label.configure(image=image)
            try:
                relative = path.relative_to(self._inventory_photo_picker_initial_dir()).as_posix()
            except Exception:
                relative = str(path)
            preview_detail.configure(text=relative)

        def attach() -> None:
            selected_paths[:] = [path_by_iid[iid] for iid in tree.selection() if iid in path_by_iid]
            if not selected_paths:
                messagebox.showinfo("Choose photo", "Select one or more unattached photos to attach.")
                return
            if len(selected_paths) > max_count:
                messagebox.showinfo("Too many photos", f"Select up to {max_count} photo(s) for this card.")
                return
            popup.destroy()

        def close() -> None:
            selected_paths.clear()
            popup.destroy()

        search_var.trace_add("write", lambda *_: render())
        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(12, 0))
        buttons.columnconfigure(0, weight=1)
        ttk.Button(buttons, text="Cancel", command=close, style="Soft.TButton").grid(row=0, column=1, padx=(0, 8))
        ttk.Button(buttons, text="Attach Selected", command=attach, style="Primary.TButton").grid(row=0, column=2)
        tree.bind("<<TreeviewSelect>>", update_preview)
        tree.bind("<Double-1>", lambda _event: attach())
        popup.protocol("WM_DELETE_WINDOW", close)
        render()
        search.focus_set()
        popup.update_idletasks()
        x = self.winfo_rootx() + max(50, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(50, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")
        self.wait_window(popup)
        return selected_paths

    def _inventory_photo_group_key_for_path(self, path: Path) -> str:
        return self._inventory_photo_capture_group_key({"filename": path.name, "relative_path": path.name})

    def _expand_inventory_photo_group_selection(self, selected_paths: list[Path], max_count: int) -> list[Path]:
        if not selected_paths or max_count <= 0:
            return []
        selected_keys = {self._inventory_photo_group_key_for_path(path) for path in selected_paths}
        selected_keys.discard("")
        expanded: list[Path] = []
        seen: set[str] = set()

        def add_path(path: Path) -> None:
            if len(expanded) >= max_count:
                return
            key = str(path)
            try:
                key = str(path.resolve())
            except Exception:
                pass
            if key in seen:
                return
            seen.add(key)
            expanded.append(path)

        for path in selected_paths:
            add_path(Path(path))
        if not selected_keys or len(expanded) >= max_count:
            return expanded
        for path in self._inventory_unattached_photo_paths():
            if len(expanded) >= max_count:
                break
            if self._inventory_photo_group_key_for_path(path) in selected_keys:
                add_path(path)
        return expanded

    def attach_photo_to_selected_inventory_row(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        records = [self.inventory_tree_records.get(iid) for iid in self.inventory_tree.selection()]
        selected = [(iid, record) for iid, record in zip(self.inventory_tree.selection(), records) if record]
        active_selected = [(iid, record) for iid, record in selected if str(record.get("status") or "").lower() == "active"]
        if len(active_selected) != 1:
            messagebox.showinfo("Choose one card", "Select one active inventory row to attach photo(s).")
            return
        iid, record = active_selected[0]
        existing_paths = [str(value or "").strip() for value in (record.get("photo_paths") or []) if str(value or "").strip()]
        remaining_slots = MAX_INVENTORY_PHOTOS_PER_CARD - len(existing_paths)
        if remaining_slots <= 0:
            messagebox.showinfo("Photo limit reached", f"This inventory row already has {MAX_INVENTORY_PHOTOS_PER_CARD} photo(s). Detach one before adding another.")
            return
        paths = self._choose_unattached_inventory_photos(max_count=remaining_slots)
        if not paths:
            return
        selected_count = len(paths)
        paths = self._expand_inventory_photo_group_selection(paths, max_count=remaining_slots)
        copied_paths: list[Path] = []
        errors: list[str] = []
        for raw_path in paths:
            try:
                copied_paths.append(self._copy_inventory_photo_attachment(Path(raw_path), record))
            except Exception as error:
                errors.append(f"{Path(raw_path).name}: {error}")
        if not copied_paths:
            self._show_copyable_error("Attach photo failed", "\n".join(errors) or "No photo files were attached.")
            return
        existing_set = set(existing_paths)
        added = 0
        for copied_path in copied_paths:
            if len(existing_paths) >= MAX_INVENTORY_PHOTOS_PER_CARD:
                break
            path_text = self._inventory_photo_storage_value(copied_path)
            if path_text not in existing_set:
                existing_paths.append(path_text)
                existing_set.add(path_text)
                added += 1
        if not added:
            self.inventory_status_var.set("Selected photo(s) were already attached to this inventory row.")
            return
        updates = {"photo_paths": existing_paths, "photo_count": len(existing_paths)}
        changed = self._update_inventory_record_by_key(str(record.get("inventory_key") or ""), updates)
        if changed:
            updated_record = dict(record)
            updated_record.update(updates)
            normalized = self._normalize_inventory_record(updated_record)
            self.inventory_tree_records[iid] = normalized
            self._refresh_inventory_tree_row(iid, normalized)
            grouped_text = f" including {len(paths) - selected_count} grouped photo(s)" if len(paths) > selected_count else ""
            self.inventory_status_var.set(f"Attached {added} photo(s) to inventory row{grouped_text}.")
            self.status_var.set(f"Attached {added} inventory photo(s).")
            self._append_activity(
                "Inventory Photo Attach",
                f"Attached {added} photo(s) to inventory row.",
                {
                    "inventory_key": normalized.get("inventory_key") or "",
                    "item_id": normalized.get("item_id") or "",
                    "cert_number": normalized.get("cert_number") or "",
                    "paths": [str(path) for path in copied_paths[:10]],
                },
            )
        if errors:
            self._show_copyable_error("Attach photo warning", "\n".join(errors[:10]))

    def detach_photo_from_selected_inventory_row(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        selected = list(self.inventory_tree.selection())
        if len(selected) != 1:
            messagebox.showinfo("Choose one card", "Select one inventory row to detach photo(s).")
            return
        iid = selected[0]
        record = self.inventory_tree_records.get(iid)
        if not record:
            return
        photo_values = [str(value or "").strip() for value in (record.get("photo_paths") or []) if str(value or "").strip()]
        if not photo_values:
            messagebox.showinfo("No photo", "This inventory row does not have an attached photo.")
            return
        remove_values: list[str] = []
        if len(photo_values) == 1:
            if not messagebox.askyesno("Detach photo", f"Detach this photo from the inventory row?\n\n{photo_values[0]}\n\nThe photo file will not be deleted."):
                return
            remove_values = photo_values
        else:
            remove_values = self._choose_inventory_photos_to_detach(photo_values)
            if not remove_values:
                return
        remaining = [value for value in photo_values if value not in set(remove_values)]
        updates = {"photo_paths": remaining, "photo_count": len(remaining)}
        changed = self._update_inventory_record_by_key(str(record.get("inventory_key") or ""), updates)
        if not changed:
            return
        self._remove_inventory_photo_state_links(str(record.get("inventory_key") or ""), remove_values)
        updated_record = dict(record)
        updated_record.update(updates)
        normalized = self._normalize_inventory_record(updated_record)
        self.inventory_tree_records[iid] = normalized
        self._refresh_inventory_tree_row(iid, normalized)
        self.inventory_status_var.set(f"Detached {len(remove_values)} photo(s) from inventory row.")
        self.status_var.set(f"Detached {len(remove_values)} inventory photo(s).")
        self._append_activity(
            "Inventory Photo Detach",
            f"Detached {len(remove_values)} photo(s) from inventory row.",
            {
                "inventory_key": normalized.get("inventory_key") or "",
                "cert_number": normalized.get("cert_number") or "",
                "paths": remove_values[:10],
            },
        )

    def _choose_inventory_photos_to_detach(self, photo_values: list[str]) -> list[str]:
        popup = tk.Toplevel(self)
        popup.title("Detach Inventory Photos")
        popup.configure(bg="#121212")
        popup.geometry("760x420")
        popup.transient(self)
        popup.grab_set()
        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(16, 14))
        frame.pack(fill=tk.BOTH, expand=True)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)
        ttk.Label(frame, text="Detach Photos", style="Section.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
        tree = ttk.Treeview(frame, columns=("path",), show="headings", selectmode="extended")
        tree.heading("path", text="Attached Photo")
        tree.column("path", width=680, anchor="w", stretch=True)
        tree.grid(row=1, column=0, sticky="nsew")
        y_scroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        y_scroll.grid(row=1, column=1, sticky="ns")
        tree.configure(yscrollcommand=y_scroll.set)
        for index, value in enumerate(photo_values):
            tree.insert("", tk.END, iid=f"photo-{index}", values=(value,))
        result: list[str] = []

        def detach() -> None:
            result[:] = [str(tree.item(iid, "values")[0]) for iid in tree.selection()]
            if not result:
                messagebox.showinfo("Choose photo", "Select one or more photos to detach.")
                return
            popup.destroy()

        def close() -> None:
            result.clear()
            popup.destroy()

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(12, 0))
        buttons.columnconfigure(0, weight=1)
        ttk.Button(buttons, text="Cancel", command=close, style="Soft.TButton").grid(row=0, column=1, padx=(0, 8))
        ttk.Button(buttons, text="Detach Selected", command=detach, style="Primary.TButton").grid(row=0, column=2)
        popup.protocol("WM_DELETE_WINDOW", close)
        popup.update_idletasks()
        x = self.winfo_rootx() + max(70, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(70, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")
        self.wait_window(popup)
        return result

    def _remove_inventory_photo_state_links(self, inventory_key: str, removed_values: list[str]) -> None:
        if not inventory_key or not removed_values:
            return
        removed_keys: set[str] = set(removed_values)
        for value in removed_values:
            for candidate in self._inventory_photo_path_candidates(value):
                removed_keys.add(str(candidate))
                try:
                    removed_keys.add(str(candidate.resolve()))
                except Exception:
                    pass
        state = self._load_inventory_photo_state()
        photos = state.get("photos")
        if not isinstance(photos, dict):
            return
        changed = False
        for entry in photos.values():
            if not isinstance(entry, dict):
                continue
            linked_keys = [str(key) for key in (entry.get("linked_keys") or [])]
            if inventory_key not in linked_keys:
                continue
            entry_path = str(entry.get("path") or entry.get("relative_path") or "")
            entry_keys = {entry_path}
            for candidate in self._inventory_photo_path_candidates(entry_path):
                entry_keys.add(str(candidate))
                try:
                    entry_keys.add(str(candidate.resolve()))
                except Exception:
                    pass
            if not (entry_keys & removed_keys):
                continue
            linked_keys = [key for key in linked_keys if key != inventory_key]
            entry["linked_keys"] = linked_keys
            entry["status"] = "linked" if linked_keys else "no_matching_inventory"
            entry["detached_at"] = datetime.now().isoformat(timespec="seconds")
            changed = True
        if changed:
            self._save_inventory_photo_state(state)

    def refresh_inventory_tab(self, reconcile: bool = False, enrich: bool = False, filtered_only: bool = False) -> None:
        perf_start = time.perf_counter()
        self._last_inventory_enrich_visible_count = 0
        self._last_inventory_enrich_changed_count = 0
        if getattr(self, "inventory_filter_after_id", None):
            try:
                self.after_cancel(self.inventory_filter_after_id)
            except tk.TclError:
                pass
            self.inventory_filter_after_id = None
        self._inventory_source_rows_cache = {}
        if reconcile and not getattr(self, "_inventory_reconcile_running", False):
            self._inventory_reconcile_running = True
            try:
                self._sync_received_inventory_to_ledger(filtered_only=filtered_only)
            finally:
                self._inventory_reconcile_running = False
        all_stored_rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        non_active_stored_rows = [record for record in all_stored_rows if str(record.get("status") or "").lower() != "active"]
        stored_rows = [record for record in all_stored_rows if str(record.get("status") or "").lower() == "active"]
        if enrich and filtered_only:
            filtered_keys = {str(record.get("inventory_key") or "") for record in self._filtered_inventory_records(stored_rows)}
            self._last_inventory_enrich_visible_count = len(filtered_keys)
            inventory_rows: list[dict[str, object]] = []
            for record in stored_rows:
                if str(record.get("inventory_key") or "") not in filtered_keys:
                    inventory_rows.append(record)
                    continue
                enriched = self._enrich_inventory_record_assignment(record, force=True)
                if enriched != record:
                    self._last_inventory_enrich_changed_count += 1
                inventory_rows.append(enriched)
            self.inventory_rows = inventory_rows
        else:
            if enrich:
                self._last_inventory_enrich_visible_count = len(stored_rows)
                inventory_rows = []
                for record in stored_rows:
                    enriched = self._enrich_inventory_record_assignment(record)
                    if enriched != record:
                        self._last_inventory_enrich_changed_count += 1
                    inventory_rows.append(enriched)
                self.inventory_rows = inventory_rows
            else:
                self.inventory_rows = stored_rows
        if enrich and self.inventory_rows != stored_rows:
            self._save_inventory_ledger([*non_active_stored_rows, *self.inventory_rows])
        self.filtered_inventory_rows = self._filtered_inventory_records(self.inventory_rows)
        if hasattr(self, "_sorted_records"):
            self.filtered_inventory_rows = self._sorted_records(
                self.filtered_inventory_rows,
                getattr(self, "inventory_sort_column", "date"),
                bool(getattr(self, "inventory_sort_descending", True)),
                "inventory",
            )
        if not hasattr(self, "inventory_tree"):
            record_performance_event(
                "inventory.refresh",
                perf_start,
                f"rows={len(self.inventory_rows)} filtered={len(self.filtered_inventory_rows)} reconcile={reconcile} enrich={enrich} filtered_only={filtered_only} tree=missing",
            )
            return
        self._refresh_person_combo_values()
        if hasattr(self, "_configure_sortable_tree_headings"):
            self._configure_sortable_tree_headings(self.inventory_tree, INVENTORY_HEADINGS, "inventory")
        self.inventory_tree.delete(*self.inventory_tree.get_children())
        self.inventory_tree_records = {}
        total_purchase = 0.0
        total_value = 0.0
        for record in self.filtered_inventory_rows:
            purchase = self._money_value(record.get("purchase_price"))
            value = self._money_value(record.get("inventory_value"))
            card_ladder = self._money_value(record.get("card_ladder_value"))
            comps = self._money_value(record.get("card_ladder_comps_average"))
            cy_value = self._money_value(record.get("cy_value"))
            if purchase is not None:
                total_purchase += purchase
            if value is not None:
                total_value += value
            values_by_column = {
                "date": record.get("date_added") or "",
                "type": record.get("item_type") or "",
                "item_id": record.get("item_id") or "",
                "person": record.get("assigned_person") or "Unassigned",
                "sport": record.get("sport") or "",
                "cert": record.get("cert_number") or "",
                "grader": record.get("grader") or "",
                "card": record.get("card_title") or "",
                "purchase": format_money(purchase),
                "card_ladder": format_money(card_ladder),
                "comps": format_money(comps),
                "cy_estimate": format_money(cy_value),
                "cy_confidence": record.get("cy_confidence") if record.get("cy_confidence") is not None else "",
                "company": record.get("best_company") or "",
                "payout": format_money(record.get("estimated_payout")),
                "paid_with": record.get("paid_with") or "",
                "source": record.get("source_sheet") or "",
                "status": record.get("status") or "",
                "photos": str(len(record.get("photo_paths") or [])),
                "notes": inventory_display_notes(record),
            }
            iid = self.inventory_tree.insert(
                "",
                tk.END,
                values=tuple(values_by_column.get(column, "") for column in self.inventory_tree["columns"]),
            )
            self.inventory_tree_records[iid] = record
        self.inventory_metric_var.set(f"Cards: {len(self.filtered_inventory_rows)}   Purchase Total: {format_money(total_purchase)}   Source Value: {format_money(total_value)}")
        self.inventory_status_var.set(f"Loaded {len(self.filtered_inventory_rows)}/{len(self.inventory_rows)} inventory card(s) from {INVENTORY_LEDGER_PATH.name}.")
        record_performance_event(
            "inventory.refresh",
            perf_start,
            f"rows={len(self.inventory_rows)} filtered={len(self.filtered_inventory_rows)} reconcile={reconcile} enrich={enrich} filtered_only={filtered_only} changed={self._last_inventory_enrich_changed_count}",
        )

    def _schedule_inventory_filter_refresh(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        if getattr(self, "inventory_filter_after_id", None):
            try:
                self.after_cancel(self.inventory_filter_after_id)
            except tk.TclError:
                pass
        self.inventory_filter_after_id = self.after(150, self.refresh_inventory_tab)

    def update_inventory_payouts(self) -> None:
        perf_start = time.perf_counter()
        try:
            assignment_start = time.perf_counter()
            self._load_player_overrides()
            self.assignment_engine = AssignmentEngine.load()
            record_performance_event("assignment.load.inventory_payouts", assignment_start, f"companies={len(self.assignment_engine.companies)}")
        except Exception:
            pass
        self.refresh_inventory_tab(enrich=True, filtered_only=True)
        visible_count = int(getattr(self, "_last_inventory_enrich_visible_count", 0) or len(getattr(self, "filtered_inventory_rows", [])))
        changed_count = int(getattr(self, "_last_inventory_enrich_changed_count", 0) or 0)
        self.inventory_status_var.set(f"Updated payouts for {visible_count} visible inventory card(s); changed {changed_count}.")
        record_performance_event("inventory.update_payouts", perf_start, f"visible={visible_count} changed={changed_count}")

    def _comp_low_outlier_pct(self) -> float | None:
        text = str(self.comp_low_outlier_pct_var.get() if hasattr(self, "comp_low_outlier_pct_var") else "").strip()
        if not text or text.lower() == "off":
            return None
        value = self._money_value(text.rstrip("%"))
        if value is not None and value > 100:
            value = 100.0
        return value if value is not None and value > 0 else None

    def _lot_purchase_base_value(self, row: WorkbookRow, source: str) -> float | None:
        label = str(source or "").strip().lower()
        if label == "card ladder value":
            return self._money_value(row.card_ladder_value)
        if label == "comps average":
            return self._money_value(row.card_ladder_comps_average)
        if label == "cy estimate":
            return self._money_value(row.cy_value)
        return None

    def _lot_purchase_allocations(self, rows: list[WorkbookRow], lot_total: float, percent: float, source: str) -> tuple[list[float], dict[str, object]]:
        target = max(float(lot_total or 0), 0.0)
        rate = max(float(percent or 0), 0.0) / 100.0
        allocations: list[float] = []
        running = 0.0
        capped = False
        value_missing = 0
        for row in rows:
            if capped:
                allocations.append(0.0)
                continue
            base_value = self._lot_purchase_base_value(row, source)
            if base_value is None:
                value_missing += 1
                planned = 0.0
            else:
                planned = round(base_value * rate, 2)
            remaining = round(target - running, 2)
            if planned >= remaining and remaining > 0:
                allocations.append(remaining)
                running = target
                capped = True
                continue
            allocations.append(planned)
            running = round(running + planned, 2)
        info = {
            "allocated": round(sum(allocations), 2),
            "target": round(target, 2),
            "remaining": round(target - sum(allocations), 2),
            "capped": capped,
            "value_missing": value_missing,
        }
        return allocations, info

    def open_lot_purchase_fill_popup(self) -> None:
        with self.state.lock:
            rows = list(self.state.rows)
        if not rows:
            messagebox.showinfo("No comp rows", "Load or add comp rows before filling purchase prices.")
            return
        total_var = tk.StringVar()
        percent_var = tk.StringVar(value="70")
        source_var = tk.StringVar(value="Comps Average")
        popup = tk.Toplevel(self)
        popup.title("Lot Price Fill")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)
        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Lot Price Fill", style="Panel.TLabel", font=("Segoe UI Semibold", 13)).grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Label(frame, text=f"Comp rows: {len(rows)}", style="Muted.TLabel").grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 14))
        ttk.Label(frame, text="Lot buy price", style="Panel.TLabel").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=(0, 8))
        ttk.Entry(frame, textvariable=total_var, width=18).grid(row=2, column=1, sticky="w", pady=(0, 8))
        ttk.Label(frame, text="Percent of value", style="Panel.TLabel").grid(row=3, column=0, sticky="w", padx=(0, 10), pady=(0, 8))
        ttk.Entry(frame, textvariable=percent_var, width=18).grid(row=3, column=1, sticky="w", pady=(0, 8))
        ttk.Label(frame, text="Value source", style="Panel.TLabel").grid(row=4, column=0, sticky="w", padx=(0, 10), pady=(0, 8))
        ttk.Combobox(
            frame,
            textvariable=source_var,
            values=COMP_LOT_VALUE_SOURCE_OPTIONS,
            width=20,
            state="readonly",
        ).grid(row=4, column=1, sticky="w", pady=(0, 8))

        def submit() -> None:
            total = self._money_value(total_var.get())
            percent = self._money_value(percent_var.get())
            if total is None or total <= 0:
                messagebox.showinfo("Lot buy price", "Enter the total lot buy price.", parent=popup)
                return
            if percent is None or percent <= 0:
                messagebox.showinfo("Percent", "Enter the percent of value to use.", parent=popup)
                return
            popup.destroy()
            self.apply_lot_purchase_fill(float(total), float(percent), source_var.get())

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=5, column=0, columnspan=2, sticky="e", pady=(12, 0))
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="Apply", command=submit, style="Primary.TButton").pack(side=tk.LEFT)
        popup.bind("<Return>", lambda _event: submit())
        popup.bind("<Escape>", lambda _event: popup.destroy())
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")

    def apply_lot_purchase_fill(self, lot_total: float, percent: float, source: str) -> None:
        with self.state.lock:
            rows = list(self.state.rows)
            allocations, info = self._lot_purchase_allocations(rows, lot_total, percent, source)
            for row, allocation in zip(rows, allocations):
                row.existing_value = allocation
        self.comp_output_saved = False
        self._refresh_comp_table(schedule_recommendations=True)
        message = f"Filled purchase prices: {format_money(info['allocated'])} across {len(rows)} comp row(s)."
        if info.get("capped"):
            message += " Lot total was hit; remaining rows were set to $0.00."
            messagebox.showwarning("Lot total hit", message)
        elif info.get("remaining"):
            message += f" Remaining unallocated balance: {format_money(info['remaining'])}."
            messagebox.showwarning("Lot total not fully allocated", message)
        elif info.get("value_missing"):
            message += f" {info['value_missing']} row(s) had no value and received $0.00."
        self.status_var.set(message)

    def open_inventory_recomp_popup(self) -> None:
        self.refresh_inventory_tab()
        visible_count = len(getattr(self, "filtered_inventory_rows", []))
        eligible_count = sum(
            1
            for record in getattr(self, "filtered_inventory_rows", [])
            if scan_to_cert(record.get("cert_number")) and str(record.get("grader") or "").strip()
        )
        cl_value_var = tk.BooleanVar(value=True)
        cl_comps_var = tk.BooleanVar(value=True)
        cy_var = tk.BooleanVar(value=False)
        strategy_var = tk.StringVar(value=self.comp_strategy_label.get() or "Average last 5")
        scope_var = tk.StringVar(value=COMP_SCOPE_EMPTY)

        popup = tk.Toplevel(self)
        popup.title("Inventory Recomp")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Inventory Recomp", style="Panel.TLabel", font=("Segoe UI Semibold", 13)).grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Label(frame, text=f"Filtered rows: {visible_count}   Eligible: {eligible_count}", style="Muted.TLabel").grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 14))
        ttk.Checkbutton(frame, text="Card Ladder value", variable=cl_value_var, style="Panel.TCheckbutton").grid(row=2, column=0, columnspan=2, sticky="w", pady=(0, 8))
        ttk.Checkbutton(frame, text="Card Ladder comps", variable=cl_comps_var, style="Panel.TCheckbutton").grid(row=3, column=0, columnspan=2, sticky="w", pady=(0, 8))
        if COMP_CY_ENABLED:
            ttk.Checkbutton(frame, text="CY estimate", variable=cy_var, style="Panel.TCheckbutton").grid(row=4, column=0, columnspan=2, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="Scope", style="Muted.TLabel").grid(row=5, column=0, sticky="w", pady=(0, 6))
        ttk.Combobox(frame, textvariable=scope_var, values=(COMP_SCOPE_EMPTY, COMP_SCOPE_ALL), width=24, state="readonly").grid(row=5, column=1, sticky="w", pady=(0, 6))
        ttk.Label(frame, text="Comp Method", style="Muted.TLabel").grid(row=6, column=0, sticky="w", pady=(0, 6))
        ttk.Combobox(frame, textvariable=strategy_var, values=list(COMP_STRATEGY_DISPLAY.keys()), width=24, state="readonly").grid(row=6, column=1, sticky="w", pady=(0, 6))

        def submit() -> None:
            features = {
                "card_ladder_value": bool(cl_value_var.get()),
                "card_ladder_comps": bool(cl_comps_var.get()),
                "cy": bool(cy_var.get()) and COMP_CY_ENABLED,
                "strategy_label": strategy_var.get(),
                "scope": scope_var.get(),
            }
            if not any(features.get(key) for key in ("card_ladder_value", "card_ladder_comps", "cy")):
                messagebox.showinfo("Choose recomp features", "Choose at least one field to refresh.", parent=popup)
                return
            popup.destroy()
            self.recomp_inventory_visible_rows(features)

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=7, column=0, columnspan=2, sticky="e", pady=(12, 0))
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="Start Recomp", command=submit, style="Primary.TButton").pack(side=tk.LEFT)
        popup.bind("<Return>", lambda _event: submit())
        popup.bind("<Escape>", lambda _event: popup.destroy())
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")

    def recomp_inventory_visible_rows(self, features: dict[str, object] | None = None) -> None:
        features = dict(features or {"card_ladder_value": True, "card_ladder_comps": True, "cy": False, "strategy_label": self.comp_strategy_label.get(), "scope": COMP_SCOPE_EMPTY})
        if not COMP_CY_ENABLED:
            features["cy"] = False
        if self.inventory_recomp_context:
            messagebox.showinfo("Inventory recomp running", "Wait for the current Inventory recomp run to finish.")
            return
        run_card_ladder = bool(features.get("card_ladder_value") or features.get("card_ladder_comps"))
        run_cy = bool(features.get("cy")) and COMP_CY_ENABLED
        if run_card_ladder:
            extension_warning = self._cardladder_extension_warning()
            if extension_warning:
                messagebox.showwarning("Reload Card Ladder extension", extension_warning)
                self.inventory_status_var.set("Reload the Card Ladder Chrome extension before recomping inventory.")
                return
        with self.state.lock:
            comp_busy = bool(
                self.state.command
                or self.state.cardladder_running
                or getattr(self.state, "cy_batch_running", False)
                or getattr(self.state, "cy_lookup_inflight", set())
                or getattr(self.state, "cy_lookup_pending", set())
            )
        if comp_busy:
            messagebox.showinfo("Comp run active", "Wait for the active comp run to finish before recomping inventory.")
            return
        self.refresh_inventory_tab()
        records = [self._normalize_inventory_record(record) for record in getattr(self, "filtered_inventory_rows", [])]
        temp_rows: list[WorkbookRow] = []
        keys_by_excel_row: dict[int, str] = {}
        for index, record in enumerate(records, start=2):
            if str(record.get("status") or "").lower() != "active":
                continue
            if not scan_to_cert(record.get("cert_number")) or not str(record.get("grader") or "").strip():
                continue
            key = str(record.get("inventory_key") or "")
            if not key:
                continue
            if not self._inventory_recomp_record_matches_scope(record, features):
                continue
            row = self._inventory_workbook_row(record, index)
            row.status = "Queued"
            temp_rows.append(row)
            keys_by_excel_row[row.excel_row] = key
        if not temp_rows:
            scope = str(features.get("scope") or COMP_SCOPE_EMPTY)
            detail = "empty selected fields" if scope == COMP_SCOPE_EMPTY else "both a cert number and grader"
            messagebox.showinfo("No eligible inventory rows", f"No visible inventory rows have {detail}.")
            self.inventory_status_var.set(f"No visible inventory rows are eligible for recomp ({scope}).")
            return
        current_sheet = self.selected_working_sheet.get() if hasattr(self, "selected_working_sheet") else ""
        self.inventory_recomp_context = {
            "rows": list(self.state.rows),
            "row_sources": dict(getattr(self, "row_sources", {})),
            "comp_sheet_sources": dict(getattr(self, "comp_sheet_sources", {})),
            "selected_working_sheet": current_sheet,
            "comp_output_saved": bool(getattr(self, "comp_output_saved", True)),
            "keys_by_excel_row": keys_by_excel_row,
            "features": features,
            "total": len(temp_rows),
            "changed": 0,
        }
        strategy_label = str(features.get("strategy_label") or self.comp_strategy_label.get() or "Average last 5")
        self.comp_strategy_label.set(strategy_label)
        self.state.set_comp_strategy(COMP_STRATEGY_DISPLAY.get(strategy_label, COMP_STRATEGY_AVERAGE), self._comp_low_outlier_pct())
        self.state.set_rows(temp_rows)
        self.row_sources = {row.excel_row: "Inventory" for row in temp_rows}
        self.comp_sheet_sources = {}
        command_id = 0
        card_ladder_command_id = 0
        if run_card_ladder:
            card_ladder_command_id = self.state.start_all_comps(requery_all=True, allow_deferred_cy=run_cy)
            command_id = card_ladder_command_id
        if run_cy:
            command_id = self.state.start_cy_lookups(temp_rows, defer=run_card_ladder)
        self._refresh_comp_table(schedule_recommendations=False)
        if run_card_ladder:
            self.after(12000, lambda queued_command_id=card_ladder_command_id: self._warn_if_extension_not_checked_in(queued_command_id))
        pieces = []
        if run_card_ladder:
            if features.get("card_ladder_value"):
                pieces.append("Card Ladder value")
            if features.get("card_ladder_comps"):
                pieces.append("Card Ladder comps")
        if run_cy:
            pieces.append("CY")
        scope = str(features.get("scope") or COMP_SCOPE_EMPTY)
        self.inventory_status_var.set(f"Queued {' and '.join(pieces)} refresh for {len(temp_rows)} visible inventory card(s) ({scope}).")
        self.status_var.set(f"Inventory recomp queued as command #{command_id}.")

    def _inventory_recomp_record_matches_scope(self, record: dict[str, object], features: dict[str, object]) -> bool:
        if str(features.get("scope") or COMP_SCOPE_EMPTY) == COMP_SCOPE_ALL:
            return True
        if features.get("card_ladder_value") and self._money_value(record.get("card_ladder_value")) is None:
            return True
        if features.get("card_ladder_comps") and self._money_value(record.get("card_ladder_comps_average")) is None:
            return True
        if features.get("cy"):
            cy_value_missing = self._money_value(record.get("cy_value")) is None
            cy_confidence_missing = not str(record.get("cy_confidence") or "").strip()
            if cy_value_missing or cy_confidence_missing:
                return True
        return False

    def _sync_inventory_recomp_results(self) -> int:
        context = self.inventory_recomp_context
        if not context:
            return 0
        keys_by_excel_row = context.get("keys_by_excel_row")
        if not isinstance(keys_by_excel_row, dict):
            return 0
        features = context.get("features") if isinstance(context.get("features"), dict) else {}
        with self.state.lock:
            comp_rows = list(self.state.rows)
        ledger = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        ledger_indexes = {str(record.get("inventory_key") or ""): index for index, record in enumerate(ledger)}
        changed = 0
        for row in comp_rows:
            if str(row.status or "").strip().lower() in {"queued", "cy queued"}:
                continue
            key = str(keys_by_excel_row.get(row.excel_row) or "")
            if not key or key not in ledger_indexes:
                continue
            index = ledger_indexes[key]
            record = dict(ledger[index])
            if row.card_title and not is_placeholder_title(row.card_title, row.grader):
                record["card_title"] = row.card_title
            if features.get("card_ladder_value", True):
                record["card_ladder_value"] = row.card_ladder_value
            if features.get("card_ladder_comps", True):
                record["card_ladder_comps_average"] = row.card_ladder_comps_average
                record["card_ladder_comps"] = row.card_ladder_comps
            if features.get("cy", True):
                record["cy_value"] = row.cy_value
                record["cy_confidence"] = row.cy_confidence
            enriched = self._enrich_inventory_record_assignment(record, force=True)
            enriched["status"] = ledger[index].get("status") or "Active"
            enriched["inventory_key"] = key
            normalized = self._normalize_inventory_record(enriched)
            if normalized != ledger[index]:
                ledger[index] = normalized
                changed += 1
        if changed:
            self._save_inventory_ledger(ledger)
            context["changed"] = int(context.get("changed") or 0) + changed
        return changed

    def _finish_inventory_recomp(self) -> None:
        context = self.inventory_recomp_context
        if not context:
            return
        changed_total = int(context.get("changed") or 0)
        total = int(context.get("total") or 0)
        original_rows = context.get("rows")
        self.inventory_recomp_context = None
        if isinstance(original_rows, list):
            self.state.set_rows(original_rows)
        self.row_sources = dict(context.get("row_sources") or {})
        self.comp_sheet_sources = dict(context.get("comp_sheet_sources") or {})
        if hasattr(self, "selected_working_sheet"):
            self.selected_working_sheet.set(str(context.get("selected_working_sheet") or ""))
        self.comp_output_saved = bool(context.get("comp_output_saved", True))
        self._refresh_comp_table(schedule_recommendations=False)
        self.refresh_inventory_tab()
        self.inventory_status_var.set(f"Recomp finished for {total} visible inventory card(s); updated {changed_total}.")
        self.status_var.set(f"Inventory recomp finished for {total} visible card(s).")

    def delete_selected_inventory_records(self) -> None:
        if not hasattr(self, "inventory_tree"):
            return
        records = [self.inventory_tree_records.get(iid) for iid in self.inventory_tree.selection()]
        keys = {str(record.get("inventory_key") or "") for record in records if record}
        keys.discard("")
        if not keys:
            messagebox.showinfo("Choose inventory", "Select one or more inventory rows to delete.")
            return
        confirmed = messagebox.askyesno(
            "Delete from inventory?",
            f"Delete {len(keys)} selected inventory item(s)?\n\nThis only removes them from Inventory.",
        )
        if not confirmed:
            return
        deleted = self._delete_inventory_records_by_keys(keys)
        self.refresh_inventory_tab()
        self.status_var.set(f"Deleted {deleted} inventory item(s).")
        if deleted:
            self._append_activity("Inventory Delete", f"Deleted {deleted} inventory item(s).", {"deleted": deleted})

    def _delete_inventory_records_by_keys(self, keys: set[str]) -> int:
        if not keys:
            return 0
        with shared_lock(CARD_PIPELINE_DIR, "inventory-delete", self.lucas_identity):
            rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
            removed = [record for record in rows if str(record.get("inventory_key") or "") in keys]
            kept = [record for record in rows if str(record.get("inventory_key") or "") not in keys]
            deleted = len(rows) - len(kept)
            if deleted:
                self._record_inventory_deleted_tombstones(removed)
                self._save_inventory_ledger(kept)
                cleanup = getattr(self, "_delete_inventory_photo_files_for_removed_records", None)
                if callable(cleanup):
                    cleanup(removed, kept)
        return deleted

    def _filtered_inventory_records(self, rows: list[dict[str, object]]) -> list[dict[str, object]]:
        person = self.inventory_person_var.get().strip().lower() if hasattr(self, "inventory_person_var") else ""
        sport_filters = self._inventory_sport_filter_values()
        grader_filters = inventory_grader_filter_values(self.inventory_grader_var.get() if hasattr(self, "inventory_grader_var") else "")
        card_year = re.sub(r"\D", "", self.inventory_year_var.get()) if hasattr(self, "inventory_year_var") else ""
        search = self.inventory_search_var.get().strip().lower() if hasattr(self, "inventory_search_var") else ""
        min_value = self._money_value(self.inventory_min_var.get()) if hasattr(self, "inventory_min_var") else None
        max_value = self._money_value(self.inventory_max_var.get()) if hasattr(self, "inventory_max_var") else None
        min_date = self._profit_record_date(self.inventory_date_min_var.get()) if hasattr(self, "inventory_date_min_var") else None
        max_date = self._profit_record_date(self.inventory_date_max_var.get()) if hasattr(self, "inventory_date_max_var") else None
        missing_title_only = bool(self.inventory_missing_title_var.get()) if hasattr(self, "inventory_missing_title_var") else False
        missing_comps_only = bool(self.inventory_missing_comps_var.get()) if hasattr(self, "inventory_missing_comps_var") else False
        missing_cl_only = bool(self.inventory_missing_cl_var.get()) if hasattr(self, "inventory_missing_cl_var") else False
        missing_photos_only = bool(self.inventory_missing_photos_var.get()) if hasattr(self, "inventory_missing_photos_var") else False
        filtered: list[dict[str, object]] = []
        for record in rows:
            if str(record.get("status") or "").lower() != "active":
                continue
            if person and person not in str(record.get("assigned_person") or "Unassigned").lower():
                continue
            if sport_filters:
                record_sport_text = str(record.get("sport") or "").strip().lower()
                record_sport = (assignment_engine.canonical_sport_label(record_sport_text) or record_sport_text).strip().lower()
                if record_sport not in sport_filters and not any(sport in record_sport_text for sport in sport_filters):
                    continue
            if grader_filters:
                record_grader = str(record.get("grader") or "").strip().upper()
                if record_grader not in grader_filters:
                    continue
            if card_year and card_year not in self._inventory_record_card_years(record):
                continue
            if search:
                searchable = f"{record.get('item_id') or ''} {record.get('cert_number') or ''} {record.get('card_title') or ''}".lower()
                if any(part not in searchable for part in search.split()):
                    continue
            value = self._money_value(record.get("inventory_value") or record.get("purchase_price")) or 0.0
            if min_value is not None and value < min_value:
                continue
            if max_value is not None and value > max_value:
                continue
            if min_date is not None or max_date is not None:
                record_date = self._profit_record_date(record.get("date_added"))
                if min_date is not None and (record_date is None or record_date < min_date):
                    continue
                if max_date is not None and (record_date is None or record_date > max_date):
                    continue
            if missing_title_only and not self._inventory_record_missing_card_description(record):
                continue
            if missing_comps_only or missing_cl_only:
                missing_value_match = False
                if missing_comps_only and self._inventory_record_missing_comps(record):
                    missing_value_match = True
                if missing_cl_only and self._inventory_record_missing_cl_value(record):
                    missing_value_match = True
                if not missing_value_match:
                    continue
            if missing_photos_only and self._inventory_photo_paths_for_record(record):
                continue
            filtered.append(record)
        return filtered

    def _inventory_record_missing_card_description(self, record: dict[str, object]) -> bool:
        title = str(record.get("card_title") or "").strip()
        if not title:
            return True
        cert = scan_to_cert(record.get("cert_number"))
        return bool(cert and scan_to_cert(title) == cert)

    def _inventory_record_missing_comps(self, record: dict[str, object]) -> bool:
        raw_value = record.get("card_ladder_comps_average")
        if raw_value is None or str(raw_value).strip() == "":
            return True
        return self._money_value(raw_value) is None

    def _inventory_record_missing_cl_value(self, record: dict[str, object]) -> bool:
        raw_value = record.get("card_ladder_value")
        if raw_value is None or str(raw_value).strip() == "":
            return True
        return self._money_value(raw_value) is None

    def _inventory_sport_filter_values(self) -> set[str]:
        if not hasattr(self, "inventory_sport_var"):
            return set()
        raw = str(self.inventory_sport_var.get() or "")
        values: set[str] = set()
        for part in re.split(r"[,;/|]", raw):
            text = part.strip().lower()
            if not text:
                continue
            values.add((assignment_engine.canonical_sport_label(text) or text).strip().lower())
        return values

    def _inventory_record_card_years(self, record: dict[str, object]) -> set[str]:
        values = (
            record.get("card_title"),
            record.get("title"),
            record.get("description"),
        )
        text = " ".join(str(value or "") for value in values)
        return set(re.findall(r"\b(?:19|20)\d{2}\b", text))

    def export_inventory(self) -> None:
        rows = self.filtered_inventory_rows if hasattr(self, "filtered_inventory_rows") else []
        if not rows:
            messagebox.showinfo("No inventory", "No inventory rows match the current filters.")
            return
        path = filedialog.asksaveasfilename(
            title="Export inventory",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=f"inventory-{datetime.now():%Y%m%d-%H%M%S}.xlsx",
        )
        if not path:
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventory"
        headers = ["Date Added", "Type", "Item ID", "Person", "Sport", "Certification Number", "Grader", "Card Description", "Purchase Price", "Paid With", "Card Ladder", "Comps", "CY Estimate", "CY Confidence", "Best Company", "Estimated Payout", "Source Sheet", "Source", "Status", "Photos", "Photo Paths", "Notes"]
        sheet.append(headers)
        for record in rows:
            sheet.append([
                record.get("date_added") or "",
                record.get("item_type") or "",
                record.get("item_id") or "",
                record.get("assigned_person") or "",
                record.get("sport") or "",
                record.get("cert_number") or "",
                record.get("grader") or "",
                record.get("card_title") or "",
                record.get("purchase_price"),
                record.get("paid_with") or "",
                record.get("card_ladder_value"),
                record.get("card_ladder_comps_average"),
                record.get("cy_value"),
                record.get("cy_confidence"),
                record.get("best_company") or "",
                record.get("estimated_payout"),
                record.get("source_sheet") or "",
                record.get("source") or "",
                record.get("status") or "",
                len(record.get("photo_paths") or []),
                "; ".join(str(path) for path in (record.get("photo_paths") or [])),
                inventory_display_notes(record),
            ])
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = "A2"
        for index, width in enumerate([14, 12, 22, 18, 14, 22, 12, 60, 16, 18, 16, 16, 16, 14, 20, 16, 28, 24, 14, 10, 45, 36], start=1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
        workbook.save(path)
        self.status_var.set(f"Exported inventory: {path}")

    def _quarter_label_for_date(self, value: object) -> str:
        record_date = self._profit_record_date(value)
        if record_date is None:
            return ""
        return f"Q{((record_date.month - 1) // 3) + 1}"

    def _empty_year_end_bucket(self) -> dict[str, object]:
        return {
            "sales": 0.0,
            "purchase": 0.0,
            "gross_profit": 0.0,
            "expenses": 0.0,
            "net_profit": 0.0,
            "inventory_value": 0.0,
            "inventory_cost": 0.0,
            "sold_count": 0,
            "active_inventory_count": 0,
            "expense_by_type": defaultdict(float),
            "quarters": {
                quarter: {
                    "sales": 0.0,
                    "purchase": 0.0,
                    "gross_profit": 0.0,
                    "expenses": 0.0,
                    "net_profit": 0.0,
                    "sold_count": 0,
                    "expense_by_type": defaultdict(float),
                }
                for quarter in ("Q1", "Q2", "Q3", "Q4")
            },
        }

    def _year_end_report_data(self, year: int) -> dict[str, object]:
        overall = self._empty_year_end_bucket()
        people: defaultdict[str, dict[str, object]] = defaultdict(self._empty_year_end_bucket)
        sales_detail: list[dict[str, object]] = []
        expense_detail: list[dict[str, object]] = []
        inventory_detail: list[dict[str, object]] = []
        expense_types: set[str] = set(EXPENSE_CATEGORY_OPTIONS)
        ledger = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
        ledger = [record for record in ledger if not self._is_manual_company_profit_backfill(record)]
        ledger, _removed = CardPipelineApp._dedupe_profit_records(self, ledger)
        for record in self._enrich_profit_records_with_people(ledger):
            record_date = self._profit_record_date(record.get("date_added"))
            if record_date is None or record_date.year != year:
                continue
            person = str(record.get("assigned_person") or "Unassigned").strip() or "Unassigned"
            quarter = self._quarter_label_for_date(record.get("date_added"))
            person_bucket = people[person]
            is_expense = str(record.get("record_type") or "").strip().lower() == "expense"
            if is_expense:
                expense_type = str(record.get("expense_type") or "Fees").strip() or "Fees"
                expense_types.add(expense_type)
                amount = abs(self._money_value(record.get("expense_amount")) or self._money_value(record.get("profit")) or 0.0)
                for bucket in (overall, person_bucket):
                    bucket["expenses"] += amount
                    bucket["net_profit"] -= amount
                    bucket["expense_by_type"][expense_type] += amount
                    if quarter:
                        bucket["quarters"][quarter]["expenses"] += amount
                        bucket["quarters"][quarter]["net_profit"] -= amount
                        bucket["quarters"][quarter]["expense_by_type"][expense_type] += amount
                expense_detail.append(
                    {
                        "date": record_date.isoformat(),
                        "quarter": quarter,
                        "person": person,
                        "expense_type": expense_type,
                        "amount": amount,
                        "related": self._expense_related_label(record),
                        "notes": record.get("notes") or "",
                    }
                )
                continue
            sale = self._money_value(record.get("sale_price")) or 0.0
            purchase = self._money_value(record.get("purchase_price")) or 0.0
            profit = self._money_value(record.get("profit"))
            if profit is None:
                profit = sale - purchase
            for bucket in (overall, person_bucket):
                bucket["sales"] += sale
                bucket["purchase"] += purchase
                bucket["gross_profit"] += profit
                bucket["net_profit"] += profit
                bucket["sold_count"] += 1
                if quarter:
                    bucket["quarters"][quarter]["sales"] += sale
                    bucket["quarters"][quarter]["purchase"] += purchase
                    bucket["quarters"][quarter]["gross_profit"] += profit
                    bucket["quarters"][quarter]["net_profit"] += profit
                    bucket["quarters"][quarter]["sold_count"] += 1
            sales_detail.append(
                {
                    "date": record_date.isoformat(),
                    "quarter": quarter,
                    "person": person,
                    "company": record.get("company") or "",
                    "card": record.get("card_title") or "",
                    "cert": record.get("cert_number") or record.get("item_id") or "",
                    "purchase": purchase,
                    "sale": sale,
                    "profit": profit,
                    "source_sheet": record.get("source_sheet") or "",
                }
            )
        inactive_statuses = {"sold", "deleted", "removed", "archived", "company sheet"}
        for record in [self._normalize_inventory_record(row) for row in self._load_inventory_ledger()]:
            status = str(record.get("status") or "Active").strip() or "Active"
            if status.lower() in inactive_statuses:
                continue
            person = str(record.get("assigned_person") or "Unassigned").strip() or "Unassigned"
            person_bucket = people[person]
            value = self._money_value(record.get("inventory_value") or record.get("estimated_payout") or record.get("purchase_price")) or 0.0
            cost = self._money_value(record.get("purchase_price")) or 0.0
            for bucket in (overall, person_bucket):
                bucket["inventory_value"] += value
                bucket["inventory_cost"] += cost
                bucket["active_inventory_count"] += 1
            inventory_detail.append(
                {
                    "person": person,
                    "type": record.get("item_type") or "",
                    "cert": record.get("cert_number") or record.get("item_id") or "",
                    "grader": record.get("grader") or "",
                    "card": record.get("card_title") or "",
                    "purchase": cost,
                    "value": value,
                    "best_company": record.get("best_company") or "",
                    "status": status,
                    "source_sheet": record.get("source_sheet") or "",
                }
            )
        return {
            "year": year,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "overall": overall,
            "people": dict(sorted(people.items(), key=lambda item: item[0].lower())),
            "expense_types": sorted(expense_types, key=str.lower),
            "sales_detail": sales_detail,
            "expense_detail": expense_detail,
            "inventory_detail": sorted(inventory_detail, key=lambda item: (str(item.get("person") or "").lower(), str(item.get("card") or "").lower())),
        }

    def _style_year_end_sheet(self, sheet, currency_columns: set[int] | None = None, integer_columns: set[int] | None = None) -> None:
        header_fill = PatternFill("solid", fgColor="1ED760")
        header_font = Font(bold=True, color="000000")
        title_font = Font(bold=True, size=14)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.font = title_font
        if sheet.max_row >= 2:
            for cell in sheet[2]:
                cell.fill = header_fill
                cell.font = header_font
        for column in currency_columns or set():
            for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=3, max_row=sheet.max_row):
                for item in cell:
                    item.number_format = '"$"#,##0.00'
        for column in integer_columns or set():
            for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=3, max_row=sheet.max_row):
                for item in cell:
                    item.number_format = '#,##0'
        for column_index in range(1, sheet.max_column + 1):
            max_length = 10
            for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=1, max_row=sheet.max_row):
                for item in cell:
                    max_length = max(max_length, min(len(str(item.value or "")), 55))
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 2, 60)
        sheet.freeze_panes = "A3"
        if sheet.max_row >= 2 and sheet.max_column >= 1:
            sheet.auto_filter.ref = f"A2:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    def _write_year_end_sheet(self, workbook: Workbook, title: str, headers: list[str], rows: list[list[object]], currency_columns: set[int] | None = None, integer_columns: set[int] | None = None):
        sheet = workbook.create_sheet(title)
        sheet.append([title])
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        self._style_year_end_sheet(sheet, currency_columns=currency_columns, integer_columns=integer_columns)
        return sheet

    def _build_year_end_report_workbook(self, data: dict[str, object]) -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)
        year = int(data.get("year") or datetime.now().year)
        overall = data["overall"]
        people: dict[str, dict[str, object]] = data["people"]
        summary_rows = [
            ["Generated At", data.get("generated_at") or ""],
            ["Report Year", year],
            ["Total Sales", overall["sales"]],
            ["Purchase Cost Sold", overall["purchase"]],
            ["Gross Profit", overall["gross_profit"]],
            ["Expenses", overall["expenses"]],
            ["Net Profit", overall["net_profit"]],
            ["Remaining Inventory Value", overall["inventory_value"]],
            ["Remaining Inventory Cost", overall["inventory_cost"]],
            ["Sold Cards", overall["sold_count"]],
            ["Active Inventory Items", overall["active_inventory_count"]],
        ]
        summary_sheet = self._write_year_end_sheet(workbook, f"{year} Summary", ["Metric", "Overall"], summary_rows)
        for row_index in range(5, 12):
            summary_sheet.cell(row_index, 2).number_format = '"$"#,##0.00'
        for row_index in range(12, 14):
            summary_sheet.cell(row_index, 2).number_format = '#,##0'
        quarterly_rows: list[list[object]] = []
        for scope, bucket in [("Overall", overall), *people.items()]:
            for quarter in ("Q1", "Q2", "Q3", "Q4"):
                q = bucket["quarters"][quarter]
                quarterly_rows.append([scope, quarter, q["sales"], q["purchase"], q["gross_profit"], q["expenses"], q["net_profit"], q["sold_count"]])
        self._write_year_end_sheet(
            workbook,
            "Quarterly",
            ["Scope", "Quarter", "Sales", "Purchase Cost Sold", "Gross Profit", "Expenses", "Net Profit", "Sold Cards"],
            quarterly_rows,
            currency_columns={3, 4, 5, 6, 7},
            integer_columns={8},
        )
        expense_rows: list[list[object]] = []
        for scope, bucket in [("Overall", overall), *people.items()]:
            for expense_type in data["expense_types"]:
                quarters = bucket["quarters"]
                values = [quarters[quarter]["expense_by_type"].get(expense_type, 0.0) for quarter in ("Q1", "Q2", "Q3", "Q4")]
                total = bucket["expense_by_type"].get(expense_type, 0.0)
                if total or scope == "Overall":
                    expense_rows.append([scope, expense_type, total, *values])
        self._write_year_end_sheet(
            workbook,
            "Expenses by Type",
            ["Scope", "Expense Type", "Year Total", "Q1", "Q2", "Q3", "Q4"],
            expense_rows,
            currency_columns={3, 4, 5, 6, 7},
        )
        person_rows = [
            [
                person,
                bucket["sales"],
                bucket["purchase"],
                bucket["gross_profit"],
                bucket["expenses"],
                bucket["net_profit"],
                bucket["inventory_value"],
                bucket["inventory_cost"],
                bucket["sold_count"],
                bucket["active_inventory_count"],
            ]
            for person, bucket in people.items()
        ]
        self._write_year_end_sheet(
            workbook,
            "Per Person",
            ["Person", "Sales", "Purchase Cost Sold", "Gross Profit", "Expenses", "Net Profit", "Remaining Inventory Value", "Remaining Inventory Cost", "Sold Cards", "Active Inventory"],
            person_rows,
            currency_columns={2, 3, 4, 5, 6, 7, 8},
            integer_columns={9, 10},
        )
        inventory_rows = [
            [row["person"], row["type"], row["cert"], row["grader"], row["card"], row["purchase"], row["value"], row["best_company"], row["status"], row["source_sheet"]]
            for row in data["inventory_detail"]
        ]
        self._write_year_end_sheet(
            workbook,
            "Inventory Left",
            ["Person", "Type", "Cert / Item ID", "Grader", "Card", "Purchase Cost", "Inventory Value", "Best Company", "Status", "Source Sheet"],
            inventory_rows,
            currency_columns={6, 7},
        )
        sales_rows = [
            [row["date"], row["quarter"], row["person"], row["company"], row["card"], row["cert"], row["purchase"], row["sale"], row["profit"], row["source_sheet"]]
            for row in data["sales_detail"]
        ]
        self._write_year_end_sheet(
            workbook,
            "Sales Detail",
            ["Date", "Quarter", "Person", "Company", "Card", "Cert / Item ID", "Purchase", "Sale", "Profit", "Source Sheet"],
            sales_rows,
            currency_columns={7, 8, 9},
        )
        expense_detail_rows = [
            [row["date"], row["quarter"], row["person"], row["expense_type"], row["amount"], row["related"], row["notes"]]
            for row in data["expense_detail"]
        ]
        self._write_year_end_sheet(
            workbook,
            "Expense Detail",
            ["Date", "Quarter", "Person", "Expense Type", "Amount", "Related", "Notes"],
            expense_detail_rows,
            currency_columns={5},
        )
        return workbook

    def export_year_end_report(self) -> None:
        current_year = datetime.now().year
        year = simpledialog.askinteger(
            "Export Year-End Report",
            "Report year:",
            initialvalue=current_year,
            minvalue=2000,
            maxvalue=current_year + 1,
            parent=self,
        )
        if not year:
            return
        path = filedialog.asksaveasfilename(
            title="Export year-end report",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=f"lucas-year-end-report-{year}-{datetime.now():%Y%m%d-%H%M%S}.xlsx",
        )
        if not path:
            return
        try:
            data = self._year_end_report_data(int(year))
            workbook = self._build_year_end_report_workbook(data)
            workbook.save(path)
        except Exception as error:
            messagebox.showerror("Export failed", str(error))
            return
        self.status_var.set(f"Exported year-end report: {path}")
        messagebox.showinfo("Export complete", f"Year-end report exported:\n\n{path}")

    def _profit_record_key(self, record: dict[str, object]) -> str:
        record_type = str(record.get("record_type") or "").strip().lower()
        if record_type == "expense":
            return "|".join(
                str(record.get(field) or "").strip().lower()
                for field in ("record_type", "expense_id", "assigned_person", "date_added", "expense_type", "expense_amount", "related_type", "source_sheet", "item_id", "cert_number", "notes")
            )
        item_id = str(record.get("item_id") or "").strip()
        if item_id:
            return "|".join(
                str(record.get(field) or "").strip().lower()
                for field in ("item_id", "company", "date_added", "weekly_sheet_name", "source_sheet")
            )
        return "|".join(
            str(record.get(field) or "").strip().lower()
            for field in ("cert_number", "company", "date_added", "weekly_sheet_name", "source_sheet")
        )

    def _profit_record_identity_keys(self, record: dict[str, object]) -> set[str]:
        normalized = dict(record)
        primary = str(normalized.get("ledger_key") or self._profit_record_key(normalized) or "").strip().lower()
        keys = {primary} if primary else set()
        if str(normalized.get("record_type") or "").strip().lower() == "expense":
            return keys
        company = str(normalized.get("company") or normalized.get("best_company") or "").strip().lower()
        source_sheet = Path(str(normalized.get("source_sheet") or "")).name.strip().lower()
        cert = scan_to_cert(normalized.get("cert_number"))
        item_id = str(normalized.get("item_id") or "").strip().lower()
        stable_id = cert or item_id
        if company and source_sheet and stable_id:
            keys.add(f"sold-card|{company}|{source_sheet}|{stable_id}")
        return keys

    def _dedupe_profit_records(self, rows: list[dict[str, object]]) -> tuple[list[dict[str, object]], int]:
        kept: list[dict[str, object]] = []
        seen: set[str] = set()
        removed = 0
        for record in rows:
            normalized = self._normalize_profit_record(record)
            keys = CardPipelineApp._profit_record_identity_keys(self, normalized)
            if keys and keys & seen:
                removed += 1
                continue
            kept.append(normalized)
            seen.update(keys)
        return kept, removed

    def _is_manual_company_profit_backfill(self, record: dict[str, object]) -> bool:
        if str(record.get("record_type") or "").strip().lower() == "expense":
            return False
        if str(record.get("source_sheet") or "").strip():
            return False
        weekly = str(record.get("weekly_sheet_name") or "").strip().lower()
        return ".xlsx" in weekly

    def _money_value(self, value: object) -> float | None:
        if value is None or value == "":
            return None
        match = re.search(r"-?[\d,.]+\s*[kK]?", str(value))
        if not match:
            return None
        text = match.group(0).strip().strip(".,").replace(",", "")
        multiplier = 1
        if text.lower().endswith("k"):
            multiplier = 1000
            text = text[:-1].strip().strip(".,")
        if re.fullmatch(r"-?\d{1,3}\.\d{3}", text):
            text = text.replace(".", "")
        try:
            return float(text) * multiplier
        except ValueError:
            return None

    def _normalize_profit_record(self, record: dict[str, object]) -> dict[str, object]:
        normalized = dict(record)
        record_type = str(normalized.get("record_type") or "").strip().lower()
        if record_type == "expense":
            amount = self._money_value(normalized.get("expense_amount") or normalized.get("amount") or normalized.get("purchase_price")) or 0.0
            expense_type = str(normalized.get("expense_type") or normalized.get("category") or "Fees").strip() or "Fees"
            if expense_type not in EXPENSE_CATEGORY_OPTIONS:
                expense_type = "Fees"
            notes = str(normalized.get("notes") or "").strip()
            related_sheet = str(normalized.get("source_sheet") or normalized.get("related_sheet") or "").strip()
            related_type = str(normalized.get("related_type") or normalized.get("tie_to") or "").strip()
            if not related_type:
                related_type = "Sheet" if related_sheet and Path(related_sheet).name.lower() != "expenses" else "General"
            if related_type not in EXPENSE_LINK_OPTIONS:
                related_type = "General"
            related_item_id = str(normalized.get("item_id") or normalized.get("related_item_id") or "").strip()
            related_cert = str(normalized.get("cert_number") or normalized.get("related_cert") or "").strip()
            normalized["record_type"] = "expense"
            normalized["expense_id"] = str(normalized.get("expense_id") or "").strip()
            normalized["expense_type"] = expense_type
            normalized["expense_amount"] = round(abs(amount), 2)
            normalized["related_type"] = related_type
            normalized["purchase_price"] = None
            normalized["sale_price"] = None
            normalized["profit"] = -round(abs(amount), 2)
            normalized["date_added"] = self._profit_local_calendar_date(
                normalized.get("date_added") or datetime.now().strftime("%Y-%m-%d"),
                normalized.get("ledger_added_at"),
            )
            normalized["company"] = f"Expense: {expense_type}"
            normalized["card_title"] = notes or expense_type
            normalized["item_id"] = related_item_id if related_type == "Card" else ""
            normalized["cert_number"] = related_cert if related_type == "Card" else ""
            normalized["weekly_sheet_name"] = ""
            normalized["source_sheet"] = related_sheet if related_type in {"Card", "Sheet"} and related_sheet else "Expenses"
            owner_for_profile = getattr(self, "_owner_for_profile", lambda person="": str(person or "").strip() or "Unassigned")
            normalized["assigned_person"] = owner_for_profile(normalized.get("assigned_person") or normalized.get("person"))
            normalized["notes"] = notes
            normalized["ledger_added_at"] = str(normalized.get("ledger_added_at") or "").strip()
            normalized["ledger_key"] = self._profit_record_key(normalized)
            return normalized
        purchase = self._money_value(normalized.get("purchase_price"))
        sale = self._money_value(normalized.get("sale_price"))
        normalized["purchase_price"] = purchase
        normalized["sale_price"] = sale
        normalized["profit"] = round(sale - purchase, 2) if sale is not None and purchase is not None else None
        normalized["date_added"] = self._profit_local_calendar_date(
            normalized.get("date_added") or datetime.now().strftime("%Y-%m-%d"),
            normalized.get("ledger_added_at"),
        )
        normalized["company"] = str(normalized.get("company") or normalized.get("best_company") or "").strip()
        normalized["item_type"] = str(normalized.get("item_type") or normalized.get("type") or "").strip()
        normalized["item_id"] = str(normalized.get("item_id") or "").strip()
        normalized["card_title"] = str(normalized.get("card_title") or "").strip()
        normalized["cert_number"] = str(normalized.get("cert_number") or "").strip()
        normalized["weekly_sheet_name"] = str(normalized.get("weekly_sheet_name") or "").strip()
        normalized["source_sheet"] = str(normalized.get("source_sheet") or "").strip()
        normalized["original_source_sheet"] = str(normalized.get("original_source_sheet") or "").strip()
        owner_for_profile = getattr(self, "_owner_for_profile", lambda person="": str(person or "").strip() or "Unassigned")
        assigned_person = owner_for_profile(normalized.get("assigned_person") or normalized.get("person"))
        if getattr(self, "_is_personal_lucas", lambda: False)():
            if str(normalized.get("source_sheet") or "").strip().lower() == "unassigned general sold":
                normalized["source_sheet"] = self._general_sold_sheet_name(assigned_person)
        normalized["assigned_person"] = assigned_person
        normalized["ledger_added_at"] = str(normalized.get("ledger_added_at") or "").strip()
        photo_paths = normalized.get("photo_paths") or normalized.get("photos") or []
        if isinstance(photo_paths, str):
            photo_paths = [part.strip() for part in re.split(r"[;\n]", photo_paths) if part.strip()]
        elif not isinstance(photo_paths, list):
            photo_paths = []
        normalized["photo_paths"] = [str(path).strip() for path in photo_paths if str(path or "").strip()]
        normalized["photo_count"] = len(normalized["photo_paths"])
        normalized["ledger_key"] = self._profit_record_key(normalized)
        return normalized

    def _person_for_profit_record(self, record: dict[str, object]) -> str:
        if getattr(self, "_is_personal_lucas", lambda: False)():
            return self._personal_default_person()
        existing = str(record.get("assigned_person") or "").strip()
        if existing and existing.lower() != "unassigned":
            return existing
        source_sheet = Path(str(record.get("source_sheet") or "")).name
        if not source_sheet:
            return existing
        for stage in ("Incoming", "Received", "Working"):
            marker = self.home_sheet_markers.get(self._home_sheet_key(stage, source_sheet), {})
            person = str(marker.get("assigned_person") or "").strip()
            if person:
                return person
        for key, marker in self.home_sheet_markers.items():
            _stage, name = self._split_home_sheet_key(key)
            if Path(name).name == source_sheet:
                person = str(marker.get("assigned_person") or "").strip()
                if person:
                    return person
        return existing

    def _enrich_profit_records_with_people(self, rows: list[dict[str, object]]) -> list[dict[str, object]]:
        enriched: list[dict[str, object]] = []
        for record in rows:
            normalized = self._normalize_profit_record(record)
            normalized["assigned_person"] = self._person_for_profit_record(normalized)
            enriched.append(normalized)
        return enriched

    def _filtered_profit_records(self, rows: list[dict[str, object]]) -> list[dict[str, object]]:
        needle = self.profit_person_var.get().strip().lower() if hasattr(self, "profit_person_var") else ""
        search = self.profit_search_var.get().strip().lower() if hasattr(self, "profit_search_var") else ""
        period = self.profit_period_var.get().strip() if hasattr(self, "profit_period_var") else "Total"
        period_start, period_end = self._profit_period_bounds(period)
        filtered: list[dict[str, object]] = []
        for record in rows:
            if needle and needle not in (str(record.get("assigned_person") or "Unassigned").lower()):
                continue
            if search:
                searchable = " ".join(
                    str(record.get(field) or "")
                    for field in ("item_id", "cert_number", "card_title", "company", "source_sheet", "weekly_sheet_name", "assigned_person")
                ).lower()
                if any(part not in searchable for part in search.split()):
                    continue
            if period_start is not None:
                sold_date = self._profit_record_date(record.get("date_added"))
                if sold_date is None or sold_date < period_start or sold_date > period_end:
                    continue
            filtered.append(record)
        return filtered

    def _profit_record_date(self, value: object):
        text = str(value or "").strip()
        if not text:
            return None
        try:
            return datetime.strptime(text[:10], "%Y-%m-%d").date()
        except ValueError:
            return None

    def _profit_today(self):
        return datetime.now().date()

    def _profit_period_bounds(self, period: str, as_of=None):
        today = as_of or self._profit_today()
        label = (period or "Total").strip().lower()
        if label in {"5 days", "5 day", "five days"}:
            return today - timedelta(days=4), today
        if label == "week":
            return today - timedelta(days=6), today
        if label in {"last 30 days", "last 30 day", "30 days", "rolling 30 days", "rolling month"}:
            return today - timedelta(days=29), today
        if label in {"calendar month", "month", "this month", "mtd"}:
            return today.replace(day=1), today
        if label in {"year", "ytd", "year to date"}:
            return today.replace(month=1, day=1), today
        return None, today

    def _canonical_profit_period(self, period: str) -> str:
        label = str(period or "").strip()
        normalized = label.lower()
        if normalized in {"month", "this month", "mtd"}:
            return "Calendar Month"
        if normalized in {"30 days", "last 30 day", "rolling 30 days", "rolling month"}:
            return "Last 30 Days"
        return label

    def _profit_period_label(self) -> str:
        period = self._canonical_profit_period(self.profit_period_var.get().strip() if hasattr(self, "profit_period_var") else DEFAULT_PROFIT_PERIOD)
        return period if period in PROFIT_PERIOD_OPTIONS else DEFAULT_PROFIT_PERIOD

    def _profit_graph_label(self) -> str:
        graph = self.profit_graph_var.get().strip() if hasattr(self, "profit_graph_var") else DEFAULT_PROFIT_GRAPH
        return graph if graph in PROFIT_GRAPH_OPTIONS else DEFAULT_PROFIT_GRAPH

    def _profit_plot_label(self) -> str:
        plot = self.profit_plot_var.get().strip() if hasattr(self, "profit_plot_var") else DEFAULT_PROFIT_PLOT
        return plot if plot in PROFIT_PLOT_OPTIONS else DEFAULT_PROFIT_PLOT

    def _profit_chart_title(self) -> str:
        plot_label = self._profit_plot_label() if hasattr(self, "_profit_plot_label") else DEFAULT_PROFIT_PLOT
        plot_suffix = " by Sport" if plot_label == "By Sport" and self._profit_graph_label() != "Profit by Company" else ""
        return f"{self._profit_graph_label()}{plot_suffix} ({self._profit_period_label()})"

    def _profit_sport_label(self, record: dict[str, object]) -> str:
        def display_sport(value: str) -> str:
            return value.upper() if value.lower() == "tcg" else value.title()
        cached = str(record.get("_profit_sport_label") or "").strip()
        if cached:
            return cached
        explicit = str(record.get("sport") or record.get("category") or "").strip()
        if explicit:
            label = display_sport(assignment_engine.canonical_sport_label(explicit) or explicit)
            record["_profit_sport_label"] = label
            return label
        title = str(record.get("card_title") or "").strip()
        if title:
            cache_key = title.lower()
            cache = self.profit_sport_cache if hasattr(self, "profit_sport_cache") else {}
            cached_title_sport = str(cache.get(cache_key) or "").strip()
            if cached_title_sport:
                record["_profit_sport_label"] = cached_title_sport
                return cached_title_sport
            parsed = assignment_engine.parse_card_for_matching(title)
            sport = assignment_engine.canonical_sport_label(parsed.get("sport") or "")
            if sport:
                label = display_sport(sport)
                cache[cache_key] = label
                record["_profit_sport_label"] = label
                return label
            cache[cache_key] = "Other"
        record["_profit_sport_label"] = "Other"
        return "Other"

    def _profit_chart_bucket_label(self, sold_date) -> str:
        period = self._profit_period_label()
        if period in {"Year", "YTD", "Total"}:
            return sold_date.strftime("%Y-%m")
        return sold_date.isoformat()

    def _profit_chart_bucket_display(self, bucket: str) -> str:
        if len(bucket) == 7 and bucket[4] == "-":
            try:
                return datetime.strptime(bucket, "%Y-%m").strftime("%b")
            except ValueError:
                return bucket
        return bucket[5:] if len(bucket) >= 10 else bucket

    def _profit_chart_bucket_range(self, rows: list[dict[str, object]], monthly: bool) -> list[str]:
        period = self._profit_period_label()
        period_start, period_end = self._profit_period_bounds(period)
        if monthly and period in {"Year", "YTD"}:
            year = (period_end or self._profit_today()).year
            month_end = 12 if period == "Year" else (period_end or self._profit_today()).month
            return [f"{year}-{month:02d}" for month in range(1, month_end + 1)]
        if monthly:
            buckets = {
                self._profit_chart_bucket_label(sold_date)
                for record in rows
                for sold_date in [self._profit_record_date(record.get("date_added"))]
                if sold_date is not None
            }
            return sorted(buckets)
        if period_start is not None and period_end is not None:
            labels: list[str] = []
            cursor = period_start
            while cursor <= period_end:
                labels.append(cursor.isoformat())
                cursor += timedelta(days=1)
            return labels
        buckets = {
            self._profit_chart_bucket_label(sold_date)
            for record in rows
            for sold_date in [self._profit_record_date(record.get("date_added"))]
            if sold_date is not None
        }
        return sorted(buckets)

    def _profit_chart_series(self, rows: list[dict[str, object]]) -> tuple[list[str], list[float]]:
        daily: dict[str, float] = {}
        sales: dict[str, float] = {}
        ratio_mode = self._profit_graph_label() == "Profit to Sales Ratio"
        for record in rows:
            profit = self._money_value(record.get("profit"))
            sold_date = self._profit_record_date(record.get("date_added"))
            if profit is None or sold_date is None:
                continue
            day = sold_date.isoformat()
            daily[day] = daily.get(day, 0.0) + float(profit)
            sale = self._money_value(record.get("sale_price"))
            if sale is not None:
                sales[day] = sales.get(day, 0.0) + float(sale)
        period_start, period_end = self._profit_period_bounds(self._profit_period_label())
        if period_start is not None:
            cursor = period_start
            while cursor <= period_end:
                daily.setdefault(cursor.isoformat(), 0.0)
                sales.setdefault(cursor.isoformat(), 0.0)
                cursor += timedelta(days=1)
        days = sorted(daily)
        if ratio_mode:
            return days, [daily[day] / sales[day] if sales.get(day) else 0.0 for day in days]
        daily_values = [daily[day] for day in days]
        if self._profit_graph_label() != "Overall Profit":
            return days, daily_values
        running = 0.0
        cumulative_values: list[float] = []
        for value in daily_values:
            running += value
            cumulative_values.append(running)
        return days, cumulative_values

    def _profit_company_label(self, record: dict[str, object]) -> str:
        label = str(record.get("company") or record.get("buyer") or "").strip()
        return label or "General Sold"

    def _profit_company_chart_series(self, rows: list[dict[str, object]]) -> tuple[list[str], list[dict[str, object]]]:
        monthly = self._profit_period_label() in {"Year", "YTD", "Total"}
        buckets = self._profit_chart_bucket_range(rows, monthly)
        profit_by_company: dict[str, dict[str, float]] = {}
        for record in rows:
            if str(record.get("record_type") or "").strip().lower() == "expense":
                continue
            profit = self._money_value(record.get("profit"))
            sold_date = self._profit_record_date(record.get("date_added"))
            if profit is None or sold_date is None:
                continue
            company = self._profit_company_label(record)
            bucket = self._profit_chart_bucket_label(sold_date)
            profit_by_company.setdefault(company, {})[bucket] = profit_by_company.setdefault(company, {}).get(bucket, 0.0) + float(profit)
        if not buckets:
            buckets = sorted({bucket for company_values in profit_by_company.values() for bucket in company_values})
        company_totals = {
            company: sum(abs(value) for value in company_values.values())
            for company, company_values in profit_by_company.items()
        }
        companies = [company for company, _total in sorted(company_totals.items(), key=lambda item: item[1], reverse=True)][:6]
        colors = ["#22c55e", "#38bdf8", "#eab308", "#fb7185", "#c084fc", "#f97316"]
        lines: list[dict[str, object]] = []
        for index, company in enumerate(companies):
            values = [profit_by_company.get(company, {}).get(bucket, 0.0) for bucket in buckets]
            lines.append({"label": company, "values": values, "color": colors[index % len(colors)]})
        return buckets, lines

    def _profit_chart_lines(self, rows: list[dict[str, object]]) -> tuple[list[str], list[dict[str, object]], bool]:
        if self._profit_graph_label() == "Profit by Company":
            labels, lines = self._profit_company_chart_series(rows)
            return labels, lines, False
        if self._profit_plot_label() != "By Sport":
            labels, values = self._profit_chart_series(rows)
            return labels, [{"label": self._profit_graph_label(), "values": values, "color": "#22c55e"}], self._profit_graph_label() == "Profit to Sales Ratio"
        monthly = self._profit_period_label() in {"Year", "YTD", "Total"}
        buckets = self._profit_chart_bucket_range(rows, monthly)
        profit_by_sport: dict[str, dict[str, float]] = {}
        sales_by_sport: dict[str, dict[str, float]] = {}
        for record in rows:
            if str(record.get("record_type") or "").strip().lower() == "expense":
                continue
            profit = self._money_value(record.get("profit"))
            sold_date = self._profit_record_date(record.get("date_added"))
            if profit is None or sold_date is None:
                continue
            sport = self._profit_sport_label(record)
            bucket = self._profit_chart_bucket_label(sold_date)
            profit_by_sport.setdefault(sport, {})[bucket] = profit_by_sport.setdefault(sport, {}).get(bucket, 0.0) + float(profit)
            sale = self._money_value(record.get("sale_price"))
            if sale is not None:
                sales_by_sport.setdefault(sport, {})[bucket] = sales_by_sport.setdefault(sport, {}).get(bucket, 0.0) + float(sale)
        if not buckets:
            buckets = sorted({bucket for sport_values in profit_by_sport.values() for bucket in sport_values})
        sport_totals = {
            sport: sum(abs(value) for value in profit_by_sport.get(sport, {}).values()) + sum(abs(value) for value in sales_by_sport.get(sport, {}).values()) * 0.01
            for sport in set(profit_by_sport) | set(sales_by_sport)
        }
        preferred = [sport for sport in ("Football", "Baseball", "Basketball") if sport in sport_totals]
        remaining = [sport for sport, _total in sorted(sport_totals.items(), key=lambda item: item[1], reverse=True) if sport not in preferred]
        sports = (preferred + remaining)[:6]
        ratio_mode = self._profit_graph_label() == "Profit to Sales Ratio"
        lines: list[dict[str, object]] = []
        for index, sport in enumerate(sports):
            values = []
            running_profit = 0.0
            for bucket in buckets:
                profit = profit_by_sport.get(sport, {}).get(bucket, 0.0)
                if ratio_mode:
                    sale = sales_by_sport.get(sport, {}).get(bucket, 0.0)
                    values.append(profit / sale if sale else 0.0)
                else:
                    if self._profit_graph_label() == "Overall Profit":
                        running_profit += profit
                        values.append(round(running_profit, 2))
                    else:
                        values.append(profit)
            color = PROFIT_SPORT_COLORS.get(sport) or ["#22c55e", "#eab308", "#38bdf8", "#fb7185", "#c084fc", "#f97316"][index % 6]
            lines.append({"label": sport, "values": values, "color": color})
        return buckets, lines, ratio_mode

    def _set_profit_view_mode(self, mode: str) -> None:
        self.profit_view_mode.set(mode)
        self.refresh_profit_tab()

    def _configure_profit_tree(self, mode: str) -> None:
        if not hasattr(self, "profit_tree"):
            return
        if mode == "Expenses":
            columns = ("date", "person", "type", "amount", "related", "notes")
            headings = {
                "date": "Date",
                "person": "Person",
                "type": "Type",
                "amount": "Amount",
                "related": "Related",
                "notes": "Notes",
            }
            widths = {"date": 95, "person": 150, "type": 120, "amount": 105, "related": 320, "notes": 320}
        elif mode == "Sold Sheets":
            columns = ("person", "sheet", "companies", "cards", "purchase", "sale", "profit", "last_sale")
            headings = {
                "person": "Person",
                "sheet": "Sold Sheet",
                "companies": "Companies",
                "cards": "Cards",
                "purchase": "Purchase",
                "sale": "Sale Price",
                "profit": "Profit",
                "last_sale": "Last Sale",
            }
            widths = {"person": 150, "sheet": 300, "companies": 220, "cards": 80, "purchase": 105, "sale": 105, "profit": 105, "last_sale": 95}
        else:
            columns = ("date", "person", "company", "card", "cert", "purchase", "sale", "profit", "sheet")
            headings = {
                "date": "Date",
                "person": "Person",
                "company": "Company",
                "card": "Card",
                "cert": "Cert / Item ID",
                "purchase": "Purchase",
                "sale": "Sale Price",
                "profit": "Profit",
                "sheet": "Company Sheet",
            }
            widths = {"date": 95, "person": 135, "company": 140, "card": 390, "cert": 100, "purchase": 105, "sale": 105, "profit": 105, "sheet": 200}
        columns = self._personal_person_last_columns(columns)
        self.profit_tree.configure(columns=columns)
        if self.profit_sort_column not in columns:
            self.profit_sort_column = columns[0]
            self.profit_sort_descending = self.profit_sort_column in {"date", "last_sale"}
        for column in columns:
            self.profit_tree.column(column, width=widths[column], minwidth=45, stretch=False)
        if hasattr(self, "_configure_sortable_tree_headings"):
            self._configure_sortable_tree_headings(self.profit_tree, headings, "profit")
        else:
            for column in columns:
                self.profit_tree.heading(column, text=headings[column], anchor=tk.W)
        self.profit_table_title_var.set(mode)

    def _expense_related_label(self, record: dict[str, object]) -> str:
        related_type = str(record.get("related_type") or "General").strip() or "General"
        source_sheet = str(record.get("source_sheet") or "").strip()
        item_id = str(record.get("item_id") or "").strip()
        cert = str(record.get("cert_number") or "").strip()
        if related_type == "Card":
            parts = [part for part in (source_sheet if source_sheet and source_sheet != "Expenses" else "", item_id or cert) if part]
            return " | ".join(parts) if parts else "Card"
        if related_type == "Sheet":
            return source_sheet if source_sheet and source_sheet != "Expenses" else "Sheet"
        return "General"

    def _expense_link_options(self, person: str = "") -> tuple[list[str], list[str], dict[str, dict[str, str]]]:
        person_filter = person.strip().lower()
        sheets: set[str] = set()
        card_options: list[str] = []
        card_lookup: dict[str, dict[str, str]] = {}
        seen_cards: set[tuple[str, str, str]] = set()
        rows = self.profit_rows if hasattr(self, "profit_rows") and self.profit_rows else [
            self._normalize_profit_record(record) for record in self._load_profit_ledger()
        ]
        for record in rows:
            if str(record.get("record_type") or "").strip().lower() == "expense":
                continue
            record_person = str(record.get("assigned_person") or "Unassigned").strip()
            if person_filter and person_filter not in record_person.lower():
                continue
            source_sheet = str(record.get("source_sheet") or "").strip()
            if not source_sheet:
                source_sheet = str(record.get("weekly_sheet_name") or "").strip()
            if not source_sheet:
                continue
            sheets.add(source_sheet)
            item_id = str(record.get("item_id") or "").strip()
            cert = str(record.get("cert_number") or "").strip()
            card_title = str(record.get("card_title") or "").strip()
            card_key = (source_sheet.lower(), item_id.lower(), cert.lower(), card_title.lower())
            if card_key in seen_cards:
                continue
            seen_cards.add(card_key)
            sale = self._money_value(record.get("sale_price"))
            label_parts = [source_sheet]
            if item_id:
                label_parts.append(item_id)
            if card_title:
                label_parts.append(card_title)
            if sale is not None:
                label_parts.append(format_money(sale))
            label = " | ".join(label_parts)
            card_options.append(label)
            card_lookup[label] = {
                "source_sheet": source_sheet,
                "item_id": item_id,
                "cert_number": cert,
            }
        return sorted(sheets, key=str.lower), sorted(card_options, key=str.lower), card_lookup

    def _profit_sheet_rows(self, rows: list[dict[str, object]]) -> list[dict[str, object]]:
        grouped: dict[tuple[str, str], dict[str, object]] = {}
        for record in rows:
            source_sheet = str(record.get("source_sheet") or "").strip() or "Unknown sheet"
            person = str(record.get("assigned_person") or "").strip() or "Unassigned"
            key = (person, source_sheet)
            group = grouped.setdefault(
                key,
                {
                    "person": person,
                    "sheet": source_sheet,
                    "companies": set(),
                    "cards": 0,
                    "purchase": 0.0,
                    "sale": 0.0,
                    "profit": 0.0,
                    "complete": 0,
                    "last_sale": "",
                },
            )
            company = str(record.get("company") or "").strip()
            if company:
                group["companies"].add(company)
            if str(record.get("record_type") or "").strip().lower() != "expense":
                group["cards"] = int(group["cards"]) + 1
            purchase = self._money_value(record.get("purchase_price"))
            sale = self._money_value(record.get("sale_price"))
            profit = self._money_value(record.get("profit"))
            if purchase is not None:
                group["purchase"] = float(group["purchase"]) + purchase
            if sale is not None:
                group["sale"] = float(group["sale"]) + sale
            if profit is not None:
                group["profit"] = float(group["profit"]) + profit
                group["complete"] = int(group["complete"]) + 1
            date = str(record.get("date_added") or "")[:10]
            if date and date > str(group["last_sale"]):
                group["last_sale"] = date
        result: list[dict[str, object]] = []
        for group in grouped.values():
            companies = sorted(group.pop("companies"), key=str.lower)
            group["companies"] = ", ".join(companies)
            result.append(group)
        return sorted(result, key=lambda item: (str(item.get("last_sale") or ""), str(item.get("person") or ""), str(item.get("sheet") or "")), reverse=True)

    def _append_profit_records(self, records: list[dict[str, object]]) -> int:
        if not records:
            return 0
        with shared_lock(CARD_PIPELINE_DIR, "profit-ledger", self.lucas_identity):
            ledger = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
            existing_keys = set().union(*(CardPipelineApp._profit_record_identity_keys(self, record) for record in ledger)) if ledger else set()
            added = 0
            for record in records:
                normalized = self._normalize_profit_record(record)
                keys = CardPipelineApp._profit_record_identity_keys(self, normalized)
                if not keys or keys & existing_keys:
                    continue
                if not str(normalized.get("ledger_added_at") or "").strip():
                    normalized["ledger_added_at"] = datetime.now().isoformat(timespec="microseconds")
                normalized["recorded_by"] = self.lucas_identity.get("display_name", "")
                normalized["recorded_machine"] = self.lucas_identity.get("machine", "")
                ledger.append(normalized)
                existing_keys.update(keys)
                added += 1
            if added:
                self._save_profit_ledger(ledger)
        return added

    def record_profit_sales(self, records: list[dict[str, object]]) -> int:
        added = CardPipelineApp._append_profit_records(self, records)
        self.refresh_profit_tab()
        return added

    def open_add_expense_popup(self) -> None:
        personal_inventory = self._is_personal_lucas()
        person_var = tk.StringVar(value=self._personal_default_person() if personal_inventory else (self.profit_person_var.get().strip() if hasattr(self, "profit_person_var") else ""))
        date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        type_var = tk.StringVar(value=EXPENSE_CATEGORY_OPTIONS[0])
        amount_var = tk.StringVar()
        link_var = tk.StringVar(value=EXPENSE_LINK_OPTIONS[0])
        sheet_var = tk.StringVar()
        item_id_var = tk.StringVar()
        cert_var = tk.StringVar()
        card_var = tk.StringVar()
        notes_var = tk.StringVar()

        popup = tk.Toplevel(self)
        popup.title("Add Expense")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Add Expense", style="Panel.TLabel", font=("Segoe UI Semibold", 12)).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 12))
        person_label = ttk.Label(frame, text="Person", style="Panel.TLabel")
        if not personal_inventory:
            person_label.grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        person_combo = ttk.Combobox(frame, textvariable=person_var, width=34)
        if not personal_inventory:
            person_combo.grid(row=1, column=1, sticky="ew", pady=(0, 10))
        self._bind_person_autocomplete(person_combo)
        ttk.Label(frame, text="Date", style="Panel.TLabel").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        ttk.Entry(frame, textvariable=date_var, width=18).grid(row=2, column=1, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="Type", style="Panel.TLabel").grid(row=3, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        ttk.Combobox(frame, textvariable=type_var, values=EXPENSE_CATEGORY_OPTIONS, width=18, state="readonly").grid(row=3, column=1, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="Amount", style="Panel.TLabel").grid(row=4, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        ttk.Entry(frame, textvariable=amount_var, width=18).grid(row=4, column=1, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="Tie To", style="Panel.TLabel").grid(row=5, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        link_combo = ttk.Combobox(frame, textvariable=link_var, values=EXPENSE_LINK_OPTIONS, width=18, state="readonly")
        link_combo.grid(row=5, column=1, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="Sheet", style="Panel.TLabel").grid(row=6, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        sheet_combo = ttk.Combobox(frame, textvariable=sheet_var, width=36, state="disabled")
        sheet_combo.grid(row=6, column=1, sticky="ew", pady=(0, 10))
        ttk.Label(frame, text="Card", style="Panel.TLabel").grid(row=7, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        card_combo = ttk.Combobox(frame, textvariable=card_var, width=50, state="disabled")
        card_combo.grid(row=7, column=1, sticky="ew", pady=(0, 10))
        ttk.Label(frame, text="Notes", style="Panel.TLabel").grid(row=8, column=0, sticky="w", padx=(0, 10), pady=(0, 14))
        ttk.Entry(frame, textvariable=notes_var, width=36).grid(row=8, column=1, sticky="ew", pady=(0, 14))
        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=9, column=0, columnspan=2, sticky="e")
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(
            buttons,
            text="Save",
            command=lambda: self._save_expense_from_popup(person_var, date_var, type_var, amount_var, link_var, sheet_var, item_id_var, cert_var, notes_var, popup),
            style="Primary.TButton",
        ).pack(side=tk.LEFT)
        card_lookup: dict[str, dict[str, str]] = {}

        def refresh_link_options(*_args) -> None:
            nonlocal card_lookup
            sheet_options, card_options, card_lookup = self._expense_link_options(person_var.get())
            related_type = link_var.get().strip()
            sheet_combo["values"] = sheet_options
            filtered_card_options = [
                option for option in card_options
                if not sheet_var.get().strip() or card_lookup.get(option, {}).get("source_sheet") == sheet_var.get().strip()
            ]
            card_combo["values"] = filtered_card_options if related_type == "Card" else card_options
            if related_type == "Sheet":
                sheet_combo.configure(state="readonly")
                card_combo.configure(state="disabled")
                card_var.set("")
                item_id_var.set("")
                cert_var.set("")
                if sheet_var.get() not in sheet_options:
                    sheet_var.set(sheet_options[0] if sheet_options else "")
            elif related_type == "Card":
                sheet_combo.configure(state="readonly")
                card_combo.configure(state="readonly")
                if sheet_var.get() not in sheet_options:
                    sheet_var.set(sheet_options[0] if sheet_options else "")
                filtered_card_options = [
                    option for option in card_options
                    if not sheet_var.get().strip() or card_lookup.get(option, {}).get("source_sheet") == sheet_var.get().strip()
                ]
                card_combo["values"] = filtered_card_options
                if card_var.get() not in filtered_card_options:
                    card_var.set(filtered_card_options[0] if filtered_card_options else "")
                selection = card_lookup.get(card_var.get(), {})
                sheet_var.set(selection.get("source_sheet", ""))
                item_id_var.set(selection.get("item_id", ""))
                cert_var.set(selection.get("cert_number", ""))
            else:
                sheet_combo.configure(state="disabled")
                card_combo.configure(state="disabled")
                sheet_var.set("")
                item_id_var.set("")
                cert_var.set("")
                card_var.set("")

        def apply_card_selection(*_args) -> None:
            selection = card_lookup.get(card_var.get(), {})
            sheet_var.set(selection.get("source_sheet", ""))
            item_id_var.set(selection.get("item_id", ""))
            cert_var.set(selection.get("cert_number", ""))

        def refresh_cards_for_sheet(*_args) -> None:
            if link_var.get().strip() == "Card":
                refresh_link_options()

        link_combo.bind("<<ComboboxSelected>>", refresh_link_options, add="+")
        person_combo.bind("<<ComboboxSelected>>", refresh_link_options, add="+")
        person_var.trace_add("write", refresh_link_options)
        sheet_combo.bind("<<ComboboxSelected>>", refresh_cards_for_sheet, add="+")
        card_combo.bind("<<ComboboxSelected>>", apply_card_selection, add="+")
        card_var.trace_add("write", apply_card_selection)
        refresh_link_options()
        frame.columnconfigure(1, weight=1)
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")

    def _save_expense_from_popup(
        self,
        person_var: tk.StringVar,
        date_var: tk.StringVar,
        type_var: tk.StringVar,
        amount_var: tk.StringVar,
        link_var: tk.StringVar,
        sheet_var: tk.StringVar,
        item_id_var: tk.StringVar,
        cert_var: tk.StringVar,
        notes_var: tk.StringVar,
        popup: tk.Toplevel,
    ) -> None:
        person = person_var.get().strip()
        if getattr(self, "_is_personal_lucas", lambda: False)():
            person = self._personal_default_person()
        person_choice = self._canonical_person_choice(person)
        if person_choice is None:
            messagebox.showinfo("Person required", "Choose an existing person for this expense.")
            return
        person = person_choice
        expense_date = date_var.get().strip()
        if self._profit_record_date(expense_date) is None:
            messagebox.showinfo("Date required", "Enter the expense date as YYYY-MM-DD.")
            return
        amount = self._money_value(amount_var.get())
        if amount is None or amount <= 0:
            messagebox.showinfo("Amount required", "Enter an expense amount greater than zero.")
            return
        expense_type = type_var.get().strip()
        if expense_type not in EXPENSE_CATEGORY_OPTIONS:
            expense_type = "Fees"
        related_type = link_var.get().strip()
        if related_type not in EXPENSE_LINK_OPTIONS:
            related_type = "General"
        related_sheet = sheet_var.get().strip()
        related_item_id = item_id_var.get().strip()
        related_cert = cert_var.get().strip()
        if related_type == "Sheet" and not related_sheet:
            messagebox.showinfo("Sheet required", "Choose the sold sheet this expense belongs to.")
            return
        if related_type == "Card" and not (related_sheet or related_item_id or related_cert):
            messagebox.showinfo("Card required", "Choose the sold card this expense belongs to.")
            return
        record = {
            "record_type": "expense",
            "expense_id": datetime.now().strftime("%Y%m%d%H%M%S%f"),
            "date_added": expense_date[:10],
            "assigned_person": person,
            "expense_type": expense_type,
            "expense_amount": amount,
            "related_type": related_type,
            "source_sheet": related_sheet,
            "item_id": related_item_id,
            "cert_number": related_cert,
            "notes": notes_var.get().strip(),
        }
        added = self.record_profit_sales([record])
        if added:
            popup.destroy()
            self.status_var.set(f"Added {expense_type} expense for {person}: {format_money(amount)}.")
            self._append_activity("Expense Add", f"Added {expense_type} expense for {person}: {format_money(amount)}.", {"person": person, "expense_type": expense_type, "amount": amount, "related_type": related_type})
        else:
            messagebox.showinfo("Expense not added", "That expense already exists in the profit ledger.")

    def _delete_profit_expense_records(self, records: list[dict[str, object]]) -> int:
        expense_keys: set[str] = set()
        for record in records:
            normalized = self._normalize_profit_record(record)
            if str(normalized.get("record_type") or "").strip().lower() != "expense":
                continue
            key = str(normalized.get("ledger_key") or self._profit_record_key(normalized) or "")
            if key:
                expense_keys.add(key)
        if not expense_keys:
            return 0
        with shared_lock(CARD_PIPELINE_DIR, "profit-expense-delete", self.lucas_identity):
            ledger = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
            kept: list[dict[str, object]] = []
            deleted = 0
            for record in ledger:
                key = str(record.get("ledger_key") or self._profit_record_key(record) or "")
                is_expense = str(record.get("record_type") or "").strip().lower() == "expense"
                if is_expense and key in expense_keys:
                    deleted += 1
                    continue
                kept.append(record)
            if deleted:
                self._save_profit_ledger(kept)
            return deleted

    def delete_selected_profit_expenses(self) -> None:
        if not hasattr(self, "profit_tree"):
            return
        selected = list(self.profit_tree.selection())
        records = [self.profit_tree_records.get(iid) for iid in selected if self.profit_tree_records.get(iid)]
        if not records:
            messagebox.showinfo("Choose expense", "Select one or more expense rows to delete.")
            return
        if any(str(record.get("record_type") or "").strip().lower() != "expense" for record in records):
            messagebox.showinfo("Delete expenses only", "Only expense rows can be deleted here. Sold cards should be refunded instead.")
            return
        confirmed = messagebox.askyesno(
            "Delete expense(s)?",
            f"Delete {len(records)} expense row(s) from the profit ledger?",
        )
        if not confirmed:
            return
        deleted = self._delete_profit_expense_records(records)
        self.refresh_profit_tab()
        self.status_var.set(f"Deleted {deleted} expense row(s) from the profit ledger.")
        if deleted:
            self._append_activity("Expense Delete", f"Deleted {deleted} expense row(s).", {"deleted": deleted})

    def refund_selected_profit_to_inventory(self) -> None:
        if not hasattr(self, "profit_tree"):
            return
        selected = list(self.profit_tree.selection())
        records = [self.profit_tree_records.get(iid) for iid in selected if self.profit_tree_records.get(iid)]
        if not records:
            messagebox.showinfo("Choose sold cards", "Select one or more sold card rows to refund.")
            return
        if any(str(record.get("record_type") or "").strip().lower() == "expense" for record in records):
            messagebox.showinfo("Cannot refund expenses", "Expense rows adjust profit only and cannot be returned to inventory.")
            return
        confirmed = messagebox.askyesno(
            "Refund selected card(s)?",
            f"Refund {len(records)} sold card(s) and return them to active inventory?",
        )
        if not confirmed:
            return
        refunded = 0
        inventory_records: list[dict[str, object]] = []
        with shared_lock(CARD_PIPELINE_DIR, "refund-inventory", self.lucas_identity):
            ledger = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
            refund_keys = {str(self._normalize_profit_record(record).get("ledger_key") or "") for record in records}
            kept = [record for record in ledger if str(record.get("ledger_key") or "") not in refund_keys]
            refunded = len(ledger) - len(kept)
            if refunded:
                self._save_profit_ledger(kept)
            for record in records:
                normalized = self._normalize_profit_record(record)
                source_sheet = str(normalized.get("source_sheet") or "")
                cert = str(normalized.get("cert_number") or "")
                if source_sheet and cert:
                    remove_company_sheet_rows_for_source(COMPANY_SHEETS_DIR, source_sheet, {cert})
                inventory_records.append(
                    self._normalize_inventory_record(
                        {
                            "date_added": datetime.now().strftime("%Y-%m-%d"),
                            "item_type": normalized.get("item_type") or ("Raw" if str(normalized.get("item_id") or "").upper().startswith("RAW-") else "Graded"),
                            "item_id": normalized.get("item_id") or "",
                            "assigned_person": normalized.get("assigned_person") or self._person_for_profit_record(normalized) or "Unassigned",
                            "sport": CardPipelineApp._inventory_sport_from_value(self, normalized.get("sport") or normalized.get("category"), normalized.get("card_title")),
                            "cert_number": normalized.get("cert_number") or "",
                            "grader": normalized.get("grader") or "",
                            "card_title": normalized.get("card_title") or "",
                            "purchase_price": normalized.get("purchase_price"),
                            "card_ladder_value": normalized.get("card_ladder_value"),
                            "card_ladder_comps_average": normalized.get("card_ladder_comps_average") or normalized.get("comps"),
                            "cy_value": normalized.get("cy_value") or normalized.get("cy_estimate"),
                            "inventory_value": normalized.get("sale_price") or normalized.get("card_ladder_value") or normalized.get("comps") or normalized.get("cy_estimate"),
                            "source_sheet": normalized.get("source_sheet") or "",
                            "source": normalized.get("source") or "",
                            "photo_paths": list(normalized.get("photo_paths") or []),
                            "status": "Active",
                            "notes": "Refunded from sold cards",
                        }
                    )
                )
            restore_photos = getattr(self, "_restore_inventory_photo_files_for_records", None)
            if callable(restore_photos):
                restore_photos(inventory_records)
            self.add_inventory_records(inventory_records)
        self.refresh_profit_tab()
        self.refresh_inventory_tab()
        self.status_var.set(f"Refunded {refunded or len(records)} card(s) back to active inventory.")
        self._append_activity("Refund", f"Refunded {refunded or len(records)} sold card(s) back to active inventory.", {"refunded": refunded or len(records)})

    def _show_profit_context_menu(self, event: tk.Event) -> None:
        if not hasattr(self, "profit_tree"):
            return
        iid = self.profit_tree.identify_row(event.y)
        if iid:
            if iid not in self.profit_tree.selection():
                self.profit_tree.selection_set(iid)
            self.profit_tree.focus(iid)
        selected = list(self.profit_tree.selection())
        records = [self.profit_tree_records.get(item) for item in selected if self.profit_tree_records.get(item)]
        if not records:
            return
        expenses = [record for record in records if str(record.get("record_type") or "").strip().lower() == "expense"]
        sold_cards = [record for record in records if str(record.get("record_type") or "").strip().lower() != "expense"]
        menu = tk.Menu(self, tearoff=0)
        if sold_cards and len(sold_cards) == len(records):
            menu.add_command(label="Refund to Inventory", command=self.refund_selected_profit_to_inventory)
        if expenses and len(expenses) == len(records):
            menu.add_command(label="Delete Expense", command=self.delete_selected_profit_expenses)
        if menu.index("end") is None:
            return
        menu.tk_popup(event.x_root, event.y_root)

    def recover_sold_ledger(self) -> None:
        if getattr(self, "_profit_recovery_running", False):
            messagebox.showinfo("Recover Sold Ledger", "Recovery is already scanning company sheets.")
            return
        self._profit_recovery_running = True
        if hasattr(self, "profit_recover_button"):
            self.profit_recover_button.configure(state=tk.DISABLED)
        self.profit_status_var.set("Recover Sold Ledger is scanning company sheets in the background...")
        self.status_var.set("Recover Sold Ledger is scanning company sheets...")

        def worker() -> None:
            started = time.perf_counter()
            try:
                records = read_company_profit_records(COMPANY_SHEETS_DIR)
                record_performance_event("profit.company_sheet_scan", started, f"records={len(records)} background=1")
                self.events.put(("profit_recovery_done", {"records": records}))
            except Exception as error:
                record_performance_event("profit.company_sheet_scan", started, f"error={error} background=1", force=True)
                self.events.put(("profit_recovery_error", {"error": str(error)}))

        threading.Thread(target=worker, daemon=True).start()

    def _merge_profit_recovery_records(self, company_profit_records: list[dict[str, object]]) -> tuple[int, int]:
        ledger = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
        pruned_manual = len([record for record in ledger if self._is_manual_company_profit_backfill(record)])
        if pruned_manual:
            ledger = [record for record in ledger if not self._is_manual_company_profit_backfill(record)]
        existing_keys = set().union(*(CardPipelineApp._profit_record_identity_keys(self, record) for record in ledger)) if ledger else set()
        backfilled = 0
        for record in company_profit_records:
            normalized = self._normalize_profit_record(record)
            keys = CardPipelineApp._profit_record_identity_keys(self, normalized)
            if not keys or keys & existing_keys:
                continue
            ledger.append(normalized)
            existing_keys.update(keys)
            backfilled += 1
        if backfilled or pruned_manual:
            with shared_lock(CARD_PIPELINE_DIR, "profit-ledger", self.lucas_identity):
                current = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
                current = [record for record in current if not self._is_manual_company_profit_backfill(record)]
                current_keys = set().union(*(CardPipelineApp._profit_record_identity_keys(self, record) for record in current)) if current else set()
                for record in ledger:
                    keys = CardPipelineApp._profit_record_identity_keys(self, record)
                    if keys and not (keys & current_keys):
                        current.append(record)
                        current_keys.update(keys)
                self._save_profit_ledger(current)
        return backfilled, pruned_manual

    def _finish_profit_recovery(self, payload: dict[str, object]) -> None:
        records = list(payload.get("records") or [])
        backfilled, pruned_manual = self._merge_profit_recovery_records(records)
        self._profit_recovery_running = False
        if hasattr(self, "profit_recover_button"):
            self.profit_recover_button.configure(state=tk.NORMAL)
        self.refresh_profit_tab()
        cleanup_suffix = f" Removed {pruned_manual} manual company row(s)." if pruned_manual else ""
        self.profit_status_var.set(f"Recover Sold Ledger scanned {len(records)} company sale row(s) and added {backfilled}.{cleanup_suffix}")
        self.status_var.set("Recover Sold Ledger complete.")
        self._append_activity("Profit Recovery", f"Recovered {backfilled} sold ledger row(s).", {"scanned": len(records), "backfilled": backfilled, "removed_manual": pruned_manual})

    def _handle_profit_recovery_error(self, payload: dict[str, object]) -> None:
        self._profit_recovery_running = False
        if hasattr(self, "profit_recover_button"):
            self.profit_recover_button.configure(state=tk.NORMAL)
        error = str(payload.get("error") or "Unknown error")
        self.profit_status_var.set(f"Recover Sold Ledger failed: {error}")
        self.status_var.set("Recover Sold Ledger failed.")
        self._show_error_with_copy("Recover Sold Ledger failed", "Could not scan company sheets.", {"error": error, "company_sheets_dir": str(COMPANY_SHEETS_DIR)})

    def refresh_profit_tab(self, deep_sync: bool = False) -> None:
        perf_start = time.perf_counter()
        ledger = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
        pruned_manual = len([record for record in ledger if self._is_manual_company_profit_backfill(record)])
        if pruned_manual:
            ledger = [record for record in ledger if not self._is_manual_company_profit_backfill(record)]
        ledger, pruned_duplicates = CardPipelineApp._dedupe_profit_records(self, ledger)
        if pruned_manual or pruned_duplicates:
            with shared_lock(CARD_PIPELINE_DIR, "profit-ledger", self.lucas_identity):
                current = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
                kept = [record for record in current if not self._is_manual_company_profit_backfill(record)]
                kept, _removed = CardPipelineApp._dedupe_profit_records(self, kept)
                if len(kept) != len(current):
                    self._save_profit_ledger(kept)
                    ledger = kept
        existing_keys = set().union(*(CardPipelineApp._profit_record_identity_keys(self, record) for record in ledger)) if ledger else set()
        backfilled = 0
        if deep_sync:
            company_backfill_start = time.perf_counter()
            company_profit_records = read_company_profit_records(COMPANY_SHEETS_DIR)
            record_performance_event("profit.company_sheet_scan", company_backfill_start, f"records={len(company_profit_records)}")
            for record in company_profit_records:
                normalized = self._normalize_profit_record(record)
                keys = CardPipelineApp._profit_record_identity_keys(self, normalized)
                if not keys or keys & existing_keys:
                    continue
                ledger.append(normalized)
                existing_keys.update(keys)
                backfilled += 1
        if backfilled:
            with shared_lock(CARD_PIPELINE_DIR, "profit-ledger", self.lucas_identity):
                current = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
                current_keys = set().union(*(CardPipelineApp._profit_record_identity_keys(self, record) for record in current)) if current else set()
                for record in ledger:
                    keys = CardPipelineApp._profit_record_identity_keys(self, record)
                    if keys and not (keys & current_keys):
                        current.append(record)
                        current_keys.update(keys)
                self._save_profit_ledger(current)
                ledger = current
        self.profit_rows = self._enrich_profit_records_with_people(ledger)
        self.profit_rows.sort(
            key=lambda record: (
                str(record.get("date_added") or ""),
                self._profit_added_sort_value(record),
                str(record.get("company") or ""),
                str(record.get("card_title") or ""),
            ),
            reverse=True,
        )
        self.filtered_profit_rows = self._filtered_profit_records(self.profit_rows)
        if not hasattr(self, "profit_tree"):
            record_performance_event(
                "profit.refresh",
                perf_start,
                f"rows={len(self.profit_rows)} filtered={len(self.filtered_profit_rows)} backfilled={backfilled} tree=missing",
            )
            return
        self._refresh_person_combo_values()
        mode = self.profit_view_mode.get()
        self._configure_profit_tree(mode)
        if mode != "Sold Sheets" and hasattr(self, "_sorted_records"):
            self.filtered_profit_rows = self._sorted_records(
                self.filtered_profit_rows,
                getattr(self, "profit_sort_column", "date"),
                bool(getattr(self, "profit_sort_descending", True)),
                "profit",
                mode,
            )
        self.profit_tree.delete(*self.profit_tree.get_children())
        self.profit_tree_records = {}
        total_purchase = 0.0
        total_sale = 0.0
        total_profit = 0.0
        total_expenses = 0.0
        total_gross_profit = 0.0
        complete_count = 0
        for record in self.filtered_profit_rows:
            is_expense = str(record.get("record_type") or "").strip().lower() == "expense"
            purchase = self._money_value(record.get("purchase_price"))
            sale = self._money_value(record.get("sale_price"))
            profit = self._money_value(record.get("profit"))
            if purchase is not None:
                total_purchase += purchase
            if sale is not None:
                total_sale += sale
            if profit is not None:
                total_profit += profit
                if is_expense:
                    total_expenses += abs(profit)
                else:
                    total_gross_profit += profit
                complete_count += 1
            if mode == "Expenses":
                if not is_expense:
                    continue
                values_by_column = {
                    "date": record.get("date_added") or "",
                    "person": record.get("assigned_person") or "Unassigned",
                    "type": record.get("expense_type") or "",
                    "amount": format_money(record.get("expense_amount")),
                    "related": self._expense_related_label(record),
                    "notes": record.get("notes") or "",
                }
                iid = self.profit_tree.insert(
                    "",
                    tk.END,
                    values=tuple(values_by_column.get(column, "") for column in self.profit_tree["columns"]),
                    tags=("profit_negative",),
                )
                self.profit_tree_records[iid] = record
            elif mode != "Sold Sheets":
                tag = "profit_negative" if profit is not None and profit < 0 else "profit_positive"
                values_by_column = {
                    "date": record.get("date_added") or "",
                    "person": record.get("assigned_person") or "Unassigned",
                    "company": record.get("company") or "",
                    "card": record.get("card_title") or "",
                    "cert": record.get("cert_number") or record.get("item_id") or "",
                    "purchase": format_money(purchase),
                    "sale": format_money(sale),
                    "profit": format_money(profit),
                    "sheet": record.get("weekly_sheet_name") or record.get("source_sheet") or "",
                }
                iid = self.profit_tree.insert(
                    "",
                    tk.END,
                    values=tuple(values_by_column.get(column, "") for column in self.profit_tree["columns"]),
                    tags=(tag,),
                )
                self.profit_tree_records[iid] = record
        if mode == "Sold Sheets":
            sheet_rows = self._profit_sheet_rows(self.filtered_profit_rows)
            if hasattr(self, "_sorted_records"):
                sheet_rows = self._sorted_records(
                    sheet_rows,
                    getattr(self, "profit_sort_column", "last_sale"),
                    bool(getattr(self, "profit_sort_descending", True)),
                    "profit_sheet",
                    mode,
                )
            for sheet_row in sheet_rows:
                profit = self._money_value(sheet_row.get("profit"))
                tag = "profit_negative" if profit is not None and profit < 0 else "profit_positive"
                values_by_column = {
                    "person": sheet_row.get("person") or "",
                    "sheet": sheet_row.get("sheet") or "",
                    "companies": sheet_row.get("companies") or "",
                    "cards": sheet_row.get("cards") or 0,
                    "purchase": format_money(sheet_row.get("purchase")),
                    "sale": format_money(sheet_row.get("sale")),
                    "profit": format_money(profit),
                    "last_sale": sheet_row.get("last_sale") or "",
                }
                self.profit_tree.insert(
                    "",
                    tk.END,
                    values=tuple(values_by_column.get(column, "") for column in self.profit_tree["columns"]),
                    tags=(tag,),
                )
        display_count = len([record for record in self.filtered_profit_rows if str(record.get("record_type") or "").strip().lower() == "expense"]) if mode == "Expenses" else len(self.filtered_profit_rows)
        if self.filtered_profit_rows:
            if mode == "Expenses":
                total_by_column = {"date": "TOTAL", "amount": format_money(total_expenses)}
            elif mode == "Sold Sheets":
                total_by_column = {
                    "person": "TOTAL",
                    "cards": f"{len(self.filtered_profit_rows)} card(s)",
                    "purchase": format_money(total_purchase),
                    "sale": format_money(total_sale),
                    "profit": format_money(total_profit),
                }
            else:
                total_by_column = {
                    "date": "TOTAL",
                    "card": f"{len(self.filtered_profit_rows)} card(s)",
                    "purchase": format_money(total_purchase),
                    "sale": format_money(total_sale),
                    "profit": format_money(total_profit),
                }
            self.profit_tree.insert(
                "",
                tk.END,
                values=tuple(total_by_column.get(column, "") for column in self.profit_tree["columns"]),
                tags=("total_row",),
            )
        self.profit_metric_var.set(
            f"{self._profit_period_label()}   Sales: {format_money(total_sale)}   Gross: {format_money(total_gross_profit)}   Expenses: {format_money(total_expenses)}   Net: {format_money(total_profit)}"
        )
        missing = len(self.filtered_profit_rows) - complete_count
        suffix = f" | {missing} card(s) missing purchase or sale price" if missing else ""
        filter_label = self.profit_person_var.get().strip()
        filter_suffix = f" | Filter: {filter_label}" if filter_label else ""
        search_label = self.profit_search_var.get().strip() if hasattr(self, "profit_search_var") else ""
        search_suffix = f" | Search: {search_label}" if search_label else ""
        period_suffix = f" | Period: {self._profit_period_label()}"
        backfill_suffix = f" | backfilled {backfilled} from company sheets" if backfilled else ""
        sync_suffix = " | deep sync checked company sheets" if deep_sync and not backfilled else ""
        cleanup_parts = []
        if pruned_manual:
            cleanup_parts.append(f"removed {pruned_manual} manual row(s)")
        if pruned_duplicates:
            cleanup_parts.append(f"removed {pruned_duplicates} duplicate row(s)")
        cleanup_suffix = f" | {'; '.join(cleanup_parts)}" if cleanup_parts else ""
        self.profit_status_var.set(f"Loaded {display_count}/{len(self.profit_rows)} profit row(s) from {PROFIT_LEDGER_PATH.name}{filter_suffix}{search_suffix}{period_suffix}{suffix}{backfill_suffix}{sync_suffix}{cleanup_suffix}.")
        self._draw_profit_chart()
        record_performance_event(
            "profit.refresh",
            perf_start,
            f"rows={len(self.profit_rows)} filtered={len(self.filtered_profit_rows)} mode={mode} deep_sync={deep_sync} backfilled={backfilled} pruned_manual={pruned_manual}",
        )

    def _profit_month_key(self, value: object) -> str:
        text = str(value or "").strip()
        try:
            return datetime.strptime(text[:10], "%Y-%m-%d").strftime("%Y-%m")
        except ValueError:
            return text[:7] if len(text) >= 7 else "Unknown"

    def _profit_chart_tooltip_value(self, value: float, percent_mode: bool) -> str:
        return f"{value * 100:.2f}%" if percent_mode else format_money(value)

    def _hide_profit_chart_tooltip(self) -> None:
        if hasattr(self, "profit_chart_canvas"):
            self.profit_chart_canvas.delete("profit_tooltip")

    def _show_profit_chart_tooltip(self, event: tk.Event) -> None:
        points = getattr(self, "profit_chart_points", [])
        if not points or not hasattr(self, "profit_chart_canvas"):
            return
        nearest = None
        nearest_distance = 14 * 14
        for point in points:
            distance = (event.x - point["x"]) ** 2 + (event.y - point["y"]) ** 2
            if distance <= nearest_distance:
                nearest = point
                nearest_distance = distance
        if not nearest:
            self._hide_profit_chart_tooltip()
            return
        canvas = self.profit_chart_canvas
        self._hide_profit_chart_tooltip()
        series = str(nearest.get("series") or "").strip()
        label = str(nearest.get("label") or "").strip()
        value = self._profit_chart_tooltip_value(float(nearest.get("value") or 0.0), bool(nearest.get("percent_mode")))
        text = f"{series}\n{label}: {value}" if series else f"{label}: {value}"
        x = min(max(float(nearest["x"]) + 12, 8), max(canvas.winfo_width() - 150, 8))
        y = max(float(nearest["y"]) - 34, 8)
        text_id = canvas.create_text(x + 8, y + 6, anchor="nw", text=text, fill="#f5f5f5", font=("Segoe UI", 9), tags=("profit_tooltip",))
        bbox = canvas.bbox(text_id)
        if not bbox:
            return
        x1, y1, x2, y2 = bbox
        rect_id = canvas.create_rectangle(x1 - 6, y1 - 4, x2 + 6, y2 + 4, fill="#111111", outline="#4b5563", tags=("profit_tooltip",))
        canvas.tag_lower(rect_id, text_id)

    def _draw_profit_chart(self) -> None:
        if not hasattr(self, "profit_chart_canvas"):
            return
        canvas = self.profit_chart_canvas
        canvas.delete("all")
        self.profit_chart_points = []
        width = max(canvas.winfo_width(), 400)
        height = max(canvas.winfo_height(), 120)
        pad_left, pad_right = 62, 22
        pad_top = 26 if height >= 170 else 18
        pad_bottom = 44 if height >= 170 else 28
        plot_w = max(width - pad_left - pad_right, 10)
        plot_h = max(height - pad_top - pad_bottom, 10)
        chart_rows = self.filtered_profit_rows if hasattr(self, "filtered_profit_rows") else self.profit_rows
        labels, chart_lines, percent_mode = self._profit_chart_lines(chart_rows)
        chart_title = self._profit_chart_title()
        if hasattr(self, "profit_chart_title_var"):
            self.profit_chart_title_var.set(chart_title)
        if not labels or not chart_lines:
            canvas.create_text(width / 2, height / 2, text="No profit data yet", fill="#b3b3b3", font=("Segoe UI", 12, "bold"))
            return
        values = [float(value) for line in chart_lines for value in line.get("values", [])] + [0.0]
        min_y = min(values)
        max_y = max(values)
        if min_y == max_y:
            min_y -= 1
            max_y += 1
        def x_at(index: int) -> float:
            if len(labels) == 1:
                return pad_left + plot_w / 2
            return pad_left + (plot_w * index / (len(labels) - 1))
        def y_at(value: float) -> float:
            return pad_top + (max_y - value) / (max_y - min_y) * plot_h
        def value_label(value: float) -> str:
            return f"{value * 100:.2f}%" if percent_mode else format_money(value)
        zero_y = y_at(0.0)
        grid_lines = 4
        for line_index in range(grid_lines + 1):
            y = pad_top + (plot_h * line_index / grid_lines)
            value = max_y - ((max_y - min_y) * line_index / grid_lines)
            canvas.create_line(pad_left, y, pad_left + plot_w, y, fill="#2f2f2f")
            canvas.create_text(8, y - 6, anchor="nw", text=value_label(value), fill="#8f8f8f", font=("Segoe UI", 8))
        vertical_lines = min(max(len(labels), 2), 8)
        for line_index in range(vertical_lines):
            x = pad_left + (plot_w * line_index / max(vertical_lines - 1, 1))
            canvas.create_line(x, pad_top, x, pad_top + plot_h, fill="#2a2a2a")
        canvas.create_line(pad_left, pad_top, pad_left, pad_top + plot_h, fill="#555555")
        canvas.create_line(pad_left, zero_y, pad_left + plot_w, zero_y, fill="#555555")
        for line in chart_lines:
            line_values = [float(value) for value in line.get("values", [])]
            color = str(line.get("color") or "#22c55e")
            series_label = str(line.get("label") or "")
            points = [(x_at(index), y_at(value)) for index, value in enumerate(line_values)]
            for first, second in zip(points, points[1:]):
                canvas.create_line(*first, *second, fill=color, width=3)
            for index, (x, y) in enumerate(points):
                value = line_values[index]
                point_color = "#ef4444" if value < 0 and not percent_mode else color
                canvas.create_oval(x - 4, y - 4, x + 4, y + 4, fill=point_color, outline="")
                self.profit_chart_points.append({
                    "x": x,
                    "y": y,
                    "label": self._profit_chart_bucket_display(labels[index]),
                    "series": series_label,
                    "value": value,
                    "percent_mode": percent_mode,
                })
        for index, label in enumerate(labels):
            if len(labels) <= 14 or index % max(1, len(labels) // 8) == 0:
                canvas.create_text(x_at(index), height - max(18, pad_bottom - 20), text=self._profit_chart_bucket_display(label), fill="#b3b3b3", font=("Segoe UI", 8))
        legend_x = pad_left
        legend_y = 8
        for line in chart_lines[:6]:
            color = str(line.get("color") or "#22c55e")
            label = str(line.get("label") or "")
            canvas.create_oval(legend_x, legend_y + 4, legend_x + 8, legend_y + 12, fill=color, outline="")
            canvas.create_text(legend_x + 12, legend_y, anchor="nw", text=label, fill=color, font=("Segoe UI", 9, "bold"))
            legend_x += max(88, len(label) * 8 + 34)

    def choose_working_folder(self) -> None:
        selected = filedialog.askdirectory(
            title="Choose WORKING SHEETS folder",
            initialdir=str(WORKING_SHEETS_DIR if WORKING_SHEETS_DIR.exists() else CARD_PIPELINE_DIR if CARD_PIPELINE_DIR.exists() else ROOT),
        )
        if not selected:
            return
        set_pipeline_from_working_dir(Path(selected))
        settings = load_app_settings()
        settings["pipeline_root"] = str(CARD_PIPELINE_DIR)
        settings["working_sheets_dir"] = str(WORKING_SHEETS_DIR)
        save_app_settings(settings)
        self.pipeline_root_var.set(str(CARD_PIPELINE_DIR))
        for directory in (WORKING_SHEETS_DIR, INCOMING_SHEETS_DIR, RECEIVED_SHEETS_DIR):
            directory.mkdir(parents=True, exist_ok=True)
        COMPANY_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
        self._load_player_overrides()
        self.assignment_engine = AssignmentEngine.load()
        self.assignment_config_status.set(self._assignment_config_status())
        self.home_sheet_markers = self._load_sheet_markers()
        self.refresh_profit_tab()
        self.status_var.set(f"Working folder set to {WORKING_SHEETS_DIR}")
        self.refresh_home()
        self.refresh_working_sheets()
        self.refresh_incoming_index()
        self.refresh_received_sheets()

    def choose_pipeline_root(self) -> None:
        self.choose_working_folder()

    def _configure_colored_button(
        self,
        button: tk.Widget,
        bg: str,
        fg: str,
        hover: str | None = None,
        pressed: str | None = None,
    ) -> None:
        setattr(button, "_lucas_bg", bg)
        setattr(button, "_lucas_fg", fg)
        setattr(button, "_lucas_hover", hover or bg)
        setattr(button, "_lucas_pressed", pressed or hover or bg)
        button.configure(bg=bg, fg=fg)

    def _make_colored_button(self, parent: tk.Widget, text: str, command, variant: str = "soft") -> tk.Label:
        palette = getattr(self, "app_palette", {})
        if variant == "primary":
            bg = str(palette.get("button") or "#1ed760")
            hover = str(palette.get("button_hover") or "#1fdf64")
            pressed = str(palette.get("button_pressed") or "#169c46")
            fg = "#000000"
        else:
            bg = str(palette.get("button") or "#1ed760")
            hover = str(palette.get("button_hover") or "#1fdf64")
            pressed = str(palette.get("button_pressed") or "#169c46")
            fg = "#000000"
        button = tk.Label(
            parent,
            text=text,
            bg=bg,
            fg=fg,
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=0,
            padx=16,
            pady=9,
            font=("Segoe UI Semibold", 10),
            cursor="hand2",
        )
        self._configure_colored_button(button, bg, fg, hover, pressed)
        button.bind("<Enter>", lambda _event: button.configure(bg=getattr(button, "_lucas_hover", bg)), add="+")
        button.bind("<Leave>", lambda _event: button.configure(bg=getattr(button, "_lucas_bg", bg)), add="+")
        button.bind("<ButtonPress-1>", lambda _event: button.configure(bg=getattr(button, "_lucas_pressed", pressed)), add="+")
        button.bind("<ButtonRelease-1>", lambda _event: (button.configure(bg=getattr(button, "_lucas_hover", hover)), command()), add="+")
        self._set_tooltip(button)
        return button

    def _make_inventory_toolbar_icon_button(self, parent: tk.Widget, icon: str, tooltip: str, command) -> tk.Canvas:
        palette = getattr(self, "app_palette", {})
        bg = str(palette.get("button") or "#1ed760")
        hover = str(palette.get("button_hover") or "#1fdf64")
        pressed = str(palette.get("button_pressed") or "#169c46")
        canvas = tk.Canvas(
            parent,
            width=38,
            height=38,
            bg=bg,
            highlightthickness=0,
            borderwidth=0,
            cursor="hand2",
        )
        setattr(canvas, "_lucas_bg", bg)
        setattr(canvas, "_lucas_hover", hover)
        setattr(canvas, "_lucas_pressed", pressed)
        setattr(canvas, "_lucas_icon", icon)
        setattr(canvas, "_lucas_command", command)

        def redraw(fill: str) -> None:
            canvas.delete("all")
            canvas.configure(bg=fill)
            if icon == "filter":
                canvas.create_polygon(9, 10, 29, 10, 22, 19, 22, 27, 16, 30, 16, 19, fill="#000000", outline="")
            else:
                canvas.create_text(19, 19, text="⚙", fill="#000000", font=("Segoe UI Semibold", 19))

        canvas._lucas_redraw = redraw  # type: ignore[attr-defined]
        redraw(bg)
        canvas.bind("<Enter>", lambda _event: redraw(getattr(canvas, "_lucas_hover", hover)), add="+")
        canvas.bind("<Leave>", lambda _event: redraw(getattr(canvas, "_lucas_bg", bg)), add="+")
        canvas.bind("<ButtonPress-1>", lambda _event: redraw(getattr(canvas, "_lucas_pressed", pressed)), add="+")
        canvas.bind("<ButtonRelease-1>", lambda event, button=canvas, run=command: self._release_inventory_icon_button(event, button, run), add="+")
        self._set_tooltip(canvas, tooltip)
        return canvas

    def _release_inventory_icon_button(self, event: tk.Event, button: tk.Canvas, command) -> None:
        redraw = getattr(button, "_lucas_redraw", None)
        if callable(redraw):
            redraw(getattr(button, "_lucas_hover", getattr(button, "_lucas_bg", "#1ed760")))
        width = max(button.winfo_width(), 1)
        height = max(button.winfo_height(), 1)
        if 0 <= event.x <= width and 0 <= event.y <= height:
            command()

    def _show_lucas_settings_menu(self, anchor: tk.Widget) -> None:
        menu = tk.Menu(self, tearoff=False, bg="#1f1f1f", fg="#ffffff", activebackground="#1ed760", activeforeground="#000000")
        menu.add_command(label="Activity Log", command=self.open_activity_log)
        menu.add_command(label="System Health", command=self.open_setup_doctor)
        menu.add_command(label="Working Folder", command=self.choose_working_folder)
        menu.add_separator()
        menu.add_command(label="Recover Sold Ledger", command=self.recover_sold_ledger)
        menu.add_command(label="Export Year-End Report", command=self.export_year_end_report)
        if self._personal_instagram_sync_enabled():
            menu.add_separator()
            menu.add_command(label="Instagram Inventory Sync", command=self.open_instagram_inventory_sync)
        menu.add_separator()
        menu.add_command(label="Mobile Help", command=self.open_mobile_connection_helper)
        try:
            menu.tk_popup(anchor.winfo_rootx(), anchor.winfo_rooty() + anchor.winfo_height())
        finally:
            menu.grab_release()

    def _bind_responsive_button_row(
        self,
        parent: tk.Widget,
        buttons: list[tk.Widget],
        min_button_width: int = 96,
        uniform_columns: bool = False,
    ) -> None:
        state = {"columns": 0}

        def relayout(_event: tk.Event | None = None) -> None:
            live_width = parent.winfo_width()
            width = live_width if live_width > 1 else max(parent.winfo_reqwidth(), min_button_width)
            gap = 8
            columns = max(1, width // (min_button_width + gap))
            columns = min(columns, len(buttons))
            if columns == state["columns"]:
                return
            state["columns"] = columns
            for column in range(max(len(buttons), 1)):
                parent.columnconfigure(column, weight=0, uniform="")
            for row in range((len(buttons) + columns - 1) // columns + 1):
                parent.rowconfigure(row, weight=0)
            for index, button in enumerate(buttons):
                row = index // columns
                column = index % columns
                padx = (0, gap) if column < columns - 1 else (0, 0)
                pady = (0, 6) if row < (len(buttons) - 1) // columns else (0, 0)
                button.grid(row=row, column=column, sticky="ew", padx=padx, pady=pady)
            for column in range(columns):
                parent.columnconfigure(column, weight=1 if uniform_columns else 0, uniform="responsive_buttons" if uniform_columns else "")

        parent.bind("<Configure>", relayout, add="+")
        parent.after_idle(relayout)

    def _build_home_tab_button(self, parent: tk.Frame, text: str, command) -> tk.Label:
        palette = self.home_tab_palette
        button = tk.Label(
            parent,
            text=text,
            bg=palette["soft_button"],
            fg=palette["muted"],
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=0,
            padx=5,
            pady=6,
            font=("Segoe UI Semibold", 8),
            cursor="hand2",
        )
        self._configure_colored_button(button, palette["soft_button"], palette["muted"], palette["soft_button_hover"], palette["border"])
        button.bind("<Enter>", lambda _event: button.configure(bg=getattr(button, "_lucas_hover", palette["soft_button_hover"])), add="+")
        button.bind("<Leave>", lambda _event: button.configure(bg=getattr(button, "_lucas_bg", palette["soft_button"])), add="+")
        button.bind("<ButtonPress-1>", lambda _event: button.configure(bg=getattr(button, "_lucas_pressed", palette["border"])), add="+")
        button.bind("<ButtonRelease-1>", lambda _event: (button.configure(bg=getattr(button, "_lucas_hover", palette["soft_button_hover"])), command()), add="+")
        self._set_tooltip(button)
        return button

    def _set_home_sheet_kind(self, kind: str) -> None:
        self.home_sheet_kind.set(kind)
        self._update_home_sheet_tabs()
        self._refresh_home_sheet_list()

    def _on_home_person_filter_changed(self) -> None:
        person = self.home_person_var.get().strip()
        for var in (self.payout_person_var, self.inventory_person_var, self.profit_person_var):
            if var.get().strip() != person:
                var.set(person)
        self._refresh_home_sheet_list()
        self._refresh_home_metrics()
        self.refresh_payouts_tab()
        self.refresh_inventory_tab()
        self.refresh_profit_tab()

    def _home_person_filter(self) -> str:
        return self.home_person_var.get().strip().lower() if hasattr(self, "home_person_var") else ""

    def _home_sheet_matches_person_filter(self, key: str) -> bool:
        needle = self._home_person_filter()
        if not needle:
            return True
        marker = self.home_sheet_markers.get(key, {})
        person = str(marker.get("assigned_person") or "Unassigned").strip().lower()
        return needle in person

    def _filtered_home_sheet_names(self, kind: str) -> list[str]:
        names = [
            name
            for name in self.home_sheet_paths.get(kind, {})
            if self._home_sheet_matches_person_filter(self._home_sheet_key(kind, name))
        ]
        sorter = getattr(self, "_sorted_home_sheet_names", None)
        return sorter(kind, names) if callable(sorter) else sorted(names, key=lambda name: name.lower())

    def _sorted_home_sheet_names(self, kind: str, names: list[str]) -> list[str]:
        mode = self.home_sheet_sort_var.get().strip().lower() if hasattr(self, "home_sheet_sort_var") else "date created"
        paths = self.home_sheet_paths.get(kind, {})
        if mode == "name":
            return sorted(names, key=lambda name: name.lower())

        def created_time(name: str) -> float:
            path = paths.get(name)
            try:
                return float(path.stat().st_ctime) if path else 0.0
            except Exception:
                return 0.0

        return sorted(names, key=lambda name: (-created_time(name), name.lower()))

    def _update_home_sheet_tabs(self) -> None:
        if not hasattr(self, "home_incoming_tab") or not hasattr(self, "home_working_tab"):
            return
        palette = self.home_tab_palette
        active_kind = self.home_sheet_kind.get()
        active = {"bg": palette["button"], "fg": "#000000", "activebackground": palette["button_hover"], "activeforeground": "#000000"}
        inactive = {"bg": palette["soft_button"], "fg": palette["muted"], "activebackground": palette["soft_button_hover"], "activeforeground": palette["text"]}
        self._set_home_tab_button_state(self.home_incoming_tab, active if active_kind == "Incoming" else inactive)
        self._set_home_tab_button_state(self.home_working_tab, active if active_kind == "Working" else inactive)
        if hasattr(self, "home_received_tab"):
            self._set_home_tab_button_state(self.home_received_tab, active if active_kind == "Received" else inactive)
        if hasattr(self, "home_edit_markers_tab"):
            self._set_home_tab_button_state(self.home_edit_markers_tab, inactive)

    def _set_home_tab_button_state(self, button: tk.Widget, colors: dict[str, str]) -> None:
        self._configure_colored_button(
            button,
            colors["bg"],
            colors["fg"],
            colors.get("activebackground"),
            colors.get("activebackground"),
        )

    def refresh_home(self, reconcile_accounted: bool = True) -> None:
        perf_start = time.perf_counter()
        self.home_sheet_paths = {"Incoming": {}, "Working": {}, "Received": {}}
        self.home_sheet_summaries = {}
        live_summary_paths: list[Path] = []
        errors: list[str] = []
        archived_count = 0
        reconciled_count = 0
        duplicate_warnings: list[str] = []
        duplicate_notices: list[str] = []
        conflict_files = self._shared_conflict_files()
        if conflict_files:
            errors.append(f"Shared conflicts: {', '.join(path.name for path in conflict_files[:3])}")
        try:
            archived = self._archive_eligible_received_sheets()
            if archived:
                archived_count = len(archived)
        except Exception as error:
            errors.append(f"Archive: {error}")
        if reconcile_accounted:
            try:
                reconciliation = self._reconcile_accounted_home_sheets()
                reconciled_count = len(reconciliation.get("moved") or [])
                duplicate_warnings = list(reconciliation.get("warnings") or [])
                duplicate_notices = list(reconciliation.get("notices") or [])
            except Exception as error:
                errors.append(f"Reconcile: {error}")
        for kind, directory in (("Incoming", INCOMING_SHEETS_DIR), ("Working", WORKING_SHEETS_DIR), ("Received", RECEIVED_SHEETS_DIR)):
            try:
                directory.mkdir(parents=True, exist_ok=True)
                paths = sorted(directory.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
            except Exception as error:
                errors.append(f"{kind}: {error}")
                continue
            self.home_sheet_paths[kind] = {path.name: path for path in paths}
            live_summary_paths.extend(paths)
            for path in paths:
                key = self._home_sheet_key(kind, path.name)
                try:
                    summary = self._summarize_home_workbook_cached(path)
                except Exception as error:
                    errors.append(f"{path.name}: {error}")
                    summary = {"name": path.name, "row_count": 0, "received_count": 0, "purchase_total": 0.0, "all_received": False, "partially_received": False}
                summary = self._enrich_home_seller_payout_summary(path, self.home_sheet_markers.get(key, {}), summary)
                self.home_sheet_summaries[key] = summary
        self._prune_home_summary_cache(live_summary_paths)
        self._refresh_home_sheet_list()
        self._refresh_home_metrics()
        self.refresh_payouts_tab()
        self._update_home_sheet_tabs()
        if errors:
            self.status_var.set(f"Home refreshed with {len(errors)} sheet issue(s).")
        elif duplicate_warnings:
            self.status_var.set(f"Home refreshed. {len(duplicate_warnings)} incoming/working sheet issue(s) need review.")
            self._warn_accounted_duplicate_sheets(duplicate_warnings)
        elif duplicate_notices:
            self.status_var.set(f"Home refreshed. {len(duplicate_notices)} partial incoming/working sheet(s) already have some accounted cards; remaining rows can still be received.")
        elif self._seller_warning_count("Working"):
            count = self._seller_warning_count("Working")
            self.status_var.set(f"Home metrics refreshed. {count} working seller sheet(s) need values.")
        elif reconciled_count:
            self.status_var.set(f"Home metrics refreshed. Moved {reconciled_count} fully accounted sheet(s) to RECEIVED SHEETS.")
        elif archived_count:
            self.status_var.set(f"Home metrics refreshed. Archived {archived_count} paid received sheet(s).")
        else:
            self.status_var.set("Home metrics refreshed.")
        total_sheets = sum(len(paths) for paths in self.home_sheet_paths.values())
        record_performance_event(
            "home.refresh",
            perf_start,
            f"sheets={total_sheets} summaries={len(self.home_sheet_summaries)} archived={archived_count} reconciled={reconciled_count} duplicate_warnings={len(duplicate_warnings)} duplicate_notices={len(duplicate_notices)} reconcile_accounted={reconcile_accounted} errors={len(errors)}",
        )

    def _home_summary_cache_key(self, path: Path) -> str:
        return os.path.normcase(str(path.resolve()))

    def _summarize_home_workbook_cached(self, path: Path) -> dict[str, object]:
        stat = path.stat()
        key = self._home_summary_cache_key(path)
        with self.home_summary_cache_lock:
            cached = self.home_summary_cache.get(key)
            if cached and cached.get("mtime_ns") == stat.st_mtime_ns and cached.get("size") == stat.st_size:
                return dict(cached.get("summary") or {})
        summary = summarize_workbook(path)
        with self.home_summary_cache_lock:
            self.home_summary_cache[key] = {"mtime_ns": stat.st_mtime_ns, "size": stat.st_size, "summary": dict(summary)}
        return summary

    def _prune_home_summary_cache(self, live_paths: list[Path]) -> None:
        live_keys = {self._home_summary_cache_key(path) for path in live_paths}
        with self.home_summary_cache_lock:
            for key in list(self.home_summary_cache):
                if key not in live_keys:
                    self.home_summary_cache.pop(key, None)

    def _accounted_source_key(self, value: object) -> str:
        return Path(str(value or "")).name.strip().lower()

    def _accounted_identity_key(self, cert: object = "", item_id: object = "") -> str:
        cert_key = scan_to_cert(cert)
        if cert_key:
            return f"cert:{cert_key}"
        item_key = str(item_id or "").strip().lower()
        if item_key:
            return f"item:{item_key}"
        return ""

    def _add_accounted_identity(self, index: dict[str, set[str]], source_sheet: object, cert: object = "", item_id: object = "") -> None:
        source_key = self._accounted_source_key(source_sheet)
        identity = self._accounted_identity_key(cert, item_id)
        if source_key and identity:
            index.setdefault(source_key, set()).add(identity)

    def _add_accounted_cert(self, index: dict[str, set[str]], source_sheet: object, cert: object) -> None:
        self._add_accounted_identity(index, source_sheet, cert=cert)

    def _accounted_sheet_cert_index(self) -> dict[str, set[str]]:
        index: dict[str, set[str]] = {}
        for record in [self._normalize_inventory_record(row) for row in self._load_inventory_ledger()]:
            self._add_accounted_identity(index, record.get("source_sheet"), record.get("cert_number"), record.get("item_id"))
        for record in [self._normalize_profit_record(row) for row in self._load_profit_ledger()]:
            self._add_accounted_identity(index, record.get("source_sheet"), record.get("cert_number"), record.get("item_id"))
            self._add_accounted_identity(index, record.get("original_source_sheet"), record.get("cert_number"), record.get("item_id"))
        for entry in self._load_activity_log():
            if str(entry.get("action") or "").strip().lower() != "inventory sold":
                continue
            details = entry.get("details") if isinstance(entry.get("details"), dict) else {}
            inventory_key = str(details.get("inventory_key") or "")
            parts = inventory_key.split("|")
            if len(parts) >= 2:
                self._add_accounted_cert(index, parts[1], parts[0])
        return index

    def _sheet_cert_set(self, path: Path) -> set[str]:
        return {
            scan_to_cert(row.get("cert_number"))
            for row in read_simple_spreadsheet(path)
            if scan_to_cert(row.get("cert_number"))
        }

    def _sheet_accounting_payload(self, path: Path) -> dict[str, object]:
        identities: set[str] = set()
        certs: set[str] = set()
        row_refs_by_identity: dict[str, tuple[str, str, int]] = {}
        for row in read_simple_spreadsheet(path):
            identity = self._accounted_identity_key(row.get("cert_number"), row.get("item_id"))
            if not identity:
                continue
            identities.add(identity)
            cert = scan_to_cert(row.get("cert_number"))
            if cert:
                certs.add(cert)
            elif str(row.get("item_id") or "").strip():
                workbook_sheet = str(row.get("workbook_sheet") or "").strip()
                try:
                    workbook_row = int(row.get("workbook_row") or 0)
                except (TypeError, ValueError):
                    workbook_row = 0
                if workbook_sheet and workbook_row > 0:
                    row_refs_by_identity[identity] = (path.name, workbook_sheet, workbook_row)
        return {"identities": identities, "certs": certs, "row_refs_by_identity": row_refs_by_identity}

    def _reconcile_accounted_home_sheets(self) -> dict[str, list[str]]:
        result: dict[str, list[str]] = {"moved": [], "warnings": [], "notices": []}
        if not CARD_PIPELINE_DIR.exists():
            return result
        accounted = self._accounted_sheet_cert_index()
        if not accounted:
            return result
        for stage, directory in (("Incoming", INCOMING_SHEETS_DIR), ("Working", WORKING_SHEETS_DIR)):
            if not directory.exists():
                continue
            for path in sorted(directory.glob("*.xlsx"), key=lambda item: item.name.lower()):
                source_key = self._accounted_source_key(path.name)
                accounted_certs = accounted.get(source_key, set())
                if not accounted_certs:
                    continue
                try:
                    sheet_payload = self._sheet_accounting_payload(path)
                except Exception as error:
                    result["warnings"].append(f"{path.name}: could not inspect duplicate/accounted rows ({error})")
                    continue
                sheet_identities = set(sheet_payload.get("identities") or set())
                if not sheet_identities:
                    continue
                matched = sheet_identities & accounted_certs
                if not matched:
                    continue
                missing = sheet_identities - accounted_certs
                if missing:
                    result["notices"].append(
                        f"{path.name}: {len(matched)}/{len(sheet_identities)} row(s) already exist in inventory/company/sold ledgers; {len(missing)} still unaccounted."
                    )
                    continue
                destination = RECEIVED_SHEETS_DIR / path.name
                if destination.exists():
                    result["warnings"].append(f"{path.name}: all rows accounted, but RECEIVED SHEETS already has that file name.")
                    continue
                try:
                    row_refs_by_identity = dict(sheet_payload.get("row_refs_by_identity") or {})
                    row_refs = {
                        row_ref
                        for identity, row_ref in row_refs_by_identity.items()
                        if identity in sheet_identities
                    }
                    mark_result = mark_received_in_workbooks([path], set(sheet_payload.get("certs") or set()), row_refs)
                    mark_errors = list(mark_result.get("errors") or [])
                    if mark_errors:
                        result["warnings"].append(f"{path.name}: all rows accounted, but receive marks could not be written: {mark_errors[0]}")
                        continue
                    old_key = self._home_sheet_key(stage, path.name)
                    marker = dict(self.home_sheet_markers.get(old_key, {}))
                    marker["all_received"] = True
                    marker.setdefault("received_at", datetime.now().isoformat(timespec="seconds"))
                    new_key = self._move_sheet_to_received(old_key)
                    if not new_key:
                        result["warnings"].append(f"{path.name}: all rows accounted, but L.U.C.A.S could not move the sheet.")
                        continue
                    self.home_sheet_markers[new_key] = marker
                    self._save_sheet_markers()
                    self._append_activity(
                        "Sheet Reconcile",
                        f"Moved fully accounted sheet {path.name} from {stage} to Received.",
                        {"sheet": path.name, "from": stage, "to": "Received", "accounted_rows": len(sheet_identities)},
                    )
                    result["moved"].append(path.name)
                except Exception as error:
                    result["warnings"].append(f"{path.name}: all rows accounted, but reconcile failed: {error}")
        return result

    def _warn_accounted_duplicate_sheets(self, warnings: list[str]) -> None:
        if not warnings:
            return
        seen = getattr(self, "_accounted_duplicate_warning_seen", set())
        fresh = [warning for warning in warnings if warning not in seen]
        if not fresh:
            return
        self._accounted_duplicate_warning_seen = set(seen) | set(fresh)
        messagebox.showwarning(
            "Incoming Sheet Already Accounted",
            "L.U.C.A.S found incoming/working sheet rows that already exist in inventory, company sheets, or sold ledgers.\n\n"
            + "\n".join(fresh[:8])
            + ("\n\nFix: do not re-add duplicate incoming sheets. Move/resolve the sheet before receiving it again." if len(fresh) <= 8 else "\n\nMore warnings were found. Check Home/System Health."),
        )

    def _shared_conflict_files(self) -> list[Path]:
        if not CARD_PIPELINE_DIR.exists():
            return []
        patterns = ("*conflict*.json", "*conflicted*.json", "*copy*.json")
        found: list[Path] = []
        for pattern in patterns:
            found.extend(CARD_PIPELINE_DIR.glob(pattern))
        return sorted({path for path in found if path.name.lower() != PROFIT_LEDGER_PATH.name.lower() and path.name.lower() != SHEET_MARKERS_PATH.name.lower()})

    def _start_startup_refresh(self) -> None:
        self.status_var.set("Loading sheet lists...")
        thread = threading.Thread(target=self._startup_refresh_worker, daemon=True)
        thread.start()

    def _refresh_startup_google_sheet_caches(self) -> dict[str, object]:
        result = {"refreshed": 0, "errors": []}
        sources = self._saved_google_sheet_sources()
        if not sources:
            return result
        try:
            with shared_lock(CARD_PIPELINE_DIR, "google-sheet-cache", self.lucas_identity):
                for source in sources:
                    url = str(source.get("url") or "").strip()
                    path = source.get("path")
                    name = str(source.get("name") or "google-sheet")
                    if not url or not path:
                        continue
                    try:
                        output_path = path_from_source_value(path, ASSIGNMENT_CONFIG_PATH.parent)
                        export_google_sheet_to_xlsx(url, output_path, interactive=False)
                        result["refreshed"] = int(result["refreshed"]) + 1
                    except Exception as error:
                        result["errors"].append(f"{name}: {error}")
        except Exception as error:
            result["errors"].append(str(error))
        return result

    def _refresh_keep_source_registry(self) -> None:
        try:
            self.state.register_keep_note_sources(self._saved_google_keep_sources())
        except Exception:
            return

    def sync_google_keep_notes(self) -> None:
        self._refresh_keep_source_registry()
        sources = self._saved_google_keep_sources()
        if not sources:
            messagebox.showinfo("Google Keep Sync", "No Google Keep note sources are saved in Company Rules.")
            self.assignment_config_status.set("No Google Keep notes are configured in Company Rules.")
            return
        opened = 0
        for source in sources:
            url = str(source.get("url") or "").strip()
            if url and self._open_google_keep_source_url(url):
                opened += 1
        if opened:
            note_word = "note" if opened == 1 else "notes"
            self.assignment_config_status.set(
                f"Opened {opened} Google Keep {note_word}. The Chrome extension will sync each note after it loads."
            )
            return
        messagebox.showinfo("Google Keep Sync", "L.U.C.A.S could not open the saved Google Keep notes.")
        self.assignment_config_status.set("Could not open the saved Google Keep notes.")

    def _open_google_keep_source_url(self, url: str) -> bool:
        if sys.platform == "darwin":
            try:
                result = subprocess.run(
                    ["open", "-a", "Google Chrome", url],
                    check=False,
                    capture_output=True,
                    text=True,
                    timeout=10,
                )
                if result.returncode == 0:
                    return True
            except Exception:
                pass
        return bool(webbrowser.open(url))

    def _saved_google_keep_sources(self) -> list[dict[str, object]]:
        try:
            raw = json.loads(ASSIGNMENT_CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            return []
        entries = raw.get("companies", raw) if isinstance(raw, dict) else raw
        if not isinstance(entries, list):
            return []
        output_dir = CARD_PIPELINE_DIR / "ASSIGNMENT RULES" / "KEEP EXPORTS"
        sources: list[dict[str, object]] = []
        seen: set[tuple[str, str]] = set()
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            for source in (
                entry.get("rules") or entry.get("rules_source") or entry.get("rulesSource"),
                entry.get("payout") or entry.get("payout_source") or entry.get("payoutSource"),
            ):
                prepared = self._google_keep_cache_source(source, output_dir)
                if not prepared:
                    continue
                key = (str(prepared.get("url") or ""), str(prepared.get("path") or ""))
                if key in seen:
                    continue
                seen.add(key)
                sources.append(prepared)
        return sources

    def _google_keep_cache_source(self, source: object, output_dir: Path) -> dict[str, object] | None:
        if isinstance(source, dict):
            url = str(source.get("url") or "").strip()
            path = source.get("path") or source.get("file")
            if str(source.get("kind") or "").strip() == "google_keep" and url:
                name = str(source.get("name") or (Path(str(path)).stem if path else "") or "Google Keep note")
                cache_path = Path(str(path)) if path else output_dir / f"{safe_filename(name)}.txt"
                try:
                    cache_path.relative_to(ASSIGNMENT_CONFIG_PATH.parent)
                    cache_path = output_dir / f"{safe_filename(name)}.txt"
                except ValueError:
                    pass
                return {"url": url, "path": str(cache_path), "name": name}
            return None
        raw = normalize_source_value(source)
        if not is_google_keep_url(raw):
            return None
        name = "Google Keep note"
        return {"url": raw, "path": str(output_dir / f"{safe_filename(name)}.txt"), "name": name}

    def _saved_google_sheet_sources(self) -> list[dict[str, object]]:
        try:
            raw = json.loads(ASSIGNMENT_CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            return []
        entries = raw.get("companies", raw) if isinstance(raw, dict) else raw
        if not isinstance(entries, list):
            return []
        output_dir = CARD_PIPELINE_DIR / "ASSIGNMENT RULES" / "SHEET EXPORTS"
        sources: list[dict[str, object]] = []
        seen: set[tuple[str, str]] = set()
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            for source in (
                entry.get("rules") or entry.get("rules_source") or entry.get("rulesSource"),
                entry.get("payout") or entry.get("payout_source") or entry.get("payoutSource"),
            ):
                prepared = self._google_sheet_cache_source(source, output_dir)
                if not prepared:
                    continue
                key = (str(prepared.get("url") or ""), str(prepared.get("path") or ""))
                if key in seen:
                    continue
                seen.add(key)
                sources.append(prepared)
        return sources

    def _google_sheet_cache_source(self, source: object, output_dir: Path) -> dict[str, object] | None:
        if isinstance(source, dict):
            url = str(source.get("url") or "").strip()
            path = source.get("path") or source.get("file")
            if str(source.get("kind") or "").strip() == "google_sheet" and url:
                name = str(source.get("name") or (Path(str(path)).stem if path else "") or "Google Sheet")
                cache_path = Path(str(path)) if path else output_dir / f"{safe_filename(name)}.xlsx"
                return {"url": url, "path": str(cache_path), "name": name}
            return None
        raw = normalize_source_value(source)
        if not raw:
            return None
        if is_google_sheet_url(raw):
            return {"url": raw, "path": str(output_dir / "Google Sheet.xlsx"), "name": "Google Sheet"}
        path = Path(raw).expanduser()
        if not path.is_absolute():
            path = ASSIGNMENT_CONFIG_PATH.parent / path
        if path.suffix.lower() != ".gsheet":
            return None
        try:
            shortcut = load_gsheet_shortcut(path)
            url = gsheet_shortcut_url(shortcut)
        except Exception:
            return None
        if not url:
            return None
        name = str(shortcut.get("name") or path.stem or "google-sheet")
        return {"url": url, "path": str(output_dir / f"{safe_filename(name)}.xlsx"), "name": name}

    def _startup_refresh_worker(self) -> None:
        perf_start = time.perf_counter()
        payload = {
            "incoming_paths": {},
            "working_paths": {},
            "received_paths": {},
            "incoming_index": {},
            "incoming_path_count": 0,
            "home_paths": {"Incoming": {}, "Working": {}, "Received": {}},
            "home_summaries": {},
            "google_sheet_cache": {"refreshed": 0, "errors": []},
            "errors": [],
        }
        errors: list[str] = payload["errors"]
        google_start = time.perf_counter()
        google_cache_result = self._refresh_startup_google_sheet_caches()
        record_performance_event(
            "startup.google_sheet_cache",
            google_start,
            f"refreshed={google_cache_result.get('refreshed') or 0} errors={len(google_cache_result.get('errors') or [])}",
        )
        payload["google_sheet_cache"] = google_cache_result
        errors.extend(google_cache_result.get("errors") or [])

        try:
            CARD_PIPELINE_DIR.mkdir(parents=True, exist_ok=True)
            WORKING_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
            working_paths = sorted(WORKING_SHEETS_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
            payload["working_paths"] = {path.name: path for path in working_paths}
            payload["home_paths"]["Working"] = {path.name: path for path in working_paths}
        except Exception as error:
            errors.append(f"Working: {error}")
            working_paths = []

        try:
            RECEIVED_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
            received_paths = sorted(RECEIVED_SHEETS_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
            payload["received_paths"] = {path.name: path for path in received_paths}
            payload["home_paths"]["Received"] = {path.name: path for path in received_paths}
        except Exception as error:
            errors.append(f"Received: {error}")
            received_paths = []

        try:
            INCOMING_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
            incoming_paths = sorted(INCOMING_SHEETS_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
            payload["incoming_paths"] = {path.name: path for path in incoming_paths}
            payload["home_paths"]["Incoming"] = {path.name: path for path in incoming_paths}
            payload["incoming_path_count"] = len(incoming_paths)
        except Exception as error:
            errors.append(f"Incoming: {error}")
            incoming_paths = []

        index: dict[str, dict[str, object]] = {}
        incoming_index_start = time.perf_counter()
        index_paths = sorted(
            [*incoming_paths, *working_paths],
            key=lambda path: (path.parent.name.lower(), path.name.lower()),
        )
        for path in index_paths:
            try:
                rows = read_simple_spreadsheet(path)
            except Exception as error:
                errors.append(f"{path.name}: {error}")
                continue
            for row in rows:
                cert = scan_to_cert(row.get("cert_number"))
                candidate = {
                    "sheet": path.name,
                    "path": path,
                    "workbook_sheet": row.get("workbook_sheet") or "",
                    "workbook_row": row.get("workbook_row"),
                    "card_title": row.get("card_title") or "",
                    "grader": row.get("grader") or "",
                    "sport": row.get("sport") or row.get("category") or "",
                    "category": row.get("category") or row.get("sport") or "",
                    "purchase_price": row.get("purchase_price"),
                    "card_ladder_value": row.get("card_ladder_value"),
                    "card_ladder_comps_average": row.get("card_ladder_comps_average"),
                    "cy_value": row.get("cy_value"),
                    "cy_confidence": row.get("cy_confidence"),
                    "card_ladder_comps": row.get("card_ladder_comps") or "",
                    "best_company": row.get("best_company") or "",
                    "estimated_payout": row.get("estimated_payout"),
                }
                receive_key = self._receive_row_ref_key(candidate)
                if receive_key:
                    candidate["receive_key"] = receive_key
                if not cert:
                    if receive_key:
                        index[receive_key] = candidate
                    continue
                existing = index.get(cert)
                if existing:
                    for key, value in candidate.items():
                        if (existing.get(key) is None or existing.get(key) == "") and value not in (None, ""):
                            existing[key] = value
                    continue
                index[cert] = candidate
        payload["incoming_index"] = index
        record_performance_event("startup.incoming_index", incoming_index_start, f"sheets={len(index_paths)} certs={len(index)}")

        summaries_start = time.perf_counter()
        live_summary_paths: list[Path] = []
        for kind, paths in (("Incoming", incoming_paths), ("Working", working_paths), ("Received", received_paths)):
            live_summary_paths.extend(paths)
            for path in paths:
                key = self._home_sheet_key(kind, path.name)
                try:
                    summary = self._summarize_home_workbook_cached(path)
                except Exception as error:
                    errors.append(f"{path.name}: {error}")
                    summary = {"name": path.name, "row_count": 0, "received_count": 0, "purchase_total": 0.0, "all_received": False, "partially_received": False}
                summary = self._enrich_home_seller_payout_summary(path, self.home_sheet_markers.get(key, {}), summary)
                payload["home_summaries"][key] = summary
        self._prune_home_summary_cache(live_summary_paths)
        record_performance_event("startup.home_summaries", summaries_start, f"summaries={len(payload['home_summaries'])}")

        payload["perf_elapsed"] = time.perf_counter() - perf_start
        record_performance_event(
            "startup.worker",
            perf_start,
            f"incoming={len(incoming_paths)} working={len(working_paths)} received={len(received_paths)} certs={len(index)} errors={len(errors)}",
            force=True,
        )
        self.events.put(("startup_refresh", payload))

    def _apply_startup_refresh(self, payload: dict[str, object]) -> None:
        perf_start = time.perf_counter()
        self.incoming_sheet_paths = dict(payload.get("incoming_paths") or {})
        self.working_sheet_paths = dict(payload.get("working_paths") or {})
        self._populate_comp_sheet_list()
        if self.comp_sheet_paths and self.selected_working_sheet.get() not in self.comp_sheet_paths:
            self.selected_working_sheet.set(next(iter(self.comp_sheet_paths)))
        self._select_working_sheet_in_list()

        self.received_sheet_paths = dict(payload.get("received_paths") or {})
        if hasattr(self, "received_sheet_combo"):
            received_names = list(self.received_sheet_paths)
            self.received_sheet_combo["values"] = received_names
            if received_names and self.selected_received_sheet.get() not in self.received_sheet_paths:
                self.selected_received_sheet.set(received_names[0])
            elif not received_names:
                self.selected_received_sheet.set("")

        self.incoming_cert_index = dict(payload.get("incoming_index") or {})
        self._match_all_review_rows()
        self._refresh_table()
        self.review_status.set(f"Indexed {len(self.incoming_cert_index)} cert(s) from {int(payload.get('incoming_path_count') or 0)} incoming sheet(s).")

        self.home_sheet_paths = dict(payload.get("home_paths") or {"Incoming": {}, "Working": {}, "Received": {}})
        self.home_sheet_summaries = dict(payload.get("home_summaries") or {})
        self._refresh_home_sheet_list()
        self._refresh_home_metrics()
        self.refresh_payouts_tab()
        self._update_home_sheet_tabs()
        google_cache = payload.get("google_sheet_cache") if isinstance(payload.get("google_sheet_cache"), dict) else {}
        refreshed_google_sheets = int((google_cache or {}).get("refreshed") or 0)
        self._refresh_keep_source_registry()
        if refreshed_google_sheets:
            self.assignment_engine = AssignmentEngine.load()
            self.assignment_config_status.set(self._assignment_config_status())

        errors = list(payload.get("errors") or [])
        if errors:
            self.status_var.set(f"Startup sheet refresh finished with {len(errors)} issue(s).")
        elif refreshed_google_sheets:
            self.status_var.set(f"Sheet lists loaded. Refreshed {refreshed_google_sheets} Google Sheet cache(s).")
        elif self._seller_warning_count("Working"):
            count = self._seller_warning_count("Working")
            self.status_var.set(f"Sheet lists loaded. {count} working seller sheet(s) need values.")
        else:
            self.status_var.set("Sheet lists loaded.")
        record_performance_event(
            "startup.apply",
            perf_start,
            f"worker_elapsed={float(payload.get('perf_elapsed') or 0):.3f}s home_summaries={len(self.home_sheet_summaries)} incoming_index={len(self.incoming_cert_index)} errors={len(errors)}",
        )

    def _refresh_home_sheet_list(self) -> None:
        if not hasattr(self, "home_sheet_list"):
            return
        kind = self.home_sheet_kind.get()
        self.home_sheet_list.delete(0, tk.END)
        for name in self._filtered_home_sheet_names(kind):
            self.home_sheet_list.insert(tk.END, name)
        if self.home_sheet_list.size():
            self.home_sheet_list.selection_set(0)
            self._load_home_selected_marker()
        else:
            self.home_selected_sheet_key = ""

    def _refresh_home_metrics(self) -> None:
        if not hasattr(self, "incoming_volume_tree"):
            return
        for tree in (self.incoming_volume_tree, self.partial_received_tree):
            tree.delete(*tree.get_children())
        incoming_names = self._filtered_home_sheet_names("Incoming")
        total_cards = 0
        total_received = 0
        total_volume = 0.0
        volume_rows: list[dict[str, object]] = []
        for name in incoming_names:
            key = self._home_sheet_key("Incoming", name)
            summary = self.home_sheet_summaries.get(key, {})
            marker = self.home_sheet_markers.get(key, {})
            total = int(summary.get("row_count") or 0)
            received = int(summary.get("received_count") or 0)
            volume = float(summary.get("purchase_total") or 0.0)
            status = self._incoming_sheet_status(marker, summary)
            total_cards += total
            total_received += received
            total_volume += volume
            volume_rows.append(
                {
                    "sheet": name,
                    "person": str(marker.get("assigned_person") or ""),
                    "cards": total,
                    "received": received,
                    "volume": volume,
                    "status": status,
                }
            )
            if summary.get("partially_received"):
                self.partial_received_tree.insert(
                    "",
                    tk.END,
                    tags=("partial_sheet",),
                    values=(
                        name,
                        f"{received}/{total}",
                        format_money(volume),
                        str(marker.get("assigned_person") or ""),
                        str(marker.get("tracking_number") or ""),
                        "Yes" if marker.get("all_received") else "",
                    ),
                )
        if hasattr(self, "_sorted_records"):
            volume_rows = self._sorted_records(
                volume_rows,
                getattr(self, "home_incoming_volume_sort_column", "sheet"),
                bool(getattr(self, "home_incoming_volume_sort_descending", False)),
                "home_incoming_volume",
            )
        if hasattr(self, "_configure_sortable_tree_headings"):
            self._configure_sortable_tree_headings(
                self.incoming_volume_tree,
                getattr(self, "incoming_volume_headings", {"sheet": "Sheet", "person": "Person", "cards": "Cards", "received": "Received", "volume": "Price Volume", "status": "Status"}),
                "home_incoming_volume",
            )
        for record in volume_rows:
            self.incoming_volume_tree.insert(
                "",
                tk.END,
                values=(
                    record.get("sheet") or "",
                    record.get("person") or "",
                    record.get("cards") or 0,
                    record.get("received") or 0,
                    format_money(record.get("volume")),
                    record.get("status") or "",
                ),
            )
        if incoming_names:
            self.incoming_volume_tree.insert(
                "",
                tk.END,
                tags=("total_divider",),
                values=("━━━━━━", "━━━━━━", "━━━━━━", "━━━━━━", "━━━━━━", "━━━━━━"),
            )
            self.incoming_volume_tree.insert(
                "",
                tk.END,
                tags=("total_row",),
                values=("TOTAL", "", total_cards, total_received, format_money(total_volume), ""),
            )

    def _incoming_sheet_status(self, marker: dict[str, object], summary: dict[str, object]) -> str:
        warning = str(summary.get("seller_payout_warning") or "").strip()
        if warning:
            return warning
        has_tracking = bool(str(marker.get("tracking_number") or "").strip())
        received = int(summary.get("received_count") or 0)
        if has_tracking or received:
            return "Awaiting Receive"
        return "Awaiting tracking"

    def _seller_warning_count(self, stage: str = "") -> int:
        count = 0
        for key, summary in getattr(self, "home_sheet_summaries", {}).items():
            kind, _name = self._split_home_sheet_key(key)
            if stage and kind != stage:
                continue
            if str(summary.get("seller_payout_warning") or "").strip():
                count += 1
        return count

    def _enrich_home_seller_payout_summary(self, path: Path, marker: dict[str, object], summary: dict[str, object]) -> dict[str, object]:
        enriched = dict(summary)
        if not self._sheet_marker_is_seller_payout(marker):
            return enriched
        try:
            rows = self._workbook_rows_from_simple_records(read_simple_spreadsheet(path), source_name=path.name)
        except Exception as error:
            enriched["seller_payout_warning"] = f"Seller payout pending: could not read sheet values ({error})"
            enriched["seller_payout_pending"] = True
            enriched["seller_payout_payable"] = False
            return enriched
        seller_summary = self._seller_payout_summary_for_rows(rows, marker)
        enriched.update(seller_summary)
        return enriched

    def _workbook_rows_from_simple_records(self, records: list[dict[str, object]], source_name: str = "") -> list[WorkbookRow]:
        rows: list[WorkbookRow] = []
        for offset, record in enumerate(records, start=2):
            cert = str(record.get("cert_number") or "")
            grader = str(record.get("grader") or infer_grader(str(record.get("card_title") or "")) or "PSA").upper()
            card = str(record.get("card_title") or "").strip()
            rows.append(
                WorkbookRow(
                    excel_row=offset,
                    cert_number=cert,
                    card_title=card,
                    grader=grader,
                    category=str(record.get("sport") or record.get("category") or "").strip(),
                    existing_value=record.get("purchase_price"),
                    card_ladder_value=record.get("card_ladder_value"),
                    card_ladder_comps_average=record.get("card_ladder_comps_average"),
                    cy_value=record.get("cy_value"),
                    cy_confidence=record.get("cy_confidence"),
            card_ladder_comps=str(record.get("card_ladder_comps") or ""),
            best_company=str(record.get("best_company") or ""),
            estimated_payout=record.get("estimated_payout"),
            received=bool(record.get("received")),
            status=str(record.get("status") or ("Ready" if cert and grader else "Needs setup")),
            notes=str(record.get("notes") or source_name or ""),
        )
            )
        return rows

    def _seller_term_for_marker(self, marker: dict[str, object]) -> dict[str, object] | None:
        if not self._sheet_marker_is_seller_payout(marker):
            return None
        seller = str(marker.get("assigned_person") or "").strip()
        sheet_type = str(marker.get("seller_sheet_type") or "").strip()
        if not sheet_type:
            return None
        rate = self._seller_terms_rate(marker.get("seller_rate"))
        deduction = self._seller_terms_rate(marker.get("seller_deduction"))
        term = self._seller_terms_match(seller, sheet_type) if seller else None
        if term:
            if rate is None:
                rate = self._seller_terms_rate(term.get("rate"))
            if deduction is None:
                deduction = self._seller_terms_rate(term.get("deduction"))
        if rate is None and deduction is None:
            return None
        return {"seller": seller, "sheet_type": sheet_type, "rate": rate, "deduction": deduction}

    def _seller_terms_value_label(self, sheet_type: str, deduction: float | None = None) -> str:
        if deduction is not None:
            return f"{sheet_type} payout"
        company_key = str(sheet_type or "").strip().lower()
        for company in getattr(self.assignment_engine, "companies", []):
            if str(getattr(company, "name", "") or "").strip().lower() != company_key:
                continue
            value_source = str(getattr(company, "value_source", "") or "").strip().lower()
            if value_source == "card_ladder":
                return "Card Ladder value"
            if value_source == "cy_estimate":
                return "CY Estimate"
            return "Comps"
        return "Comps/Card Ladder/CY value"

    def _seller_payout_summary_for_rows(self, rows: list[WorkbookRow], marker: dict[str, object]) -> dict[str, object]:
        term = self._seller_term_for_marker(marker)
        if not term:
            return {"seller_payout_total": 0.0, "seller_payout_ready_count": 0, "seller_payout_missing_count": len(rows), "seller_payout_pending": bool(rows), "seller_payout_payable": False}
        sheet_type = str(term.get("sheet_type") or "").strip()
        rate = self._seller_terms_rate(term.get("rate"))
        deduction = self._seller_terms_rate(term.get("deduction"))
        payout_total = 0.0
        ready_count = 0
        missing_count = 0
        for row in rows:
            seller_price = (
                self._seller_terms_company_price(row, sheet_type, deduction=deduction)
                if deduction is not None
                else self._seller_terms_company_price(row, sheet_type, rate=rate)
            )
            if seller_price is None:
                missing_count += 1
                continue
            payout_total += seller_price
            ready_count += 1
        value_label = self._seller_terms_value_label(sheet_type, deduction=deduction)
        pending = missing_count > 0
        warning = ""
        if pending:
            warning = f"Seller owed money but no {value_label} input to determine seller payout"
            if ready_count:
                warning = f"{warning} ({missing_count} row(s) pending)"
        return {
            "seller_payout_total": round(payout_total, 2),
            "seller_payout_ready_count": ready_count,
            "seller_payout_missing_count": missing_count,
            "seller_payout_pending": pending,
            "seller_payout_payable": bool(rows) and not pending and ready_count > 0,
            "seller_payout_warning": warning,
            "seller_payout_value_label": value_label,
        }

    def refresh_payouts_tab(self) -> None:
        perf_start = time.perf_counter()
        if not hasattr(self, "payout_summary_tree"):
            record_performance_event("payouts.refresh", perf_start, "tree=missing")
            return
        self._refresh_person_combo_values()
        self.payout_summary_tree.delete(*self.payout_summary_tree.get_children())
        self.payout_detail_tree.delete(*self.payout_detail_tree.get_children())
        self.payout_summary_people = {}
        self.payout_detail_keys = {}

        balances: dict[str, dict[str, float | int]] = {}
        detail_count = 0
        filter_person = self.payout_person_var.get().strip().lower()
        for item in self._payout_sheet_items():
            person = item["person"] or "Unassigned"
            if filter_person and filter_person not in person.lower():
                continue
            balance = balances.setdefault(
                person,
                {
                    "sheets": 0,
                    "cards": 0,
                    "expenses": 0.0,
                    "total_net_profit": 0.0,
                    "unpaid_sheets": 0,
                    "unpaid_cards": 0,
                    "unpaid_expenses": 0.0,
                    "unpaid_net_profit": 0.0,
                    "balance": 0.0,
                },
            )
            balance["sheets"] = int(balance["sheets"]) + 1
            balance["cards"] = int(balance["cards"]) + int(item["row_count"])
            balance["expenses"] = float(balance["expenses"]) + float(item.get("expense_total") or 0.0)
            balance["total_net_profit"] = float(balance["total_net_profit"]) + float(item.get("net_profit_total") or 0.0)
            if not item["paid"] and item.get("payable", True):
                balance["unpaid_sheets"] = int(balance["unpaid_sheets"]) + 1
                balance["unpaid_cards"] = int(balance["unpaid_cards"]) + int(item["row_count"])
                balance["unpaid_expenses"] = float(balance["unpaid_expenses"]) + float(item.get("expense_total") or 0.0)
                balance["unpaid_net_profit"] = float(balance["unpaid_net_profit"]) + float(item.get("net_profit_total") or 0.0)
                balance["balance"] = float(balance["balance"]) + float(item["payout_balance"])
            iid = f"payout:{detail_count}"
            self.payout_detail_keys[iid] = str(item["key"])
            self.payout_detail_tree.insert(
                "",
                tk.END,
                iid=iid,
                values=(
                    item["name"],
                    item["stage"],
                    item["person"],
                    item["row_count"],
                    f"{item['received_count']}/{item['row_count']}",
                    format_money(float(item["payout_balance"])),
                    item["status"],
                ),
            )
            detail_count += 1

        for index, (person, values) in enumerate(sorted(balances.items(), key=lambda pair: (-float(pair[1]["balance"]), pair[0].lower()))):
            iid = f"payout-summary:{index}"
            self.payout_summary_people[iid] = person
            self.payout_summary_tree.insert(
                "",
                tk.END,
                iid=iid,
                values=(
                    person,
                    int(values["sheets"]),
                    int(values["cards"]),
                    format_money(float(values["expenses"])),
                    format_money(float(values["total_net_profit"])),
                    format_money(float(values["unpaid_net_profit"])),
                    format_money(float(values["balance"])),
                ),
            )

        total_balance = sum(float(values["balance"]) for values in balances.values())
        total_sheets = sum(int(values["sheets"]) for values in balances.values())
        total_cards = sum(int(values["cards"]) for values in balances.values())
        total_expenses = sum(float(values["expenses"]) for values in balances.values())
        total_net_profit = sum(float(values["total_net_profit"]) for values in balances.values())
        total_unpaid_net_profit = sum(float(values["unpaid_net_profit"]) for values in balances.values())
        if balances:
            self.payout_summary_tree.insert(
                "",
                tk.END,
                tags=("total_divider",),
                values=("------", "------", "------", "------", "------", "------", "------"),
            )
            self.payout_summary_tree.insert(
                "",
                tk.END,
                tags=("total_row",),
                values=("TOTAL", total_sheets, total_cards, format_money(total_expenses), format_money(total_net_profit), format_money(total_unpaid_net_profit), format_money(total_balance)),
            )
        filter_label = self.payout_person_var.get().strip()
        suffix = f" | Filter: {filter_label}" if filter_label else ""
        self.payout_status_var.set(f"{detail_count} payment sheet(s) | Active balance: {format_money(total_balance)}{suffix}")
        record_performance_event(
            "payouts.refresh",
            perf_start,
            f"details={detail_count} people={len(balances)} total_net={total_net_profit:.2f} unpaid_net={total_unpaid_net_profit:.2f} expenses={total_expenses:.2f} balance={total_balance:.2f}",
        )

    def _payout_sheet_items(self) -> list[dict[str, object]]:
        items: list[dict[str, object]] = []
        seller_names = self._seller_terms_seller_names()
        realized_profit_groups = self._realized_profit_groups_by_person_sheet()
        loose_expense_groups = self._loose_expense_adjustments_by_person()
        for stage in ("Incoming", "Received"):
            for name in self.home_sheet_paths.get(stage, {}):
                key = self._home_sheet_key(stage, name)
                marker = self.home_sheet_markers.get(key, {})
                summary = self.home_sheet_summaries.get(key, {})
                paid = bool(marker.get("paid"))
                row_count = int(summary.get("row_count") or 0)
                received_count = int(summary.get("received_count") or 0)
                if stage == "Received":
                    received_count = int(summary.get("received_count") or row_count)
                status = "Paid" if paid else self._payout_sheet_status(stage, marker, summary)
                person = str(marker.get("assigned_person") or "").strip()
                person_key = person.lower()
                is_seller_payout = self._sheet_marker_is_seller_payout(marker) or bool(person_key and person_key in seller_names)
                if not is_seller_payout or stage != "Received":
                    continue
                purchase_total = float(summary.get("purchase_total") or 0.0)
                estimated_payout_total = float(summary.get("estimated_payout_total") or 0.0)
                realized_profit_total = float(realized_profit_groups.get((person.lower(), Path(name).name.lower()), {}).get("profit") or 0.0)
                has_live_seller_summary = "seller_payout_total" in summary or "seller_payout_pending" in summary
                seller_payable = bool(summary.get("seller_payout_payable")) if has_live_seller_summary else True
                seller_pending = bool(summary.get("seller_payout_pending")) if has_live_seller_summary else False
                if self._sheet_marker_is_seller_payout(marker) and has_live_seller_summary:
                    payout_balance = round(float(summary.get("seller_payout_total") or 0.0), 2)
                    value_label = str(summary.get("seller_payout_value_label") or "seller terms")
                    payout_basis = f"Seller terms from {value_label}"
                    if seller_pending:
                        status = str(summary.get("seller_payout_warning") or "Seller payout pending values")
                else:
                    payout_balance, payout_basis = self._active_payout_balance(
                        person,
                        purchase_total,
                        estimated_payout_total,
                        seller_names,
                        realized_profit_total=realized_profit_total,
                        seller_payout=True,
                    )
                    seller_payable = True
                items.append(
                    {
                        "key": key,
                        "stage": stage,
                        "name": name,
                        "person": person,
                        "paid": paid,
                        "row_count": row_count,
                        "received_count": received_count,
                        "purchase_total": purchase_total,
                        "estimated_payout_total": estimated_payout_total,
                        "estimated_profit": round(estimated_payout_total - purchase_total, 2),
                        "realized_profit_total": round(realized_profit_total, 2),
                        "expense_total": 0.0,
                        "net_profit_total": round(realized_profit_total, 2),
                        "payout_balance": payout_balance,
                        "payout_basis": payout_basis,
                        "payable": seller_payable,
                        "status": status,
                    }
                )
        for (person_key, source_key), group in sorted(realized_profit_groups.items(), key=lambda pair: (pair[0][0], pair[0][1])):
            if self._source_sheet_is_seller_payout(str(group.get("source_sheet") or ""), str(group.get("person") or ""), seller_names):
                continue
            person = str(group.get("person") or "").strip() or "Unassigned"
            source_sheet = str(group.get("source_sheet") or "").strip() or "Sold Cards"
            key = self._sold_payout_key(person, source_sheet)
            marker = self.home_sheet_markers.get(key, {})
            paid_groups = self._payout_realized_groups_for_marker(person, source_sheet, group, marker)
            for paid, payout_group in paid_groups:
                realized_profit_total = float(payout_group.get("profit") or 0.0)
                if realized_profit_total == 0:
                    continue
                payout_balance, payout_basis = self._active_payout_balance(
                    person,
                    float(payout_group.get("purchase_total") or 0.0),
                    float(payout_group.get("sale_total") or 0.0),
                    seller_names,
                    realized_profit_total=realized_profit_total,
                    seller_payout=False,
                )
                items.append(
                    {
                        "key": key,
                        "stage": "Sold",
                        "name": source_sheet,
                        "person": person,
                        "paid": paid,
                        "paid_at": str(marker.get("paid_at") or "") if paid else "",
                        "row_count": int(payout_group.get("row_count") or 0),
                        "received_count": int(payout_group.get("row_count") or 0),
                        "purchase_total": float(payout_group.get("purchase_total") or 0.0),
                        "estimated_payout_total": float(payout_group.get("sale_total") or 0.0),
                        "estimated_profit": round(float(payout_group.get("sale_total") or 0.0) - float(payout_group.get("purchase_total") or 0.0), 2),
                        "realized_profit_total": round(realized_profit_total, 2),
                        "expense_total": round(float(payout_group.get("expense_total") or 0.0), 2),
                        "net_profit_total": round(realized_profit_total, 2),
                        "payout_balance": payout_balance,
                        "payout_basis": payout_basis,
                        "status": "Paid" if paid else "Sold",
                    }
                )
        for person_key, group in sorted(loose_expense_groups.items(), key=lambda pair: pair[1]["person"].lower()):
            person = str(group.get("person") or "").strip() or "Unassigned"
            source_sheet = "Expense Adjustments"
            key = self._sold_payout_key(person, source_sheet)
            marker = self.home_sheet_markers.get(key, {})
            paid_groups = self._payout_realized_groups_for_marker(person, source_sheet, group, marker)
            for paid, payout_group in paid_groups:
                net_profit_total = float(payout_group.get("profit") or 0.0)
                if net_profit_total == 0:
                    continue
                payout_balance = min(0.0, round(net_profit_total / 2.0, 2))
                items.append(
                    {
                        "key": key,
                        "stage": "Sold",
                        "name": source_sheet,
                        "person": person,
                        "paid": paid,
                        "paid_at": str(marker.get("paid_at") or "") if paid else "",
                        "row_count": 0,
                        "received_count": 0,
                        "purchase_total": 0.0,
                        "estimated_payout_total": 0.0,
                        "estimated_profit": round(net_profit_total, 2),
                        "realized_profit_total": round(net_profit_total, 2),
                        "expense_total": round(float(payout_group.get("expense_total") or 0.0), 2),
                        "net_profit_total": round(net_profit_total, 2),
                        "payout_balance": payout_balance,
                        "payout_basis": "Expense adjustment to team net profit",
                        "status": "Paid" if paid else "Expenses",
                    }
                )
        return items

    def _realized_profit_totals_by_person_sheet(self) -> dict[tuple[str, str], float]:
        return {
            key: float(group.get("profit") or 0.0)
            for key, group in self._realized_profit_groups_by_person_sheet().items()
        }

    def _profit_record_payout_time(self, record: dict[str, object]) -> datetime | None:
        for field in ("ledger_added_at", "created_at", "recorded_at"):
            text = str(record.get(field) or "").strip()
            if not text:
                continue
            try:
                return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
            except ValueError:
                continue
        date_text = str(record.get("date_added") or "").strip()[:10]
        if date_text:
            try:
                return datetime.strptime(date_text, "%Y-%m-%d") + timedelta(hours=23, minutes=59, seconds=59)
            except ValueError:
                return None
        return None

    def _empty_realized_profit_group(self, person: str, source_sheet: str) -> dict[str, object]:
        return {
            "person": person,
            "source_sheet": source_sheet,
            "row_count": 0,
            "purchase_total": 0.0,
            "sale_total": 0.0,
            "expense_total": 0.0,
            "profit": 0.0,
            "records": [],
        }

    def _add_profit_record_to_realized_group(self, group: dict[str, object], record: dict[str, object]) -> None:
        profit = self._money_value(record.get("profit"))
        if profit is None:
            return
        is_expense = str(record.get("record_type") or "").strip().lower() == "expense"
        purchase = self._money_value(record.get("purchase_price")) or 0.0
        sale = self._money_value(record.get("sale_price")) or 0.0
        if is_expense:
            group["expense_total"] = float(group["expense_total"]) + abs(profit)
        else:
            group["row_count"] = int(group["row_count"]) + 1
            group["purchase_total"] = float(group["purchase_total"]) + purchase
            group["sale_total"] = float(group["sale_total"]) + sale
        group["profit"] = float(group["profit"]) + profit
        records = group.setdefault("records", [])
        if isinstance(records, list):
            records.append(record)

    def _payout_realized_groups_for_marker(
        self,
        person: str,
        source_sheet: str,
        group: dict[str, object],
        marker: dict[str, object],
    ) -> list[tuple[bool, dict[str, object]]]:
        if not marker.get("paid"):
            return [(False, group)]
        paid_at_text = str(marker.get("paid_at") or "").strip()
        if not paid_at_text:
            return [(True, group)]
        try:
            paid_at = datetime.fromisoformat(paid_at_text.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            return [(True, group)]
        paid_group = self._empty_realized_profit_group(person, source_sheet)
        open_group = self._empty_realized_profit_group(person, source_sheet)
        records = group.get("records")
        if not isinstance(records, list):
            return [(True, group)]
        for record in records:
            record_time = self._profit_record_payout_time(record)
            target = open_group if record_time and record_time > paid_at else paid_group
            self._add_profit_record_to_realized_group(target, record)
        result: list[tuple[bool, dict[str, object]]] = []
        if float(paid_group.get("profit") or 0.0):
            result.append((True, paid_group))
        if float(open_group.get("profit") or 0.0):
            result.append((False, open_group))
        return result or [(True, group)]

    def _realized_profit_groups_by_person_sheet(self) -> dict[tuple[str, str], dict[str, object]]:
        totals: dict[tuple[str, str], float] = defaultdict(float)
        groups: dict[tuple[str, str], dict[str, object]] = {}
        for record in self._enrich_profit_records_with_people(self._load_profit_ledger()):
            person = str(record.get("assigned_person") or "").strip()
            source_sheet = Path(str(record.get("source_sheet") or "")).name.strip()
            profit = self._money_value(record.get("profit"))
            if not person or not source_sheet or profit is None:
                continue
            is_expense = str(record.get("record_type") or "").strip().lower() == "expense"
            if is_expense:
                related_type = str(record.get("related_type") or "").strip()
                if source_sheet.lower() == "expenses" or (related_type and related_type not in {"Card", "Sheet"}):
                    continue
            key = (person.lower(), source_sheet.lower())
            group = groups.setdefault(
                key,
                self._empty_realized_profit_group(person, source_sheet),
            )
            purchase = self._money_value(record.get("purchase_price")) or 0.0
            sale = self._money_value(record.get("sale_price")) or 0.0
            if is_expense:
                group["expense_total"] = float(group["expense_total"]) + abs(profit)
            else:
                group["row_count"] = int(group["row_count"]) + 1
                group["purchase_total"] = float(group["purchase_total"]) + purchase
                group["sale_total"] = float(group["sale_total"]) + sale
            totals[key] += profit
            group["profit"] = totals[key]
            records = group.setdefault("records", [])
            if isinstance(records, list):
                records.append(record)
        return groups

    def _loose_expense_adjustments_by_person(self) -> dict[str, dict[str, object]]:
        groups: dict[str, dict[str, object]] = {}
        for record in self._enrich_profit_records_with_people(self._load_profit_ledger()):
            if str(record.get("record_type") or "").strip().lower() != "expense":
                continue
            person = str(record.get("assigned_person") or "").strip()
            source_sheet = Path(str(record.get("source_sheet") or "")).name.strip()
            related_type = str(record.get("related_type") or "").strip()
            is_loose = source_sheet.lower() == "expenses" or not source_sheet or (related_type and related_type not in {"Card", "Sheet"})
            if not person or not is_loose:
                continue
            profit = self._money_value(record.get("profit"))
            amount = abs(profit if profit is not None else (self._money_value(record.get("expense_amount")) or 0.0))
            if amount <= 0:
                continue
            key = person.lower()
            group = groups.setdefault(key, self._empty_realized_profit_group(person, "Expense Adjustments"))
            group["expense_total"] = float(group["expense_total"]) + amount
            group["profit"] = float(group["profit"]) - amount
            records = group.setdefault("records", [])
            if isinstance(records, list):
                records.append(record)
        return groups

    def _sold_payout_key(self, person: str, source_sheet: str) -> str:
        return self._home_sheet_key("Sold", f"{str(person or '').strip()}|{Path(str(source_sheet or '')).name}")

    def _split_sold_payout_name(self, name: str) -> tuple[str, str]:
        if "|" not in name:
            return "", name
        person, source_sheet = name.split("|", 1)
        return person, source_sheet

    def _seller_terms_seller_names(self) -> set[str]:
        return {
            str(term.get("seller") or "").strip().lower()
            for term in self._load_seller_terms()
            if str(term.get("seller") or "").strip()
        }

    def _sheet_marker_is_seller_payout(self, marker: dict[str, object]) -> bool:
        return bool(marker.get("seller_terms_applied") or marker.get("seller_sheet_type"))

    def _source_sheet_is_seller_payout(self, source_sheet: str, person: str = "", seller_names: set[str] | None = None) -> bool:
        source_name = Path(str(source_sheet or "")).name.lower()
        if not source_name:
            return False
        saw_marker = False
        for stage in ("Incoming", "Received", "Working"):
            target_key = self._home_sheet_key(stage, source_name).lower()
            marker = self.home_sheet_markers.get(self._home_sheet_key(stage, source_name), {})
            if not marker:
                marker = next((candidate for key, candidate in self.home_sheet_markers.items() if str(key).lower() == target_key), {})
            if marker:
                saw_marker = True
                if self._sheet_marker_is_seller_payout(marker):
                    return True
        if saw_marker:
            return False
        seller_names = seller_names if seller_names is not None else self._seller_terms_seller_names()
        return bool(str(person or "").strip().lower() in seller_names)

    def _active_payout_balance(
        self,
        person: str,
        purchase_total: float,
        estimated_payout_total: float,
        seller_names: set[str] | None = None,
        realized_profit_total: float | None = None,
        seller_payout: bool | None = None,
    ) -> tuple[float, str]:
        normalized_person = str(person or "").strip().lower()
        seller_names = seller_names if seller_names is not None else self._seller_terms_seller_names()
        if seller_payout is True or (seller_payout is None and normalized_person and normalized_person in seller_names):
            return round(float(purchase_total or 0.0), 2), "Seller purchase total"
        realized_profit = float(realized_profit_total or 0.0)
        return round(realized_profit / 2.0, 2), "Team half sold profit"

    def _payout_sheet_status(self, stage: str, marker: dict[str, object], summary: dict[str, object]) -> str:
        received_count = int(summary.get("received_count") or 0)
        if stage == "Received" or marker.get("all_received") or summary.get("all_received"):
            return "Unpaid"
        if received_count:
            return "Partially Received"
        return "Unreceived"

    def _network_mode_enabled(self) -> bool:
        var = getattr(self, "create_network_mode_var", None)
        return bool(var.get()) if var is not None else False

    def _set_create_network_controls_visible(self, visible: bool) -> None:
        widgets = (
            getattr(self, "network_seller_label", None),
            getattr(self, "seller_terms_seller_combo", None),
            getattr(self, "network_sheet_type_label", None),
            getattr(self, "seller_terms_sheet_type_combo", None),
        )
        for widget in widgets:
            if widget is None:
                continue
            if visible:
                widget.grid()
            else:
                widget.grid_remove()

    def _toggle_create_network_mode(self) -> None:
        enabled = self._network_mode_enabled()
        settings = load_app_settings()
        settings["network_mode"] = enabled
        save_app_settings(settings)
        self.app_settings = settings
        self._set_create_network_controls_visible(enabled)
        if not enabled:
            restored = self._restore_create_seller_term_prices()
            if restored:
                self._refresh_table()
            self.status_var.set("Network Mode off. Seller terms hidden.")
        else:
            self.status_var.set("Network Mode on. Seller and Sheet Type are available in Create.")
            self.apply_create_seller_terms(show_status=False)

    def _load_seller_terms(self) -> list[dict[str, object]]:
        if not SELLER_TERMS_PATH.exists():
            return []
        try:
            with SELLER_TERMS_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
        except Exception:
            return []
        terms: list[dict[str, object]] = []
        for row in rows:
            normalized = {re.sub(r"[^a-z0-9]+", "", str(key or "").lower()): value for key, value in row.items()}
            seller = str(normalized.get("seller") or normalized.get("person") or normalized.get("name") or "").strip()
            sheet_type = str(normalized.get("sheettype") or normalized.get("type") or normalized.get("company") or "").strip()
            value_source = str(normalized.get("valuesource") or normalized.get("source") or "").strip()
            rate = self._seller_terms_rate(normalized.get("sellerrate") or normalized.get("rate") or normalized.get("payout") or normalized.get("percentage"))
            deduction = self._seller_terms_rate(normalized.get("deduction") or normalized.get("sellerdeduction") or normalized.get("deductionpercent") or normalized.get("deductionpercentage"))
            if seller and sheet_type and (rate is not None or deduction is not None):
                terms.append({"seller": seller, "sheet_type": sheet_type, "value_source": value_source, "rate": rate, "deduction": deduction})
        return terms

    def _refresh_seller_terms_dropdowns(self) -> None:
        if hasattr(self, "seller_terms_sheet_type_combo"):
            sheet_types = sorted({str(term.get("sheet_type") or "") for term in self._load_seller_terms() if term.get("sheet_type")}, key=str.lower)
            self.seller_terms_sheet_type_combo["values"] = sheet_types

    def _seller_terms_rate(self, value: object) -> float | None:
        raw = str(value or "").strip()
        text = raw.replace("%", "").strip()
        if not text:
            return None
        try:
            numeric = float(text)
        except ValueError:
            return None
        rate = numeric / 100 if "%" in raw or numeric > 1 else numeric
        return rate if rate >= 0 else None

    def _seller_terms_match(self, seller: str, sheet_type: str) -> dict[str, object] | None:
        seller_key = seller.strip().lower()
        type_key = sheet_type.strip().lower()
        if not seller_key or not type_key:
            return None
        for term in self._load_seller_terms():
            if str(term.get("seller") or "").strip().lower() == seller_key and str(term.get("sheet_type") or "").strip().lower() == type_key:
                return term
        return None

    def _seller_terms_company_decision(self, row: WorkbookRow, company_name: str):
        company_key = company_name.strip().lower()
        decisions = list(self.assignment_engine.evaluate(row))
        if not company_key:
            return None, decisions
        for decision in decisions:
            if decision.company.strip().lower() == company_key:
                return decision, decisions
        return None, decisions

    def _seller_terms_company_price(self, row: WorkbookRow, company_name: str, rate: float | None = None, deduction: float | None = None) -> float | None:
        decision, _decisions = self._seller_terms_company_decision(row, company_name)
        if decision is None:
            return None
        if decision.source_value is None:
            return None
        if deduction is not None:
            if decision.payout is None:
                return None
            return max(0.0, round(decision.payout - (decision.source_value * deduction), 2))
        if rate is not None:
            if not getattr(decision, "accepted", False):
                return None
            return round(decision.source_value * rate, 2)
        return None

    def _seller_terms_no_match_details(self, rows: list[WorkbookRow], company_name: str, limit: int = 5) -> str:
        details: list[str] = []
        for row in rows[:limit]:
            decision, decisions = self._seller_terms_company_decision(row, company_name)
            label_parts = [str(row.cert_number or "").strip(), str(row.card_title or "").strip()]
            label = " / ".join(part for part in label_parts if part) or f"row {getattr(row, 'excel_row', '') or '?'}"
            if not decisions:
                details.append(f"{label}: no assignment companies are loaded.")
            elif decision is None:
                details.append(f"{label}: {company_name} is not an active/loaded assignment company.")
            elif decision.source_value is None:
                details.append(f"{label}: {decision.reason or 'missing value for this company'}.")
            elif decision.payout is None:
                source = format_money(decision.source_value) if decision.source_value is not None else "no source value"
                details.append(f"{label}: {decision.reason or 'no payout'} ({source}).")
            elif not decision.accepted:
                details.append(f"{label}: {decision.reason or 'not accepted by company rules'}.")
            else:
                details.append(f"{label}: no seller price was produced.")
        if len(rows) > limit:
            details.append(f"...and {len(rows) - limit} more row(s).")
        return "\n".join(details)

    def _seller_terms_no_match_message(self, rows: list[WorkbookRow], seller_sheet_type: str, deduction: float | None) -> str:
        basis = "actual payout" if deduction is not None else "source value"
        details = self._seller_terms_no_match_details(rows, seller_sheet_type)
        message = (
            f"No Create rows produced a {seller_sheet_type} seller payout. "
            f"L.U.C.A.S found the People Rules row, but this Sheet Type needs a {seller_sheet_type} {basis} on at least one card."
        )
        if details:
            message = f"{message}\n\nRow reasons:\n{details}"
        return message

    def _restore_create_seller_term_prices(self) -> int:
        restored = 0
        for row in self.intake_rows:
            if hasattr(row, "_seller_terms_base_purchase"):
                base_value = getattr(row, "_seller_terms_base_purchase")
                if row.existing_value != base_value:
                    row.existing_value = base_value
                    restored += 1
        return restored

    def apply_create_seller_terms(self, show_status: bool = True) -> int:
        if not self._network_mode_enabled():
            restored = self._restore_create_seller_term_prices()
            if restored:
                self._refresh_table()
            return 0
        seller = self.seller_terms_seller_var.get().strip() if hasattr(self, "seller_terms_seller_var") else ""
        sheet_type = self.seller_terms_sheet_type_var.get().strip() if hasattr(self, "seller_terms_sheet_type_var") else ""
        if not seller or not sheet_type:
            restored = self._restore_create_seller_term_prices()
            if restored:
                self._refresh_table()
            if show_status and (seller or sheet_type):
                self.status_var.set("Seller terms need both Seller and Sheet Type.")
            return 0
        term = self._seller_terms_match(seller, sheet_type)
        if not term:
            restored = self._restore_create_seller_term_prices()
            if restored:
                self._refresh_table()
            if show_status:
                self.status_var.set(f"No seller terms found for {seller} / {sheet_type}.")
            return 0
        rate = self._money_value(term.get("rate"))
        deduction = self._money_value(term.get("deduction"))
        if deduction is None and rate is None:
            if show_status:
                self.status_var.set(f"Seller terms for {seller} / {sheet_type} need either a Deduction or Seller Rate.")
            return 0
        changed = 0
        skipped = 0
        for row in self.intake_rows:
            if not hasattr(row, "_seller_terms_base_purchase"):
                setattr(row, "_seller_terms_base_purchase", row.existing_value)
            if deduction is not None:
                seller_price = self._seller_terms_company_price(row, sheet_type, deduction=deduction)
            else:
                seller_price = self._seller_terms_company_price(row, sheet_type, rate=rate)
            if seller_price is None:
                skipped += 1
                continue
            if row.existing_value != seller_price:
                row.existing_value = seller_price
                changed += 1
        if changed:
            self._refresh_table()
        if show_status:
            if deduction is not None:
                suffix = f" ({skipped} skipped: no matching {sheet_type} payout)" if skipped else ""
                self.status_var.set(f"Applied seller terms: {seller} / {sheet_type} payout minus {deduction:.0%}.{suffix}")
            else:
                suffix = f" ({skipped} skipped: no matching {sheet_type} source value)" if skipped else ""
                self.status_var.set(f"Applied seller terms: {seller} / {sheet_type} at {rate:.0%} of {sheet_type} rule value.{suffix}")
        return changed

    def _known_assigned_people(self) -> list[str]:
        people = {
            str(marker.get("assigned_person") or "").strip()
            for marker in self.home_sheet_markers.values()
            if CardPipelineApp._is_real_person_name(marker.get("assigned_person"))
        }
        return sorted(people, key=str.lower)

    @staticmethod
    def _is_real_person_name(value: object) -> bool:
        person = str(value or "").strip()
        return bool(person and person.lower() != "unassigned")

    def _known_people(self) -> list[str]:
        people_set = {
            person
            for person in self._known_assigned_people()
            if CardPipelineApp._is_real_person_name(person)
        }
        people_set.update(
            str(term.get("seller") or "").strip()
            for term in self._load_seller_terms()
            if CardPipelineApp._is_real_person_name(term.get("seller"))
        )
        if hasattr(self, "profit_rows"):
            people_set.update(
                str(record.get("assigned_person") or "").strip()
                for record in self.profit_rows
                if CardPipelineApp._is_real_person_name(record.get("assigned_person"))
            )
        if hasattr(self, "inventory_rows"):
            people_set.update(
                str(record.get("assigned_person") or "").strip()
                for record in self.inventory_rows
                if CardPipelineApp._is_real_person_name(record.get("assigned_person"))
            )
        for record in self._load_profit_ledger():
            person = str(record.get("assigned_person") or record.get("person") or "").strip()
            if CardPipelineApp._is_real_person_name(person):
                people_set.add(person)
        for record in self._load_inventory_ledger():
            person = str(record.get("assigned_person") or record.get("person") or "").strip()
            if CardPipelineApp._is_real_person_name(person):
                people_set.add(person)
        return sorted(people_set, key=str.lower)

    def _refresh_person_combo_values(self, filter_text: str = "") -> None:
        people = self._known_people()
        if filter_text:
            needle = filter_text.strip().lower()
            people = [person for person in people if needle in person.lower()]
        filter_people = [""] + people
        if hasattr(self, "payout_person_combo"):
            self.payout_person_combo["values"] = filter_people
        if hasattr(self, "home_person_combo"):
            self.home_person_combo["values"] = filter_people
        if hasattr(self, "profit_person_combo"):
            self.profit_person_combo["values"] = filter_people
        if hasattr(self, "inventory_person_combo"):
            self.inventory_person_combo["values"] = filter_people
        if hasattr(self, "seller_terms_seller_combo"):
            self.seller_terms_seller_combo["values"] = filter_people

    def _person_combo_values(self, allow_blank: bool = False) -> list[str]:
        people = self._known_people()
        return ([""] if allow_blank else []) + people

    def _bind_person_autocomplete(self, combo: ttk.Combobox, refresh_callback=None, allow_blank: bool = False) -> None:
        combo["values"] = self._person_combo_values(allow_blank=allow_blank)
        if not getattr(self, "_is_personal_lucas", lambda: False)():
            combo.configure(state="readonly")
        combo.configure(postcommand=lambda widget=combo: self._refresh_person_combo_widget(widget, allow_blank=allow_blank))
        combo.bind("<FocusIn>", lambda _event, widget=combo: self._refresh_person_combo_widget(widget, allow_blank=allow_blank), add="+")
        if str(combo.cget("state")) != "readonly":
            combo.bind("<KeyRelease>", lambda event, widget=combo: self._filter_person_combo(widget, event, refresh_callback=refresh_callback, allow_blank=allow_blank), add="+")

    def _refresh_person_combo_widget(self, combo: ttk.Combobox, allow_blank: bool = False) -> None:
        if str(combo.cget("state")) == "readonly":
            combo["values"] = self._person_combo_values(allow_blank=allow_blank)
            return
        typed = combo.get().strip().lower()
        people = self._person_combo_values(allow_blank=allow_blank)
        if typed:
            people = [person for person in people if typed in person.lower()]
        combo["values"] = people

    def _filter_person_combo(self, combo: ttk.Combobox, event, refresh_callback=None, allow_blank: bool = False) -> None:
        if event.keysym in {"Up", "Down", "Left", "Right", "Return", "KP_Enter", "Escape", "Tab"}:
            return
        self._refresh_person_combo_widget(combo, allow_blank=allow_blank)
        if refresh_callback:
            refresh_callback()

    def _canonical_person_choice(self, person: object, allow_blank: bool = False) -> str | None:
        if getattr(self, "_is_personal_lucas", lambda: False)():
            return self._personal_default_person()
        text = str(person or "").strip()
        if not text:
            return "" if allow_blank else None
        people_by_key = {known.strip().lower(): known.strip() for known in self._known_people()}
        return people_by_key.get(text.lower())

    def delete_person_records(self, person: str) -> dict[str, int]:
        target = person.strip().lower()
        counts = {"markers": 0, "inventory": 0, "profit": 0}
        if not target:
            return counts
        for marker in self.home_sheet_markers.values():
            if str(marker.get("assigned_person") or "").strip().lower() == target:
                marker["assigned_person"] = ""
                counts["markers"] += 1
        if counts["markers"]:
            self._save_sheet_markers()

        inventory_rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        for record in inventory_rows:
            if str(record.get("assigned_person") or "").strip().lower() == target:
                record["assigned_person"] = "Unassigned"
                counts["inventory"] += 1
        if counts["inventory"]:
            self._save_inventory_ledger(inventory_rows)

        profit_rows = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
        for record in profit_rows:
            if str(record.get("assigned_person") or "").strip().lower() == target:
                record["assigned_person"] = ""
                counts["profit"] += 1
        if counts["profit"]:
            self._save_profit_ledger(profit_rows)
        return counts

    def open_delete_person_dialog(self) -> None:
        people = [person for person in self._known_people() if person.lower() != "unassigned"]
        selected = self.payout_person_var.get().strip() if hasattr(self, "payout_person_var") else ""
        if selected and selected not in people and selected.lower() != "unassigned":
            people.insert(0, selected)
        if not people:
            messagebox.showinfo("No people", "No assigned people are available to delete.")
            return
        person_var = tk.StringVar(value=selected if selected in people else people[0])
        popup = tk.Toplevel(self)
        popup.title("Delete Person")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Delete Person", style="Panel.TLabel", font=("Segoe UI Semibold", 12)).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 12))
        ttk.Label(frame, text="Person", style="Panel.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(0, 12))
        combo = ttk.Combobox(frame, textvariable=person_var, values=people, width=34, state="readonly")
        combo.grid(row=1, column=1, sticky="ew", pady=(0, 12))

        def apply_delete() -> None:
            person = person_var.get().strip()
            if not person:
                return
            confirmed = messagebox.askyesno(
                "Delete person?",
                f"Remove {person} from sheet assignments, inventory, and profit records?",
                parent=popup,
            )
            if not confirmed:
                return
            counts = self.delete_person_records(person)
            for var in (self.payout_person_var, self.inventory_person_var, self.profit_person_var):
                if var.get().strip().lower() == person.lower():
                    var.set("")
            self.refresh_home()
            self.refresh_inventory_tab()
            self.refresh_profit_tab()
            self.refresh_payouts_tab()
            popup.destroy()
            self.status_var.set(
                f"Deleted person {person}: {counts['markers']} sheet marker(s), {counts['inventory']} inventory row(s), {counts['profit']} profit row(s)."
            )

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=2, column=0, columnspan=2, sticky="e")
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="Delete", command=apply_delete, style="Primary.TButton").pack(side=tk.LEFT)
        frame.columnconfigure(1, weight=1)
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")

    def _selected_payout_keys(self) -> list[str]:
        if not hasattr(self, "payout_detail_tree"):
            return []
        return [self.payout_detail_keys.get(iid, "") for iid in self.payout_detail_tree.selection() if self.payout_detail_keys.get(iid)]

    def _payout_history_items_for_person(self, person: str) -> list[dict[str, object]]:
        person_key = str(person or "").strip().lower()
        if not person_key:
            return []
        source_items = [
            item
            for item in self._payout_sheet_items()
            if str(item.get("person") or "Unassigned").strip().lower() == person_key
        ]
        paid_batches: dict[str, dict[str, object]] = {}
        open_items: list[dict[str, object]] = []
        for item in source_items:
            if not item.get("paid"):
                open_items.append(item)
                continue
            key = str(item.get("key") or "")
            marker = self.home_sheet_markers.get(key, {})
            paid_at = str(item.get("paid_at") or marker.get("paid_at") or "Paid").strip() or "Paid"
            batch = paid_batches.setdefault(
                paid_at,
                {
                    "key": f"paid-batch:{paid_at}",
                    "stage": "Paid",
                    "name": f"Total paid at {paid_at}",
                    "person": person,
                    "paid": True,
                    "paid_at": paid_at if paid_at != "Paid" else "",
                    "row_count": 0,
                    "received_count": 0,
                    "purchase_total": 0.0,
                    "estimated_payout_total": 0.0,
                    "estimated_profit": 0.0,
                    "realized_profit_total": 0.0,
                    "expense_total": 0.0,
                    "net_profit_total": 0.0,
                    "payout_balance": 0.0,
                    "payout_basis": "Total paid balance",
                    "status": "Paid",
                },
            )
            for field in ("row_count", "received_count"):
                batch[field] = int(batch.get(field) or 0) + int(item.get(field) or 0)
            for field in ("purchase_total", "estimated_payout_total", "estimated_profit", "realized_profit_total", "expense_total", "net_profit_total", "payout_balance"):
                batch[field] = round(float(batch.get(field) or 0.0) + float(item.get(field) or 0.0), 2)
        return [
            *sorted(paid_batches.values(), key=lambda item: str(item.get("paid_at") or item.get("name") or ""), reverse=True),
            *sorted(open_items, key=lambda item: (str(item.get("stage") or ""), str(item.get("name") or "").lower())),
        ]

    def open_payout_history_menu(self, event=None) -> None:
        if event is not None:
            row_id = self.payout_summary_tree.identify_row(event.y)
            if not row_id:
                return
            self.payout_summary_tree.selection_set(row_id)
        selected = self.payout_summary_tree.selection()
        if not selected:
            return
        person = self.payout_summary_people.get(selected[0])
        if not person or person in {"TOTAL", "━━━━━━", "------"}:
            return
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="View Historical Payouts", command=lambda person=person: self.open_payout_history_popup(person))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def open_payout_history_popup(self, person: str) -> None:
        items = self._payout_history_items_for_person(person)
        popup = tk.Toplevel(self)
        popup.title(f"Payout History - {person}")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.geometry("1080x520")

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(14, 12))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=f"Payout History: {person}", style="Panel.TLabel", font=("Segoe UI Semibold", 12)).pack(anchor=tk.W)
        totals = {
            "paid": sum(float(item.get("payout_balance") or 0.0) for item in items if item.get("paid")),
            "unpaid": sum(float(item.get("payout_balance") or 0.0) for item in items if not item.get("paid")),
        }
        ttk.Label(
            frame,
            text=f"{len(items)} payout row(s) | Paid: {format_money(totals['paid'])} | Open: {format_money(totals['unpaid'])}",
            style="Muted.TLabel",
        ).pack(anchor=tk.W, pady=(4, 10))
        tree = self._build_home_tree(
            frame,
            columns=("status", "stage", "sheet", "cards", "expenses", "net_profit", "balance", "basis", "paid_at"),
            headings={
                "status": "Status",
                "stage": "Type",
                "sheet": "Sheet / Group",
                "cards": "Cards",
                "expenses": "Expenses",
                "net_profit": "Net Profit",
                "balance": "Balance",
                "basis": "Basis",
                "paid_at": "Paid At",
            },
            widths={"status": 90, "stage": 85, "sheet": 260, "cards": 70, "expenses": 100, "net_profit": 110, "balance": 110, "basis": 180, "paid_at": 150},
            height=16,
        )
        tree.tag_configure("paid", foreground="#b8f7ce")
        tree.tag_configure("open", foreground="#fff3b0")
        tree.tag_configure("negative", foreground="#ffd1d1")
        for index, item in enumerate(items):
            paid = bool(item.get("paid"))
            balance = float(item.get("payout_balance") or 0.0)
            tag = "paid" if paid else ("negative" if balance < 0 else "open")
            tree.insert(
                "",
                tk.END,
                iid=f"history:{index}",
                values=(
                    "Paid" if paid else "Open",
                    item.get("stage") or "",
                    item.get("name") or "",
                    item.get("row_count") or 0,
                    format_money(float(item.get("expense_total") or 0.0)),
                    format_money(float(item.get("net_profit_total") or 0.0)),
                    format_money(balance),
                    item.get("payout_basis") or "",
                    item.get("paid_at") or "",
                ),
                tags=(tag,),
            )
        if not items:
            tree.insert("", tk.END, values=("No payouts", "", "", "", "", "", "", "", ""))
        ttk.Button(frame, text="Close", command=popup.destroy, style="Soft.TButton").pack(anchor=tk.E, pady=(10, 0))

    def mark_payout_person_paid(self, event=None) -> None:
        if event is not None:
            row_id = self.payout_summary_tree.identify_row(event.y)
            if not row_id:
                return
            self.payout_summary_tree.selection_set(row_id)
        selected = self.payout_summary_tree.selection()
        if not selected:
            return
        person = self.payout_summary_people.get(selected[0])
        if not person:
            return
        if person in {"TOTAL", "━━━━━━", "------"}:
            return
        matching_items = [
            item
            for item in self._payout_sheet_items()
            if not item["paid"] and item.get("payable", True) and (item["person"] or "Unassigned") == person
        ]
        if not matching_items:
            self.payout_status_var.set(f"No unpaid sheets found for {person}.")
            return
        total_balance = sum(float(item["payout_balance"]) for item in matching_items)
        total_cards = sum(int(item["row_count"]) for item in matching_items)
        self.open_mark_payout_person_paid_popup(person, matching_items, total_cards, total_balance)

    def open_mark_payout_person_paid_popup(
        self,
        person: str,
        matching_items: list[dict[str, object]],
        total_cards: int,
        total_balance: float,
    ) -> None:
        confirmed_var = tk.BooleanVar(value=False)
        popup = tk.Toplevel(self)
        popup.title("Mark Payout Paid")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Mark Payout Paid", style="Panel.TLabel", font=("Segoe UI Semibold", 12)).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 2))
        ttk.Label(frame, text=person, style="Muted.TLabel").grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 14))
        ttk.Label(frame, text="Sheets", style="Panel.TLabel").grid(row=2, column=0, sticky="w", padx=(0, 18), pady=(0, 8))
        ttk.Label(frame, text=str(len(matching_items)), style="Panel.TLabel").grid(row=2, column=1, sticky="w", pady=(0, 8))
        ttk.Label(frame, text="Cards", style="Panel.TLabel").grid(row=3, column=0, sticky="w", padx=(0, 18), pady=(0, 8))
        ttk.Label(frame, text=str(total_cards), style="Panel.TLabel").grid(row=3, column=1, sticky="w", pady=(0, 8))
        ttk.Label(frame, text="Total Balance", style="Panel.TLabel").grid(row=4, column=0, sticky="w", padx=(0, 18), pady=(0, 14))
        ttk.Label(frame, text=format_money(total_balance), style="Panel.TLabel", font=("Segoe UI Semibold", 11)).grid(row=4, column=1, sticky="w", pady=(0, 14))

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=6, column=0, columnspan=2, sticky="e")
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        confirm_button = ttk.Button(
            buttons,
            text="Confirm Mark Paid",
            style="Primary.TButton",
            state=tk.DISABLED,
            command=lambda: self._apply_payout_person_paid(person, matching_items, total_balance, popup),
        )
        confirm_button.pack(side=tk.LEFT)

        def toggle_confirm() -> None:
            confirm_button.configure(state=tk.NORMAL if confirmed_var.get() else tk.DISABLED)

        ttk.Checkbutton(
            frame,
            text="I confirm this person's full active balance has been paid.",
            variable=confirmed_var,
            command=toggle_confirm,
            style="Panel.TCheckbutton",
        ).grid(row=5, column=0, columnspan=2, sticky="w", pady=(0, 14))
        frame.columnconfigure(1, weight=1)
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")

    def _apply_payout_person_paid(
        self,
        person: str,
        matching_items: list[dict[str, object]],
        total_balance: float,
        popup: tk.Toplevel | None = None,
    ) -> None:
        for item in matching_items:
            key = str(item["key"])
            kind, _name = self._split_home_sheet_key(key)
            marker = dict(self.home_sheet_markers.get(key, {}))
            summary = self.home_sheet_summaries.get(key, {})
            marker["assigned_person"] = str(item["person"] or "").strip()
            marker["paid"] = True
            marker["paid_at"] = datetime.now().isoformat(timespec="seconds")
            marker["all_received"] = bool(marker.get("all_received") or summary.get("all_received") or kind == "Received")
            marker["tracking_number"] = str(marker.get("tracking_number") or "")
            self.home_sheet_markers[key] = marker
        self._save_sheet_markers()
        self.refresh_home()
        if popup is not None:
            popup.destroy()
        self.status_var.set(f"Marked {len(matching_items)} sheet(s) paid for {person}: {format_money(total_balance)}.")

    def open_payout_marker_editor(self, event=None) -> None:
        if event is not None:
            row_id = self.payout_detail_tree.identify_row(event.y)
            if not row_id:
                return
            self.payout_detail_tree.selection_set(row_id)
        keys = self._selected_payout_keys()
        if not keys:
            return
        key = keys[0]
        kind, name = self._split_home_sheet_key(key)
        marker = self.home_sheet_markers.get(key, {})
        summary = self.home_sheet_summaries.get(key, {})
        if kind == "Sold":
            key_person, source_sheet = self._split_sold_payout_name(name)
            person = str(marker.get("assigned_person") or key_person).strip()
            display_name = source_sheet
            realized_profit_total = self._realized_profit_totals_by_person_sheet().get((person.lower(), Path(source_sheet).name.lower()), 0.0)
        else:
            person = str(marker.get("assigned_person") or "").strip()
            display_name = name
            realized_profit_total = self._realized_profit_totals_by_person_sheet().get((person.lower(), Path(name).name.lower()), 0.0)
        payout_item = self._payout_item_for_key(key)
        payable = bool(payout_item.get("payable", True)) if payout_item else True
        if payout_item:
            balance = float(payout_item.get("payout_balance") or 0.0)
            basis = str(payout_item.get("payout_basis") or "")
            status = str(payout_item.get("status") or "")
        else:
            balance, basis = self._active_payout_balance(
                person,
                float(summary.get("purchase_total") or 0.0),
                float(summary.get("estimated_payout_total") or 0.0),
                realized_profit_total=realized_profit_total,
            )
            status = ""
        paid_var = tk.BooleanVar(value=bool(marker.get("paid")))
        person_var = tk.StringVar(value=str(marker.get("assigned_person") or "").strip())

        popup = tk.Toplevel(self)
        popup.title("Payout Sheet")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=display_name, style="Panel.TLabel", font=("Segoe UI Semibold", 12)).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 2))
        subtitle = f"{kind} | Balance: {format_money(balance)} | {basis}"
        if status and status not in {"Paid", "Unpaid", "Sold"}:
            subtitle = f"{subtitle} | {status}"
        ttk.Label(frame, text=subtitle, style="Muted.TLabel").grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 14))
        ttk.Label(frame, text="Assigned Person", style="Panel.TLabel").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        person_combo = ttk.Combobox(frame, textvariable=person_var, width=34)
        person_combo.grid(row=2, column=1, sticky="ew", pady=(0, 10))
        self._bind_person_autocomplete(person_combo, allow_blank=True)
        paid_state = tk.NORMAL if payable or bool(marker.get("paid")) else tk.DISABLED
        ttk.Checkbutton(frame, text="Paid", variable=paid_var, state=paid_state, style="Panel.TCheckbutton").grid(row=3, column=0, columnspan=2, sticky="w", pady=(0, 14))
        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=4, column=0, columnspan=2, sticky="e")
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(
            buttons,
            text="Save",
            command=lambda: self.save_payout_sheet_marker(key, person_var.get().strip(), bool(paid_var.get()), popup),
            style="Primary.TButton",
        ).pack(side=tk.LEFT)
        frame.columnconfigure(1, weight=1)
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")

    def _payout_item_for_key(self, key: str) -> dict[str, object] | None:
        for item in self._payout_sheet_items():
            if str(item.get("key") or "") == key:
                return item
        return None

    def save_payout_sheet_marker(self, key: str, person: str, paid: bool, popup: tk.Toplevel | None = None) -> None:
        marker = dict(self.home_sheet_markers.get(key, {}))
        kind, _name = self._split_home_sheet_key(key)
        summary = self.home_sheet_summaries.get(key, {})
        payout_item = self._payout_item_for_key(key)
        if paid and payout_item and not bool(payout_item.get("payable", True)):
            messagebox.showinfo("Payout not ready", str(payout_item.get("status") or "This seller payout is still waiting on required values."))
            return
        person_choice = self._canonical_person_choice(person, allow_blank=True)
        if person_choice is None:
            messagebox.showinfo("Person required", "Choose an existing person.")
            return
        marker["assigned_person"] = person_choice
        marker["paid"] = bool(paid)
        if paid:
            marker.setdefault("paid_at", datetime.now().isoformat(timespec="seconds"))
        else:
            marker.pop("paid_at", None)
        marker["all_received"] = bool(marker.get("all_received") or summary.get("all_received") or kind == "Received")
        marker["tracking_number"] = str(marker.get("tracking_number") or "")
        self.home_sheet_markers[key] = marker
        self._save_sheet_markers()
        self.refresh_home()
        if popup is not None:
            popup.destroy()
        self.status_var.set(f"Updated payout marker for {self._split_home_sheet_key(key)[1]}.")

    def _load_home_selected_marker(self) -> None:
        if not hasattr(self, "home_sheet_list"):
            return
        selected = self.home_sheet_list.curselection()
        if not selected:
            return
        kind = self.home_sheet_kind.get()
        name = str(self.home_sheet_list.get(selected[0]))
        key = self._home_sheet_key(kind, name)
        self.home_selected_sheet_key = key
        summary = self.home_sheet_summaries.get(key, {})
        warning = str(summary.get("seller_payout_warning") or "").strip()
        if warning:
            self.status_var.set(warning)

    def _show_home_sheet_context_menu(self, event) -> str:
        if not hasattr(self, "home_sheet_list"):
            return "break"
        index = self.home_sheet_list.nearest(event.y)
        if index < 0 or index >= self.home_sheet_list.size():
            return "break"
        self.home_sheet_list.selection_clear(0, tk.END)
        self.home_sheet_list.selection_set(index)
        self.home_sheet_list.activate(index)
        self._load_home_selected_marker()
        menu = tk.Menu(self, tearoff=False, bg="#1f1f1f", fg="#ffffff", activebackground="#1ed760", activeforeground="#000000")
        kind, _name = self._split_home_sheet_key(self.home_selected_sheet_key)
        menu.add_command(label="Review Sheet", command=self.review_selected_home_sheet)
        menu.add_separator()
        move_menu = tk.Menu(menu, tearoff=False, bg="#1f1f1f", fg="#ffffff", activebackground="#1ed760", activeforeground="#000000")
        for target_stage in ("Incoming", "Working", "Received"):
            move_menu.add_command(
                label=f"Move to {target_stage}",
                command=lambda stage=target_stage: self.move_selected_home_sheet_to_stage(stage),
                state=tk.DISABLED if target_stage == kind else tk.NORMAL,
            )
        menu.add_cascade(label="Move Sheet", menu=move_menu)
        menu.add_separator()
        menu.add_command(label="Delete Sheet", command=self.delete_selected_home_sheet)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        return "break"

    def review_selected_home_sheet(self) -> None:
        if not self.home_selected_sheet_key:
            messagebox.showinfo("Choose sheet", "Choose a sheet on Home before reviewing.")
            return
        kind, name = self._split_home_sheet_key(self.home_selected_sheet_key)
        if kind not in {"Incoming", "Working", "Received"} or not name:
            messagebox.showinfo("Cannot review", "Only Incoming, Working, and Received sheets can be reviewed from Home.")
            return
        path = self._sheet_path_for_stage(kind, name)
        if not self._sheet_path_is_visible_home_sheet(kind, path):
            messagebox.showerror("Review blocked", f"Review is only allowed inside {kind} sheets.")
            return
        if not path.exists():
            messagebox.showerror("Review failed", f"Sheet not found: {path}")
            return
        try:
            preview = self._home_sheet_preview_data(path)
        except Exception as error:
            messagebox.showerror("Review failed", str(error))
            return
        self._open_home_sheet_review_popup(kind, name, preview)

    def _home_sheet_preview_data(self, path: Path, max_rows: int = 500) -> dict[str, object]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            max_column = max(int(sheet.max_column or 0), 1)
            columns = ["row", *[get_column_letter(index) for index in range(1, max_column + 1)]]
            headings = {"row": "#", **{column: column for column in columns[1:]}}
            rows: list[tuple[object, ...]] = []
            for row_index, row_values in enumerate(
                sheet.iter_rows(min_row=1, max_row=min(int(sheet.max_row or 0), max_rows), max_col=max_column, values_only=True),
                start=1,
            ):
                rows.append((row_index, *row_values))
            truncated = bool((sheet.max_row or 0) > max_rows)
            return {
                "sheet_title": sheet.title,
                "columns": columns,
                "headings": headings,
                "rows": rows,
                "truncated": truncated,
                "max_rows": max_rows,
            }
        finally:
            workbook.close()

    def _home_sheet_preview_value(self, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M")
        return str(value)

    def _open_home_sheet_review_popup(self, kind: str, name: str, preview: dict[str, object]) -> None:
        popup = tk.Toplevel(self)
        popup.title(f"Review Sheet - {name}")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.geometry("1180x650")
        popup.minsize(820, 420)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(16, 14))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=name, style="Panel.TLabel", font=("Segoe UI Semibold", 12)).pack(anchor="w")
        subtitle = f"{kind} | {preview.get('sheet_title') or 'Sheet'}"
        if preview.get("truncated"):
            subtitle += f" | Showing first {preview.get('max_rows')} rows"
        ttk.Label(frame, text=subtitle, style="Muted.TLabel").pack(anchor="w", pady=(2, 12))

        table_frame = ttk.Frame(frame, style="Panel.TFrame")
        table_frame.pack(fill=tk.BOTH, expand=True)
        columns = tuple(str(column) for column in preview.get("columns") or ("row",))
        headings = preview.get("headings") if isinstance(preview.get("headings"), dict) else {}
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        y_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        for column in columns:
            tree.heading(column, text=str(headings.get(column, column)), anchor=tk.W)
            if column == "row":
                tree.column(column, width=58, minwidth=48, stretch=False, anchor=tk.E)
            else:
                tree.column(column, width=150, minwidth=80, stretch=False, anchor=tk.W)
        for row_values in preview.get("rows") or []:
            tree.insert("", tk.END, values=[self._home_sheet_preview_value(value) for value in row_values])
        self._bind_context_menu(tree, lambda event, preview_tree=tree: self._show_home_sheet_review_context_menu(event, preview_tree))

        actions = ttk.Frame(frame, style="Panel.TFrame")
        actions.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(actions, text="Close", command=popup.destroy, style="Primary.TButton").pack(side=tk.RIGHT)
        popup.bind("<Escape>", lambda _event: popup.destroy())

    def _show_home_sheet_review_context_menu(self, event: tk.Event, tree: ttk.Treeview) -> str:
        row_id = tree.identify_row(event.y)
        if not row_id:
            return "break"
        column_id = tree.identify_column(event.x)
        if row_id not in tree.selection():
            tree.selection_set(row_id)
            tree.focus(row_id)
        menu = tk.Menu(self, tearoff=False, bg="#1f1f1f", fg="#ffffff", activebackground="#1ed760", activeforeground="#000000")
        menu.add_command(label="Copy Cell", command=lambda row=row_id, column=column_id: self.copy_tree_cell_value(tree, row, column, "sheet cell"))
        menu.add_command(label="Copy Row", command=lambda row=row_id: self.copy_tree_row_values(tree, row, "sheet row"))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
        return "break"

    def move_selected_home_sheet_to_stage(self, target_stage: str) -> None:
        move_started = time.perf_counter()
        move_phases: list[str] = []
        if not self.home_selected_sheet_key:
            messagebox.showinfo("Choose sheet", "Choose a sheet on Home before moving.")
            return
        source_stage, name = self._split_home_sheet_key(self.home_selected_sheet_key)
        if source_stage not in {"Incoming", "Working", "Received"} or not name:
            messagebox.showinfo("Cannot move", "Only Incoming, Working, and Received sheets can be moved from Home.")
            return
        if target_stage not in {"Incoming", "Working", "Received"}:
            messagebox.showinfo("Cannot move", "Choose Incoming, Working, or Received.")
            return
        if source_stage == target_stage:
            return
        path = self._sheet_path_for_stage(source_stage, name)
        if not self._sheet_path_is_visible_home_sheet(source_stage, path):
            messagebox.showerror("Move blocked", f"Move is only allowed inside {source_stage} sheets.")
            return
        if not self._confirm_home_stage_move(source_stage, target_stage, name):
            return
        try:
            lock_started = time.perf_counter()
            with shared_lock(CARD_PIPELINE_DIR, "sheet-stage-move", self.lucas_identity):
                move_phases.append(f"lock_wait={time.perf_counter() - lock_started:.3f}s")
                phase_started = time.perf_counter()
                moved_key, cleanup = self._move_home_sheet_to_stage(self.home_selected_sheet_key, target_stage)
                move_phases.append(f"move={time.perf_counter() - phase_started:.3f}s")
                self.home_selected_sheet_key = moved_key
                self.home_sheet_kind.set(target_stage)
                phase_started = time.perf_counter()
                self._save_sheet_markers()
                move_phases.append(f"save_markers={time.perf_counter() - phase_started:.3f}s")
        except Exception as error:
            record_performance_event("home.stage_move.failed", move_started, f"sheet={name} from={source_stage} to={target_stage} error={error}", force=True)
            messagebox.showerror("Move failed", str(error))
            return
        phase_started = time.perf_counter()
        self._refresh_after_home_stage_move(name, source_stage, target_stage)
        move_phases.append(f"refresh={time.perf_counter() - phase_started:.3f}s")
        cleanup_note = ""
        if cleanup:
            cleanup_note = (
                f" Cleared {cleanup.get('received_rows_cleared', 0)} received mark(s), "
                f"removed {cleanup.get('company_rows_removed', 0)} company row(s), "
                f"removed {cleanup.get('profit_rows_removed', 0)} profit ledger row(s), "
                f"and removed {cleanup.get('inventory_rows_removed', 0)} inventory row(s)."
            )
        self.status_var.set(f"Moved {name} from {source_stage} to {target_stage}.{cleanup_note}")
        phase_started = time.perf_counter()
        self._append_activity("Sheet Move", f"Moved {name} from {source_stage} to {target_stage}.", {"sheet": name, "from": source_stage, "to": target_stage, "cleanup": cleanup})
        move_phases.append(f"activity={time.perf_counter() - phase_started:.3f}s")
        record_performance_event(
            "home.stage_move.total",
            move_started,
            f"sheet={name} from={source_stage} to={target_stage} {' '.join(move_phases)}",
            force=True,
        )

    def _confirm_home_stage_move(self, source_stage: str, target_stage: str, name: str) -> bool:
        if source_stage == "Received" and target_stage != "Received":
            confirmed = messagebox.askyesno(
                "Move received sheet?",
                (
                    f"Move this sheet from Received to {target_stage}?\n\n{name}\n\n"
                    "This will clear received/paid markers and remove inventory, company-sheet, and profit rows created from this sheet."
                ),
                icon="warning",
            )
            if not confirmed:
                return False
            return messagebox.askyesno(
                "Remove received side effects?",
                (
                    "Confirm one more time before L.U.C.A.S removes rows tied to this received sheet.\n\n"
                    f"Sheet: {name}\n"
                    "Affected data: active inventory rows, company-sheet rows, and profit ledger rows created from this sheet."
                ),
                icon="warning",
            )
        return messagebox.askyesno(
            "Move sheet?",
            f"Move this sheet from {source_stage} to {target_stage}?\n\n{name}",
        )

    def _refresh_after_home_stage_move(self, sheet_name: str, source_stage: str, target_stage: str) -> None:
        perf_start = time.perf_counter()
        phase_started = time.perf_counter()
        self.refresh_working_sheets()
        record_performance_event("home.stage_move.refresh_working_sheets", phase_started, f"sheet={sheet_name} from={source_stage} to={target_stage}")
        phase_started = time.perf_counter()
        self.refresh_received_sheets()
        record_performance_event("home.stage_move.refresh_received_sheets", phase_started, f"sheet={sheet_name} from={source_stage} to={target_stage}")
        phase_started = time.perf_counter()
        if target_stage in {"Incoming", "Working"}:
            self.refresh_incoming_index()
        elif source_stage in {"Incoming", "Working"}:
            self._drop_sheet_from_incoming_index(sheet_name)
        record_performance_event("home.stage_move.refresh_index", phase_started, f"sheet={sheet_name} from={source_stage} to={target_stage}")
        phase_started = time.perf_counter()
        self._refresh_home_after_stage_move(sheet_name, source_stage, target_stage)
        record_performance_event("home.stage_move.refresh_home_memory", phase_started, f"sheet={sheet_name} from={source_stage} to={target_stage}")
        phase_started = time.perf_counter()
        self._refresh_table()
        record_performance_event("home.stage_move.refresh_table", phase_started, f"sheet={sheet_name} from={source_stage} to={target_stage}")
        record_performance_event(
            "home.stage_move_refresh",
            perf_start,
            f"sheet={sheet_name} from={source_stage} to={target_stage}",
        )

    def _refresh_home_after_stage_move(self, sheet_name: str, source_stage: str, target_stage: str) -> None:
        source_key = self._home_sheet_key(source_stage, sheet_name)
        target_key = self._home_sheet_key(target_stage, sheet_name)
        target_path = self._sheet_path_for_stage(target_stage, sheet_name)

        self.home_sheet_paths.setdefault(source_stage, {}).pop(sheet_name, None)
        target_paths = self.home_sheet_paths.setdefault(target_stage, {})
        if target_path.exists():
            target_paths.pop(sheet_name, None)
            self.home_sheet_paths[target_stage] = {sheet_name: target_path, **target_paths}

        summary = self.home_sheet_summaries.pop(source_key, None)
        if summary is None and target_path.exists():
            try:
                summary = self._summarize_home_workbook_cached(target_path)
            except Exception:
                summary = None
        if summary is not None:
            self.home_sheet_summaries[target_key] = self._enrich_home_seller_payout_summary(
                target_path,
                self.home_sheet_markers.get(target_key, {}),
                dict(summary),
            )

        self._refresh_home_sheet_list()
        self._refresh_home_metrics()
        self.refresh_payouts_tab()
        self._update_home_sheet_tabs()

    def _drop_sheet_from_incoming_index(self, sheet_name: str) -> None:
        target = Path(str(sheet_name or "")).name
        index = getattr(self, "incoming_cert_index", None)
        if not target or not isinstance(index, dict):
            return
        kept: dict[str, dict[str, object]] = {}
        for cert, row in index.items():
            row_sheet = Path(str((row or {}).get("sheet") or "")).name
            if row_sheet != target:
                kept[cert] = row
        self.incoming_cert_index = kept
        self._match_all_review_rows()
        if hasattr(self, "review_status"):
            self.review_status.set(f"Indexed {len(kept)} cert(s) after moving {target}.")

    def delete_selected_home_sheet(self) -> None:
        if not self.home_selected_sheet_key:
            messagebox.showinfo("Choose sheet", "Choose a sheet on Home before deleting.")
            return
        kind, name = self._split_home_sheet_key(self.home_selected_sheet_key)
        if kind not in {"Incoming", "Working", "Received"} or not name:
            messagebox.showinfo("Cannot delete", "Only Incoming, Working, and Received sheets can be deleted from Home.")
            return
        path = self._sheet_path_for_stage(kind, name)
        if not self._sheet_path_is_visible_home_sheet(kind, path):
            messagebox.showerror("Delete blocked", f"Delete is only allowed inside {kind} sheets.")
            return
        if not path.exists():
            messagebox.showerror("Delete failed", f"Sheet not found: {path}")
            return
        confirmed = messagebox.askyesno(
            "Delete sheet?",
            f"Delete this {kind.lower()} sheet?\n\n{name}\n\nThis archives the .xlsx file for {DELETED_ARCHIVE_RETENTION_DAYS} days, then it can be removed from the deleted archive.",
            icon="warning",
        )
        if not confirmed:
            return
        try:
            with shared_lock(CARD_PIPELINE_DIR, "sheet-delete", self.lucas_identity):
                archive_path = self._archive_deleted_file(
                    path,
                    DELETED_SHEETS_DIR,
                    "home_sheet_deleted",
                    {"sheet": name, "stage": kind},
                )
                inventory_rows_removed = self._remove_inventory_rows_for_source(name)
                self._delete_sheet_marker(self.home_selected_sheet_key)
                self._save_sheet_markers()
        except Exception as error:
            messagebox.showerror("Delete failed", str(error))
            return
        self.home_selected_sheet_key = ""
        self.refresh_pipeline()
        self.refresh_home()
        self.status_var.set(f"Archived deleted {kind.lower()} sheet for {DELETED_ARCHIVE_RETENTION_DAYS} days: {name}. Removed {inventory_rows_removed} inventory row(s).")
        self._append_activity(
            "Sheet Delete",
            f"Archived deleted {kind.lower()} sheet {name} for {DELETED_ARCHIVE_RETENTION_DAYS} days.",
            {"sheet": name, "stage": kind, "archive_path": str(archive_path), "inventory_rows_removed": inventory_rows_removed},
        )

    def open_sheet_marker_editor(self) -> None:
        if not self.home_selected_sheet_key:
            messagebox.showinfo("Choose sheet", "Choose a sheet on Home before editing markers.")
            return
        kind, name = self._split_home_sheet_key(self.home_selected_sheet_key)
        marker = self.home_sheet_markers.get(self.home_selected_sheet_key, {})
        summary = self.home_sheet_summaries.get(self.home_selected_sheet_key, {})
        incoming_proper_var = tk.BooleanVar(value=(kind == "Incoming"))
        all_received_var = tk.BooleanVar(value=bool(marker.get("all_received") or summary.get("all_received")))
        tracking_var = tk.StringVar(value=str(marker.get("tracking_number") or ""))
        personal_profile = self._is_personal_lucas()
        person_var = tk.StringVar(value=self._personal_default_person() if personal_profile else str(marker.get("assigned_person") or ""))

        popup = tk.Toplevel(self)
        popup.title("Edit Sheet Markers")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.resizable(False, False)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(18, 16))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=name, style="Panel.TLabel", font=("Segoe UI Semibold", 12)).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 2))
        ttk.Label(frame, text=kind, style="Muted.TLabel").grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 14))
        ttk.Checkbutton(frame, text="Incoming", variable=incoming_proper_var, style="Panel.TCheckbutton").grid(row=2, column=0, columnspan=2, sticky="w", pady=(0, 10))
        ttk.Label(frame, text="Tracking Number", style="Panel.TLabel").grid(row=3, column=0, sticky="w", padx=(0, 10), pady=(0, 10))
        ttk.Entry(frame, textvariable=tracking_var, width=34).grid(row=3, column=1, sticky="ew", pady=(0, 10))
        ttk.Checkbutton(frame, text="All Received", variable=all_received_var, style="Panel.TCheckbutton").grid(row=4, column=0, columnspan=2, sticky="w", pady=(0, 10))
        if not personal_profile:
            ttk.Label(frame, text="Assigned Person", style="Panel.TLabel").grid(row=5, column=0, sticky="w", padx=(0, 10), pady=(0, 14))
            person_combo = ttk.Combobox(frame, textvariable=person_var, width=34)
            person_combo.grid(row=5, column=1, sticky="ew", pady=(0, 14))
            self._bind_person_autocomplete(person_combo, allow_blank=True)
        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=6, column=0, columnspan=2, sticky="e")
        ttk.Button(buttons, text="Cancel", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(
            buttons,
            text="Save Markers",
            command=lambda: self.save_home_sheet_markers(
                {
                    "incoming_proper": bool(incoming_proper_var.get()),
                    "tracking_number": tracking_var.get().strip(),
                    "all_received": bool(all_received_var.get()),
                    "assigned_person": person_var.get().strip(),
                },
                popup,
            ),
            style="Primary.TButton",
        ).pack(side=tk.LEFT)
        frame.columnconfigure(1, weight=1)
        popup.update_idletasks()
        x = self.winfo_rootx() + max(80, (self.winfo_width() - popup.winfo_width()) // 2)
        y = self.winfo_rooty() + max(80, (self.winfo_height() - popup.winfo_height()) // 2)
        popup.geometry(f"+{x}+{y}")

    def save_home_sheet_markers(self, marker: dict[str, object], popup: tk.Toplevel | None = None) -> None:
        if not self.home_selected_sheet_key:
            messagebox.showinfo("Choose sheet", "Choose a sheet on Home before saving markers.")
            return
        existing_marker = dict(self.home_sheet_markers.get(self.home_selected_sheet_key, {}))
        incoming_proper = bool(marker.get("incoming_proper"))
        old_assigned_person = str(existing_marker.get("assigned_person") or "").strip()
        updated_marker = dict(existing_marker)
        updated_marker["paid"] = bool(existing_marker.get("paid"))
        updated_marker["tracking_number"] = str(marker.get("tracking_number") or "").strip()
        updated_marker["all_received"] = bool(marker.get("all_received"))
        updated_marker["assigned_person"] = (
            self._personal_default_person()
            if self._is_personal_lucas()
            else str(marker.get("assigned_person") or "").strip()
        )
        if not self._is_personal_lucas():
            person_choice = self._canonical_person_choice(updated_marker["assigned_person"], allow_blank=True)
            if person_choice is None:
                messagebox.showinfo("Person required", "Choose an existing person.")
                return
            updated_marker["assigned_person"] = person_choice
        if bool(updated_marker.get("seller_terms_applied") or updated_marker.get("seller_sheet_type")):
            sheet_type = str(updated_marker.get("seller_sheet_type") or "").strip()
            assigned_person = str(updated_marker.get("assigned_person") or "").strip()
            if not assigned_person or not sheet_type:
                messagebox.showinfo("Seller terms required", "Network Mode seller sheets need an assigned person and Sheet Type.")
                return
            if assigned_person.lower() != old_assigned_person.lower():
                seller_term = self._seller_terms_match(assigned_person, sheet_type)
                if not seller_term:
                    messagebox.showinfo(
                        "People Rule missing",
                        f"No People Rule was found for {assigned_person} / {sheet_type}. Update People Rules before changing this seller sheet.",
                    )
                    return
                updated_marker["seller_terms_applied"] = True
                updated_marker["seller_rate"] = seller_term.get("rate")
                updated_marker["seller_deduction"] = seller_term.get("deduction")
        marker = updated_marker
        key = self.home_selected_sheet_key
        source_kind, _ = self._split_home_sheet_key(key)
        moved = False
        inventory_rows_added = 0
        inventory_candidate_rows = 0
        inventory_rows_reassigned = 0
        profit_rows_reassigned = 0
        try:
            with shared_lock(CARD_PIPELINE_DIR, "receive-company-sheets", self.lucas_identity):
                selected_kind, _selected_name = self._split_home_sheet_key(key)
                if selected_kind == "Received":
                    marker["all_received"] = True
                    marker.setdefault("received_at", existing_marker.get("received_at") or datetime.now().isoformat(timespec="seconds"))
                elif marker["all_received"]:
                    moved_key = self._move_sheet_to_received(key)
                    if moved_key:
                        self._delete_sheet_marker(key)
                        key = moved_key
                        self.home_selected_sheet_key = key
                        moved = True
                elif incoming_proper:
                    moved_key = self._move_working_sheet_to_incoming(key)
                    if moved_key:
                        self._delete_sheet_marker(key)
                        key = moved_key
                        self.home_selected_sheet_key = key
                        self.home_sheet_kind.set("Incoming")
                        moved = True
                if moved:
                    current_kind, _current_name = self._split_home_sheet_key(key)
                    marker = self._marker_for_stage(marker, current_kind)
                self.home_sheet_markers[key] = marker
                _current_kind, current_name = self._split_home_sheet_key(key)
                if old_assigned_person != str(marker.get("assigned_person") or "").strip():
                    inventory_rows_reassigned = self._retarget_inventory_rows_for_source(current_name, str(marker.get("assigned_person") or ""))
                    profit_rows_reassigned = self._retarget_profit_rows_for_source(current_name, str(marker.get("assigned_person") or ""))
                if _current_kind == "Received" and marker["all_received"] and str(marker.get("assigned_person") or "").strip():
                    inventory_rows_added, inventory_candidate_rows = self._sync_received_sheet_inventory_to_ledger(
                        _current_kind,
                        self._sheet_path_for_stage(_current_kind, current_name),
                        str(marker.get("assigned_person") or ""),
                    )
                self._save_sheet_markers()
        except Exception as error:
            messagebox.showerror("Save failed", str(error))
            return
        refresh_marker_save = getattr(self, "_refresh_after_home_marker_save", None)
        if callable(refresh_marker_save):
            refresh_marker_save(current_name, _current_kind, moved=moved, source_stage=source_kind)
        if hasattr(self, "inventory_tree") and (inventory_rows_added or inventory_rows_reassigned):
            self.refresh_inventory_tab(enrich=True)
        if profit_rows_reassigned and hasattr(self, "profit_tree"):
            self.refresh_profit_tab()
        if popup is not None:
            popup.destroy()
        reassigned_notes = []
        if inventory_rows_reassigned:
            reassigned_notes.append(f"{inventory_rows_reassigned} inventory row(s)")
        if profit_rows_reassigned:
            reassigned_notes.append(f"{profit_rows_reassigned} profit row(s)")
        inventory_note = f" Reassigned {', '.join(reassigned_notes)}." if reassigned_notes else ""
        if inventory_rows_added:
            inventory_note += f" Added {inventory_rows_added} inventory row(s)."
        elif inventory_candidate_rows:
            inventory_note += " Inventory was already up to date."
        self.status_var.set(("Sheet markers saved and moved." if moved else "Sheet markers saved.") + inventory_note)

    def _refresh_after_home_marker_save(self, sheet_name: str, stage: str, moved: bool = False, source_stage: str = "") -> None:
        if moved:
            self._refresh_after_home_stage_move(sheet_name, source_stage, stage)
            return
        key = self._home_sheet_key(stage, sheet_name)
        path = self._sheet_path_for_stage(stage, sheet_name)
        if path.exists():
            try:
                summary = self._summarize_home_workbook_cached(path)
            except Exception:
                summary = self.home_sheet_summaries.get(key, {})
            self.home_sheet_summaries[key] = self._enrich_home_seller_payout_summary(
                path,
                self.home_sheet_markers.get(key, {}),
                dict(summary),
            )
        self._refresh_home_sheet_list()
        self._refresh_home_metrics()
        self.refresh_payouts_tab()
        self._update_home_sheet_tabs()

    def _home_sheet_key(self, kind: str, name: str) -> str:
        return f"{kind}|{name}"

    def _split_home_sheet_key(self, key: str) -> tuple[str, str]:
        if "|" not in key:
            return "", key
        kind, name = key.split("|", 1)
        return kind, name

    def _move_working_sheet_to_incoming(self, key: str) -> str:
        moved_key, _cleanup = self._move_home_sheet_to_stage(key, "Incoming")
        return moved_key

    def _move_sheet_to_received(self, key: str) -> str:
        moved_key, _cleanup = self._move_home_sheet_to_stage(key, "Received")
        return moved_key

    def _move_received_sheet_to_incoming(self, key: str) -> str:
        moved_key, _cleanup = self._move_home_sheet_to_stage(key, "Incoming")
        return moved_key

    def _assign_sheet_to_seller(
        self,
        stage: str,
        sheet_name: str,
        seller: str,
        sheet_type: str = "",
        term: dict[str, object] | None = None,
    ) -> bool:
        person = str(seller or "").strip()
        if stage not in {"Working", "Incoming", "Received"} or not sheet_name or not person:
            return False
        key = self._home_sheet_key(stage, sheet_name)
        marker = dict(self._load_sheet_markers().get(key, {}))
        marker.update(dict(self.home_sheet_markers.get(key, {})))
        marker["assigned_person"] = person
        sheet_type = str(sheet_type or "").strip()
        if sheet_type:
            marker["seller_terms_applied"] = True
            marker["seller_sheet_type"] = sheet_type
            if term:
                rate = term.get("rate")
                deduction = term.get("deduction")
                if rate is not None:
                    marker["seller_rate"] = rate
                if deduction is not None:
                    marker["seller_deduction"] = deduction
        marker = self._marker_for_stage(marker, stage)
        self.home_sheet_markers[key] = marker
        self._save_sheet_markers()
        return True

    def _move_home_sheet_to_stage(self, key: str, target_stage: str) -> tuple[str, dict[str, int]]:
        perf_start = time.perf_counter()
        source_stage, name = self._split_home_sheet_key(key)
        if source_stage not in {"Incoming", "Working", "Received"} or target_stage not in {"Incoming", "Working", "Received"} or not name:
            return "", {}
        if source_stage == target_stage:
            return key, {}
        source = self._sheet_path_for_stage(source_stage, name)
        phase_started = time.perf_counter()
        if not source.exists():
            raise FileNotFoundError(f"{source_stage} sheet not found: {source}")
        record_performance_event("home.stage_move.source_exists", phase_started, f"sheet={name} from={source_stage} to={target_stage}")
        target_dir = {
            "Incoming": INCOMING_SHEETS_DIR,
            "Working": WORKING_SHEETS_DIR,
            "Received": RECEIVED_SHEETS_DIR,
        }[target_stage]
        phase_started = time.perf_counter()
        target_dir.mkdir(parents=True, exist_ok=True)
        record_performance_event("home.stage_move.target_mkdir", phase_started, f"sheet={name} from={source_stage} to={target_stage}")
        destination = target_dir / source.name
        phase_started = time.perf_counter()
        if destination.exists():
            raise FileExistsError(f"{target_stage} sheet already exists: {destination.name}")
        record_performance_event("home.stage_move.destination_exists", phase_started, f"sheet={name} from={source_stage} to={target_stage}")

        cleanup: dict[str, int] = {}
        moving_out_of_received = source_stage == "Received" and target_stage != "Received"
        if moving_out_of_received:
            phase_started = time.perf_counter()
            cleanup = self._cleanup_sheet_received_side_effects(source.name, source)
            record_performance_event("home.stage_move.cleanup", phase_started, f"sheet={name} from={source_stage} to={target_stage}")
        phase_started = time.perf_counter()
        shutil.move(str(source), str(destination))
        record_performance_event("home.stage_move.shutil_move", phase_started, f"sheet={name} from={source_stage} to={target_stage}", force=True)

        phase_started = time.perf_counter()
        old_marker = dict(self.home_sheet_markers.get(key, {}))
        if self._is_personal_lucas():
            old_marker["assigned_person"] = self._personal_default_person()
        self._delete_sheet_marker(key)
        new_key = self._home_sheet_key(target_stage, destination.name)
        self.home_sheet_markers[new_key] = self._marker_for_stage(old_marker, target_stage)
        record_performance_event("home.stage_move.marker_update", phase_started, f"sheet={name} from={source_stage} to={target_stage}")
        record_performance_event("home.stage_move.move_function", perf_start, f"sheet={name} from={source_stage} to={target_stage}", force=True)
        return new_key, cleanup

    def _marker_for_stage(self, marker: dict[str, object], stage: str) -> dict[str, object]:
        normalized = dict(marker)
        if stage == "Received":
            normalized["all_received"] = True
            normalized.setdefault("received_at", datetime.now().isoformat(timespec="seconds"))
        else:
            normalized["all_received"] = False
            normalized["paid"] = False
            for field in ("received_at", "archived_at", "archived_from"):
                normalized.pop(field, None)
        return normalized

    def _cleanup_sheet_received_side_effects(self, source_sheet_name: str, source_path: Path) -> dict[str, int]:
        clear_result = clear_received_in_workbooks([source_path])
        company_result = remove_company_sheet_rows_for_source(COMPANY_SHEETS_DIR, source_sheet_name)
        profit_rows_removed = self._remove_profit_ledger_rows_for_source(source_sheet_name)
        inventory_rows_removed = self._remove_inventory_rows_for_source(source_sheet_name)
        errors = list(clear_result.get("errors") or []) + list(company_result.get("errors") or [])
        if errors:
            raise RuntimeError("Move cleanup failed: " + "; ".join(str(error) for error in errors[:5]))
        return {
            "received_rows_cleared": int(clear_result.get("rows_cleared") or 0),
            "company_rows_removed": int(company_result.get("rows_removed") or 0),
            "profit_rows_removed": profit_rows_removed,
            "inventory_rows_removed": inventory_rows_removed,
        }

    def _remove_profit_ledger_rows_for_source(self, source_sheet_name: str) -> int:
        source_name = Path(str(source_sheet_name or "")).name
        if not source_name:
            return 0
        ledger = [self._normalize_profit_record(record) for record in self._load_profit_ledger()]
        kept = [
            record
            for record in ledger
            if Path(str(record.get("source_sheet") or "")).name != source_name
        ]
        removed = len(ledger) - len(kept)
        if removed:
            self._save_profit_ledger(kept)
        return removed

    def _remove_inventory_rows_for_source(self, source_sheet_name: str) -> int:
        source_name = Path(str(source_sheet_name or "")).name.strip().lower()
        if not source_name:
            return 0
        ledger = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
        kept = [
            record
            for record in ledger
            if Path(str(record.get("source_sheet") or "")).name.strip().lower() != source_name
        ]
        removed = len(ledger) - len(kept)
        if removed:
            self._save_inventory_ledger(kept)
        return removed

    def _sheet_path_for_stage(self, kind: str, name: str) -> Path:
        if kind == "Working":
            return self.home_sheet_paths.get("Working", {}).get(name) or WORKING_SHEETS_DIR / name
        if kind == "Incoming":
            return self.home_sheet_paths.get("Incoming", {}).get(name) or INCOMING_SHEETS_DIR / name
        if kind == "Received":
            return self.received_sheet_paths.get(name) or RECEIVED_SHEETS_DIR / name
        return Path(name)

    def _sheet_path_is_visible_home_sheet(self, kind: str, path: Path) -> bool:
        expected = {
            "Incoming": INCOMING_SHEETS_DIR,
            "Working": WORKING_SHEETS_DIR,
            "Received": RECEIVED_SHEETS_DIR,
        }.get(kind)
        if expected is None:
            return False
        try:
            path.resolve().relative_to(expected.resolve())
            return path.suffix.lower() == ".xlsx"
        except Exception:
            return False

    def _move_fully_received_sheets_to_received(self, paths: list[Path]) -> list[str]:
        moved: list[str] = []
        for path in paths:
            if not path.exists():
                continue
            try:
                summary = summarize_workbook(path)
            except Exception:
                continue
            if not summary.get("all_received"):
                continue
            parent = path.parent.resolve()
            kind = "Incoming" if parent == INCOMING_SHEETS_DIR.resolve() else "Working" if parent == WORKING_SHEETS_DIR.resolve() else ""
            if not kind:
                continue
            old_key = self._home_sheet_key(kind, path.name)
            marker = dict(self.home_sheet_markers.get(old_key, {}))
            marker["all_received"] = True
            new_key = self._move_sheet_to_received(old_key)
            if new_key:
                self._delete_sheet_marker(old_key)
                self.home_sheet_markers[new_key] = marker
                moved.append(path.name)
        return moved

    def _archive_eligible_received_sheets(self) -> list[str]:
        if not RECEIVED_SHEETS_DIR.exists():
            return []
        archived: list[str] = []
        cutoff = datetime.now() - timedelta(days=14)
        RECEIVED_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
        ARCHIVED_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
        with shared_lock(CARD_PIPELINE_DIR, "sheet-archive", self.lucas_identity):
            latest_markers = self._load_sheet_markers()
            self.home_sheet_markers.update(latest_markers)
            for path in sorted(RECEIVED_SHEETS_DIR.glob("*.xlsx"), key=lambda item: item.name.lower()):
                key = self._home_sheet_key("Received", path.name)
                marker = dict(self.home_sheet_markers.get(key, {}))
                if not marker.get("paid"):
                    continue
                if not self._received_sheet_is_archive_age(path, marker, cutoff):
                    continue
                received_at = self._received_at_for_archive(path, marker)
                destination = self._unique_archive_path(path.name)
                shutil.move(str(path), str(destination))
                archived_key = self._home_sheet_key("Archived", destination.name)
                marker["all_received"] = True
                marker["paid"] = True
                marker.setdefault("received_at", received_at.isoformat(timespec="seconds"))
                marker["archived_at"] = datetime.now().isoformat(timespec="seconds")
                marker["archived_from"] = path.name
                self._delete_sheet_marker(key)
                self.home_sheet_markers[archived_key] = marker
                archived.append(destination.name)
            if archived:
                self._save_sheet_markers()
        return archived

    def _received_sheet_is_archive_age(self, path: Path, marker: dict[str, object], cutoff: datetime) -> bool:
        received_at = self._received_at_for_archive(path, marker)
        return received_at <= cutoff

    def _received_at_for_archive(self, path: Path, marker: dict[str, object]) -> datetime:
        raw = str(marker.get("received_at") or "").strip()
        if raw:
            try:
                return datetime.fromisoformat(raw)
            except ValueError:
                pass
        try:
            return datetime.fromtimestamp(path.stat().st_mtime)
        except OSError:
            return datetime.now()

    def _unique_archive_path(self, name: str) -> Path:
        destination = ARCHIVED_SHEETS_DIR / name
        if not destination.exists():
            return destination
        for index in range(2, 1000):
            candidate = ARCHIVED_SHEETS_DIR / f"{destination.stem}-{index}{destination.suffix}"
            if not candidate.exists():
                return candidate
        return ARCHIVED_SHEETS_DIR / f"{destination.stem}-{datetime.now().strftime('%Y%m%d%H%M%S')}{destination.suffix}"

    def _load_sheet_markers(self) -> dict[str, dict[str, object]]:
        try:
            if not SHEET_MARKERS_PATH.exists():
                return {}
            raw = json.loads(SHEET_MARKERS_PATH.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                return {str(key): dict(value) for key, value in raw.items() if isinstance(value, dict)}
        except Exception:
            return {}
        return {}

    def _delete_sheet_marker(self, key: str) -> None:
        if not key:
            return
        self.home_sheet_markers.pop(key, None)
        self.deleted_sheet_marker_keys.add(key)

    def _save_sheet_markers(self) -> None:
        with shared_lock(CARD_PIPELINE_DIR, "sheet-markers", self.lucas_identity):
            latest = self._load_sheet_markers()
            latest.update(self.home_sheet_markers)
            for key in self.deleted_sheet_marker_keys - set(self.home_sheet_markers):
                latest.pop(key, None)
            self.deleted_sheet_marker_keys.clear()
            self.home_sheet_markers = latest
            atomic_write_json(SHEET_MARKERS_PATH, self.home_sheet_markers)

    def _show_mode(self) -> None:
        for child in self.mode_host.winfo_children():
            child.destroy()
        mode = self.input_mode.get()
        if mode == "Barcode Scanner":
            self._build_barcode_mode()
            self.after(100, self._arm_scanner)
        elif mode == "Manual Entry":
            self._build_manual_intake_mode()
        elif mode == "Photo OCR":
            self._build_file_mode(photo=True)
        else:
            self._build_file_mode(photo=False)
        if hasattr(self, "intake_tree"):
            self._refresh_table()
        self.after_idle(self._refresh_tab_scroll_bindings)

    def _show_review_mode(self) -> None:
        if not hasattr(self, "review_mode_host"):
            return
        for child in self.review_mode_host.winfo_children():
            child.destroy()
        if self.review_mode.get() == "Manual Receive":
            self._build_manual_review_mode()
        else:
            self._build_automatic_review_mode()
        self._refresh_table()
        self.after_idle(self._refresh_tab_scroll_bindings)

    def _build_manual_review_mode(self) -> None:
        self.review_mode_host.columnconfigure(8, weight=1)
        ttk.Label(self.review_mode_host, text="Double-click cells in the Receive table to enter certs, raw Item IDs, or adjust matched details.", style="Muted.TLabel").grid(row=0, column=0, columnspan=9, sticky="w")

    def _build_manual_intake_mode(self) -> None:
        self.scanning_station_active = False
        self.scan_entry = None
        self.mode_host.columnconfigure(8, weight=1)
        ttk.Label(self.mode_host, text="Use the + Add row line in the Create table, then double-click cells to enter certs, card details, purchase, or CY fields.", style="Muted.TLabel").grid(row=0, column=0, columnspan=9, sticky="w")

    def _build_automatic_review_mode(self) -> None:
        self.review_mode_host.columnconfigure(8, weight=1)
        ttk.Label(self.review_mode_host, text="Input", style="Panel.TLabel").grid(row=0, column=0, sticky="w")
        selector = ttk.Combobox(
            self.review_mode_host,
            textvariable=self.review_input_mode,
            state="readonly",
            values=["Barcode Scanner", "Photo OCR"],
            width=18,
        )
        selector.grid(row=0, column=1, sticky="w", padx=(8, 16))
        selector.bind("<<ComboboxSelected>>", lambda _event: self._show_review_mode())
        if self.review_input_mode.get() == "Photo OCR":
            self._build_review_photo_controls(start_col=2)
        else:
            self._build_review_barcode_controls(start_col=2)

    def _build_review_barcode_controls(self, start_col: int) -> None:
        self.review_station_button = ttk.Button(self.review_mode_host, text="Enter Receive Scanning Mode", command=self.toggle_review_scanning, style="Primary.TButton")
        self.review_station_button.grid(row=0, column=start_col, sticky="w", padx=(0, 14))
        ttk.Label(self.review_mode_host, text="Scan", style="Panel.TLabel").grid(row=0, column=start_col + 1, sticky="w")
        self.review_scan_entry = ttk.Entry(self.review_mode_host, textvariable=self.review_scan_cert, width=28)
        self.review_scan_entry.grid(row=0, column=start_col + 2, sticky="w", padx=(8, 14))
        self.review_scan_entry.bind("<Return>", lambda _event: self.add_review_scanned_row())
        self.review_scan_entry.bind("<KP_Enter>", lambda _event: self.add_review_scanned_row())
        self._set_review_station_controls()
        if self.review_scanning_active:
            self.after(100, self._arm_review_scanner)

    def _build_review_photo_controls(self, start_col: int) -> None:
        self.review_scanning_active = False
        self.review_scan_entry = None
        ttk.Button(self.review_mode_host, text="Add Receive Photos", command=self.add_review_photos, style="Soft.TButton").grid(row=0, column=start_col, sticky="w", padx=(0, 8))
        ttk.Button(self.review_mode_host, text="Scan Receive Photos", command=self.scan_review_photos, style="Primary.TButton").grid(row=0, column=start_col + 1, sticky="w", padx=(0, 8))
        ttk.Button(self.review_mode_host, text="Clear Receive Photos", command=self.clear_review_photos, style="Soft.TButton").grid(row=0, column=start_col + 2, sticky="w")
        ttk.Label(self.review_mode_host, textvariable=self.review_photo_status, style="Muted.TLabel").grid(row=1, column=0, columnspan=9, sticky="w", pady=(10, 0))

    def _build_barcode_mode(self) -> None:
        self.mode_host.columnconfigure(7, weight=1)
        self.station_button = ttk.Button(self.mode_host, text="Enter Scanning Station Mode", command=self.toggle_scanning_station, style="Primary.TButton")
        self.station_button.grid(row=0, column=0, sticky="w", padx=(0, 14))
        ttk.Label(self.mode_host, text="Scan", style="Panel.TLabel").grid(row=0, column=1, sticky="w")
        self.scan_entry = ttk.Entry(self.mode_host, textvariable=self.scan_cert, width=28)
        self.scan_entry.grid(row=0, column=2, sticky="w", padx=(8, 14))
        self.scan_entry.bind("<Return>", lambda _event: self.add_scanned_row())
        self.scan_entry.bind("<KP_Enter>", lambda _event: self.add_scanned_row())
        ttk.Label(self.mode_host, text="Grader", style="Panel.TLabel").grid(row=0, column=3, sticky="w")
        ttk.Combobox(self.mode_host, textvariable=self.scan_grader, values=["PSA", "BGS", "SGC", "CGC"], state="readonly", width=8).grid(row=0, column=4, sticky="w", padx=(8, 14))
        ttk.Label(self.mode_host, textvariable=self.scan_status, style="Muted.TLabel").grid(row=1, column=0, columnspan=8, sticky="w", pady=(10, 0))
        self._set_station_controls()

    def _build_file_mode(self, photo: bool) -> None:
        if photo:
            self._build_photo_mode()
            return
        label = "Photo OCR Export" if photo else "Spreadsheet"
        self.mode_host.columnconfigure(1, weight=1)
        ttk.Label(self.mode_host, text=label, style="Panel.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Entry(self.mode_host, textvariable=self.file_path).grid(row=0, column=1, sticky="ew", padx=(0, 8))
        ttk.Button(self.mode_host, text="Browse", command=self.browse_file, style="Soft.TButton").grid(row=0, column=2, sticky="e")
        ttk.Label(self.mode_host, text="Sheet", style="Panel.TLabel").grid(row=0, column=3, sticky="w", padx=(14, 8))
        self.sheet_combo = ttk.Combobox(self.mode_host, textvariable=self.sheet_name, state="readonly", width=18)
        self.sheet_combo.grid(row=0, column=4, sticky="w")
        ttk.Button(self.mode_host, text="Load Rows", command=self.load_file_rows, style="Primary.TButton").grid(row=0, column=5, sticky="e", padx=(14, 0))

    def _build_photo_mode(self) -> None:
        self.mode_host.columnconfigure(4, weight=1)
        ttk.Button(self.mode_host, text="Add Photos", command=self.add_photos, style="Soft.TButton").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Button(self.mode_host, text="Add Folder", command=self.add_photo_folder, style="Soft.TButton").grid(row=0, column=1, sticky="w", padx=(0, 8))
        ttk.Button(self.mode_host, text="Scan Photos", command=self.scan_photos, style="Primary.TButton").grid(row=0, column=2, sticky="w", padx=(0, 8))
        ttk.Button(self.mode_host, text="Clear Photos", command=self.clear_photos, style="Soft.TButton").grid(row=0, column=3, sticky="w", padx=(0, 8))
        ttk.Label(self.mode_host, textvariable=self.photo_status, style="Muted.TLabel").grid(row=1, column=0, columnspan=5, sticky="w", pady=(10, 0))

    def add_scanned_row(self) -> None:
        if not self.scanning_station_active:
            self.scan_status.set("Click Enter Scanning Station Mode before scanning.")
            return
        cert = scan_to_cert(self.scan_cert.get())
        if not cert:
            self.scan_status.set("No cert detected. Scan again.")
            self._arm_scanner()
            return
        grader = self.scan_grader.get().strip().upper()
        card = self.scan_card.get().strip()
        added_rows = self._append_rows([
            {
                "cert_number": cert,
                "grader": grader,
                "card_title": card,
                "purchase_price": None,
                "source": "Barcode",
                "notes": "" if cert and grader else "Missing cert or grader",
            }
        ])
        self.scan_cert.set("")
        self.scan_card.set("")
        self.scan_status.set(f"Added row {len(self.intake_rows) + 1}: {cert}. Scanner ready for next cert.")
        self.status_var.set(f"Added scanned card {cert}.")
        if added_rows:
            self._select_excel_row(added_rows[-1])
        self._arm_scanner()

    def add_manual_intake_row(self) -> int | None:
        added_rows = self._append_rows([
            {
                "cert_number": "",
                "grader": "",
                "card_title": "",
                "purchase_price": None,
                "source": "Manual",
                "notes": "Manual create row",
            }
        ])
        if added_rows:
            row_id = str(added_rows[-1])
            self.intake_tree.selection_set(row_id)
            self.intake_tree.focus(row_id)
            self.intake_tree.see(row_id)
            self.status_var.set("Manual create row added. Double-click cells to edit it.")
            return added_rows[-1]
        return None

    def browse_file(self) -> None:
        path = filedialog.askopenfilename(title="Choose workbook", filetypes=[("Excel workbook", "*.xlsx")])
        if not path:
            return
        self.file_path.set(path)
        try:
            names = workbook_sheet_names(Path(path))
            self.sheet_combo["values"] = names
            if names:
                self.sheet_name.set(names[0])
        except Exception as error:
            messagebox.showerror("Workbook error", str(error))

    def load_file_rows(self) -> None:
        path = Path(self.file_path.get())
        if not path.exists():
            messagebox.showinfo("Choose file", "Choose a workbook first.")
            return
        try:
            if self.input_mode.get() == "Photo OCR":
                rows = read_photo_export(path, self.sheet_name.get() or None)
            else:
                rows = read_simple_spreadsheet(path, self.sheet_name.get() or None)
        except Exception as error:
            messagebox.showerror("Load failed", self._create_sheet_load_error(error))
            return
        if not rows:
            messagebox.showinfo("No usable rows", self._create_sheet_no_rows_message(path))
            self.status_var.set(f"No usable rows found in {path.name}.")
            return
        self._append_rows(rows)
        self.status_var.set(f"Loaded {len(rows)} row(s) from {path.name}.")

    def _create_sheet_load_error(self, error: Exception) -> str:
        raw = str(error).strip()
        if isinstance(error, (TypeError, ValueError, KeyError, IndexError)):
            return (
                "This sheet does not match the Create import format.\n\n"
                "Expected either a simple sheet with Cert #, Card Description, and Purchase Price columns, "
                "or a Photo OCR export with certification/card fields.\n\n"
                f"Details: {raw or type(error).__name__}"
            )
        return raw or type(error).__name__

    def _create_sheet_no_rows_message(self, path: Path) -> str:
        return (
            f"No usable card rows were found in {path.name}.\n\n"
            "Create can import a simple sheet with Cert #, Card Description, and Purchase Price columns, "
            "or a Photo OCR export with certification/card fields."
        )

    def toggle_scanning_station(self) -> None:
        self.scanning_station_active = not self.scanning_station_active
        self._set_station_controls()
        if self.scanning_station_active:
            self.scan_status.set("Scanning station armed. Scan certs now; each scan adds the next row.")
            self._arm_scanner()
        else:
            self.scan_status.set("Scanning station is off.")

    def _set_station_controls(self) -> None:
        if not hasattr(self, "station_button"):
            return
        self.station_button.configure(text="Exit Scanning Station Mode" if self.scanning_station_active else "Enter Scanning Station Mode")
        if self.scan_entry is not None:
            self.scan_entry.configure(state=tk.NORMAL if self.scanning_station_active else tk.DISABLED)

    def add_photos(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Choose card photos",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.webp *.bmp"), ("All files", "*.*")],
        )
        self._add_photo_paths([Path(path) for path in paths])

    def add_photo_folder(self) -> None:
        folder = filedialog.askdirectory(title="Choose photo folder")
        if not folder:
            return
        extensions = {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
        self._add_photo_paths([path for path in Path(folder).iterdir() if path.suffix.lower() in extensions])

    def clear_photos(self) -> None:
        if self.photo_worker and self.photo_worker.is_alive():
            messagebox.showinfo("Scan running", "Wait for the photo scan to finish before clearing photos.")
            return
        self.photo_paths = []
        self.photo_status.set("No photos selected.")

    def _add_photo_paths(self, paths: list[Path]) -> None:
        existing = {path.resolve() for path in self.photo_paths if path.exists()}
        added = 0
        for path in paths:
            if not path.exists() or path.resolve() in existing:
                continue
            self.photo_paths.append(path)
            existing.add(path.resolve())
            added += 1
        self.photo_status.set(f"{len(self.photo_paths)} photo(s) selected. Added {added}.")

    def scan_photos(self) -> None:
        if self.photo_worker and self.photo_worker.is_alive():
            messagebox.showinfo("Scan running", "Photo scan is already running.")
            return
        if not self.photo_paths:
            messagebox.showinfo("No photos", "Add photos before scanning.")
            return
        if genai is None or identify_cards_sync is None:
            messagebox.showerror("Missing dependency", "Photo OCR dependencies are not available.")
            return
        self._load_photo_env()
        api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
        if not api_key:
            messagebox.showerror("Missing GOOGLE_API_KEY", "Create .env in the L.U.C.A.S project folder or set GOOGLE_API_KEY.")
            return
        self.photo_client = make_photo_ocr_client(api_key)
        self.photo_status.set(f"Scanning 0/{len(self.photo_paths)} photo(s)...")
        self.photo_worker = threading.Thread(target=self._photo_scan_worker, daemon=True)
        self.photo_worker.start()

    def _photo_scan_worker(self) -> None:
        total = len(self.photo_paths)
        detected_total = 0
        for index, path in enumerate(list(self.photo_paths), start=1):
            try:
                self.events.put(("photo_status", f"Scanning {index}/{total}: {path.name}..."))
                image_b64 = base64.b64encode(path.read_bytes()).decode("utf-8")
                cards = identify_cards_sync(
                    self.photo_client,
                    image_b64,
                    progress_callback=lambda message, i=index, n=total, p=path: self.events.put(
                        ("photo_status", f"Scanning {i}/{n}: {p.name} - {message}")
                    ),
                )
                self._inventory_photo_rescue_single_bgs_cert(cards, image_b64, client=self.photo_client)
                rows = [self._photo_card_to_row(path, card) for card in cards if self._photo_card_has_inventory(card)]
                detected_total += len(rows)
                self.events.put(("photo_rows", rows))
                self.events.put(("photo_status", f"Scanning {index}/{total}: {path.name} -> {len(rows)} card row(s)."))
            except (TemporaryModelUnavailable, ModelQuotaExceeded, ModelResponseParseError) as error:
                self.events.put(("photo_status", f"{path.name}: {error}"))
            except Exception as error:
                self.events.put(("photo_status", f"{path.name}: {error}"))
        self.events.put(("photo_status", f"Photo scan complete. Added {detected_total} card row(s)."))

    def _photo_card_to_row(self, path: Path, card: dict) -> dict[str, object]:
        grader = normalize_grader(card.get("grading_company"))
        title = build_card_title(
            {
                "description": "",
                "year": card.get("year"),
                "set": card.get("set"),
                "player": card.get("player"),
                "card_number": card.get("card_number"),
                "parallel": card.get("parallel"),
                "subset": card.get("subset") or card.get("attributes"),
                "grader": grader,
                "grade": card.get("grade"),
            }
        )
        cert = scan_to_cert(card.get("cert_number"))
        notes = clean_part(card.get("position") or "")
        if not any(card.get(key) for key in ("cert_number", "player", "year", "set", "card_number", "parallel", "subset", "grade", "label_text")):
            review_note = f"OCR review needed: {card.get('error')}" if card.get("error") else "detected slab - review needed"
            notes = clean_part("; ".join(part for part in (notes, review_note) if part))
        return {
            "cert_number": cert,
            "grader": grader or infer_grader(title),
            "card_title": title,
            "purchase_price": None,
            "source": f"Photo: {path.name}",
            "notes": notes,
        }

    def _photo_card_has_inventory(self, card: dict) -> bool:
        if any(card.get(key) for key in ("cert_number", "player", "year", "set", "card_number", "parallel", "subset", "grade", "label_text")):
            return True
        return bool(card.get("is_graded_slab") or card.get("detection_confidence") or card.get("error"))

    def _load_photo_env(self) -> None:
        if not load_dotenv:
            return
        load_dotenv(ROOT / ".env", override=False)
        load_dotenv(PHOTO_APP_DIR / ".env", override=False)
        load_dotenv(PHOTO_APP_ROOT / ".env", override=False)

    def _inventory_photo_source_folder(self) -> Path:
        configured = str(self.app_settings.get("inventory_photo_folder") or "").strip() if hasattr(self, "app_settings") else ""
        return Path(configured).expanduser() if configured else INVENTORY_PHOTOS_DIR

    def _inventory_photo_shared_folder(self) -> Path:
        return INVENTORY_PHOTOS_DIR

    def _inventory_photo_picker_initial_dir(self) -> Path:
        for folder in (self._inventory_photo_source_folder(), self._inventory_photo_shared_folder(), ROOT):
            try:
                folder.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
            if folder.exists():
                return folder
        return ROOT

    def choose_inventory_photo_folder(self) -> None:
        current = self._inventory_photo_source_folder()
        try:
            current.mkdir(parents=True, exist_ok=True)
        except Exception:
            current = INVENTORY_PHOTOS_DIR
        selected = filedialog.askdirectory(
            title="Choose Inventory Photo Folder",
            initialdir=str(current if current.exists() else ROOT),
        )
        if not selected:
            return
        folder = Path(selected).expanduser()
        settings = load_app_settings()
        settings["inventory_photo_folder"] = str(folder)
        save_app_settings(settings)
        self.app_settings = settings
        try:
            count = len(self._inventory_photo_files(folder))
        except Exception:
            count = 0
        self.inventory_status_var.set(f"Inventory photo folder set to {folder}. Found {count} photo file(s).")
        self.status_var.set(f"Inventory photo folder set to {folder}.")

    def _inventory_photo_status(self, message: str, background: bool = False) -> None:
        if background:
            self.events.put(("inventory_photo_status", message))
            return
        self.inventory_status_var.set(message)

    def _prepare_inventory_photo_scan_folder(self, manual: bool = False, background: bool = False) -> Path | None:
        source = self._inventory_photo_source_folder()
        shared = self._inventory_photo_shared_folder()
        if not source.exists():
            try:
                source_is_shared = source.resolve(strict=False) == shared.resolve(strict=False)
            except Exception:
                source_is_shared = source == shared
            if source_is_shared:
                shared.mkdir(parents=True, exist_ok=True)
                return shared
            if manual:
                messagebox.showerror("Photo folder missing", f"Inventory photo folder does not exist:\n{source}")
            self._inventory_photo_status(f"Inventory photo folder does not exist: {source}", background=background)
            return None
        try:
            source_resolved = source.resolve()
            shared_resolved = shared.resolve() if shared.exists() else shared
        except Exception:
            source_resolved = source
            shared_resolved = shared
        if source_resolved == shared_resolved:
            return shared
        try:
            source_images = self._inventory_photo_paths(source)
            total = len(source_images)
            shared.mkdir(parents=True, exist_ok=True)
            copied = 0
            skipped = 0
            if not total:
                self._inventory_photo_status(f"No inventory photos found in {source}. Scanning {shared}...", background=background)
                if not background:
                    self.update_idletasks()
            for index, source_path in enumerate(source_images, start=1):
                try:
                    relative = source_path.relative_to(source)
                except Exception:
                    relative = Path(source_path.name)
                if relative.is_absolute() or ".." in relative.parts:
                    relative = Path(source_path.name)
                destination = shared / relative
                destination.parent.mkdir(parents=True, exist_ok=True)
                self._inventory_photo_status(f"Mirroring inventory photos: {index}/{total} {source_path.name}", background=background)
                if background:
                    self.events.put(("status", f"Mirroring inventory photos: {index}/{total}"))
                else:
                    self.status_var.set(f"Mirroring inventory photos: {index}/{total}")
                    self.update_idletasks()
                if destination.exists():
                    try:
                        if destination.stat().st_size == source_path.stat().st_size:
                            skipped += 1
                            continue
                    except Exception:
                        pass
                shutil.copy2(source_path, destination)
                copied += 1
            self._inventory_photo_status(f"Mirrored {copied} photo(s) to shared folder; skipped {skipped}. Scanning {shared}...", background=background)
            if background:
                self.events.put(("status", f"Mirrored inventory photos to {shared}."))
            else:
                self.status_var.set(f"Mirrored inventory photos to {shared}.")
                self.update_idletasks()
            return shared
        except Exception as error:
            if manual:
                messagebox.showerror("Mirror failed", f"Could not mirror inventory photos:\n{error}")
            self._inventory_photo_status(f"Inventory photo mirror failed: {error}", background=background)
            return None

    def mirror_inventory_photos_to_shared(self) -> None:
        if self.inventory_photo_worker and self.inventory_photo_worker.is_alive():
            messagebox.showinfo("Scan running", "Inventory photo scan is already running.")
            return
        folder = self._prepare_inventory_photo_scan_folder(manual=True)
        if folder:
            self.scan_inventory_photos(manual=True, folder=folder)

    def _load_inventory_photo_state(self) -> dict[str, object]:
        if not INVENTORY_PHOTO_STATE_PATH.exists():
            return {"version": 1, "photos": {}}
        try:
            raw = json.loads(INVENTORY_PHOTO_STATE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {"version": 1, "photos": {}}
        if not isinstance(raw, dict):
            return {"version": 1, "photos": {}}
        raw.setdefault("version", 1)
        if not isinstance(raw.get("photos"), dict):
            raw["photos"] = {}
        return raw

    def _save_inventory_photo_state(self, state: dict[str, object]) -> None:
        INVENTORY_PHOTO_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
        state["updated_at"] = datetime.now().isoformat(timespec="seconds")
        atomic_write_json(INVENTORY_PHOTO_STATE_PATH, state)

    def _inventory_photo_file_hash(self, path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _inventory_photo_paths(self, folder: Path) -> list[Path]:
        allowed = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}
        if not folder.exists():
            return []
        return [
            path
            for path in sorted(folder.rglob("*"), key=lambda item: str(item).lower())
            if path.is_file() and path.suffix.lower() in allowed
        ]

    def _inventory_photo_files(self, folder: Path) -> list[dict[str, object]]:
        images: list[dict[str, object]] = []
        if hasattr(self, "_inventory_photo_paths"):
            paths = self._inventory_photo_paths(folder)
        else:
            allowed = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}
            paths = [
                path
                for path in sorted(folder.rglob("*"), key=lambda item: str(item).lower())
                if path.is_file() and path.suffix.lower() in allowed
            ] if folder.exists() else []
        for path in paths:
            try:
                stat = path.stat()
                images.append(
                    {
                        "path": str(path),
                        "relative_path": path.relative_to(folder).as_posix(),
                        "filename": path.name,
                        "size": stat.st_size,
                        "modified": int(stat.st_mtime),
                        "sha256": self._inventory_photo_file_hash(path),
                    }
                )
            except Exception:
                continue
        return images

    def _inventory_photo_base64(self, path: Path) -> str:
        if path.suffix.lower() not in {".heic", ".heif"}:
            return base64.b64encode(path.read_bytes()).decode("utf-8")
        try:
            import pillow_heif

            pillow_heif.register_heif_opener()
        except Exception as error:
            raise RuntimeError("HEIC inventory photos need pillow-heif. Rerun install_dependencies, or change the iPhone Shortcut to export JPEG files.") from error
        from PIL import Image, ImageOps
        import io

        with Image.open(path) as opened:
            converted = ImageOps.exif_transpose(opened)
            if converted.mode not in {"RGB", "L"}:
                converted = converted.convert("RGB")
            buffer = io.BytesIO()
            converted.save(buffer, format="JPEG", quality=92)
        return base64.b64encode(buffer.getvalue()).decode("utf-8")

    def _inventory_photo_certs_from_cards(self, cards: list[dict]) -> set[str]:
        certs: set[str] = set()
        for card in cards:
            if not isinstance(card, dict):
                continue
            cert_values = [
                card.get("cert_number"),
                card.get("cert"),
                card.get("certification_number"),
                card.get("item_id"),
            ]
            for value in cert_values:
                text = str(value or "")
                direct = scan_to_cert(text)
                if direct:
                    certs.add(direct)
            label_text = str(card.get("label_text") or "")
            for match in re.findall(r"\b\d{5,12}\b", label_text):
                cert = scan_to_cert(match)
                if cert:
                    certs.add(cert)
        return certs

    def _inventory_photo_rescue_single_bgs_cert(self, cards: list[dict], image_b64: str, client: object | None = None) -> set[str]:
        if _verify_cert_only_sync is None or len(cards) != 1 or not isinstance(cards[0], dict):
            return set()
        card = cards[0]
        if scan_to_cert(card.get("cert_number")):
            return set()
        company = str(card.get("grading_company") or "").strip().upper()
        label_text = str(card.get("label_text") or "")
        if company != "BGS" and "BGS" not in label_text.upper() and "BECKETT" not in label_text.upper():
            return set()
        ocr_client = client or getattr(self, "inventory_photo_client", None) or getattr(self, "photo_client", None)
        if ocr_client is None:
            return set()
        try:
            verification = _verify_cert_only_sync(ocr_client, image_b64)
        except Exception as error:
            app_debug_log(f"BGS full-photo cert rescue skipped: {error}")
            return set()
        verified_cert = scan_to_cert(verification.get("cert_number"))
        if not verified_cert:
            return set()
        verified_company = str(verification.get("grading_company") or "").strip().upper()
        if verified_company and verified_company not in {"BGS", "BECKETT", "UNKNOWN"}:
            return set()
        card["cert_number"] = verified_cert
        card["cert_verified"] = "YES"
        if verification.get("label_text") and not card.get("label_text"):
            card["label_text"] = str(verification.get("label_text") or "")
        return {verified_cert}

    def _active_inventory_keys_by_cert(self, rows: list[dict[str, object]]) -> dict[str, list[str]]:
        by_cert: dict[str, list[str]] = defaultdict(list)
        for record in rows:
            if str(record.get("status") or "").lower() != "active":
                continue
            cert = scan_to_cert(record.get("cert_number"))
            key = str(record.get("inventory_key") or "")
            if cert and key:
                by_cert[cert].append(key)
        return by_cert

    def _inventory_photo_match_keys(
        self,
        certs: set[str],
        cards: list[dict],
        keys_by_cert: dict[str, list[str]],
        records_by_key: dict[str, dict[str, object]],
    ) -> set[str]:
        matched = {key for cert in certs for key in keys_by_cert.get(cert, [])}
        if matched:
            return matched
        if certs:
            return set()
        for card in cards:
            if not isinstance(card, dict):
                continue
            key = self._inventory_photo_best_title_match(card, records_by_key)
            if key:
                matched.add(key)
        return matched

    def _inventory_photo_best_title_match(self, card: dict, records_by_key: dict[str, dict[str, object]]) -> str:
        card_text = self._inventory_photo_card_match_text(card)
        if not card_text:
            return ""
        card_compact = self._compact_match_text(card_text)
        card_tokens = self._match_text_tokens(card_text)
        filename_hint_tokens = self._match_text_tokens(card.get("photo_filename_hint"))
        player_compact = self._compact_match_text(card.get("player") or card.get("subject") or "")
        year = str(card.get("year") or "").strip()
        set_compact = self._compact_match_text(card.get("set") or "")
        number_compact = self._compact_match_text(card.get("card_number") or "")
        parallel_compact = self._compact_match_text(card.get("parallel") or card.get("subset") or "")
        grade = str(card.get("grade") or "").strip()
        grader = self._compact_match_text(card.get("grading_company") or card.get("grader") or "")
        scored: list[tuple[float, str]] = []
        for key, record in records_by_key.items():
            title = str(record.get("card_title") or record.get("title") or "").strip()
            if not title:
                continue
            title_compact = self._compact_match_text(title)
            title_tokens = self._match_text_tokens(title)
            score = difflib.SequenceMatcher(None, card_compact, title_compact).ratio()
            evidence = 0.0
            if player_compact and player_compact in title_compact:
                evidence += 4.0
            if year and re.search(rf"\b{re.escape(year)}\b", title):
                evidence += 2.0
            if set_compact and set_compact in title_compact:
                evidence += 2.0
            if number_compact and number_compact in title_compact:
                evidence += 1.0
            if parallel_compact and parallel_compact in title_compact:
                evidence += 1.0
            if grade and re.search(rf"\b{re.escape(grade)}\b", title):
                evidence += 0.75
            if grader and grader in title_compact:
                evidence += 0.75
            overlap = card_tokens & title_tokens
            if len(overlap) >= 3:
                overlap_ratio = len(overlap) / max(1, len(card_tokens))
                if overlap_ratio >= 0.5:
                    evidence += min(5.0, len(overlap) * 1.0)
            if filename_hint_tokens and len(filename_hint_tokens & title_tokens) >= 3:
                hint_ratio = len(filename_hint_tokens & title_tokens) / max(1, len(filename_hint_tokens))
                if hint_ratio >= 0.75:
                    evidence += 1.5
            scored.append((evidence + score, key))
        if not scored:
            return ""
        scored.sort(reverse=True)
        best_score, best_key = scored[0]
        second_score = scored[1][0] if len(scored) > 1 else 0.0
        return best_key if best_score >= 5.5 and best_score >= second_score + 0.75 else ""

    def _inventory_photo_card_match_text(self, card: dict) -> str:
        values = [
            card.get("player"),
            card.get("subject"),
            card.get("year"),
            card.get("set"),
            card.get("card_number"),
            card.get("parallel"),
            card.get("subset"),
            card.get("grade"),
            card.get("grading_company"),
            card.get("label_text"),
            card.get("photo_filename_hint"),
        ]
        return " ".join(str(value or "") for value in values if str(value or "").strip())

    def _match_text_tokens(self, value: object) -> set[str]:
        tokens = set(re.findall(r"[a-z0-9]+", str(value or "").lower()))
        stop_words = {
            "the", "and", "with", "for", "card", "cards", "photo", "photos", "front", "back",
            "image", "img", "shot", "scan", "scans", "psa", "bgs", "sgc", "cgc", "gem", "mint",
        }
        return {token for token in tokens if (len(token) >= 3 or token.isdigit()) and token not in stop_words}

    def _compact_match_text(self, value: object) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())

    def _inventory_photo_filename_hint(self, image: dict[str, object]) -> str:
        stem = Path(str(image.get("relative_path") or image.get("filename") or "")).stem.strip()
        if not stem:
            return ""
        stem = re.sub(r"(?i)^\[[^\]]+\][-_ ]*", "", stem)
        stem = re.sub(r"(?i)^card\[\d+\][-_ ]*\[\d+\][-_ ]*", "", stem)
        stem = re.sub(r"(?i)^card[-_ ]*\d+[-_ ]*\d+[-_ ]*", "", stem)
        stem = re.sub(r"(?i)^card\[\d+\][-_ ]*", "", stem)
        stem = re.sub(r"(?i)^card[-_ ]*\d+[-_ ]*", "", stem)
        stem = re.sub(r"(?i)^\[\d+\][-_ ]*", "", stem)
        stem = re.sub(r"(?i)^(?:photo|img|shot|scan|p)?[-_ ]*\d+[-_ ]*", "", stem)
        stem = re.sub(r"(?i)[-_ ]+(?:photo|img|shot|scan|p)[-_ ]*\d+$", "", stem)
        stem = re.sub(r"[\[\]{}()]+", " ", stem)
        stem = re.sub(r"[_-]+", " ", stem)
        stem = re.sub(r"\s+", " ", stem).strip()
        return stem if len(self._match_text_tokens(stem)) >= 2 else ""


    def _link_inventory_photo_to_keys(self, keys: set[str], photo_path: Path) -> int:
        if not keys:
            return 0
        storage_value = getattr(self, "_inventory_photo_storage_value", None)
        path_text = storage_value(photo_path) if callable(storage_value) else str(photo_path)
        with shared_lock(CARD_PIPELINE_DIR, "inventory-photo-link", self.lucas_identity):
            rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
            changed = 0
            for record in rows:
                if str(record.get("inventory_key") or "") not in keys:
                    continue
                photo_paths = list(record.get("photo_paths") or [])
                if path_text not in photo_paths:
                    if len(photo_paths) >= MAX_INVENTORY_PHOTOS_PER_CARD:
                        continue
                    photo_paths.append(path_text)
                    record["photo_paths"] = photo_paths
                    record["photo_count"] = len(photo_paths)
                    changed += 1
            if changed:
                self._save_inventory_ledger(rows)
        return changed

    def _inventory_photo_relative_path(self, path: Path) -> Path | None:
        try:
            resolved = path.resolve()
        except Exception:
            resolved = path
        for root in (self._inventory_photo_shared_folder(), self._inventory_photo_source_folder()):
            try:
                return resolved.relative_to(root.resolve())
            except Exception:
                continue
        parts = list(path.parts)
        for index, part in enumerate(parts):
            if part.lower() == "inventory photos" and index + 1 < len(parts):
                return Path(*parts[index + 1 :])
        return None

    def _inventory_photo_storage_value(self, path: Path) -> str:
        relative = self._inventory_photo_relative_path(path)
        if relative and not relative.is_absolute() and ".." not in relative.parts:
            return relative.as_posix()
        return str(path)

    def _inventory_photo_path_candidates(self, path_value: object) -> list[Path]:
        text = str(path_value or "").strip()
        path = Path(text).expanduser()
        candidates: list[Path] = []
        if path.is_absolute():
            candidates.append(path)
            relative = self._inventory_photo_relative_path(path)
            if relative and not relative.is_absolute() and ".." not in relative.parts:
                candidates.extend([self._inventory_photo_shared_folder() / relative, self._inventory_photo_source_folder() / relative])
            elif path.name:
                candidates.extend([self._inventory_photo_shared_folder() / path.name, self._inventory_photo_source_folder() / path.name])
        else:
            candidates.extend([self._inventory_photo_shared_folder() / path, self._inventory_photo_source_folder() / path, path])
        seen: set[str] = set()
        unique: list[Path] = []
        for candidate in candidates:
            key = str(candidate)
            if key in seen:
                continue
            seen.add(key)
            unique.append(candidate)
        return unique

    def _resolve_inventory_photo_path(self, path_value: object) -> Path:
        candidates = self._inventory_photo_path_candidates(path_value)
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return candidates[0] if candidates else Path(str(path_value or "")).expanduser()

    def _inventory_record_references_photo(self, record: dict[str, object], photo_path: Path) -> bool:
        expected = self._inventory_photo_storage_value(photo_path)
        try:
            expected_resolved = str(photo_path.resolve()).lower()
        except Exception:
            expected_resolved = str(photo_path).lower()
        for value in record.get("photo_paths") or []:
            if str(value) == expected:
                return True
            for candidate in self._inventory_photo_path_candidates(value):
                try:
                    if str(candidate.resolve()).lower() == expected_resolved:
                        return True
                except Exception:
                    continue
        return False

    def _inventory_photo_scan_can_skip(self, existing: dict[str, object], records_by_key: dict[str, dict[str, object]], photo_path: Path) -> bool:
        if not existing:
            return False
        status = str(existing.get("status") or "").strip()
        linked_keys = [str(key).strip() for key in (existing.get("linked_keys") or []) if str(key).strip()]
        if status not in {"linked", "missing_from_album", "archived_from_album"} or not linked_keys:
            return False
        for key in linked_keys:
            record = records_by_key.get(str(key))
            if record and self._inventory_record_references_photo(record, photo_path):
                return True
        return True

    def _inventory_photo_capture_group_key(self, image: dict[str, object]) -> str:
        stem = Path(str(image.get("relative_path") or image.get("filename") or "")).stem.strip()
        if not stem:
            return ""
        match = re.match(r"(?i)^(.+?(?:card|group|item)[-_\s\[]*\d+\]?)(?:[-_\s\[]+(?:photo|img|shot|p)?[-_\s\[]*\d+\]?)*(?:[-_\s]+.*)?$", stem)
        if not match:
            return ""
        return self._compact_match_text(match.group(1))

    def _inventory_photo_scan_group_nearby_unmatched(
        self,
        images: list[dict[str, object]],
        photos: dict[str, object],
    ) -> int:
        ordered = sorted(
            [image for image in images if str(image.get("sha256") or "")],
            key=lambda image: (int(image.get("modified") or 0), str(image.get("relative_path") or image.get("filename") or "")),
        )
        if not ordered:
            return 0
        linked_count = 0
        grouped_hashes: set[str] = set()
        for index, image in enumerate(ordered):
            sha = str(image.get("sha256") or "")
            entry = photos.get(sha)
            if not isinstance(entry, dict) or str(entry.get("status") or "") != "linked":
                continue
            if entry.get("auto_grouped_from"):
                continue
            anchor_keys = {str(key) for key in (entry.get("linked_keys") or []) if str(key).strip()}
            if not anchor_keys:
                continue
            anchor_group_key = self._inventory_photo_capture_group_key(image)
            latest_rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
            records_by_key = {str(record.get("inventory_key") or ""): record for record in latest_rows if str(record.get("status") or "").lower() == "active"}
            remaining = MAX_INVENTORY_PHOTOS_PER_CARD
            for key in anchor_keys:
                record = records_by_key.get(key)
                if record:
                    remaining = min(remaining, max(0, MAX_INVENTORY_PHOTOS_PER_CARD - len(record.get("photo_paths") or [])))
            if remaining <= 0:
                continue
            anchor_modified = int(image.get("modified") or 0)
            candidates: list[tuple[int, int, dict[str, object]]] = []
            for candidate_index, candidate in enumerate(ordered):
                if candidate_index == index:
                    continue
                candidate_sha = str(candidate.get("sha256") or "")
                if not candidate_sha or candidate_sha in grouped_hashes:
                    continue
                candidate_entry = photos.get(candidate_sha)
                if not isinstance(candidate_entry, dict):
                    continue
                if candidate_entry.get("linked_keys") or str(candidate_entry.get("status") or "") == "linked":
                    continue
                candidate_certs = {scan_to_cert(cert) for cert in (candidate_entry.get("certs") or []) if scan_to_cert(cert)}
                if candidate_certs:
                    continue
                same_group = bool(anchor_group_key and self._inventory_photo_capture_group_key(candidate) == anchor_group_key)
                if not same_group:
                    continue
                distance = abs(int(candidate.get("modified") or 0) - anchor_modified)
                candidates.append((distance, abs(candidate_index - index), candidate))
            candidates.sort(key=lambda item: (item[0], item[1]))
            for _distance, _position_distance, candidate in candidates[:remaining]:
                candidate_sha = str(candidate.get("sha256") or "")
                candidate_path = Path(str(candidate.get("path") or ""))
                changed = self._link_inventory_photo_to_keys(anchor_keys, candidate_path)
                if not changed:
                    continue
                candidate_entry = photos.get(candidate_sha)
                if not isinstance(candidate_entry, dict):
                    candidate_entry = {}
                photos[candidate_sha] = {
                    **candidate,
                    **candidate_entry,
                    "linked_keys": sorted(anchor_keys),
                    "status": "linked",
                    "auto_grouped_from": str(image.get("relative_path") or image.get("filename") or sha),
                    "auto_grouped_at": datetime.now().isoformat(timespec="seconds"),
                    "last_seen": datetime.now().isoformat(timespec="seconds"),
                }
                grouped_hashes.add(candidate_sha)
                linked_count += changed
        return linked_count

    def scan_inventory_photos(self, manual: bool = False, folder: Path | None = None) -> None:
        if self.inventory_photo_worker and self.inventory_photo_worker.is_alive():
            if manual:
                messagebox.showinfo("Scan running", "Inventory photo scan is already running.")
            return
        if not manual and folder is None:
            self.inventory_status_var.set("Inventory photo background scan queued.")
            self.inventory_photo_worker = threading.Thread(target=self._inventory_photo_background_scan_worker, daemon=True)
            self.inventory_photo_worker.start()
            return
        folder = Path(folder).expanduser() if folder else self._prepare_inventory_photo_scan_folder(manual=manual)
        if not folder:
            return
        folder.mkdir(parents=True, exist_ok=True)
        if genai is None or identify_cards_sync is None:
            self.inventory_status_var.set("Inventory photo scan unavailable: missing photo OCR dependencies.")
            return
        self._load_photo_env()
        api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
        if not api_key:
            self.inventory_status_var.set("Inventory photo scan unavailable: missing GOOGLE_API_KEY.")
            if manual:
                messagebox.showerror("Missing GOOGLE_API_KEY", "Create .env in the L.U.C.A.S project folder or set GOOGLE_API_KEY.")
            return
        self.inventory_photo_client = make_photo_ocr_client(api_key)
        self.inventory_status_var.set(f"Scanning inventory photos in {folder}...")
        self.inventory_photo_worker = threading.Thread(target=self._inventory_photo_scan_worker, args=(folder, False), daemon=True)
        self.inventory_photo_worker.start()

    def _inventory_photo_background_scan_worker(self) -> None:
        folder = self._prepare_inventory_photo_scan_folder(manual=False, background=True)
        if not folder:
            return
        try:
            folder.mkdir(parents=True, exist_ok=True)
        except Exception as error:
            self.events.put(("inventory_photo_status", f"Inventory photo scan unavailable: {error}"))
            return
        if genai is None or identify_cards_sync is None:
            self.events.put(("inventory_photo_status", "Inventory photo scan unavailable: missing photo OCR dependencies."))
            return
        self._load_photo_env()
        api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
        if not api_key:
            self.events.put(("inventory_photo_status", "Inventory photo scan unavailable: missing GOOGLE_API_KEY."))
            return
        self.inventory_photo_client = make_photo_ocr_client(api_key)
        self.events.put(("inventory_photo_status", f"Inventory photo background scan started in {folder}."))
        self._inventory_photo_scan_worker(folder, True)

    def _inventory_photo_scan_worker(self, folder: Path, background: bool = False) -> None:
        started = time.perf_counter()
        linked = 0
        scanned = 0
        errors: list[str] = []
        skipped = 0
        grouped = 0
        state = self._load_inventory_photo_state()
        photos = state.setdefault("photos", {})
        try:
            images = self._inventory_photo_files(folder)
            total = len(images)
            if not background:
                self.events.put(("inventory_photo_status", f"Inventory photo scan starting in {folder}: 0/{total} file(s)."))
            current_hashes = {str(image.get("sha256") or "") for image in images}
            for sha, record in list(photos.items()):
                if sha and sha not in current_hashes and isinstance(record, dict) and not record.get("removed_at"):
                    record["status"] = "missing_from_album"
                    record["removed_at"] = datetime.now().isoformat(timespec="seconds")
            rows = [self._normalize_inventory_record(record) for record in self._load_inventory_ledger()]
            keys_by_cert = self._active_inventory_keys_by_cert(rows)
            active_records = [record for record in rows if str(record.get("status") or "").lower() == "active"]
            records_by_key = {str(record.get("inventory_key") or ""): record for record in active_records}
            sold_cert_source = getattr(self, "_sold_inventory_cert_numbers", None)
            sold_certs = sold_cert_source() if callable(sold_cert_source) else set()
            sold_photo_source = getattr(self, "_sold_inventory_photo_used_keys", None)
            sold_photo_paths, sold_photo_hashes = (
                sold_photo_source()
                if callable(sold_photo_source)
                else (set(), set())
            )
            last_background_status = 0.0
            for index, image in enumerate(images, start=1):
                image_label = str(image.get("relative_path") or image.get("filename") or "photo")
                status_message = f"Inventory photo scan: {index}/{total} {image_label}"
                now = time.monotonic()
                if not background or index == total or now - last_background_status >= 15:
                    self.events.put(("inventory_photo_status", status_message if not background else f"Inventory photo background scan: {index}/{total}"))
                    last_background_status = now
                app_debug_log(status_message)
                sha = str(image.get("sha256") or "")
                if not sha:
                    continue
                existing = photos.get(sha) if isinstance(photos.get(sha), dict) else {}
                image_path = Path(str(image.get("path") or ""))
                can_skip = getattr(self, "_inventory_photo_scan_can_skip", None)
                sold_cert_match = getattr(self, "_inventory_photo_state_matches_sold_cert", None)
                sold_cert_skip = bool(callable(sold_cert_match) and sold_cert_match(existing, sold_certs))
                sold_photo_match = getattr(self, "_inventory_photo_image_matches_sold_photo", None)
                sold_photo_skip = bool(callable(sold_photo_match) and sold_photo_match(image, sold_photo_paths, sold_photo_hashes))
                references_photo = getattr(self, "_inventory_record_references_photo", None)
                attached_records = [
                    record
                    for record in active_records
                    if callable(references_photo) and references_photo(record, image_path)
                ]
                if (callable(can_skip) and can_skip(existing, records_by_key, image_path)) or sold_cert_skip or sold_photo_skip or attached_records:
                    skipped += 1
                    existing["last_seen"] = datetime.now().isoformat(timespec="seconds")
                    if sold_cert_skip or sold_photo_skip:
                        existing["status"] = "sold_inventory"
                    elif attached_records:
                        existing["status"] = "linked"
                        existing["linked_keys"] = sorted(
                            {
                                str(record.get("inventory_key") or "")
                                for record in attached_records
                                if str(record.get("inventory_key") or "").strip()
                            }
                        )
                    photos[sha] = existing
                    self._save_inventory_photo_state(state)
                    continue
                try:
                    image_b64 = self._inventory_photo_base64(image_path)
                    cards = identify_cards_sync(self.inventory_photo_client, image_b64)
                    certs = self._inventory_photo_certs_from_cards(cards)
                    if not certs:
                        bgs_rescue = getattr(self, "_inventory_photo_rescue_single_bgs_cert", None)
                        if callable(bgs_rescue):
                            certs.update(bgs_rescue(cards, image_b64))
                    scanned += 1
                except Exception as error:
                    errors.append(f"{image.get('relative_path')}: {error}")
                    photos[sha] = {**image, "cards": [], "certs": [], "linked_keys": [], "status": "ocr_error", "error": str(error), "last_seen": datetime.now().isoformat(timespec="seconds")}
                    self._save_inventory_photo_state(state)
                    app_debug_log(f"Inventory photo scan OCR error: {image_label}: {error}")
                    continue
                filename_hint = self._inventory_photo_filename_hint(image)
                match_cards = list(cards)
                if filename_hint and not certs:
                    match_cards.append({"label_text": filename_hint, "photo_filename_hint": filename_hint})
                matched_keys = self._inventory_photo_match_keys(certs, match_cards, keys_by_cert, records_by_key)
                if matched_keys:
                    linked += self._link_inventory_photo_to_keys(matched_keys, image_path)
                photos[sha] = {
                    **image,
                    "cards": cards,
                    "certs": sorted(certs),
                    "filename_hint": filename_hint,
                    "linked_keys": sorted(matched_keys),
                    "status": "linked" if matched_keys else "no_matching_inventory",
                    "last_seen": datetime.now().isoformat(timespec="seconds"),
                }
                self._save_inventory_photo_state(state)
            if isinstance(photos, dict):
                grouped = self._inventory_photo_scan_group_nearby_unmatched(images, photos)
                if grouped:
                    linked += grouped
            self._save_inventory_photo_state(state)
            grouped_text = f", grouped {grouped}" if grouped else ""
            self.events.put(("inventory_photo_status", f"Inventory photo scan complete: {len(images)} file(s), skipped {skipped}, OCR scanned {scanned}, linked {linked}{grouped_text}."))
            if linked:
                self.events.put(("inventory_refresh", None))
        except Exception as error:
            errors.append(str(error))
            self.events.put(("inventory_photo_status", f"Inventory photo scan failed: {error}"))
        finally:
            if errors:
                record_performance_event("inventory.photos.errors", started, " | ".join(errors[:5]), force=True)
            record_performance_event("inventory.photos.scan", started, f"skipped={skipped} scanned={scanned} linked={linked} errors={len(errors)}")

    def refresh_incoming_index(self) -> None:
        try:
            INCOMING_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
            WORKING_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
            paths = sorted(
                [*INCOMING_SHEETS_DIR.glob("*.xlsx"), *WORKING_SHEETS_DIR.glob("*.xlsx")],
                key=lambda path: (path.parent.name.lower(), path.name.lower()),
            )
            ensure_raw_ids = getattr(self, "_ensure_raw_item_ids_in_sheet_paths", None)
            raw_id_result = ensure_raw_ids(paths) if callable(ensure_raw_ids) else {"ids_added": 0}
        except Exception as error:
            self.incoming_cert_index = {}
            self.review_status.set(f"Incoming sheets unavailable: {error}")
            return
        index: dict[str, dict[str, object]] = {}
        for path in paths:
            try:
                rows = read_simple_spreadsheet(path)
            except Exception:
                continue
            for row in rows:
                cert = scan_to_cert(row.get("cert_number"))
                candidate = {
                    "item_id": row.get("item_id") or "",
                    "cert_number": cert,
                    "sheet": path.name,
                    "path": path,
                    "workbook_sheet": row.get("workbook_sheet") or "",
                    "workbook_row": row.get("workbook_row"),
                    "card_title": row.get("card_title") or "",
                    "grader": row.get("grader") or "",
                    "sport": row.get("sport") or row.get("category") or "",
                    "category": row.get("category") or row.get("sport") or "",
                    "purchase_price": row.get("purchase_price"),
                    "card_ladder_value": row.get("card_ladder_value"),
                    "card_ladder_comps_average": row.get("card_ladder_comps_average"),
                    "cy_value": row.get("cy_value"),
                    "cy_confidence": row.get("cy_confidence"),
                    "card_ladder_comps": row.get("card_ladder_comps") or "",
                    "best_company": row.get("best_company") or "",
                    "estimated_payout": row.get("estimated_payout"),
                }
                receive_key = self._receive_row_ref_key(candidate)
                if receive_key:
                    candidate["receive_key"] = receive_key
                if not cert:
                    if receive_key:
                        index[receive_key] = candidate
                    continue
                existing = index.get(cert)
                if existing:
                    for key, value in candidate.items():
                        if (existing.get(key) is None or existing.get(key) == "") and value not in (None, ""):
                            existing[key] = value
                    continue
                index[cert] = candidate
        self.incoming_cert_index = index
        self._match_all_review_rows()
        self._refresh_table()
        cert_count = sum(1 for key in index if not str(key).startswith("raw:"))
        raw_count = len(index) - cert_count
        raw_added = int(raw_id_result.get("ids_added") or 0)
        raw_note = f" Added {raw_added} missing raw ID(s)." if raw_added else ""
        self.review_status.set(f"Indexed {cert_count} cert(s) and {raw_count} raw row(s) from {len(paths)} incoming/working sheet(s).{raw_note}")

    def refresh_received_sheets(self) -> None:
        try:
            RECEIVED_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
            paths = sorted(RECEIVED_SHEETS_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
        except Exception as error:
            self.received_sheet_paths = {}
            if hasattr(self, "received_sheet_combo"):
                self.received_sheet_combo["values"] = []
            self.review_status.set(f"Received sheets unavailable: {error}")
            return
        self.received_sheet_paths = {path.name: path for path in paths}
        if hasattr(self, "received_sheet_combo"):
            names = list(self.received_sheet_paths)
            self.received_sheet_combo["values"] = names
            if names and self.selected_received_sheet.get() not in self.received_sheet_paths:
                self.selected_received_sheet.set(names[0])
            elif not names:
                self.selected_received_sheet.set("")

    def load_selected_received_sheet_for_review(self) -> None:
        name = self.selected_received_sheet.get()
        path = self.received_sheet_paths.get(name)
        if not path:
            messagebox.showinfo("Choose sheet", "Choose a received sheet to load into Assignment.")
            return
        self.review_status.set(f"Loading received sheet: {name}...")
        self.status_var.set(f"Loading received sheet: {name}...")
        threading.Thread(target=self._load_received_sheet_worker, args=(name, path), daemon=True).start()

    def _load_received_sheet_worker(self, name: str, path: Path) -> None:
        try:
            rows = read_simple_spreadsheet(path)
        except Exception as error:
            self.events.put(("load_received_sheet_error", {"name": name, "error": str(error)}))
            return
        self.events.put(("load_received_sheet_done", {"name": name, "rows": rows}))

    def _apply_loaded_received_sheet(self, name: str, rows: list[dict[str, object]]) -> None:
        review_rows = []
        for row in rows:
            review_rows.append(
                {
                    "cert_number": row.get("cert_number"),
                    "grader": row.get("grader"),
                    "card_title": row.get("card_title"),
                    "purchase_price": row.get("purchase_price"),
                    "card_ladder_value": row.get("card_ladder_value"),
                    "card_ladder_comps_average": row.get("card_ladder_comps_average"),
                    "cy_value": row.get("cy_value"),
                    "cy_confidence": row.get("cy_confidence"),
                    "card_ladder_comps": row.get("card_ladder_comps") or "",
                    "best_company": row.get("best_company") or "",
                    "estimated_payout": row.get("estimated_payout"),
                    "source": f"Received Sheet: {name}",
                    "sheet_source": name,
                    "status": "Received",
                    "notes": "Loaded from received sheet",
                }
            )
        added = self._append_review_rows(review_rows, schedule_recommendations=True)
        self.review_status.set(f"Loaded {len(added)} row(s) from {name}.")
        self.status_var.set(f"Loaded received sheet: {name}")

    def add_manual_review_row(self) -> int | None:
        added_rows = self._append_review_rows([
            {
                "cert_number": "",
                "grader": "",
                "card_title": "",
                "purchase_price": None,
                "source": "Manual",
                "notes": "Manual assignment",
            }
        ])
        if added_rows:
            row_id = str(added_rows[-1])
            target_tree = self.receive_tree if hasattr(self, "receive_tree") else self.review_tree
            target_tree.selection_set(row_id)
            target_tree.focus(row_id)
            target_tree.see(row_id)
            self.review_status.set("Manual row added. Double-click cells to edit it.")
            return added_rows[-1]
        return None

    def toggle_review_scanning(self) -> None:
        self.review_scanning_active = not self.review_scanning_active
        self._set_review_station_controls()
        if self.review_scanning_active:
            self.review_status.set("Receive scanning mode armed. Scan received certs now.")
            self._arm_review_scanner()
        else:
            self.review_status.set("Receive station is off.")

    def _set_review_station_controls(self) -> None:
        if not hasattr(self, "review_station_button"):
            return
        self.review_station_button.configure(text="Exit Receive Scanning Mode" if self.review_scanning_active else "Enter Receive Scanning Mode")
        if self.review_scan_entry is not None:
            self.review_scan_entry.configure(state=tk.NORMAL if self.review_scanning_active else tk.DISABLED)

    def add_review_scanned_row(self) -> None:
        if not self.review_scanning_active:
            self.review_status.set("Click Enter Receive Scanning Mode before scanning.")
            return
        raw_input = str(self.review_scan_cert.get() or "").strip()
        cert_candidate = "" if raw_input.upper().startswith("RAW-") else scan_to_cert(raw_input)
        cert = cert_candidate if re.search(r"\d", cert_candidate) else ""
        if not cert:
            matches: list[dict[str, object]] = []
            if raw_input.upper().startswith("RAW-"):
                matches = self._incoming_raw_matches({"item_id": raw_input})
                if not matches:
                    self.refresh_incoming_index()
                    matches = self._incoming_raw_matches({"item_id": raw_input})
            elif raw_input:
                matches = self._incoming_title_matches(raw_input)
                if not matches:
                    self.refresh_incoming_index()
                    matches = self._incoming_title_matches(raw_input)
            if matches:
                rows = [self._receive_match_to_review_payload(match, "Receive Search") for match in matches]
                self._append_review_rows(rows)
                self.review_scan_cert.set("")
                label = raw_input if len(raw_input) <= 60 else f"{raw_input[:57]}..."
                self.review_status.set(f"Matched {len(matches)} incoming row(s) for {label}. Ready for next scan.")
                self._arm_review_scanner()
                return
            self.review_status.set("No cert, raw ID, or matching incoming card found. Scan or type again.")
            self._arm_review_scanner()
            return
        self._append_review_rows([
            {
                "cert_number": cert,
                "grader": "",
                "card_title": "",
                "purchase_price": None,
                "source": "Receive Barcode",
                "notes": "Received",
            }
        ])
        self.review_scan_cert.set("")
        self.review_status.set(f"Received {cert}. Ready for next scan.")
        self._arm_review_scanner()

    def add_review_photos(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Choose receive photos",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.webp *.bmp"), ("All files", "*.*")],
        )
        self._add_review_photo_paths([Path(path) for path in paths])

    def clear_review_photos(self) -> None:
        if self.review_photo_worker and self.review_photo_worker.is_alive():
            messagebox.showinfo("Scan running", "Wait for the receive photo scan to finish before clearing photos.")
            return
        self.review_photo_paths = []
        self.review_photo_status.set("No receive photos selected.")

    def _add_review_photo_paths(self, paths: list[Path]) -> None:
        existing = {path.resolve() for path in self.review_photo_paths if path.exists()}
        added = 0
        for path in paths:
            if not path.exists() or path.resolve() in existing:
                continue
            self.review_photo_paths.append(path)
            existing.add(path.resolve())
            added += 1
        self.review_photo_status.set(f"{len(self.review_photo_paths)} receive photo(s) selected. Added {added}.")

    def scan_review_photos(self) -> None:
        if self.review_photo_worker and self.review_photo_worker.is_alive():
            messagebox.showinfo("Scan running", "Receive photo scan is already running.")
            return
        if not self.review_photo_paths:
            messagebox.showinfo("No photos", "Add receive photos before scanning.")
            return
        if genai is None or identify_cards_sync is None:
            messagebox.showerror("Missing dependency", "Photo OCR dependencies are not available.")
            return
        self._load_photo_env()
        api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
        if not api_key:
            messagebox.showerror("Missing GOOGLE_API_KEY", "Create .env in the L.U.C.A.S project folder or set GOOGLE_API_KEY.")
            return
        self.photo_client = make_photo_ocr_client(api_key)
        self.review_photo_status.set(f"Scanning 0/{len(self.review_photo_paths)} receive photo(s)...")
        self.review_photo_worker = threading.Thread(target=self._review_photo_scan_worker, daemon=True)
        self.review_photo_worker.start()

    def _review_photo_scan_worker(self) -> None:
        total = len(self.review_photo_paths)
        detected_total = 0
        for index, path in enumerate(list(self.review_photo_paths), start=1):
            try:
                self.events.put(("review_status", f"Scanning {index}/{total}: {path.name}..."))
                image_b64 = base64.b64encode(path.read_bytes()).decode("utf-8")
                cards = identify_cards_sync(
                    self.photo_client,
                    image_b64,
                    progress_callback=lambda message, i=index, n=total, p=path: self.events.put(
                        ("review_status", f"Scanning {i}/{n}: {p.name} - {message}")
                    ),
                )
                self._inventory_photo_rescue_single_bgs_cert(cards, image_b64, client=self.photo_client)
                rows = [self._photo_card_to_review_row(path, card) for card in cards if self._photo_card_has_inventory(card)]
                detected_total += len(rows)
                self.events.put(("review_rows", rows))
                self.events.put(("review_status", f"Scanning {index}/{total}: {path.name} -> {len(rows)} receive row(s)."))
            except Exception as error:
                self.events.put(("review_status", f"{path.name}: {error}"))
        self.events.put(("review_status", f"Receive photo scan complete. Added {detected_total} row(s)."))

    def _photo_card_to_review_row(self, path: Path, card: dict) -> dict[str, object]:
        row = self._photo_card_to_row(path, card)
        row["source"] = f"Receive Photo: {path.name}"
        row["notes"] = "Received"
        return row

    def _append_review_rows(self, rows: list[dict[str, object]], schedule_recommendations: bool = False) -> list[int]:
        existing = list(self.review_rows)
        start = len(existing) + 2
        added_excel_rows: list[int] = []
        refreshed_incoming_index = False
        for offset, row in enumerate(rows):
            cert = scan_to_cert(row.get("cert_number"))
            match = self._incoming_match(cert) if cert else self._incoming_raw_match(row)
            stale_assignment_match = (
                bool(match)
                and not str(match.get("best_company") or "").strip()
                and match.get("estimated_payout") is None
            )
            if cert and (not match or stale_assignment_match) and not refreshed_incoming_index:
                self.refresh_incoming_index()
                refreshed_incoming_index = True
                match = self._incoming_match(cert)
            grader = str(row.get("grader") or match.get("grader") or infer_grader(str(row.get("card_title") or ""))).upper()
            card = str(row.get("card_title") or match.get("card_title") or "").strip()
            category = str(row.get("sport") or row.get("category") or match.get("sport") or match.get("category") or "").strip()
            purchase_price = row.get("purchase_price") if row.get("purchase_price") is not None else match.get("purchase_price")
            card_ladder_value = row.get("card_ladder_value") if row.get("card_ladder_value") is not None else match.get("card_ladder_value")
            comps_average = row.get("card_ladder_comps_average") if row.get("card_ladder_comps_average") is not None else match.get("card_ladder_comps_average")
            cy_value = row.get("cy_value") if row.get("cy_value") is not None else match.get("cy_value")
            cy_confidence = row.get("cy_confidence") if row.get("cy_confidence") is not None else match.get("cy_confidence")
            comp_details = str(row.get("card_ladder_comps") or match.get("card_ladder_comps") or "")
            best_company = str(row.get("best_company") or match.get("best_company") or "").strip()
            estimated_payout = row.get("estimated_payout") if row.get("estimated_payout") is not None else match.get("estimated_payout")
            sheet_source = str(row.get("sheet_source") or match.get("sheet") or ("NO SHEET FOUND" if not match else ""))
            status = str(row.get("status") or ("Received" if match else ("Needs raw match" if not cert else "Received - no incoming match")))
            excel_row = start + offset
            workbook_row = WorkbookRow(
                excel_row=excel_row,
                cert_number=cert,
                card_title=card,
                grader=grader,
                item_id=str(row.get("item_id") or match.get("item_id") or ""),
                category=category,
                existing_value=purchase_price,
                card_ladder_value=card_ladder_value,
                card_ladder_comps_average=comps_average,
                cy_value=cy_value,
                cy_confidence=cy_confidence,
                card_ladder_comps=comp_details,
                best_company=best_company,
                estimated_payout=estimated_payout,
                received=bool(row.get("received")),
                status=status,
                notes=str(row.get("notes") or ""),
            )
            if match:
                self._attach_receive_match_to_row(workbook_row, match)
            self._ensure_receive_row_assignment(workbook_row)
            existing.append(workbook_row)
            self.review_sources[excel_row] = str(row.get("source") or "")
            self.review_sheet_sources[excel_row] = sheet_source
            added_excel_rows.append(excel_row)
        self.review_rows = existing
        self._refresh_table(schedule_recommendations=schedule_recommendations)
        return added_excel_rows

    def _incoming_match(self, cert: str) -> dict[str, object]:
        return self.incoming_cert_index.get(scan_to_cert(cert), {})

    def _incoming_index_candidates(self) -> list[dict[str, object]]:
        seen: set[str] = set()
        candidates: list[dict[str, object]] = []
        for key, candidate in self.incoming_cert_index.items():
            if not isinstance(candidate, dict):
                continue
            candidate_key = str(candidate.get("receive_key") or key or self._receive_row_ref_key(candidate)).strip().lower()
            if candidate_key in seen:
                continue
            seen.add(candidate_key)
            candidates.append(candidate)
        return candidates

    def _normalize_receive_search_text(self, value: object) -> str:
        return re.sub(r"\s+", " ", str(value or "").strip().lower())

    def _incoming_title_matches(self, query: str, limit: int = 25) -> list[dict[str, object]]:
        normalized_query = self._normalize_receive_search_text(query)
        if not normalized_query:
            return []
        query_tokens = [token for token in re.split(r"[^a-z0-9]+", normalized_query) if token]
        matches: list[dict[str, object]] = []
        for candidate in self._incoming_index_candidates():
            title = self._normalize_receive_search_text(candidate.get("card_title"))
            if not title:
                continue
            if normalized_query in title or (query_tokens and all(token in title for token in query_tokens)):
                matches.append(candidate)
                if len(matches) >= limit:
                    break
        return matches

    def _incoming_raw_matches(self, row: dict[str, object]) -> list[dict[str, object]]:
        explicit_key = str(row.get("receive_key") or "").strip().lower()
        if explicit_key and explicit_key in self.incoming_cert_index:
            return [self.incoming_cert_index[explicit_key]]
        item_id = str(row.get("item_id") or "").strip().lower()
        if item_id:
            matches = [
                candidate
                for candidate in self._incoming_index_candidates()
                if str(candidate.get("item_id") or "").strip().lower() == item_id
            ]
            if matches:
                return matches
        sheet_source = Path(str(row.get("sheet_source") or "")).name.strip().lower()
        title = self._normalize_receive_search_text(row.get("card_title"))
        if not title:
            return []
        matches = []
        for candidate in self._incoming_index_candidates():
            if sheet_source and Path(str(candidate.get("sheet") or "")).name.strip().lower() != sheet_source:
                continue
            if self._normalize_receive_search_text(candidate.get("card_title")) == title:
                matches.append(candidate)
        return matches

    def _receive_match_to_review_payload(self, match: dict[str, object], source: str) -> dict[str, object]:
        return {
            "item_id": match.get("item_id") or "",
            "cert_number": match.get("cert_number") or "",
            "grader": match.get("grader") or "",
            "sport": match.get("sport") or match.get("category") or "",
            "card_title": match.get("card_title") or "",
            "purchase_price": match.get("purchase_price"),
            "card_ladder_value": match.get("card_ladder_value"),
            "card_ladder_comps_average": match.get("card_ladder_comps_average"),
            "cy_value": match.get("cy_value"),
            "cy_confidence": match.get("cy_confidence"),
            "card_ladder_comps": match.get("card_ladder_comps") or "",
            "best_company": match.get("best_company") or "",
            "estimated_payout": match.get("estimated_payout"),
            "sheet_source": match.get("sheet") or "",
            "receive_key": match.get("receive_key") or self._receive_row_ref_key(match),
            "source": source,
            "notes": "Received",
        }

    def _receive_match_option_label(self, match: dict[str, object]) -> str:
        title = str(match.get("card_title") or "").strip() or "(no title)"
        sheet = Path(str(match.get("sheet") or "")).name
        identifier = str(match.get("cert_number") or match.get("item_id") or "").strip()
        parts = [title]
        if sheet:
            parts.append(sheet)
        if identifier:
            parts.append(identifier)
        return " | ".join(parts)

    def _refresh_receive_card_autocomplete(self, editor: ttk.Combobox, query: str) -> None:
        matches = self._incoming_title_matches(query, limit=30) if query.strip() else []
        labels: list[str] = []
        label_counts: dict[str, int] = {}
        self.receive_cell_autocomplete_matches = {}
        for match in matches:
            label = self._receive_match_option_label(match)
            label_counts[label] = label_counts.get(label, 0) + 1
            if label_counts[label] > 1:
                label = f"{label} #{label_counts[label]}"
            labels.append(label)
            self.receive_cell_autocomplete_matches[label] = match
        editor.configure(values=labels)

    def _selected_receive_autocomplete_match(self, value: str) -> dict[str, object]:
        match = self.receive_cell_autocomplete_matches.get(value)
        if match:
            return match
        title = str(value or "").split(" | ", 1)[0].strip()
        if not title:
            return {}
        matches = [
            candidate
            for candidate in self._incoming_title_matches(title, limit=2)
            if self._normalize_receive_search_text(candidate.get("card_title")) == self._normalize_receive_search_text(title)
        ]
        return matches[0] if len(matches) == 1 else {}

    def _receive_row_ref_key(self, match: dict[str, object]) -> str:
        sheet_name = str(match.get("sheet") or "").strip()
        workbook_sheet = str(match.get("workbook_sheet") or "").strip()
        try:
            workbook_row = int(match.get("workbook_row") or 0)
        except (TypeError, ValueError):
            workbook_row = 0
        if not sheet_name or not workbook_sheet or workbook_row <= 0:
            return ""
        return f"raw:{sheet_name.lower()}:{workbook_sheet.lower()}:{workbook_row}"

    def _incoming_raw_match(self, row: dict[str, object]) -> dict[str, object]:
        if not hasattr(self, "_incoming_raw_matches"):
            explicit_key = str(row.get("receive_key") or "").strip().lower()
            if explicit_key and explicit_key in self.incoming_cert_index:
                return self.incoming_cert_index[explicit_key]
            item_id = str(row.get("item_id") or "").strip().lower()
            if item_id:
                matches = [
                    candidate
                    for key, candidate in self.incoming_cert_index.items()
                    if str(key).startswith("raw:")
                    and str(candidate.get("item_id") or "").strip().lower() == item_id
                ]
                if len(matches) == 1:
                    return matches[0]
            sheet_source = Path(str(row.get("sheet_source") or "")).name.strip().lower()
            title = re.sub(r"\s+", " ", str(row.get("card_title") or "").strip().lower())
            if not title:
                return {}
            matches = []
            for key, candidate in self.incoming_cert_index.items():
                if not str(key).startswith("raw:"):
                    continue
                if sheet_source and Path(str(candidate.get("sheet") or "")).name.strip().lower() != sheet_source:
                    continue
                candidate_title = re.sub(r"\s+", " ", str(candidate.get("card_title") or "").strip().lower())
                if candidate_title == title:
                    matches.append(candidate)
            return matches[0] if len(matches) == 1 else {}
        matches = self._incoming_raw_matches(row)
        return matches[0] if len(matches) == 1 else {}

    def _attach_receive_match_to_row(self, row: WorkbookRow, match: dict[str, object]) -> None:
        receive_key = str(match.get("receive_key") or self._receive_row_ref_key(match)).strip()
        if receive_key:
            setattr(row, "_receive_key", receive_key)
        setattr(row, "_receive_sheet", str(match.get("sheet") or ""))
        setattr(row, "_receive_workbook_sheet", str(match.get("workbook_sheet") or ""))
        try:
            setattr(row, "_receive_workbook_row", int(match.get("workbook_row") or 0))
        except (TypeError, ValueError):
            setattr(row, "_receive_workbook_row", 0)

    def _receive_row_ref(self, row: WorkbookRow) -> tuple[str, str, int] | None:
        sheet_name = str(getattr(row, "_receive_sheet", "") or self.review_sheet_sources.get(row.excel_row, "") or "").strip()
        workbook_sheet = str(getattr(row, "_receive_workbook_sheet", "") or "").strip()
        try:
            workbook_row = int(getattr(row, "_receive_workbook_row", 0) or 0)
        except (TypeError, ValueError):
            workbook_row = 0
        if not sheet_name or not workbook_sheet or workbook_row <= 0:
            return None
        return (Path(sheet_name).name, workbook_sheet, workbook_row)

    def _receive_row_was_marked(
        self,
        row: WorkbookRow,
        marked_certs: set[str],
        marked_row_refs: set[tuple[str, str, int]],
    ) -> bool:
        cert = scan_to_cert(row.cert_number)
        if cert:
            return cert in marked_certs
        row_ref = self._receive_row_ref(row)
        if not row_ref:
            return False
        normalized_ref = (row_ref[0].strip().lower(), row_ref[1].strip().lower(), int(row_ref[2]))
        return normalized_ref in marked_row_refs

    def _match_all_review_rows(self) -> None:
        for row in self.review_rows:
            match = self._incoming_match(row.cert_number) if scan_to_cert(row.cert_number) else self._incoming_raw_match(
                {
                    "card_title": row.card_title,
                    "item_id": row.item_id,
                    "sheet_source": self.review_sheet_sources.get(row.excel_row, ""),
                    "receive_key": getattr(row, "_receive_key", ""),
                }
            )
            self.review_sheet_sources[row.excel_row] = str(match.get("sheet") or "NO SHEET FOUND")
            if match:
                self._attach_receive_match_to_row(row, match)
                if is_placeholder_title(row.card_title, row.grader) and match.get("card_title"):
                    row.card_title = str(match.get("card_title") or "")
                if not row.category and (match.get("sport") or match.get("category")):
                    row.category = str(match.get("sport") or match.get("category") or "")
                if not row.grader and match.get("grader"):
                    row.grader = str(match.get("grader") or "")
                if row.existing_value is None and match.get("purchase_price") is not None:
                    row.existing_value = match.get("purchase_price")
                if row.card_ladder_value is None and match.get("card_ladder_value") is not None:
                    row.card_ladder_value = match.get("card_ladder_value")
                if row.card_ladder_comps_average is None and match.get("card_ladder_comps_average") is not None:
                    row.card_ladder_comps_average = match.get("card_ladder_comps_average")
                if row.cy_value is None and match.get("cy_value") is not None:
                    row.cy_value = match.get("cy_value")
                if row.cy_confidence is None and match.get("cy_confidence") is not None:
                    row.cy_confidence = match.get("cy_confidence")
                if not row.card_ladder_comps and match.get("card_ladder_comps"):
                    row.card_ladder_comps = str(match.get("card_ladder_comps") or "")
                if not row.best_company and match.get("best_company"):
                    row.best_company = str(match.get("best_company") or "")
                if row.estimated_payout is None and match.get("estimated_payout") is not None:
                    row.estimated_payout = match.get("estimated_payout")
                self._ensure_receive_row_assignment(row)
                row.status = "Received"
            elif row.status == "Received":
                row.status = "Received - no incoming match"

    def _ensure_receive_row_assignment(self, row: WorkbookRow) -> None:
        recommendation = self.assignment_engine.recommend(row, person=getattr(self, "_assignment_person_for_row", lambda _row: "")(row))
        if recommendation.payout is None:
            if not row.best_company:
                row.best_company = NO_COMPANY_TAKES_LABEL
                row.estimated_payout = None
            return
        row.best_company = recommendation.company
        row.estimated_payout = recommendation.payout

    def clear_review_rows(self) -> None:
        self.review_rows = []
        self.review_sources = {}
        self.review_sheet_sources = {}
        self._refresh_table()
        self.review_status.set("Receive/assignment rows cleared.")

    def _clear_received_rows(
        self,
        marked_certs: set[str],
        marked_row_refs: set[tuple[str, str, int]] | None = None,
    ) -> int:
        normalized_refs = {
            (str(sheet_file).strip().lower(), str(sheet_name).strip().lower(), int(row_index))
            for sheet_file, sheet_name, row_index in (marked_row_refs or set())
        }
        if not marked_certs and not normalized_refs:
            return 0
        remaining: list[WorkbookRow] = []
        new_sources: dict[int, str] = {}
        new_sheet_sources: dict[int, str] = {}
        cleared = 0
        for row in self.review_rows:
            old_excel_row = row.excel_row
            if self._receive_row_was_marked(row, marked_certs, normalized_refs):
                cleared += 1
                continue
            row.excel_row = len(remaining) + 2
            remaining.append(row)
            if old_excel_row in self.review_sources:
                new_sources[row.excel_row] = self.review_sources[old_excel_row]
            if old_excel_row in self.review_sheet_sources:
                new_sheet_sources[row.excel_row] = self.review_sheet_sources[old_excel_row]
        if cleared:
            self.review_rows = remaining
            self.review_sources = new_sources
            self.review_sheet_sources = new_sheet_sources
            self._cancel_cell_edit()
            self._refresh_table(schedule_recommendations=False)
        return cleared

    def _clear_received_rows_for_certs(self, marked_certs: set[str]) -> int:
        return self._clear_received_rows(marked_certs)

    def delete_selected_review_rows(self) -> None:
        tree = self.review_tree
        if hasattr(self, "receive_tree") and self.receive_tree.selection():
            tree = self.receive_tree
        elif self.review_tree.selection():
            tree = self.review_tree
        deleted = self._delete_selected_rows(
            tree,
            self.review_rows,
            self.review_sources,
            self.review_sheet_sources,
        )
        if deleted:
            self.review_status.set(f"Deleted {deleted} receive/assignment row(s).")
            self.status_var.set(f"Deleted {deleted} receive/assignment row(s).")
        else:
            self.review_status.set("Select receive or assignment rows to delete.")

    def mark_review_received_in_sheets(self) -> None:
        certs = {scan_to_cert(row.cert_number) for row in self.review_rows if scan_to_cert(row.cert_number)}
        row_refs = {row_ref for row in self.review_rows if not scan_to_cert(row.cert_number) for row_ref in [self._receive_row_ref(row)] if row_ref}
        if not certs and not row_refs:
            messagebox.showinfo("No received cards", "Scan/load certed cards or load/match raw rows in Receive before marking sheets.")
            return
        paths: list[Path] = []
        errors: list[str] = []
        for directory in (INCOMING_SHEETS_DIR, WORKING_SHEETS_DIR):
            try:
                directory.mkdir(parents=True, exist_ok=True)
                paths.extend(sorted(directory.glob("*.xlsx"), key=lambda path: path.name.lower()))
            except Exception as error:
                errors.append(f"{directory}: {error}")
        if not paths:
            messagebox.showinfo("No sheets found", "No incoming or working sheets were found to update.")
            return
        marked_certs: set[str] = set()
        marked_row_refs: set[tuple[str, str, int]] = set()
        try:
            with shared_lock(CARD_PIPELINE_DIR, "receive-company-sheets", self.lucas_identity):
                result = mark_received_in_workbooks(paths, certs, row_refs)
                errors.extend(result.get("errors") or [])
                rows_marked = int(result.get("rows_marked") or 0)
                files_updated = int(result.get("files_updated") or 0)
                marked_certs = set(result.get("certs_marked") or set())
                marked_row_refs = set(result.get("row_refs_marked") or set())
                certs_marked = len(marked_certs)
                raw_rows_marked = len(marked_row_refs)
                company_rows_added = 0
                company_rows_missing_company = 0
                inventory_rows_added = 0
                if rows_marked:
                    company_rows = [
                        row
                        for row in self.review_rows
                        if row.company_pile and self._receive_row_was_marked(row, marked_certs, marked_row_refs)
                    ]
                    inventory_rows = [
                        row
                        for row in self.review_rows
                        if not row.company_pile and self._receive_row_was_marked(row, marked_certs, marked_row_refs)
                    ]
                    self._apply_recommendations_to_rows(company_rows, force=True)
                    eligible_company_rows = [row for row in company_rows if self._row_has_assignable_company(row)]
                    company_rows_missing_company = len(company_rows) - len(eligible_company_rows)
                    if eligible_company_rows:
                        company_result = append_company_sheet_rows(
                            COMPANY_SHEETS_DIR,
                            eligible_company_rows,
                            self.review_sources,
                            self.review_sheet_sources,
                            sheet_name_lookup=self._company_sheet_name_lookup_for_rows(eligible_company_rows),
                        )
                        company_rows_added = int(company_result.get("rows_added") or 0)
                        self.record_profit_sales(list(company_result.get("added_records") or []))
                        errors.extend(company_result.get("errors") or [])
                    if inventory_rows:
                        inventory_records = [
                            self._inventory_record_from_row(
                                row,
                                person=self._assignment_person_for_row(row),
                                source_sheet=self.review_sheet_sources.get(row.excel_row, ""),
                                source=self.review_sources.get(row.excel_row, ""),
                                notes="Received without company pile",
                            )
                            for row in inventory_rows
                        ]
                        inventory_rows_added = self.add_inventory_records(inventory_records)
                moved_received = self._move_fully_received_sheets_to_received(paths)
                if moved_received:
                    self._save_sheet_markers()
        except Exception as error:
            messagebox.showerror("Shared folder busy", str(error))
            self.status_var.set(f"Receive update failed: {error}")
            return
        self.refresh_incoming_index()
        self.refresh_working_sheets()
        self.refresh_received_sheets()
        cleared_receive_rows = self._clear_received_rows(marked_certs, marked_row_refs)
        if rows_marked:
            self.review_status.set(f"Marked {rows_marked} row(s) received across {files_updated} sheet file(s).")
            self.status_var.set(f"Marked {certs_marked}/{len(certs)} cert(s) and {raw_rows_marked}/{len(row_refs)} raw row(s); cleared {cleared_receive_rows} receive row(s).")
        else:
            self.review_status.set("No matching cert/raw rows were found in incoming or working sheets.")
            self.status_var.set("No sheet rows marked received.")
        if moved_received:
            self.status_var.set(f"Moved {len(moved_received)} fully received sheet(s) to RECEIVED SHEETS.")
        if company_rows_added:
            self.status_var.set(f"Added {company_rows_added} card(s) to weekly company sheet(s).")
        if inventory_rows_added:
            self.status_var.set(f"Added {inventory_rows_added} received card(s) to active inventory.")
        elif company_rows_missing_company:
            self.status_var.set(f"{company_rows_missing_company} checked company pile card(s) had no Best Company.")
        self.refresh_home()
        summary_lines = [
            f"Marked rows: {rows_marked}",
            f"Updated sheet files: {files_updated}",
            f"Matched certs: {certs_marked}/{len(certs)}",
        ]
        if row_refs:
            summary_lines.append(f"Matched raw rows: {raw_rows_marked}/{len(row_refs)}")
        if moved_received:
            summary_lines.append(f"Moved to received: {len(moved_received)}")
        if company_rows_added:
            summary_lines.append(f"Company sheet rows added: {company_rows_added}")
        if inventory_rows_added:
            summary_lines.append(f"Inventory rows added: {inventory_rows_added}")
        if company_rows_missing_company:
            summary_lines.append(f"Company pile rows missing Best Company: {company_rows_missing_company}")
        if cleared_receive_rows:
            summary_lines.append(f"Cleared receive rows: {cleared_receive_rows}")
        self._append_activity(
            "Receive",
            f"Marked {rows_marked} row(s) received across {files_updated} sheet file(s).",
            {
                "rows_marked": rows_marked,
                "files_updated": files_updated,
                "certs_marked": certs_marked,
                "total_certs": len(certs),
                "raw_rows_marked": raw_rows_marked,
                "total_raw_rows": len(row_refs),
                "company_rows_added": company_rows_added,
                "inventory_rows_added": inventory_rows_added,
                "receive_rows_cleared": cleared_receive_rows,
                "moved_received": len(moved_received),
                "warnings": errors[:8],
            },
        )
        if errors:
            messagebox.showwarning("Mark received completed with warnings", "\n".join(summary_lines + ["", "Warnings:", *errors[:8]]))
        else:
            messagebox.showinfo("Mark received complete", "\n".join(summary_lines))

    def _apply_recommendations_to_rows(self, rows: list[WorkbookRow], force: bool = False) -> None:
        for row in rows:
            if not force and row.best_company and row.estimated_payout is not None:
                continue
            recommendation = self.assignment_engine.recommend(row, person=getattr(self, "_assignment_person_for_row", lambda _row: "")(row))
            if recommendation.payout is None:
                self._record_unassigned_player(row)
                if force:
                    row.best_company = NO_COMPANY_TAKES_LABEL
                    row.estimated_payout = None
                elif not row.best_company:
                    row.best_company = NO_COMPANY_TAKES_LABEL
                    row.estimated_payout = None
                continue
            row.best_company = recommendation.company
            row.estimated_payout = recommendation.payout

    def _row_has_assignable_company(self, row: WorkbookRow) -> bool:
        company = str(row.best_company or "").strip()
        return bool(company) and company.upper() != NO_COMPANY_TAKES_LABEL

    def _ensure_company_sheet_folders(self) -> None:
        try:
            COMPANY_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
            for company in self.assignment_engine.companies:
                folder_name = self._safe_company_folder_name(company.name)
                if folder_name:
                    (COMPANY_SHEETS_DIR / folder_name).mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

    def _weekly_company_sheet_timer(self) -> None:
        self._ensure_weekly_company_sheets_due()
        self.after(5 * 60 * 1000, self._weekly_company_sheet_timer)

    def _company_sheet_reset_schedules(self) -> dict[str, dict[str, str]]:
        raw = {}
        if isinstance(getattr(self, "app_settings", None), dict):
            raw = self.app_settings.get("company_sheet_reset_schedules") or {}
        schedules: dict[str, dict[str, str]] = {}
        if isinstance(raw, dict):
            for company, config in raw.items():
                company_name = str(company or "").strip()
                if not company_name or not isinstance(config, dict):
                    continue
                weekday = str(config.get("weekday") or DEFAULT_COMPANY_RESET_WEEKDAY).strip().title()
                if weekday not in COMPANY_RESET_WEEKDAYS:
                    weekday = DEFAULT_COMPANY_RESET_WEEKDAY
                time_text = str(config.get("time") or DEFAULT_COMPANY_RESET_TIME).strip()
                try:
                    time_text = parse_company_reset_time(time_text).strftime("%H:%M")
                except ValueError:
                    time_text = DEFAULT_COMPANY_RESET_TIME
                schedules[company_name] = {"weekday": weekday, "time": time_text}
        for company in getattr(getattr(self, "assignment_engine", None), "companies", []) or []:
            company_name = str(getattr(company, "name", "") or "").strip()
            if not company_name:
                continue
            weekday = str(getattr(company, "reset_weekday", "") or DEFAULT_COMPANY_RESET_WEEKDAY).strip().title()
            if weekday not in COMPANY_RESET_WEEKDAYS:
                weekday = DEFAULT_COMPANY_RESET_WEEKDAY
            time_text = str(getattr(company, "reset_time", "") or DEFAULT_COMPANY_RESET_TIME).strip()
            try:
                time_text = parse_company_reset_time(time_text).strftime("%H:%M")
            except ValueError:
                time_text = DEFAULT_COMPANY_RESET_TIME
            schedules[company_name] = {"weekday": weekday, "time": time_text}
        return schedules

    def _company_sheet_schedule_for_company(self, company: str) -> dict[str, str]:
        schedules = self._company_sheet_reset_schedules()
        company_name = str(company or "").strip()
        return schedules.get(company_name) or {"weekday": DEFAULT_COMPANY_RESET_WEEKDAY, "time": DEFAULT_COMPANY_RESET_TIME}

    def _company_sheet_week_start_for_company(self, company: str, now: datetime | None = None) -> datetime.date:
        schedule = self._company_sheet_schedule_for_company(company)
        return company_sheet_week_start_for_schedule(now or datetime.now(), schedule.get("weekday"), schedule.get("time"))

    def _company_week_start_lookup(self, companies: list[str], now: datetime | None = None) -> dict[str, datetime.date]:
        moment = now or datetime.now()
        return {company: self._company_sheet_week_start_for_company(company, moment) for company in companies if str(company or "").strip()}

    def _company_sheet_name_lookup_for_rows(self, rows: list[WorkbookRow], now: datetime | None = None) -> dict[str, str]:
        companies = sorted({clean_part(getattr(row, "best_company", "")) for row in rows if clean_part(getattr(row, "best_company", ""))}, key=str.lower)
        return {
            company: f"Week of {self._company_sheet_week_start_for_company(company, now):%Y-%m-%d}"
            for company in companies
        }

    def _ensure_weekly_company_sheets_due(self, now: datetime | None = None) -> dict[str, object]:
        now = now or datetime.now()
        try:
            markers = json.loads(WEEKLY_COMPANY_SHEETS_PATH.read_text(encoding="utf-8")) if WEEKLY_COMPANY_SHEETS_PATH.exists() else {}
        except Exception:
            markers = {}
        weeks = markers.get("weeks") if isinstance(markers, dict) else {}
        if not isinstance(weeks, dict):
            weeks = {}
        companies = [company.name for company in self.assignment_engine.companies if company.name]
        company_names = sorted(companies, key=str.lower)
        week_start_lookup = self._company_week_start_lookup(companies, now)
        marker_key = "|".join(f"{company}:{week_start_lookup[company].isoformat()}" for company in company_names if company in week_start_lookup)
        marker = weeks.get(marker_key) if isinstance(weeks.get(marker_key), dict) else {}
        if marker.get("companies") == company_names:
            return {"created": [], "existing": [], "errors": [], "skipped": True, "week_start": marker_key}
        result = ensure_company_weekly_sheets(COMPANY_SHEETS_DIR, companies, week_start_lookup=week_start_lookup)
        errors = list(result.get("errors") or [])
        if not errors:
            weeks[marker_key] = {
                "created_at": now.isoformat(timespec="seconds"),
                "week_start": marker_key,
                "week_start_by_company": {company: week_start.isoformat() for company, week_start in week_start_lookup.items()},
                "reset_schedules": self._company_sheet_reset_schedules(),
                "company_count": len(companies),
                "companies": company_names,
                "created_count": len(result.get("created") or []),
                "existing_count": len(result.get("existing") or []),
            }
            atomic_write_json(WEEKLY_COMPANY_SHEETS_PATH, {"weeks": weeks})
            created_count = len(result.get("created") or [])
            if created_count:
                self.status_var.set(f"Created {created_count} weekly company sheet tab(s) for week of {marker_key}.")
        return {**result, "skipped": False, "week_start": marker_key}

    def _safe_company_folder_name(self, name: str) -> str:
        return re.sub(r"[<>:\"/\\|?*]+", " ", str(name or "")).strip()[:140].strip()

    def _arm_review_scanner(self) -> None:
        if self.review_mode.get() != "Automatic Receive" or self.review_scan_entry is None:
            return
        try:
            self.review_scan_entry.focus_set()
            self.review_scan_entry.icursor(tk.END)
        except tk.TclError:
            pass

    def _cardladder_extension_warning(self) -> str:
        with self.state.lock:
            last_seen = self.state.last_seen_extension
            extension_version = self.state.extension_version
            manifest_version = self.state.extension_manifest_version
            extension_name = self.state.extension_name
            extension_url = self.state.extension_url
        if not last_seen:
            return ""
        loaded_label = extension_name or "Card Ladder extension"
        version_label = extension_version or "unversioned/old"
        manifest_label = manifest_version or "unknown"
        if extension_version != EXPECTED_CARDLADDER_EXTENSION_VERSION:
            return (
                f"{loaded_label} checked in at {last_seen}, but it is {version_label} "
                f"(manifest {manifest_label}).\n\n"
                f"Expected helper version: {EXPECTED_CARDLADDER_EXTENSION_VERSION} "
                f"(manifest {EXPECTED_CARDLADDER_MANIFEST_VERSION}).\n\n"
                "Open chrome://extensions, remove or disable the old Card Ladder Auto-Comp helper, "
                f"then Load unpacked from:\n{CARDLADDER_EXTENSION_DIR}\n\n"
                f"Chrome extension URL seen by the app: {extension_url or 'unknown'}"
            )
        return ""

    def run_all_comps(self) -> None:
        app_debug_log(f"run_all_comps_clicked rows={len(self.state.rows)}")
        if not self.state.rows:
            app_debug_log("run_all_comps_blocked reason=no_rows")
            messagebox.showinfo("No comp sheet loaded", "Choose and load an incoming or working sheet in the Comp tab first.")
            return
        requery_all = self.comp_scope_label.get() == COMP_SCOPE_ALL
        source_label = self.comp_source_label.get()
        if not COMP_CY_ENABLED and source_label != COMP_SOURCE_CARD_LADDER:
            source_label = COMP_SOURCE_CARD_LADDER
            self.comp_source_label.set(COMP_SOURCE_CARD_LADDER)
        run_card_ladder = source_label in {COMP_SOURCE_BOTH, COMP_SOURCE_CARD_LADDER}
        run_cy = COMP_CY_ENABLED and source_label in {COMP_SOURCE_BOTH, COMP_SOURCE_CY}
        card_ladder_eligible = [
            row
            for row in self.state.rows
            if run_card_ladder and row.cert_number and row.grader and (requery_all or not row_has_comp_data(row))
        ]
        cy_eligible = [
            row
            for row in self.state.rows
            if run_cy and row.cert_number and row.grader and (requery_all or row.cy_value is None)
        ]
        if card_ladder_eligible:
            extension_warning = self._cardladder_extension_warning()
            if extension_warning:
                messagebox.showwarning("Reload Card Ladder extension", extension_warning)
                self.status_var.set("Reload the Card Ladder Chrome extension before comping.")
                return
        if not card_ladder_eligible and not cy_eligible:
            if requery_all:
                message = f"No rows have both a cert number and company ready for {source_label}."
            else:
                message = f"No rows are missing {source_label} data. Switch Run Scope to Recomp All if you want to refresh every row."
            app_debug_log(f"run_all_comps_blocked reason=no_eligible source={source_label} requery_all={requery_all}")
            messagebox.showinfo("No eligible rows", message)
            self.status_var.set(message)
            return
        self.state.set_comp_strategy(COMP_STRATEGY_DISPLAY.get(self.comp_strategy_label.get(), COMP_STRATEGY_AVERAGE), self._comp_low_outlier_pct())
        self.pending_comp_assignment_row_ids = {id(row) for row in [*card_ladder_eligible, *cy_eligible]}
        command_id = 0
        card_ladder_command_id = 0
        if card_ladder_eligible:
            card_ladder_command_id = self.state.start_all_comps(requery_all=requery_all, allow_deferred_cy=False)
            command_id = card_ladder_command_id
        if cy_eligible:
            command_id = self.state.start_cy_lookups(cy_eligible, defer=bool(card_ladder_eligible))
        self.comp_output_saved = False
        self._refresh_table()
        if card_ladder_eligible:
            self.after(12000, lambda queued_command_id=card_ladder_command_id: self._warn_if_extension_not_checked_in(queued_command_id))
        pieces = []
        if card_ladder_eligible:
            pieces.append(f"{len(card_ladder_eligible)} Card Ladder")
        if cy_eligible:
            pieces.append(f"{len(cy_eligible)} CY")
        self.status_var.set(f"Queued {' and '.join(pieces)} row(s) using {self.comp_scope_label.get()} as command #{command_id}.")
        app_debug_log(
            f"run_all_comps_queued command={command_id} source={source_label} "
            f"card_ladder={len(card_ladder_eligible)} cy={len(cy_eligible)}"
        )

    def _warn_if_extension_not_checked_in(self, command_id: int) -> None:
        extension_warning = self._cardladder_extension_warning()
        if extension_warning:
            messagebox.showwarning("Reload Card Ladder extension", extension_warning)
            self.status_var.set("Reload the Card Ladder Chrome extension before comping.")
            return
        with self.state.lock:
            command_pending = bool(self.state.command and self.state.command.get("id") == command_id)
        if not command_pending:
            return
        messagebox.showwarning(
            "Card Ladder extension not connected",
            "The rows were queued, but the Card Ladder Chrome extension has not checked in. Make sure Chrome is open, logged into Card Ladder, and the bundled extension is loaded.",
        )

    def stop_comp_run(self) -> None:
        self.pending_comp_assignment_row_ids = set()
        self.state.request_cancel()
        self.comp_output_saved = False
        self._refresh_table()
        self.status_var.set("Stop requested. Card Ladder and queued CY work were cancelled.")

    def clear_comp_rows(self) -> None:
        if self.state.rows and not self.comp_output_saved:
            confirmed = messagebox.askyesno(
                "Clear unsaved comp rows?",
                "These comp rows have not been saved as an output. Clear them anyway?",
                icon=messagebox.WARNING,
            )
            if not confirmed:
                self.status_var.set("Clear comp rows cancelled.")
                return
        self.state.set_rows([])
        self.row_sources = {}
        self.comp_sheet_sources = {}
        self.pending_comp_assignment_row_ids = set()
        self.selected_working_sheet.set("")
        self.loaded_comp_sheet_label = ""
        self.comp_output_saved = True
        self._cancel_cell_edit()
        try:
            self.working_sheet_list.selection_clear(0, tk.END)
        except tk.TclError:
            pass
        self._refresh_table()
        self.status_var.set("Comp rows cleared.")

    def recalculate_comp_method(self, _event=None) -> None:
        strategy = COMP_STRATEGY_DISPLAY.get(self.comp_strategy_label.get(), COMP_STRATEGY_AVERAGE)
        low_outlier_pct = self._comp_low_outlier_pct()
        self.state.set_comp_strategy(strategy, low_outlier_pct)
        updated = 0
        with self.state.lock:
            for row in self.state.rows:
                comps = parse_formatted_comps(row.card_ladder_comps)
                if not comps:
                    continue
                row.card_ladder_comps_average = comp_price(comps, strategy, low_outlier_pct)
                row.card_ladder_comps = format_comps(comps, strategy, low_outlier_pct)
                updated += 1
        if updated:
            self.comp_output_saved = False
        self._refresh_table(schedule_recommendations=bool(updated))
        if updated:
            suffix = f" and low comp filter {low_outlier_pct:g}% of average" if low_outlier_pct else ""
            self.status_var.set(f"Recalculated {updated} comp row(s) with {self.comp_strategy_label.get()}{suffix}.")
        elif self.state.rows:
            self.status_var.set("Comp method updated. No stored comp details were available to recalculate.")
        else:
            self.status_var.set("Comp method updated.")

    def save_output(self) -> None:
        if not self.state.rows:
            messagebox.showinfo("No rows", "Load or scan cards before saving.")
            return
        self._apply_recommendations()
        default = default_output_path(ROOT)
        path = filedialog.asksaveasfilename(
            title="Save pipeline workbook",
            initialdir=str(default.parent),
            initialfile=default.name,
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return
        try:
            with shared_lock(CARD_PIPELINE_DIR, "workbook-writes", self.lucas_identity):
                write_pipeline_output(Path(path), self.state.rows, self.row_sources)
        except Exception as error:
            messagebox.showerror("Save failed", str(error))
            return
        self.comp_output_saved = True
        self.status_var.set(f"Saved {path}")

    def _marker_for_sheet_name(self, sheet_name: str, stages: tuple[str, ...] = ("Working", "Incoming", "Received")) -> tuple[str, dict[str, object]]:
        source_name = Path(str(sheet_name or "")).name
        if not source_name:
            return "", {}
        for stage in stages:
            key = self._home_sheet_key(stage, source_name)
            marker = self.home_sheet_markers.get(key, {})
            if marker:
                return key, marker
        source_key = source_name.lower()
        for key, marker in self.home_sheet_markers.items():
            stage, name = self._split_home_sheet_key(key)
            if stage in stages and Path(name).name.lower() == source_key:
                return key, marker
        return "", {}

    def _apply_seller_terms_to_rows_for_marker(self, rows: list[WorkbookRow], marker: dict[str, object]) -> int:
        term = self._seller_term_for_marker(marker)
        if not term:
            return 0
        sheet_type = str(term.get("sheet_type") or "").strip()
        rate = self._seller_terms_rate(term.get("rate"))
        deduction = self._seller_terms_rate(term.get("deduction"))
        changed = 0
        for row in rows:
            seller_price = (
                self._seller_terms_company_price(row, sheet_type, deduction=deduction)
                if deduction is not None
                else self._seller_terms_company_price(row, sheet_type, rate=rate)
            )
            if seller_price is None:
                continue
            if row.existing_value != seller_price:
                row.existing_value = seller_price
                changed += 1
        return changed

    def save_comp_to_source_sheet(self) -> None:
        if not self.state.rows:
            messagebox.showinfo("No rows", "Load an incoming or working sheet before saving back to its source.")
            return
        label = self.selected_working_sheet.get().strip() if hasattr(self, "selected_working_sheet") else ""
        stage, name, path = self._comp_sheet_info(label)
        if not path:
            messagebox.showinfo("No source sheet", "Choose and load an incoming or working sheet before saving back to its source.")
            return
        key, marker = self._marker_for_sheet_name(path.name, (stage,) if stage else ("Working", "Incoming"))
        seller_updates = self._apply_seller_terms_to_rows_for_marker(self.state.rows, marker)
        try:
            with shared_lock(CARD_PIPELINE_DIR, "workbook-writes", self.lucas_identity):
                write_working_sheet(path, self.state.rows, self.row_sources)
        except Exception as error:
            messagebox.showerror("Save failed", str(error))
            return
        self.comp_output_saved = True
        self._refresh_table(schedule_recommendations=False)
        self.refresh_home()
        suffix = f" Seller prices updated on {seller_updates} row(s)." if key and seller_updates else ""
        stage_label = f"{stage.lower()} " if stage else ""
        self.status_var.set(f"Saved current comp rows back to {stage_label}{path.name}.{suffix}")

    def save_working_sheet(self) -> None:
        if not self.intake_rows:
            messagebox.showinfo("No create rows", "Scan or load cards in Create before saving a working sheet.")
            return
        title = self.working_sheet_title.get().strip()
        if not title:
            messagebox.showinfo("Title required", "Enter a working sheet title first.")
            return
        seller = ""
        seller_sheet_type = ""
        seller_term: dict[str, object] | None = None
        if self._network_mode_enabled() and hasattr(self, "seller_terms_seller_var"):
            seller = self.seller_terms_seller_var.get().strip()
            seller_sheet_type = self.seller_terms_sheet_type_var.get().strip() if hasattr(self, "seller_terms_sheet_type_var") else ""
            if seller or seller_sheet_type:
                if not seller or not seller_sheet_type:
                    messagebox.showinfo(
                        "Seller terms required",
                        "Network Mode seller buys need both Seller and Sheet Type. Leave both blank for a normal Open Team sheet.",
                    )
                    return
                seller_term = self._seller_terms_match(seller, seller_sheet_type)
                if not seller_term:
                    messagebox.showinfo(
                        "Seller terms not found",
                        f"No People Rules were found for {seller} / {seller_sheet_type}. Open People Rules or leave both fields blank.",
                    )
                    return
                rate = self._money_value(seller_term.get("rate"))
                deduction = self._money_value(seller_term.get("deduction"))
                applicable_rows = 0
                for row in self.intake_rows:
                    seller_price = (
                        self._seller_terms_company_price(row, seller_sheet_type, deduction=deduction)
                        if deduction is not None
                        else self._seller_terms_company_price(row, seller_sheet_type, rate=rate)
                    )
                    if seller_price is not None:
                        applicable_rows += 1
                if applicable_rows <= 0:
                    self.status_var.set(self._seller_terms_no_match_message(self.intake_rows, seller_sheet_type, deduction))
        self.apply_create_seller_terms(show_status=False)
        try:
            with shared_lock(CARD_PIPELINE_DIR, "workbook-writes", self.lucas_identity):
                path = working_sheet_path(WORKING_SHEETS_DIR, title)
                raw_ids_added = self._ensure_raw_item_ids_for_rows(self.intake_rows)
                write_working_sheet(path, self.intake_rows, self.intake_sources)
            if seller:
                self._assign_sheet_to_seller("Working", path.name, seller, seller_sheet_type, seller_term)
        except Exception as error:
            messagebox.showerror("Save failed", str(error))
            return
        seller_note = f" Assigned to {seller} for payouts." if seller else ""
        if seller and seller_term:
            term_summary = self._seller_payout_summary_for_rows(self.intake_rows, {"assigned_person": seller, "seller_terms_applied": True, "seller_sheet_type": seller_sheet_type, "seller_rate": seller_term.get("rate"), "seller_deduction": seller_term.get("deduction")})
            if term_summary.get("seller_payout_pending"):
                seller_note = f"{seller_note} {term_summary.get('seller_payout_warning')}."
        raw_note = f" Added {raw_ids_added} raw item ID(s)." if raw_ids_added else ""
        self.status_var.set(f"Saved working sheet: {path}.{seller_note}{raw_note}")
        self.intake_rows = []
        self.intake_sources = {}
        self.intake_sheet_sources = {}
        self.working_sheet_title.set("")
        self._refresh_table()
        self.refresh_home()
        self._append_activity("Create", f"Saved working sheet {path.name}.", {"sheet": path.name, "seller": seller, "sheet_type": seller_sheet_type})

    def refresh_pipeline(self) -> None:
        self.refresh_working_sheets()
        self.refresh_home()
        self._refresh_table()

    def refresh_working_sheets(self) -> None:
        try:
            CARD_PIPELINE_DIR.mkdir(parents=True, exist_ok=True)
            INCOMING_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
            WORKING_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
            incoming_paths = sorted(INCOMING_SHEETS_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
            working_paths = sorted(WORKING_SHEETS_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
        except Exception as error:
            self.incoming_sheet_paths = {}
            self.working_sheet_paths = {}
            self.comp_sheet_paths = {}
            self.comp_sheet_stages = {}
            if hasattr(self, "working_sheet_list"):
                self.working_sheet_list.delete(0, tk.END)
            self.status_var.set(f"Comp sheets unavailable: {error}")
            return
        self.incoming_sheet_paths = {path.name: path for path in incoming_paths}
        self.working_sheet_paths = {path.name: path for path in working_paths}
        self._populate_comp_sheet_list()
        if self.comp_sheet_paths and self.selected_working_sheet.get() not in self.comp_sheet_paths:
            self.selected_working_sheet.set(next(iter(self.comp_sheet_paths)))
        self._select_working_sheet_in_list()
        self.status_var.set(f"Found {len(incoming_paths)} incoming and {len(working_paths)} working sheet(s).")

    def _comp_sheet_label(self, stage: str, name: str) -> str:
        return f"{stage} / {name}"

    def _populate_comp_sheet_list(self) -> None:
        self.comp_sheet_paths = {}
        self.comp_sheet_stages = {}
        for stage, paths in (("Incoming", self.incoming_sheet_paths), ("Working", self.working_sheet_paths)):
            for name, path in paths.items():
                label = self._comp_sheet_label(stage, name)
                self.comp_sheet_paths[label] = path
                self.comp_sheet_stages[label] = stage
        if hasattr(self, "working_sheet_list"):
            self.working_sheet_list.delete(0, tk.END)
            for label in self.comp_sheet_paths:
                self.working_sheet_list.insert(tk.END, label)

    def _comp_sheet_info(self, label: str) -> tuple[str, str, Path | None]:
        label = str(label or "").strip()
        comp_sheet_paths = getattr(self, "comp_sheet_paths", {})
        comp_sheet_stages = getattr(self, "comp_sheet_stages", {})
        working_sheet_paths = getattr(self, "working_sheet_paths", {})
        incoming_sheet_paths = getattr(self, "incoming_sheet_paths", {})
        path = comp_sheet_paths.get(label)
        if path:
            return comp_sheet_stages.get(label, ""), path.name, path
        if label in working_sheet_paths:
            return "Working", label, working_sheet_paths.get(label)
        if label in incoming_sheet_paths:
            return "Incoming", label, incoming_sheet_paths.get(label)
        return "", label, None

    def load_selected_working_sheet(self) -> None:
        label = self._selected_working_sheet_name()
        stage, name, path = self._comp_sheet_info(label)
        if not path:
            messagebox.showinfo("Choose sheet", "Choose an incoming or working sheet first.")
            return
        self.status_var.set(f"Loading {stage.lower()} sheet: {name}...")
        threading.Thread(target=self._load_working_sheet_worker, args=(label, stage, name, path), daemon=True).start()

    def _load_working_sheet_worker(self, label: str, stage: str, name: str, path: Path) -> None:
        try:
            rows = read_simple_spreadsheet(path)
        except Exception as error:
            self.events.put(("load_working_sheet_error", {"name": name, "stage": stage, "error": str(error)}))
            return
        self.events.put(("load_working_sheet_done", {"label": label, "stage": stage, "name": name, "rows": rows}))

    def _apply_loaded_working_sheet(self, name: str, rows: list[dict[str, object]], stage: str = "Working", label: str = "") -> None:
        workbook_rows: list[WorkbookRow] = []
        sources: dict[int, str] = {}
        for offset, row in enumerate(rows, start=2):
            cert = str(row.get("cert_number") or "")
            grader = str(row.get("grader") or infer_grader(str(row.get("card_title") or "")) or "PSA").upper()
            card = str(row.get("card_title") or "").strip()
            workbook_rows.append(
                WorkbookRow(
                    excel_row=offset,
                    cert_number=cert,
                    card_title=card,
                    grader=grader,
                    item_id=str(row.get("item_id") or ""),
                    category=str(row.get("sport") or row.get("category") or "").strip(),
                    existing_value=row.get("purchase_price"),
                    card_ladder_value=row.get("card_ladder_value"),
                    card_ladder_comps_average=row.get("card_ladder_comps_average"),
                    cy_value=row.get("cy_value"),
                    cy_confidence=row.get("cy_confidence"),
                    card_ladder_comps=str(row.get("card_ladder_comps") or ""),
                    best_company=str(row.get("best_company") or ""),
                    estimated_payout=row.get("estimated_payout"),
                    received=bool(row.get("received")),
                    status=str(row.get("status") or ("Ready" if cert and grader else "Needs setup")),
                    notes=str(row.get("notes") or ""),
                )
            )
            sources[offset] = str(row.get("source") or name)
        self.state.set_rows(workbook_rows)
        self.row_sources = sources
        self.comp_sheet_sources = {}
        self.comp_output_saved = True
        self.loaded_comp_sheet_label = label or self._comp_sheet_label(stage, name)
        self._refresh_table(schedule_recommendations=any(row.card_ladder_comps_average is not None for row in workbook_rows))
        self.selected_working_sheet.set(self.loaded_comp_sheet_label)
        self._select_working_sheet_in_list()
        self.status_var.set(f"Loaded {stage.lower()} sheet: {name}")

    def _selected_working_sheet_name(self) -> str:
        if hasattr(self, "working_sheet_list"):
            selected = self.working_sheet_list.curselection()
            if selected:
                return str(self.working_sheet_list.get(selected[0]))
        return self.selected_working_sheet.get()

    def _select_working_sheet_in_list(self) -> None:
        if not hasattr(self, "working_sheet_list"):
            return
        target = self.selected_working_sheet.get()
        self.working_sheet_list.selection_clear(0, tk.END)
        for index, name in enumerate(self.working_sheet_list.get(0, tk.END)):
            if name == target:
                self.working_sheet_list.selection_set(index)
                self.working_sheet_list.see(index)
                break

    def clear_rows(self) -> None:
        self.intake_rows = []
        self.intake_sources = {}
        self.intake_sheet_sources = {}
        self._refresh_table()
        self.status_var.set("Create rows cleared.")

    def delete_selected_intake_rows(self) -> None:
        deleted = self._delete_selected_rows(
            self.intake_tree,
            self.intake_rows,
            self.intake_sources,
            self.intake_sheet_sources,
        )
        if deleted:
            self.status_var.set(f"Deleted {deleted} create row(s).")
        else:
            self.status_var.set("Select create rows to delete.")

    def delete_selected_comp_rows(self) -> None:
        with self.state.lock:
            rows = list(self.state.rows)
        deleted = self._delete_selected_rows(
            self.comp_tree,
            rows,
            self.row_sources,
            self.comp_sheet_sources,
        )
        if deleted:
            self.comp_output_saved = False
            self.status_var.set(f"Deleted {deleted} comp row(s). Save back to source sheet to persist.")
        else:
            self.status_var.set("Select comp rows to delete.")

    def add_comp_row(self) -> None:
        label = str(getattr(self, "loaded_comp_sheet_label", "") or "").strip()
        if not label:
            messagebox.showinfo("No comp sheet loaded", "Load an incoming or working sheet in the Comp tab before adding rows.")
            return
        stage, name, path = self._comp_sheet_info(label)
        if not path:
            messagebox.showinfo("No source sheet", "Load an incoming or working sheet in the Comp tab before adding rows.")
            return
        with self.state.lock:
            rows = list(self.state.rows)
        excel_row = max((int(getattr(row, "excel_row", 1) or 1) for row in rows), default=1) + 1
        rows.append(
            WorkbookRow(
                excel_row=excel_row,
                cert_number="",
                card_title="",
                grader="",
                status="Needs setup",
            )
        )
        self.state.set_rows(rows)
        self.row_sources[excel_row] = name or path.name
        self.comp_sheet_sources.pop(excel_row, None)
        self.comp_output_saved = False
        self._cancel_cell_edit()
        self._refresh_comp_table(schedule_recommendations=False)
        try:
            self.comp_tree.selection_set(str(excel_row))
            self.comp_tree.focus(str(excel_row))
            self.comp_tree.see(str(excel_row))
        except tk.TclError:
            pass
        self.status_var.set("Added blank comp row. Edit it, then Save Back to Source Sheet to persist.")

    def _append_rows(self, rows: list[dict[str, object]]) -> list[int]:
        existing = list(self.intake_rows)
        start = len(existing) + 2
        added_excel_rows: list[int] = []
        for offset, row in enumerate(rows):
            cert = str(row.get("cert_number") or "")
            grader = str(row.get("grader") or infer_grader(str(row.get("card_title") or ""))).upper()
            card = str(row.get("card_title") or "").strip()
            status = "Ready" if cert and grader else "Needs setup"
            notes = str(row.get("notes") or "")
            excel_row = start + offset
            existing.append(
                WorkbookRow(
                    excel_row=excel_row,
                    cert_number=cert,
                    card_title=card,
                    grader=grader,
                    category=str(row.get("sport") or row.get("category") or "").strip(),
                    existing_value=row.get("purchase_price"),
                    status=status,
                    notes=notes,
                )
            )
            self.intake_sources[excel_row] = str(row.get("source") or "")
            self.intake_sheet_sources[excel_row] = ""
            added_excel_rows.append(excel_row)
        self.intake_rows = existing
        self.apply_create_seller_terms(show_status=False)
        self._refresh_table()
        return added_excel_rows

    def _apply_recommendations(self) -> None:
        for row in [*self.state.rows, *self.review_rows]:
            recommendation = self.assignment_engine.recommend(row, person=getattr(self, "_assignment_person_for_row", lambda _row: "")(row))
            if recommendation.payout is None:
                self._record_unassigned_player(row)
                row.best_company = NO_COMPANY_TAKES_LABEL
                row.estimated_payout = None
                continue
            row.best_company = recommendation.company
            row.estimated_payout = recommendation.payout

    def _load_player_overrides(self) -> int:
        return assignment_engine.load_player_sport_overrides(PLAYER_OVERRIDES_PATH)

    def _load_unassigned_players(self) -> dict[str, dict[str, object]]:
        try:
            raw = json.loads(UNASSIGNED_PLAYERS_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
        entries = raw.get("entries", raw) if isinstance(raw, dict) else {}
        return entries if isinstance(entries, dict) else {}

    def _save_unassigned_players(self, entries: dict[str, dict[str, object]]) -> None:
        UNASSIGNED_PLAYERS_PATH.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_json(UNASSIGNED_PLAYERS_PATH, {"entries": entries})

    def _record_unassigned_players(self, rows: list[WorkbookRow]) -> None:
        for row in rows:
            self._record_unassigned_player(row)

    def _record_unassigned_player(self, row: WorkbookRow) -> None:
        best_company = str(row.best_company or "").strip()
        if row.estimated_payout is not None or (best_company and best_company.upper() != NO_COMPANY_TAKES_LABEL):
            return
        if row.card_ladder_comps_average is None and row.card_ladder_value is None:
            return
        title = str(row.card_title or "").strip()
        if not title:
            return
        player_guess = self._guess_unassigned_player(row)
        if not player_guess:
            return
        key = re.sub(r"[^a-z0-9]+", " ", player_guess.lower()).strip() or re.sub(r"[^a-z0-9]+", " ", title.lower()).strip()
        if not key:
            return
        try:
            with shared_lock(CARD_PIPELINE_DIR, "unassigned-players", self.lucas_identity):
                entries = self._load_unassigned_players()
                existing = dict(entries.get(key, {}))
                existing["player"] = existing.get("player") or player_guess
                existing["last_title"] = title
                existing["sample_titles"] = self._append_unique_sample(existing.get("sample_titles"), title)
                existing["count"] = int(existing.get("count") or 0) + 1
                existing["last_seen"] = datetime.now().isoformat(timespec="seconds")
                existing["source"] = str(row.cert_number or "")
                entries[key] = existing
                self._save_unassigned_players(entries)
        except Exception:
            pass

    def _append_unique_sample(self, samples: object, value: str) -> list[str]:
        result = [str(item) for item in samples] if isinstance(samples, list) else []
        if value not in result:
            result.append(value)
        return result[-5:]

    def _guess_unassigned_player(self, row: WorkbookRow) -> str:
        title = str(row.card_title or "").strip()
        if not title:
            return ""
        grader_pattern = r"\b(?:PSA|BGS|SGC|CGC)\b"
        before_grader = re.split(grader_pattern, title, flags=re.I)[0].strip()
        before_grader = self._strip_card_variant_tail(before_grader)
        number_match = re.search(r"\b\d+[A-Za-z]?\s+([A-Z][A-Za-z'.-]+(?:\s+[A-Z][A-Za-z'.-]+){1,3})\s*$", before_grader)
        if number_match:
            return number_match.group(1).strip()
        words = re.findall(r"[A-Z][A-Za-z'.-]+", before_grader)
        stop_words = {
            "Topps", "Panini", "Bowman", "Donruss", "Optic", "Prizm", "Select", "Mosaic", "Contenders",
            "Chrome", "Finest", "Upper", "Deck", "Fleer", "Score", "Absolute", "Elite", "Stadium",
            "Club", "Uniformity", "Dominance", "Rookie", "Prospect", "Autographs", "Variation",
        }
        candidates = [word for word in words if word not in stop_words and not re.fullmatch(r"[IVX]+", word)]
        return " ".join(candidates[-2:]).strip() if len(candidates) >= 2 else ""

    def _strip_card_variant_tail(self, value: str) -> str:
        text = str(value or "").strip()
        tail_terms = (
            "chrome-refractor",
            "chrome refractor",
            "refractor",
            "silver",
            "prizm",
            "mosaic",
            "green",
            "blue",
            "red",
            "gold",
            "orange",
            "purple",
            "black",
            "white",
            "pink",
            "aqua",
            "teal",
            "auto",
            "autograph",
            "rookie",
        )
        changed = True
        while changed:
            changed = False
            for term in tail_terms:
                pattern = rf"(?:[- ]+{re.escape(term)})$"
                stripped = re.sub(pattern, "", text, flags=re.I).strip()
                if stripped != text:
                    text = stripped
                    changed = True
        return text

    def _auto_categorize_unassigned_players(self) -> dict[str, object]:
        entries = self._load_unassigned_players()
        resolved: list[tuple[str, str, str, str]] = []
        unresolved: list[str] = []
        errors: list[str] = []
        for key, entry in list(entries.items()):
            player = str(entry.get("player") or "").strip()
            if not player:
                unresolved.append(str(key))
                continue
            try:
                category, evidence = self._search_unassigned_player_category(entry)
            except Exception as error:
                errors.append(f"{player}: {error}")
                continue
            if category:
                resolved.append((str(key), player, category, evidence))
            else:
                unresolved.append(str(key))
        if resolved:
            for _key, player, category, _evidence in resolved:
                self._write_player_category_override(player, category)
            with shared_lock(CARD_PIPELINE_DIR, "unassigned-players", self.lucas_identity):
                latest = self._load_unassigned_players()
                for key, _player, _category, _evidence in resolved:
                    latest.pop(key, None)
                self._save_unassigned_players(latest)
            self.assignment_engine = AssignmentEngine.load()
        return {
            "resolved": len(resolved),
            "unresolved": len(unresolved),
            "errors": errors,
            "details": resolved,
        }

    def _search_unassigned_player_category(self, entry: dict[str, object]) -> tuple[str, str]:
        player = str(entry.get("player") or "").strip()
        title = str(entry.get("last_title") or "").strip()
        local_category, local_evidence = self._local_unassigned_player_category(entry)
        if local_category:
            return local_category, local_evidence
        query = " ".join(part for part in (player, title, "sports cards category") if part)
        text = " ".join(part for part in (player, title, self._category_research_text(player, query)) if part)
        return self._infer_category_from_web_text(text)

    def _local_unassigned_player_category(self, entry: dict[str, object]) -> tuple[str, str]:
        explicit = str(entry.get("sport") or entry.get("category") or "").strip()
        normalized = assignment_engine.canonical_sport_label(explicit)
        if normalized:
            return normalized, "saved entry category"
        player = str(entry.get("player") or "").strip()
        titles = [str(entry.get("last_title") or "").strip()]
        samples = entry.get("sample_titles")
        if isinstance(samples, list):
            titles.extend(str(sample or "").strip() for sample in samples)
        for title in [item for item in titles if item]:
            parsed = assignment_engine.parse_card_for_matching(title)
            sport = assignment_engine.canonical_sport_label(parsed.get("sport") or "")
            parsed_player = str(parsed.get("playerName") or "").strip()
            if sport and (not player or not parsed_player or parsed_player.lower() == player.lower() or player.lower() in title.lower()):
                return sport, "card title parser"
        local_text = " ".join(part for part in [player, *titles] if part)
        return self._infer_category_from_web_text(local_text)

    def _category_research_text(self, player: str, query: str) -> str:
        parts: list[str] = []
        for fetcher in (
            lambda: self._wikipedia_search_text(player),
            lambda: self._wikidata_search_text(player),
            lambda: self._duckduckgo_search_text(query),
        ):
            try:
                text = fetcher()
            except Exception:
                continue
            if text:
                parts.append(text)
        return " ".join(parts)

    def _wikipedia_search_text(self, player: str) -> str:
        search_url = "https://en.wikipedia.org/w/api.php?" + urllib.parse.urlencode({
            "action": "query",
            "list": "search",
            "srsearch": player,
            "format": "json",
            "srlimit": "3",
        })
        payload = self._read_json_url(search_url)
        snippets: list[str] = []
        for item in ((payload.get("query") or {}).get("search") or [])[:3]:
            title = str(item.get("title") or "")
            snippet = re.sub(r"<[^>]+>", " ", str(item.get("snippet") or ""))
            snippets.append(f"{title} {snippet}")
            if title:
                summary_url = "https://en.wikipedia.org/api/rest_v1/page/summary/" + urllib.parse.quote(title.replace(" ", "_"), safe="")
                try:
                    summary = self._read_json_url(summary_url)
                    snippets.append(" ".join(str(summary.get(key) or "") for key in ("title", "description", "extract")))
                except Exception:
                    pass
        return " ".join(snippets)

    def _wikidata_search_text(self, player: str) -> str:
        search_url = "https://www.wikidata.org/w/api.php?" + urllib.parse.urlencode({
            "action": "wbsearchentities",
            "search": player,
            "language": "en",
            "format": "json",
            "limit": "5",
        })
        payload = self._read_json_url(search_url)
        return " ".join(
            " ".join(str(item.get(key) or "") for key in ("label", "description"))
            for item in payload.get("search") or []
        )

    def _duckduckgo_search_text(self, query: str) -> str:
        return self._html_search_text(f"https://duckduckgo.com/html/?q={urllib.parse.quote_plus(query)}")

    def _web_search_text(self, query: str) -> str:
        return self._duckduckgo_search_text(query)

    def _read_json_url(self, url: str) -> dict[str, object]:
        request = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) LUCAS/1.0",
                "Accept": "application/json,text/plain,*/*",
            },
        )
        with urllib.request.urlopen(request, timeout=12) as response:
            return json.loads(response.read(200000).decode("utf-8", errors="replace"))

    def _html_search_text(self, url: str) -> str:
        request = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) LUCAS/1.0",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            },
        )
        with urllib.request.urlopen(request, timeout=12) as response:
            raw = response.read(200000).decode("utf-8", errors="replace")
        text = re.sub(r"<script\b.*?</script>|<style\b.*?</style>", " ", raw, flags=re.I | re.S)
        text = re.sub(r"<[^>]+>", " ", text)
        return re.sub(r"\s+", " ", html.unescape(text)).strip()

    def _infer_category_from_web_text(self, text: str) -> tuple[str, str]:
        haystack = re.sub(r"\s+", " ", str(text or "").lower())
        scores: dict[str, int] = {}
        for category, terms in ASSIGNMENT_CATEGORY_WEB_SIGNALS.items():
            score = 0
            for term in terms:
                cleaned = str(term).lower()
                matches = len(re.findall(rf"\b{re.escape(cleaned)}\b", haystack))
                if not matches:
                    continue
                score += matches * (4 if cleaned == category else 2 if " " in cleaned else 1)
            if score:
                scores[category] = score
        if not scores:
            return "", ""
        ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
        best_category, best_score = ranked[0]
        second_score = ranked[1][1] if len(ranked) > 1 else 0
        if best_score < 2 or (second_score and best_score < second_score + 2):
            return "", ""
        return best_category, f"score {best_score}"

    def open_unassigned_players_dialog(self) -> None:
        entries = self._load_unassigned_players()
        popup = tk.Toplevel(self)
        popup.title("Unassigned Players")
        popup.configure(bg="#1f1f1f")
        popup.transient(self)
        popup.grab_set()
        popup.geometry("980x560")
        popup.minsize(860, 480)

        frame = ttk.Frame(popup, style="Panel.TFrame", padding=(16, 14))
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text="Unassigned Players", style="Panel.TLabel", font=("Segoe UI Semibold", 13)).grid(row=0, column=0, columnspan=4, sticky="w")
        ttk.Label(frame, text="Search the player, choose a category, then save to teach assignment matching.", style="Muted.TLabel").grid(row=1, column=0, columnspan=4, sticky="w", pady=(2, 12))

        tree = ttk.Treeview(frame, columns=("player", "count", "last_seen", "title"), show="headings", selectmode="browse")
        headings = {"player": "Player", "count": "Count", "last_seen": "Last Seen", "title": "Sample Card"}
        widths = {"player": 180, "count": 60, "last_seen": 150, "title": 510}
        for column in headings:
            tree.heading(column, text=headings[column], anchor=tk.W)
            tree.column(column, width=widths[column], minwidth=50, stretch=column == "title")
        tree.grid(row=2, column=0, columnspan=4, sticky="nsew")
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        scrollbar.grid(row=2, column=4, sticky="ns")
        tree.configure(yscrollcommand=scrollbar.set)

        player_var = tk.StringVar()
        category_var = tk.StringVar(value=ASSIGNMENT_CATEGORY_OPTIONS[0])
        selected_key = tk.StringVar()
        ttk.Label(frame, text="Player", style="Panel.TLabel").grid(row=3, column=0, sticky="w", pady=(12, 4))
        player_entry = ttk.Entry(frame, textvariable=player_var)
        player_entry.grid(row=4, column=0, sticky="ew", padx=(0, 10))
        ttk.Label(frame, text="Category", style="Panel.TLabel").grid(row=3, column=1, sticky="w", pady=(12, 4))
        category_combo = ttk.Combobox(frame, textvariable=category_var, values=ASSIGNMENT_CATEGORY_OPTIONS, state="readonly", width=18)
        category_combo.grid(row=4, column=1, sticky="ew", padx=(0, 10))

        entry_by_iid: dict[str, dict[str, object]] = {}

        def refresh_tree() -> None:
            tree.delete(*tree.get_children())
            entry_by_iid.clear()
            latest = self._load_unassigned_players()
            for key, entry in sorted(latest.items(), key=lambda item: str(item[1].get("last_seen") or ""), reverse=True):
                iid = str(key)
                entry_by_iid[iid] = entry
                tree.insert("", tk.END, iid=iid, values=(
                    entry.get("player") or "",
                    entry.get("count") or 0,
                    entry.get("last_seen") or "",
                    entry.get("last_title") or "",
                ))

        def select_entry(_event=None) -> None:
            selected = tree.selection()
            if not selected:
                return
            iid = selected[0]
            selected_key.set(iid)
            entry = entry_by_iid.get(iid, {})
            player_var.set(str(entry.get("player") or ""))

        def search_selected() -> None:
            key = selected_key.get()
            entry = entry_by_iid.get(key, {})
            query = " ".join(part for part in (player_var.get().strip(), "sports cards", str(entry.get("last_title") or "")) if part)
            if not query.strip():
                return
            webbrowser.open(f"https://www.google.com/search?q={urllib.parse.quote_plus(query)}")

        def save_mapping() -> None:
            player = player_var.get().strip()
            category = category_var.get().strip()
            key = selected_key.get()
            if not player or not category or not key:
                messagebox.showinfo("Unassigned player", "Choose an entry, enter the player name, and choose a category.")
                return
            self.save_player_category_override(player, category, key)
            refresh_tree()

        def remove_entry() -> None:
            key = selected_key.get()
            if not key:
                return
            with shared_lock(CARD_PIPELINE_DIR, "unassigned-players", self.lucas_identity):
                latest = self._load_unassigned_players()
                latest.pop(key, None)
                self._save_unassigned_players(latest)
            selected_key.set("")
            player_var.set("")
            refresh_tree()

        def auto_categorize_all() -> None:
            auto_button.configure(state=tk.DISABLED)
            self.status_var.set("Auto-categorizing unassigned players...")

            def worker() -> None:
                try:
                    result = self._auto_categorize_unassigned_players()
                except Exception as error:
                    self.after(0, lambda: finish_auto({"resolved": 0, "unresolved": 0, "errors": [str(error)]}))
                    return
                self.after(0, lambda: finish_auto(result))

            def finish_auto(result: dict[str, object]) -> None:
                auto_button.configure(state=tk.NORMAL)
                refresh_tree()
                resolved = int(result.get("resolved") or 0)
                unresolved = int(result.get("unresolved") or 0)
                errors = list(result.get("errors") or [])
                self.assignment_config_status.set(self._assignment_config_status())
                if resolved:
                    self._schedule_assignment_recommendations(delay_ms=50)
                if errors:
                    self.status_var.set(f"Auto-categorized {resolved}; {unresolved} left; {len(errors)} search issue(s).")
                    messagebox.showwarning("Auto Categorize", "\n".join([f"Resolved: {resolved}", f"Left: {unresolved}", "", "Issues:", *[str(error) for error in errors[:8]]]))
                else:
                    self.status_var.set(f"Auto-categorized {resolved} unassigned player(s); {unresolved} left.")
                    messagebox.showinfo("Auto Categorize", f"Resolved: {resolved}\nLeft unresolved: {unresolved}")

            threading.Thread(target=worker, daemon=True).start()

        tree.bind("<<TreeviewSelect>>", select_entry)
        ttk.Button(frame, text="Web Search", command=search_selected, style="Soft.TButton").grid(row=4, column=2, sticky="ew", padx=(0, 10))
        ttk.Button(frame, text="Save Category", command=save_mapping, style="Primary.TButton").grid(row=4, column=3, sticky="ew")
        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.grid(row=5, column=0, columnspan=4, sticky="e", pady=(12, 0))
        auto_button = ttk.Button(buttons, text="Auto Categorize All", command=auto_categorize_all, style="Primary.TButton")
        auto_button.pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="Remove", command=remove_entry, style="Soft.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(buttons, text="Close", command=popup.destroy, style="Soft.TButton").pack(side=tk.LEFT)
        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(3, weight=1)
        frame.rowconfigure(2, weight=1)
        refresh_tree()
        if tree.get_children():
            first = tree.get_children()[0]
            tree.selection_set(first)
            tree.focus(first)
            select_entry()

    def save_player_category_override(self, player: str, category: str, unassigned_key: str = "") -> None:
        self._write_player_category_override(player, category)
        self.assignment_engine = AssignmentEngine.load()
        if unassigned_key:
            with shared_lock(CARD_PIPELINE_DIR, "unassigned-players", self.lucas_identity):
                entries = self._load_unassigned_players()
                entries.pop(unassigned_key, None)
                self._save_unassigned_players(entries)
        self.assignment_config_status.set(self._assignment_config_status())
        self._schedule_assignment_recommendations(delay_ms=50)
        self.status_var.set(f"Saved {player} as {category}. Assignment will recalculate.")

    def _write_player_category_override(self, player: str, category: str) -> None:
        with shared_lock(CARD_PIPELINE_DIR, "assignment-player-overrides", self.lucas_identity):
            try:
                payload = json.loads(PLAYER_OVERRIDES_PATH.read_text(encoding="utf-8"))
            except Exception:
                payload = {}
            players = payload.get("players") if isinstance(payload, dict) else {}
            if not isinstance(players, dict):
                players = {}
            players[player] = {"sport": category, "displayName": player}
            payload = {"players": players}
            PLAYER_OVERRIDES_PATH.parent.mkdir(parents=True, exist_ok=True)
            atomic_write_json(PLAYER_OVERRIDES_PATH, payload)
        assignment_engine.add_player_sport_hint(player, category, display_name=player)

    def _queue_assignment_recommendations(self) -> None:
        row_ids = self.assignment_recommendation_row_ids
        self.assignment_recommendation_row_ids = None
        self.assignment_recommendation_after_id = None
        all_rows = [*self.state.rows, *self.review_rows]
        rows = [row for row in all_rows if id(row) in row_ids] if row_ids is not None else all_rows
        if not rows or not self.assignment_engine.companies:
            self.assignment_progress_value.set(0)
            return
        self.assignment_recommendation_job += 1
        job_id = self.assignment_recommendation_job
        self.assignment_recommendation_running = True
        self.assignment_progress_value.set(0)
        total = len(rows)
        self.review_status.set(f"Calculating assignment recommendations: 0/{total}...")
        self.status_var.set("Calculating assignment recommendations...")
        threading.Thread(target=self._assignment_recommendations_worker, args=(job_id, rows), daemon=True).start()

    def _schedule_assignment_recommendations(self, delay_ms: int = 700, row_ids: set[int] | None = None) -> None:
        if not self.assignment_engine.companies:
            self.assignment_progress_value.set(0)
            return
        if row_ids is None:
            self.assignment_recommendation_row_ids = None
        elif self.assignment_recommendation_row_ids is not None:
            self.assignment_recommendation_row_ids.update(row_ids)
        elif self.assignment_recommendation_after_id is None:
            self.assignment_recommendation_row_ids = set(row_ids)
        if self.assignment_recommendation_after_id is not None:
            try:
                self.after_cancel(self.assignment_recommendation_after_id)
            except tk.TclError:
                pass
        self.assignment_recommendation_after_id = self.after(delay_ms, self._queue_assignment_recommendations)

    def _assignment_recommendations_worker(self, job_id: int, rows: list[WorkbookRow]) -> None:
        perf_start = time.perf_counter()
        total = len(rows)
        results: list[tuple[int, str, float | None]] = []
        progress_step = max(1, total // 25)
        try:
            for index, row in enumerate(rows, start=1):
                recommendation = self.assignment_engine.recommend(row, person=getattr(self, "_assignment_person_for_row", lambda _row: "")(row))
                results.append((id(row), recommendation.company, recommendation.payout))
                if index == total or index % progress_step == 0:
                    self.events.put(("assignment_recommendations_progress", {"job_id": job_id, "done": index, "total": total}))
        except Exception as error:
            record_performance_event("assignment.recommendations", perf_start, f"job={job_id} rows={total} error={error}", force=True)
            self.events.put(("assignment_recommendations_error", {"job_id": job_id, "error": str(error)}))
            return
        record_performance_event("assignment.recommendations", perf_start, f"job={job_id} rows={total} results={len(results)}")
        self.events.put(("assignment_recommendations_done", {"job_id": job_id, "total": total, "results": results}))

    def _assignment_person_for_row(self, row: WorkbookRow) -> str:
        if getattr(self, "_is_personal_lucas", lambda: False)():
            return self._personal_default_person()
        selected_working_sheet = getattr(self, "selected_working_sheet", None)
        source_sheet = (
            getattr(self, "comp_sheet_sources", {}).get(row.excel_row)
            or getattr(self, "review_sheet_sources", {}).get(row.excel_row)
            or getattr(self, "intake_sheet_sources", {}).get(row.excel_row)
            or getattr(self, "row_sources", {}).get(row.excel_row)
            or getattr(self, "review_sources", {}).get(row.excel_row)
            or getattr(self, "intake_sources", {}).get(row.excel_row)
            or (selected_working_sheet.get() if selected_working_sheet is not None else "")
            or ""
        )
        source_name = Path(str(source_sheet or "")).name
        if not source_name or source_name == "NO SHEET FOUND":
            return ""
        for stage in ("Working", "Incoming", "Received"):
            marker = self.home_sheet_markers.get(self._home_sheet_key(stage, source_name), {})
            person = str(marker.get("assigned_person") or "").strip()
            if person:
                return person
        for key, marker in self.home_sheet_markers.items():
            _stage, name = self._split_home_sheet_key(key)
            if Path(name).name == source_name:
                person = str(marker.get("assigned_person") or "").strip()
                if person:
                    return person
        return ""

    def _apply_assignment_recommendation_results(self, payload: dict[str, object]) -> None:
        if int(payload.get("job_id") or 0) != self.assignment_recommendation_job:
            return
        results = {
            int(row_id): (str(company or ""), payout)
            for row_id, company, payout in list(payload.get("results") or [])
        }
        filled = 0
        comp_rows_updated = False
        review_rows_updated = False
        unresolved_review_rows: list[WorkbookRow] = []
        state_row_ids = {id(row) for row in self.state.rows}
        review_row_ids = {id(row) for row in self.review_rows}
        for row in [*self.state.rows, *self.review_rows]:
            if id(row) not in results:
                continue
            company, payout = results.get(id(row), ("", None))
            if payout is not None:
                row.best_company = company
                row.estimated_payout = payout
                filled += 1
                if id(row) in state_row_ids:
                    comp_rows_updated = True
                if id(row) in review_row_ids:
                    review_rows_updated = True
            else:
                row.best_company = NO_COMPANY_TAKES_LABEL
                row.estimated_payout = None
                if id(row) in review_row_ids:
                    unresolved_review_rows.append(row)
                    review_rows_updated = True
        self._record_unassigned_players(unresolved_review_rows)
        if comp_rows_updated:
            self.comp_output_saved = False
        total = int(payload.get("total") or 0)
        self.assignment_recommendation_running = False
        self.assignment_progress_value.set(100 if total else 0)
        self.review_status.set(f"Assignment recommendations complete: {filled}/{total} row(s) populated.")
        self.status_var.set(f"Assignment recommendations complete: {filled}/{total} row(s) populated.")
        if comp_rows_updated and not review_rows_updated:
            self._refresh_comp_table(schedule_recommendations=False)
        else:
            self._refresh_table(schedule_recommendations=False)

    def _apply_assignment_to_comp_rows(self, row_ids: set[int]) -> int:
        if not row_ids or not self.assignment_engine.companies:
            return 0
        perf_start = time.perf_counter()
        updated = 0
        for row in self.state.rows:
            if id(row) not in row_ids:
                continue
            recommendation = self.assignment_engine.recommend(row, person=getattr(self, "_assignment_person_for_row", lambda _row: "")(row))
            if recommendation.payout is not None:
                row.best_company = recommendation.company
                row.estimated_payout = recommendation.payout
            else:
                row.best_company = NO_COMPANY_TAKES_LABEL
                row.estimated_payout = None
            updated += 1
        if updated:
            self.comp_output_saved = False
        record_performance_event("assignment.apply_comp_rows", perf_start, f"rows={len(row_ids)} updated={updated}")
        return updated

    def _assignment_person_for_row(self, row: WorkbookRow) -> str:
        if getattr(self, "_is_personal_lucas", lambda: False)():
            return self._personal_default_person()
        selected_working_sheet = getattr(self, "selected_working_sheet", None)
        source_sheet = (
            getattr(self, "comp_sheet_sources", {}).get(row.excel_row)
            or getattr(self, "review_sheet_sources", {}).get(row.excel_row)
            or getattr(self, "intake_sheet_sources", {}).get(row.excel_row)
            or getattr(self, "row_sources", {}).get(row.excel_row)
            or getattr(self, "review_sources", {}).get(row.excel_row)
            or getattr(self, "intake_sources", {}).get(row.excel_row)
            or (selected_working_sheet.get() if selected_working_sheet is not None else "")
            or ""
        )
        source_name = Path(str(source_sheet or "")).name
        if not source_name or source_name == "NO SHEET FOUND":
            return ""
        for stage in ("Working", "Incoming", "Received"):
            marker = self.home_sheet_markers.get(self._home_sheet_key(stage, source_name), {})
            person = str(marker.get("assigned_person") or "").strip()
            if person:
                return person
        for key, marker in self.home_sheet_markers.items():
            _stage, name = self._split_home_sheet_key(key)
            if Path(name).name == source_name:
                person = str(marker.get("assigned_person") or "").strip()
                if person:
                    return person
        return ""

    def _update_assignment_recommendation_progress(self, payload: dict[str, object]) -> None:
        if int(payload.get("job_id") or 0) != self.assignment_recommendation_job:
            return
        done = int(payload.get("done") or 0)
        total = int(payload.get("total") or 0)
        percent = (done / total * 100) if total else 0
        self.assignment_progress_value.set(percent)
        self.review_status.set(f"Calculating assignment recommendations: {done}/{total}...")

    def _handle_assignment_recommendation_error(self, payload: dict[str, object]) -> None:
        if int(payload.get("job_id") or 0) != self.assignment_recommendation_job:
            return
        error = str(payload.get("error") or "Unknown error")
        self.assignment_recommendation_running = False
        self.assignment_progress_value.set(0)
        self.review_status.set(f"Assignment recommendations failed: {error}")
        self.status_var.set(f"Assignment recommendations failed: {error}")

    def reload_assignment_rules(self) -> None:
        perf_start = time.perf_counter()
        self._load_player_overrides()
        assignment_start = time.perf_counter()
        self.assignment_engine = AssignmentEngine.load()
        record_performance_event("assignment.load.reload", assignment_start, f"companies={len(self.assignment_engine.companies)}")
        self._refresh_keep_source_registry()
        self._ensure_company_sheet_folders()
        self._ensure_weekly_company_sheets_due()
        self.assignment_config_status.set(self._assignment_config_status())
        self._refresh_table(schedule_recommendations=True)
        self.review_status.set("Assignment rules reloaded.")
        self.status_var.set("Assignment rules reloaded.")
        record_performance_event("assignment.reload", perf_start, f"companies={len(self.assignment_engine.companies)}")

    def open_assignment_rules(self) -> None:
        open_assignment_rules_dialog(self, CARD_PIPELINE_DIR, self.reload_assignment_rules)

    def open_people_rules(self) -> None:
        open_people_rules_dialog(self, CARD_PIPELINE_DIR, self._assignment_company_health_payload(), self._after_people_rules_saved)

    def _after_people_rules_saved(self) -> None:
        self._refresh_seller_terms_dropdowns()
        self._refresh_person_combo_values()
        self.assignment_config_status.set(self._assignment_config_status())
        self.status_var.set("People rules saved.")

    def _assignment_config_status(self) -> str:
        if self.assignment_engine.error:
            return f"Company Rules config error: {self.assignment_engine.error}"
        count = len(self.assignment_engine.companies)
        if not count:
            return "Assignment companies: none configured. Add assignment_companies.json to enable best-company payouts."
        return f"Assignment companies loaded: {count}"

    def _refresh_table(self, schedule_recommendations: bool = False) -> None:
        self._render_rows(self.intake_tree, self.intake_rows, self.intake_sources)
        self._render_rows(self.comp_tree, self.state.rows, self.row_sources, self.comp_sheet_sources)
        self._render_rows(self.receive_tree, self.review_rows, self.review_sources, self.review_sheet_sources)
        self._render_rows(self.review_tree, self.review_rows, self.review_sources, self.review_sheet_sources)
        completed = sum(1 for row in self.state.rows if row.card_ladder_value is not None)
        self.summary_var.set(f"{len(self.intake_rows)} create rows | Loaded comp rows: {len(self.state.rows)} | Card Ladder values: {completed}")
        if schedule_recommendations:
            self._schedule_assignment_recommendations()

    def _refresh_comp_table(self, schedule_recommendations: bool = False, recommendation_row_ids: set[int] | None = None) -> None:
        self._render_rows(self.comp_tree, self.state.rows, self.row_sources, self.comp_sheet_sources)
        completed = sum(1 for row in self.state.rows if row.card_ladder_value is not None)
        self.summary_var.set(f"{len(self.intake_rows)} create rows | Loaded comp rows: {len(self.state.rows)} | Card Ladder values: {completed}")
        if schedule_recommendations:
            self._schedule_assignment_recommendations(row_ids=recommendation_row_ids)

    def _render_rows(self, tree: ttk.Treeview, rows: list[WorkbookRow], sources: dict[int, str], sheet_sources: dict[int, str] | None = None) -> None:
        self._remember_column_widths(tree)
        tree.delete(*tree.get_children())
        duplicate_certs = self._duplicate_certs(rows)
        columns = self._tree_columns(tree)
        for row in rows:
            tags = []
            if row.cert_number and row.cert_number in duplicate_certs:
                tags.append("duplicate_cert")
            if (sheet_sources or {}).get(row.excel_row) == "NO SHEET FOUND":
                tags.append("no_sheet_found")
            tree.insert(
                "",
                tk.END,
                iid=str(row.excel_row),
                tags=tuple(tags),
                values=tuple(self._row_display_value(row, col, sources, sheet_sources) for col in columns),
            )
        add_row_iid = ""
        should_show_add_row = False
        if tree is self.intake_tree and self.input_mode.get() == "Manual Entry":
            add_row_iid = ADD_INTAKE_ROW_IID
            should_show_add_row = True
        elif tree is self.comp_tree:
            add_row_iid = ADD_COMP_ROW_IID
            should_show_add_row = True
        elif self._is_receive_tree(tree) and self.review_mode.get() == "Manual Receive":
            add_row_iid = ADD_REVIEW_ROW_IID
            should_show_add_row = True
        if should_show_add_row:
            add_values = []
            for col in columns:
                if col == "excel_row":
                    add_values.append("+")
                elif col == "card_title":
                    add_values.append("Add row")
                else:
                    add_values.append("")
            tree.insert(
                "",
                tk.END,
                iid=add_row_iid,
                tags=("add_review_row",),
                values=tuple(add_values),
            )
        self._restore_column_widths(tree)

    def _tree_columns(self, tree: ttk.Treeview) -> tuple[str, ...]:
        return tuple(getattr(tree, "_display_columns", DISPLAY_COLUMNS))

    def _is_receive_tree(self, tree: ttk.Treeview) -> bool:
        return hasattr(self, "receive_tree") and tree is self.receive_tree

    def _is_review_row_tree(self, tree: ttk.Treeview) -> bool:
        return self._is_receive_tree(tree) or (hasattr(self, "review_tree") and tree is self.review_tree)

    def _row_display_value(
        self,
        row: WorkbookRow,
        column: str,
        sources: dict[int, str],
        sheet_sources: dict[int, str] | None,
    ) -> object:
        if column == "excel_row":
            return row.excel_row
        if column == "source":
            return sources.get(row.excel_row, "")
        if column == "person":
            return self._assignment_person_for_row(row)
        if column == "item_id":
            return getattr(row, "item_id", "")
        if column == "sheet_source":
            return (sheet_sources or {}).get(row.excel_row, "")
        if column == "cert_number":
            return row.cert_number
        if column == "grader":
            return row.grader
        if column == "category":
            return row.category
        if column == "card_title":
            return row.card_title
        if column == "purchase_price":
            return format_money(row.existing_value if isinstance(row.existing_value, (int, float)) else None)
        if column == "card_ladder_value":
            return format_money(row.card_ladder_value)
        if column == "card_ladder_comps_average":
            return format_money(row.card_ladder_comps_average)
        if column == "cy_value":
            return format_money(row.cy_value)
        if column == "cy_confidence":
            return "" if row.cy_confidence is None else row.cy_confidence
        if column == "best_company":
            return row.best_company
        if column == "estimated_payout":
            return format_money(row.estimated_payout)
        if column == "received":
            return "X" if getattr(row, "received", False) else ""
        if column == "status":
            return row.status
        if column == "company_pile":
            return "[x]" if row.company_pile else "[ ]"
        return ""

    def _delete_selected_table_rows(self, event) -> str | None:
        tree = event.widget
        if tree is self.intake_tree:
            self.delete_selected_intake_rows()
            return "break"
        if tree is self.comp_tree:
            self.delete_selected_comp_rows()
            return "break"
        if self._is_review_row_tree(tree):
            self.delete_selected_review_rows()
            return "break"
        return None

    def _delete_selected_rows(
        self,
        tree: ttk.Treeview,
        rows: list[WorkbookRow],
        sources: dict[int, str],
        sheet_sources: dict[int, str],
    ) -> int:
        selected_rows = {
            int(iid)
            for iid in tree.selection()
            if str(iid).isdigit() and str(iid) not in {ADD_INTAKE_ROW_IID, ADD_REVIEW_ROW_IID, ADD_COMP_ROW_IID}
        }
        if not selected_rows:
            return 0
        remaining: list[WorkbookRow] = []
        new_sources: dict[int, str] = {}
        new_sheet_sources: dict[int, str] = {}
        for next_excel_row, row in enumerate((row for row in rows if row.excel_row not in selected_rows), start=2):
            old_excel_row = row.excel_row
            row.excel_row = next_excel_row
            remaining.append(row)
            if old_excel_row in sources:
                new_sources[next_excel_row] = sources[old_excel_row]
            if old_excel_row in sheet_sources:
                new_sheet_sources[next_excel_row] = sheet_sources[old_excel_row]
        if tree is self.intake_tree:
            self.intake_rows = remaining
            self.intake_sources = new_sources
            self.intake_sheet_sources = new_sheet_sources
        elif tree is self.comp_tree:
            self.state.set_rows(remaining)
            self.row_sources = new_sources
            self.comp_sheet_sources = new_sheet_sources
        elif self._is_review_row_tree(tree):
            self.review_rows = remaining
            self.review_sources = new_sources
            self.review_sheet_sources = new_sheet_sources
        else:
            return 0
        self._cancel_cell_edit()
        self._refresh_table(schedule_recommendations=(tree is self.comp_tree or tree is self.review_tree))
        return len(selected_rows)

    def _duplicate_certs(self, rows: list[WorkbookRow]) -> set[str]:
        counts: dict[str, int] = {}
        for row in rows:
            cert = str(row.cert_number or "").strip().upper()
            if not cert:
                continue
            counts[cert] = counts.get(cert, 0) + 1
        return {cert for cert, count in counts.items() if count > 1}

    def _remember_column_widths(self, tree: ttk.Treeview) -> None:
        widths = self.column_widths_by_tree.setdefault(id(tree), {})
        for col in self._tree_columns(tree):
            try:
                widths[col] = int(tree.column(col, "width"))
            except tk.TclError:
                pass

    def _restore_column_widths(self, tree: ttk.Treeview) -> None:
        widths = self.column_widths_by_tree.get(id(tree), {})
        for col in self._tree_columns(tree):
            if col in widths:
                tree.column(col, width=widths[col])

    def _select_excel_row(self, excel_row: int) -> None:
        iid = str(excel_row)
        if self.intake_tree.exists(iid):
            self.intake_tree.selection_set(iid)
            self.intake_tree.focus(iid)
            self.intake_tree.see(iid)

    def _handle_table_click(self, event):
        tree = event.widget
        row_id = tree.identify_row(event.y)
        column_id = tree.identify_column(event.x)
        if tree is self.intake_tree and row_id == ADD_INTAKE_ROW_IID:
            self.add_manual_intake_row()
            return "break"
        if tree is self.comp_tree and row_id == ADD_COMP_ROW_IID:
            self.add_comp_row()
            return "break"
        if self._is_receive_tree(tree) and row_id == ADD_REVIEW_ROW_IID:
            self.add_manual_review_row()
            return "break"
        if self._is_receive_tree(tree) and row_id and column_id:
            column_index = int(column_id.replace("#", "")) - 1
            columns = self._tree_columns(tree)
            if 0 <= column_index < len(columns) and columns[column_index] == "company_pile":
                self._toggle_company_pile(row_id)
                return "break"
        return None

    def _toggle_company_pile(self, row_id: str) -> None:
        if not str(row_id).isdigit():
            return
        excel_row = int(row_id)
        for row in self.review_rows:
            if row.excel_row != excel_row:
                continue
            row.company_pile = not row.company_pile
            self._refresh_table()
            self.review_status.set("Company pile checked." if row.company_pile else "Company pile unchecked.")
            return

    def _begin_cell_edit(self, event) -> None:
        tree = event.widget
        row_id = tree.identify_row(event.y)
        column_id = tree.identify_column(event.x)
        if tree is self.intake_tree and row_id == ADD_INTAKE_ROW_IID:
            self.add_manual_intake_row()
            return
        if tree is self.comp_tree and row_id == ADD_COMP_ROW_IID:
            self.add_comp_row()
            return
        if self._is_receive_tree(tree) and row_id == ADD_REVIEW_ROW_IID:
            self.add_manual_review_row()
            return
        if not row_id or not column_id:
            return
        column_index = int(column_id.replace("#", "")) - 1
        columns = self._tree_columns(tree)
        if column_index < 0 or column_index >= len(columns):
            return
        column = columns[column_index]
        if column not in EDITABLE_COLUMNS:
            return
        bbox = tree.bbox(row_id, column_id)
        if not bbox:
            return
        self._cancel_cell_edit()
        x, y, width, height = bbox
        current = tree.set(row_id, column)
        is_receive_card_autocomplete = self._is_review_row_tree(tree) and column == "card_title"
        editor: ttk.Entry | ttk.Combobox
        if is_receive_card_autocomplete:
            editor = ttk.Combobox(tree, values=(), width=max(24, width // 8))
            self._refresh_receive_card_autocomplete(editor, current)
        else:
            editor = ttk.Entry(tree)
        editor.insert(0, current)
        editor.select_range(0, tk.END)
        editor.place(x=x, y=y, width=width, height=height)
        editor.focus_set()
        self.cell_editor = editor
        self.cell_edit = (tree, row_id, column)
        editor.bind("<Return>", lambda _event: self._commit_cell_edit())
        editor.bind("<KP_Enter>", lambda _event: self._commit_cell_edit())
        editor.bind("<Escape>", lambda _event: self._cancel_cell_edit())
        editor.bind("<FocusOut>", lambda _event: self._commit_cell_edit())
        if is_receive_card_autocomplete:
            editor.bind("<KeyRelease>", lambda event, widget=editor: self._on_receive_card_autocomplete_key(event, widget), add="+")
            editor.bind("<<ComboboxSelected>>", lambda _event: self._commit_cell_edit(), add="+")

    def _commit_cell_edit(self) -> None:
        if not self.cell_editor or not self.cell_edit:
            return
        tree, row_id, column = self.cell_edit
        value = self.cell_editor.get()
        current = tree.set(row_id, column)
        selected_match = self._selected_receive_autocomplete_match(value) if self._is_review_row_tree(tree) and column == "card_title" else {}
        if selected_match:
            value = str(selected_match.get("card_title") or value)
        self._destroy_cell_editor()
        if value.strip() == str(current or "").strip() and not selected_match:
            return
        excel_row = int(row_id)
        self._apply_cell_value(tree, excel_row, column, value)
        if selected_match:
            self._apply_receive_match_to_existing_row(tree, excel_row, selected_match)
        if tree is self.comp_tree:
            self.comp_output_saved = False
        self._refresh_table(schedule_recommendations=self._edit_affects_assignment(tree, column))
        if tree.exists(row_id):
            tree.selection_set(row_id)
            tree.focus(row_id)
            tree.see(row_id)
        self.status_var.set(f"Updated row {excel_row}.")

    def _on_receive_card_autocomplete_key(self, event, editor: ttk.Combobox) -> None:
        if event.keysym in {"Return", "KP_Enter", "Escape", "Up", "Down", "Left", "Right", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R"}:
            return
        self._refresh_receive_card_autocomplete(editor, editor.get())

    def _apply_receive_match_to_existing_row(self, tree: ttk.Treeview, excel_row: int, match: dict[str, object]) -> None:
        if not self._is_review_row_tree(tree) or not match:
            return
        self.review_sheet_sources[excel_row] = str(match.get("sheet") or "NO SHEET FOUND")
        for row in self.review_rows:
            if row.excel_row != excel_row:
                continue
            self._attach_receive_match_to_row(row, match)
            row.status = "Received"
            if match.get("item_id"):
                row.item_id = str(match.get("item_id") or "")
            if match.get("cert_number"):
                row.cert_number = str(match.get("cert_number") or "")
            if match.get("card_title"):
                row.card_title = str(match.get("card_title") or "")
            if match.get("grader"):
                row.grader = str(match.get("grader") or "")
            if match.get("sport") or match.get("category"):
                row.category = str(match.get("sport") or match.get("category") or "")
            if match.get("purchase_price") is not None:
                row.existing_value = match.get("purchase_price")
            if match.get("card_ladder_value") is not None:
                row.card_ladder_value = match.get("card_ladder_value")
            if match.get("card_ladder_comps_average") is not None:
                row.card_ladder_comps_average = match.get("card_ladder_comps_average")
            if match.get("cy_value") is not None:
                row.cy_value = match.get("cy_value")
            if match.get("cy_confidence") is not None:
                row.cy_confidence = match.get("cy_confidence")
            if match.get("card_ladder_comps"):
                row.card_ladder_comps = str(match.get("card_ladder_comps") or "")
            if match.get("best_company"):
                row.best_company = str(match.get("best_company") or "")
            if match.get("estimated_payout") is not None:
                row.estimated_payout = match.get("estimated_payout")
            self._ensure_receive_row_assignment(row)
            self.review_status.set(f"Matched receive row to {match.get('sheet') or 'incoming sheet'}.")
            return

    def _edit_affects_assignment(self, tree: ttk.Treeview, column: str) -> bool:
        if tree is not self.comp_tree and tree is not self.review_tree:
            return False
        return column in {
            "cert_number",
            "grader",
            "category",
            "card_title",
            "card_ladder_value",
            "card_ladder_comps_average",
            "cy_value",
        }

    def _cancel_cell_edit(self) -> None:
        self._destroy_cell_editor()

    def _destroy_cell_editor(self) -> None:
        if self.cell_editor is not None:
            try:
                self.cell_editor.destroy()
            except tk.TclError:
                pass
        self.cell_editor = None
        self.cell_edit = None

    def _apply_cell_value(self, tree: ttk.Treeview, excel_row: int, column: str, value: str) -> None:
        clean_value = value.strip()
        if tree is self.comp_tree:
            target_rows = self.state.rows
            target_sources = self.row_sources
            target_sheet_sources = self.comp_sheet_sources
        elif self._is_review_row_tree(tree):
            target_rows = self.review_rows
            target_sources = self.review_sources
            target_sheet_sources = self.review_sheet_sources
        else:
            target_rows = self.intake_rows
            target_sources = self.intake_sources
            target_sheet_sources = self.intake_sheet_sources
        if column == "source":
            target_sources[excel_row] = clean_value
            return
        if column == "sheet_source":
            target_sheet_sources[excel_row] = clean_value
            return
        for row in target_rows:
            if row.excel_row != excel_row:
                continue
            previous_cert = scan_to_cert(row.cert_number)
            if column == "item_id":
                row.item_id = clean_value
            elif column == "cert_number":
                row.cert_number = scan_to_cert(clean_value)
            elif column == "grader":
                row.grader = normalize_grader(clean_value) or clean_value.upper()
            elif column == "category":
                row.category = clean_value
            elif column == "card_title":
                row.card_title = clean_value
                inferred = infer_grader(row.card_title)
                if inferred:
                    row.grader = inferred
                if not row.category:
                    row.category = str(assignment_engine.parse_card_for_matching(row.card_title).get("sport") or "")
            elif column == "purchase_price":
                row.existing_value = self._parse_money_text(clean_value)
            elif column == "card_ladder_value":
                row.card_ladder_value = self._parse_money_text(clean_value)
            elif column == "card_ladder_comps_average":
                row.card_ladder_comps_average = self._parse_money_text(clean_value)
            elif column == "cy_value":
                row.cy_value = self._parse_money_text(clean_value)
            elif column == "cy_confidence":
                row.cy_confidence = clean_value
            row.status = "Ready" if (row.cert_number and row.grader) or row.item_id else "Needs setup"
            if tree is self.intake_tree:
                if column == "purchase_price":
                    setattr(row, "_seller_terms_base_purchase", row.existing_value)
                if column in {"purchase_price", "card_ladder_value", "card_ladder_comps_average", "cy_value"}:
                    self.apply_create_seller_terms(show_status=False)
            if self._is_review_row_tree(tree) and (
                (column == "cert_number" and scan_to_cert(row.cert_number) != previous_cert)
                or column == "item_id"
                or (column == "card_title" and not scan_to_cert(row.cert_number) and not row.item_id and clean_value)
            ):
                if scan_to_cert(row.cert_number):
                    match = self._incoming_match(row.cert_number)
                else:
                    raw_match_query = {
                        "item_id": row.item_id,
                        "card_title": row.card_title,
                        "sheet_source": target_sheet_sources.get(excel_row, ""),
                        "receive_key": "" if column in {"item_id", "card_title"} else getattr(row, "_receive_key", ""),
                    }
                    match = self._incoming_raw_match(raw_match_query)
                    if not match and column == "card_title":
                        title_matches = self._incoming_title_matches(row.card_title, limit=2)
                        if len(title_matches) == 1:
                            match = title_matches[0]
                target_sheet_sources[excel_row] = str(match.get("sheet") or "NO SHEET FOUND")
                if match:
                    self._attach_receive_match_to_row(row, match)
                    row.status = "Received"
                    if not row.item_id and match.get("item_id"):
                        row.item_id = str(match.get("item_id") or "")
                    if is_placeholder_title(row.card_title, row.grader) and match.get("card_title"):
                        row.card_title = str(match.get("card_title") or "")
                    if not row.category and (match.get("sport") or match.get("category")):
                        row.category = str(match.get("sport") or match.get("category") or "")
                    if row.existing_value is None and match.get("purchase_price") is not None:
                        row.existing_value = match.get("purchase_price")
                    if row.card_ladder_value is None and match.get("card_ladder_value") is not None:
                        row.card_ladder_value = match.get("card_ladder_value")
                    if row.card_ladder_comps_average is None and match.get("card_ladder_comps_average") is not None:
                        row.card_ladder_comps_average = match.get("card_ladder_comps_average")
                    if row.cy_value is None and match.get("cy_value") is not None:
                        row.cy_value = match.get("cy_value")
                    if row.cy_confidence is None and match.get("cy_confidence") is not None:
                        row.cy_confidence = match.get("cy_confidence")
                    if not row.card_ladder_comps and match.get("card_ladder_comps"):
                        row.card_ladder_comps = str(match.get("card_ladder_comps") or "")
                elif row.cert_number or row.item_id:
                    target_sheet_sources[excel_row] = "NO SHEET FOUND"
                    setattr(row, "_receive_key", "")
                    setattr(row, "_receive_sheet", "")
                    setattr(row, "_receive_workbook_sheet", "")
                    setattr(row, "_receive_workbook_row", 0)
                    row.status = "Received - no incoming match"
            if not row.cert_number and not row.item_id:
                row.notes = "Missing cert"
            elif row.cert_number and not row.grader:
                row.notes = "Missing grader"
            elif row.notes in {"Missing cert", "Missing grader", "Missing cert or grader"}:
                row.notes = ""
            return

    def _poll_events(self) -> None:
        try:
            while True:
                event = self.events.get_nowait()
                if event == "refresh":
                    self._refresh_table()
                elif event == "comp_refresh":
                    self.comp_output_saved = False
                    with self.state.lock:
                        comp_running = bool(
                            self.state.cardladder_running
                            or getattr(self.state, "cy_batch_running", False)
                            or getattr(self.state, "cy_lookup_inflight", set())
                            or getattr(self.state, "cy_lookup_pending", set())
                        )
                        updated_row_ids = set(getattr(self.state, "updated_row_ids", set()))
                        self.state.updated_row_ids = set()
                    if self.inventory_recomp_context:
                        changed = self._sync_inventory_recomp_results()
                        if not comp_running:
                            self._finish_inventory_recomp()
                        elif changed:
                            self.refresh_inventory_tab()
                            self.inventory_status_var.set(f"Inventory recomp running; updated {changed} card(s) so far.")
                        continue
                    scoped_row_ids = updated_row_ids & self.pending_comp_assignment_row_ids
                    assigned = self._apply_assignment_to_comp_rows(scoped_row_ids)
                    if not comp_running:
                        self.pending_comp_assignment_row_ids = set()
                    self._refresh_comp_table(schedule_recommendations=False)
                    if assigned:
                        self.status_var.set(f"Updated assignment for {assigned} comped row(s).")
                elif isinstance(event, tuple):
                    kind, payload = event
                    if kind == "photo_rows":
                        self._append_rows(payload)
                    elif kind == "status":
                        self.status_var.set(str(payload))
                    elif kind == "photo_status":
                        self.photo_status.set(str(payload))
                        self.status_var.set(str(payload))
                    elif kind == "review_rows":
                        self._append_review_rows(payload)
                    elif kind == "review_status":
                        self.review_photo_status.set(str(payload))
                        self.review_status.set(str(payload))
                        self.status_var.set(str(payload))
                    elif kind == "inventory_photo_status":
                        self.inventory_status_var.set(str(payload))
                        self.status_var.set(str(payload))
                    elif kind == "startup_refresh":
                        self._apply_startup_refresh(payload)
                    elif kind == "load_working_sheet_done":
                        self._apply_loaded_working_sheet(
                            str(payload.get("name") or ""),
                            list(payload.get("rows") or []),
                            str(payload.get("stage") or "Working"),
                            str(payload.get("label") or ""),
                        )
                    elif kind == "load_working_sheet_error":
                        self.status_var.set(f"Comp sheet load failed: {payload.get('error')}")
                        messagebox.showerror("Load failed", str(payload.get("error") or "Unknown error"))
                    elif kind == "load_received_sheet_done":
                        self._apply_loaded_received_sheet(str(payload.get("name") or ""), list(payload.get("rows") or []))
                    elif kind == "load_received_sheet_error":
                        self.review_status.set(f"Received sheet load failed: {payload.get('error')}")
                        self.status_var.set(f"Received sheet load failed: {payload.get('error')}")
                        messagebox.showerror("Load failed", str(payload.get("error") or "Unknown error"))
                    elif kind == "assignment_recommendations_progress":
                        self._update_assignment_recommendation_progress(payload)
                    elif kind == "assignment_recommendations_done":
                        self._apply_assignment_recommendation_results(payload)
                    elif kind == "assignment_recommendations_error":
                        self._handle_assignment_recommendation_error(payload)
                    elif kind == "profit_recovery_done":
                        self._finish_profit_recovery(payload)
                    elif kind == "profit_recovery_error":
                        self._handle_profit_recovery_error(payload)
                    elif kind == "inventory_refresh":
                        self.refresh_inventory_tab(enrich=True)
                        if payload:
                            self.status_var.set(str(payload))
                    elif kind == "profit_refresh":
                        self.refresh_profit_tab()
                        self.refresh_payouts_tab()
                        self.status_var.set(str(payload))
        except queue.Empty:
            pass
        self.after(200, self._poll_events)

    def _parse_money_text(self, value: str) -> float | None:
        text = value.strip().replace("$", "").replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None

    def _arm_scanner(self) -> None:
        if self.input_mode.get() != "Barcode Scanner" or self.scan_entry is None:
            return
        try:
            self.scan_entry.focus_set()
            self.scan_entry.icursor(tk.END)
        except tk.TclError:
            pass

    def destroy(self) -> None:
        self._stop_instagram_background_tunnel()
        self.bridge.stop()
        super().destroy()


def first_number(*values: float | None) -> float | None:
    for value in values:
        if value is not None:
            return value
    return None


def is_placeholder_title(card_title: str, grader: str) -> bool:
    title = str(card_title or "").strip()
    company = str(grader or "").strip()
    if not title:
        return True
    return bool(company and title.upper() == company.upper())


if __name__ == "__main__":
    CardPipelineApp().mainloop()
