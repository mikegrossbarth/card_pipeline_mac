from __future__ import annotations

import csv
import json
import os
import re
import unicodedata
import urllib.parse
import urllib.request
import urllib.error
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from google_sheets_import import GoogleSheetsAuthError, export_google_sheet_to_xlsx, read_google_sheet_tabs, read_google_sheet_text


ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "assignment_companies.json"
PLAYER_SPORT_DATA_PATH = ROOT / "assignment_player_sport_data.js"
CATEGORY_ALIASES = {
    "football": ["football", "nfl"],
    "soccer": ["soccer", "futbol", "premier league", "uefa", "fifa"],
    "baseball": ["baseball", "mlb", "world series"],
    "basketball": ["basketball", "b-ball", "bball", "nba"],
    "hockey": ["hockey", "nhl"],
    "wnba": ["wnba"],
    "pokemon": ["pokemon", "poke"],
    "one piece": ["one piece", "onepiece", "one_piece", "1 piece"],
    "wwe": ["wwe", "wrestling", "wwf"],
    "f1": ["f1", "formula 1", "formula one", "formula1"],
    "marvel": ["marvel"],
    "disney": ["disney"],
    "star wars": ["star wars", "starwars"],
    "ufc": ["ufc", "mma"],
}
SINGLE_TOKEN_CONTEXT_REQUIRED_CATEGORIES = {
    "pokemon",
    "one piece",
    "marvel",
    "disney",
    "star wars",
}
PLAYER_SPORT_HINTS = {
    "victor wembanyama": "basketball",
    "stephen curry": "basketball",
    "steph curry": "basketball",
    "nikola jokic": "basketball",
    "luka doncic": "basketball",
    "giannis antetokounmpo": "basketball",
    "anthony edwards": "basketball",
    "kevin durant": "basketball",
    "ja morant": "basketball",
    "jayson tatum": "basketball",
    "shai gilgeous alexander": "basketball",
    "kemba walker": "basketball",
    "lebron james": "basketball",
    "michael jordan": "basketball",
    "kobe bryant": "basketball",
    "larry bird": "basketball",
    "magic johnson": "basketball",
    "kareem abdul jabbar": "basketball",
    "kareem abdul-jabbar": "basketball",
    "shaquille o'neal": "basketball",
    "tim duncan": "basketball",
    "wilt chamberlain": "basketball",
    "jerry west": "basketball",
    "tom brady": "football",
    "patrick mahomes": "football",
    "cj stroud": "football",
    "joe montana": "football",
    "jerry rice": "football",
    "barry sanders": "football",
    "peyton manning": "football",
    "dan marino": "football",
    "walter payton": "football",
    "randy moss": "football",
    "john elway": "football",
    "emmitt smith": "football",
    "deion sanders": "football",
    "aaron donald": "football",
    "shohei ohtani": "baseball",
    "shoehi ohtani": "baseball",
    "mike trout": "baseball",
    "aaron judge": "baseball",
    "luis robert": "baseball",
    "babe ruth": "baseball",
    "mickey mantle": "baseball",
    "lou gehrig": "baseball",
    "lou gherig": "baseball",
    "hank aaron": "baseball",
    "ken griffey jr": "baseball",
    "willie mays": "baseball",
    "sandy koufax": "baseball",
    "nolan ryan": "baseball",
    "randy johnson": "baseball",
    "ichiro suzuki": "baseball",
    "cal ripken jr": "baseball",
    "roberto clemente": "baseball",
    "jackie robinson": "baseball",
    "clayton kershaw": "baseball",
    "mookie betts": "baseball",
    "shintaro fujinami": "baseball",
    "yusniel diaz": "baseball",
    "jasson dominguez": "baseball",
    "ricardo olivar": "baseball",
    "chipper jones": "baseball",
    "ronald acuna jr": "baseball",
    "ronald acuna jr.": "baseball",
    "ronald acuña jr": "baseball",
    "ronald acuña jr.": "baseball",
    "lionel messi": "soccer",
    "cristiano ronaldo": "soccer",
    "erling haaland": "soccer",
    "connor bedard": "hockey",
    "wayne gretzky": "hockey",
    "aja wilson": "wnba",
    "a'ja wilson": "wnba",
    "caitlin clark": "wnba",
    "angel reese": "wnba",
    "sabrina ionescu": "wnba",
    "breanna stewart": "wnba",
    "diana taurasi": "wnba",
    "sue bird": "wnba",
    "candace parker": "wnba",
    "napheesa collier": "wnba",
    "kelsey plum": "wnba",
    "aliyah boston": "wnba",
    "paige bueckers": "wnba",
    "skylar diggins": "wnba",
    "elena delle donne": "wnba",
    "brittney griner": "wnba",
    "maya moore": "wnba",
    "lisa leslie": "wnba",
    "sheryl swoopes": "wnba",
}
PLAYER_DISPLAY_NAMES: dict[str, str] = {}
PLAYER_TEAM_HINTS: dict[str, list[str]] = {}
PARTIAL_PLAYER_HINTS: dict[str, dict[str, str]] = {}
PARTIAL_PLAYER_TOKEN_OVERRIDES = {
    "judge": "aaron judge",
}
SORTED_PLAYER_KEYS: list[str] = []
DISTINCTIVE_FIRST_NAMES = {
    "lebron",
    "kareem",
    "magic",
    "kobe",
    "shaquille",
    "hakeem",
    "giannis",
    "nikola",
    "dwyane",
    "kawhi",
    "dirk",
    "dolph",
    "manu",
    "shai",
    "peyton",
    "emmitt",
    "ladainian",
    "deion",
    "shoeless",
    "ichiro",
    "satchel",
    "jimmie",
    "yogi",
    "honus",
    "pedro",
}
AMBIGUOUS_PARTIAL_TOKENS = {
    "john",
    "joe",
    "bob",
    "jim",
    "mike",
    "steve",
    "david",
    "chris",
    "paul",
    "luis",
    "james",
    "thomas",
    "johnson",
    "brown",
    "white",
    "green",
    "young",
    "rose",
    "king",
    "hill",
    "bell",
    "reed",
    "allen",
    "george",
    "parker",
    "wilson",
    "martinez",
    "robinson",
    "jackson",
    "orange",
    "purple",
    "blue",
    "red",
    "gold",
    "silver",
    "black",
}
PRODUCT_WORDS = {
    "topps",
    "bowman",
    "chrome",
    "sapphire",
    "finest",
    "heritage",
    "stadium",
    "club",
    "panini",
    "donruss",
    "optic",
    "prizm",
    "select",
    "mosaic",
    "contenders",
    "national",
    "treasures",
    "flawless",
    "immaculate",
    "obsidian",
    "revolution",
    "absolute",
    "elite",
    "upper",
    "deck",
    "sp",
    "young",
    "guns",
    "pokemon",
    "one",
    "piece",
    "orange",
    "purple",
    "blue",
    "red",
    "green",
    "gold",
    "silver",
    "black",
    "white",
    "pink",
    "aqua",
    "teal",
    "bronze",
    "vinyl",
    "wave",
    "shimmer",
    "choice",
    "auto",
}
GOAT_PAYOUT_PLAYERS = {
    "tom brady",
    "stephen curry",
    "steph curry",
    "michael jordan",
    "lebron james",
    "kobe bryant",
    "lionel messi",
    "cristiano ronaldo",
    "shohei ohtani",
    "shoehi ohtani",
    "patrick mahomes",
    "joe montana",
    "jerry rice",
    "barry sanders",
    "peyton manning",
    "dan marino",
    "walter payton",
    "randy moss",
    "john elway",
    "emmitt smith",
    "deion sanders",
    "aaron donald",
    "babe ruth",
    "mickey mantle",
    "lou gehrig",
    "lou gherig",
    "hank aaron",
    "ken griffey jr",
    "willie mays",
    "sandy koufax",
    "nolan ryan",
    "randy johnson",
    "ichiro suzuki",
    "cal ripken jr",
    "roberto clemente",
    "jackie robinson",
    "aaron judge",
    "clayton kershaw",
    "mike trout",
    "mookie betts",
    "larry bird",
    "magic johnson",
    "kareem abdul jabbar",
    "kareem abdul-jabbar",
    "shaquille o'neal",
    "tim duncan",
    "wilt chamberlain",
    "jerry west",
    "kevin durant",
    "nikola jokic",
}


def initialize_player_sport_data() -> None:
    load_extension_player_sport_data()
    for player in PLAYER_SPORT_HINTS:
        PLAYER_DISPLAY_NAMES.setdefault(player, title_case_name(player))
    rebuild_partial_player_hints()
    SORTED_PLAYER_KEYS[:] = sorted(PLAYER_SPORT_HINTS, key=len, reverse=True)


def load_extension_player_sport_data() -> None:
    if not PLAYER_SPORT_DATA_PATH.exists():
        return
    try:
        raw = PLAYER_SPORT_DATA_PATH.read_text(encoding="utf-8")
        match = re.search(r"window\.AutoSheetReviewPlayerSports\s*=\s*(\{.*?\})\s*;\s*\}\)\(\);", raw, re.S)
        if not match:
            return
        payload = json.loads(match.group(1))
    except Exception:
        return
    players = payload.get("players") if isinstance(payload, dict) else {}
    if not isinstance(players, dict):
        return
    for player, value in players.items():
        key = normalize_player_key(player)
        if not key:
            continue
        if isinstance(value, str):
            sport = canonical_sport_label(value) or clean_rule_text(value)
            display_name = title_case_name(key)
            teams: list[str] = []
        elif isinstance(value, dict):
            sport = canonical_sport_label(value.get("sport") or "") or clean_rule_text(value.get("sport") or "")
            display_name = str(value.get("displayName") or value.get("display_name") or title_case_name(key)).strip()
            teams_value = value.get("teams") if isinstance(value.get("teams"), list) else [value.get("team")] if value.get("team") else []
            teams = [str(team or "").strip() for team in teams_value if str(team or "").strip()]
        else:
            continue
        if not sport:
            continue
        PLAYER_SPORT_HINTS[key] = sport
        PLAYER_DISPLAY_NAMES[key] = display_name
        if teams:
            PLAYER_TEAM_HINTS[key] = teams
        base_key = player_base_name(key)
        if base_key and base_key != key:
            PLAYER_SPORT_HINTS.setdefault(base_key, sport)
            PLAYER_DISPLAY_NAMES.setdefault(base_key, re.sub(r"\s+(?:jr|sr|ii|iii|iv|v)\.?$", "", display_name, flags=re.I).strip())
            if teams:
                PLAYER_TEAM_HINTS.setdefault(base_key, teams)


def load_player_sport_overrides(path: Path) -> int:
    try:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        return 0
    players = payload.get("players") if isinstance(payload, dict) else payload
    if not isinstance(players, dict):
        return 0
    loaded = 0
    for player, value in players.items():
        if isinstance(value, str):
            sport = value
            display_name = title_case_name(clean_rule_text(player))
        elif isinstance(value, dict):
            sport = value.get("sport") or value.get("category") or ""
            display_name = str(value.get("displayName") or value.get("display_name") or player).strip()
        else:
            continue
        if add_player_sport_hint(player, sport, display_name=display_name):
            loaded += 1
    if loaded:
        rebuild_partial_player_hints()
        SORTED_PLAYER_KEYS[:] = sorted(PLAYER_SPORT_HINTS, key=len, reverse=True)
        _find_known_player_sports_cached.cache_clear()
    return loaded


def add_player_sport_hint(player: Any, sport: Any, display_name: str = "") -> bool:
    key = normalize_player_key(player)
    category = canonical_sport_label(str(sport or "")) or clean_rule_text(sport)
    if not key or not category:
        return False
    PLAYER_SPORT_HINTS[key] = category
    PLAYER_DISPLAY_NAMES[key] = str(display_name or title_case_name(key)).strip()
    base_key = player_base_name(key)
    if base_key and base_key != key:
        PLAYER_SPORT_HINTS.setdefault(base_key, category)
        PLAYER_DISPLAY_NAMES.setdefault(base_key, re.sub(r"\s+(?:jr|sr|ii|iii|iv|v)\.?$", "", PLAYER_DISPLAY_NAMES[key], flags=re.I).strip())
    rebuild_partial_player_hints()
    SORTED_PLAYER_KEYS[:] = sorted(PLAYER_SPORT_HINTS, key=len, reverse=True)
    _find_known_player_sports_cached.cache_clear()
    return True


def rebuild_partial_player_hints() -> None:
    token_map: dict[str, list[dict[str, str]]] = {}
    for player, sport in PLAYER_SPORT_HINTS.items():
        parts = clean_rule_text(player).split()
        if not parts:
            continue
        tokens = [parts[-1], *[part for part in parts if is_distinctive_first_name(part)]]
        for token in tokens:
            if len(token) < 4 or is_ambiguous_partial_token(token) or token in PRODUCT_WORDS:
                continue
            token_map.setdefault(token, []).append({
                "key": player,
                "playerName": PLAYER_DISPLAY_NAMES.get(player) or title_case_name(player),
                "sport": sport,
            })
    PARTIAL_PLAYER_HINTS.clear()
    for token, hints in token_map.items():
        override_key = PARTIAL_PLAYER_TOKEN_OVERRIDES.get(token)
        override = next((hint for hint in hints if hint["key"] == override_key), None) if override_key else None
        if override:
            PARTIAL_PLAYER_HINTS[token] = override
            continue
        sports = {hint["sport"] for hint in hints}
        if len(sports) == 1:
            PARTIAL_PLAYER_HINTS[token] = sorted(hints, key=lambda hint: len(hint["playerName"]))[0]


def normalize_player_key(value: Any) -> str:
    return clean_rule_text(value)


def player_base_name(value: Any) -> str:
    return re.sub(r"\s+(?:jr|sr|ii|iii|iv|v)\.?$", "", clean_rule_text(value)).strip()


def is_distinctive_first_name(token: str) -> bool:
    return token in DISTINCTIVE_FIRST_NAMES


def is_ambiguous_partial_token(token: str) -> bool:
    return token in AMBIGUOUS_PARTIAL_TOKENS


def title_case_name(value: Any) -> str:
    return " ".join(word[:1].upper() + word[1:] for word in str(value or "").split())


@dataclass
class AssignmentRule:
    matcher: str = ""
    min_price: float | None = None
    max_price: float | None = None
    block: bool = False


@dataclass
class GradeRule:
    allowed: bool = True
    min_grade: float | None = None
    max_grade: float | None = None


@dataclass
class CompanyRules:
    include: list[str] = field(default_factory=list)
    exclude: list[str] = field(default_factory=list)
    ranges: list[AssignmentRule] = field(default_factory=list)
    blocks: list[AssignmentRule] = field(default_factory=list)
    grade_rules: dict[str, GradeRule] = field(default_factory=dict)
    rule_groups: list["CompanyRules"] = field(default_factory=list)
    goat_players: set[str] = field(default_factory=set)
    goat_ranges: list[AssignmentRule] = field(default_factory=list)
    accept_all: bool = False


@dataclass
class PayoutTier:
    min_price: float = 0
    max_price: float | None = None
    rate: float = 0
    matcher: str = ""


@dataclass
class AssignmentCompany:
    name: str
    rules: CompanyRules
    payout_tiers: list[PayoutTier]
    value_source: str = "comps"


@dataclass
class AssignmentRecommendation:
    company: str = ""
    payout: float | None = None
    source_value: float | None = None


@dataclass
class AssignmentDecision:
    company: str
    accepted: bool
    payout: float | None = None
    reason: str = ""
    source_value: float | None = None


class AssignmentEngine:
    def __init__(self, companies: list[AssignmentCompany] | None = None, error: str = "") -> None:
        self.companies = companies or []
        self.error = error

    @classmethod
    def load(cls, config_path: Path | None = None) -> "AssignmentEngine":
        path = Path(config_path or CONFIG_PATH)
        if not path.exists():
            return cls([])
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
            entries = raw.get("companies", raw) if isinstance(raw, dict) else raw
            companies: list[AssignmentCompany] = []
            errors: list[str] = []
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                name = str(entry.get("name") or "Unnamed company").strip()
                try:
                    company = load_company(entry, path.parent)
                except Exception as error:
                    errors.append(f"{name}: {error}")
                    continue
                if company is not None:
                    companies.append(company)
            return cls(companies, "; ".join(errors))
        except Exception as error:
            return cls([], str(error))

    def recommend(self, row: Any) -> AssignmentRecommendation:
        default_source_value = assignment_value(row)
        if default_source_value is None:
            return AssignmentRecommendation()

        candidates: list[AssignmentRecommendation] = []
        for decision in self.evaluate(row):
            if not decision.accepted or decision.payout is None:
                continue
            candidates.append(AssignmentRecommendation(decision.company, round(decision.payout, 2), decision.source_value))
        return max(candidates, key=lambda item: item.payout or 0) if candidates else AssignmentRecommendation(source_value=default_source_value)

    def evaluate(self, row: Any) -> list[AssignmentDecision]:
        grader = str(getattr(row, "grader", "") or "")
        decisions: list[AssignmentDecision] = []
        for company in self.companies:
            source_value = company_assignment_value(row, company)
            if source_value is None:
                decisions.append(AssignmentDecision(company.name, False, reason="missing comp/card ladder value"))
                continue
            card_text = card_row_text(row, source_value)
            if not company_accepts(company.rules, card_text, source_value, grader):
                decisions.append(AssignmentDecision(company.name, False, reason="card does not match company rules", source_value=source_value))
                continue
            payout = payout_for_value(company.payout_tiers, source_value, card_text, company.rules)
            if payout is None:
                decisions.append(AssignmentDecision(company.name, True, None, "accepted, but no payout tier matched", source_value))
                continue
            decisions.append(AssignmentDecision(company.name, True, payout, "accepted and payout tier matched", source_value))
        return decisions


def load_company(entry: dict[str, Any], base_dir: Path) -> AssignmentCompany | None:
    name = str(entry.get("name") or "").strip()
    if not name or entry.get("active") is False:
        return None
    rules_text = read_source_text(entry.get("rules") or entry.get("rules_source") or entry.get("rulesSource"), base_dir)
    payout_text = read_source_text(entry.get("payout") or entry.get("payout_source") or entry.get("payoutSource"), base_dir)
    rules = parse_rules(rules_text, accept_all=bool(entry.get("accept_all")))
    payout_tiers = parse_payouts(payout_text)
    if not payout_tiers and entry.get("rate") is not None:
        rate = parse_rate(entry.get("rate"))
        if rate is not None:
            payout_tiers = [PayoutTier(rate=rate)]
    return AssignmentCompany(
        name=name,
        rules=rules,
        payout_tiers=payout_tiers,
        value_source=company_value_source(entry),
    )


def assignment_value(row: Any) -> float | None:
    comps = to_number(getattr(row, "card_ladder_comps_average", None))
    cl_value = to_number(getattr(row, "card_ladder_value", None))
    return comps if comps is not None else cl_value


def company_assignment_value(row: Any, company: AssignmentCompany) -> float | None:
    comps = to_number(getattr(row, "card_ladder_comps_average", None))
    cl_value = to_number(getattr(row, "card_ladder_value", None))
    if company.value_source == "card_ladder":
        return cl_value
    return comps if comps is not None else cl_value


def company_value_source(entry: dict[str, Any]) -> str:
    raw = str(
        entry.get("value_source")
        or entry.get("valueSource")
        or entry.get("assignment_value_source")
        or entry.get("assignmentValueSource")
        or ""
    ).strip().lower()
    if raw in {"card_ladder", "cardladder", "cl", "card ladder", "card_ladder_value"}:
        return "card_ladder"
    return "comps"


def card_row_text(row: Any, source_value: float) -> str:
    parts = [
        getattr(row, "cert_number", ""),
        getattr(row, "grader", ""),
        getattr(row, "card_title", ""),
        f"${source_value}",
    ]
    return " ".join(str(part or "") for part in parts).strip()


def company_accepts(rules: CompanyRules, text: str, price: float, grader: str) -> bool:
    if not rules.accept_all and not (rules.include or rules.exclude or rules.ranges or rules.blocks or rules.grade_rules or rules.rule_groups):
        return False
    haystack = clean_text(text)
    grade_company, grade = parse_grade(text, grader)

    for rule in rules.blocks:
        if rule_matches(rule, haystack, price):
            return False
    if any(term_matches(term, haystack) for term in rules.exclude):
        return False
    if rules.rule_groups:
        return any(company_accepts(group, text, price, grader) for group in rules.rule_groups)

    if rules.grade_rules:
        grade_rule = rules.grade_rules.get(clean_text(grade_company))
        if grade_rule is None or not grade_rule.allowed:
            return False
        if grade is not None and grade_rule.min_grade is not None and grade < grade_rule.min_grade:
            return False
        if grade is not None and grade_rule.max_grade is not None and grade > grade_rule.max_grade:
            return False

    if rules.include and not any(term_matches(term, haystack) for term in rules.include):
        return False
    if rules.ranges:
        return any(rule_matches(rule, haystack, price) for rule in rules.ranges)
    return True


def payout_for_value(tiers: list[PayoutTier], value: float, text: str = "", rules: CompanyRules | None = None) -> float | None:
    haystack = clean_text(text)
    payouts: list[float] = []
    for tier in tiers:
        if value < tier.min_price:
            continue
        if tier.max_price is not None and value > tier.max_price:
            continue
        if tier.matcher and not payout_category_matches(tier.matcher, haystack, rules, value):
            continue
        payouts.append(value * tier.rate)
    return max(payouts) if payouts else None


def read_source_text(source: Any, base_dir: Path, interactive_google: bool = False) -> str:
    if isinstance(source, dict):
        return read_structured_source_text(source, base_dir, interactive_google=interactive_google)
    raw = normalize_source_value(source)
    if not raw:
        return ""
    if raw.startswith(("http://", "https://")):
        return read_url_text(raw)
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = base_dir / path
    try:
        if not path.exists():
            return ""
    except OSError as error:
        raise ValueError(f"Invalid local source path: {raw}") from error
    if path.suffix.lower() == ".gsheet":
        exported = materialize_gsheet_shortcut(path, path.parent / "LUCAS SHEET EXPORTS")
        return read_workbook_text(exported)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_workbook_text(path)
    try:
        return path.read_text(encoding="utf-8-sig")
    except OSError as error:
        raise ValueError(f"Could not open local source path: {raw}") from error


def read_structured_source_text(source: dict[str, Any], base_dir: Path, interactive_google: bool = False) -> str:
    kind = str(source.get("kind") or "").strip()
    path_value = source.get("path") or source.get("file")
    url = str(source.get("url") or "").strip()
    sheet_name = str(source.get("sheet_name") or source.get("sheet") or "").strip()
    if kind == "google_sheet" and url:
        path = path_from_source_value(path_value, base_dir) if path_value else None
        try:
            sheets = read_google_sheet_tabs(url, interactive=interactive_google, sheet_name=sheet_name)
            if sheet_name:
                return workbook_values_text(sheets, sheet_name=sheet_name)
            synthesized = synthesize_workbook_rules(sheets)
            if synthesized:
                return "\n".join(synthesized)
            return workbook_values_text(sheets)
        except GoogleSheetsAuthError:
            if not path or not path.exists():
                raise
        except Exception:
            if not path or not path.exists():
                raise
        try:
            if path:
                materialize_google_sheet_url_to_path(url, path)
                return read_workbook_text(path, sheet_name=sheet_name)
            output_dir = base_dir / "ASSIGNMENT RULES" / "SHEET EXPORTS"
            exported = materialize_google_sheet_url(url, output_dir, str(source.get("name") or "google-sheet"), unique=False)
            return read_workbook_text(exported, sheet_name=sheet_name)
        except Exception:
            if path and path.exists():
                return read_workbook_text(path, sheet_name=sheet_name)
            raise
    if sheet_name and path_value:
        path = path_from_source_value(path_value, base_dir)
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            return read_workbook_text(path, sheet_name=sheet_name)
    return read_source_text(path_value or url or source.get("doc_id"), base_dir, interactive_google=interactive_google)


def path_from_source_value(value: Any, base_dir: Path) -> Path:
    raw = normalize_source_value(value)
    path = Path(raw).expanduser()
    return path if path.is_absolute() else base_dir / path


def normalize_source_value(source: Any) -> str:
    raw = str(source or "").strip().strip('"').strip("'")
    if raw.startswith("file://"):
        parsed = urllib.parse.urlparse(raw)
        raw = urllib.parse.unquote(parsed.path or "")
        if os.name == "nt" and re.match(r"^/[A-Za-z]:/", raw):
            raw = raw[1:]
    if os.name == "nt" and re.match(r"^/[A-Za-z]:[\\/]", raw):
        raw = raw[1:]
    return raw


def load_gsheet_shortcut(path: Path) -> dict[str, Any]:
    try:
        raw = path.read_bytes()
    except OSError as error:
        raise ValueError(
            f"Could not open this .gsheet shortcut locally: {path}. Google Drive may be exposing it as an unreadable placeholder."
        ) from error
    try:
        shortcut = json.loads(raw.decode("utf-8-sig"))
    except UnicodeDecodeError as error:
        raise ValueError(
            f"This .gsheet shortcut is not UTF-8 JSON: {path}. Paste the Google Sheet URL so L.U.C.A.S can export it."
        ) from error
    except json.JSONDecodeError as error:
        raise ValueError(
            f"This .gsheet file is not readable shortcut JSON: {path}. Paste the Google Sheet URL so L.U.C.A.S can export it."
        ) from error
    return shortcut


def read_gsheet_shortcut_text(path: Path) -> str:
    shortcut = load_gsheet_shortcut(path)
    url = gsheet_shortcut_url(shortcut)
    if not url:
        raise ValueError(
            "This .gsheet shortcut does not contain readable sheet data. Choose a synced/exported .xlsx or .csv copy from Google Drive."
        )
    text = read_url_text(url)
    if not text.strip():
        raise ValueError(
            "Google returned no CSV rows for this .gsheet shortcut. Export or sync the sheet as .xlsx/.csv and choose that file."
        )
    return text


def materialize_gsheet_shortcut(path: Path, output_dir: Path) -> Path:
    shortcut = load_gsheet_shortcut(path)
    url = gsheet_shortcut_url(shortcut)
    if not url:
        raise ValueError("This .gsheet shortcut does not contain a Google Sheet URL or document id.")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{safe_filename(str(shortcut.get('name') or path.stem or 'google-sheet'))}.xlsx"
    try:
        return export_google_sheet_to_xlsx(url, output_path, interactive=False)
    except Exception:
        return materialize_google_sheet_url(url, output_dir, str(shortcut.get("name") or path.stem or "google-sheet"), unique=False)


def materialize_google_sheet_url(url: str, output_dir: Path, name: str = "google-sheet", unique: bool = True) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = safe_filename(name)
    output_path = output_dir / f"{stem}.xlsx"
    if unique:
        output_path = unique_export_path(output_path)
    materialize_google_sheet_url_to_path(url, output_path)
    return output_path


def materialize_google_sheet_url_to_path(url: str, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    download_google_sheet_xlsx(url, output_path)


def download_google_sheet_xlsx(url: str, output_path: Path) -> None:
    request_url = google_sheet_xlsx_url(url)
    try:
        with urllib.request.urlopen(request_url, timeout=40) as response:
            data = response.read()
    except urllib.error.HTTPError as error:
        if error.code in {401, 403}:
            raise ValueError(
                "Google rejected the sheet export because this sheet is private. Share/export access is required, or L.U.C.A.S needs an authenticated Google import connection."
            ) from error
        raise
    if not data.startswith(b"PK"):
        text = data[:500].decode("utf-8", errors="replace")
        if re.search(r"<!doctype html|<html[\s>]", text, re.I):
            raise ValueError(
                "Google returned a web page instead of an XLSX export. Open the sheet in Google Drive and save/download it as .xlsx or .csv."
            )
        raise ValueError("Google did not return a valid XLSX export for this sheet.")
    output_path.write_bytes(data)


def google_sheet_xlsx_url(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    if "docs.google.com" not in parsed.netloc or "/spreadsheets/" not in parsed.path:
        return url
    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if not match:
        return url
    return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"


def unique_export_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = path.with_name(f"{path.stem}-{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    return path


def safe_filename(value: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9._ -]+", "-", value).strip(" .-")
    return safe or "google-sheet"


def gsheet_shortcut_url(shortcut: dict[str, Any]) -> str:
    url = str(shortcut.get("url") or "").strip()
    if url:
        return url
    doc_id = str(shortcut.get("doc_id") or "").strip()
    resource_id = str(shortcut.get("resource_id") or "").strip()
    if not doc_id and resource_id.startswith("spreadsheet:"):
        doc_id = resource_id.split(":", 1)[1].strip()
    if doc_id:
        return f"https://docs.google.com/spreadsheets/d/{doc_id}/edit"
    return ""


def read_url_text(url: str) -> str:
    request_url = google_sheet_csv_url(url) or url
    with urllib.request.urlopen(request_url, timeout=20) as response:
        data = response.read()
    text = data.decode("utf-8-sig", errors="replace")
    if re.search(r"^\s*<!doctype html|<html[\s>]", text, re.I):
        raise ValueError(
            "Google returned a web page instead of sheet rows. Choose a synced/exported .xlsx or .csv file from your Drive folder."
        )
    return text


def google_sheet_csv_url(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    if "docs.google.com" not in parsed.netloc or "/spreadsheets/" not in parsed.path:
        return ""
    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if not match:
        return ""
    query = urllib.parse.parse_qs(parsed.query)
    gid = query.get("gid", ["0"])[0]
    return f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=csv&gid={gid}"


def read_workbook_text(path: Path, sheet_name: str = "") -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected_sheets = [
            sheet for sheet in workbook.worksheets
            if not sheet_name or sheet.title.lower() == sheet_name.lower()
        ]
        if sheet_name and not selected_sheets:
            raise ValueError(f"Workbook does not contain a sheet named {sheet_name}.")
        sheets = [(sheet.title, worksheet_values(sheet)) for sheet in selected_sheets]
        if not sheet_name:
            synthesized = synthesize_workbook_rules(sheets)
            if synthesized:
                return "\n".join(synthesized)
        return workbook_values_text(sheets, sheet_name=sheet_name)
    finally:
        workbook.close()


def worksheet_values(sheet: Any) -> list[list[Any]]:
    values: list[list[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        cells = list(row)
        while cells and cells[-1] in (None, ""):
            cells.pop()
        values.append(cells)
    return values


def workbook_values_text(sheets: list[tuple[str, list[list[Any]]]], sheet_name: str = "") -> str:
    lines: list[str] = []
    for title, values in sheets:
        lines.append(f"# {title}")
        for row in values:
            cells = [str(cell).strip() for cell in row if cell not in (None, "")]
            if cells:
                lines.append(" ".join(cells))
    if sheet_name and not lines:
        raise ValueError(f"Workbook does not contain a sheet named {sheet_name}.")
    return "\n".join(lines)


def synthesize_workbook_rules(sheets: list[tuple[str, list[list[Any]]]]) -> list[str]:
    goat_players: list[str] = []
    for title, values in sheets:
        if re.search(r"goats?", title, re.I):
            goat_players.extend(extract_player_names_from_values(values))
    context = {"goatPlayers": unique_values(goat_players)}

    rules: list[str] = []
    rules.extend(f"goat-player: {player}" for player in context["goatPlayers"])
    for title, values in sheets:
        cleaned_title = clean_rule_label(title)
        if re.search(r"^(comping standards|payouts)$", cleaned_title, re.I):
            continue
        if re.search(r"do not buy|never buy", cleaned_title, re.I):
            rules.extend(synthesize_do_not_buy_rules(values))
            continue
        sheet_rules = synthesize_arena_club_rules(values, context)
        if not sheet_rules:
            sheet_rules = synthesize_generic_sheet_rules(title, values)
        if sheet_rules:
            rules.extend(sheet_rules)
        else:
            rules.extend(synthesize_generic_sheet_rules(title, values))
    return unique_values(rules)


def synthesize_arena_club_rules(values: list[list[Any]], context: dict[str, Any]) -> list[str]:
    rules: list[str] = []
    rules.extend(synthesize_arena_club_category_table_rules(values, context))

    header_row_index = next(
        (
            index for index, row in enumerate(values)
            if any(re.search(r"brady|kobe|lebron|kaboom|downtown|goats?|color blast|manga", str(cell or ""), re.I) for cell in row)
        ),
        -1,
    )
    if header_row_index >= 0:
        headers = values[header_row_index]
        range_row = next((row for row in values[header_row_index + 1:] if any(parse_sheet_range(cell) for cell in row)), [])
        for index, header in enumerate(headers):
            range_value = range_row[index] if index < len(range_row) else None
            parsed_range = parse_sheet_range(range_value)
            if not parsed_range:
                continue
            for label in expand_header_label(header):
                if is_goat_rule_label(label) and context.get("goatPlayers"):
                    for player in context["goatPlayers"]:
                        rules.append(format_rule(player, parsed_range))
                        rules.append(f"goat-range: {format_rule(player, parsed_range)}")
                else:
                    rules.append(format_rule(label, parsed_range))

    rules.extend(synthesize_arena_club_sport_range_rules(values))
    return unique_values(rules)


def synthesize_arena_club_category_table_rules(values: list[list[Any]], context: dict[str, Any]) -> list[str]:
    rules: list[str] = []
    for row in values:
        label = normalize_rule_label(row[0] if row else "")
        parsed_range = parse_sheet_range(row[1] if len(row) > 1 else None)
        if not label or not parsed_range or is_generic_rule_label(label) or parse_sheet_range(label):
            continue
        for expanded_label in expand_header_label(label):
            if is_goat_rule_label(expanded_label) and context.get("goatPlayers"):
                for player in context["goatPlayers"]:
                    rules.append(format_rule(player, parsed_range))
                    rules.append(f"goat-range: {format_rule(player, parsed_range)}")
            else:
                rules.append(format_rule(expanded_label, parsed_range))
    return rules


def synthesize_arena_club_sport_range_rules(values: list[list[Any]]) -> list[str]:
    rules: list[str] = []
    for row_index, row in enumerate(values):
        for column_index, cell in enumerate(row):
            if not re.match(r"^price ranges?$", str(cell or "").strip(), re.I):
                continue
            parsed_range = parse_sheet_range(row[column_index + 1] if column_index + 1 < len(row) else None)
            sport = find_nearest_sport_label(values, row_index, column_index)
            if parsed_range and sport:
                rules.append(format_rule(sport, parsed_range))
    return rules


def synthesize_generic_sheet_rules(title: str, values: list[list[Any]]) -> list[str]:
    rules: list[str] = []
    title_hint = clean_rule_text(title)
    for row_index, row in enumerate(values):
        for column_index, cell in enumerate(row):
            parsed_range = parse_sheet_range(cell)
            if not parsed_range:
                continue
            label = find_rule_label(values, row_index, column_index, title_hint)
            if label:
                rules.append(format_rule(label, parsed_range))
    return unique_values(rules)


def synthesize_do_not_buy_rules(values: list[list[Any]]) -> list[str]:
    rules: list[str] = []
    max_columns = max((len(row) for row in values), default=0)
    section_max_by_column: list[float | None] = [None] * max_columns
    for row in values:
        apply_do_not_buy_section_headings(row, section_max_by_column)
        for column_index, cell in enumerate(row):
            text = re.sub(r"^\d+[.)]\s*", "", clean_rule_label(cell)).strip()
            if not text or parse_do_not_buy_section(text):
                continue
            over_match = (
                re.match(r"^(.+?)\s+(?:cards\s+)?over\s+\$?([\d,]+(?:\.\d+)?k?)\+?(?:\s+value)?", text, re.I)
                or re.match(r"^(.+?)\s+\$?([\d,]+(?:\.\d+)?k?)\+$", text, re.I)
            )
            if over_match:
                rules.append(f"block: {over_match.group(1).strip()} over {format_rule_number(parse_money(over_match.group(2)) or 0)}")
                continue
            section_max = section_max_by_column[column_index] if column_index < len(section_max_by_column) else None
            rules.append(f"block: {text} over {format_rule_number(section_max)}" if section_max is not None else f"block: {text}")
    return unique_values(rules)


def apply_do_not_buy_section_headings(row: list[Any], section_max_by_column: list[float | None]) -> None:
    headings = [
        (index, section)
        for index, cell in enumerate(row)
        for section in [parse_do_not_buy_section(clean_rule_label(cell))]
        if section and not section.get("keepCurrentPrice")
    ]
    for heading_index, (index, section) in enumerate(headings):
        next_index = headings[heading_index + 1][0] if heading_index + 1 < len(headings) else len(section_max_by_column)
        for column in range(index, next_index):
            section_max_by_column[column] = section.get("maxPrice")


def parse_do_not_buy_section(value: Any) -> dict[str, Any] | None:
    text = str(value or "").strip()
    threshold_match = re.search(r"(?:do\s+not|don't|dont|not|never|avoid).{0,40}?\bover\s+\$?([\d,]+(?:\.\d+)?k?)", text, re.I)
    if threshold_match:
        return {"maxPrice": parse_money(threshold_match.group(1))}
    if re.match(r"^(?:basketball|football|baseball|soccer|hockey|wnba|collegiate|vintage|notes?)$", text, re.I):
        return {"keepCurrentPrice": True}
    if re.match(r"^(?:do not buy|don't buy|dont buy|never buy|players to never buy|players to avoid|players to not buy|players not to buy|basketball|football|baseball|soccer|hockey|wnba|collegiate|vintage|currently avoiding(?: buying)?|pausing/limiting|notes?)$", text, re.I):
        return {"maxPrice": None}
    return None


def extract_player_names_from_values(values: list[list[Any]]) -> list[str]:
    players: list[str] = []
    for row in values:
        for cell in row:
            text = re.sub(r"^\d+\.?\s*", "", clean_rule_label(cell))
            text = re.sub(r"\s+\$?\d[\d,]*(?:\.\d+)?k?\s*(?:-|–|—|to|through|thru)\s*\$?\d[\d,]*(?:\.\d+)?k?.*$", "", text, flags=re.I).strip()
            if is_likely_player_name(text):
                players.append(text)
    return unique_values(players)


def is_likely_player_name(value: Any) -> bool:
    text = str(value or "").strip()
    if not text or len(text) > 48:
        return False
    if re.search(r"price|range|grade|sport|dupes|qty|conf|goats?|tab|notes?", text, re.I):
        return False
    return re.match(r"^[A-Za-z][A-Za-z'. -]+(?:\s+[A-Za-z'. -]+)+$", text) is not None


def expand_header_label(value: Any) -> list[str]:
    label = normalize_rule_label(value)
    if not label:
        return []
    if re.search(r"tom brady.*kobe bryant|kobe bryant.*tom brady", label, re.I):
        return ["Tom Brady", "Kobe Bryant"]
    return [label]


def normalize_rule_label(value: Any) -> str:
    label = clean_rule_label(value)
    label = re.sub(r"&", " ", label)
    label = re.sub(r"\bBRADY\b", "Tom Brady", label, flags=re.I)
    label = re.sub(r"\bKOBE\b", "Kobe Bryant", label, flags=re.I)
    label = re.sub(r"\bKabooms\b", "Kaboom", label, flags=re.I)
    label = re.sub(r"\bGOATS\b", "GOAT", label, flags=re.I)
    return label.strip()


def is_goat_rule_label(value: Any) -> bool:
    return re.search(r"\bgoats?\b", str(value or ""), re.I) is not None


def find_nearest_sport_label(values: list[list[Any]], row_index: int, column_index: int) -> str:
    for index in range(row_index - 1, -1, -1):
        row = values[index]
        candidate = normalize_sport_label(row[column_index] if column_index < len(row) else "")
        if candidate:
            return candidate
    return ""


def normalize_sport_label(value: Any) -> str:
    label = clean_rule_label(value).lower()
    sports = {
        "basketball": "Basketball",
        "baseball": "Baseball",
        "football": "Football",
        "soccer": "Soccer",
        "wwe": "WWE",
        "wrestling": "WWE",
        "wwf": "WWE",
        "f1": "F1",
        "formula 1": "F1",
        "formula one": "F1",
        "formula1": "F1",
        "marvel": "Marvel",
        "disney": "Disney",
        "star wars": "Star Wars",
        "starwars": "Star Wars",
        "ufc": "UFC",
        "mma": "UFC",
        "hockey": "Hockey",
        "pokemon": "Pokemon",
        "poke": "Pokemon",
    }
    return sports.get(label, "")


def find_rule_label(values: list[list[Any]], row_index: int, column_index: int, title_hint: str) -> str:
    row = values[row_index] if row_index < len(values) else []
    left_values = [clean_rule_label(value) for value in reversed(row[:column_index]) if clean_rule_label(value)]
    for left_label in left_values:
        if not re.search(r"price ranges?|range|dupes|qty|conf", left_label, re.I):
            return normalize_rule_label(left_label)
    for index in range(row_index - 1, -1, -1):
        row_above = values[index]
        above = clean_rule_label(row_above[column_index] if column_index < len(row_above) else "")
        if above and not parse_sheet_range(above):
            return normalize_rule_label(above)
    return normalize_rule_label(title_hint)


def parse_sheet_range(value: Any) -> dict[str, float] | None:
    match = re.search(r"\$?\s*(\d[\d,]*(?:\.\d+)?k?)\s*(?:-|–|—|to|through|thru)\s*\$?\s*(\d[\d,]*(?:\.\d+)?k?)", str(value or ""), re.I)
    if not match:
        return None
    return {"min": parse_money(match.group(1)) or 0, "max": parse_money(match.group(2)) or 0}


def format_rule(label: str, parsed_range: dict[str, float]) -> str:
    return f"{label} ${format_rule_number(parsed_range['min'])}-${format_rule_number(parsed_range['max'])}"


def format_rule_number(value: float | None) -> str:
    if value is None:
        return ""
    return str(int(value)) if float(value).is_integer() else str(value)


def clean_rule_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def unique_values(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def parse_rules(text: str, accept_all: bool = False) -> CompanyRules:
    stripped = str(text or "").strip()
    if not stripped:
        return CompanyRules(accept_all=accept_all)
    try:
        payload = json.loads(stripped)
        if isinstance(payload, dict):
            return parse_rule_dict(payload, accept_all)
    except Exception:
        pass

    rules = CompanyRules(accept_all=accept_all)
    for line in source_lines(stripped):
        lowered = line.lower()
        key_value = re.match(r"^([^:=]+)\s*[:=]\s*(.+)$", line)
        if key_value:
            key = normalize_key(key_value.group(1))
            value = key_value.group(2).strip()
            if key in {"include", "includekeywords", "keywords", "sports", "sport"}:
                rules.include.extend(split_values(value))
                continue
            if key in {"exclude", "excludekeywords"}:
                rules.exclude.extend(split_values(value))
                continue
            if key in {"block", "blocks", "blockrules", "donotbuy", "neverbuy"}:
                rules.blocks.append(parse_rule_line(value, block=True))
                continue
            if key in {"goatplayer", "goatplayers"}:
                rules.goat_players.update(clean_rule_text(player) for player in split_values(value))
                continue
            if key in {"goatrange", "goatranges"}:
                parsed_goat_range = parse_rule_line(value)
                if parsed_goat_range.min_price is not None or parsed_goat_range.max_price is not None:
                    rules.goat_ranges.append(parsed_goat_range)
                continue
            if key in {"minprice", "minimumprice"}:
                rules.ranges.append(AssignmentRule(min_price=parse_money(value)))
                continue
            if key in {"maxprice", "maximumprice"}:
                rules.ranges.append(AssignmentRule(max_price=parse_money(value)))
                continue
        if "block:" in lowered:
            rules.blocks.append(parse_rule_line(line.split(":", 1)[1], block=True))
            continue
        parsed = parse_rule_line(line)
        if parsed.min_price is not None or parsed.max_price is not None:
            rules.ranges.append(parsed)
    return rules


def parse_rule_dict(payload: dict[str, Any], accept_all: bool = False) -> CompanyRules:
    rules = CompanyRules(accept_all=accept_all or bool(payload.get("accept_all")))
    rules.include.extend(split_values(payload.get("include") or payload.get("includeKeywords") or payload.get("sports") or payload.get("sport")))
    rules.exclude.extend(split_values(payload.get("exclude") or payload.get("excludeKeywords")))
    rules.goat_players.update(clean_rule_text(player) for player in split_values(payload.get("goatPlayers") or payload.get("goat_players")))
    for item in payload.get("goatRanges") or payload.get("goat_ranges") or []:
        if isinstance(item, dict):
            rules.goat_ranges.append(AssignmentRule(
                matcher=str(item.get("matcher") or item.get("player") or "").strip(),
                min_price=to_number(item.get("min") or item.get("minPrice")),
                max_price=to_number(item.get("max") or item.get("maxPrice")),
            ))
    for block in split_values(payload.get("blocks") or payload.get("blockRules")):
        rules.blocks.append(parse_rule_line(block, block=True))
    for item in payload.get("ranges") or payload.get("rangeRules") or []:
        if isinstance(item, dict):
            rules.ranges.append(AssignmentRule(
                matcher=str(item.get("matcher") or item.get("sport") or "").strip(),
                min_price=to_number(item.get("min") or item.get("minPrice")),
                max_price=to_number(item.get("max") or item.get("maxPrice")),
            ))
    grades = payload.get("grades") or {}
    if isinstance(grades, dict):
        for company, grade_payload in grades.items():
            if isinstance(grade_payload, dict):
                rules.grade_rules[clean_text(company)] = GradeRule(
                    allowed=grade_payload.get("allowed") is not False,
                    min_grade=to_number(grade_payload.get("min")),
                    max_grade=to_number(grade_payload.get("max")),
                )
    custom_rules = payload.get("rules") or payload.get("customRules") or []
    if isinstance(custom_rules, list):
        rules.rule_groups.extend(parse_custom_rule_group(item) for item in custom_rules if isinstance(item, dict))
        rules.rule_groups = [group for group in rules.rule_groups if group.include or group.ranges or group.grade_rules or group.accept_all]
    return rules


def parse_custom_rule_group(payload: dict[str, Any]) -> CompanyRules:
    group = CompanyRules()
    sports = split_values(payload.get("sports") or payload.get("sport"))
    if payload.get("sportOther"):
        sports.append(str(payload.get("sportOther")).strip())
    group.include.extend(sport for sport in sports if sport and sport != "custom")
    price_ranges = payload.get("priceRanges") or payload.get("ranges") or []
    if isinstance(price_ranges, list):
        for price_range in price_ranges:
            if not isinstance(price_range, dict):
                continue
            min_price = to_number(price_range.get("min") or price_range.get("minPrice"))
            max_price = to_number(price_range.get("max") or price_range.get("maxPrice"))
            if min_price is None and max_price is None:
                continue
            group.ranges.append(AssignmentRule(min_price=min_price, max_price=max_price))
    grades = payload.get("grades") or {}
    if isinstance(grades, dict):
        for company, grade_payload in grades.items():
            if isinstance(grade_payload, dict):
                group.grade_rules[clean_text(company)] = GradeRule(
                    allowed=grade_payload.get("allowed") is not False,
                    min_grade=to_number(grade_payload.get("min")),
                    max_grade=to_number(grade_payload.get("max")),
                )
    return group


def parse_rule_line(line: str, block: bool = False) -> AssignmentRule:
    text = str(line or "").strip()
    over_match = re.match(r"(.+?)\s+(?:over|above)\s+\$?\s*([\d,.]+k?)\+?$", text, re.I)
    if over_match:
        return AssignmentRule(matcher=rule_matcher_label(over_match.group(1)), min_price=parse_money(over_match.group(2)), block=block)
    range_match = re.search(r"\$?\s*([\d,.]+k?)\s*(?:-|to|through|thru|–|—)\s*\$?\s*([\d,.]+k?)", text, re.I)
    if range_match:
        matcher = f"{text[:range_match.start()]} {text[range_match.end():]}".strip(" -:|")
        return AssignmentRule(matcher=rule_matcher_label(matcher), min_price=parse_money(range_match.group(1)), max_price=parse_money(range_match.group(2)), block=block)
    return AssignmentRule(matcher=text, block=block)


def rule_matcher_label(value: Any) -> str:
    label = str(value or "").strip()
    if is_generic_rule_label(label):
        return ""
    return label


def is_generic_rule_label(value: Any) -> bool:
    text = clean_rule_label(value).lower()
    return text in {
        "price",
        "prices",
        "price range",
        "price ranges",
        "range",
        "ranges",
        "value",
        "value range",
        "value ranges",
        "comp",
        "comps",
        "comp range",
        "comp ranges",
    }


def parse_payouts(text: str) -> list[PayoutTier]:
    stripped = str(text or "").strip()
    if not stripped:
        return []
    try:
        payload = json.loads(stripped)
        return parse_payout_json(payload)
    except Exception:
        pass

    tiers: list[PayoutTier] = []
    for line in source_lines(stripped):
        table_tier = parse_payout_table_line(line)
        if table_tier:
            tiers.append(table_tier)
            continue
        rate = parse_rate(line)
        if rate is None:
            continue
        range_match = re.search(r"\$?\s*([\d,.]+k?)\s*(?:-|to|through|thru|–|—)\s*\$?\s*([\d,.]+k?)", line, re.I)
        if range_match:
            tiers.append(PayoutTier(parse_money(range_match.group(1)) or 0, parse_money(range_match.group(2)), rate))
            continue
        min_match = re.search(r"(?:over|above|min(?:imum)?)\s+\$?\s*([\d,.]+k?)", line, re.I)
        if min_match:
            tiers.append(PayoutTier(parse_money(min_match.group(1)) or 0, None, rate))
            continue
        column_numbers = payout_row_numbers(line)
        if len(column_numbers) >= 2:
            tiers.append(PayoutTier(column_numbers[0], column_numbers[1], rate))
            continue
        if len(column_numbers) == 1:
            tiers.append(PayoutTier(column_numbers[0], None, rate))
            continue
        tiers.append(PayoutTier(0, None, rate))
    return sorted(tiers, key=lambda tier: tier.min_price, reverse=True)


def parse_payout_json(payload: Any) -> list[PayoutTier]:
    if isinstance(payload, dict) and payload.get("rate") is not None:
        rate = parse_rate(payload.get("rate"))
        return [PayoutTier(rate=rate)] if rate is not None else []
    tiers_payload = payload.get("tiers") if isinstance(payload, dict) else payload
    tiers: list[PayoutTier] = []
    for item in tiers_payload or []:
        if not isinstance(item, dict):
            continue
        rate = parse_rate(item.get("rate") or item.get("payout"))
        if rate is None:
            continue
        tiers.append(PayoutTier(
            min_price=to_number(item.get("min") or item.get("minPrice")) or 0,
            max_price=to_number(item.get("max") or item.get("maxPrice")),
            rate=rate,
            matcher=str(item.get("matcher") or item.get("category") or item.get("sport") or "").strip(),
        ))
    return sorted(tiers, key=lambda tier: tier.min_price, reverse=True)


def parse_payout_table_line(line: str) -> PayoutTier | None:
    text = str(line or "").strip()
    if not text or re.search(r"\bcategory\b|\bvalue range\b|\bpayout\b", text, re.I):
        return None
    range_match = re.search(r"\$?\s*([\d,.]+k?)\s*(?:-|to|through|thru|â€“|â€”)\s*\$?\s*([\d,.]+k?)", text, re.I)
    if not range_match:
        range_match = re.search(r"\$?\s*([\d,.]+k?)\s*(?:-|\u2013|\u2014|to|through|thru)\s*\$?\s*([\d,.]+k?)", text, re.I)
    if not range_match:
        return None
    before = text[:range_match.start()].strip(" -:|")
    after = text[range_match.end():].strip()
    rate = parse_payout_table_rate(after)
    if rate is None:
        return None
    return PayoutTier(
        min_price=parse_money(range_match.group(1)) or 0,
        max_price=parse_money(range_match.group(2)),
        rate=rate,
        matcher=normalize_payout_category(before),
    )


def parse_payout_table_rate(value: Any) -> float | None:
    text = str(value or "").strip()
    percent_match = re.search(r"(\d+(?:\.\d+)?)\s*%", text)
    if percent_match:
        return float(percent_match.group(1)) / 100
    number_match = re.search(r"\b(\d+(?:\.\d+)?)\b", text)
    if not number_match:
        return None
    number = float(number_match.group(1))
    return number / 100 if number > 2 else number


def payout_row_numbers(line: str) -> list[float]:
    text = re.sub(r"\d+(?:\.\d+)?\s*%", " ", str(line or ""))
    numbers: list[float] = []
    for match in re.finditer(r"\$?\s*([\d,.]+k?)", text, re.I):
        number = parse_money(match.group(1))
        if number is not None:
            numbers.append(number)
    return numbers


def normalize_payout_category(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    text = re.sub(r"^\*+", "", text).strip()
    replacements = {
        "kabooms": "Kaboom",
        "downtowns": "Downtown",
        "mangas": "Manga",
        "color blasts": "Color Blast",
    }
    return replacements.get(text.lower(), text)


def payout_category_matches(category: str, haystack: str, rules: CompanyRules | None = None, value: float | None = None) -> bool:
    expected = clean_rule_text(category)
    if not expected:
        return True
    parsed_value = parse_card_for_matching(haystack)
    if is_goat_payout_category(category):
        goat_players = (rules.goat_players if rules and rules.goat_players else set()) or GOAT_PAYOUT_PLAYERS
        if not goat_players:
            return False
        return any(
            player_matches_value(player, parsed_value)
            for player in goat_players
        )
    return (
        matcher_matches_text(category, haystack)
        or player_matches_value(category, parsed_value)
        or sport_label_matches_value(category, parsed_value, haystack)
    )


def is_goat_payout_category(category: Any) -> bool:
    text = clean_rule_text(category)
    return bool(re.search(r"\bgoats?\b", text))


def parse_card_for_matching(text: str) -> dict[str, Any]:
    raw = str(text or "")
    player_matches = find_known_player_sports(raw)
    player = player_matches[0]["playerName"] if player_matches else ""
    sport = infer_sport(raw, player)
    team_matches = find_known_player_teams(raw, player_matches)
    return {
        "text": raw,
        "playerName": player,
        "sport": sport,
        "sportCorrelations": player_matches,
        "team": team_matches[0]["team"] if team_matches else "",
        "teamCorrelations": team_matches,
    }


def infer_sport(raw: str, player_name: str = "") -> str:
    haystack = clean_rule_text(raw)
    for sport, aliases in CATEGORY_ALIASES.items():
        if any(text_contains_clean_term(haystack, alias) for alias in aliases):
            return sport
    player_key = clean_rule_text(player_name)
    if player_key in PLAYER_SPORT_HINTS:
        return PLAYER_SPORT_HINTS[player_key]
    compact_player_key = compact_name_key(player_name)
    for candidate, sport in PLAYER_SPORT_HINTS.items():
        if compact_name_key(candidate) == compact_player_key:
            return sport
    return ""


def find_known_player_sports(raw: str) -> list[dict[str, str]]:
    return [dict(item) for item in _find_known_player_sports_cached(str(raw or ""))]


@lru_cache(maxsize=4096)
def _find_known_player_sports_cached(raw: str) -> tuple[tuple[tuple[str, str], ...], ...]:
    haystack = f" {clean_rule_text(raw)} "
    compact_haystack = compact_name_key(raw)
    seen: set[str] = set()
    exact_matches: list[dict[str, str]] = []
    player_keys = SORTED_PLAYER_KEYS or sorted(PLAYER_SPORT_HINTS, key=len, reverse=True)
    for player in player_keys:
        key = clean_rule_text(player)
        compact_key = compact_name_key(player)
        sport = PLAYER_SPORT_HINTS[player]
        if not single_token_player_match_allowed(key, sport, haystack):
            continue
        compact_match_allowed = bool(compact_key and len(compact_key) >= 6)
        if f" {key} " not in haystack and not (compact_match_allowed and compact_key in compact_haystack):
            continue
        if any(player_name_contains(existing["key"], key) for existing in exact_matches):
            continue
        unique_key = f"{key}:{sport}"
        if unique_key in seen:
            continue
        seen.add(unique_key)
        exact_matches.append({
            "key": key,
            "playerName": PLAYER_DISPLAY_NAMES.get(key) or title_case_name(key),
            "sport": sport,
        })
    if exact_matches:
        return tuple(tuple(sorted(match.items())) for match in exact_matches)

    partial_matches: list[dict[str, str]] = []
    for token in sorted(PARTIAL_PLAYER_HINTS, key=len, reverse=True):
        if f" {token} " not in haystack:
            continue
        hint = PARTIAL_PLAYER_HINTS[token]
        if not single_token_player_match_allowed(token, hint["sport"], haystack):
            continue
        unique_key = f"{hint['key']}:{hint['sport']}"
        if unique_key in seen:
            continue
        seen.add(unique_key)
        partial_matches.append(dict(hint))
    return tuple(tuple(sorted(match.items())) for match in partial_matches)


def single_token_player_match_allowed(key: str, sport: str, haystack: str) -> bool:
    if len(str(key or "").split()) != 1:
        return True
    category = canonical_sport_label(sport) or clean_rule_text(sport)
    if category not in SINGLE_TOKEN_CONTEXT_REQUIRED_CATEGORIES:
        return True
    aliases = CATEGORY_ALIASES.get(category, [category])
    return any(text_contains_clean_term(haystack, alias) for alias in aliases)


def find_known_player_teams(raw: str, sport_correlations: list[dict[str, str]] | None = None) -> list[dict[str, str]]:
    seen: set[str] = set()
    teams: list[dict[str, str]] = []
    correlations = sport_correlations if sport_correlations is not None else find_known_player_sports(raw)
    for correlation in correlations:
        key = clean_rule_text(correlation.get("key") or correlation.get("playerName") or "")
        for team in PLAYER_TEAM_HINTS.get(key, []):
            unique_key = f"{key}:{clean_rule_text(team)}"
            if unique_key in seen:
                continue
            seen.add(unique_key)
            teams.append({
                "key": key,
                "playerName": correlation.get("playerName") or PLAYER_DISPLAY_NAMES.get(key) or title_case_name(key),
                "sport": correlation.get("sport") or PLAYER_SPORT_HINTS.get(key, ""),
                "team": team,
            })
    return teams


def player_name_contains(longer_player: str, shorter_player: str) -> bool:
    longer = clean_rule_text(longer_player)
    shorter = clean_rule_text(shorter_player)
    if longer != shorter and shorter in longer:
        return True
    compact_longer = compact_name_key(longer_player)
    compact_shorter = compact_name_key(shorter_player)
    return bool(compact_longer and compact_shorter and compact_longer != compact_shorter and compact_shorter in compact_longer)


def matcher_matches_text(matcher: str, haystack: str) -> bool:
    cleaned = clean_rule_text(matcher)
    aliases = aliases_for(cleaned)
    if len(aliases) > 1:
        return any(text_contains_clean_term(haystack, alias) for alias in aliases)
    terms = [term for term in cleaned.split() if len(term) >= 2]
    return all(text_contains_clean_term(haystack, term) for term in terms) if terms else True


def player_matches_value(expected_player: str, value: dict[str, Any]) -> bool:
    expected = clean_rule_text(expected_player)
    if not expected:
        return False
    parsed_player = clean_rule_text(value.get("playerName") or "")
    compact_expected = compact_name_key(expected_player)
    compact_parsed_player = compact_name_key(value.get("playerName") or "")
    if parsed_player and parsed_player == expected:
        return True
    if compact_expected and compact_parsed_player and compact_parsed_player == compact_expected:
        return True
    return any(
        clean_rule_text(correlation.get("playerName") or "") == expected
        or clean_rule_text(correlation.get("key") or "") == expected
        or compact_name_key(correlation.get("playerName") or "") == compact_expected
        or compact_name_key(correlation.get("key") or "") == compact_expected
        for correlation in value.get("sportCorrelations") or []
    )


def sport_label_matches_value(expected_sport: str, value: dict[str, Any], haystack: str) -> bool:
    if not canonical_sport_label(expected_sport):
        return False
    return sport_matches_value(expected_sport, value, haystack)


def sport_matches_value(expected_sport: str, value: dict[str, Any], haystack: str) -> bool:
    expected = clean_rule_text(expected_sport)
    parsed_sport = clean_rule_text(value.get("sport") or "")
    parsed_player = clean_rule_text(value.get("playerName") or "")
    if parsed_player and parsed_player == expected:
        return True
    if player_matches_value(expected_sport, value):
        return True
    if parsed_sport and (parsed_sport == expected or any(clean_rule_text(alias) == parsed_sport for alias in aliases_for(expected))):
        return True
    if any(
        clean_rule_text(correlation.get("sport") or "") == expected
        or any(clean_rule_text(alias) == clean_rule_text(correlation.get("sport") or "") for alias in aliases_for(expected))
        for correlation in value.get("sportCorrelations") or []
    ):
        return True
    return matcher_matches_text(expected_sport, haystack)


def canonical_sport_label(value: str) -> str:
    key = clean_rule_text(value)
    if not key:
        return ""
    for sport, aliases in CATEGORY_ALIASES.items():
        if clean_rule_text(sport) == key or any(clean_rule_text(alias) == key for alias in aliases):
            return sport
    return ""


def aliases_for(value: str) -> list[str]:
    key = clean_rule_text(value)
    for sport, aliases in CATEGORY_ALIASES.items():
        if clean_rule_text(sport) == key or any(clean_rule_text(alias) == key for alias in aliases):
            return aliases
    return [key]


def text_contains_clean_term(haystack: str, term: str) -> bool:
    cleaned_haystack = f" {clean_rule_text(haystack)} "
    cleaned_term = clean_rule_text(term)
    if not cleaned_term:
        return False
    return re.search(rf"\s{re.escape(cleaned_term)}s?(?=\s)", cleaned_haystack) is not None


def clean_rule_text(value: Any) -> str:
    text = strip_accents(str(value or "")).lower()
    text = re.sub(r"[\u0300-\u036f]", "", text)
    text = re.sub(r"\b(fill|bid|range|buy|pay|up to|acceptable|target|min|max|price)\b", " ", text)
    text = re.sub(r"[:|,;()[\]{}]+", " ", text)
    text = re.sub(r"[^a-z0-9/.' -]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def strip_accents(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", str(value or ""))
        if not unicodedata.combining(char)
    )


def compact_name_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_rule_text(value))


def source_lines(text: str) -> list[str]:
    rows: list[str] = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        try:
            csv_line = re.sub(r"(?<=\d),(?=\d{3}\b)", "", line)
            parsed = next(csv.reader([csv_line]))
            line = " ".join(cell.strip() for cell in parsed if cell.strip())
        except Exception:
            pass
        rows.append(line)
    return rows


def rule_matches(rule: AssignmentRule, haystack: str, price: float) -> bool:
    if rule.matcher and not term_matches(rule.matcher, haystack):
        return False
    if rule.min_price is not None and price < rule.min_price:
        return False
    if rule.max_price is not None and price > rule.max_price:
        return False
    return True


def term_matches(term: str, haystack: str) -> bool:
    words = clean_text(term).split()
    if not words:
        return True
    aliases = {
        "b ball": ["basketball", "nba"],
        "bball": ["basketball", "nba"],
        "poke": ["pokemon"],
        "one piece": ["onepiece", "1 piece"],
        "wwe": ["wrestling", "wwf"],
        "f1": ["formula 1", "formula one", "formula1"],
        "star wars": ["starwars"],
        "ufc": ["mma"],
    }
    options = [words]
    alias_text = " ".join(words)
    if canonical_sport_label(alias_text) and sport_term_matches_known_player(alias_text, haystack):
        return True
    if matcher_matches_text(alias_text, haystack):
        return True
    options.extend(alias.split() for alias in aliases.get(alias_text, []))
    return any(all(word_matches(word, haystack) for word in option) for option in options)


def sport_term_matches_known_player(sport: str, haystack: str) -> bool:
    expected = canonical_sport_label(sport)
    if not expected:
        return False
    return any(canonical_sport_label(match.get("sport") or "") == expected for match in find_known_player_sports(haystack))


def word_matches(word: str, haystack: str) -> bool:
    candidates = {word}
    if len(word) > 3 and word.endswith("s"):
        candidates.add(word[:-1])
    return any(re.search(rf"\b{re.escape(candidate)}s?\b", haystack) for candidate in candidates)


def parse_grade(text: str, fallback_grader: str = "") -> tuple[str, float | None]:
    match = re.search(r"\b(PSA|BGS|SGC|CGC)\s*(?:g(?:rade)?\s*)?([0-9]+(?:[._][0-9])?)?\b", text, re.I)
    company = match.group(1).upper() if match else str(fallback_grader or "").upper()
    grade = to_number(match.group(2).replace("_", ".") if match and match.group(2) else None)
    return company, grade


def parse_rate(value: Any) -> float | None:
    text = str(value or "").strip()
    match = re.search(r"(\d+(?:\.\d+)?)\s*%", text)
    if match:
        return float(match.group(1)) / 100
    number = to_number(text)
    if number is None:
        return None
    return number / 100 if number > 1 else number


def parse_money(value: Any) -> float | None:
    text = str(value or "").strip().lower().replace("$", "").replace(",", "")
    if not text:
        return None
    multiplier = 1000 if text.endswith("k") else 1
    text = text.removesuffix("k")
    return to_number(text) * multiplier if to_number(text) is not None else None


def to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except Exception:
        return None


def split_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in str(value).split(",") if item.strip()]


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9/.' -]+", " ", strip_accents(str(value or "")).lower())).strip()


initialize_player_sport_data()
