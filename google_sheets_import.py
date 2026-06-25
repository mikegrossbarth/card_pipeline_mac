from __future__ import annotations

import json
import os
import secrets
import ssl
import time
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parent
TOKEN_PATH = ROOT / "lucas_google_sheets_token.json"
SCOPES = ("https://www.googleapis.com/auth/spreadsheets.readonly",)
AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
TOKEN_URL = "https://oauth2.googleapis.com/token"
_ENV_LOADED = False
LAST_OAUTH_DIAGNOSTICS: dict[str, Any] = {}
_SSL_CONTEXT: ssl.SSLContext | None = None


class GoogleSheetsAuthError(RuntimeError):
    pass


def google_ssl_context() -> ssl.SSLContext:
    global _SSL_CONTEXT
    if _SSL_CONTEXT is not None:
        return _SSL_CONTEXT
    cafile = os.environ.get("SSL_CERT_FILE") or os.environ.get("REQUESTS_CA_BUNDLE") or ""
    if not cafile:
        try:
            import certifi

            cafile = certifi.where()
            if cafile:
                os.environ.setdefault("SSL_CERT_FILE", cafile)
                os.environ.setdefault("REQUESTS_CA_BUNDLE", cafile)
        except Exception:
            cafile = ""
    LAST_OAUTH_DIAGNOSTICS["ssl_cert_file"] = cafile or "system default"
    try:
        _SSL_CONTEXT = ssl.create_default_context(cafile=cafile or None)
    except Exception:
        _SSL_CONTEXT = ssl.create_default_context()
    return _SSL_CONTEXT


def google_ssl_error_message(error: BaseException) -> str:
    LAST_OAUTH_DIAGNOSTICS["ssl_error"] = str(error)
    LAST_OAUTH_DIAGNOSTICS.setdefault("ssl_cert_file", os.environ.get("SSL_CERT_FILE") or os.environ.get("REQUESTS_CA_BUNDLE") or "system default")
    return (
        "Google HTTPS certificate verification failed: unable to get local issuer certificate. "
        "On macOS, run Install Certificates.command for the Python version used by L.U.C.A.S, "
        "then rerun install_dependencies.sh or update certifi in the app environment."
    )


def authorize_google_sheets(interactive: bool = True) -> dict[str, Any]:
    client_id, client_secret = oauth_client_config()
    token = load_token()
    if token and token_matches_client(token, client_id):
        token = refresh_token_if_needed(token, client_id, client_secret)
        if token.get("access_token"):
            return token
    if not interactive:
        raise GoogleSheetsAuthError(
            "Google Sheets is not connected yet. Open Assignment Rules and click Connect Google, then try again."
        )
    return run_desktop_oauth(client_id, client_secret)


def read_google_sheet_text(url: str, interactive: bool = False, sheet_name: str = "") -> str:
    sheets = read_google_sheet_tabs(url, interactive=interactive, sheet_name=sheet_name)
    lines: list[str] = []
    for title, values in sheets:
        lines.append(f"# {title}")
        for row in values:
            cells = [str(cell).strip() for cell in row if str(cell).strip()]
            if cells:
                lines.append(" ".join(cells))
    return "\n".join(lines)


def read_google_sheet_tabs(url: str, interactive: bool = False, sheet_name: str = "") -> list[tuple[str, list[list[Any]]]]:
    spreadsheet_id = spreadsheet_id_from_url(url)
    if not spreadsheet_id:
        raise ValueError("Use a Google Sheets URL for this rules or payout source.")
    token = authorize_google_sheets(interactive=interactive)
    access_token = str(token.get("access_token") or "")
    metadata = sheets_api_json(
        f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}"
        "?fields=properties.title,sheets(properties(title,sheetId,gridProperties(rowCount,columnCount)))",
        access_token,
    )
    tabs: list[tuple[str, list[list[Any]]]] = []
    found_sheet = False
    for sheet in metadata.get("sheets") or []:
        title = str(((sheet or {}).get("properties") or {}).get("title") or "").strip()
        if not title:
            continue
        if sheet_name and title.lower() != sheet_name.lower():
            continue
        found_sheet = True
        values = read_sheet_values(spreadsheet_id, title, access_token)
        tabs.append((title, values))
    if sheet_name and not found_sheet:
        raise ValueError(f"Google Sheet does not contain a tab named {sheet_name}.")
    return tabs


def export_google_sheet_to_xlsx(url: str, output_path: Path, interactive: bool = False) -> Path:
    tabs = read_google_sheet_tabs(url, interactive=interactive)
    if not tabs:
        raise ValueError("Google Sheet returned no tabs to export.")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_titles: set[str] = set()
    for title, values in tabs:
        sheet = workbook.create_sheet(unique_sheet_title(title, used_titles))
        for row in values:
            sheet.append(list(row))
        if sheet.max_row:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output_path)
    workbook.close()
    return output_path


def unique_sheet_title(title: str, used_titles: set[str]) -> str:
    cleaned = clean_sheet_title(title)
    candidate = cleaned
    index = 2
    while candidate.lower() in used_titles:
        suffix = f" {index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(candidate.lower())
    return candidate


def clean_sheet_title(title: str) -> str:
    cleaned = "".join(" " if char in "[]:*?/\\\\" else char for char in str(title or "").strip())
    cleaned = " ".join(cleaned.split()) or "Sheet"
    return cleaned[:31]


def spreadsheet_id_from_url(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "/" not in text and " " not in text and len(text) > 20:
        return text
    try:
        parsed = urllib.parse.urlparse(text)
    except Exception:
        return ""
    if "docs.google.com" not in parsed.netloc or "/spreadsheets/" not in parsed.path:
        return ""
    parts = parsed.path.split("/")
    try:
        index = parts.index("d")
    except ValueError:
        return ""
    return parts[index + 1] if index + 1 < len(parts) else ""


def oauth_client_config() -> tuple[str, str]:
    load_local_env_files()
    client_id = (
        os.environ.get("GOOGLE_SHEETS_OAUTH_CLIENT_ID")
        or os.environ.get("GOOGLE_OAUTH_CLIENT_ID")
        or ""
    ).strip()
    client_secret = (
        os.environ.get("GOOGLE_SHEETS_OAUTH_CLIENT_SECRET")
        or os.environ.get("GOOGLE_OAUTH_CLIENT_SECRET")
        or ""
    ).strip()
    if not client_id:
        raise GoogleSheetsAuthError(
            "Missing GOOGLE_SHEETS_OAUTH_CLIENT_ID in .env. Create a Google OAuth Desktop client and add its client ID."
        )
    return client_id, client_secret


def load_local_env_files() -> None:
    global _ENV_LOADED
    if _ENV_LOADED:
        return
    _ENV_LOADED = True
    for path in (ROOT / ".env", ROOT / "photo_tool" / "app" / ".env", ROOT / "photo_tool" / ".env"):
        load_simple_env(path)


def load_simple_env(path: Path) -> None:
    try:
        lines = path.read_text(encoding="utf-8-sig").splitlines()
    except OSError:
        return
    for raw_line in lines:
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if not key or key in os.environ:
            continue
        os.environ[key] = clean_env_value(value)


def clean_env_value(value: str) -> str:
    text = value.strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1]
    return text


def load_token() -> dict[str, Any]:
    try:
        payload = json.loads(TOKEN_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def save_token(token: dict[str, Any]) -> None:
    LAST_OAUTH_DIAGNOSTICS["token_save_attempted"] = True
    LAST_OAUTH_DIAGNOSTICS["token_path"] = str(TOKEN_PATH)
    TOKEN_PATH.write_text(json.dumps(token, indent=2), encoding="utf-8")
    LAST_OAUTH_DIAGNOSTICS["token_saved"] = True


def token_matches_client(token: dict[str, Any], client_id: str) -> bool:
    saved_client_id = str(token.get("client_id") or "")
    return not saved_client_id or saved_client_id == client_id


def refresh_token_if_needed(token: dict[str, Any], client_id: str, client_secret: str) -> dict[str, Any]:
    expires_at = float(token.get("expires_at") or 0)
    if token.get("access_token") and expires_at > time.time() + 90:
        return token
    refresh_token = str(token.get("refresh_token") or "")
    if not refresh_token:
        return token
    payload = {
        "client_id": client_id,
        "refresh_token": refresh_token,
        "grant_type": "refresh_token",
    }
    if client_secret:
        payload["client_secret"] = client_secret
    refreshed = post_form(TOKEN_URL, payload)
    merged = {**token, **refreshed, "client_id": client_id}
    merged["refresh_token"] = refreshed.get("refresh_token") or refresh_token
    merged["expires_at"] = time.time() + int(refreshed.get("expires_in") or 3600)
    save_token(merged)
    return merged


def run_desktop_oauth(client_id: str, client_secret: str) -> dict[str, Any]:
    LAST_OAUTH_DIAGNOSTICS.clear()
    LAST_OAUTH_DIAGNOSTICS.update(
        {
            "callback_received": False,
            "authorization_code_received": False,
            "token_exchange_attempted": False,
            "token_save_attempted": False,
            "token_saved": False,
            "token_path": str(TOKEN_PATH),
        }
    )
    state = secrets.token_urlsafe(24)
    server = OAuthCallbackServer(("127.0.0.1", 0), OAuthCallbackHandler)
    server.timeout = 1
    server.expected_state = state
    redirect_uri = f"http://127.0.0.1:{server.server_port}/oauth2callback"
    params = {
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "response_type": "code",
        "scope": " ".join(SCOPES),
        "access_type": "offline",
        "prompt": "consent",
        "state": state,
    }
    webbrowser.open(f"{AUTH_URL}?{urllib.parse.urlencode(params)}")
    try:
        deadline = time.time() + 180
        while not server.done and time.time() < deadline:
            server.handle_request()
    finally:
        server.server_close()
    if server.error:
        LAST_OAUTH_DIAGNOSTICS["error"] = server.error
        raise GoogleSheetsAuthError(server.error)
    if not server.code:
        LAST_OAUTH_DIAGNOSTICS["error"] = "Timed out before approval finished."
        raise GoogleSheetsAuthError(
            "Google Sheets OAuth timed out before approval finished. Try Connect Google again and complete the Allow screen in the browser."
        )
    LAST_OAUTH_DIAGNOSTICS["callback_received"] = True
    LAST_OAUTH_DIAGNOSTICS["authorization_code_received"] = True

    payload = {
        "client_id": client_id,
        "code": server.code,
        "grant_type": "authorization_code",
        "redirect_uri": redirect_uri,
    }
    if client_secret:
        payload["client_secret"] = client_secret
    LAST_OAUTH_DIAGNOSTICS["token_exchange_attempted"] = True
    token = post_form(TOKEN_URL, payload)
    token["client_id"] = client_id
    token["expires_at"] = time.time() + int(token.get("expires_in") or 3600)
    save_token(token)
    return token


class OAuthCallbackServer(ThreadingHTTPServer):
    expected_state: str = ""
    code: str = ""
    error: str = ""
    done: bool = False


class OAuthCallbackHandler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path != "/oauth2callback":
            self.send_response(204)
            self.end_headers()
            return
        LAST_OAUTH_DIAGNOSTICS["callback_received"] = True
        query = urllib.parse.parse_qs(parsed.query)
        state = query.get("state", [""])[0]
        if state != self.server.expected_state:
            self.server.error = "Google Sheets OAuth state did not match. Try connecting again."
            self.server.done = True
            self.send_oauth_response(False)
            return
        oauth_error = query.get("error", [""])[0]
        if oauth_error:
            description = query.get("error_description", [""])[0]
            detail = f": {urllib.parse.unquote_plus(description)}" if description else ""
            self.server.error = f"Google Sheets OAuth failed: {oauth_error}{detail}"
            self.server.done = True
            self.send_oauth_response(False)
            return
        self.server.code = query.get("code", [""])[0]
        LAST_OAUTH_DIAGNOSTICS["authorization_code_received"] = bool(self.server.code)
        if not self.server.code:
            self.server.error = (
                "Google Sheets OAuth finished without an authorization code. Try Connect Google again, choose the Google account that can open the sheet, and click Allow."
            )
        self.server.done = True
        self.send_oauth_response(bool(self.server.code))

    def send_oauth_response(self, success: bool) -> None:
        body = (
            "<html><body><h2>Google Sheets connected.</h2>"
            "<p>You can close this browser tab and return to L.U.C.A.S.</p></body></html>"
            if success
            else "<html><body><h2>Google Sheets connection failed.</h2>"
            "<p>Return to L.U.C.A.S and try again.</p></body></html>"
        )
        data = body.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, _format: str, *_args: Any) -> None:
        return


def sheets_api_json(url: str, access_token: str) -> dict[str, Any]:
    request = urllib.request.Request(url, headers={"Authorization": f"Bearer {access_token}"})
    try:
        with urllib.request.urlopen(request, timeout=20, context=google_ssl_context()) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        text = error.read().decode("utf-8", errors="replace")
        if error.code in {401, 403}:
            raise GoogleSheetsAuthError(
                f"Google Sheets authorization failed ({error.code}). Connect Google again or confirm this account can open the sheet."
            ) from error
        raise ValueError(f"Google Sheets API failed ({error.code}): {text[:200]}") from error
    except ssl.SSLCertVerificationError as error:
        raise GoogleSheetsAuthError(google_ssl_error_message(error)) from error
    except urllib.error.URLError as error:
        if isinstance(error.reason, ssl.SSLCertVerificationError):
            raise GoogleSheetsAuthError(google_ssl_error_message(error.reason)) from error
        raise


def read_sheet_values(spreadsheet_id: str, title: str, access_token: str) -> list[list[Any]]:
    encoded_title = urllib.parse.quote(title, safe="")
    payload = sheets_api_json(
        f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{encoded_title}",
        access_token,
    )
    values = payload.get("values") or []
    return values if isinstance(values, list) else []


def post_form(url: str, payload: dict[str, str]) -> dict[str, Any]:
    data = urllib.parse.urlencode(payload).encode("utf-8")
    request = urllib.request.Request(url, data=data, method="POST")
    request.add_header("Content-Type", "application/x-www-form-urlencoded")
    try:
        with urllib.request.urlopen(request, timeout=30, context=google_ssl_context()) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        text = error.read().decode("utf-8", errors="replace")
        LAST_OAUTH_DIAGNOSTICS["error"] = f"Google OAuth token request failed ({error.code}): {text[:220]}"
        raise GoogleSheetsAuthError(f"Google OAuth token request failed ({error.code}): {text[:220]}") from error
    except ssl.SSLCertVerificationError as error:
        message = google_ssl_error_message(error)
        LAST_OAUTH_DIAGNOSTICS["error"] = message
        raise GoogleSheetsAuthError(message) from error
    except urllib.error.URLError as error:
        if isinstance(error.reason, ssl.SSLCertVerificationError):
            message = google_ssl_error_message(error.reason)
            LAST_OAUTH_DIAGNOSTICS["error"] = message
            raise GoogleSheetsAuthError(message) from error
        raise
