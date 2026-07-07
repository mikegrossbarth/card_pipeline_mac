from __future__ import annotations

import re
import time
from datetime import date, datetime, time as datetime_time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import range_boundaries


DEFAULT_SHEET = "Cards"
RECEIVED_HEADER = "RECEIVED"
RECEIVED_FILL = PatternFill("solid", fgColor="C6EFCE")
RECEIVED_FONT = Font(color="006100")

PHOTO_EXPORT_HEADERS = {
    "cert": ("certificationnumber", "certnumber", "cert"),
    "description": ("carddescription", "card", "description"),
    "card_number": ("cardnumber",),
    "player": ("playersubject", "player", "subject"),
    "year": ("year",),
    "set": ("set",),
    "subset": ("subset",),
    "parallel": ("parallel",),
    "grader": ("gradingcompany", "grader", "gradingco"),
    "grade": ("grade",),
    "source": ("sourcephoto", "sourcefile"),
}

PHOTO_EXPORT_POSITIONS = {
    "cert": 1,
    "description": 2,
    "card_number": 3,
    "player": 4,
    "year": 5,
    "set": 6,
    "subset": 7,
    "parallel": 8,
    "grader": 9,
    "grade": 10,
    "source": 16,
}

CERT_HEADERS = ("certificationnumber", "certnumber", "cert", "certification", "cert#")
GRADER_HEADERS = ("company", "gradingcompany", "grader", "gradingco", "gradingcompanyname")
CARD_HEADERS = ("carddescription", "card", "description", "title", "cardtitle", "item", "itemtitle")
SPORT_HEADERS = ("sport", "sports", "category", "cardcategory", "assignmentcategory")
DATE_HEADERS = ("date", "dateadded", "createddate", "receiveddate", "purchasedate")
PURCHASE_PRICE_HEADERS = ("purchaseprice", "purchase", "price", "cost", "buyprice", "paid")
CARD_LADDER_VALUE_HEADERS = (
    "cardladdervalue",
    "cardladder",
    "clvalue",
    "cardladderprice",
    "laddervalue",
    "ladderprice",
    "value",
)
COMPS_AVERAGE_HEADERS = (
    "comps",
    "comp",
    "compsvalue",
    "compvalue",
    "compprice",
    "compsprice",
    "compsavg",
    "compavg",
    "averagecomp",
    "averagecomps",
    "avgcomp",
    "avgcomps",
    "cardladdercomps",
    "cardladdercompsaverage",
    "cardladdercompaverage",
    "clcomps",
    "clcomp",
    "compsaverage",
    "compaverage",
)
CY_VALUE_HEADERS = (
    "cyvalue",
    "cyestimate",
    "estimate",
    "cybuyprice",
    "courtyardvalue",
    "courtyardestimate",
    "courtyardbuyprice",
    "cardyardvalue",
    "cardyardestimate",
    "cardyardbuyprice",
)
CY_CONFIDENCE_HEADERS = (
    "cyconfidence",
    "confidence",
    "courtyardconfidence",
    "cardyardconfidence",
)
COMP_DETAILS_HEADERS = ("cardladdercompdetails", "compdetails", "cardladdercompsdetail", "compsdetails")
BEST_COMPANY_HEADERS = ("bestcompany", "assignedcompany", "companyassignment")
ESTIMATED_PAYOUT_HEADERS = ("estimatedpayout", "estpayout", "payout")
STATUS_HEADERS = ("compstatus", "status", "assignmentstatus")
NOTES_HEADERS = ("notes", "note")
SOURCE_HEADERS = ("source", "sourcephoto", "sourcefile", "file")
SIMPLE_HEADER_ALIASES = (
    CERT_HEADERS
    + GRADER_HEADERS
    + CARD_HEADERS
    + SPORT_HEADERS
    + DATE_HEADERS
    + PURCHASE_PRICE_HEADERS
    + CARD_LADDER_VALUE_HEADERS
    + COMPS_AVERAGE_HEADERS
    + CY_VALUE_HEADERS
    + CY_CONFIDENCE_HEADERS
    + BEST_COMPANY_HEADERS
    + ESTIMATED_PAYOUT_HEADERS
)

COMPANY_SHEET_HEADERS = [
    "Date Added",
    "Source Sheet",
    "Source",
    "Certification Number",
    "Grader",
    "Card Description",
    "Purchase Price",
    "Card Ladder Value",
    "Comps",
    "CY Estimate",
    "CY Confidence",
    "Best Company",
    "Estimated Payout",
    "Status",
    "Notes",
]
CY_COMPANY_SHEET_HEADERS = [
    "Grader",
    "Cert",
    "Description",
    "Grade",
    "Purchase",
    "Estimate",
    "Confidence",
    "Date Added",
    "Source Sheet",
    "Source",
    "Sport",
    "Card Ladder Value",
    "Comps",
    "Best Company",
    "Estimated Payout",
    "Status",
    "Notes",
]
FANATICS_COMPANY_SHEET_HEADERS = [
    "Category",
    "Card",
    "Grade",
    "Cert #",
    "CL Value",
    "Payout",
    "Date Added",
    "Source Sheet",
    "Source",
    "Grader",
    "Purchase Price",
    "Comps",
    "CY Estimate",
    "CY Confidence",
    "Best Company",
    "Status",
    "Notes",
]


def read_simple_spreadsheet(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows: list[dict[str, Any]] = []
        header_row = _simple_header_row(sheet)
        has_header = bool(header_row)
        headers = _header_map_for_row(sheet, header_row or 1) if has_header else {}
        start_row = (header_row or 0) + 1 if has_header else 1
        cert_fallback = None if has_header else 1
        card_fallback = None if has_header else 2
        purchase_fallback = None if has_header else 3
        source_fallback = None if has_header else 4
        for row_index in range(start_row, _sheet_max_row(sheet) + 1):
            date_added = clean_part(_cell_by_header(sheet, row_index, headers, DATE_HEADERS, None))
            cert = normalize_cert(_cell_by_header(sheet, row_index, headers, CERT_HEADERS, cert_fallback))
            grader = normalize_grader(_cell_by_header(sheet, row_index, headers, GRADER_HEADERS, None))
            card = clean_part(_cell_by_header(sheet, row_index, headers, CARD_HEADERS, card_fallback))
            sport = clean_part(_cell_by_header(sheet, row_index, headers, SPORT_HEADERS, None))
            purchase_price = parse_money(_cell_by_header(sheet, row_index, headers, PURCHASE_PRICE_HEADERS, purchase_fallback))
            card_ladder_value = parse_money(_cell_by_header(sheet, row_index, headers, CARD_LADDER_VALUE_HEADERS, None))
            comps_average = parse_money(_cell_by_header(sheet, row_index, headers, COMPS_AVERAGE_HEADERS, None))
            cy_value = parse_money(_cell_by_header(sheet, row_index, headers, CY_VALUE_HEADERS, None))
            cy_confidence = _cell_by_header(sheet, row_index, headers, CY_CONFIDENCE_HEADERS, None)
            comp_details = clean_part(_cell_by_header(sheet, row_index, headers, COMP_DETAILS_HEADERS, None))
            best_company = clean_part(_cell_by_header(sheet, row_index, headers, BEST_COMPANY_HEADERS, None))
            estimated_payout = parse_money(_cell_by_header(sheet, row_index, headers, ESTIMATED_PAYOUT_HEADERS, None))
            status = clean_part(_cell_by_header(sheet, row_index, headers, STATUS_HEADERS, None))
            notes = clean_part(_cell_by_header(sheet, row_index, headers, NOTES_HEADERS, None))
            source = clean_part(_cell_by_header(sheet, row_index, headers, SOURCE_HEADERS, source_fallback))
            received = _is_received_value(_cell_by_header(sheet, row_index, headers, (_normalize_header(RECEIVED_HEADER),), None))
            if not cert and not card and purchase_price is None:
                continue
            grader = grader or infer_grader(card)
            rows.append(
                {
                    "cert_number": cert,
                    "card_title": card,
                    "sport": sport,
                    "grader": grader,
                    "purchase_price": purchase_price,
                    "card_ladder_value": card_ladder_value,
                    "card_ladder_comps_average": comps_average,
                    "cy_value": cy_value,
                    "cy_confidence": cy_confidence,
                    "card_ladder_comps": comp_details,
                    "best_company": best_company,
                    "estimated_payout": estimated_payout,
                    "source": source or f"{path.name}:{row_index}",
                    "workbook_sheet": sheet.title,
                    "workbook_row": row_index,
                    "date_added": date_added,
                    "received": received,
                    "status": status,
                    "notes": notes or _setup_notes(cert, card, grader),
                }
            )
        return rows
    finally:
        workbook.close()


def read_photo_export(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        schema = _detect_photo_schema(sheet)
        rows: list[dict[str, Any]] = []
        for row_index in range(schema["first_data_row"], _sheet_max_row(sheet) + 1):
            source = _source_row(sheet, row_index, schema["headers"])
            cert = normalize_cert(source.get("cert"))
            card = build_card_title(source)
            grader = normalize_grader(source.get("grader")) or infer_grader(card)
            if not cert and not card:
                continue
            rows.append(
                {
                    "cert_number": cert,
                    "card_title": card,
                    "grader": grader,
                    "purchase_price": None,
                    "source": clean_part(source.get("source")) or f"{path.name}:{row_index}",
                    "notes": _setup_notes(cert, card, grader),
                }
            )
        return rows
    finally:
        workbook.close()


def workbook_sheet_names(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def summarize_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        row_count = 0
        received_count = 0
        purchase_total = 0.0
        estimated_payout_total = 0.0
        for sheet in workbook.worksheets:
            header_row = 1 if _looks_like_simple_header(sheet) else None
            headers = _header_map_for_row(sheet, 1) if header_row else {}
            first_data_row = 2 if header_row else 1
            cert_col = _cert_column(sheet) or 1
            card_col = _card_column(sheet) or 2
            price_col = _price_column(sheet) or 3
            payout_col = next((headers.get(header) for header in ESTIMATED_PAYOUT_HEADERS if headers.get(header)), None)
            received_col = _received_column(sheet)
            for row_index in range(first_data_row, _sheet_max_row(sheet) + 1):
                cert = normalize_cert(sheet.cell(row_index, cert_col).value)
                card = clean_part(sheet.cell(row_index, card_col).value)
                purchase_price = parse_money(sheet.cell(row_index, price_col).value)
                estimated_payout = parse_money(sheet.cell(row_index, payout_col).value) if payout_col else None
                if not cert and not card and purchase_price is None:
                    continue
                row_count += 1
                if purchase_price is not None:
                    purchase_total += purchase_price
                if estimated_payout is not None:
                    estimated_payout_total += estimated_payout
                if received_col and _is_received_value(sheet.cell(row_index, received_col).value):
                    received_count += 1
        return {
            "path": path,
            "name": path.name,
            "row_count": row_count,
            "received_count": received_count,
            "purchase_total": purchase_total,
            "estimated_payout_total": estimated_payout_total,
            "all_received": bool(row_count and received_count == row_count),
            "partially_received": bool(received_count and received_count < row_count),
        }
    finally:
        workbook.close()


def write_pipeline_output(path: Path, rows: list[Any], source_lookup: dict[int, str] | None = None) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DEFAULT_SHEET
    headers = [
        "Source",
        "Certification Number",
        "Sport",
        "Card Description",
        "Purchase Price",
        "Card Ladder Value",
        "Comps",
        "CY Estimate",
        "CY Confidence",
        "Card Ladder Comp Details",
        "Card Ladder Screenshot",
        "Best Company",
        "Estimated Payout",
        "Comp Status",
        "Notes",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                (source_lookup or {}).get(row.excel_row, ""),
                row.cert_number,
                getattr(row, "category", ""),
                row.card_title,
                row.existing_value,
                row.card_ladder_value,
                row.card_ladder_comps_average,
                row.cy_value,
                row.cy_confidence,
                row.card_ladder_comps,
                row.card_ladder_screenshot,
                row.best_company,
                row.estimated_payout,
                row.status,
                row.notes,
            ]
        )

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [18, 22, 14, 62, 16, 18, 14, 14, 16, 58, 42, 18, 18, 20, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    workbook.save(path)
    return path


def append_company_sheet_rows(
    directory: Path,
    rows: list[Any],
    source_lookup: dict[int, str] | None = None,
    sheet_source_lookup: dict[int, str] | None = None,
    sheet_name_lookup: dict[str, str] | None = None,
) -> dict[str, Any]:
    result = {"files_updated": 0, "rows_added": 0, "skipped": 0, "errors": [], "added_records": []}
    grouped: dict[str, list[Any]] = {}
    for row in rows:
        company = clean_part(getattr(row, "best_company", ""))
        if not company:
            result["skipped"] += 1
            continue
        grouped.setdefault(company, []).append(row)
    if not grouped:
        return result
    directory.mkdir(parents=True, exist_ok=True)
    for company, company_rows in grouped.items():
        path = company_workbook_path(directory, company)
        sheet_name = (sheet_name_lookup or {}).get(company) or company_weekly_sheet_name()
        try:
            append_result = append_rows_to_company_sheet(path, company_rows, source_lookup, sheet_source_lookup, sheet_name=sheet_name)
            added = int(append_result.get("added") or 0)
            if added:
                result["files_updated"] += 1
                result["rows_added"] += added
                result["added_records"].extend(append_result.get("records") or [])
        except Exception as error:
            result["errors"].append(f"{company}: {error}")
    return result


def company_workbook_path(directory: Path, company: str) -> Path:
    safe_company = safe_filename(company) or "Company"
    company_dir = directory / safe_company
    company_dir.mkdir(parents=True, exist_ok=True)
    return company_dir / f"{safe_company}.xlsx"


def company_week_start_for_time(moment: datetime | None = None) -> date:
    moment = moment or datetime.now()
    current_monday = (moment - timedelta(days=moment.weekday())).date()
    if moment.weekday() == 0 and moment.time() < datetime_time(20, 0):
        return current_monday - timedelta(days=7)
    return current_monday


def company_weekly_sheet_name(today: date | datetime | None = None) -> str:
    if isinstance(today, datetime):
        start = company_week_start_for_time(today)
    elif today is None:
        start = company_week_start_for_time()
    else:
        start = week_start(today)
    return f"Week of {start:%Y-%m-%d}"


def company_weekly_sheet_path(directory: Path, company: str, today: date | datetime | None = None) -> Path:
    if isinstance(today, datetime):
        start = company_week_start_for_time(today)
    elif today is None:
        start = company_week_start_for_time()
    else:
        start = week_start(today)
    safe_company = safe_filename(company) or "Company"
    company_dir = directory / safe_company
    company_dir.mkdir(parents=True, exist_ok=True)
    return company_dir / f"{safe_company} WEEK OF {start:%Y-%m-%d}.xlsx"


def ensure_company_weekly_sheets(
    directory: Path,
    companies: list[str],
    week_start_date: date | None = None,
    week_start_lookup: dict[str, date] | None = None,
) -> dict[str, Any]:
    result = {"created": [], "existing": [], "errors": []}
    for company in companies:
        company = clean_part(company)
        if not company:
            continue
        try:
            path = company_workbook_path(directory, company)
            company_week_start = (week_start_lookup or {}).get(company) or week_start_date or date.today()
            sheet_name = company_weekly_sheet_name(company_week_start)
            created = ensure_company_workbook_sheet(path, sheet_name)
            sheet_ref = f"{path}:{sheet_name}"
            if created:
                result["created"].append(sheet_ref)
                continue
            result["existing"].append(sheet_ref)
        except Exception as error:
            result["errors"].append(f"{company}: {error}")
    return result


def create_empty_company_sheet(path: Path) -> Path:
    ensure_company_workbook_sheet(path, "Cards")
    return path


def ensure_company_workbook_sheet(path: Path, sheet_name: str) -> bool:
    path.parent.mkdir(parents=True, exist_ok=True)
    created = False
    if path.exists():
        workbook = load_workbook(path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            ensure_company_sheet_layout(sheet, path.parent.name)
            workbook.save(path)
            workbook.close()
            return False
        if workbook.sheetnames == ["Sheet"] and _sheet_max_row(workbook.active) == 1 and workbook.active.cell(1, 1).value is None:
            sheet = workbook.active
            sheet.title = sheet_name
        else:
            sheet = workbook.create_sheet(sheet_name)
        created = True
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        created = True
    if _sheet_max_row(sheet) <= 1 and sheet.cell(1, 1).value is None:
        set_company_sheet_headers(sheet, path.parent.name)
    style_company_sheet_header(sheet)
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)
    workbook.close()
    return created


def week_start(day: date) -> date:
    return day - timedelta(days=day.weekday())


def append_rows_to_company_sheet(path: Path, rows: list[Any], source_lookup: dict[int, str] | None = None, sheet_source_lookup: dict[int, str] | None = None, sheet_name: str | None = None) -> dict[str, Any]:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        workbook = load_workbook(path)
        if sheet_name:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
                set_company_sheet_headers(sheet, path.parent.name)
                style_company_sheet_header(sheet)
        else:
            sheet = workbook.active
        ensure_company_sheet_layout(sheet, path.parent.name)
        existing_certs = existing_sheet_certs(sheet)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name or "Cards"
        set_company_sheet_headers(sheet, path.parent.name)
        style_company_sheet_header(sheet)
        existing_certs = set()

    added = 0
    added_records: list[dict[str, Any]] = []
    today_text = datetime.now().strftime("%Y-%m-%d")
    company = path.parent.name
    for row in rows:
        cert = normalize_cert(getattr(row, "cert_number", ""))
        if cert and cert in existing_certs:
            continue
        source_sheet = (sheet_source_lookup or {}).get(row.excel_row, "")
        source = (source_lookup or {}).get(row.excel_row, "")
        purchase_price = parse_money(getattr(row, "existing_value", None))
        sale_price = parse_money(getattr(row, "estimated_payout", None))
        append_company_row(sheet, row, company, today_text, source_sheet, source, cert, purchase_price, sale_price)
        added_records.append(
            {
                "date_added": today_text,
                "company": company,
                "weekly_sheet": str(path),
                "weekly_sheet_name": f"{path.name}:{sheet.title}",
                "source_sheet": source_sheet,
                "source": source,
                "cert_number": cert,
                "grader": clean_part(getattr(row, "grader", "")),
                "card_title": clean_part(getattr(row, "card_title", "")),
                "purchase_price": purchase_price,
                "sale_price": sale_price,
                "card_ladder_value": parse_money(getattr(row, "card_ladder_value", None)),
                "comps": parse_money(getattr(row, "card_ladder_comps_average", None)),
                "cy_estimate": parse_money(getattr(row, "cy_value", None)),
                "cy_value": parse_money(getattr(row, "cy_value", None)),
                "cy_confidence": getattr(row, "cy_confidence", None),
                "best_company": clean_part(getattr(row, "best_company", "")),
                "status": clean_part(getattr(row, "status", "")),
                "notes": clean_part(getattr(row, "notes", "")),
            }
        )
        if cert:
            existing_certs.add(cert)
        added += 1
    if added:
        style_company_sheet_header(sheet)
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)
    workbook.close()
    return {"added": added, "records": added_records}


def existing_sheet_certs(sheet) -> set[str]:
    headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
    cert_col = headers.get("certificationnumber") or headers.get("certnumber") or headers.get("cert")
    if not cert_col:
        return set()
    return {
        normalize_cert(sheet.cell(row_index, cert_col).value)
        for row_index in range(2, _sheet_max_row(sheet) + 1)
        if normalize_cert(sheet.cell(row_index, cert_col).value)
    }


def style_company_sheet_header(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    if _is_cy_company_sheet(sheet):
        widths = [14, 18, 62, 12, 14, 14, 14, 14, 28, 22, 14, 18, 14, 18, 18, 20, 38]
    elif _is_fanatics_company_sheet(sheet):
        widths = [16, 62, 14, 18, 14, 14, 14, 28, 22, 14, 16, 14, 14, 16, 18, 20, 38]
    else:
        widths = [14, 28, 22, 22, 14, 62, 16, 18, 14, 14, 16, 18, 18, 20, 38]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    if _is_cy_company_sheet(sheet):
        for index in range(8, len(CY_COMPANY_SHEET_HEADERS) + 1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].hidden = True


def set_company_sheet_headers(sheet, company: str = "") -> None:
    if is_cy_company(company):
        headers = CY_COMPANY_SHEET_HEADERS
    elif is_fanatics_company(company):
        headers = FANATICS_COMPANY_SHEET_HEADERS
    else:
        headers = COMPANY_SHEET_HEADERS
    for index, header in enumerate(headers, start=1):
        sheet.cell(1, index).value = header


def is_cy_company(company: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", "", clean_part(company).lower())
    return normalized in {"cy", "courtyard", "courtyardcards", "courtyardcard"} or normalized.startswith("courtyard")


def is_fanatics_company(company: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", "", clean_part(company).lower())
    return normalized == "fanatics"


def _is_cy_company_sheet(sheet) -> bool:
    headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
    return bool(headers.get("cert") and headers.get("estimate") and headers.get("confidence"))


def _is_fanatics_company_sheet(sheet) -> bool:
    headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
    return all(headers.get(_normalize_header(header)) for header in FANATICS_COMPANY_SHEET_HEADERS[:6])


def ensure_company_sheet_layout(sheet, company: str = "") -> None:
    if is_fanatics_company(company):
        ensure_fanatics_company_sheet_layout(sheet)
        return
    if is_cy_company(company):
        headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
        has_data = _sheet_max_row(sheet) >= 2 and any(sheet.cell(row_index, 1).value is not None for row_index in range(2, _sheet_max_row(sheet) + 1))
        if (not headers or sheet.cell(1, 1).value is None) or (not has_data and not _is_cy_company_sheet(sheet)):
            set_company_sheet_headers(sheet, company)
            return
    ensure_cy_columns_after_comps(sheet)


def ensure_fanatics_company_sheet_layout(sheet) -> None:
    if _sheet_max_row(sheet) <= 1 and sheet.cell(1, 1).value is None:
        set_company_sheet_headers(sheet, "Fanatics")
        return
    if _is_fanatics_company_sheet(sheet):
        return
    headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
    if not headers:
        set_company_sheet_headers(sheet, "Fanatics")
        return
    row_values: list[list[Any]] = []
    max_row = _sheet_max_row(sheet)
    for row_index in range(2, max_row + 1):
        card = clean_part(_cell_by_header(sheet, row_index, headers, CARD_HEADERS, 6))
        grader = normalize_grader(_cell_by_header(sheet, row_index, headers, GRADER_HEADERS, 5))
        values = [
            clean_part(_cell_by_header(sheet, row_index, headers, SPORT_HEADERS, None)),
            card,
            _fanatics_grade_label(grader, card),
            normalize_cert(_cell_by_header(sheet, row_index, headers, CERT_HEADERS, 4)),
            _cell_by_header(sheet, row_index, headers, CARD_LADDER_VALUE_HEADERS, 8),
            _cell_by_header(sheet, row_index, headers, ESTIMATED_PAYOUT_HEADERS, 13),
            _cell_by_header(sheet, row_index, headers, ("dateadded", "date"), 1),
            clean_part(_cell_by_header(sheet, row_index, headers, ("sourcesheet",), 2)),
            clean_part(_cell_by_header(sheet, row_index, headers, SOURCE_HEADERS, 3)),
            grader,
            _cell_by_header(sheet, row_index, headers, PURCHASE_PRICE_HEADERS, 7),
            _cell_by_header(sheet, row_index, headers, COMPS_AVERAGE_HEADERS, 9),
            _cell_by_header(sheet, row_index, headers, CY_VALUE_HEADERS, 10),
            _cell_by_header(sheet, row_index, headers, CY_CONFIDENCE_HEADERS, 11),
            clean_part(_cell_by_header(sheet, row_index, headers, BEST_COMPANY_HEADERS, 12)),
            clean_part(_cell_by_header(sheet, row_index, headers, STATUS_HEADERS, 14)),
            clean_part(_cell_by_header(sheet, row_index, headers, NOTES_HEADERS, 15)),
        ]
        if any(value not in (None, "") for value in values):
            row_values.append(values)
    max_col = max(sheet.max_column, len(FANATICS_COMPANY_SHEET_HEADERS))
    for column in range(1, max_col + 1):
        sheet.cell(1, column).value = None
    set_company_sheet_headers(sheet, "Fanatics")
    for offset, values in enumerate(row_values, start=2):
        for column, value in enumerate(values, start=1):
            sheet.cell(offset, column).value = value
        for column in range(len(values) + 1, max_col + 1):
            sheet.cell(offset, column).value = None
    for row_index in range(len(row_values) + 2, max_row + 1):
        for column in range(1, max_col + 1):
            sheet.cell(row_index, column).value = None


def _fanatics_grade_label(grader: str, card_title: str) -> str:
    text = clean_part(card_title)
    match = re.search(r"\b(PSA|BGS|SGC|CGC)\s*([0-9]+(?:\.[0-9]+)?)\b", text, re.I)
    if match:
        return f"{match.group(1).upper()} {match.group(2)}"
    company = normalize_grader(grader)
    return company


def append_company_row(sheet, row: Any, company: str, today_text: str, source_sheet: str, source: str, cert: str, purchase_price: float | None, sale_price: float | None) -> None:
    if _is_cy_company_sheet(sheet):
        sheet.append(
            [
                getattr(row, "grader", ""),
                cert,
                getattr(row, "card_title", ""),
                "",
                purchase_price,
                getattr(row, "cy_value", None),
                getattr(row, "cy_confidence", None),
                today_text,
                source_sheet,
                source,
                getattr(row, "category", ""),
                getattr(row, "card_ladder_value", None),
                getattr(row, "card_ladder_comps_average", None),
                getattr(row, "best_company", ""),
                sale_price,
                getattr(row, "status", ""),
                getattr(row, "notes", ""),
            ]
        )
        return
    if _is_fanatics_company_sheet(sheet):
        card_title = getattr(row, "card_title", "")
        grader = getattr(row, "grader", "")
        sheet.append(
            [
                getattr(row, "category", ""),
                card_title,
                _fanatics_grade_label(grader, card_title),
                cert,
                getattr(row, "card_ladder_value", None),
                sale_price,
                today_text,
                source_sheet,
                source,
                grader,
                purchase_price,
                getattr(row, "card_ladder_comps_average", None),
                getattr(row, "cy_value", None),
                getattr(row, "cy_confidence", None),
                getattr(row, "best_company", ""),
                getattr(row, "status", ""),
                getattr(row, "notes", ""),
            ]
        )
        return
    sheet.append(
        [
            today_text,
            source_sheet,
            source,
            cert,
            getattr(row, "grader", ""),
            getattr(row, "card_title", ""),
            purchase_price,
            getattr(row, "card_ladder_value", None),
            getattr(row, "card_ladder_comps_average", None),
            getattr(row, "cy_value", None),
            getattr(row, "cy_confidence", None),
            getattr(row, "best_company", ""),
            sale_price,
            getattr(row, "status", ""),
            getattr(row, "notes", ""),
        ]
    )


def ensure_cy_columns_after_comps(sheet) -> None:
    headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
    comps_col = next((headers[alias] for alias in COMPS_AVERAGE_HEADERS if alias in headers), None)
    if not comps_col:
        return
    if not any(alias in headers for alias in CY_VALUE_HEADERS):
        sheet.insert_cols(comps_col + 1)
        cell = sheet.cell(1, comps_col + 1)
        cell.value = "CY Estimate"
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.font = Font(color="FFFFFF", bold=True)
        sheet.column_dimensions[cell.column_letter].width = 14
        headers = _header_map_for_row(sheet, 1)
    cy_col = next((headers[alias] for alias in CY_VALUE_HEADERS if alias in headers), comps_col + 1)
    cy_header = str(sheet.cell(1, cy_col).value or "").strip().lower()
    if cy_header == "cy value":
        sheet.cell(1, cy_col).value = "CY Estimate"
    if not any(alias in headers for alias in CY_CONFIDENCE_HEADERS):
        sheet.insert_cols(cy_col + 1)
        cell = sheet.cell(1, cy_col + 1)
        cell.value = "CY Confidence"
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.font = Font(color="FFFFFF", bold=True)
        sheet.column_dimensions[cell.column_letter].width = 16
    sheet.auto_filter.ref = sheet.dimensions


def read_company_profit_records(directory: Path) -> list[dict[str, Any]]:
    if not directory.exists():
        return []
    records: list[dict[str, Any]] = []
    for path in sorted(directory.glob("*/*.xlsx")):
        company = path.parent.name
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            for sheet in workbook.worksheets:
                if _sheet_max_row(sheet) < 2:
                    continue
                sheet_label = f"{path.name}:{sheet.title}"
                if len(workbook.worksheets) == 1 and sheet.title in {"Cards", "Sheet"}:
                    sheet_label = path.name
                headers = _header_map_for_row(sheet, 1)
                for row_index in range(2, _sheet_max_row(sheet) + 1):
                    cert = normalize_cert(_cell_by_header(sheet, row_index, headers, CERT_HEADERS, 4))
                    card = clean_part(_cell_by_header(sheet, row_index, headers, CARD_HEADERS, 6))
                    purchase = parse_money(_cell_by_header(sheet, row_index, headers, PURCHASE_PRICE_HEADERS, 7))
                    sale = parse_money(_cell_by_header(sheet, row_index, headers, ESTIMATED_PAYOUT_HEADERS, 12))
                    source_sheet = clean_part(_cell_by_header(sheet, row_index, headers, ("sourcesheet",), 2))
                    if not cert and not card and purchase is None and sale is None:
                        continue
                    if not source_sheet:
                        continue
                    records.append(
                        {
                            "date_added": clean_part(_cell_by_header(sheet, row_index, headers, ("dateadded", "date"), 1)),
                            "company": company,
                            "weekly_sheet": str(path),
                            "weekly_sheet_name": sheet_label,
                            "source_sheet": source_sheet,
                            "source": clean_part(_cell_by_header(sheet, row_index, headers, SOURCE_HEADERS, 3)),
                            "cert_number": cert,
                            "grader": normalize_grader(_cell_by_header(sheet, row_index, headers, GRADER_HEADERS, 5)),
                            "card_title": card,
                            "purchase_price": purchase,
                            "sale_price": sale,
                            "card_ladder_value": parse_money(_cell_by_header(sheet, row_index, headers, CARD_LADDER_VALUE_HEADERS, 8)),
                            "comps": parse_money(_cell_by_header(sheet, row_index, headers, COMPS_AVERAGE_HEADERS, 9)),
                            "cy_estimate": parse_money(_cell_by_header(sheet, row_index, headers, CY_VALUE_HEADERS, 10)),
                            "cy_value": parse_money(_cell_by_header(sheet, row_index, headers, CY_VALUE_HEADERS, 10)),
                            "cy_confidence": _cell_by_header(sheet, row_index, headers, CY_CONFIDENCE_HEADERS, 11),
                            "best_company": clean_part(_cell_by_header(sheet, row_index, headers, BEST_COMPANY_HEADERS, 12)) or company,
                            "status": clean_part(_cell_by_header(sheet, row_index, headers, STATUS_HEADERS, 14)),
                            "notes": clean_part(_cell_by_header(sheet, row_index, headers, NOTES_HEADERS, 15)),
                        }
                    )
        finally:
            workbook.close()
    return records


def write_working_sheet(path: Path, rows: list[Any], source_lookup: dict[int, str] | None = None) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = DEFAULT_SHEET
    headers = [
        "Certification Number",
        "Company",
        "Sport",
        "Card Description",
        "Purchase Price",
        "Card Ladder Value",
        "Comps",
        "CY Estimate",
        "CY Confidence",
        "Best Company",
        "Estimated Payout",
        "Status",
        "Source",
        RECEIVED_HEADER,
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([
            row.cert_number,
            row.grader,
            getattr(row, "category", ""),
            row.card_title,
            row.existing_value,
            row.card_ladder_value,
            row.card_ladder_comps_average,
            row.cy_value,
            row.cy_confidence,
            getattr(row, "best_company", ""),
            getattr(row, "estimated_payout", None),
            getattr(row, "status", ""),
            (source_lookup or {}).get(row.excel_row, ""),
            "X" if getattr(row, "received", False) else "",
        ])
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for letter, width in {"A": 22, "B": 14, "C": 14, "D": 62, "E": 16, "F": 18, "G": 14, "H": 14, "I": 16, "J": 18, "K": 18, "L": 18, "M": 38, "N": 14}.items():
        sheet.column_dimensions[letter].width = width
    workbook.save(path)
    return path


def mark_received_in_workbooks(
    paths: list[Path],
    certs: set[str],
    row_refs: set[tuple[str, str, int]] | None = None,
) -> dict[str, Any]:
    target_certs = {normalize_cert(cert) for cert in certs if normalize_cert(cert)}
    target_row_refs = {
        (str(sheet_file or "").strip().lower(), str(sheet_name or "").strip().lower(), int(row_index))
        for sheet_file, sheet_name, row_index in (row_refs or set())
        if str(sheet_file or "").strip() and str(sheet_name or "").strip() and int(row_index or 0) > 0
    }
    result = {
        "files_scanned": 0,
        "files_updated": 0,
        "rows_marked": 0,
        "certs_marked": set(),
        "row_refs_marked": set(),
        "errors": [],
    }
    if not target_certs and not target_row_refs:
        return result

    for path in paths:
        result["files_scanned"] += 1
        try:
            workbook = load_workbook(path)
        except Exception as error:
            result["errors"].append(f"{path.name}: {error}")
            continue

        changed = False
        try:
            for sheet in workbook.worksheets:
                cert_col = _cert_column(sheet)
                if not cert_col:
                    continue
                header_row = 1 if _looks_like_simple_header(sheet) else None
                received_col = _ensure_received_column(sheet, header_row)
                first_data_row = 2 if header_row else 1
                for row_index in range(first_data_row, _sheet_max_row(sheet) + 1):
                    cert = normalize_cert(sheet.cell(row_index, cert_col).value)
                    row_ref = (path.name.strip().lower(), sheet.title.strip().lower(), row_index)
                    if cert not in target_certs and row_ref not in target_row_refs:
                        continue
                    sheet.cell(row_index, received_col).value = "X"
                    for col_index in range(1, _sheet_max_column(sheet) + 1):
                        cell = sheet.cell(row_index, col_index)
                        cell.fill = RECEIVED_FILL
                        cell.font = RECEIVED_FONT
                    result["rows_marked"] += 1
                    if cert:
                        result["certs_marked"].add(cert)
                    if row_ref in target_row_refs:
                        result["row_refs_marked"].add(row_ref)
                    changed = True
            if changed:
                workbook.save(path)
                result["files_updated"] += 1
        except Exception as error:
            result["errors"].append(f"{path.name}: {error}")
        finally:
            workbook.close()
    return result


def clear_received_in_workbooks(paths: list[Path]) -> dict[str, Any]:
    result = {
        "files_scanned": 0,
        "files_updated": 0,
        "rows_cleared": 0,
        "errors": [],
    }
    for path in paths:
        result["files_scanned"] += 1
        try:
            workbook = load_workbook(path)
        except Exception as error:
            result["errors"].append(f"{path.name}: {error}")
            continue

        changed = False
        try:
            for sheet in workbook.worksheets:
                headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
                received_col = headers.get("received")
                if not received_col:
                    continue
                for row_index in range(2, _sheet_max_row(sheet) + 1):
                    received_cell = sheet.cell(row_index, received_col)
                    if str(received_cell.value or "").strip():
                        received_cell.value = ""
                        result["rows_cleared"] += 1
                        changed = True
                    for col_index in range(1, _sheet_max_column(sheet) + 1):
                        cell = sheet.cell(row_index, col_index)
                        if cell.fill == RECEIVED_FILL:
                            cell.fill = PatternFill(fill_type=None)
                            changed = True
            if changed:
                workbook.save(path)
                result["files_updated"] += 1
        except Exception as error:
            result["errors"].append(f"{path.name}: {error}")
        finally:
            workbook.close()
    return result


def remove_company_sheet_rows_for_source(directory: Path, source_sheet_name: str, cert_numbers: set[str] | None = None) -> dict[str, Any]:
    result = {
        "files_scanned": 0,
        "files_updated": 0,
        "rows_removed": 0,
        "errors": [],
    }
    source_name = Path(str(source_sheet_name or "")).name
    if not source_name or not directory.exists():
        return result
    target_certs = {normalize_cert(cert) for cert in (cert_numbers or set()) if normalize_cert(cert)}
    for path in sorted(directory.glob("*/*.xlsx")):
        result["files_scanned"] += 1
        try:
            workbook = load_workbook(path)
        except Exception as error:
            result["errors"].append(f"{path.name}: {error}")
            continue
        changed = False
        try:
            for sheet in workbook.worksheets:
                headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
                source_col = headers.get("sourcesheet")
                cert_col = headers.get("certificationnumber") or headers.get("certnumber") or headers.get("cert")
                if not source_col:
                    continue
                for row_index in range(_sheet_max_row(sheet), 1, -1):
                    row_source = Path(clean_part(sheet.cell(row_index, source_col).value)).name
                    row_cert = normalize_cert(sheet.cell(row_index, cert_col).value) if cert_col else ""
                    if row_source == source_name and (not target_certs or row_cert in target_certs):
                        sheet.delete_rows(row_index, 1)
                        result["rows_removed"] += 1
                        changed = True
                if changed:
                    sheet.auto_filter.ref = sheet.dimensions
            if changed:
                workbook.save(path)
                result["files_updated"] += 1
        except Exception as error:
            result["errors"].append(f"{path.name}: {error}")
        finally:
            workbook.close()
    return result


def working_sheet_path(directory: Path, title: str) -> Path:
    safe = safe_filename(title) or time.strftime("working-sheet-%Y%m%d-%H%M%S")
    return directory / f"{safe}.xlsx"


def safe_filename(value: str) -> str:
    text = re.sub(r"[<>:\"/\\|?*]+", " ", str(value or "")).strip()
    text = re.sub(r"\s+", " ", text)
    return text[:140].strip()


def default_output_path(root: Path) -> Path:
    stamp = time.strftime("%Y%m%d-%H%M%S")
    return root / "outputs" / f"card-pipeline-comps-{stamp}.xlsx"


def scan_to_cert(value: Any) -> str:
    text = str(value or "")
    candidates = re.findall(r"\d{6,12}", text)
    if candidates:
        return max(candidates, key=len)
    return normalize_cert(text)


def normalize_cert(value: Any) -> str:
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"[^0-9A-Z]", "", str(value or ""), flags=re.I)


def infer_grader(card_title: str) -> str:
    match = re.search(r"\b(PSA|BGS|SGC|CGC|BECKETT)\b", str(card_title or ""), re.I)
    if not match:
        return ""
    return normalize_grader(match.group(1))


def normalize_grader(value: Any) -> str:
    text = clean_part(value).upper()
    aliases = {"BECKETT": "BGS", "BVG": "BGS", "PSA": "PSA", "BGS": "BGS", "SGC": "SGC", "CGC": "CGC"}
    return aliases.get(text, "")


def parse_money(value: Any) -> float | None:
    if value is None or value == "":
        return None
    match = re.search(r"[\d,]+(?:\.\d{1,2})?", str(value))
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def format_money(value: float | None) -> str:
    return "" if value is None else f"${value:,.2f}"


def build_card_title(row: dict[str, Any]) -> str:
    description = clean_part(row.get("description", ""))
    grade = clean_grade(row.get("grade", ""))
    if description:
        parts = [description]
    else:
        parts = [
            clean_part(row.get("year")),
            clean_part(row.get("set")),
            clean_part(row.get("player")),
            _card_number_part(row.get("card_number")),
            clean_part(row.get("parallel")),
            clean_part(row.get("subset")),
        ]
    title = " ".join(part for part in parts if part)
    if grade and title and not re.search(rf"(?<!\d){re.escape(grade)}(?!\d)", title):
        parts.append(grade)
    return re.sub(r"\s+", " ", " ".join(part for part in parts if part)).strip()


def clean_grade(value: Any) -> str:
    numbers = re.findall(r"\d+(?:\.\d+)?", str(value or ""))
    return numbers[-1] if numbers else ""


def clean_part(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _looks_like_simple_header(sheet) -> bool:
    return _simple_header_row(sheet) is not None


def _simple_header_row(sheet) -> int | None:
    best_row = None
    best_score = 0
    for row_index in range(1, min(_sheet_max_row(sheet), 10) + 1):
        headers = _header_map_for_row(sheet, row_index)
        score = _simple_header_score(headers)
        if score > best_score:
            best_row = row_index
            best_score = score
    if best_row and best_score >= 2:
        return best_row
    if best_row and best_score >= 1:
        first = " ".join(clean_part(sheet.cell(best_row, col).value).lower() for col in range(1, min(_sheet_max_column(sheet), 8) + 1))
        if any(token in first for token in ("cert", "card", "description", "purchase", "price", "comp", "ladder", "date")):
            return best_row
    first = " ".join(clean_part(sheet.cell(1, col).value).lower() for col in range(1, min(_sheet_max_column(sheet), 8) + 1))
    return any(token in first for token in ("cert", "card", "description", "purchase", "price", "comp", "ladder"))


def _simple_header_score(headers: dict[str, int]) -> int:
    groups = (
        CERT_HEADERS,
        GRADER_HEADERS,
        CARD_HEADERS,
        SPORT_HEADERS,
        DATE_HEADERS,
        PURCHASE_PRICE_HEADERS,
        CARD_LADDER_VALUE_HEADERS,
        COMPS_AVERAGE_HEADERS,
        CY_VALUE_HEADERS,
        CY_CONFIDENCE_HEADERS,
        BEST_COMPANY_HEADERS,
        ESTIMATED_PAYOUT_HEADERS,
        STATUS_HEADERS,
        NOTES_HEADERS,
        SOURCE_HEADERS,
    )
    return sum(1 for aliases in groups if any(alias in headers for alias in aliases))


def _cert_column(sheet) -> int | None:
    headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
    for alias in CERT_HEADERS:
        if alias in headers:
            return headers[alias]
    if _looks_like_simple_header(sheet):
        return None
    return 1 if _sheet_max_column(sheet) >= 1 else None


def _card_column(sheet) -> int | None:
    headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
    for alias in CARD_HEADERS:
        if alias in headers:
            return headers[alias]
    return None


def _price_column(sheet) -> int | None:
    headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
    for alias in PURCHASE_PRICE_HEADERS:
        if alias in headers:
            return headers[alias]
    return None


def _received_column(sheet) -> int | None:
    headers = _header_map_for_row(sheet, 1) if _sheet_max_row(sheet) else {}
    return headers.get(_normalize_header(RECEIVED_HEADER))


def _ensure_received_column(sheet, header_row: int | None) -> int:
    headers = _header_map_for_row(sheet, header_row or 1) if header_row else {}
    existing = headers.get(_normalize_header(RECEIVED_HEADER))
    if existing:
        return existing
    col = _sheet_max_column(sheet) + 1
    if header_row:
        header_cell = sheet.cell(header_row, col)
        header_cell.value = RECEIVED_HEADER
        header_cell.fill = PatternFill("solid", fgColor="111827")
        header_cell.font = Font(color="FFFFFF", bold=True)
        sheet.column_dimensions[header_cell.column_letter].width = 14
        sheet.auto_filter.ref = sheet.dimensions
    return col


def _is_received_value(value: Any) -> bool:
    text = clean_part(value).upper()
    return text in {"X", "Y", "YES", "TRUE", "1", "RECEIVED"}


def _setup_notes(cert: str, card: str, grader: str) -> str:
    notes = []
    if not cert:
        notes.append("Missing cert")
    if not card:
        notes.append("Missing card description")
    if not grader:
        notes.append("Missing grader")
    return "; ".join(notes)


def _card_number_part(value: Any) -> str:
    text = clean_part(value)
    if not text:
        return ""
    return text if text.startswith("#") else f"#{text}"


def _detect_photo_schema(sheet) -> dict[str, Any]:
    best_row = None
    best_headers: dict[str, int] = {}
    best_score = 0
    for row_index in range(1, min(_sheet_max_row(sheet), 10) + 1):
        headers = _header_map_for_row(sheet, row_index)
        score = _header_score(headers)
        if score > best_score:
            best_row = row_index
            best_headers = headers
            best_score = score
    if best_row and best_score >= 3:
        return {"headers": best_headers, "first_data_row": best_row + 1}
    return {"headers": {}, "first_data_row": 1}


def _header_score(headers: dict[str, int]) -> int:
    return sum(1 for aliases in PHOTO_EXPORT_HEADERS.values() if any(alias in headers for alias in aliases))


def _header_map_for_row(sheet, row_index: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, _sheet_max_column(sheet) + 1):
        value = sheet.cell(row_index, col).value
        if value:
            headers[_normalize_header(value)] = col
    return headers


def _sheet_max_row(sheet) -> int:
    row_count, _column_count = _sheet_bounds(sheet)
    return row_count


def _sheet_max_column(sheet) -> int:
    _row_count, column_count = _sheet_bounds(sheet)
    return column_count


def _sheet_bounds(sheet) -> tuple[int, int]:
    max_row = getattr(sheet, "max_row", None)
    max_column = getattr(sheet, "max_column", None)
    row_count = max_row if isinstance(max_row, int) and max_row >= 1 else 0
    column_count = max_column if isinstance(max_column, int) and max_column >= 1 else 0
    if row_count and column_count:
        return row_count, column_count

    dimension = ""
    for kwargs in ({}, {"force": True}):
        try:
            dimension = str(sheet.calculate_dimension(**kwargs) or "")
            break
        except TypeError:
            continue
        except Exception:
            dimension = ""
            break
    if not dimension:
        try:
            reset = getattr(sheet, "reset_dimensions", None)
            if callable(reset):
                reset()
                dimension = str(sheet.calculate_dimension(force=True) or "")
        except Exception:
            dimension = ""
    if dimension:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(dimension)
            row_count = row_count or max(max_row, min_row)
            column_count = column_count or max(max_col, min_col)
        except Exception:
            pass
    return row_count, column_count


def _source_row(sheet, row_index: int, headers: dict[str, int]) -> dict[str, Any]:
    return {
        key: _cell(sheet, row_index, headers, aliases, PHOTO_EXPORT_POSITIONS.get(key))
        for key, aliases in PHOTO_EXPORT_HEADERS.items()
    }


def _cell(sheet, row_index: int, headers: dict[str, int], aliases: tuple[str, ...], fallback_col: int | None) -> Any:
    for alias in aliases:
        col = headers.get(alias)
        if _valid_column(col):
            return sheet.cell(row_index, col).value
    return sheet.cell(row_index, fallback_col).value if _valid_column(fallback_col) else ""


def _cell_by_header(sheet, row_index: int, headers: dict[str, int], aliases: tuple[str, ...], fallback_col: int | None) -> Any:
    for alias in aliases:
        col = headers.get(alias)
        if _valid_column(col):
            return sheet.cell(row_index, col).value
    return sheet.cell(row_index, fallback_col).value if _valid_column(fallback_col) else ""


def _valid_column(value: Any) -> bool:
    return isinstance(value, int) and value >= 1


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())
