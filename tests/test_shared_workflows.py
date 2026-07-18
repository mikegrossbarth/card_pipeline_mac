from __future__ import annotations

import base64
import json
import os
import queue
import socket
import sys
import threading
import time
import types
import unittest
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
PHOTO_APP = ROOT / "photo_tool" / "app"
if str(PHOTO_APP) not in sys.path:
    sys.path.insert(0, str(PHOTO_APP))

import app
import assignment_config_ui
import assignment_engine
import bridge_server
import cardladder_ocr
import google_sheets_import
import lucas_diagnostics
from comp_engine.workbook_io import WorkbookRow
from intake_io import append_company_sheet_rows, company_weekly_sheet_name, ensure_company_weekly_sheets, mark_received_in_workbooks, parse_money as intake_parse_money, scan_to_cert, read_company_profit_records, read_simple_spreadsheet, write_pipeline_output, write_working_sheet
from shared_state import atomic_write_json, local_identity, read_json, shared_lock


if "google" not in sys.modules:
    google_module = types.ModuleType("google")
    genai_module = types.ModuleType("google.genai")
    genai_types_module = types.ModuleType("google.genai.types")

    class _Client:
        pass

    class _Part:
        @staticmethod
        def from_bytes(data=None, mime_type=None):
            return {"data": data, "mime_type": mime_type}

    class _GenerateContentConfig:
        def __init__(self, **kwargs):
            self.kwargs = kwargs

    class _ThinkingConfig:
        def __init__(self, **kwargs):
            self.kwargs = kwargs

    genai_module.Client = _Client
    genai_types_module.Part = _Part
    genai_types_module.GenerateContentConfig = _GenerateContentConfig
    genai_types_module.ThinkingConfig = _ThinkingConfig
    genai_module.types = genai_types_module
    google_module.genai = genai_module
    sys.modules["google"] = google_module
    sys.modules["google.genai"] = genai_module
    sys.modules["google.genai.types"] = genai_types_module

import multi_card_extraction


class SharedStateTests(unittest.TestCase):
    def test_card_ladder_money_parsers_treat_three_digit_decimal_and_k_suffix_as_thousands(self) -> None:
        class MoneyDummy:
            _money_value = app.CardPipelineApp._money_value

        self.assertEqual(bridge_server.parse_value("$15.920"), 15920.0)
        self.assertEqual(cardladder_ocr.parse_money("$15.920"), 15920.0)
        self.assertEqual(bridge_server.parse_value("$20.27k"), 20270.0)
        self.assertEqual(cardladder_ocr.parse_money("$20.27k"), 20270.0)
        self.assertEqual(MoneyDummy()._money_value("$20.27k"), 20270.0)
        self.assertEqual(intake_parse_money("$20.27k"), 20270.0)
        self.assertEqual(bridge_server.parse_value("$20.27"), 20.27)
        self.assertEqual(cardladder_ocr.parse_money("$20.27"), 20.27)
        self.assertEqual(MoneyDummy()._money_value("$20.27"), 20.27)

    def test_scan_to_cert_preserves_long_psa_cert_numbers(self) -> None:
        self.assertEqual(scan_to_cert("1401017991290"), "1401017991290")
        self.assertEqual(scan_to_cert("PSA Cert 1401017991290"), "1401017991290")

    def test_card_ladder_profile_title_strips_trailing_close_ui_text(self) -> None:
        raw_titles = [
            "2024 Panini Prizm Silver Caitlin Clark Close",
            "2024 Panini Prizm Silver Caitlin Clark Close PSA 10",
            "2024 Panini Prizm Silver Caitlin Clark Close $125.00",
            "2024 Panini Prizm Silver Caitlin Clark Close search_off",
        ]
        for cleaner in (bridge_server.clean_profile_title, cardladder_ocr.clean_profile_title):
            for raw in raw_titles:
                with self.subTest(cleaner=getattr(cleaner, "__module__", ""), raw=raw):
                    self.assertEqual(cleaner(raw), "2024 Panini Prizm Silver Caitlin Clark")

    def test_card_ladder_generic_short_profile_title_requires_review(self) -> None:
        reason = bridge_server.generic_profile_review_reason(
            "2024 Topps",
            "PSA",
            "9",
            {"resultCount": 13},
        )
        self.assertIn("overly broad profile title", reason)
        self.assertIn("2024 Topps PSA 9", reason)

    def test_card_ladder_generic_short_profile_title_is_not_saved_to_row(self) -> None:
        row = WorkbookRow(excel_row=7, cert_number="148874718", card_title="", grader="PSA")
        bridge_server.BridgeState()._apply_cardladder_result_to_row(
            row,
            {
                "status": "ok",
                "value": 60,
                "ocr": {
                    "profileTitle": "2024 Topps",
                    "profileGrader": "PSA",
                    "profileGrade": "9",
                    "resultCount": 13,
                    "comps": [{"title": "Shohei Ohtani 2024 Topps Heritage PSA 9", "price": "$60"}],
                },
            },
        )
        self.assertEqual(row.card_title, "")
        self.assertIsNone(row.card_ladder_value)
        self.assertEqual(row.status, "Card Ladder review")
        self.assertIn("overly broad profile title", row.notes)

    def test_google_ssl_context_uses_certifi_when_no_cert_env_is_set(self) -> None:
        with TemporaryDirectory() as tmp:
            cafile = Path(tmp) / "cacert.pem"
            cafile.write_text("", encoding="utf-8")
            fake_certifi = types.SimpleNamespace(where=lambda: str(cafile))
            old_context = google_sheets_import._SSL_CONTEXT
            old_diagnostics = dict(google_sheets_import.LAST_OAUTH_DIAGNOSTICS)
            old_ssl_cert = os.environ.pop("SSL_CERT_FILE", None)
            old_requests_bundle = os.environ.pop("REQUESTS_CA_BUNDLE", None)
            try:
                google_sheets_import._SSL_CONTEXT = None
                sentinel_context = object()
                with patch.dict(sys.modules, {"certifi": fake_certifi}), patch(
                    "google_sheets_import.ssl.create_default_context",
                    return_value=sentinel_context,
                ) as create_context:
                    context = google_sheets_import.google_ssl_context()
                self.assertIs(context, sentinel_context)
                create_context.assert_called_once_with(cafile=str(cafile))
                self.assertEqual(os.environ.get("SSL_CERT_FILE"), str(cafile))
                self.assertEqual(os.environ.get("REQUESTS_CA_BUNDLE"), str(cafile))
                self.assertEqual(google_sheets_import.LAST_OAUTH_DIAGNOSTICS["ssl_cert_file"], str(cafile))
            finally:
                google_sheets_import._SSL_CONTEXT = old_context
                google_sheets_import.LAST_OAUTH_DIAGNOSTICS.clear()
                google_sheets_import.LAST_OAUTH_DIAGNOSTICS.update(old_diagnostics)
                if old_ssl_cert is not None:
                    os.environ["SSL_CERT_FILE"] = old_ssl_cert
                else:
                    os.environ.pop("SSL_CERT_FILE", None)
                if old_requests_bundle is not None:
                    os.environ["REQUESTS_CA_BUNDLE"] = old_requests_bundle
                else:
                    os.environ.pop("REQUESTS_CA_BUNDLE", None)

    def test_shared_lock_serializes_concurrent_writers(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            events: list[tuple[str, str, float]] = []

            def worker(name: str, delay: float) -> None:
                with shared_lock(root, "same-file", {"display_name": name, "machine": name}):
                    events.append((name, "enter", time.time()))
                    time.sleep(delay)
                    events.append((name, "exit", time.time()))

            first = threading.Thread(target=worker, args=("A", 0.25))
            second = threading.Thread(target=worker, args=("B", 0.01))
            first.start()
            time.sleep(0.05)
            second.start()
            first.join()
            second.join()

            self.assertEqual([event[:2] for event in events], [("A", "enter"), ("A", "exit"), ("B", "enter"), ("B", "exit")])

    def test_atomic_json_write_and_local_identity(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            settings_path = root / "lucas_settings.json"
            identity = local_identity(settings_path)
            self.assertTrue(identity["user_id"])
            self.assertTrue(identity["machine"])

            state_path = root / "state.json"
            atomic_write_json(state_path, {"a": 1})
            atomic_write_json(state_path, {"b": [1, 2, 3]})
            self.assertEqual(read_json(state_path, {}), {"b": [1, 2, 3]})
            self.assertFalse(list(root.glob("*.tmp")))

    def test_personal_lucas_marker_edit_forces_mikey(self) -> None:
        class StatusVar:
            def set(self, value):
                self.value = value

        class PersonalMarkerDummy:
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _marker_for_stage = app.CardPipelineApp._marker_for_stage
            save_home_sheet_markers = app.CardPipelineApp.save_home_sheet_markers
            _personal_default_person = app.CardPipelineApp._personal_default_person

            def _is_personal_lucas(self):
                return True

            def _save_sheet_markers(self):
                self.saved_markers = True

            def _delete_sheet_marker(self, _key):
                pass

            def _move_sheet_to_received(self, _key):
                return ""

            def _move_working_sheet_to_incoming(self, _key):
                return ""

            def _retarget_inventory_rows_for_source(self, _source_sheet_name, _assigned_person):
                self.retargeted_inventory_person = _assigned_person
                return 0

            def _retarget_profit_rows_for_source(self, _source_sheet_name, _assigned_person):
                self.retargeted_profit_person = _assigned_person
                return 0

            def _sync_received_sheet_inventory_to_ledger(self, *_args):
                return 0, 0

            def refresh_working_sheets(self):
                self.refresh_working_calls = getattr(self, "refresh_working_calls", 0) + 1

            def refresh_received_sheets(self):
                self.refresh_received_calls = getattr(self, "refresh_received_calls", 0) + 1

            def refresh_incoming_index(self):
                self.refresh_incoming_index_calls = getattr(self, "refresh_incoming_index_calls", 0) + 1

            def refresh_home(self):
                self.refresh_home_calls = getattr(self, "refresh_home_calls", 0) + 1

        dummy = PersonalMarkerDummy()
        dummy.home_selected_sheet_key = "Incoming|Personal Lot.xlsx"
        dummy.home_sheet_markers = {"Incoming|Personal Lot.xlsx": {"assigned_person": "Kevin Hambone"}}
        dummy.home_sheet_summaries = {}
        dummy.home_sheet_kind = type("Kind", (), {"set": lambda self, _value: None})()
        dummy.lucas_identity = {"display_name": "Tester"}
        dummy.status_var = StatusVar()
        dummy.saved_markers = False

        with patch.object(app, "shared_lock", lambda *_args, **_kwargs: __import__("contextlib").nullcontext()):
            dummy.save_home_sheet_markers(
                {
                    "incoming_proper": True,
                    "tracking_number": "",
                    "all_received": False,
                    "assigned_person": "Kevin Hambone",
                }
            )

        marker = dummy.home_sheet_markers["Incoming|Personal Lot.xlsx"]
        self.assertTrue(dummy.saved_markers)
        self.assertEqual(marker["assigned_person"], "Mikey")
        self.assertEqual(dummy.retargeted_inventory_person, "Mikey")
        self.assertEqual(dummy.retargeted_profit_person, "Mikey")
        self.assertEqual(getattr(dummy, "refresh_home_calls", 0), 0)
        self.assertEqual(getattr(dummy, "refresh_incoming_index_calls", 0), 0)
        self.assertEqual(getattr(dummy, "refresh_working_calls", 0), 0)
        self.assertEqual(getattr(dummy, "refresh_received_calls", 0), 0)

    def test_personal_lucas_right_click_move_forces_mikey(self) -> None:
        class PersonalMoveDummy:
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _sheet_path_for_stage = app.CardPipelineApp._sheet_path_for_stage
            _marker_for_stage = app.CardPipelineApp._marker_for_stage
            _move_home_sheet_to_stage = app.CardPipelineApp._move_home_sheet_to_stage
            _personal_default_person = app.CardPipelineApp._personal_default_person

            def _is_personal_lucas(self):
                return True

            def _delete_sheet_marker(self, key):
                self.home_sheet_markers.pop(key, None)

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            incoming_dir = root / "INCOMING SHEETS"
            working_dir = root / "WORKING SHEETS"
            received_dir = root / "RECEIVED SHEETS"
            incoming_dir.mkdir(parents=True)
            working_dir.mkdir()
            received_dir.mkdir()
            (working_dir / "Personal Lot.xlsx").write_text("placeholder", encoding="utf-8")

            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_received = app.RECEIVED_SHEETS_DIR
            app.INCOMING_SHEETS_DIR = incoming_dir
            app.WORKING_SHEETS_DIR = working_dir
            app.RECEIVED_SHEETS_DIR = received_dir
            dummy = PersonalMoveDummy()
            dummy.home_sheet_markers = {"Working|Personal Lot.xlsx": {"assigned_person": "Kevin Hambone"}}
            dummy.home_sheet_paths = {"Incoming": {}, "Working": {"Personal Lot.xlsx": working_dir / "Personal Lot.xlsx"}, "Received": {}}
            try:
                moved_key, cleanup = dummy._move_home_sheet_to_stage("Working|Personal Lot.xlsx", "Incoming")

                self.assertEqual(moved_key, "Incoming|Personal Lot.xlsx")
                self.assertEqual(cleanup, {})
                self.assertEqual(dummy.home_sheet_markers[moved_key]["assigned_person"], "Mikey")
            finally:
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working
                app.RECEIVED_SHEETS_DIR = old_received

    def test_personal_lucas_profit_records_default_unassigned_to_mikey(self) -> None:
        class ProfitDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _general_sold_sheet_name = app.CardPipelineApp._general_sold_sheet_name
            _personal_default_person = app.CardPipelineApp._personal_default_person
            _owner_for_profile = app.CardPipelineApp._owner_for_profile

            def __init__(self, personal: bool):
                self.personal = personal

            def _is_personal_lucas(self):
                return self.personal

        personal = ProfitDummy(True)
        team = ProfitDummy(False)

        personal_record = personal._normalize_profit_record(
            {
                "assigned_person": "Kevin Hambone",
                "cert_number": "123",
                "company": "General Sold",
                "date_added": "2026-07-15",
                "weekly_sheet_name": "Inventory Sale",
                "source_sheet": "Unassigned General Sold",
                "purchase_price": 10,
                "sale_price": 15,
            }
        )
        personal_expense = personal._normalize_profit_record(
            {
                "record_type": "expense",
                "assigned_person": "Kevin Hambone",
                "date_added": "2026-07-15",
                "expense_type": "Shipping",
                "expense_amount": 4,
            }
        )
        team_record = team._normalize_profit_record(
            {
                "assigned_person": "Kevin Hambone",
                "cert_number": "123",
                "company": "General Sold",
                "date_added": "2026-07-15",
                "weekly_sheet_name": "Inventory Sale",
                "source_sheet": "Unassigned General Sold",
                "purchase_price": 10,
                "sale_price": 15,
            }
        )

        self.assertEqual(personal_record["assigned_person"], "Mikey")
        self.assertEqual(personal_record["source_sheet"], "Mikey General Sold")
        self.assertTrue(personal_record["ledger_key"].endswith("|mikey general sold"))
        self.assertEqual(personal_expense["assigned_person"], "Mikey")
        self.assertEqual(personal._general_sold_sheet_name("Kevin Hambone"), "Mikey General Sold")
        self.assertEqual(team_record["assigned_person"], "Kevin Hambone")
        self.assertEqual(team._general_sold_sheet_name(""), "Unassigned General Sold")

    def test_team_person_choices_must_already_exist(self) -> None:
        class PersonChoiceDummy:
            _canonical_person_choice = app.CardPipelineApp._canonical_person_choice
            _person_combo_values = app.CardPipelineApp._person_combo_values
            _personal_default_person = app.CardPipelineApp._personal_default_person

            def __init__(self, personal: bool = False):
                self.personal = personal

            def _is_personal_lucas(self):
                return self.personal

            def _known_people(self):
                return ["James Copeland", "Kevin Hambone"]

        team = PersonChoiceDummy()
        self.assertEqual(team._canonical_person_choice("james copeland"), "James Copeland")
        self.assertEqual(team._canonical_person_choice("", allow_blank=True), "")
        self.assertIsNone(team._canonical_person_choice("New Person"))
        self.assertIsNone(team._canonical_person_choice(""))
        self.assertEqual(team._person_combo_values(allow_blank=True), ["", "James Copeland", "Kevin Hambone"])

        personal = PersonChoiceDummy(personal=True)
        self.assertEqual(personal._canonical_person_choice("New Person"), "Mikey")

    def test_personal_lucas_inventory_and_assignment_default_to_mikey(self) -> None:
        class PersonalInventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_sport_from_value = app.CardPipelineApp._inventory_sport_from_value
            _assignment_person_for_row = app.CardPipelineApp._assignment_person_for_row
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _personal_default_person = app.CardPipelineApp._personal_default_person
            _owner_for_profile = app.CardPipelineApp._owner_for_profile

            def _is_personal_lucas(self):
                return True

        class TeamInventoryDummy(PersonalInventoryDummy):
            def _is_personal_lucas(self):
                return False

        personal = PersonalInventoryDummy()
        personal.home_sheet_markers = {"Incoming|Lot.xlsx": {"assigned_person": "Kevin Hambone"}}
        personal.selected_working_sheet = types.SimpleNamespace(get=lambda: "Lot.xlsx")
        team = TeamInventoryDummy()

        personal_record = personal._normalize_inventory_record(
            {"assigned_person": "Kevin Hambone", "cert_number": "123", "source_sheet": "Lot.xlsx", "card_title": "Test Card"}
        )
        team_record = team._normalize_inventory_record(
            {"assigned_person": "Kevin Hambone", "cert_number": "123", "source_sheet": "Lot.xlsx", "card_title": "Test Card"}
        )

        self.assertEqual(personal_record["assigned_person"], "Mikey")
        self.assertEqual(personal_record["inventory_key"], "123|lot.xlsx|mikey")
        self.assertEqual(personal._assignment_person_for_row(WorkbookRow(excel_row=2, cert_number="", card_title="", grader="")), "Mikey")
        self.assertEqual(team_record["assigned_person"], "Kevin Hambone")
        self.assertEqual(team_record["inventory_key"], "123|lot.xlsx|kevin hambone")

    def test_home_sheet_sort_modes(self) -> None:
        class SortVar:
            def __init__(self, value):
                self.value = value

            def get(self):
                return self.value

        class FakePath:
            def __init__(self, created):
                self.created = created

            def stat(self):
                return types.SimpleNamespace(st_ctime=self.created)

        class SortDummy:
            _sorted_home_sheet_names = app.CardPipelineApp._sorted_home_sheet_names

        dummy = SortDummy()
        dummy.home_sheet_paths = {
            "Incoming": {
                "zeta.xlsx": FakePath(10),
                "alpha.xlsx": FakePath(30),
                "middle.xlsx": FakePath(20),
            }
        }
        dummy.home_sheet_sort_var = SortVar("Name")
        self.assertEqual(dummy._sorted_home_sheet_names("Incoming", ["zeta.xlsx", "alpha.xlsx", "middle.xlsx"]), ["alpha.xlsx", "middle.xlsx", "zeta.xlsx"])

        dummy.home_sheet_sort_var = SortVar("Date Created")
        self.assertEqual(dummy._sorted_home_sheet_names("Incoming", ["zeta.xlsx", "alpha.xlsx", "middle.xlsx"]), ["alpha.xlsx", "middle.xlsx", "zeta.xlsx"])


class WorkbookCompanyProfitTests(unittest.TestCase):
    def test_company_sheet_week_start_rolls_forward_monday_at_8pm(self) -> None:
        self.assertEqual(app.company_sheet_week_start_for_time(datetime(2026, 6, 13, 23, 59)).isoformat(), "2026-06-08")
        self.assertEqual(app.company_sheet_week_start_for_time(datetime(2026, 6, 14, 0, 0)).isoformat(), "2026-06-08")
        self.assertEqual(app.company_sheet_week_start_for_time(datetime(2026, 6, 14, 23, 59)).isoformat(), "2026-06-08")
        self.assertEqual(app.company_sheet_week_start_for_time(datetime(2026, 6, 15, 19, 59)).isoformat(), "2026-06-08")
        self.assertEqual(app.company_sheet_week_start_for_time(datetime(2026, 6, 15, 20, 0)).isoformat(), "2026-06-15")
        self.assertEqual(app.company_sheet_week_start_for_time(datetime(2026, 6, 16, 8, 0)).isoformat(), "2026-06-15")
        self.assertEqual(company_weekly_sheet_name(datetime(2026, 6, 15, 19, 59)), "Week of 2026-06-08")
        self.assertEqual(company_weekly_sheet_name(datetime(2026, 6, 15, 20, 0)), "Week of 2026-06-15")

    def test_ensure_company_weekly_sheets_creates_blank_company_tabs(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp) / "COMPANY SHEETS"

            result = ensure_company_weekly_sheets(root, ["Arena Club", "Fanatics"], app.company_sheet_week_start_for_time(datetime(2026, 6, 15, 20, 0)))
            repeat = ensure_company_weekly_sheets(root, ["Arena Club", "Fanatics"], app.company_sheet_week_start_for_time(datetime(2026, 6, 15, 20, 0)))

            self.assertEqual(len(result["created"]), 2)
            self.assertEqual(len(repeat["existing"]), 2)
            arena_path = root / "Arena Club" / "Arena Club.xlsx"
            self.assertTrue(arena_path.exists())
            workbook = load_workbook(arena_path, read_only=True, data_only=True)
            self.assertIn("Week of 2026-06-15", workbook.sheetnames)
            workbook.close()
            rows = read_simple_spreadsheet(arena_path)
            self.assertEqual(rows, [])

    def test_cy_sheet_estimate_and_confidence_are_preserved_separately_from_purchase(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "cy.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Grader", "Cert", "Description", "Grade", "Purchase", "Estimate", "Confidence"])
            sheet.append(["CGC", "1401045404276", "2022 Paradigm Trigger Unown VSTAR", "g10", 17.37, 19.30, 4])
            workbook.save(path)

            rows = read_simple_spreadsheet(path)

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["purchase_price"], 17.37)
            self.assertEqual(rows[0]["cy_value"], 19.30)
            self.assertEqual(rows[0]["cy_confidence"], 4)

    def test_sheet_loader_uses_headers_when_date_column_shifts_values(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "shifted.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Date", "Cert", "Grader", "Card", "Comp Value", "Purchase Price", "CL Value"])
            sheet.append(["2026-06-24", "137915162", "PSA", "Test Card PSA 10", 120, 45, 150])
            workbook.save(path)

            rows = read_simple_spreadsheet(path)

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["cert_number"], "137915162")
            self.assertEqual(rows[0]["purchase_price"], 45)
            self.assertEqual(rows[0]["card_ladder_comps_average"], 120)
            self.assertEqual(rows[0]["card_ladder_value"], 150)
            self.assertEqual(rows[0]["date_added"], "2026-06-24")

    def test_sheet_loader_detects_header_row_below_title_rows(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "title-row.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Network mode test upload"])
            sheet.append(["Date", "Cert", "Grader", "Card", "Comps", "Purchase"])
            sheet.append(["2026-06-24", "22222222", "PSA", "Another Test Card PSA 9", 80, 25])
            workbook.save(path)

            rows = read_simple_spreadsheet(path)

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["purchase_price"], 25)
            self.assertEqual(rows[0]["card_ladder_comps_average"], 80)

    def test_courtyard_weekly_sheet_uses_cy_ingest_format(self) -> None:
        with TemporaryDirectory() as tmp:
            company_dir = Path(tmp) / "COMPANY SHEETS"
            rows = [
                WorkbookRow(
                    excel_row=2,
                    cert_number="1401045404276",
                    grader="CGC",
                    category="pokemon",
                    card_title="2022 Paradigm Trigger Unown VSTAR CGC 10",
                    existing_value=17.37,
                    card_ladder_value=22,
                    card_ladder_comps_average=21,
                    cy_value=19.30,
                    cy_confidence=4,
                    best_company="CourtYard",
                    estimated_payout=18.33,
                    status="Received",
                )
            ]

            result = append_company_sheet_rows(company_dir, rows, {2: "cy.xlsx:2"}, {2: "cy.xlsx"})

            self.assertEqual(result["rows_added"], 1)
            path = company_dir / "CourtYard" / "CourtYard.xlsx"
            workbook = load_workbook(path, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            try:
                first_headers = [sheet.cell(1, column).value for column in range(1, 8)]
                self.assertEqual(first_headers, ["Grader", "Cert", "Description", "Grade", "Purchase", "Estimate", "Confidence"])
                self.assertTrue(sheet.column_dimensions["H"].hidden)
            finally:
                workbook.close()

            imported = read_simple_spreadsheet(path)
            self.assertEqual(imported[0]["grader"], "CGC")
            self.assertEqual(imported[0]["cert_number"], "1401045404276")
            self.assertEqual(imported[0]["purchase_price"], 17.37)
            self.assertEqual(imported[0]["cy_value"], 19.30)
            self.assertEqual(imported[0]["cy_confidence"], 4)
            profit_records = read_company_profit_records(company_dir)
            self.assertEqual(len(profit_records), 1)
            self.assertEqual(profit_records[0]["sale_price"], 18.33)
            self.assertEqual(profit_records[0]["source_sheet"], "cy.xlsx")

    def test_fanatics_weekly_sheet_uses_fanatics_front_columns(self) -> None:
        with TemporaryDirectory() as tmp:
            company_dir = Path(tmp) / "COMPANY SHEETS"
            rows = [
                WorkbookRow(
                    excel_row=2,
                    cert_number="153415486",
                    grader="PSA",
                    category="BASEBALL",
                    card_title="2012 Topps Pro Debut #164 Jose Ramirez PSA 10",
                    existing_value=120,
                    card_ladder_value=212,
                    card_ladder_comps_average=205,
                    best_company="Fanatics",
                    estimated_payout=197.16,
                    status="Received",
                )
            ]

            result = append_company_sheet_rows(company_dir, rows, {2: "fanatics.xlsx:2"}, {2: "fanatics.xlsx"})

            self.assertEqual(result["rows_added"], 1)
            path = company_dir / "Fanatics" / "Fanatics.xlsx"
            workbook = load_workbook(path, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            try:
                first_headers = [sheet.cell(1, column).value for column in range(1, 7)]
                first_values = [sheet.cell(2, column).value for column in range(1, 7)]
                self.assertEqual(first_headers, ["Category", "Card", "Grade", "Cert #", "CL Value", "Payout"])
                self.assertEqual(first_values, ["BASEBALL", "2012 Topps Pro Debut #164 Jose Ramirez PSA 10", "PSA 10", "153415486", 212, 197.16])
                self.assertEqual(sheet.cell(1, 8).value, "Source Sheet")
                self.assertEqual(sheet.cell(2, 8).value, "fanatics.xlsx")
            finally:
                workbook.close()

            imported = read_simple_spreadsheet(path)
            self.assertEqual(imported[0]["sport"], "BASEBALL")
            self.assertEqual(imported[0]["cert_number"], "153415486")
            self.assertEqual(imported[0]["card_ladder_value"], 212)
            self.assertEqual(imported[0]["estimated_payout"], 197.16)
            profit_records = read_company_profit_records(company_dir)
            self.assertEqual(len(profit_records), 1)
            self.assertEqual(profit_records[0]["purchase_price"], 120)
            self.assertEqual(profit_records[0]["sale_price"], 197.16)
            self.assertEqual(profit_records[0]["source_sheet"], "fanatics.xlsx")

    def test_existing_fanatics_sheet_migrates_to_front_column_format(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp) / "COMPANY SHEETS"
            path = root / "Fanatics" / "Fanatics.xlsx"
            path.parent.mkdir(parents=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Week of 2026-06-22"
            sheet.append(["Date Added", "Source Sheet", "Source", "Certification Number", "Grader", "Card Description", "Purchase Price", "Card Ladder Value", "Comps", "CY Estimate", "CY Confidence", "Best Company", "Estimated Payout", "Status", "Notes"])
            sheet.append(["2026-06-22", "Lot A.xlsx", "LUCAS", "153415486", "PSA", "2012 Topps Pro Debut #164 Jose Ramirez PSA 10", 120, 212, 205, "", "", "Fanatics", 197.16, "Received", ""])
            workbook.save(path)
            workbook.close()

            ensure_company_weekly_sheets(root, ["Fanatics"], date(2026, 6, 22))

            workbook = load_workbook(path, data_only=True)
            sheet = workbook["Week of 2026-06-22"]
            try:
                first_headers = [sheet.cell(1, column).value for column in range(1, 7)]
                first_values = [sheet.cell(2, column).value for column in range(1, 7)]
                self.assertEqual(first_headers, ["Category", "Card", "Grade", "Cert #", "CL Value", "Payout"])
                self.assertEqual(first_values, [None, "2012 Topps Pro Debut #164 Jose Ramirez PSA 10", "PSA 10", "153415486", 212, 197.16])
                self.assertEqual(sheet.cell(2, 8).value, "Lot A.xlsx")
                self.assertEqual(sheet.cell(2, 11).value, 120)
            finally:
                workbook.close()

    def test_working_sheet_round_trips_manual_sport(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "manual-sport.xlsx"
            rows = [
                WorkbookRow(
                    excel_row=2,
                    cert_number="0010355805",
                    grader="BGS",
                    category="baseball",
                    card_title="Generic Prospect Auto BGS 9.5",
                    existing_value=25,
                )
            ]

            write_working_sheet(path, rows)
            loaded = read_simple_spreadsheet(path)

            self.assertEqual(loaded[0]["sport"], "baseball")

    def test_working_sheet_round_trips_received_marker(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "received-marker.xlsx"
            rows = [
                WorkbookRow(
                    excel_row=2,
                    cert_number="11111111",
                    grader="PSA",
                    card_title="Received Card PSA 10",
                    received=True,
                ),
                WorkbookRow(
                    excel_row=3,
                    cert_number="22222222",
                    grader="PSA",
                    card_title="Not Received Card PSA 9",
                    received=False,
                ),
            ]

            write_working_sheet(path, rows)
            loaded = read_simple_spreadsheet(path)

            self.assertTrue(loaded[0]["received"])
            self.assertFalse(loaded[1]["received"])

    def test_receive_company_append_dedupes_and_profit_backfills(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            working_dir = root / "WORKING SHEETS"
            company_dir = root / "COMPANY SHEETS"
            working_dir.mkdir()

            source_rows = [
                WorkbookRow(excel_row=2, cert_number="11111111", grader="PSA", card_title="Card One PSA 10", existing_value=50),
                WorkbookRow(excel_row=3, cert_number="22222222", grader="PSA", card_title="Card Two PSA 9", existing_value=75),
            ]
            sheet_path = working_dir / "test.xlsx"
            write_working_sheet(sheet_path, source_rows)

            receive_result = mark_received_in_workbooks([sheet_path], {"11111111"})
            self.assertEqual(receive_result["rows_marked"], 1)
            self.assertIn("11111111", receive_result["certs_marked"])
            self.assertEqual(len(read_simple_spreadsheet(sheet_path)), 2)

            sold_row = WorkbookRow(
                excel_row=2,
                cert_number="11111111",
                grader="PSA",
                card_title="Card One PSA 10",
                existing_value=50,
                card_ladder_value=100,
                card_ladder_comps_average=100,
                cy_value=88,
                cy_confidence=4,
                best_company="Arena Club",
                estimated_payout=90,
                company_pile=True,
                status="Received",
            )
            first_append = append_company_sheet_rows(company_dir, [sold_row], {2: "test.xlsx:2"}, {2: "test.xlsx"})
            second_append = append_company_sheet_rows(company_dir, [sold_row], {2: "test.xlsx:2"}, {2: "test.xlsx"})

            self.assertEqual(first_append["rows_added"], 1)
            self.assertEqual(second_append["rows_added"], 0)
            self.assertEqual(len(first_append["added_records"]), 1)

            profit_records = read_company_profit_records(company_dir)
            self.assertEqual(len(profit_records), 1)
            self.assertEqual(profit_records[0]["cy_value"], 88.0)
            self.assertEqual(profit_records[0]["cy_confidence"], 4)
            self.assertEqual(profit_records[0]["purchase_price"], 50.0)
            self.assertEqual(profit_records[0]["sale_price"], 90.0)

    def test_manual_company_sheet_rows_without_source_sheet_do_not_backfill_profit(self) -> None:
        with TemporaryDirectory() as tmp:
            company_dir = Path(tmp) / "COMPANY SHEETS"
            path = company_dir / "Arena Club" / "Arena Club.xlsx"
            path.parent.mkdir(parents=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Week of 2026-06-22"
            sheet.append(["Date Added", "Source Sheet", "Source", "Certification Number", "Grader", "Card Description", "Purchase Price", "Card Ladder Value", "Comps", "CY Estimate", "CY Confidence", "Best Company", "Estimated Payout", "Status", "Notes"])
            sheet.append(["2026-06-22", "", "Manual", "999", "PSA", "Manual Card PSA 10", 10, "", 20, "", "", "Arena Club", 18, "Received", ""])
            sheet.append(["2026-06-22", "Lot A.xlsx", "LUCAS", "111", "PSA", "LUCAS Card PSA 10", 10, "", 20, "", "", "Arena Club", 18, "Received", ""])
            workbook.save(path)

            records = read_company_profit_records(company_dir)

            self.assertEqual(len(records), 1)
            self.assertEqual(records[0]["cert_number"], "111")

    def test_pipeline_output_round_trips_cy_value(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "comps.xlsx"
            rows = [
                WorkbookRow(
                    excel_row=2,
                    cert_number="11111111",
                    grader="PSA",
                    card_title="Card One PSA 10",
                    existing_value=50,
                    card_ladder_value=100,
                    card_ladder_comps_average=95,
                    cy_value=87.5,
                    cy_confidence=3,
                    status="Card Ladder OK",
                )
            ]

            write_pipeline_output(path, rows, {2: "source"})
            reloaded = read_simple_spreadsheet(path)

            self.assertEqual(reloaded[0]["cy_value"], 87.5)
            self.assertEqual(reloaded[0]["cy_confidence"], 3)


class CYLookupTests(unittest.TestCase):
    def test_cardladder_extension_error_is_recorded_on_row(self) -> None:
        state = bridge_server.BridgeState()
        row = WorkbookRow(excel_row=2, cert_number="156327815", grader="PSA", card_title="")
        result = {
            "excelRow": 2,
            "certNumber": "156327815",
            "grader": "PSA",
            "status": "extension_error",
            "error": "Card Ladder lookup failed after the cert opened.",
            "ocr": {"debugImage": ""},
            "extensionVersion": app.EXPECTED_CARDLADDER_EXTENSION_VERSION,
        }

        state._apply_cardladder_result_to_row(row, result)

        self.assertEqual(row.status, "Card Ladder extension error")
        self.assertIsNone(row.card_ladder_value)
        self.assertEqual(row.card_ladder_comps, "")
        self.assertIn("lookup failed", row.notes)

    def test_cardladder_partial_capture_preserves_value_and_profile(self) -> None:
        state = bridge_server.BridgeState()
        row = WorkbookRow(excel_row=2, cert_number="45039796", grader="PSA", card_title="")
        result = {
            "excelRow": 2,
            "certNumber": "45039796",
            "grader": "PSA",
            "status": "partial_comp_capture",
            "value": 550,
            "error": "Only captured 0 comp(s); expected 2. Re-run this row.",
            "ocr": {
                "profileTitle": "1965 Topps 350 Mickey Mantle",
                "profileGrader": "PSA",
                "profileGrade": "4.5",
                "resultCount": 12,
                "comps": [],
                "debugImage": "data:image/png;base64,debug",
            },
        }

        state._apply_cardladder_result_to_row(row, result)

        self.assertEqual(row.card_ladder_value, 550)
        self.assertEqual(row.card_title, "1965 Topps 350 Mickey Mantle PSA 4.5")
        self.assertEqual(row.card_ladder_comps, "")
        self.assertIsNone(row.card_ladder_comps_average)
        self.assertEqual(row.card_ladder_screenshot, "data:image/png;base64,debug")
        self.assertEqual(row.status, "Card Ladder partial capture")
        self.assertIn("Only captured 0 comp", row.notes)

    def test_cardladder_result_fills_blank_sport_from_profile_title(self) -> None:
        state = bridge_server.BridgeState()
        row = WorkbookRow(excel_row=2, cert_number="99505674", grader="PSA", card_title="")
        result = {
            "excelRow": 2,
            "certNumber": "99505674",
            "grader": "PSA",
            "status": "ok",
            "value": 73,
            "ocr": {
                "profileTitle": "2022 Panini Donruss 202 Chet Holmgren Yellow Holo Laser",
                "profileGrader": "PSA",
                "profileGrade": "9",
                "comps": [{"title": "2022 Panini Donruss 202 Chet Holmgren Yellow Holo Laser PSA 9", "date_sold": "Jun 1, 2026", "price": "$73.00"}],
            },
        }

        state._apply_cardladder_result_to_row(row, result)

        self.assertEqual(row.card_title, "2022 Panini Donruss 202 Chet Holmgren Yellow Holo Laser PSA 9")
        self.assertEqual(row.category, "basketball")

    def test_cardladder_result_does_not_overwrite_manual_sport(self) -> None:
        state = bridge_server.BridgeState()
        row = WorkbookRow(excel_row=2, cert_number="99505674", grader="PSA", card_title="", category="baseball")
        result = {
            "excelRow": 2,
            "certNumber": "99505674",
            "grader": "PSA",
            "status": "ok",
            "value": 73,
            "ocr": {
                "profileTitle": "2022 Panini Donruss 202 Chet Holmgren Yellow Holo Laser",
                "profileGrader": "PSA",
                "profileGrade": "9",
                "comps": [{"title": "2022 Panini Donruss 202 Chet Holmgren Yellow Holo Laser PSA 9", "date_sold": "Jun 1, 2026", "price": "$73.00"}],
            },
        }

        state._apply_cardladder_result_to_row(row, result)

        self.assertEqual(row.category, "baseball")

    def test_cardladder_only_result_does_not_trigger_cy_lookup_on_mac(self) -> None:
        state = bridge_server.BridgeState()
        row = WorkbookRow(
            excel_row=2,
            cert_number="11111111",
            grader="PSA",
            card_title="Card One PSA 10",
        )
        state.set_rows([row])

        with patch.object(bridge_server, "cy_lookup_enabled", return_value=True), \
                patch.object(bridge_server, "lookup_cy_buy_price", return_value=(87.5, 4, "ok")) as lookup:
            state.post_cardladder_result(
                {
                    "excelRow": 2,
                    "certNumber": "11111111",
                    "value": "$100",
                    "status": "ok",
                    "ocr": {"comps": [{"price": "$90", "date_sold": "2026-06-01"}]},
                }
            )
            lookup.assert_not_called()

        self.assertEqual(row.card_ladder_value, 100)
        self.assertIsNone(row.cy_value)
        self.assertIsNone(row.cy_confidence)
        self.assertEqual(row.status, "Card Ladder OK")
        self.assertNotIn("CY value:", row.notes)

    def test_cardladder_result_triggers_cy_lookup_when_cy_is_requested(self) -> None:
        state = bridge_server.BridgeState()
        row = WorkbookRow(
            excel_row=2,
            cert_number="11111111",
            grader="PSA",
            card_title="Card One PSA 10",
        )
        state.set_rows([row])
        state.cardladder_allows_cy = True

        with patch.object(bridge_server, "cy_lookup_enabled", return_value=True), \
                patch.object(bridge_server, "lookup_cy_buy_price", return_value=(87.5, 4, "ok")):
            state.post_cardladder_result(
                {
                    "excelRow": 2,
                    "certNumber": "11111111",
                    "value": "$100",
                    "status": "ok",
                    "ocr": {"comps": [{"price": "$90", "date_sold": "2026-06-01"}]},
                }
            )

            deadline = time.time() + 2
            while row.cy_value is None and time.time() < deadline:
                time.sleep(0.01)

        self.assertEqual(row.card_ladder_value, 100)
        self.assertEqual(row.cy_value, 87.5)
        self.assertEqual(row.cy_confidence, 4)
        self.assertEqual(row.status, "Card Ladder OK")
        self.assertIn("CY value: $87.50", row.notes)

    def test_cardladder_status_survives_cy_unavailable(self) -> None:
        state = bridge_server.BridgeState()
        row = WorkbookRow(
            excel_row=2,
            cert_number="11111111",
            grader="PSA",
            card_title="Card One PSA 10",
        )
        state.set_rows([row])
        state.cardladder_allows_cy = True

        with patch.object(bridge_server, "cy_lookup_enabled", return_value=True), \
                patch.object(bridge_server, "lookup_cy_buy_price", return_value=(None, "not found")):
            state.post_cardladder_result(
                {
                    "excelRow": 2,
                    "certNumber": "11111111",
                    "value": "$100",
                    "status": "ok",
                    "ocr": {"comps": [{"price": "$90", "date_sold": "2026-06-01"}]},
                }
            )

            deadline = time.time() + 2
            while "CY lookup:" not in row.notes and time.time() < deadline:
                time.sleep(0.01)

        self.assertEqual(row.card_ladder_value, 100)
        self.assertIsNone(row.cy_value)
        self.assertEqual(row.status, "Card Ladder OK")
        self.assertIn("CY lookup: not found", row.notes)

    def test_cy_lookup_waits_until_cardladder_finishes(self) -> None:
        state = bridge_server.BridgeState()
        row = WorkbookRow(
            excel_row=2,
            cert_number="11111111",
            grader="PSA",
            card_title="Card One PSA 10",
        )
        state.set_rows([row])
        state.cardladder_running = True
        state.cardladder_allows_cy = True

        with patch.object(bridge_server, "cy_lookup_enabled", return_value=True), \
                patch.object(bridge_server, "lookup_cy_buy_price", return_value=(87.5, "ok")) as lookup:
            state.post_cardladder_result(
                {
                    "excelRow": 2,
                    "certNumber": "11111111",
                    "value": "$100",
                    "status": "ok",
                    "ocr": {"comps": [{"price": "$90", "date_sold": "2026-06-01"}]},
                }
            )
            time.sleep(0.05)
            lookup.assert_not_called()
            self.assertIsNone(row.cy_value)

            state.finish_cardladder({})
            deadline = time.time() + 2
            while row.cy_value is None and time.time() < deadline:
                time.sleep(0.01)

        self.assertEqual(row.cy_value, 87.5)

    def test_stop_run_clears_pending_cy_lookup_after_cardladder(self) -> None:
        state = bridge_server.BridgeState()
        row = WorkbookRow(
            excel_row=2,
            cert_number="11111111",
            grader="PSA",
            card_title="Card One PSA 10",
        )
        state.set_rows([row])
        state.cardladder_running = True
        state.cardladder_allows_cy = True
        state.cy_lookup_generation += 1
        state.cy_lookup_pending.add(row.excel_row)
        row.status = "CY queued"

        state.request_cancel()
        state.finish_cardladder({})

        self.assertFalse(state.cy_lookup_pending)
        self.assertFalse(state.cy_lookup_inflight)
        self.assertIsNone(row.cy_value)
        self.assertEqual(row.status, "CY cancelled")

    def test_cy_only_batch_leaves_app_open_by_default_after_last_lookup(self) -> None:
        state = bridge_server.BridgeState()
        rows = [
            WorkbookRow(excel_row=2, cert_number="11111111", grader="PSA", card_title="Card One PSA 10"),
        ]
        state.set_rows(rows)

        with patch.object(bridge_server, "cy_lookup_enabled", return_value=True), \
                patch.object(bridge_server, "lookup_cy_buy_price", return_value=(87.5, 4, "ok")), \
                patch.object(bridge_server, "close_cy_adapter") as close_cy:
            state.start_cy_lookups(rows)
            deadline = time.time() + 2
            while any(row.cy_value is None for row in rows) and time.time() < deadline:
                time.sleep(0.01)

        self.assertEqual([row.cy_value for row in rows], [87.5])
        self.assertEqual(rows[0].cy_confidence, 4)
        self.assertEqual(rows[0].status, "CY OK")
        self.assertEqual(close_cy.call_count, 0)

    def test_cy_only_batch_can_close_app_after_last_lookup_when_enabled(self) -> None:
        state = bridge_server.BridgeState()
        rows = [
            WorkbookRow(excel_row=2, cert_number="11111111", grader="PSA", card_title="Card One PSA 10"),
        ]
        state.set_rows(rows)

        with patch.object(bridge_server, "cy_lookup_enabled", return_value=True), \
                patch.object(bridge_server, "cy_close_after_batch_enabled", return_value=True), \
                patch.object(bridge_server, "lookup_cy_buy_price", return_value=(87.5, 4, "ok")), \
                patch.object(bridge_server, "close_cy_adapter") as close_cy:
            state.start_cy_lookups(rows)
            deadline = time.time() + 2
            while any(row.cy_value is None for row in rows) and time.time() < deadline:
                time.sleep(0.01)
            while close_cy.call_count == 0 and time.time() < deadline:
                time.sleep(0.01)

        self.assertEqual([row.cy_value for row in rows], [87.5])
        self.assertGreaterEqual(close_cy.call_count, 1)

    def test_cy_only_unavailable_sets_cy_unavailable_status(self) -> None:
        state = bridge_server.BridgeState()
        rows = [
            WorkbookRow(excel_row=2, cert_number="11111111", grader="PSA", card_title="Card One PSA 10"),
        ]
        state.set_rows(rows)

        with patch.object(bridge_server, "cy_lookup_enabled", return_value=True), \
                patch.object(bridge_server, "lookup_cy_buy_price", return_value=(None, "not found")), \
                patch.object(bridge_server, "close_cy_adapter"):
            state.start_cy_lookups(rows)
            deadline = time.time() + 2
            while rows[0].status == "CY queued" and time.time() < deadline:
                time.sleep(0.01)

        self.assertIsNone(rows[0].cy_value)
        self.assertEqual(rows[0].status, "CY unavailable")
        self.assertIn("CY lookup: not found", rows[0].notes)

    def test_cy_gui_lookup_is_serialized(self) -> None:
        entered: list[str] = []
        active = 0
        max_active = 0
        lock = threading.Lock()

        class FakeAdapter:
            def submit_cert_lookup(self, cert_number: str, slab_type: str) -> dict:
                nonlocal active, max_active
                with lock:
                    entered.append(cert_number)
                    active += 1
                    max_active = max(max_active, active)
                time.sleep(0.05)
                with lock:
                    active -= 1
                return {"cy_buy_price": "87.50", "message": "ok"}

        old_lock = bridge_server._CY_LOOKUP_LOCK
        bridge_server._CY_LOOKUP_LOCK = threading.Lock()
        try:
            with patch.object(bridge_server, "get_cy_adapter", return_value=FakeAdapter()):
                threads = [
                    threading.Thread(target=bridge_server.lookup_cy_buy_price, args=(cert, "PSA"))
                    for cert in ("11111111", "22222222")
                ]
                for thread in threads:
                    thread.start()
                for thread in threads:
                    thread.join(timeout=2)
        finally:
            bridge_server._CY_LOOKUP_LOCK = old_lock

        self.assertEqual(set(entered), {"11111111", "22222222"})
        self.assertEqual(max_active, 1)

    def test_cy_lookup_is_disabled_off_mac(self) -> None:
        self.assertFalse(bridge_server.cy_lookup_enabled("win32"))
        self.assertFalse(bridge_server.cy_lookup_enabled("linux"))


class GoogleSheetCacheTests(unittest.TestCase):
    def test_authenticated_google_sheet_export_writes_xlsx_cache(self) -> None:
        def fake_tabs(_url: str, interactive: bool = False, sheet_name: str = ""):
            return [
                ("Rules/Main*?", [["Category", "Value"], ["Baseball", "100"]]),
                ("Payouts", [["CATEGORY", "YOUR PAYOUT %"], ["Baseball", "90%"]]),
            ]

        with TemporaryDirectory() as tmp, patch.object(google_sheets_import, "read_google_sheet_tabs", side_effect=fake_tabs):
            output_path = Path(tmp) / "cache.xlsx"
            google_sheets_import.export_google_sheet_to_xlsx("https://docs.google.com/spreadsheets/d/test/edit", output_path)

            from openpyxl import load_workbook

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Rules Main", "Payouts"])
                self.assertEqual(workbook["Payouts"].cell(2, 2).value, "90%")
            finally:
                workbook.close()

    def test_oauth_callback_ignores_unrelated_browser_requests(self) -> None:
        server = google_sheets_import.OAuthCallbackServer(("127.0.0.1", 0), google_sheets_import.OAuthCallbackHandler)
        server.timeout = 1
        server.expected_state = "state-123"
        base_url = f"http://127.0.0.1:{server.server_port}"

        def serve_once() -> None:
            server.handle_request()

        try:
            thread = threading.Thread(target=serve_once)
            thread.start()
            with urllib.request.urlopen(f"{base_url}/favicon.ico", timeout=5) as response:
                self.assertEqual(response.status, 204)
            thread.join(timeout=5)
            self.assertFalse(thread.is_alive())
            self.assertFalse(server.done)
            self.assertEqual(server.code, "")
            self.assertEqual(server.error, "")

            thread = threading.Thread(target=serve_once)
            thread.start()
            with urllib.request.urlopen(f"{base_url}/oauth2callback?state=state-123&code=abc123", timeout=5) as response:
                self.assertEqual(response.status, 200)
                body = response.read().decode("utf-8")
            thread.join(timeout=5)
            self.assertFalse(thread.is_alive())
            self.assertTrue(server.done)
            self.assertEqual(server.code, "abc123")
            self.assertEqual(server.error, "")
            self.assertIn("Google Sheets connected", body)
        finally:
            server.server_close()

    def test_startup_google_sheet_cache_refresh_discovers_and_exports_sources(self) -> None:
        class Dummy:
            _saved_google_sheet_sources = app.CardPipelineApp._saved_google_sheet_sources
            _google_sheet_cache_source = app.CardPipelineApp._google_sheet_cache_source
            _refresh_startup_google_sheet_caches = app.CardPipelineApp._refresh_startup_google_sheet_caches

        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_config = app.ASSIGNMENT_CONFIG_PATH
            app.CARD_PIPELINE_DIR = tmp_path / "CARD_PIPELINE"
            app.ASSIGNMENT_CONFIG_PATH = tmp_path / "assignment_companies.json"
            cache_path = app.CARD_PIPELINE_DIR / "ASSIGNMENT RULES" / "SHEET EXPORTS" / "rules.xlsx"
            app.ASSIGNMENT_CONFIG_PATH.write_text(
                json.dumps(
                    {
                        "companies": [
                            {
                                "name": "Test",
                                "rules": {
                                    "kind": "google_sheet",
                                    "url": "https://docs.google.com/spreadsheets/d/abc/edit",
                                    "path": str(cache_path),
                                    "name": "Rules",
                                },
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            calls: list[tuple[str, Path, bool]] = []

            def fake_export(url: str, output_path: Path, interactive: bool = False) -> Path:
                calls.append((url, Path(output_path), interactive))
                Path(output_path).parent.mkdir(parents=True, exist_ok=True)
                Path(output_path).write_bytes(b"fake")
                return Path(output_path)

            dummy = Dummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                with patch.object(app, "export_google_sheet_to_xlsx", side_effect=fake_export):
                    result = dummy._refresh_startup_google_sheet_caches()
                self.assertEqual(result["refreshed"], 1)
                self.assertEqual(result["errors"], [])
                self.assertEqual(calls[0][0], "https://docs.google.com/spreadsheets/d/abc/edit")
                self.assertFalse(calls[0][2])
                self.assertTrue(cache_path.exists())
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.ASSIGNMENT_CONFIG_PATH = old_config

    def test_startup_google_sheet_cache_materializes_url_sources(self) -> None:
        class Dummy:
            _saved_google_sheet_sources = app.CardPipelineApp._saved_google_sheet_sources
            _google_sheet_cache_source = app.CardPipelineApp._google_sheet_cache_source
            _refresh_startup_google_sheet_caches = app.CardPipelineApp._refresh_startup_google_sheet_caches

        with TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_config = app.ASSIGNMENT_CONFIG_PATH
            app.CARD_PIPELINE_DIR = tmp_path / "CARD_PIPELINE"
            app.ASSIGNMENT_CONFIG_PATH = tmp_path / "assignment_companies.json"
            app.ASSIGNMENT_CONFIG_PATH.write_text(
                json.dumps(
                    {
                        "companies": [
                            {
                                "name": "Rules URL",
                                "rules": "https://docs.google.com/spreadsheets/d/raw-url/edit",
                            },
                            {
                                "name": "Payout URL",
                                "payout": {
                                    "kind": "google_sheet",
                                    "url": "https://docs.google.com/spreadsheets/d/dict-url/edit",
                                    "name": "Payout URL",
                                },
                            },
                        ]
                    }
                ),
                encoding="utf-8",
            )
            calls: list[tuple[str, Path, bool]] = []

            def fake_export(url: str, output_path: Path, interactive: bool = False) -> Path:
                calls.append((url, Path(output_path), interactive))
                Path(output_path).parent.mkdir(parents=True, exist_ok=True)
                Path(output_path).write_bytes(b"fake")
                return Path(output_path)

            dummy = Dummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                with patch.object(app, "export_google_sheet_to_xlsx", side_effect=fake_export):
                    result = dummy._refresh_startup_google_sheet_caches()
                self.assertEqual(result["refreshed"], 2)
                self.assertEqual({call[0] for call in calls}, {
                    "https://docs.google.com/spreadsheets/d/raw-url/edit",
                    "https://docs.google.com/spreadsheets/d/dict-url/edit",
                })
                self.assertTrue(all(call[1].parent.name == "SHEET EXPORTS" for call in calls))
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.ASSIGNMENT_CONFIG_PATH = old_config

    def test_google_keep_source_reads_cached_note_text(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            cache_path = root / "ASSIGNMENT RULES" / "KEEP EXPORTS" / "note.txt"
            cache_path.parent.mkdir(parents=True)
            cache_path.write_text("Basketball $10-$100 90%\n", encoding="utf-8")

            text = assignment_engine.read_source_text(
                {
                    "kind": "google_keep",
                    "url": "https://keep.google.com/u/0/#NOTE/abc123",
                    "path": str(cache_path),
                    "name": "Rules",
                },
                root,
            )

            self.assertIn("Basketball $10-$100 90%", text)

    def test_saved_google_keep_source_uses_pipeline_cache_folder(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "ASSIGNMENT RULES" / "KEEP EXPORTS"
            source = app.CardPipelineApp._google_keep_cache_source(
                object(),
                "https://keep.google.com/u/0/#NOTE/abc123",
                output_dir,
            )

            self.assertIsNotNone(source)
            self.assertEqual(Path(source["path"]).parent, output_dir)

    def test_saved_google_keep_source_rehomes_repo_cache_path(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "CARD_PIPELINE" / "ASSIGNMENT RULES" / "KEEP EXPORTS"
            old_config = app.ASSIGNMENT_CONFIG_PATH
            app.ASSIGNMENT_CONFIG_PATH = root / "assignment_companies.json"
            try:
                stale_path = root / "ASSIGNMENT RULES" / "KEEP EXPORTS" / "Google Keep note.txt"
                source = app.CardPipelineApp._google_keep_cache_source(
                    object(),
                    {
                        "kind": "google_keep",
                        "url": "https://keep.google.com/u/0/#NOTE/abc123",
                        "path": str(stale_path),
                        "name": "Google Keep note",
                    },
                    output_dir,
                )
            finally:
                app.ASSIGNMENT_CONFIG_PATH = old_config

            self.assertIsNotNone(source)
            self.assertEqual(Path(source["path"]).parent, output_dir)

    def test_sync_google_keep_notes_opens_saved_sources(self) -> None:
        class Status:
            value = ""

            def set(self, value: str) -> None:
                self.value = value

        class Dummy:
            def __init__(self) -> None:
                self.assignment_config_status = Status()
                self.refreshed = False
                self.opened: list[str] = []

            def _refresh_keep_source_registry(self) -> None:
                self.refreshed = True

            def _saved_google_keep_sources(self) -> list[dict[str, object]]:
                return [
                    {"url": "https://keep.google.com/u/0/#NOTE/abc123", "path": "rules.txt"},
                    {"url": "https://keep.google.com/u/0/#NOTE/def456", "path": "payouts.txt"},
                ]

            def _open_google_keep_source_url(self, url: str) -> bool:
                self.opened.append(url)
                return True

        dummy = Dummy()
        app.CardPipelineApp.sync_google_keep_notes(dummy)

        self.assertTrue(dummy.refreshed)
        self.assertEqual(
            dummy.opened,
            [
                "https://keep.google.com/u/0/#NOTE/abc123",
                "https://keep.google.com/u/0/#NOTE/def456",
            ],
        )
        self.assertIn("Opened 2 Google Keep notes", dummy.assignment_config_status.value)

    def test_bridge_replaces_matching_google_keep_cache(self) -> None:
        with TemporaryDirectory() as tmp:
            cache_path = Path(tmp) / "rules.txt"
            bridge = app.BridgeState()
            bridge.register_keep_note_sources(
                [
                    {
                        "url": "https://keep.google.com/u/0/#NOTE/abc123",
                        "path": str(cache_path),
                        "name": "Rules",
                    }
                ]
            )

            result = bridge.post_google_keep_note(
                {
                    "url": "https://keep.google.com/u/0/#NOTE/abc123",
                    "title": "Rules",
                    "text": "Football $20-$200 95%",
                }
            )

            self.assertTrue(result["ok"])
            self.assertEqual(cache_path.read_text(encoding="utf-8").strip(), "Football $20-$200 95%")

    def test_bridge_poll_keeps_google_keep_sources_private(self) -> None:
        bridge = app.BridgeState()
        bridge.register_keep_note_sources(
            [
                {
                    "url": "https://keep.google.com/u/0/#NOTE/abc123",
                    "path": "rules.txt",
                    "name": "Rules",
                }
            ]
        )

        poll_result = bridge.extension_poll({"extensionVersion": app.EXPECTED_CARDLADDER_EXTENSION_VERSION})
        status_result = bridge.snapshot()

        self.assertNotIn("keepNoteSources", poll_result)
        self.assertNotIn("keepNoteSources", status_result)
        self.assertEqual(bridge.keep_note_sources[0]["url"], "https://keep.google.com/u/0/#NOTE/abc123")

    def test_google_keep_hash_note_url_matches_notes_url(self) -> None:
        self.assertTrue(
            bridge_server.keep_urls_match(
                "https://keep.google.com/u/0/#NOTE/abc123",
                "https://keep.google.com/u/0/notes/abc123",
            )
        )

    def test_raw_google_sheet_url_uses_authenticated_reader(self) -> None:
        sheet_url = "https://docs.google.com/spreadsheets/d/private-sheet-id/edit#gid=0"
        with TemporaryDirectory() as tmp:
            with patch.object(assignment_engine, "read_google_sheet_text", return_value="Arena Club rules") as reader:
                text = assignment_engine.read_source_text(sheet_url, Path(tmp), interactive_google=True)

        self.assertEqual(text, "Arena Club rules")
        reader.assert_called_once_with(sheet_url, interactive=True)


class DiagnosticsTests(unittest.TestCase):
    def test_setup_doctor_reports_pipeline_and_google_paths(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            for name in ("WORKING SHEETS", "INCOMING SHEETS", "RECEIVED SHEETS", "ASSIGNMENT RULES", "COMPANY SHEETS"):
                (root / name).mkdir(parents=True)

            rows = lucas_diagnostics.setup_doctor_results(
                root,
                {"extensionVersion": "0.1.5", "expectedExtensionVersion": "0.1.5"},
                "Mac",
            )
            by_name = {row["name"]: row for row in rows}

            self.assertEqual(by_name["Shared pipeline folders"]["status"], "OK")
            self.assertEqual(by_name["Card Ladder helper version"]["status"], "OK")
            self.assertIn("lucas_google_sheets_token.json", by_name["Sheets OAuth token"]["detail"])
            self.assertIn("LUCAS Mac", lucas_diagnostics.lucas_version_label("Mac"))


class AssignmentEngineTests(unittest.TestCase):
    def test_payout_table_line_parses_range_without_year_context(self) -> None:
        tier = assignment_engine.parse_payout_table_line("Football $10 - $25 135%")

        self.assertIsNotNone(tier)
        assert tier is not None
        self.assertEqual(tier.matcher.lower(), "football")
        self.assertEqual(tier.min_price, 10)
        self.assertEqual(tier.max_price, 25)
        self.assertAlmostEqual(tier.rate, 1.35)

    def test_kemba_walker_title_infers_basketball_without_short_name_false_matches(self) -> None:
        title = "2019 Panini Contenders Optic Uniformity 21 Kemba Walker PSA 10"
        matches = assignment_engine.find_known_player_sports(title)

        self.assertTrue(any(match["key"] == "kemba walker" and match["sport"] == "basketball" for match in matches))
        self.assertFalse(any(match["sport"] in {"disney", "marvel", "star wars"} for match in matches))
        self.assertEqual(assignment_engine.infer_sport(title, "Kemba Walker"), "basketball")

    def test_shintaro_fujinami_title_infers_baseball_not_wwe(self) -> None:
        title = "2022 Bowman Npb 82 Shintaro Fujinami Chrome-Refractor PSA 10"
        matches = assignment_engine.find_known_player_sports(title)

        self.assertTrue(any(match["key"] == "shintaro fujinami" and match["sport"] == "baseball" for match in matches))
        self.assertFalse(any(match["key"] == "tatsumi fujinami" and match["sport"] == "wwe" for match in matches))
        self.assertEqual(assignment_engine.infer_sport(title, "Shintaro Fujinami"), "baseball")

    def test_curated_player_sport_hint_wins_over_generated_name_collision(self) -> None:
        parsed = assignment_engine.parse_card_for_matching("1996 Topps Chrome Michael Jordan PSA 10")

        self.assertEqual(parsed["playerName"], "Michael Jordan")
        self.assertEqual(parsed["sport"], "basketball")

    def test_bowman_chrome_prospect_title_infers_baseball_without_known_player(self) -> None:
        parsed = assignment_engine.parse_card_for_matching(
            "2017 Bowman Chrome Prospect Autographs Gold Shimmer Refractors #CPACF Clint Frazier BGS 9.5"
        )

        self.assertEqual(parsed["sport"], "baseball")

    def test_rule_category_aliases_remember_common_shorthand(self) -> None:
        self.assertEqual(assignment_engine.infer_sport("B-Ball $17-50 PSA 10"), "basketball")
        self.assertEqual(assignment_engine.infer_sport("bball $50-250 SGC 9"), "basketball")
        self.assertEqual(assignment_engine.infer_sport("1 Piece $15-50 PSA 10"), "one piece")
        self.assertEqual(assignment_engine.infer_sport("Poke $10-100 CGC 10"), "pokemon")

    def test_keep_rule_parser_handles_cy_shorthand_ranges_and_no_blocks(self) -> None:
        rules = assignment_engine.parse_rules(
            "\n".join(
                [
                    "SPORTS: PSA BGS SGC",
                    "1 Piece $1.75-2.5k None 5 4+",
                    "Poke $10-100(FIRM)3 No Limit 3+",
                    "- NO 7-10k CGC",
                    "- DO NOT BUY ANY MARIO OR LUIGI",
                ]
            )
        )

        self.assertEqual(rules.include, [])
        self.assertEqual((rules.ranges[0].min_price, rules.ranges[0].max_price), (1750.0, 2500.0))
        poke_rule = next(rule for rule in rules.ranges if rule.min_price == 10.0 and rule.max_price == 100.0)
        self.assertEqual(assignment_engine.canonical_sport_label(poke_rule.matcher), "pokemon")
        self.assertTrue(assignment_engine.rule_matches(poke_rule, "Pokemon test PSA 10", 86.0))
        self.assertTrue(any(rule.block and rule.min_price == 7000.0 and rule.max_price == 10000.0 for rule in rules.blocks))
        self.assertTrue(any(rule.block and "MARIO OR LUIGI" in rule.matcher for rule in rules.blocks))

    def test_sports_titles_do_not_match_single_word_entertainment_characters(self) -> None:
        examples = {
            "2023 Topps Chrome Platinum Anniversary Autographs Yd Yusniel Diaz PSA 9": ("Yusniel Diaz", "baseball"),
            "2021 Bowman Chrome Futurist Jd Jasson Dominguez Aqua Refractor PSA 10": ("Jasson Dominguez", "baseball"),
            "2023 Bowman Chrome Prospect Autographs Cparo Ricardo Olivar Speckle Refractor PSA 10": ("Ricardo Olivar", "baseball"),
            "1995 Zenith 111 Chipper Jones PSA 9": ("Chipper Jones", "baseball"),
        }

        for title, expected in examples.items():
            with self.subTest(title=title):
                parsed = assignment_engine.parse_card_for_matching(title)
                self.assertEqual((parsed["playerName"], parsed["sport"]), expected)

    def test_world_series_title_infers_baseball_without_false_character_match(self) -> None:
        parsed = assignment_engine.parse_card_for_matching("1990 Score #702 World Series Game 3 PSA 9")

        self.assertEqual(parsed["sport"], "baseball")
        self.assertEqual(parsed["playerName"], "")
        self.assertNotEqual(parsed["sport"], "disney")

    def test_tim_tebow_title_infers_football(self) -> None:
        parsed = assignment_engine.parse_card_for_matching("2010 Donruss Rated Rookies 95 Tim Tebow PSA 10")

        self.assertEqual(parsed["sport"], "football")
        self.assertEqual(parsed["playerName"], "Tim Tebow")

    def test_player_sport_overrides_teach_unknown_players(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "overrides.json"
            path.write_text(
                json.dumps({"players": {"Test Override Player": {"sport": "basketball", "displayName": "Test Override Player"}}}),
                encoding="utf-8",
            )

            loaded = assignment_engine.load_player_sport_overrides(path)

        self.assertEqual(loaded, 1)
        self.assertEqual(assignment_engine.infer_sport("2024 Panini Test Override Player PSA 10", "Test Override Player"), "basketball")

    def test_recommendation_chooses_highest_payout_among_accepted_companies(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="1",
            grader="PSA",
            card_title="2019 Panini Prizm Stephen Curry Silver PSA 10",
            card_ladder_comps_average=100,
            card_ladder_value=150,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Lower",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("basketball", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.9, "NBA")],
                ),
                assignment_engine.AssignmentCompany(
                    "Higher",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("basketball", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.95, "NBA")],
                ),
                assignment_engine.AssignmentCompany(
                    "Rejected",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("baseball", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 1.0, "MLB")],
                ),
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(recommendation.company, "Higher")
        self.assertEqual(recommendation.payout, 95)
        self.assertTrue(decisions["Lower"].accepted)
        self.assertFalse(decisions["Rejected"].accepted)

    def test_unlicensed_rule_matches_known_unlicensed_patterns(self) -> None:
        rules = assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("unlicensed", 0, 1000)])
        examples = [
            "2021 Panini Prizm Mike Trout Red PSA 10",
            "2020 Panini Contenders Draft Picks Stephen Curry PSA 10",
            "2023 Leaf Metal Tom Brady Autograph PSA 10",
        ]

        for title in examples:
            with self.subTest(title=title):
                self.assertTrue(assignment_engine.company_accepts(rules, title, 100, "PSA"))

        self.assertFalse(assignment_engine.company_accepts(rules, "2019 Panini Prizm Stephen Curry Silver PSA 10", 100, "PSA"))

    def test_grade_bound_rules_reject_rows_with_missing_grade(self) -> None:
        rules = assignment_engine.CompanyRules(
            accept_all=True,
            grade_rules={"psa": assignment_engine.GradeRule(allowed=True, min_grade=10)},
        )

        self.assertTrue(assignment_engine.company_accepts(rules, "2019 Panini Prizm Stephen Curry Silver PSA 10", 100, "PSA"))
        self.assertFalse(assignment_engine.company_accepts(rules, "2019 Panini Prizm Stephen Curry Silver", 100, "PSA"))

    def test_assignment_rejection_names_block_rule(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="152304499",
            grader="PSA",
            card_title="2025 Topps Chrome #251 Cooper Flagg PSA 10",
            card_ladder_comps_average=725.60,
            category="basketball",
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Arena Club",
                    assignment_engine.CompanyRules(
                        ranges=[assignment_engine.AssignmentRule("Basketball", 500, 999)],
                        blocks=[assignment_engine.AssignmentRule("2025 Topps Chrome Cooper Flagg 10", block=True, grade_companies=("PSA",))],
                    ),
                    [assignment_engine.PayoutTier(500, 999, 0.95, "Basketball")],
                )
            ]
        )

        decision = engine.evaluate(row)[0]

        self.assertFalse(decision.accepted)
        self.assertIn("blocked by rule", decision.reason)
        self.assertIn("2025 Topps Chrome Cooper Flagg 10", decision.reason)

    def test_assignment_rejection_names_value_gap_for_matching_category(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="152304527",
            grader="PSA",
            card_title="2025 Topps Chrome #254 Kon Knueppel PSA 10 Xfractor",
            card_ladder_comps_average=303,
            category="basketball",
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Arena Club",
                    assignment_engine.CompanyRules(
                        ranges=[
                            assignment_engine.AssignmentRule("Basketball", 100, 299),
                            assignment_engine.AssignmentRule("Basketball", 500, 999),
                        ]
                    ),
                    [assignment_engine.PayoutTier(100, 999, 0.95, "Basketball")],
                )
            ]
        )

        decision = engine.evaluate(row)[0]

        self.assertFalse(decision.accepted)
        self.assertIn("matched Basketball", decision.reason)
        self.assertIn("$303", decision.reason)
        self.assertIn("$100 to $299", decision.reason)
        self.assertIn("$500 to $999", decision.reason)

    def test_company_year_range_rejects_cards_outside_manual_range(self) -> None:
        rules = assignment_engine.parse_rules(
            json.dumps(
                {
                    "minYear": "1990",
                    "rules": [
                        {
                            "sports": ["football"],
                            "priceRanges": [{"min": "100", "max": "2000"}],
                        }
                    ],
                }
            )
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Fanatics",
                    rules,
                    [assignment_engine.PayoutTier(100, 2000, 0.93, "Football")],
                    value_source="card_ladder",
                )
            ]
        )
        modern = WorkbookRow(
            excel_row=2,
            cert_number="59840219",
            grader="PSA",
            card_title="2018 Panini Prizm Stained Glass SG8 Josh Allen PSA 9",
            card_ladder_value=1527,
            category="football",
        )
        too_old = WorkbookRow(
            excel_row=3,
            cert_number="123",
            grader="PSA",
            card_title="1989 Score Barry Sanders PSA 9",
            card_ladder_value=1527,
            category="football",
        )

        modern_decision, old_decision = engine.evaluate(modern)[0], engine.evaluate(too_old)[0]

        self.assertTrue(modern_decision.accepted)
        self.assertFalse(old_decision.accepted)
        self.assertIn("card year 1989", old_decision.reason)
        self.assertIn("1990 to current", old_decision.reason)

    def test_person_payout_policy_locks_companies_and_overrides_rates(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="1",
            grader="PSA",
            card_title="2019 Panini Prizm Stephen Curry Silver PSA 10",
            card_ladder_comps_average=100,
        )
        policies = assignment_engine.parse_person_policies(
            "Person,Company,Min,Max,Rate\n"
            "Lucas,Arena Club,0,,95%\n"
            "Mikey,Arena Club,0,,90%\n"
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Arena Club",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("basketball", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.8, "NBA")],
                ),
                assignment_engine.AssignmentCompany(
                    "Other Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("basketball", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 1.0, "NBA")],
                ),
            ],
            person_policies=policies,
        )

        lucas = engine.recommend(row, person="Lucas")
        mikey = engine.recommend(row, person="Mikey")
        unlisted = engine.recommend(row, person="Unlisted")
        lucas_decisions = {decision.company: decision for decision in engine.evaluate(row, person="Lucas")}

        self.assertEqual(lucas.company, "Arena Club")
        self.assertEqual(lucas.payout, 95)
        self.assertEqual(mikey.company, "Arena Club")
        self.assertEqual(mikey.payout, 90)
        self.assertEqual(unlisted.company, "Other Buyer")
        self.assertEqual(unlisted.payout, 100)
        self.assertFalse(lucas_decisions["Other Buyer"].accepted)
        self.assertIn("not allowed", lucas_decisions["Other Buyer"].reason)

    def test_assignment_engine_loads_person_payout_source_file(self) -> None:
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            policy_path = root / "person-payouts.csv"
            config_path = root / "assignment_companies.json"
            policy_path.write_text(
                "Person,Company,Rate\nLucas,Arena Club,95%\n",
                encoding="utf-8",
            )
            config_path.write_text(
                json.dumps(
                    {
                        "person_payouts_source": str(policy_path),
                        "companies": [
                            {
                                "name": "Arena Club",
                                "accept_all": True,
                                "rate": "80%",
                            }
                        ],
                    }
                ),
                encoding="utf-8",
            )

            engine = assignment_engine.AssignmentEngine.load(config_path)
            row = WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="Test Card", card_ladder_comps_average=100)

            self.assertEqual(engine.recommend(row, person="Lucas").payout, 95)
            self.assertEqual(engine.recommend(row, person="Other").payout, 80)

    def test_assignment_engine_loads_company_sheet_reset_schedule(self) -> None:
        with TemporaryDirectory() as tmp:
            config_path = Path(tmp) / "assignment_companies.json"
            config_path.write_text(
                json.dumps(
                    {
                        "companies": [
                            {
                                "name": "Arena Club",
                                "accept_all": True,
                                "rate": "80%",
                                "reset_weekday": "Wednesday",
                                "reset_time": "12:30",
                            }
                        ],
                    }
                ),
                encoding="utf-8",
            )

            engine = assignment_engine.AssignmentEngine.load(config_path)

        self.assertEqual(engine.companies[0].reset_weekday, "Wednesday")
        self.assertEqual(engine.companies[0].reset_time, "12:30")

    def test_company_can_prefer_card_ladder_value_over_comps(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="4",
            grader="PSA",
            card_title="2020 Panini Prizm Patrick Mahomes PSA 10",
            card_ladder_comps_average=100,
            card_ladder_value=150,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Comps Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.95, "NFL")],
                    value_source="comps",
                ),
                assignment_engine.AssignmentCompany(
                    "CL Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.9, "NFL")],
                    value_source="card_ladder",
                ),
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(decisions["Comps Buyer"].source_value, 100)
        self.assertEqual(decisions["Comps Buyer"].payout, 95)
        self.assertEqual(decisions["CL Buyer"].source_value, 150)
        self.assertEqual(decisions["CL Buyer"].payout, 135)
        self.assertEqual(recommendation.company, "CL Buyer")
        self.assertEqual(recommendation.payout, 135)

    def test_manual_row_category_drives_assignment_matching(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="1",
            grader="BGS",
            category="baseball",
            card_title="Generic Prospect Auto BGS 9.5",
            card_ladder_comps_average=100,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Baseball Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("baseball", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.9, "baseball")],
                )
            ]
        )

        recommendation = engine.recommend(row)

        self.assertEqual(recommendation.company, "Baseball Buyer")
        self.assertEqual(recommendation.payout, 90)

    def test_card_ladder_value_source_rejects_company_when_cl_missing(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="5",
            grader="PSA",
            card_title="2020 Panini Prizm Patrick Mahomes PSA 10",
            card_ladder_comps_average=100,
            card_ladder_value=None,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Comps Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 0.9, "NFL")],
                    value_source="comps",
                ),
                assignment_engine.AssignmentCompany(
                    "CL Required Buyer",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 500, 1.0, "NFL")],
                    value_source="card_ladder",
                ),
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(recommendation.company, "Comps Buyer")
        self.assertEqual(recommendation.payout, 90)
        self.assertFalse(decisions["CL Required Buyer"].accepted)
        self.assertIsNone(decisions["CL Required Buyer"].source_value)
        self.assertIn("missing", decisions["CL Required Buyer"].reason)

    def test_company_can_require_cy_estimate_value(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="6",
            grader="CGC",
            card_title="2022 Paradigm Trigger Unown VSTAR CGC 10",
            card_ladder_comps_average=100,
            card_ladder_value=110,
            cy_value=150,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Comps Buyer",
                    assignment_engine.CompanyRules(accept_all=True),
                    [assignment_engine.PayoutTier(10, 500, 0.9)],
                    value_source="comps",
                ),
                assignment_engine.AssignmentCompany(
                    "CY Buyer",
                    assignment_engine.CompanyRules(accept_all=True),
                    [assignment_engine.PayoutTier(10, 500, 0.85)],
                    value_source="cy_estimate",
                ),
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(decisions["Comps Buyer"].source_value, 100)
        self.assertEqual(decisions["CY Buyer"].source_value, 150)
        self.assertEqual(recommendation.company, "CY Buyer")
        self.assertEqual(recommendation.payout, 127.5)

    def test_default_company_value_source_falls_back_to_cy_after_comps_and_card_ladder(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="8",
            grader="CGC",
            card_title="2022 Paradigm Trigger Unown VSTAR CGC 10",
            card_ladder_comps_average=None,
            card_ladder_value=None,
            cy_value=150,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Default Buyer",
                    assignment_engine.CompanyRules(accept_all=True),
                    [assignment_engine.PayoutTier(10, 500, 0.8)],
                    value_source="comps",
                )
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(decisions["Default Buyer"].source_value, 150)
        self.assertEqual(recommendation.company, "Default Buyer")
        self.assertEqual(recommendation.payout, 120)

    def test_courtyard_value_source_alias_uses_cy_estimate(self) -> None:
        company = assignment_engine.load_company(
            {"name": "CY Alias Buyer", "accept_all": True, "rate": "80%", "value_source": "CourtYard"},
            Path("."),
        )
        row = WorkbookRow(
            excel_row=2,
            cert_number="9",
            grader="CGC",
            card_title="2022 Paradigm Trigger Unown VSTAR CGC 10",
            card_ladder_comps_average=100,
            card_ladder_value=110,
            cy_value=150,
        )
        engine = assignment_engine.AssignmentEngine([company])

        recommendation = engine.recommend(row)

        self.assertEqual(company.value_source, "cy_estimate")
        self.assertEqual(recommendation.company, "CY Alias Buyer")
        self.assertEqual(recommendation.payout, 120)

    def test_cy_estimate_value_source_rejects_company_when_cy_missing(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="7",
            grader="CGC",
            card_title="2022 Paradigm Trigger Unown VSTAR CGC 10",
            card_ladder_comps_average=100,
            cy_value=None,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "CY Required Buyer",
                    assignment_engine.CompanyRules(accept_all=True),
                    [assignment_engine.PayoutTier(10, 500, 1.0)],
                    value_source="cy_estimate",
                )
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(recommendation.company, "")
        self.assertFalse(decisions["CY Required Buyer"].accepted)
        self.assertIsNone(decisions["CY Required Buyer"].source_value)
        self.assertIn("missing", decisions["CY Required Buyer"].reason)

    def test_goat_payout_category_uses_payout_range_not_rule_goat_range(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="2",
            grader="PSA",
            card_title="2019 Panini Mosaic Stephen Curry Green Mosaic PSA 10",
            card_ladder_comps_average=85.61,
        )
        rules = assignment_engine.CompanyRules(
            ranges=[assignment_engine.AssignmentRule("basketball", 10, 500)],
            goat_players={"stephen curry"},
            goat_ranges=[assignment_engine.AssignmentRule("Stephen Curry", 100, 7500)],
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Goat Buyer",
                    rules,
                    [assignment_engine.PayoutTier(50, 99, 0.95, "GOATS")],
                )
            ]
        )

        recommendation = engine.recommend(row)

        self.assertEqual(recommendation.company, "Goat Buyer")
        self.assertEqual(recommendation.payout, 81.33)

    def test_date_weighted_uses_newest_comp_when_two_newest_are_more_than_seven_days_apart(self) -> None:
        newest = datetime.now() - timedelta(days=2)
        older = newest - timedelta(days=11)
        comps = [
            {"date_sold": (older - timedelta(days=30)).strftime("%b %d, %Y"), "price": "$8.00", "title": "Test Card PSA 10"},
            {"date_sold": newest.strftime("%b %d, %Y"), "price": "$25.00", "title": "Test Card PSA 10"},
            {"date_sold": older.strftime("%b %d, %Y"), "price": "$10.00", "title": "Test Card PSA 10"},
            {"date_sold": (older - timedelta(days=45)).strftime("%b %d, %Y"), "price": "$2.00", "title": "Test Card PSA 10"},
        ]

        self.assertEqual(app.comp_price(comps, app.COMP_STRATEGY_STALE_NEWEST), 25.0)

    def test_date_weighted_averages_when_two_newest_are_within_seven_days(self) -> None:
        newest = datetime.now() - timedelta(days=2)
        second = newest - timedelta(days=4)
        comps = [
            {"date_sold": (second - timedelta(days=30)).strftime("%b %d, %Y"), "price": "$8.00", "title": "Test Card PSA 10"},
            {"date_sold": newest.strftime("%b %d, %Y"), "price": "$25.00", "title": "Test Card PSA 10"},
            {"date_sold": second.strftime("%b %d, %Y"), "price": "$10.00", "title": "Test Card PSA 10"},
            {"date_sold": (second - timedelta(days=45)).strftime("%b %d, %Y"), "price": "$2.00", "title": "Test Card PSA 10"},
        ]

        self.assertEqual(app.comp_price(comps, app.COMP_STRATEGY_STALE_NEWEST), 11.25)

    def test_date_weighted_uses_newest_comp_when_newest_sale_is_stale(self) -> None:
        newest = datetime.now() - timedelta(days=8)
        second = newest - timedelta(days=4)
        comps = [
            {"date_sold": newest.strftime("%b %d, %Y"), "price": "$25.00", "title": "Test Card PSA 10"},
            {"date_sold": second.strftime("%b %d, %Y"), "price": "$10.00", "title": "Test Card PSA 10"},
            {"date_sold": (second - timedelta(days=30)).strftime("%b %d, %Y"), "price": "$8.00", "title": "Test Card PSA 10"},
        ]

        self.assertEqual(app.comp_price(comps, app.COMP_STRATEGY_STALE_NEWEST), 25.0)

    def test_date_weighted_dedupes_same_sale_before_comparing_newest_dates(self) -> None:
        comps = [
            {"date_sold": "Apr 5, 2026", "price": "$11.50", "source": "EBAY", "title": "ZANDGEMPORIUM Pokemon Magnezone Stormfront Holo Rare #5 PSA 7 $12.02"},
            {"date_sold": "Apr 5, 2026", "price": "$12.02", "source": "EBAY", "title": "ZANDGEMPORIUM Pokemon Magnezone Stormfront Holo Rare #5 PSA 7"},
            {"date_sold": "Aug 24, 2025", "price": "$10.50", "source": "EBAY", "title": "ALMAR ENTERPRISES 2008 POKEMON DIAMOND & PEARL STORMFRONT MAGNEZONE LV. 44 #5/100 RARE HOLO PSA 7"},
            {"date_sold": "Oct 2, 2022", "price": "$11.50", "source": "EBAY", "title": "Pokemon Magnezone D&P Stormfront Holo Rare #5 PSA 7 -454"},
        ]

        self.assertEqual(app.comp_price(comps, app.COMP_STRATEGY_STALE_NEWEST), 12.02)

    def test_date_weighted_prefers_clean_duplicate_when_title_contains_conflicting_price(self) -> None:
        comps = [
            {
                "date_sold": "Jun 1, 2025",
                "price": "$493.63",
                "sale_type": "Best Offer",
                "source": "EBAY",
                "title": "POKECARDS-2-TRUST 2022-23 Donruss Chet Holmgren RC #202 - Yellow Holo Laser /25 PSA 9 Rookie Pop 3- $130.50",
            },
            {
                "date_sold": "Jun 1, 2025",
                "price": "$130.50",
                "sale_type": "Best Offer",
                "source": "EBAY",
                "title": "POKECARDS-2-TRUST 2022-23 Donruss Chet Holmgren RC #202 - Yellow Holo Laser /25 PSA 9 Rookie Pop 3",
            },
            {
                "date_sold": "May 4, 2025",
                "price": "$141.25",
                "sale_type": "Auction",
                "source": "EBAY",
                "title": "POKECARDS-2-TRUST 2022-23 Donruss Chet Holmgren RC #202 - Yellow Holo Laser /25 PSA 9 Rookie Pop 3",
            },
        ]

        self.assertEqual(app.comp_price(comps, app.COMP_STRATEGY_STALE_NEWEST), 130.5)

    def test_date_weighted_prefers_best_same_day_stale_newest_comp(self) -> None:
        comps = [
            {
                "source": "BECKETT",
                "title": ", Autograph Grade: 10, Profile: 2014 Panini Flawless Greats Patches Autographs Gold #22 Joe Montana (Pop 2) $2,100.00",
                "date_sold": "Sep 4, 2025",
                "sale_type": "Auction",
                "price": "$1,424.00",
            },
            {
                "source": "ALT",
                "title": "(CONFIRMED PAID) 2014 Panini Flawless Greats Patch Autograph Gold Joe Montana #22 BGS 9 Auto 10 /10",
                "date_sold": "Sep 4, 2025",
                "sale_type": "Auction",
                "price": "$2,760.00",
            },
            {
                "source": "EBAY",
                "title": "225 BREAKERS 2014 Panini Flawless Joe Montana Greats Patch On Card AUTO Gold /10 BGS 9",
                "date_sold": "Jun 18, 2025",
                "sale_type": "Auction",
                "price": "$1,424.00",
            },
        ]

        self.assertEqual(app.comp_price(comps, app.COMP_STRATEGY_STALE_NEWEST), 2760.0)

    def test_cardladder_result_rejects_obvious_wrong_card_comp_rows(self) -> None:
        bridge = app.BridgeState()
        row = WorkbookRow(
            excel_row=2,
            cert_number="99505674",
            grader="PSA",
            card_title="2022 Panini Donruss 202 Chet Holmgren Yellow Holo Laser PSA 9",
        )
        result = {
            "excelRow": 2,
            "certNumber": "99505674",
            "grader": "PSA",
            "status": "ok",
            "value": 73,
            "ocr": {
                "profileTitle": "2022 Panini Donruss 202 Chet Holmgren Yellow Holo Laser",
                "profileGrader": "PSA",
                "profileGrade": "9",
                "resultCount": 2,
                "comps": [
                    {
                        "source": "EBAY",
                        "title": "2022 Panini Donruss Chet Holmgren Yellow Holo Laser #202 PSA 9",
                        "date_sold": "Jun 1, 2026",
                        "sale_type": "Auction",
                        "price": "$73.00",
                    },
                    {
                        "source": "EBAY",
                        "title": "2022 Panini Prizm Chet Holmgren Gold Rookie #266 PSA 10",
                        "date_sold": "Jun 2, 2026",
                        "sale_type": "Auction",
                        "price": "$314.75",
                    },
                ],
            },
        }

        bridge._apply_cardladder_result_to_row(row, result)

        self.assertEqual(row.card_ladder_comps_average, 73.0)
        self.assertIn("Yellow Holo Laser", row.card_ladder_comps)
        self.assertNotIn("Prizm", row.card_ladder_comps)
        self.assertIn("Rejected 1 likely wrong-card", row.notes)

    def test_date_weighted_normalizes_equivalent_date_formats_before_dedupe(self) -> None:
        comps = [
            {"date_sold": "Sept. 4, 2025", "price": "$90.00", "source": "EBAY", "sale_type": "Auction", "title": "Test Card PSA 10 $120.00"},
            {"date_sold": "Sep 4, 2025", "price": "$120.00", "source": "EBAY", "sale_type": "Auction", "title": "Test Card PSA 10"},
            {"date_sold": "8/1/25", "price": "$50.00", "source": "EBAY", "sale_type": "Auction", "title": "Older Test Card PSA 10"},
        ]

        self.assertEqual(app.comp_price(comps, app.COMP_STRATEGY_STALE_NEWEST), 120.0)

    def test_date_weighted_parses_date_embedded_in_ocr_label_text(self) -> None:
        comps = [
            {"date_sold": "Date Sold: Jun 1, 2025 Type Auction", "price": "$25.00", "title": "Test Card PSA 10"},
            {"date_sold": "5/20/25", "price": "$10.00", "title": "Test Card PSA 10"},
        ]

        self.assertEqual(app.comp_price(comps, app.COMP_STRATEGY_STALE_NEWEST), 25.0)

    def test_date_weighted_excludes_undated_comps_when_dated_comps_are_available(self) -> None:
        newest = datetime.now() - timedelta(days=2)
        second = newest - timedelta(days=4)
        comps = [
            {"date_sold": newest.strftime("%b %d, %Y"), "price": "$25.00", "title": "Test Card PSA 10"},
            {"date_sold": second.strftime("%b %d, %Y"), "price": "$15.00", "title": "Test Card PSA 10"},
            {"date_sold": "May 1", "price": "$500.00", "title": "Test Card PSA 10 missing year"},
        ]

        self.assertEqual(app.comp_price(comps, app.COMP_STRATEGY_STALE_NEWEST), 20.0)

    def test_accepted_company_without_matching_payout_cannot_win(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="3",
            grader="PSA",
            card_title="2022 Panini Prizm Patrick Mahomes PSA 10",
            card_ladder_comps_average=80,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "No Payout Match",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(100, 500, 1.0, "NFL")],
                ),
                assignment_engine.AssignmentCompany(
                    "Valid Payout",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("football", 10, 500)]),
                    [assignment_engine.PayoutTier(10, 99, 0.9, "NFL")],
                ),
            ]
        )

        recommendation = engine.recommend(row)
        decisions = {decision.company: decision for decision in engine.evaluate(row)}

        self.assertEqual(recommendation.company, "Valid Payout")
        self.assertIsNone(decisions["No Payout Match"].payout)
        self.assertIn("no payout tier", decisions["No Payout Match"].reason)

    def test_sport_payout_does_not_use_conflicting_single_token_player_hint(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="67474515",
            grader="PSA",
            category="baseball",
            card_title="2021 Bowman Sapphire Edition Chrome Prospects Bcp71 Blaze Jordan PSA 10",
            card_ladder_comps_average=57.31,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Arena",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("baseball", 10, 500)]),
                    [
                        assignment_engine.PayoutTier(51, 99, 0.82, "Baseball"),
                        assignment_engine.PayoutTier(51, 99, 0.95, "Soccer"),
                    ],
                )
            ]
        )

        decision = engine.evaluate(row)[0]

        self.assertTrue(decision.accepted)
        self.assertAlmostEqual(decision.payout or 0, 46.9942)
        self.assertEqual(decision.payout_rate, 0.82)
        self.assertEqual(decision.payout_category, "Baseball")

    def test_assignment_decision_reports_selected_payout_tier(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="141955355",
            grader="PSA",
            category="baseball",
            card_title="1963 Topps #200 Mickey Mantle PSA 2",
            card_ladder_comps_average=425,
        )
        engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Arena Club",
                    assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("baseball", 10, 1000)]),
                    [
                        assignment_engine.PayoutTier(0, 1000, 0.95, "Baseball"),
                        assignment_engine.PayoutTier(0, 1000, 0.9, "Soccer"),
                    ],
                )
            ]
        )

        decision = engine.evaluate(row)[0]

        self.assertTrue(decision.accepted)
        self.assertEqual(decision.payout, 403.75)
        self.assertEqual(decision.payout_rate, 0.95)
        self.assertEqual(decision.payout_category, "Baseball")

    def test_dnb_over_50_section_only_blocks_cards_above_50(self) -> None:
        values = [
            ["", "", "", "Do Not Buy these Players", "", "", "", "Do Not Buy these Players over $50"],
            ["", "Do not buy", "", "Basketball", "Football", "Baseball", "", "Basketball", "Football", "Baseball"],
            ["", "", "", "", "", "", "", "", "Jayden Daniels", ""],
        ]
        rules_text = "\n".join(["football $10-$500", *assignment_engine.synthesize_do_not_buy_rules(values)])
        rules = assignment_engine.parse_rules(rules_text)
        engine = assignment_engine.AssignmentEngine([
            assignment_engine.AssignmentCompany("Arena", rules, [assignment_engine.PayoutTier(10, 500, 0.9)])
        ])
        low_value = WorkbookRow(
            excel_row=2,
            cert_number="",
            grader="PSA",
            category="football",
            card_title="2024 Panini Absolute Jayden Daniels PSA 9",
            card_ladder_comps_average=50,
        )
        high_value = WorkbookRow(
            excel_row=3,
            cert_number="",
            grader="PSA",
            category="football",
            card_title="2024 Panini Absolute Jayden Daniels PSA 9",
            card_ladder_comps_average=51,
        )

        self.assertIn("block: Jayden Daniels over 50", rules_text)
        self.assertTrue(engine.evaluate(low_value)[0].accepted)
        self.assertFalse(engine.evaluate(high_value)[0].accepted)

    def test_dnb_over_50_blocks_misspelled_penix_jr_above_threshold(self) -> None:
        rules = assignment_engine.parse_rules("football $10-$500\nblock: Football Mchael Penix Jr. over 50")
        engine = assignment_engine.AssignmentEngine([
            assignment_engine.AssignmentCompany("Arena", rules, [assignment_engine.PayoutTier(10, 500, 0.9)])
        ])
        low_value = WorkbookRow(
            excel_row=2,
            cert_number="",
            grader="PSA",
            category="football",
            card_title="2024 Panini Prizm Michael Penix Jr PSA 10",
            card_ladder_comps_average=50,
        )
        high_value = WorkbookRow(
            excel_row=3,
            cert_number="",
            grader="PSA",
            category="football",
            card_title="2024 Panini Prizm Michael Penix Jr. PSA 10",
            card_ladder_comps_average=51,
        )

        self.assertTrue(engine.evaluate(low_value)[0].accepted)
        self.assertFalse(engine.evaluate(high_value)[0].accepted)

    def test_broad_modern_dnb_rules_apply_year_and_grade_constraints(self) -> None:
        rules = assignment_engine.parse_rules(
            "football $10-$500\n"
            "baseball $10-$500\n"
            "block: Ultra modern less than PSA/BGS 9\n"
            "block: Modern SGC over 30"
        )
        engine = assignment_engine.AssignmentEngine([
            assignment_engine.AssignmentCompany("Arena", rules, [assignment_engine.PayoutTier(10, 500, 0.9)])
        ])
        ultra_psa8 = WorkbookRow(
            excel_row=2,
            cert_number="",
            grader="PSA",
            category="football",
            card_title="2024 Panini Prizm Test Player PSA 8",
            card_ladder_comps_average=20,
        )
        ultra_psa9 = WorkbookRow(
            excel_row=3,
            cert_number="",
            grader="PSA",
            category="football",
            card_title="2024 Panini Prizm Test Player PSA 9",
            card_ladder_comps_average=20,
        )
        modern_sgc = WorkbookRow(
            excel_row=4,
            cert_number="",
            grader="SGC",
            category="baseball",
            card_title="2015 Topps Test Player SGC 10",
            card_ladder_comps_average=31,
        )
        ultra_sgc = WorkbookRow(
            excel_row=5,
            cert_number="",
            grader="SGC",
            category="baseball",
            card_title="2024 Topps Test Player SGC 10",
            card_ladder_comps_average=31,
        )

        self.assertFalse(engine.evaluate(ultra_psa8)[0].accepted)
        self.assertTrue(engine.evaluate(ultra_psa9)[0].accepted)
        self.assertFalse(engine.evaluate(modern_sgc)[0].accepted)
        self.assertTrue(engine.evaluate(ultra_sgc)[0].accepted)


class AppSharedWorkflowLogicTests(unittest.TestCase):
    def test_bridge_only_hands_commands_to_expected_extension_version(self) -> None:
        bridge = app.BridgeState()
        row = WorkbookRow(excel_row=2, cert_number="123", grader="CGC", card_title="Test Card CGC 9")
        bridge.set_rows([row])
        bridge.start_all_comps(requery_all=True)

        stale = bridge.extension_poll({"extensionVersion": "older-helper"})
        self.assertIsNone(stale["command"])

        current = bridge.extension_poll({"extensionVersion": app.EXPECTED_CARDLADDER_EXTENSION_VERSION})
        self.assertIsNotNone(current["command"])
        self.assertEqual(current["command"]["type"], "RUN_ALL_COMPS")

    def test_mobile_pin_is_reused_from_settings(self) -> None:
        settings = {"mobile_pin": "346200"}

        self.assertEqual(app.ensure_mobile_pin(settings), "346200")
        self.assertEqual(settings["mobile_pin"], "346200")

    def test_mobile_host_prefers_configured_stable_host(self) -> None:
        with patch.dict(app.os.environ, {"LUCAS_MOBILE_HOST": "Lucas-Mac.local"}):
            self.assertEqual(app.mobile_app_host({"mobile_host": "192.168.1.246"}), "Lucas-Mac.local")

    def test_mobile_host_uses_macos_local_hostname_before_lan_ip(self) -> None:
        result = types.SimpleNamespace(stdout="Michaels-MacBook\n")
        with patch.object(app.subprocess, "run", return_value=result):
            self.assertEqual(app.mobile_app_host({}), "Michaels-MacBook.local")

    def test_bridge_keeps_default_mobile_port_stable(self) -> None:
        bridge = app.BridgeServer(app.BridgeState())

        self.assertEqual(bridge.port, 8765)
        self.assertFalse(bridge.allow_port_fallback)

    def test_instagram_media_route_supports_head_requests(self) -> None:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.bind(("127.0.0.1", 0))
            port = sock.getsockname()[1]

        state = app.BridgeState()
        state.instagram_media_resolver = lambda photo_id: (b"jpg-bytes", "image/jpeg") if photo_id == "abc" else None
        bridge = app.BridgeServer(state, host="127.0.0.1", port=port)
        bridge.start()
        self.assertTrue(bridge.started, bridge.error)
        try:
            request = urllib.request.Request(
                f"http://127.0.0.1:{port}{state.instagram_media_path('abc', 'front.jpg')}",
                method="HEAD",
            )
            with urllib.request.urlopen(request, timeout=5) as response:
                self.assertEqual(response.status, 200)
                self.assertEqual(response.headers.get("content-type"), "image/jpeg")
                self.assertEqual(response.headers.get("content-length"), str(len(b"jpg-bytes")))
                self.assertEqual(response.read(), b"")
        finally:
            bridge.stop()

    def test_mobile_bridge_port_uses_profile_specific_defaults(self) -> None:
        self.assertEqual(app.mobile_bridge_port({}, Path("lucas_settings.json")), 8765)
        self.assertEqual(
            app.mobile_bridge_port(
                {"pipeline_root": "/Users/test/Library/CloudStorage/Drive/LUCAS_PERSONAL"},
                Path("lucas_settings.michael.json"),
            ),
            8766,
        )

    def test_mobile_bridge_port_can_be_overridden(self) -> None:
        with patch.dict(app.os.environ, {"LUCAS_MOBILE_PORT": "8777"}):
            self.assertEqual(app.mobile_bridge_port({}, Path("lucas_settings.json")), 8777)

    def test_mobile_public_app_url_requires_https_and_appends_profile(self) -> None:
        self.assertEqual(
            app.mobile_public_app_url("personal", {"mobile_public_url": "https://lucas.example.com"}),
            "https://lucas.example.com/mobile/personal",
        )
        self.assertEqual(
            app.mobile_public_app_url("team", {"mobile_public_url": "https://lucas.example.com/mobile"}),
            "https://lucas.example.com/mobile/team",
        )
        self.assertEqual(
            app.mobile_public_app_url("personal", {"mobile_public_url": "https://lucas.example.com/mobile/personal"}),
            "https://lucas.example.com/mobile/personal",
        )
        self.assertEqual(app.mobile_public_app_url("personal", {"mobile_public_url": "http://192.168.1.244:8766"}), "")

    def test_mobile_public_app_url_prefers_profile_setting_over_global_env(self) -> None:
        with patch.dict(app.os.environ, {"LUCAS_MOBILE_PUBLIC_URL": "https://personal.example.com"}):
            self.assertEqual(
                app.mobile_public_app_url("team", {"mobile_public_url": "https://team.example.com"}),
                "https://team.example.com/mobile/team",
            )
        with patch.dict(app.os.environ, {"LUCAS_TEAM_MOBILE_PUBLIC_URL": "https://team-env.example.com"}):
            self.assertEqual(
                app.mobile_public_app_url("team", {"mobile_public_url": "https://team.example.com"}),
                "https://team-env.example.com/mobile/team",
            )

    def test_bridge_rejects_untrusted_browser_origin(self) -> None:
        self.assertFalse(bridge_server.request_origin_allowed("https://example.com", "127.0.0.1:8765"))
        self.assertFalse(bridge_server.request_origin_allowed("file://local", "127.0.0.1:8765"))

    def test_bridge_allows_same_origin_and_extension_origins(self) -> None:
        self.assertTrue(bridge_server.request_origin_allowed("", "192.168.1.246:8765"))
        self.assertTrue(bridge_server.request_origin_allowed("http://192.168.1.246:8765", "192.168.1.246:8765"))
        self.assertTrue(bridge_server.request_origin_allowed("chrome-extension://test-extension", "127.0.0.1:8765"))

    def test_bridge_local_only_detection_blocks_lan_clients(self) -> None:
        self.assertTrue(bridge_server.is_loopback_address("127.0.0.1"))
        self.assertTrue(bridge_server.is_loopback_address("::1"))
        self.assertFalse(bridge_server.is_loopback_address("192.168.1.246"))
        self.assertIn("/status", bridge_server.BRIDGE_LOCAL_ONLY_PATH_PREFIXES)
        self.assertIn("/result/cardladder", bridge_server.BRIDGE_LOCAL_ONLY_PATH_PREFIXES)

    def test_add_comp_row_does_not_deadlock_on_bridge_state_lock(self) -> None:
        class FakeTree:
            def selection_set(self, _iid: str) -> None:
                self.selected = _iid

            def focus(self, _iid: str) -> None:
                self.focused = _iid

            def see(self, _iid: str) -> None:
                self.seen = _iid

        class Status:
            def __init__(self) -> None:
                self.value = ""

            def set(self, value: str) -> None:
                self.value = value

        class Dummy:
            add_comp_row = app.CardPipelineApp.add_comp_row

            def _comp_sheet_info(self, _label: str):
                return "Working", "Lot.xlsx", Path("Lot.xlsx")

            def _cancel_cell_edit(self) -> None:
                self.cancelled_edit = True

            def _refresh_comp_table(self, schedule_recommendations: bool = False) -> None:
                self.refreshed_with = schedule_recommendations

        dummy = Dummy()
        dummy.loaded_comp_sheet_label = "Working / Lot.xlsx"
        dummy.state = app.BridgeState()
        dummy.state.set_rows([WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="First")])
        dummy.row_sources = {2: "Lot.xlsx"}
        dummy.comp_sheet_sources = {}
        dummy.comp_tree = FakeTree()
        dummy.comp_output_saved = True
        dummy.status_var = Status()

        thread = threading.Thread(target=dummy.add_comp_row, daemon=True)
        thread.start()
        thread.join(1.0)

        self.assertFalse(thread.is_alive(), "add_comp_row deadlocked while updating BridgeState rows")
        self.assertEqual([row.excel_row for row in dummy.state.rows], [2, 3])
        self.assertFalse(dummy.comp_output_saved)
        self.assertEqual(dummy.row_sources[3], "Lot.xlsx")


    def test_delete_selected_comp_rows_rekeys_sources_for_save_back(self) -> None:
        class FakeTree:
            def selection(self) -> tuple[str, ...]:
                return ("3",)

        class Status:
            def __init__(self) -> None:
                self.value = ""

            def set(self, value: str) -> None:
                self.value = value

        class Dummy:
            delete_selected_comp_rows = app.CardPipelineApp.delete_selected_comp_rows
            _delete_selected_rows = app.CardPipelineApp._delete_selected_rows

            def _cancel_cell_edit(self) -> None:
                self.cancelled_edit = True

            def _refresh_table(self, schedule_recommendations: bool = False) -> None:
                self.refreshed_with = schedule_recommendations

        dummy = Dummy()
        dummy.comp_tree = FakeTree()
        dummy.intake_tree = object()
        dummy._is_review_row_tree = lambda _tree: False
        dummy.state = app.BridgeState()
        dummy.state.set_rows(
            [
                WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="First"),
                WorkbookRow(excel_row=3, cert_number="2", grader="PSA", card_title="Delete Me"),
                WorkbookRow(excel_row=4, cert_number="3", grader="PSA", card_title="Last"),
            ]
        )
        dummy.row_sources = {2: "source-2", 3: "source-3", 4: "source-4"}
        dummy.comp_sheet_sources = {2: "sheet-2", 3: "sheet-3", 4: "sheet-4"}
        dummy.status_var = Status()
        dummy.comp_output_saved = True

        dummy.delete_selected_comp_rows()

        self.assertEqual([row.cert_number for row in dummy.state.rows], ["1", "3"])
        self.assertEqual([row.excel_row for row in dummy.state.rows], [2, 3])
        self.assertEqual(dummy.row_sources, {2: "source-2", 3: "source-4"})
        self.assertEqual(dummy.comp_sheet_sources, {2: "sheet-2", 3: "sheet-4"})
        self.assertFalse(dummy.comp_output_saved)
        self.assertIn("Save back to source sheet", dummy.status_var.value)

    def test_unassigned_player_is_recorded_for_unmatched_valued_row(self) -> None:
        class Dummy:
            _load_unassigned_players = app.CardPipelineApp._load_unassigned_players
            _save_unassigned_players = app.CardPipelineApp._save_unassigned_players
            _record_unassigned_player = app.CardPipelineApp._record_unassigned_player
            _append_unique_sample = app.CardPipelineApp._append_unique_sample
            _guess_unassigned_player = app.CardPipelineApp._guess_unassigned_player
            _strip_card_variant_tail = app.CardPipelineApp._strip_card_variant_tail

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_unassigned = app.UNASSIGNED_PLAYERS_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.UNASSIGNED_PLAYERS_PATH = Path(tmp) / "unassigned_players.json"
            dummy = Dummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            row = WorkbookRow(
                excel_row=2,
                cert_number="",
                grader="PSA",
                card_title="2024 Panini Prizm 12 Jalen Madeup PSA 10",
                card_ladder_comps_average=25,
            )
            try:
                dummy._record_unassigned_player(row)
                entries = json.loads(app.UNASSIGNED_PLAYERS_PATH.read_text(encoding="utf-8"))["entries"]
                self.assertIn("jalen madeup", entries)
                self.assertEqual(entries["jalen madeup"]["player"], "Jalen Madeup")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.UNASSIGNED_PLAYERS_PATH = old_unassigned

    def test_unassigned_player_records_failed_assignment_even_with_wrong_partial_match(self) -> None:
        class Dummy:
            _load_unassigned_players = app.CardPipelineApp._load_unassigned_players
            _save_unassigned_players = app.CardPipelineApp._save_unassigned_players
            _record_unassigned_player = app.CardPipelineApp._record_unassigned_player
            _append_unique_sample = app.CardPipelineApp._append_unique_sample
            _guess_unassigned_player = app.CardPipelineApp._guess_unassigned_player
            _strip_card_variant_tail = app.CardPipelineApp._strip_card_variant_tail

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_unassigned = app.UNASSIGNED_PLAYERS_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.UNASSIGNED_PLAYERS_PATH = Path(tmp) / "unassigned_players.json"
            dummy = Dummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            row = WorkbookRow(
                excel_row=2,
                cert_number="",
                grader="PSA",
                card_title="2022 Bowman Npb 82 Shintaro Fujinami Chrome-Refractor PSA 10",
                card_ladder_comps_average=25,
            )
            try:
                dummy._record_unassigned_player(row)
                entries = json.loads(app.UNASSIGNED_PLAYERS_PATH.read_text(encoding="utf-8"))["entries"]
                self.assertIn("shintaro fujinami", entries)
                self.assertEqual(entries["shintaro fujinami"]["player"], "Shintaro Fujinami")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.UNASSIGNED_PLAYERS_PATH = old_unassigned

    def test_failed_assignment_is_labeled_nobody_takes_and_recorded(self) -> None:
        class Dummy:
            _apply_recommendations = app.CardPipelineApp._apply_recommendations
            _assignment_person_for_row = app.CardPipelineApp._assignment_person_for_row
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _load_unassigned_players = app.CardPipelineApp._load_unassigned_players
            _save_unassigned_players = app.CardPipelineApp._save_unassigned_players
            _record_unassigned_player = app.CardPipelineApp._record_unassigned_player
            _append_unique_sample = app.CardPipelineApp._append_unique_sample
            _guess_unassigned_player = app.CardPipelineApp._guess_unassigned_player
            _strip_card_variant_tail = app.CardPipelineApp._strip_card_variant_tail

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_unassigned = app.UNASSIGNED_PLAYERS_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.UNASSIGNED_PLAYERS_PATH = Path(tmp) / "unassigned_players.json"
            row = WorkbookRow(
                excel_row=2,
                cert_number="",
                grader="PSA",
                card_title="2023 Topps Chrome Platinum Anniversary Autographs Yd Yusniel Diaz PSA 9",
                card_ladder_comps_average=20,
            )
            dummy = Dummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.state = types.SimpleNamespace(rows=[row])
            dummy.review_rows = []
            dummy.assignment_engine = assignment_engine.AssignmentEngine(
                [
                    assignment_engine.AssignmentCompany(
                        "Basketball Buyer",
                        assignment_engine.CompanyRules(ranges=[assignment_engine.AssignmentRule("basketball", 10, 500)]),
                        [assignment_engine.PayoutTier(10, 500, 0.9, "NBA")],
                    )
                ]
            )
            try:
                dummy._apply_recommendations()
                self.assertEqual(row.best_company, app.NO_COMPANY_TAKES_LABEL)
                self.assertIsNone(row.estimated_payout)
                entries = json.loads(app.UNASSIGNED_PLAYERS_PATH.read_text(encoding="utf-8"))["entries"]
                self.assertIn("yusniel diaz", entries)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.UNASSIGNED_PLAYERS_PATH = old_unassigned

    def test_assignment_person_for_row_uses_sheet_marker(self) -> None:
        class Dummy:
            _assignment_person_for_row = app.CardPipelineApp._assignment_person_for_row
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key

        dummy = Dummy()
        dummy.comp_sheet_sources = {}
        dummy.review_sheet_sources = {}
        dummy.intake_sheet_sources = {}
        dummy.row_sources = {2: "Lot A.xlsx"}
        dummy.review_sources = {}
        dummy.intake_sources = {}
        dummy.selected_working_sheet = types.SimpleNamespace(get=lambda: "")
        dummy.home_sheet_markers = {"Working|Lot A.xlsx": {"assigned_person": "Lucas"}}
        row = WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="Test", card_ladder_comps_average=100)

        self.assertEqual(dummy._assignment_person_for_row(row), "Lucas")

    def test_receive_index_reads_working_sheet_assignment_values(self) -> None:
        class FieldVar:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class Dummy:
            refresh_incoming_index = app.CardPipelineApp.refresh_incoming_index
            _incoming_match = app.CardPipelineApp._incoming_match
            _match_all_review_rows = app.CardPipelineApp._match_all_review_rows
            _ensure_receive_row_assignment = app.CardPipelineApp._ensure_receive_row_assignment
            _attach_receive_match_to_row = app.CardPipelineApp._attach_receive_match_to_row
            _incoming_raw_match = app.CardPipelineApp._incoming_raw_match
            _receive_row_ref_key = app.CardPipelineApp._receive_row_ref_key

            def _refresh_table(self, schedule_recommendations=False):
                self.refreshed = True

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            incoming_dir = root / "INCOMING SHEETS"
            working_dir = root / "WORKING SHEETS"
            incoming_dir.mkdir(parents=True)
            working_dir.mkdir(parents=True)
            write_working_sheet(
                working_dir / "Working Lot.xlsx",
                [
                    WorkbookRow(
                        excel_row=2,
                        cert_number="12345678",
                        card_title="Test Card PSA 10",
                        grader="PSA",
                        best_company="Fanatics",
                        estimated_payout=88.0,
                    )
                ],
                {2: "Manual"},
            )

            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            app.INCOMING_SHEETS_DIR = incoming_dir
            app.WORKING_SHEETS_DIR = working_dir
            dummy = Dummy()
            dummy.incoming_cert_index = {}
            dummy.review_rows = [WorkbookRow(excel_row=2, cert_number="12345678", card_title="", grader="")]
            dummy.review_sheet_sources = {}
            dummy.review_status = FieldVar()
            dummy.refreshed = False
            try:
                dummy.refresh_incoming_index()
                self.assertTrue(dummy.refreshed)
                self.assertEqual(dummy.review_rows[0].best_company, "Fanatics")
                self.assertEqual(dummy.review_rows[0].estimated_payout, 88.0)
                self.assertEqual(dummy.review_sheet_sources[2], "Working Lot.xlsx")
                self.assertIn("incoming/working", dummy.review_status.value)
            finally:
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working

    def test_mark_received_can_target_blank_cert_raw_row_by_workbook_row(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "Raw Lot.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Cards"
            sheet.append(["Cert #", "Grader", "Card", "Purchase"])
            sheet.append(["", "", "Raw Test Card", 12])
            workbook.save(path)

            result = mark_received_in_workbooks([path], set(), {("Raw Lot.xlsx", "Cards", 2)})

            self.assertEqual(result["rows_marked"], 1)
            self.assertEqual(result["row_refs_marked"], {("raw lot.xlsx", "cards", 2)})
            saved = load_workbook(path)
            try:
                headers = [cell.value for cell in saved["Cards"][1]]
                received_col = headers.index("RECEIVED") + 1
                self.assertEqual(saved["Cards"].cell(2, received_col).value, "X")
            finally:
                saved.close()

    def test_receive_index_matches_raw_rows_by_unique_title_and_keeps_row_ref(self) -> None:
        class FieldVar:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class Dummy:
            refresh_incoming_index = app.CardPipelineApp.refresh_incoming_index
            _append_review_rows = app.CardPipelineApp._append_review_rows
            _incoming_match = app.CardPipelineApp._incoming_match
            _incoming_raw_match = app.CardPipelineApp._incoming_raw_match
            _receive_row_ref_key = app.CardPipelineApp._receive_row_ref_key
            _receive_row_ref = app.CardPipelineApp._receive_row_ref
            _receive_row_was_marked = app.CardPipelineApp._receive_row_was_marked
            _attach_receive_match_to_row = app.CardPipelineApp._attach_receive_match_to_row
            _match_all_review_rows = app.CardPipelineApp._match_all_review_rows
            _ensure_receive_row_assignment = app.CardPipelineApp._ensure_receive_row_assignment

            def _refresh_table(self, schedule_recommendations=False):
                self.refreshed = True

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            incoming_dir = root / "INCOMING SHEETS"
            working_dir = root / "WORKING SHEETS"
            incoming_dir.mkdir(parents=True)
            working_dir.mkdir(parents=True)
            write_working_sheet(
                incoming_dir / "Raw Lot.xlsx",
                [
                    WorkbookRow(
                        excel_row=2,
                        cert_number="",
                        card_title="Raw Test Card",
                        grader="",
                        existing_value=12,
                    )
                ],
                {2: "Manual"},
            )

            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            app.INCOMING_SHEETS_DIR = incoming_dir
            app.WORKING_SHEETS_DIR = working_dir
            dummy = Dummy()
            dummy.assignment_engine = types.SimpleNamespace(
                recommend=lambda row, person="": assignment_engine.AssignmentRecommendation("", None, None)
            )
            dummy.incoming_cert_index = {}
            dummy.review_rows = []
            dummy.review_sources = {}
            dummy.review_sheet_sources = {}
            dummy.review_status = FieldVar()
            dummy.refreshed = False
            try:
                dummy.refresh_incoming_index()
                self.assertIn("raw row", dummy.review_status.value)
                added = dummy._append_review_rows([{"card_title": "Raw Test Card", "source": "Manual"}])
                self.assertEqual(added, [2])
                self.assertEqual(dummy.review_rows[0].status, "Received")
                self.assertEqual(dummy.review_sheet_sources[2], "Raw Lot.xlsx")
                self.assertEqual(dummy._receive_row_ref(dummy.review_rows[0]), ("Raw Lot.xlsx", "Cards", 2))
            finally:
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working

    def test_manual_receive_can_match_raw_row_by_item_id(self) -> None:
        class FieldVar:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class Dummy:
            refresh_incoming_index = app.CardPipelineApp.refresh_incoming_index
            _append_review_rows = app.CardPipelineApp._append_review_rows
            _incoming_match = app.CardPipelineApp._incoming_match
            _incoming_raw_match = app.CardPipelineApp._incoming_raw_match
            _receive_row_ref_key = app.CardPipelineApp._receive_row_ref_key
            _receive_row_ref = app.CardPipelineApp._receive_row_ref
            _attach_receive_match_to_row = app.CardPipelineApp._attach_receive_match_to_row
            _match_all_review_rows = app.CardPipelineApp._match_all_review_rows
            _ensure_receive_row_assignment = app.CardPipelineApp._ensure_receive_row_assignment
            _row_display_value = app.CardPipelineApp._row_display_value

            def _refresh_table(self, schedule_recommendations=False):
                self.refreshed = True

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            incoming_dir = root / "INCOMING SHEETS"
            working_dir = root / "WORKING SHEETS"
            incoming_dir.mkdir(parents=True)
            working_dir.mkdir(parents=True)
            raw_id = "RAW-20260709-0007"
            write_working_sheet(
                incoming_dir / "Raw Lot.xlsx",
                [
                    WorkbookRow(
                        excel_row=2,
                        cert_number="",
                        item_id=raw_id,
                        card_title="Raw Item ID Test",
                        grader="",
                        existing_value=14,
                    )
                ],
                {2: "Manual"},
            )

            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            app.INCOMING_SHEETS_DIR = incoming_dir
            app.WORKING_SHEETS_DIR = working_dir
            dummy = Dummy()
            dummy.assignment_engine = types.SimpleNamespace(
                recommend=lambda row, person="": assignment_engine.AssignmentRecommendation("", None, None)
            )
            dummy.incoming_cert_index = {}
            dummy.review_rows = []
            dummy.review_sources = {}
            dummy.review_sheet_sources = {}
            dummy.review_status = FieldVar()
            dummy.refreshed = False
            try:
                self.assertIn("item_id", app.RECEIVE_COLUMNS)
                self.assertIn("item_id", app.EDITABLE_COLUMNS)
                dummy.refresh_incoming_index()
                added = dummy._append_review_rows([{"item_id": raw_id, "source": "Manual"}])
                self.assertEqual(added, [2])
                self.assertEqual(dummy.review_rows[0].status, "Received")
                self.assertEqual(dummy.review_rows[0].item_id, raw_id)
                self.assertEqual(dummy.review_rows[0].card_title, "Raw Item ID Test")
                self.assertEqual(dummy.review_sheet_sources[2], "Raw Lot.xlsx")
                self.assertEqual(dummy._receive_row_ref(dummy.review_rows[0]), ("Raw Lot.xlsx", "Cards", 2))
                self.assertEqual(dummy._row_display_value(dummy.review_rows[0], "item_id", {}, {}), raw_id)
            finally:
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working

    def test_receive_scan_accepts_duplicate_raw_item_id_and_adds_all_matches(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        class Dummy:
            add_review_scanned_row = app.CardPipelineApp.add_review_scanned_row
            _append_review_rows = app.CardPipelineApp._append_review_rows
            _incoming_match = app.CardPipelineApp._incoming_match
            _incoming_raw_match = app.CardPipelineApp._incoming_raw_match
            _incoming_raw_matches = app.CardPipelineApp._incoming_raw_matches
            _incoming_index_candidates = app.CardPipelineApp._incoming_index_candidates
            _normalize_receive_search_text = app.CardPipelineApp._normalize_receive_search_text
            _receive_match_to_review_payload = app.CardPipelineApp._receive_match_to_review_payload
            _receive_row_ref_key = app.CardPipelineApp._receive_row_ref_key
            _attach_receive_match_to_row = app.CardPipelineApp._attach_receive_match_to_row
            _ensure_receive_row_assignment = app.CardPipelineApp._ensure_receive_row_assignment

            def refresh_incoming_index(self):
                self.refresh_count += 1

            def _incoming_title_matches(self, query, limit=25):
                return []

            def _arm_review_scanner(self):
                self.armed = True

            def _refresh_table(self, schedule_recommendations=False):
                self.refreshed = True

        raw_id = "RAW-MIKEY-20260710-0001"
        dummy = Dummy()
        dummy.assignment_engine = types.SimpleNamespace(
            recommend=lambda row, person="": assignment_engine.AssignmentRecommendation("", None, None)
        )
        dummy.review_scanning_active = True
        dummy.review_scan_cert = FieldVar(raw_id)
        dummy.review_status = FieldVar()
        dummy.review_rows = []
        dummy.review_sources = {}
        dummy.review_sheet_sources = {}
        dummy.refresh_count = 0
        dummy.armed = False
        dummy.refreshed = False
        dummy.incoming_cert_index = {
            "raw:caliken.xlsx:cards:2": {
                "item_id": raw_id,
                "sheet": "caliken.xlsx",
                "workbook_sheet": "Cards",
                "workbook_row": 2,
                "card_title": "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1",
                "sport": "baseball",
                "receive_key": "raw:caliken.xlsx:cards:2",
            },
            "raw:owen.xlsx:cards:2": {
                "item_id": raw_id,
                "sheet": "owen.xlsx",
                "workbook_sheet": "Cards",
                "workbook_row": 2,
                "card_title": "2015 Panini Immaculate Kobe Bryant Auto /60",
                "sport": "basketball",
                "receive_key": "raw:owen.xlsx:cards:2",
            },
        }

        dummy.add_review_scanned_row()

        self.assertEqual(len(dummy.review_rows), 2)
        self.assertEqual(dummy.review_rows[0].card_title, "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1")
        self.assertEqual(dummy.review_rows[1].card_title, "2015 Panini Immaculate Kobe Bryant Auto /60")
        self.assertEqual(dummy.review_sheet_sources[2], "caliken.xlsx")
        self.assertEqual(dummy.review_sheet_sources[3], "owen.xlsx")
        self.assertIn("Matched 2 incoming row", dummy.review_status.value)
        self.assertEqual(dummy.review_scan_cert.value, "")

    def test_receive_scan_accepts_card_description_search(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        class Dummy:
            add_review_scanned_row = app.CardPipelineApp.add_review_scanned_row
            _append_review_rows = app.CardPipelineApp._append_review_rows
            _incoming_match = app.CardPipelineApp._incoming_match
            _incoming_raw_match = app.CardPipelineApp._incoming_raw_match
            _incoming_raw_matches = app.CardPipelineApp._incoming_raw_matches
            _incoming_index_candidates = app.CardPipelineApp._incoming_index_candidates
            _normalize_receive_search_text = app.CardPipelineApp._normalize_receive_search_text
            _incoming_title_matches = app.CardPipelineApp._incoming_title_matches
            _receive_match_to_review_payload = app.CardPipelineApp._receive_match_to_review_payload
            _receive_row_ref_key = app.CardPipelineApp._receive_row_ref_key
            _attach_receive_match_to_row = app.CardPipelineApp._attach_receive_match_to_row
            _ensure_receive_row_assignment = app.CardPipelineApp._ensure_receive_row_assignment

            def refresh_incoming_index(self):
                self.refresh_count += 1

            def _arm_review_scanner(self):
                self.armed = True

            def _refresh_table(self, schedule_recommendations=False):
                self.refreshed = True

        dummy = Dummy()
        dummy.assignment_engine = types.SimpleNamespace(
            recommend=lambda row, person="": assignment_engine.AssignmentRecommendation("", None, None)
        )
        dummy.review_scanning_active = True
        dummy.review_scan_cert = FieldVar("gunnar henderson")
        dummy.review_status = FieldVar()
        dummy.review_rows = []
        dummy.review_sources = {}
        dummy.review_sheet_sources = {}
        dummy.refresh_count = 0
        dummy.armed = False
        dummy.refreshed = False
        dummy.incoming_cert_index = {
            "raw:gunnar.xlsx:cards:2": {
                "item_id": "RAW-MIKEY-20260710-0007",
                "sheet": "gunnar.xlsx",
                "workbook_sheet": "Cards",
                "workbook_row": 2,
                "card_title": "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1",
                "sport": "baseball",
                "receive_key": "raw:gunnar.xlsx:cards:2",
            },
            "raw:kobe.xlsx:cards:2": {
                "item_id": "RAW-MIKEY-20260710-0008",
                "sheet": "kobe.xlsx",
                "workbook_sheet": "Cards",
                "workbook_row": 2,
                "card_title": "2015 Panini Immaculate Kobe Bryant Auto /60",
                "sport": "basketball",
                "receive_key": "raw:kobe.xlsx:cards:2",
            },
        }

        dummy.add_review_scanned_row()

        self.assertEqual(len(dummy.review_rows), 1)
        self.assertEqual(dummy.review_rows[0].item_id, "RAW-MIKEY-20260710-0007")
        self.assertEqual(dummy.review_rows[0].card_title, "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1")
        self.assertEqual(dummy.review_sheet_sources[2], "gunnar.xlsx")
        self.assertIn("Matched 1 incoming row", dummy.review_status.value)

    def test_receive_autocomplete_labels_and_resolves_selected_match(self) -> None:
        class Dummy:
            _incoming_index_candidates = app.CardPipelineApp._incoming_index_candidates
            _incoming_title_matches = app.CardPipelineApp._incoming_title_matches
            _normalize_receive_search_text = app.CardPipelineApp._normalize_receive_search_text
            _receive_match_option_label = app.CardPipelineApp._receive_match_option_label
            _selected_receive_autocomplete_match = app.CardPipelineApp._selected_receive_autocomplete_match

        dummy = Dummy()
        dummy.receive_cell_autocomplete_matches = {}
        match = {
            "item_id": "RAW-MIKEY-20260710-0007",
            "sheet": "gunnar.xlsx",
            "card_title": "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1",
            "sport": "baseball",
            "receive_key": "raw:gunnar.xlsx:cards:2",
        }
        label = dummy._receive_match_option_label(match)
        dummy.receive_cell_autocomplete_matches[label] = match
        dummy.incoming_cert_index = {"raw:gunnar.xlsx:cards:2": match}

        self.assertEqual(label, "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1 | gunnar.xlsx | RAW-MIKEY-20260710-0007")
        self.assertEqual(dummy._selected_receive_autocomplete_match(label), match)

    def test_receive_autocomplete_applies_selected_match_to_existing_row(self) -> None:
        class FieldVar:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class Dummy:
            _apply_receive_match_to_existing_row = app.CardPipelineApp._apply_receive_match_to_existing_row
            _attach_receive_match_to_row = app.CardPipelineApp._attach_receive_match_to_row
            _receive_row_ref_key = app.CardPipelineApp._receive_row_ref_key
            _ensure_receive_row_assignment = app.CardPipelineApp._ensure_receive_row_assignment
            _is_review_row_tree = lambda self, tree: True

        row = WorkbookRow(excel_row=2, cert_number="", item_id="", grader="", card_title="")
        dummy = Dummy()
        dummy.review_rows = [row]
        dummy.review_sheet_sources = {}
        dummy.review_status = FieldVar()
        dummy.assignment_engine = types.SimpleNamespace(
            recommend=lambda row, person="": assignment_engine.AssignmentRecommendation("", None, None)
        )
        match = {
            "item_id": "RAW-MIKEY-20260710-0007",
            "sheet": "gunnar.xlsx",
            "workbook_sheet": "Cards",
            "workbook_row": 2,
            "card_title": "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1",
            "sport": "baseball",
            "purchase_price": 400,
            "best_company": "Arena Club",
            "estimated_payout": 380,
        }

        dummy._apply_receive_match_to_existing_row(object(), 2, match)

        self.assertEqual(row.item_id, "RAW-MIKEY-20260710-0007")
        self.assertEqual(row.card_title, "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1")
        self.assertEqual(row.category, "baseball")
        self.assertEqual(row.existing_value, 400)
        self.assertEqual(row.best_company, "Arena Club")
        self.assertEqual(row.estimated_payout, 380)
        self.assertEqual(dummy.review_sheet_sources[2], "gunnar.xlsx")
        self.assertEqual(getattr(row, "_receive_workbook_row"), 2)

    def test_receive_barcode_refreshes_stale_match_without_assignment_values(self) -> None:
        class Dummy:
            _append_review_rows = app.CardPipelineApp._append_review_rows
            _incoming_match = app.CardPipelineApp._incoming_match
            _ensure_receive_row_assignment = app.CardPipelineApp._ensure_receive_row_assignment
            _attach_receive_match_to_row = app.CardPipelineApp._attach_receive_match_to_row
            _incoming_raw_match = app.CardPipelineApp._incoming_raw_match
            _receive_row_ref_key = app.CardPipelineApp._receive_row_ref_key

            def refresh_incoming_index(self):
                self.refresh_count += 1
                self.incoming_cert_index["12345678"] = {
                    "sheet": "Assigned Lot.xlsx",
                    "card_title": "Assigned Test Card PSA 10",
                    "grader": "PSA",
                    "best_company": "Fanatics",
                    "estimated_payout": 88.0,
                }

            def _refresh_table(self, schedule_recommendations=False):
                self.refreshed = True

        dummy = Dummy()
        dummy.assignment_engine = types.SimpleNamespace(
            recommend=lambda row, person="": assignment_engine.AssignmentRecommendation("Fanatics", 88.0, 100.0)
        )
        dummy.incoming_cert_index = {
            "12345678": {
                "sheet": "Thin Startup Lot.xlsx",
                "card_title": "Assigned Test Card PSA 10",
                "grader": "PSA",
                "best_company": "",
                "estimated_payout": None,
            }
        }
        dummy.review_rows = []
        dummy.review_sources = {}
        dummy.review_sheet_sources = {}
        dummy.refresh_count = 0
        dummy.refreshed = False

        dummy._append_review_rows([{"cert_number": "12345678", "source": "Receive Barcode", "notes": "Received"}])

        self.assertEqual(dummy.refresh_count, 1)
        self.assertTrue(dummy.refreshed)
        self.assertEqual(dummy.review_rows[0].best_company, "Fanatics")
        self.assertEqual(dummy.review_rows[0].estimated_payout, 88.0)
        self.assertEqual(dummy.review_sheet_sources[2], "Assigned Lot.xlsx")

    def test_receive_row_recalculates_assignment_when_sheet_match_has_values_but_no_company(self) -> None:
        class Dummy:
            _append_review_rows = app.CardPipelineApp._append_review_rows
            _incoming_match = app.CardPipelineApp._incoming_match
            _ensure_receive_row_assignment = app.CardPipelineApp._ensure_receive_row_assignment
            _attach_receive_match_to_row = app.CardPipelineApp._attach_receive_match_to_row
            _incoming_raw_match = app.CardPipelineApp._incoming_raw_match
            _receive_row_ref_key = app.CardPipelineApp._receive_row_ref_key

            def refresh_incoming_index(self):
                self.refresh_count += 1

            def _refresh_table(self, schedule_recommendations=False):
                self.refreshed = True

        dummy = Dummy()
        dummy.assignment_engine = types.SimpleNamespace(
            recommend=lambda row, person="": assignment_engine.AssignmentRecommendation("Arena Club", 369.89, 410.99)
        )
        dummy.incoming_cert_index = {
            "21366909": {
                "sheet": "NASHVILLE_KEVIN_HAMBONE.xlsx",
                "card_title": "1989 Star Griffey Jr. 10 Ken Griffey Jr. Mariners-Yellow Back PSA 10",
                "grader": "PSA",
                "purchase_price": 350.0,
                "card_ladder_value": 410.99,
                "card_ladder_comps_average": 410.99,
                "best_company": "",
                "estimated_payout": None,
            }
        }
        dummy.review_rows = []
        dummy.review_sources = {}
        dummy.review_sheet_sources = {}
        dummy.refresh_count = 0
        dummy.refreshed = False

        dummy._append_review_rows([{"cert_number": "21366909", "source": "Receive Barcode", "notes": "Received"}])

        self.assertEqual(dummy.refresh_count, 1)
        self.assertEqual(dummy.review_rows[0].best_company, "Arena Club")
        self.assertEqual(dummy.review_rows[0].estimated_payout, 369.89)

    def test_receive_row_recomputes_stale_stored_assignment_values(self) -> None:
        class Dummy:
            _ensure_receive_row_assignment = app.CardPipelineApp._ensure_receive_row_assignment

        row = WorkbookRow(
            excel_row=14,
            cert_number="152491672",
            grader="PSA",
            category="baseball",
            card_title="2018 Topps Allen and Ginter World Talent #WT-24 Shohei Ohtani World Talent PSA 10",
            existing_value=430,
            card_ladder_value=415.03,
            card_ladder_comps_average=500.02,
            best_company="FANATICS",
            estimated_payout=385.98,
        )
        dummy = Dummy()
        dummy.assignment_engine = types.SimpleNamespace(
            recommend=lambda row, person="": assignment_engine.AssignmentRecommendation("ARENA CLUB", 460.02, 500.02)
        )

        dummy._ensure_receive_row_assignment(row)

        self.assertEqual(row.best_company, "ARENA CLUB")
        self.assertEqual(row.estimated_payout, 460.02)

    def test_scoped_assignment_results_leave_unreturned_rows_unchanged(self) -> None:
        class FieldVar:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class Dummy:
            _apply_assignment_recommendation_results = app.CardPipelineApp._apply_assignment_recommendation_results

            def __init__(self):
                self.assignment_recommendation_job = 7
                self.assignment_recommendation_running = True
                self.assignment_progress_value = FieldVar()
                self.review_status = FieldVar()
                self.status_var = FieldVar()
                self.comp_output_saved = True
                self.review_rows = []
                self.state = types.SimpleNamespace(
                    rows=[
                        WorkbookRow(1, "111", "Changed Card", "PSA", best_company="Old A", estimated_payout=1),
                        WorkbookRow(2, "222", "Untouched Card", "PSA", best_company="Keep Me", estimated_payout=22),
                    ]
                )
                self.refreshed_comp = False

            def _record_unassigned_players(self, rows):
                self.unassigned_rows = rows

            def _refresh_comp_table(self, schedule_recommendations=False):
                self.refreshed_comp = True

            def _refresh_table(self, schedule_recommendations=False):
                raise AssertionError("scoped comp-only results should not refresh every table")

        dummy = Dummy()
        target = dummy.state.rows[0]

        dummy._apply_assignment_recommendation_results(
            {"job_id": 7, "total": 1, "results": [(id(target), "Arena Club", 95.0)]}
        )

        self.assertEqual(dummy.state.rows[0].best_company, "Arena Club")
        self.assertEqual(dummy.state.rows[0].estimated_payout, 95.0)
        self.assertEqual(dummy.state.rows[1].best_company, "Keep Me")
        self.assertEqual(dummy.state.rows[1].estimated_payout, 22)
        self.assertTrue(dummy.refreshed_comp)

    def test_nobody_takes_is_not_an_assignable_company(self) -> None:
        class Dummy:
            _row_has_assignable_company = app.CardPipelineApp._row_has_assignable_company

        dummy = Dummy()
        nobody_row = WorkbookRow(excel_row=2, cert_number="", grader="PSA", card_title="Test Card", best_company=app.NO_COMPANY_TAKES_LABEL)
        buyer_row = WorkbookRow(excel_row=3, cert_number="", grader="PSA", card_title="Test Card", best_company="Arena Club")

        self.assertFalse(dummy._row_has_assignable_company(nobody_row))
        self.assertTrue(dummy._row_has_assignable_company(buyer_row))

    def test_unassigned_auto_categorize_uses_web_text_and_saves_override(self) -> None:
        class Dummy:
            _load_unassigned_players = app.CardPipelineApp._load_unassigned_players
            _save_unassigned_players = app.CardPipelineApp._save_unassigned_players
            _auto_categorize_unassigned_players = app.CardPipelineApp._auto_categorize_unassigned_players
            _search_unassigned_player_category = app.CardPipelineApp._search_unassigned_player_category
            _local_unassigned_player_category = app.CardPipelineApp._local_unassigned_player_category
            _category_research_text = app.CardPipelineApp._category_research_text
            _infer_category_from_web_text = app.CardPipelineApp._infer_category_from_web_text
            _write_player_category_override = app.CardPipelineApp._write_player_category_override
            def _wikipedia_search_text(self, _player: str) -> str:
                return "Example Player is an NBA basketball guard who appears on Panini basketball cards."
            def _wikidata_search_text(self, _player: str) -> str:
                return ""
            def _duckduckgo_search_text(self, _query: str) -> str:
                return ""

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_unassigned = app.UNASSIGNED_PLAYERS_PATH
            old_overrides = app.PLAYER_OVERRIDES_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.UNASSIGNED_PLAYERS_PATH = Path(tmp) / "unassigned_players.json"
            app.PLAYER_OVERRIDES_PATH = Path(tmp) / "assignment_player_overrides.json"
            dummy = Dummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                dummy._save_unassigned_players({
                    "example player": {
                        "player": "Example Player",
                        "last_title": "2024 Panini Prizm 12 Example Player PSA 10",
                    }
                })
                result = dummy._auto_categorize_unassigned_players()
                overrides = json.loads(app.PLAYER_OVERRIDES_PATH.read_text(encoding="utf-8"))
                remaining = json.loads(app.UNASSIGNED_PLAYERS_PATH.read_text(encoding="utf-8"))["entries"]
                self.assertEqual(result["resolved"], 1)
                self.assertEqual(overrides["players"]["Example Player"]["sport"], "basketball")
                self.assertEqual(remaining, {})
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.UNASSIGNED_PLAYERS_PATH = old_unassigned
                app.PLAYER_OVERRIDES_PATH = old_overrides

    def test_unassigned_auto_categorize_uses_card_title_before_web(self) -> None:
        class Dummy:
            _load_unassigned_players = app.CardPipelineApp._load_unassigned_players
            _save_unassigned_players = app.CardPipelineApp._save_unassigned_players
            _auto_categorize_unassigned_players = app.CardPipelineApp._auto_categorize_unassigned_players
            _search_unassigned_player_category = app.CardPipelineApp._search_unassigned_player_category
            _local_unassigned_player_category = app.CardPipelineApp._local_unassigned_player_category
            _category_research_text = app.CardPipelineApp._category_research_text
            _infer_category_from_web_text = app.CardPipelineApp._infer_category_from_web_text
            _write_player_category_override = app.CardPipelineApp._write_player_category_override

            def _wikipedia_search_text(self, _player: str) -> str:
                raise AssertionError("web search should not run when card title resolves locally")

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_unassigned = app.UNASSIGNED_PLAYERS_PATH
            old_overrides = app.PLAYER_OVERRIDES_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.UNASSIGNED_PLAYERS_PATH = Path(tmp) / "unassigned_players.json"
            app.PLAYER_OVERRIDES_PATH = Path(tmp) / "assignment_player_overrides.json"
            dummy = Dummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                dummy._save_unassigned_players({
                    "yusniel diaz": {
                        "player": "Yusniel Diaz",
                        "last_title": "2023 Topps Chrome Platinum Anniversary Autographs Yd Yusniel Diaz PSA 9",
                    }
                })
                result = dummy._auto_categorize_unassigned_players()
                overrides = json.loads(app.PLAYER_OVERRIDES_PATH.read_text(encoding="utf-8"))
                self.assertEqual(result["resolved"], 1)
                self.assertEqual(overrides["players"]["Yusniel Diaz"]["sport"], "baseball")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.UNASSIGNED_PLAYERS_PATH = old_unassigned
                app.PLAYER_OVERRIDES_PATH = old_overrides

    def test_sheet_marker_save_merges_latest_and_honors_tombstones(self) -> None:
        class MarkerDummy:
            _load_sheet_markers = app.CardPipelineApp._load_sheet_markers
            _save_sheet_markers = app.CardPipelineApp._save_sheet_markers
            _delete_sheet_marker = app.CardPipelineApp._delete_sheet_marker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_markers = app.SHEET_MARKERS_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.SHEET_MARKERS_PATH = Path(tmp) / "sheet_markers.json"
            app.SHEET_MARKERS_PATH.write_text(
                json.dumps(
                    {
                        "Incoming|A.xlsx": {"assigned_person": "A"},
                        "Incoming|B.xlsx": {"assigned_person": "B"},
                    }
                ),
                encoding="utf-8",
            )
            dummy = MarkerDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.home_sheet_markers = {"Incoming|C.xlsx": {"assigned_person": "C"}}
            dummy.deleted_sheet_marker_keys = {"Incoming|A.xlsx"}
            try:
                dummy._save_sheet_markers()
                saved = json.loads(app.SHEET_MARKERS_PATH.read_text(encoding="utf-8"))
                self.assertNotIn("Incoming|A.xlsx", saved)
                self.assertEqual(saved["Incoming|B.xlsx"]["assigned_person"], "B")
                self.assertEqual(saved["Incoming|C.xlsx"]["assigned_person"], "C")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.SHEET_MARKERS_PATH = old_markers

    def test_home_sheet_preview_data_reads_workbook_contents(self) -> None:
        class PreviewDummy:
            _home_sheet_preview_data = app.CardPipelineApp._home_sheet_preview_data

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "Lot A.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Cards"
            sheet.append(["Certification Number", "Grader", "Card Description"])
            sheet.append(["123", "PSA", "Test Card"])
            workbook.save(path)

            preview = PreviewDummy()._home_sheet_preview_data(path)

        self.assertEqual(preview["sheet_title"], "Cards")
        self.assertEqual(preview["columns"], ["row", "A", "B", "C"])
        self.assertEqual(preview["rows"][0], (1, "Certification Number", "Grader", "Card Description"))
        self.assertEqual(preview["rows"][1], (2, "123", "PSA", "Test Card"))
        self.assertFalse(preview["truncated"])

    def test_create_seller_assignment_pays_only_after_receive(self) -> None:
        class SellerSheetDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _load_sheet_markers = app.CardPipelineApp._load_sheet_markers
            _save_sheet_markers = app.CardPipelineApp._save_sheet_markers
            _delete_sheet_marker = app.CardPipelineApp._delete_sheet_marker
            _marker_for_stage = app.CardPipelineApp._marker_for_stage
            _sheet_path_for_stage = app.CardPipelineApp._sheet_path_for_stage
            _move_home_sheet_to_stage = app.CardPipelineApp._move_home_sheet_to_stage
            _assign_sheet_to_seller = app.CardPipelineApp._assign_sheet_to_seller
            _active_payout_balance = app.CardPipelineApp._active_payout_balance
            _payout_sheet_status = app.CardPipelineApp._payout_sheet_status
            _payout_sheet_items = app.CardPipelineApp._payout_sheet_items
            _sheet_marker_is_seller_payout = app.CardPipelineApp._sheet_marker_is_seller_payout
            _source_sheet_is_seller_payout = app.CardPipelineApp._source_sheet_is_seller_payout
            _sheet_marker_is_seller_payout = app.CardPipelineApp._sheet_marker_is_seller_payout
            _source_sheet_is_seller_payout = app.CardPipelineApp._source_sheet_is_seller_payout
            _realized_profit_totals_by_person_sheet = app.CardPipelineApp._realized_profit_totals_by_person_sheet
            _realized_profit_groups_by_person_sheet = app.CardPipelineApp._realized_profit_groups_by_person_sheet
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _money_value = app.CardPipelineApp._money_value
            _person_for_profit_record = app.CardPipelineApp._person_for_profit_record
            _enrich_profit_records_with_people = app.CardPipelineApp._enrich_profit_records_with_people
            _loose_expense_adjustments_by_person = app.CardPipelineApp._loose_expense_adjustments_by_person
            def _seller_terms_seller_names(self):
                return {"john seller"}
            def _load_profit_ledger(self):
                return []

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            incoming_dir = root / "INCOMING SHEETS"
            working_dir = root / "WORKING SHEETS"
            received_dir = root / "RECEIVED SHEETS"
            working_dir.mkdir(parents=True)
            incoming_dir.mkdir()
            received_dir.mkdir()
            (working_dir / "Lot A.xlsx").write_text("placeholder", encoding="utf-8")

            old_pipeline = app.CARD_PIPELINE_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_received = app.RECEIVED_SHEETS_DIR
            old_markers = app.SHEET_MARKERS_PATH
            app.CARD_PIPELINE_DIR = root
            app.INCOMING_SHEETS_DIR = incoming_dir
            app.WORKING_SHEETS_DIR = working_dir
            app.RECEIVED_SHEETS_DIR = received_dir
            app.SHEET_MARKERS_PATH = root / "sheet_markers.json"
            dummy = SellerSheetDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.home_sheet_markers = {}
            dummy.deleted_sheet_marker_keys = set()
            dummy.home_sheet_paths = {"Incoming": {}, "Working": {"Lot A.xlsx": working_dir / "Lot A.xlsx"}, "Received": {}}
            dummy.received_sheet_paths = {}
            dummy.home_sheet_summaries = {}
            try:
                self.assertTrue(dummy._assign_sheet_to_seller("Working", "Lot A.xlsx", "John Seller", "Arena Club", {"rate": 0.9}))
                self.assertEqual(dummy.home_sheet_markers["Working|Lot A.xlsx"]["assigned_person"], "John Seller")
                self.assertTrue(dummy.home_sheet_markers["Working|Lot A.xlsx"]["seller_terms_applied"])
                self.assertEqual(dummy.home_sheet_markers["Working|Lot A.xlsx"]["seller_sheet_type"], "Arena Club")
                self.assertEqual(dummy.home_sheet_markers["Working|Lot A.xlsx"]["seller_rate"], 0.9)

                moved_key, cleanup = dummy._move_home_sheet_to_stage("Working|Lot A.xlsx", "Incoming")
                self.assertEqual(moved_key, "Incoming|Lot A.xlsx")
                self.assertEqual(cleanup, {})
                self.assertEqual(dummy.home_sheet_markers[moved_key]["assigned_person"], "John Seller")
                dummy.home_sheet_paths = {"Incoming": {"Lot A.xlsx": incoming_dir / "Lot A.xlsx"}, "Working": {}, "Received": {}}
                dummy.home_sheet_summaries = {moved_key: {"row_count": 2, "received_count": 0, "purchase_total": 123.45, "estimated_payout_total": 200.0}}

                payout_items = dummy._payout_sheet_items()
                self.assertEqual(payout_items, [])

                received_key, cleanup = dummy._move_home_sheet_to_stage(moved_key, "Received")
                self.assertEqual(received_key, "Received|Lot A.xlsx")
                self.assertEqual(cleanup, {})
                self.assertEqual(dummy.home_sheet_markers[received_key]["assigned_person"], "John Seller")
                dummy.home_sheet_paths = {"Incoming": {}, "Working": {}, "Received": {"Lot A.xlsx": received_dir / "Lot A.xlsx"}}
                dummy.home_sheet_summaries = {received_key: {"row_count": 2, "received_count": 2, "purchase_total": 123.45, "estimated_payout_total": 200.0}}

                payout_items = dummy._payout_sheet_items()
                self.assertEqual(len(payout_items), 1)
                self.assertEqual(payout_items[0]["person"], "John Seller")
                self.assertEqual(payout_items[0]["stage"], "Received")
                self.assertEqual(payout_items[0]["purchase_total"], 123.45)
                self.assertEqual(payout_items[0]["payout_balance"], 123.45)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working
                app.RECEIVED_SHEETS_DIR = old_received
                app.SHEET_MARKERS_PATH = old_markers

    def test_active_payout_balance_uses_seller_or_team_member_rule(self) -> None:
        class PayoutDummy:
            _active_payout_balance = app.CardPipelineApp._active_payout_balance

        dummy = PayoutDummy()
        sellers = {"john seller"}

        self.assertEqual(dummy._active_payout_balance("John Seller", 80.0, 150.0, sellers), (80.0, "Seller purchase total"))
        self.assertEqual(dummy._active_payout_balance("Kevin Hambone", 80.0, 150.0, sellers), (0.0, "Team half sold profit"))
        self.assertEqual(dummy._active_payout_balance("Kevin Hambone", 80.0, 150.0, sellers, realized_profit_total=70.0), (35.0, "Team half sold profit"))
        self.assertEqual(dummy._active_payout_balance("Kevin Hambone", 100.0, 80.0, sellers, realized_profit_total=-20.0), (-10.0, "Team half sold profit"))

    def test_sheet_marker_controls_seller_payout_classification(self) -> None:
        class PayoutDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _sheet_marker_is_seller_payout = app.CardPipelineApp._sheet_marker_is_seller_payout
            _source_sheet_is_seller_payout = app.CardPipelineApp._source_sheet_is_seller_payout

            def __init__(self):
                self.home_sheet_markers = {
                    "Received|Seller Lot.xlsx": {"assigned_person": "John Seller", "seller_terms_applied": True},
                    "Received|Team Lot.xlsx": {"assigned_person": "John Seller"},
                }

            def _seller_terms_seller_names(self):
                return {"john seller"}

        dummy = PayoutDummy()
        self.assertTrue(dummy._source_sheet_is_seller_payout("Seller Lot.xlsx", "John Seller"))
        self.assertFalse(dummy._source_sheet_is_seller_payout("Team Lot.xlsx", "John Seller"))
        self.assertTrue(dummy._source_sheet_is_seller_payout("Legacy Missing Marker.xlsx", "John Seller"))

    def test_known_people_includes_seller_terms_people(self) -> None:
        class PeopleDummy:
            _known_people = app.CardPipelineApp._known_people

            def _known_assigned_people(self):
                return ["Kevin Hambone", "Unassigned"]

            def _load_seller_terms(self):
                return [
                    {"seller": "John Seller", "sheet_type": "Arena Club", "rate": 0.9},
                    {"seller": "Unassigned", "sheet_type": "Fanatics", "rate": 0.9},
                ]

            def _load_profit_ledger(self):
                return [{"assigned_person": "Unassigned"}, {"assigned_person": "James Copeland"}]

            def _load_inventory_ledger(self):
                return [{"assigned_person": "Unassigned"}, {"assigned_person": "Tyler Hamlin"}]

        self.assertEqual(PeopleDummy()._known_people(), ["James Copeland", "John Seller", "Kevin Hambone", "Tyler Hamlin"])

    def test_save_working_sheet_requires_valid_network_seller_terms(self) -> None:
        class Var:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class SaveDummy:
            save_working_sheet = app.CardPipelineApp.save_working_sheet
            _network_mode_enabled = app.CardPipelineApp._network_mode_enabled
            _money_value = app.CardPipelineApp._money_value

            def __init__(self):
                self.intake_rows = [WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="Test", existing_value=10)]
                self.working_sheet_title = Var("Network Lot")
                self.create_network_mode_var = Var(True)
                self.seller_terms_seller_var = Var("John Seller")
                self.seller_terms_sheet_type_var = Var("Arena Club")
                self.applied_terms = False

            def _seller_terms_match(self, seller, sheet_type):
                return None

            def _seller_terms_company_price(self, row, company_name, rate=None, deduction=None):
                return 90.0

            def apply_create_seller_terms(self, show_status=True):
                self.applied_terms = True
                return 1

        dummy = SaveDummy()
        with patch.object(app.messagebox, "showinfo") as showinfo:
            dummy.save_working_sheet()
        self.assertTrue(showinfo.called)
        self.assertFalse(dummy.applied_terms)

    def test_seller_terms_pending_until_required_values_exist(self) -> None:
        class SellerSummaryDummy:
            _money_value = app.CardPipelineApp._money_value
            _seller_terms_rate = app.CardPipelineApp._seller_terms_rate
            _seller_terms_match = lambda self, seller, sheet_type: {"seller": seller, "sheet_type": sheet_type, "deduction": 0.1}
            _sheet_marker_is_seller_payout = app.CardPipelineApp._sheet_marker_is_seller_payout
            _seller_term_for_marker = app.CardPipelineApp._seller_term_for_marker
            _seller_terms_value_label = app.CardPipelineApp._seller_terms_value_label
            _seller_terms_company_decision = app.CardPipelineApp._seller_terms_company_decision
            _seller_terms_company_price = app.CardPipelineApp._seller_terms_company_price
            _seller_payout_summary_for_rows = app.CardPipelineApp._seller_payout_summary_for_rows

            def __init__(self, decision):
                self.assignment_engine = types.SimpleNamespace(
                    companies=[types.SimpleNamespace(name="Arena Club", value_source="comps")],
                    evaluate=lambda row: [decision],
                )

        marker = {"assigned_person": "John Seller", "seller_terms_applied": True, "seller_sheet_type": "Arena Club", "seller_deduction": 0.1}
        row = WorkbookRow(excel_row=2, cert_number="137915162", grader="PSA", card_title="Test Card PSA 10")
        pending_decision = types.SimpleNamespace(company="Arena Club", accepted=False, payout=None, source_value=None, reason="missing comp/card ladder value")
        pending = SellerSummaryDummy(pending_decision)._seller_payout_summary_for_rows([row], marker)
        self.assertTrue(pending["seller_payout_pending"])
        self.assertFalse(pending["seller_payout_payable"])
        self.assertEqual(pending["seller_payout_total"], 0.0)
        self.assertIn("Seller owed money but no Arena Club payout input", pending["seller_payout_warning"])

        ready_decision = types.SimpleNamespace(company="Arena Club", accepted=True, payout=95.0, source_value=100.0, reason="accepted")
        ready = SellerSummaryDummy(ready_decision)._seller_payout_summary_for_rows([row], marker)
        self.assertFalse(ready["seller_payout_pending"])
        self.assertTrue(ready["seller_payout_payable"])
        self.assertEqual(ready["seller_payout_total"], 85.0)

    def test_seller_rate_requires_company_acceptance(self) -> None:
        class SellerPriceDummy:
            _seller_terms_company_price = app.CardPipelineApp._seller_terms_company_price
            _seller_terms_company_decision = app.CardPipelineApp._seller_terms_company_decision

            def __init__(self, decision):
                self.assignment_engine = types.SimpleNamespace(evaluate=lambda row: [decision])

        row = WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="Rejected Card")
        rejected = types.SimpleNamespace(company="Arena Club", accepted=False, payout=None, source_value=100.0, reason="card does not match company rules")
        accepted = types.SimpleNamespace(company="Arena Club", accepted=True, payout=None, source_value=100.0, reason="accepted, but no payout tier matched")

        self.assertIsNone(SellerPriceDummy(rejected)._seller_terms_company_price(row, "Arena Club", rate=0.9))
        self.assertEqual(SellerPriceDummy(accepted)._seller_terms_company_price(row, "Arena Club", rate=0.9), 90.0)

    def test_save_home_sheet_markers_preserves_seller_metadata(self) -> None:
        class MarkerDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _sheet_marker_is_seller_payout = app.CardPipelineApp._sheet_marker_is_seller_payout
            _marker_for_stage = app.CardPipelineApp._marker_for_stage
            save_home_sheet_markers = app.CardPipelineApp.save_home_sheet_markers
            _is_personal_lucas = lambda self: False

            def _seller_terms_match(self, seller, sheet_type):
                return {"seller": seller, "sheet_type": sheet_type, "rate": 0.85, "deduction": None}

            def _move_received_sheet_to_incoming(self, key):
                return ""

            def _move_sheet_to_received(self, key):
                return ""

            def _move_working_sheet_to_incoming(self, key):
                return ""

            def _retarget_inventory_rows_for_source(self, source_sheet, assigned_person):
                return 0

            def _save_sheet_markers(self):
                self.saved = True

            def refresh_working_sheets(self):
                pass

            def refresh_received_sheets(self):
                pass

            def refresh_incoming_index(self):
                pass

            def refresh_home(self):
                pass

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            app.CARD_PIPELINE_DIR = Path(tmp)
            dummy = MarkerDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.home_selected_sheet_key = "Working|Lot.xlsx"
            dummy.home_sheet_kind = types.SimpleNamespace(set=lambda value: None)
            dummy.status_var = types.SimpleNamespace(set=lambda value: None)
            dummy.home_sheet_markers = {
                "Working|Lot.xlsx": {
                    "assigned_person": "John Seller",
                    "seller_terms_applied": True,
                    "seller_sheet_type": "Arena Club",
                    "seller_rate": 0.9,
                    "paid": False,
                }
            }
            dummy.home_sheet_summaries = {"Working|Lot.xlsx": {"row_count": 1}}
            try:
                dummy.save_home_sheet_markers(
                    {
                        "incoming_proper": False,
                        "tracking_number": "TRACK",
                        "all_received": False,
                        "assigned_person": "John Seller",
                    }
                )
                marker = dummy.home_sheet_markers["Working|Lot.xlsx"]
                self.assertTrue(marker["seller_terms_applied"])
                self.assertEqual(marker["seller_sheet_type"], "Arena Club")
                self.assertEqual(marker["seller_rate"], 0.9)
                self.assertEqual(marker["tracking_number"], "TRACK")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline

    def test_payout_history_filters_person_and_keeps_paid_rows(self) -> None:
        class Dummy:
            _payout_history_items_for_person = app.CardPipelineApp._payout_history_items_for_person

            def _payout_sheet_items(self):
                return [
                    {"key": "Sold|Kevin Hambone|Open.xlsx", "person": "Kevin Hambone", "name": "Open.xlsx", "stage": "Sold", "paid": False, "row_count": 1, "net_profit_total": 20, "payout_balance": 10},
                    {"key": "Sold|Kevin Hambone|Paid A.xlsx", "person": "Kevin Hambone", "name": "Paid A.xlsx", "stage": "Sold", "paid": True, "paid_at": "2026-07-14T10:00:00", "row_count": 1, "net_profit_total": 40, "payout_balance": 20},
                    {"key": "Sold|Kevin Hambone|Paid B.xlsx", "person": "Kevin Hambone", "name": "Paid B.xlsx", "stage": "Sold", "paid": True, "paid_at": "2026-07-14T10:00:00", "row_count": 2, "net_profit_total": 60, "payout_balance": 30},
                    {"key": "Sold|James Copeland|Other.xlsx", "person": "James Copeland", "name": "Other.xlsx", "stage": "Sold", "paid": True, "paid_at": "2026-07-14T11:00:00", "row_count": 1, "payout_balance": 30},
                ]

            home_sheet_markers = {}

        items = Dummy()._payout_history_items_for_person("Kevin Hambone")

        self.assertEqual([item["name"] for item in items], ["Total paid at 2026-07-14T10:00:00", "Open.xlsx"])
        self.assertEqual(items[0]["row_count"], 3)
        self.assertEqual(items[0]["payout_balance"], 50.0)
        self.assertEqual(sum(float(item["payout_balance"]) for item in items), 60.0)

    def test_save_payout_marker_blocks_pending_seller_paid(self) -> None:
        class PayoutDummy:
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            save_payout_sheet_marker = app.CardPipelineApp.save_payout_sheet_marker

            def _payout_item_for_key(self, key):
                return {"key": key, "payable": False, "status": "Seller owed money but no Comps input"}

        dummy = PayoutDummy()
        dummy.home_sheet_markers = {"Received|Lot.xlsx": {"assigned_person": "John Seller", "paid": False}}
        dummy.home_sheet_summaries = {"Received|Lot.xlsx": {"all_received": True}}
        with patch.object(app.messagebox, "showinfo") as showinfo:
            dummy.save_payout_sheet_marker("Received|Lot.xlsx", "John Seller", True)
        self.assertTrue(showinfo.called)
        self.assertFalse(dummy.home_sheet_markers["Received|Lot.xlsx"]["paid"])

    def test_save_comp_to_source_sheet_recalculates_seller_purchase(self) -> None:
        class Var:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class SourceSaveDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _sheet_marker_is_seller_payout = app.CardPipelineApp._sheet_marker_is_seller_payout
            _seller_terms_rate = app.CardPipelineApp._seller_terms_rate
            _seller_terms_match = lambda self, seller, sheet_type: {"seller": seller, "sheet_type": sheet_type, "rate": 0.9}
            _seller_term_for_marker = app.CardPipelineApp._seller_term_for_marker
            _seller_terms_company_decision = app.CardPipelineApp._seller_terms_company_decision
            _seller_terms_company_price = app.CardPipelineApp._seller_terms_company_price
            _marker_for_sheet_name = app.CardPipelineApp._marker_for_sheet_name
            _apply_seller_terms_to_rows_for_marker = app.CardPipelineApp._apply_seller_terms_to_rows_for_marker
            _comp_sheet_info = app.CardPipelineApp._comp_sheet_info
            save_comp_to_source_sheet = app.CardPipelineApp.save_comp_to_source_sheet

            def _refresh_table(self, schedule_recommendations=False):
                pass

            def refresh_home(self):
                pass

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            working_dir = root / "WORKING SHEETS"
            working_dir.mkdir(parents=True)
            path = working_dir / "Seller Lot.xlsx"
            write_working_sheet(path, [WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="Card", existing_value=None, card_ladder_comps_average=100.0)], {2: "manual"})

            old_pipeline = app.CARD_PIPELINE_DIR
            app.CARD_PIPELINE_DIR = root
            dummy = SourceSaveDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            row = WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="Card", existing_value=None, card_ladder_comps_average=100.0)
            dummy.state = types.SimpleNamespace(rows=[row])
            dummy.row_sources = {2: "manual"}
            dummy.selected_working_sheet = Var("Seller Lot.xlsx")
            dummy.working_sheet_paths = {"Seller Lot.xlsx": path}
            dummy.home_sheet_markers = {"Working|Seller Lot.xlsx": {"assigned_person": "John Seller", "seller_terms_applied": True, "seller_sheet_type": "Arena Club", "seller_rate": 0.9}}
            dummy.assignment_engine = types.SimpleNamespace(
                evaluate=lambda _row: [types.SimpleNamespace(company="Arena Club", accepted=True, payout=95.0, source_value=100.0, reason="accepted")]
            )
            dummy.status_var = Var("")
            dummy.status_var.set = lambda value: setattr(dummy.status_var, "value", value)
            try:
                dummy.save_comp_to_source_sheet()
                saved = read_simple_spreadsheet(path)
                self.assertEqual(saved[0]["purchase_price"], 90.0)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline

    def test_comp_sheet_list_combines_incoming_and_working_sheets(self) -> None:
        class FakeListbox:
            def __init__(self):
                self.items: list[str] = []

            def delete(self, _start, _end):
                self.items = []

            def insert(self, _index, value):
                self.items.append(value)

        class CompSheetDummy:
            _comp_sheet_label = app.CardPipelineApp._comp_sheet_label
            _populate_comp_sheet_list = app.CardPipelineApp._populate_comp_sheet_list

        dummy = CompSheetDummy()
        dummy.incoming_sheet_paths = {"Incoming Lot.xlsx": Path("/tmp/Incoming Lot.xlsx")}
        dummy.working_sheet_paths = {"Working Lot.xlsx": Path("/tmp/Working Lot.xlsx")}
        dummy.working_sheet_list = FakeListbox()

        dummy._populate_comp_sheet_list()

        self.assertEqual(dummy.working_sheet_list.items, ["Incoming / Incoming Lot.xlsx", "Working / Working Lot.xlsx"])
        self.assertEqual(dummy.comp_sheet_stages["Incoming / Incoming Lot.xlsx"], "Incoming")
        self.assertEqual(dummy.comp_sheet_paths["Working / Working Lot.xlsx"], Path("/tmp/Working Lot.xlsx"))

    def test_save_comp_to_source_sheet_can_write_incoming_sheet(self) -> None:
        class Var:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class SourceSaveDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _sheet_marker_is_seller_payout = app.CardPipelineApp._sheet_marker_is_seller_payout
            _seller_terms_rate = app.CardPipelineApp._seller_terms_rate
            _seller_term_for_marker = app.CardPipelineApp._seller_term_for_marker
            _seller_terms_company_decision = app.CardPipelineApp._seller_terms_company_decision
            _seller_terms_company_price = app.CardPipelineApp._seller_terms_company_price
            _marker_for_sheet_name = app.CardPipelineApp._marker_for_sheet_name
            _apply_seller_terms_to_rows_for_marker = app.CardPipelineApp._apply_seller_terms_to_rows_for_marker
            _comp_sheet_info = app.CardPipelineApp._comp_sheet_info
            save_comp_to_source_sheet = app.CardPipelineApp.save_comp_to_source_sheet

            def _refresh_table(self, schedule_recommendations=False):
                pass

            def refresh_home(self):
                pass

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            incoming_dir = root / "INCOMING SHEETS"
            incoming_dir.mkdir(parents=True)
            path = incoming_dir / "Incoming Lot.xlsx"
            write_working_sheet(path, [WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="Card", existing_value=12.0)], {2: "manual"})

            old_pipeline = app.CARD_PIPELINE_DIR
            app.CARD_PIPELINE_DIR = root
            dummy = SourceSaveDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            row = WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="Card", existing_value=22.0, card_ladder_value=44.0)
            dummy.state = types.SimpleNamespace(rows=[row])
            dummy.row_sources = {2: "manual"}
            dummy.selected_working_sheet = Var("Incoming / Incoming Lot.xlsx")
            dummy.comp_sheet_paths = {"Incoming / Incoming Lot.xlsx": path}
            dummy.comp_sheet_stages = {"Incoming / Incoming Lot.xlsx": "Incoming"}
            dummy.working_sheet_paths = {}
            dummy.incoming_sheet_paths = {"Incoming Lot.xlsx": path}
            dummy.home_sheet_markers = {}
            dummy.status_var = Var("")
            dummy.status_var.set = lambda value: setattr(dummy.status_var, "value", value)
            try:
                dummy.save_comp_to_source_sheet()
                saved = read_simple_spreadsheet(path)
                self.assertEqual(saved[0]["purchase_price"], 22.0)
                self.assertEqual(saved[0]["card_ladder_value"], 44.0)
                self.assertIn("incoming Incoming Lot.xlsx", dummy.status_var.value)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline

    def test_seller_terms_health_reports_duplicates_and_company_issues(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "seller_terms.csv"
            path.write_text(
                "Seller,Sheet Type,Seller Rate,Deduction\n"
                "John,Arena Club,90%,\n"
                "John,Arena Club,85%,\n"
                "Mary,Fanatics,,5%\n"
                "No Company,Unknown,90%,\n"
                "Bad Rate,Arena Club,nope,\n",
                encoding="utf-8",
            )
            lines = assignment_config_ui.seller_terms_health_lines(
                path,
                [
                    {"name": "Arena Club", "active": True},
                    {"name": "Fanatics", "active": False},
                ],
            )
        text = "\n".join(lines)
        self.assertIn("3 valid row(s)", text)
        self.assertIn("duplicate Seller/Sheet Type", text)
        self.assertIn("inactive assignment company", text)
        self.assertIn("is not an assignment company", text)
        self.assertIn("invalid Seller Rate", text)

    def test_people_rules_percent_fields_display_as_numbers_only(self) -> None:
        self.assertEqual(assignment_config_ui.seller_terms_percent_display("10%"), "10")
        self.assertEqual(assignment_config_ui.seller_terms_percent_display("0.1"), "10")
        self.assertEqual(assignment_config_ui.seller_terms_percent_display("90"), "90")
        self.assertTrue(assignment_config_ui.seller_terms_percent_input_is_number("10.5"))
        self.assertFalse(assignment_config_ui.seller_terms_percent_input_is_number("10%"))
        self.assertFalse(assignment_config_ui.seller_terms_percent_input_is_number("ten"))

    def test_paid_received_sheets_archive_after_two_weeks_only_when_paid(self) -> None:
        class ArchiveDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _load_sheet_markers = app.CardPipelineApp._load_sheet_markers
            _save_sheet_markers = app.CardPipelineApp._save_sheet_markers
            _delete_sheet_marker = app.CardPipelineApp._delete_sheet_marker
            _archive_eligible_received_sheets = app.CardPipelineApp._archive_eligible_received_sheets
            _received_sheet_is_archive_age = app.CardPipelineApp._received_sheet_is_archive_age
            _received_at_for_archive = app.CardPipelineApp._received_at_for_archive
            _unique_archive_path = app.CardPipelineApp._unique_archive_path

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            received_dir = root / "RECEIVED SHEETS"
            archive_dir = root / "ARCHIVED SHEETS"
            received_dir.mkdir()
            paid_old = received_dir / "paid-old.xlsx"
            unpaid_old = received_dir / "unpaid-old.xlsx"
            paid_recent = received_dir / "paid-recent.xlsx"
            for path in (paid_old, unpaid_old, paid_recent):
                path.write_text("placeholder", encoding="utf-8")

            old_pipeline = app.CARD_PIPELINE_DIR
            old_received = app.RECEIVED_SHEETS_DIR
            old_archive = app.ARCHIVED_SHEETS_DIR
            old_markers = app.SHEET_MARKERS_PATH
            app.CARD_PIPELINE_DIR = root
            app.RECEIVED_SHEETS_DIR = received_dir
            app.ARCHIVED_SHEETS_DIR = archive_dir
            app.SHEET_MARKERS_PATH = root / "sheet_markers.json"
            old_date = "2026-05-01T12:00:00"
            recent_date = datetime.now().isoformat(timespec="seconds")
            app.SHEET_MARKERS_PATH.write_text(
                json.dumps(
                    {
                        "Received|paid-old.xlsx": {"paid": True, "all_received": True, "received_at": old_date},
                        "Received|unpaid-old.xlsx": {"paid": False, "all_received": True, "received_at": old_date},
                        "Received|paid-recent.xlsx": {"paid": True, "all_received": True, "received_at": recent_date},
                    }
                ),
                encoding="utf-8",
            )
            dummy = ArchiveDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.home_sheet_markers = dummy._load_sheet_markers()
            dummy.deleted_sheet_marker_keys = set()
            try:
                archived = dummy._archive_eligible_received_sheets()
                saved = json.loads(app.SHEET_MARKERS_PATH.read_text(encoding="utf-8"))

                self.assertEqual(archived, ["paid-old.xlsx"])
                self.assertFalse(paid_old.exists())
                self.assertTrue((archive_dir / "paid-old.xlsx").exists())
                self.assertTrue(unpaid_old.exists())
                self.assertTrue(paid_recent.exists())
                self.assertNotIn("Received|paid-old.xlsx", saved)
                self.assertIn("Archived|paid-old.xlsx", saved)
                self.assertIn("Received|unpaid-old.xlsx", saved)
                self.assertIn("Received|paid-recent.xlsx", saved)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.RECEIVED_SHEETS_DIR = old_received
                app.ARCHIVED_SHEETS_DIR = old_archive
                app.SHEET_MARKERS_PATH = old_markers

    def test_deleted_sheet_file_archives_for_two_weeks(self) -> None:
        class DeleteArchiveDummy:
            _deleted_archive_metadata_path = app.CardPipelineApp._deleted_archive_metadata_path
            _unique_deleted_archive_path = app.CardPipelineApp._unique_deleted_archive_path
            _archive_deleted_file = app.CardPipelineApp._archive_deleted_file
            _purge_expired_deleted_archive = app.CardPipelineApp._purge_expired_deleted_archive

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            sheet = root / "WORKING SHEETS" / "Lot A.xlsx"
            sheet.parent.mkdir(parents=True)
            sheet.write_text("placeholder", encoding="utf-8")
            old_deleted_archive = app.DELETED_ARCHIVE_DIR
            old_deleted_sheets = app.DELETED_SHEETS_DIR
            app.DELETED_ARCHIVE_DIR = root / "DELETED ARCHIVE"
            app.DELETED_SHEETS_DIR = app.DELETED_ARCHIVE_DIR / "SHEETS"
            dummy = DeleteArchiveDummy()
            try:
                archive_path = dummy._archive_deleted_file(sheet, app.DELETED_SHEETS_DIR, "home_sheet_deleted", {"stage": "Working"})
                metadata_path = archive_path.with_name(f"{archive_path.name}.archive.json")
                metadata = json.loads(metadata_path.read_text(encoding="utf-8"))

                self.assertFalse(sheet.exists())
                self.assertTrue(archive_path.exists())
                self.assertEqual(archive_path.parent.parent, app.DELETED_SHEETS_DIR)
                self.assertEqual(metadata["original_path"], str(sheet))
                self.assertEqual(metadata["reason"], "home_sheet_deleted")
                self.assertEqual(metadata["retention_days"], 14)
            finally:
                app.DELETED_ARCHIVE_DIR = old_deleted_archive
                app.DELETED_SHEETS_DIR = old_deleted_sheets

    def test_team_payout_uses_sold_profit_not_unsold_estimated_profit(self) -> None:
        class PayoutDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _person_for_profit_record = app.CardPipelineApp._person_for_profit_record
            _enrich_profit_records_with_people = app.CardPipelineApp._enrich_profit_records_with_people
            _sold_payout_key = app.CardPipelineApp._sold_payout_key
            _realized_profit_totals_by_person_sheet = app.CardPipelineApp._realized_profit_totals_by_person_sheet
            _realized_profit_groups_by_person_sheet = app.CardPipelineApp._realized_profit_groups_by_person_sheet
            _loose_expense_adjustments_by_person = app.CardPipelineApp._loose_expense_adjustments_by_person
            _active_payout_balance = app.CardPipelineApp._active_payout_balance
            _payout_sheet_status = app.CardPipelineApp._payout_sheet_status
            _payout_sheet_items = app.CardPipelineApp._payout_sheet_items
            _sheet_marker_is_seller_payout = app.CardPipelineApp._sheet_marker_is_seller_payout
            _source_sheet_is_seller_payout = app.CardPipelineApp._source_sheet_is_seller_payout
            _profit_record_payout_time = app.CardPipelineApp._profit_record_payout_time
            _empty_realized_profit_group = app.CardPipelineApp._empty_realized_profit_group
            _add_profit_record_to_realized_group = app.CardPipelineApp._add_profit_record_to_realized_group
            _payout_realized_groups_for_marker = app.CardPipelineApp._payout_realized_groups_for_marker

            def __init__(self):
                self.home_sheet_paths = {"Incoming": {"Lot A.xlsx": Path("Lot A.xlsx")}, "Received": {}}
                self.home_sheet_markers = {"Incoming|Lot A.xlsx": {"assigned_person": "Kevin Hambone"}}
                self.home_sheet_summaries = {
                    "Incoming|Lot A.xlsx": {
                        "row_count": 2,
                        "received_count": 0,
                        "purchase_total": 80.0,
                        "estimated_payout_total": 150.0,
                    }
                }
                self.ledger = []

            def _seller_terms_seller_names(self):
                return set()

            def _load_profit_ledger(self):
                return self.ledger

        dummy = PayoutDummy()
        payout_items = dummy._payout_sheet_items()
        self.assertEqual(payout_items, [])

        dummy.ledger = [
            {
                "assigned_person": "Kevin Hambone",
                "source_sheet": "Lot A.xlsx",
                "purchase_price": 80.0,
                "sale_price": 150.0,
                "company": "Arena Club",
                "cert_number": "123",
                "date_added": "2026-06-19",
            },
            {
                "record_type": "expense",
                "assigned_person": "Kevin Hambone",
                "source_sheet": "Lot A.xlsx",
                "expense_amount": 20.0,
                "profit": -20.0,
                "date_added": "2026-06-19",
            },
            {
                "record_type": "expense",
                "assigned_person": "Kevin Hambone",
                "source_sheet": "",
                "expense_amount": 10.0,
                "profit": -10.0,
                "date_added": "2026-06-19",
            }
        ]
        payout_items = dummy._payout_sheet_items()
        self.assertEqual(len(payout_items), 2)
        by_name = {item["name"]: item for item in payout_items}
        self.assertEqual(by_name["Lot A.xlsx"]["stage"], "Sold")
        self.assertEqual(by_name["Lot A.xlsx"]["row_count"], 1)
        self.assertEqual(by_name["Lot A.xlsx"]["expense_total"], 20.0)
        self.assertEqual(by_name["Lot A.xlsx"]["realized_profit_total"], 50.0)
        self.assertEqual(by_name["Lot A.xlsx"]["payout_balance"], 25.0)
        self.assertEqual(by_name["Expense Adjustments"]["expense_total"], 10.0)
        self.assertEqual(by_name["Expense Adjustments"]["net_profit_total"], -10.0)
        self.assertEqual(by_name["Expense Adjustments"]["payout_balance"], -5.0)
        self.assertEqual(sum(float(item["payout_balance"]) for item in payout_items), 20.0)

    def test_team_payout_reopens_balance_for_sales_after_paid_marker(self) -> None:
        class PayoutDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _person_for_profit_record = app.CardPipelineApp._person_for_profit_record
            _enrich_profit_records_with_people = app.CardPipelineApp._enrich_profit_records_with_people
            _sold_payout_key = app.CardPipelineApp._sold_payout_key
            _realized_profit_groups_by_person_sheet = app.CardPipelineApp._realized_profit_groups_by_person_sheet
            _loose_expense_adjustments_by_person = app.CardPipelineApp._loose_expense_adjustments_by_person
            _active_payout_balance = app.CardPipelineApp._active_payout_balance
            _payout_sheet_status = app.CardPipelineApp._payout_sheet_status
            _payout_sheet_items = app.CardPipelineApp._payout_sheet_items
            _sheet_marker_is_seller_payout = app.CardPipelineApp._sheet_marker_is_seller_payout
            _source_sheet_is_seller_payout = app.CardPipelineApp._source_sheet_is_seller_payout
            _profit_record_payout_time = app.CardPipelineApp._profit_record_payout_time
            _empty_realized_profit_group = app.CardPipelineApp._empty_realized_profit_group
            _add_profit_record_to_realized_group = app.CardPipelineApp._add_profit_record_to_realized_group
            _payout_realized_groups_for_marker = app.CardPipelineApp._payout_realized_groups_for_marker

            def __init__(self):
                key = self._sold_payout_key("Kevin Hambone", "Lot A.xlsx")
                self.home_sheet_paths = {"Incoming": {}, "Received": {}}
                self.home_sheet_markers = {key: {"assigned_person": "Kevin Hambone", "paid": True, "paid_at": "2026-07-14T10:00:00"}}
                self.home_sheet_summaries = {}
                self.ledger = [
                    {
                        "assigned_person": "Kevin Hambone",
                        "source_sheet": "Lot A.xlsx",
                        "purchase_price": 100.0,
                        "sale_price": 160.0,
                        "company": "Fanatics",
                        "cert_number": "111",
                        "date_added": "2026-07-14",
                        "ledger_added_at": "2026-07-14T09:30:00",
                    },
                    {
                        "assigned_person": "Kevin Hambone",
                        "source_sheet": "Lot A.xlsx",
                        "purchase_price": 40.0,
                        "sale_price": 100.0,
                        "company": "Fanatics",
                        "cert_number": "222",
                        "date_added": "2026-07-14",
                        "ledger_added_at": "2026-07-14T10:30:00",
                    },
                ]

            def _seller_terms_seller_names(self):
                return set()

            def _load_profit_ledger(self):
                return self.ledger

        items = PayoutDummy()._payout_sheet_items()
        self.assertEqual(len(items), 2)
        paid_items = [item for item in items if item["paid"]]
        open_items = [item for item in items if not item["paid"]]
        self.assertEqual(paid_items[0]["payout_balance"], 30.0)
        self.assertEqual(open_items[0]["payout_balance"], 30.0)
        self.assertEqual(open_items[0]["status"], "Sold")

    def test_moving_received_sheet_back_clears_received_profit_and_company_rows(self) -> None:
        class MoveDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _sheet_path_for_stage = app.CardPipelineApp._sheet_path_for_stage
            _delete_sheet_marker = app.CardPipelineApp._delete_sheet_marker
            _marker_for_stage = app.CardPipelineApp._marker_for_stage
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _remove_profit_ledger_rows_for_source = app.CardPipelineApp._remove_profit_ledger_rows_for_source
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _remove_inventory_rows_for_source = app.CardPipelineApp._remove_inventory_rows_for_source
            _cleanup_sheet_received_side_effects = app.CardPipelineApp._cleanup_sheet_received_side_effects
            _move_home_sheet_to_stage = app.CardPipelineApp._move_home_sheet_to_stage

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            incoming_dir = root / "INCOMING SHEETS"
            working_dir = root / "WORKING SHEETS"
            received_dir = root / "RECEIVED SHEETS"
            company_dir = root / "COMPANY SHEETS"
            received_dir.mkdir(parents=True)
            working_dir.mkdir()

            source_path = received_dir / "Lot A.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Certification Number", "Card Description", "RECEIVED"])
            sheet.append(["123", "Test Card", "X"])
            workbook.save(source_path)

            company_path = company_dir / "Arena Club" / "Arena Club.xlsx"
            company_path.parent.mkdir(parents=True)
            company_workbook = Workbook()
            company_sheet = company_workbook.active
            company_sheet.title = "Week of 2026-06-15"
            company_sheet.append(["Date Added", "Source Sheet", "Source", "Certification Number", "Grader", "Card Description", "Purchase Price", "Card Ladder Value", "Comps", "CY Estimate", "CY Confidence", "Best Company", "Estimated Payout", "Status", "Notes"])
            company_sheet.append(["2026-06-17", "Lot A.xlsx", "", "123", "PSA", "Test Card", 40, "", 100, "", "", "Arena Club", 90, "Received", ""])
            company_workbook.save(company_path)

            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_received = app.RECEIVED_SHEETS_DIR
            old_company = app.COMPANY_SHEETS_DIR
            old_ledger = app.PROFIT_LEDGER_PATH
            old_inventory_ledger = app.INVENTORY_LEDGER_PATH
            app.INCOMING_SHEETS_DIR = incoming_dir
            app.WORKING_SHEETS_DIR = working_dir
            app.RECEIVED_SHEETS_DIR = received_dir
            app.COMPANY_SHEETS_DIR = company_dir
            app.PROFIT_LEDGER_PATH = root / "profit_ledger.json"
            app.INVENTORY_LEDGER_PATH = root / "inventory_ledger.json"
            app.PROFIT_LEDGER_PATH.write_text(
                json.dumps([
                    {
                        "date_added": "2026-06-17",
                        "company": "Arena Club",
                        "weekly_sheet_name": "Arena Club.xlsx:Week of 2026-06-15",
                        "source_sheet": "Lot A.xlsx",
                        "cert_number": "123",
                        "card_title": "Test Card",
                        "purchase_price": 40,
                        "sale_price": 90,
                    }
                ]),
                encoding="utf-8",
            )
            dummy = MoveDummy()
            dummy.home_sheet_paths = {"Incoming": {}, "Working": {}, "Received": {}}
            dummy.received_sheet_paths = {"Lot A.xlsx": source_path}
            dummy.home_sheet_markers = {
                "Received|Lot A.xlsx": {
                    "assigned_person": "Lucas",
                    "paid": True,
                    "all_received": True,
                    "received_at": "2026-06-17T10:00:00",
                }
            }
            dummy.deleted_sheet_marker_keys = set()
            try:
                moved_key, cleanup = dummy._move_home_sheet_to_stage("Received|Lot A.xlsx", "Working")

                self.assertEqual(moved_key, "Working|Lot A.xlsx")
                self.assertTrue((working_dir / "Lot A.xlsx").exists())
                moved_workbook = load_workbook(working_dir / "Lot A.xlsx", data_only=True)
                try:
                    self.assertEqual(moved_workbook.active.cell(2, 3).value, None)
                finally:
                    moved_workbook.close()
                company_rows = read_company_profit_records(company_dir)
                ledger = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                marker = dummy.home_sheet_markers[moved_key]
                self.assertEqual(company_rows, [])
                self.assertEqual(ledger, [])
                self.assertFalse(marker["paid"])
                self.assertFalse(marker["all_received"])
                self.assertNotIn("received_at", marker)
                self.assertEqual(cleanup["received_rows_cleared"], 1)
                self.assertEqual(cleanup["company_rows_removed"], 1)
                self.assertEqual(cleanup["profit_rows_removed"], 1)
            finally:
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working
                app.RECEIVED_SHEETS_DIR = old_received
                app.COMPANY_SHEETS_DIR = old_company
                app.PROFIT_LEDGER_PATH = old_ledger
                app.INVENTORY_LEDGER_PATH = old_inventory_ledger

    def test_received_sheet_move_requires_second_cleanup_confirmation(self) -> None:
        class MoveConfirmDummy:
            _confirm_home_stage_move = app.CardPipelineApp._confirm_home_stage_move

        dummy = MoveConfirmDummy()
        with patch.object(app.messagebox, "askyesno", side_effect=[True, False]) as askyesno:
            self.assertFalse(dummy._confirm_home_stage_move("Received", "Incoming", "Lot A.xlsx"))
        self.assertEqual(askyesno.call_count, 2)

        with patch.object(app.messagebox, "askyesno", return_value=True) as askyesno:
            self.assertTrue(dummy._confirm_home_stage_move("Incoming", "Working", "Lot A.xlsx"))
        self.assertEqual(askyesno.call_count, 1)

    def test_profit_sales_are_deduped_and_delta_is_recorded(self) -> None:
        class ProfitDummy:
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _profit_record_identity_keys = app.CardPipelineApp._profit_record_identity_keys
            _money_value = app.CardPipelineApp._money_value
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            record_profit_sales = app.CardPipelineApp.record_profit_sales
            refresh_profit_tab = lambda self: None

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_ledger = app.PROFIT_LEDGER_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            dummy = ProfitDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                record = {
                    "date_added": "2026-06-11",
                    "company": "Arena Club",
                    "weekly_sheet_name": "Arena WEEK.xlsx",
                    "source_sheet": "source.xlsx",
                    "cert_number": "123",
                    "card_title": "Test Card",
                    "purchase_price": "$40.00",
                    "sale_price": "$90.00",
                }
                self.assertEqual(dummy.record_profit_sales([record]), 1)
                self.assertEqual(dummy.record_profit_sales([record]), 0)
                ledger = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                self.assertEqual(len(ledger), 1)
                self.assertEqual(ledger[0]["profit"], 50.0)
                self.assertEqual(ledger[0]["recorded_by"], "Tester")
                self.assertRegex(ledger[0]["ledger_added_at"], r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.PROFIT_LEDGER_PATH = old_ledger

    def test_profit_sales_dedupe_inventory_sale_against_company_sheet_recovery(self) -> None:
        class ProfitDummy:
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _profit_record_identity_keys = app.CardPipelineApp._profit_record_identity_keys
            _money_value = app.CardPipelineApp._money_value
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            record_profit_sales = app.CardPipelineApp.record_profit_sales
            refresh_profit_tab = lambda self: None

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_ledger = app.PROFIT_LEDGER_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            dummy = ProfitDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                inventory_sale = {
                    "date_added": "2026-06-20",
                    "company": "Arena Club",
                    "weekly_sheet_name": "Inventory Sale",
                    "source_sheet": "Lot A.xlsx",
                    "cert_number": "123",
                    "card_title": "Test Card",
                    "purchase_price": 40,
                    "sale_price": 90,
                }
                recovered_company_sale = dict(
                    inventory_sale,
                    date_added="2026-06-22",
                    weekly_sheet_name="Arena Club.xlsx:Week of 2026-06-22",
                )

                self.assertEqual(dummy.record_profit_sales([inventory_sale]), 1)
                self.assertEqual(dummy.record_profit_sales([recovered_company_sale]), 0)
                ledger = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                self.assertEqual(len(ledger), 1)
                self.assertEqual(ledger[0]["weekly_sheet_name"], "Inventory Sale")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.PROFIT_LEDGER_PATH = old_ledger

    def test_profit_refresh_skips_company_scan_until_deep_sync(self) -> None:
        class ProfitRefreshDummy:
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _profit_record_identity_keys = app.CardPipelineApp._profit_record_identity_keys
            _is_manual_company_profit_backfill = app.CardPipelineApp._is_manual_company_profit_backfill
            _money_value = app.CardPipelineApp._money_value
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _enrich_profit_records_with_people = app.CardPipelineApp._enrich_profit_records_with_people
            _filtered_profit_records = app.CardPipelineApp._filtered_profit_records
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _profit_period_bounds = app.CardPipelineApp._profit_period_bounds
            _profit_today = lambda self: datetime(2026, 6, 20).date()
            refresh_profit_tab = app.CardPipelineApp.refresh_profit_tab

            def _person_for_profit_record(self, record):
                return str(record.get("assigned_person") or "")

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_company = app.COMPANY_SHEETS_DIR
            old_ledger = app.PROFIT_LEDGER_PATH
            app.CARD_PIPELINE_DIR = root
            app.COMPANY_SHEETS_DIR = root / "COMPANY SHEETS"
            app.PROFIT_LEDGER_PATH = root / "profit_ledger.json"
            dummy = ProfitRefreshDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.profit_person_var = types.SimpleNamespace(get=lambda: "")
            dummy.profit_search_var = types.SimpleNamespace(get=lambda: "")
            dummy.profit_period_var = types.SimpleNamespace(get=lambda: "Total")
            company_record = {
                "assigned_person": "Kevin Hambone",
                "date_added": "2026-06-20",
                "company": "Arena Club",
                "weekly_sheet_name": "Arena Club.xlsx:Week of 2026-06-15",
                "source_sheet": "Lot A.xlsx",
                "cert_number": "123",
                "card_title": "Test Card",
                "purchase_price": 40,
                "sale_price": 90,
            }
            try:
                with patch("app.read_company_profit_records", side_effect=AssertionError("normal refresh scanned company sheets")):
                    dummy.refresh_profit_tab()
                self.assertEqual(dummy.profit_rows, [])
                self.assertFalse(app.PROFIT_LEDGER_PATH.exists())

                with patch("app.read_company_profit_records", return_value=[company_record]) as reader:
                    dummy.refresh_profit_tab(deep_sync=True)
                reader.assert_called_once_with(app.COMPANY_SHEETS_DIR)
                ledger = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                self.assertEqual(len(ledger), 1)
                self.assertEqual(ledger[0]["cert_number"], "123")
                self.assertEqual(ledger[0]["profit"], 50.0)

                existing_inventory_sale = dummy._normalize_profit_record(dict(company_record, weekly_sheet_name="Inventory Sale"))
                app.PROFIT_LEDGER_PATH.write_text(json.dumps([existing_inventory_sale]), encoding="utf-8")
                with patch("app.read_company_profit_records", return_value=[company_record]):
                    dummy.refresh_profit_tab(deep_sync=True)
                ledger = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                self.assertEqual(len(ledger), 1)
                self.assertEqual(ledger[0]["weekly_sheet_name"], "Inventory Sale")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.COMPANY_SHEETS_DIR = old_company
                app.PROFIT_LEDGER_PATH = old_ledger

    def test_home_workbook_summary_cache_reuses_unchanged_files(self) -> None:
        class HomeSummaryDummy:
            _home_summary_cache_key = app.CardPipelineApp._home_summary_cache_key
            _summarize_home_workbook_cached = app.CardPipelineApp._summarize_home_workbook_cached
            _prune_home_summary_cache = app.CardPipelineApp._prune_home_summary_cache

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "Lot A.xlsx"
            path.write_text("first", encoding="utf-8")
            dummy = HomeSummaryDummy()
            dummy.home_summary_cache = {}
            dummy.home_summary_cache_lock = threading.Lock()
            with patch("app.summarize_workbook", side_effect=[{"name": "first"}, {"name": "second"}]) as summarizer:
                self.assertEqual(dummy._summarize_home_workbook_cached(path), {"name": "first"})
                self.assertEqual(dummy._summarize_home_workbook_cached(path), {"name": "first"})
                self.assertEqual(summarizer.call_count, 1)

                path.write_text("second and larger", encoding="utf-8")
                self.assertEqual(dummy._summarize_home_workbook_cached(path), {"name": "second"})
                self.assertEqual(summarizer.call_count, 2)

                dummy._prune_home_summary_cache([])
                self.assertEqual(dummy.home_summary_cache, {})

    def test_accounted_incoming_sheet_reconciles_to_received_without_duplicate_inventory(self) -> None:
        class ReconcileDummy:
            _accounted_identity_key = app.CardPipelineApp._accounted_identity_key
            _add_accounted_identity = app.CardPipelineApp._add_accounted_identity
            _accounted_source_key = app.CardPipelineApp._accounted_source_key
            _add_accounted_cert = app.CardPipelineApp._add_accounted_cert
            _accounted_sheet_cert_index = app.CardPipelineApp._accounted_sheet_cert_index
            _sheet_cert_set = app.CardPipelineApp._sheet_cert_set
            _sheet_accounting_payload = app.CardPipelineApp._sheet_accounting_payload
            _reconcile_accounted_home_sheets = app.CardPipelineApp._reconcile_accounted_home_sheets
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _marker_for_stage = app.CardPipelineApp._marker_for_stage
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_record_key = app.CardPipelineApp._inventory_record_key

            def _load_inventory_ledger(self):
                return self.inventory

            def _load_profit_ledger(self):
                return self.profit

            def _load_activity_log(self):
                return self.activity

            def _save_sheet_markers(self):
                pass

            def _append_activity(self, action, summary, details=None):
                self.activity.append({"action": action, "summary": summary, "details": details or {}})

            def _move_sheet_to_received(self, key):
                stage, name = self._split_home_sheet_key(key)
                source = {"Incoming": app.INCOMING_SHEETS_DIR, "Working": app.WORKING_SHEETS_DIR}[stage] / name
                destination = app.RECEIVED_SHEETS_DIR / name
                app.RECEIVED_SHEETS_DIR.mkdir(parents=True, exist_ok=True)
                app.shutil.move(str(source), str(destination))
                marker = self._marker_for_stage(dict(self.home_sheet_markers.get(key, {})), "Received")
                self.home_sheet_markers.pop(key, None)
                new_key = self._home_sheet_key("Received", name)
                self.home_sheet_markers[new_key] = marker
                return new_key

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_received = app.RECEIVED_SHEETS_DIR
            app.CARD_PIPELINE_DIR = root
            app.INCOMING_SHEETS_DIR = root / "INCOMING SHEETS"
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            app.RECEIVED_SHEETS_DIR = root / "RECEIVED SHEETS"
            app.INCOMING_SHEETS_DIR.mkdir(parents=True)
            app.WORKING_SHEETS_DIR.mkdir(parents=True)
            app.RECEIVED_SHEETS_DIR.mkdir(parents=True)
            try:
                sheet_path = app.INCOMING_SHEETS_DIR / "Lot A.xlsx"
                write_working_sheet(
                    sheet_path,
                    [
                        WorkbookRow(excel_row=2, cert_number="111", grader="PSA", card_title="Inventory Card", existing_value=10),
                        WorkbookRow(excel_row=3, cert_number="222", grader="PSA", card_title="Sold Card", existing_value=20),
                    ],
                    {2: "manual", 3: "manual"},
                )
                dummy = ReconcileDummy()
                dummy.home_sheet_markers = {"Incoming|Lot A.xlsx": {"assigned_person": "Kevin Hambone"}}
                dummy.inventory = [{"cert_number": "111", "source_sheet": "Lot A.xlsx", "status": "Active"}]
                dummy.profit = [{"cert_number": "222", "source_sheet": "Kevin Hambone General Sold", "original_source_sheet": "Lot A.xlsx", "purchase_price": 20, "sale_price": 30}]
                dummy.activity = []

                result = dummy._reconcile_accounted_home_sheets()

                self.assertEqual(result["moved"], ["Lot A.xlsx"])
                self.assertEqual(result["warnings"], [])
                self.assertFalse(sheet_path.exists())
                self.assertTrue((app.RECEIVED_SHEETS_DIR / "Lot A.xlsx").exists())
                self.assertTrue(all(row["received"] for row in read_simple_spreadsheet(app.RECEIVED_SHEETS_DIR / "Lot A.xlsx")))
                self.assertTrue(dummy.home_sheet_markers["Received|Lot A.xlsx"]["all_received"])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working
                app.RECEIVED_SHEETS_DIR = old_received

    def test_accounted_incoming_sheet_with_unaccounted_raw_row_does_not_reconcile_to_received(self) -> None:
        class ReconcileDummy:
            _accounted_identity_key = app.CardPipelineApp._accounted_identity_key
            _add_accounted_identity = app.CardPipelineApp._add_accounted_identity
            _accounted_source_key = app.CardPipelineApp._accounted_source_key
            _add_accounted_cert = app.CardPipelineApp._add_accounted_cert
            _accounted_sheet_cert_index = app.CardPipelineApp._accounted_sheet_cert_index
            _sheet_accounting_payload = app.CardPipelineApp._sheet_accounting_payload
            _reconcile_accounted_home_sheets = app.CardPipelineApp._reconcile_accounted_home_sheets
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key

            def _load_inventory_ledger(self):
                return self.inventory

            def _load_profit_ledger(self):
                return []

            def _load_activity_log(self):
                return []

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_received = app.RECEIVED_SHEETS_DIR
            app.CARD_PIPELINE_DIR = root
            app.INCOMING_SHEETS_DIR = root / "INCOMING SHEETS"
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            app.RECEIVED_SHEETS_DIR = root / "RECEIVED SHEETS"
            app.INCOMING_SHEETS_DIR.mkdir(parents=True)
            app.WORKING_SHEETS_DIR.mkdir(parents=True)
            app.RECEIVED_SHEETS_DIR.mkdir(parents=True)
            try:
                sheet_path = app.INCOMING_SHEETS_DIR / "Raw Lot.xlsx"
                write_working_sheet(
                    sheet_path,
                    [
                        WorkbookRow(excel_row=2, cert_number="111", grader="PSA", card_title="Accounted Cert", existing_value=10),
                        WorkbookRow(excel_row=3, cert_number="", item_id="RAW-MIKEY-20260709-0003", grader="PSA", card_title="2018 Topps Fire Shohei Ohtani Gold", existing_value=325),
                    ],
                    {2: "manual", 3: "manual"},
                )
                dummy = ReconcileDummy()
                dummy.inventory = [{"cert_number": "111", "source_sheet": "Raw Lot.xlsx", "status": "Active"}]

                result = dummy._reconcile_accounted_home_sheets()

                self.assertEqual(result["moved"], [])
                self.assertEqual(result["warnings"], [])
                self.assertEqual(len(result["notices"]), 1)
                self.assertIn("1/2 row(s) already exist", result["notices"][0])
                self.assertTrue(sheet_path.exists())
                self.assertFalse((app.RECEIVED_SHEETS_DIR / "Raw Lot.xlsx").exists())
                rows = read_simple_spreadsheet(sheet_path)
                self.assertFalse(rows[1]["received"])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working
                app.RECEIVED_SHEETS_DIR = old_received

    def test_partial_accounted_incoming_sheet_notices_without_moving(self) -> None:
        class ReconcileDummy:
            _accounted_identity_key = app.CardPipelineApp._accounted_identity_key
            _add_accounted_identity = app.CardPipelineApp._add_accounted_identity
            _accounted_source_key = app.CardPipelineApp._accounted_source_key
            _add_accounted_cert = app.CardPipelineApp._add_accounted_cert
            _accounted_sheet_cert_index = app.CardPipelineApp._accounted_sheet_cert_index
            _sheet_cert_set = app.CardPipelineApp._sheet_cert_set
            _sheet_accounting_payload = app.CardPipelineApp._sheet_accounting_payload
            _reconcile_accounted_home_sheets = app.CardPipelineApp._reconcile_accounted_home_sheets
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key

            def _load_inventory_ledger(self):
                return self.inventory

            def _load_profit_ledger(self):
                return []

            def _load_activity_log(self):
                return []

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_received = app.RECEIVED_SHEETS_DIR
            app.CARD_PIPELINE_DIR = root
            app.INCOMING_SHEETS_DIR = root / "INCOMING SHEETS"
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            app.RECEIVED_SHEETS_DIR = root / "RECEIVED SHEETS"
            app.INCOMING_SHEETS_DIR.mkdir(parents=True)
            app.WORKING_SHEETS_DIR.mkdir(parents=True)
            app.RECEIVED_SHEETS_DIR.mkdir(parents=True)
            try:
                sheet_path = app.INCOMING_SHEETS_DIR / "Lot B.xlsx"
                write_working_sheet(
                    sheet_path,
                    [
                        WorkbookRow(excel_row=2, cert_number="111", grader="PSA", card_title="Duplicate Card", existing_value=10),
                        WorkbookRow(excel_row=3, cert_number="333", grader="PSA", card_title="New Card", existing_value=20),
                    ],
                    {2: "manual", 3: "manual"},
                )
                dummy = ReconcileDummy()
                dummy.inventory = [{"cert_number": "111", "source_sheet": "Lot B.xlsx", "status": "Active"}]

                result = dummy._reconcile_accounted_home_sheets()

                self.assertEqual(result["moved"], [])
                self.assertEqual(result["warnings"], [])
                self.assertEqual(len(result["notices"]), 1)
                self.assertIn("1/2 row(s) already exist", result["notices"][0])
                self.assertTrue(sheet_path.exists())
                self.assertFalse((app.RECEIVED_SHEETS_DIR / "Lot B.xlsx").exists())
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working
                app.RECEIVED_SHEETS_DIR = old_received

    def test_expense_records_deduct_from_person_profit(self) -> None:
        class ExpenseDummy:
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _money_value = app.CardPipelineApp._money_value
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            record_profit_sales = app.CardPipelineApp.record_profit_sales
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _profit_today = lambda self: datetime(2026, 6, 19).date()
            _profit_period_bounds = app.CardPipelineApp._profit_period_bounds
            _canonical_profit_period = app.CardPipelineApp._canonical_profit_period
            _profit_period_label = app.CardPipelineApp._profit_period_label
            _profit_graph_label = app.CardPipelineApp._profit_graph_label
            _profit_chart_title = app.CardPipelineApp._profit_chart_title
            _filtered_profit_records = app.CardPipelineApp._filtered_profit_records
            _profit_chart_series = app.CardPipelineApp._profit_chart_series
            _expense_related_label = app.CardPipelineApp._expense_related_label
            _expense_link_options = app.CardPipelineApp._expense_link_options
            _delete_profit_expense_records = app.CardPipelineApp._delete_profit_expense_records
            refresh_profit_tab = lambda self: None

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_ledger = app.PROFIT_LEDGER_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            dummy = ExpenseDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.profit_person_var = types.SimpleNamespace(get=lambda: "Kevin")
            dummy.profit_period_var = types.SimpleNamespace(get=lambda: "YTD")
            dummy.profit_graph_var = types.SimpleNamespace(get=lambda: "Daily Trend")
            try:
                expense = {
                    "record_type": "expense",
                    "expense_id": "expense-1",
                    "assigned_person": "Kevin Hambone",
                    "date_added": "2026-06-18",
                    "expense_type": "Shipping",
                    "expense_amount": 25,
                    "related_type": "Card",
                    "source_sheet": "Lot A.xlsx",
                    "cert_number": "123",
                    "notes": "Shipment label",
                }
                sale = {
                    "assigned_person": "Kevin Hambone",
                    "cert_number": "123",
                    "card_title": "Test Card",
                    "source_sheet": "Lot A.xlsx",
                    "company": "Arena",
                    "purchase_price": 50,
                    "sale_price": 100,
                    "date_added": "2026-06-18",
                }
                self.assertEqual(dummy.record_profit_sales([expense, sale]), 2)
                ledger = [dummy._normalize_profit_record(record) for record in dummy._load_profit_ledger()]
                expense_row = next(record for record in ledger if record.get("record_type") == "expense")
                self.assertEqual(expense_row["profit"], -25)
                self.assertEqual(expense_row["company"], "Expense: Shipping")
                self.assertEqual(expense_row["source_sheet"], "Lot A.xlsx")
                self.assertEqual(expense_row["cert_number"], "123")
                self.assertEqual(dummy._expense_related_label(expense_row), "Lot A.xlsx | 123")
                raw_expense_row = dict(expense_row, source_sheet="Raw Inventory", cert_number="", item_id="RAW-20260624-0001")
                self.assertEqual(dummy._expense_related_label(raw_expense_row), "Raw Inventory | RAW-20260624-0001")
                sheets, cards, lookup = dummy._expense_link_options("Kevin")
                self.assertEqual(sheets, ["Lot A.xlsx"])
                self.assertEqual(len(cards), 1)
                self.assertIn("Lot A.xlsx | Test Card | $100.00", cards)
                self.assertEqual(lookup[cards[0]]["source_sheet"], "Lot A.xlsx")
                self.assertEqual(lookup[cards[0]]["cert_number"], "123")
                filtered = dummy._filtered_profit_records(ledger)
                _days, values = dummy._profit_chart_series(filtered)
                self.assertEqual(sum(values), 25)
                self.assertEqual(dummy._delete_profit_expense_records([expense_row]), 1)
                ledger_after_delete = [dummy._normalize_profit_record(record) for record in dummy._load_profit_ledger()]
                self.assertEqual(len(ledger_after_delete), 1)
                self.assertNotEqual(ledger_after_delete[0].get("record_type"), "expense")
                self.assertEqual(ledger_after_delete[0]["profit"], 50)
                self.assertEqual(dummy._delete_profit_expense_records([ledger_after_delete[0]]), 0)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.PROFIT_LEDGER_PATH = old_ledger

    def test_inventory_records_are_deduped_and_reactivated(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            add_inventory_records = app.CardPipelineApp.add_inventory_records
            _enrich_inventory_record_assignment = lambda self, record: record
            refresh_inventory_tab = lambda self: None

        with TemporaryDirectory() as tmp:
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = InventoryDummy()
            try:
                record = {
                    "assigned_person": "Lucas",
                    "cert_number": "123",
                    "card_title": "Test Card",
                    "source_sheet": "Lot A.xlsx",
                    "purchase_price": 40,
                    "inventory_value": 100,
                    "status": "Refunded",
                }
                self.assertEqual(dummy.add_inventory_records([record]), 1)
                self.assertEqual(dummy.add_inventory_records([{**record, "status": "Active"}]), 0)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(len(ledger), 1)
                self.assertEqual(ledger[0]["status"], "Active")
            finally:
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_inventory_record_from_row_preserves_manual_sport_for_unknown_title(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_sport_from_value = app.CardPipelineApp._inventory_sport_from_value
            _inventory_record_from_row = app.CardPipelineApp._inventory_record_from_row

        row = WorkbookRow(
            excel_row=2,
            cert_number="0010355805",
            grader="BGS",
            card_title="2017 Bowman Chrome Prospect Autographs Gold Shimmer Refractors #CPACF Clint Frazier BGS 9.5",
            category="baseball",
            card_ladder_comps_average=432,
        )

        record = InventoryDummy()._inventory_record_from_row(row, "James Copeland", "JAMES_NASHVILLE_INVENTORY_2.xlsx")

        self.assertEqual(record["sport"], "baseball")

    def test_inventory_record_normalizes_common_sport_aliases(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record

        dummy = InventoryDummy()

        self.assertEqual(dummy._normalize_inventory_record({"sport": "b-ball"})["sport"], "basketball")
        self.assertEqual(dummy._normalize_inventory_record({"category": "Poke"})["sport"], "pokemon")

    def test_inventory_filter_searches_cert_and_card_title(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records

        dummy = InventoryDummy()
        dummy.inventory_person_var = FieldVar("")
        dummy.inventory_sport_var = FieldVar("")
        dummy.inventory_min_var = FieldVar("")
        dummy.inventory_max_var = FieldVar("")
        rows = [
            {"status": "Active", "cert_number": "151740304", "card_title": "2024 Prizm Victor Wembanyama Silver PSA 10", "inventory_value": 100},
            {"status": "Active", "cert_number": "222", "card_title": "2019 Panini Mosaic Stephen Curry Green PSA 10", "inventory_value": 90},
            {"status": "Sold", "cert_number": "333", "card_title": "Hidden Sold Card", "inventory_value": 80},
        ]

        dummy.inventory_search_var = FieldVar("174030")
        self.assertEqual([row["cert_number"] for row in dummy._filtered_inventory_records(rows)], ["151740304"])

        dummy.inventory_search_var = FieldVar("curry green")
        self.assertEqual([row["cert_number"] for row in dummy._filtered_inventory_records(rows)], ["222"])

        dummy.inventory_search_var = FieldVar("hidden sold")
        self.assertEqual(dummy._filtered_inventory_records(rows), [])

    def test_inventory_filter_finds_missing_card_descriptions(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _inventory_record_missing_card_description = app.CardPipelineApp._inventory_record_missing_card_description
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records

        dummy = InventoryDummy()
        dummy.inventory_person_var = FieldVar("")
        dummy.inventory_sport_var = FieldVar("")
        dummy.inventory_grader_var = FieldVar("")
        dummy.inventory_year_var = FieldVar("")
        dummy.inventory_search_var = FieldVar("")
        dummy.inventory_min_var = FieldVar("")
        dummy.inventory_max_var = FieldVar("")
        dummy.inventory_date_min_var = FieldVar("")
        dummy.inventory_date_max_var = FieldVar("")
        dummy.inventory_missing_title_var = FieldVar(True)
        dummy.inventory_missing_photos_var = FieldVar(False)
        rows = [
            {"status": "Active", "cert_number": "111", "card_title": "", "inventory_value": 50},
            {"status": "Active", "cert_number": "222", "card_title": "222", "inventory_value": 50},
            {"status": "Active", "cert_number": "333", "card_title": "2024 Prizm Real Card PSA 10", "inventory_value": 50},
            {"status": "Sold", "cert_number": "444", "card_title": "", "inventory_value": 50},
        ]

        self.assertEqual([row["cert_number"] for row in dummy._filtered_inventory_records(rows)], ["111", "222"])

    def test_inventory_sport_filter_accepts_multiple_checked_sports(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records

        dummy = InventoryDummy()
        dummy.inventory_person_var = FieldVar("")
        dummy.inventory_sport_var = FieldVar("football, basketball, baseball")
        dummy.inventory_search_var = FieldVar("")
        dummy.inventory_min_var = FieldVar("")
        dummy.inventory_max_var = FieldVar("")
        rows = [
            {"status": "Active", "cert_number": "1", "sport": "football", "card_title": "Football Card", "inventory_value": 100},
            {"status": "Active", "cert_number": "2", "sport": "basketball", "card_title": "Basketball Card", "inventory_value": 100},
            {"status": "Active", "cert_number": "3", "sport": "baseball", "card_title": "Baseball Card", "inventory_value": 100},
            {"status": "Active", "cert_number": "4", "sport": "pokemon", "card_title": "Pokemon Card", "inventory_value": 100},
        ]

        filtered = dummy._filtered_inventory_records(rows)

        self.assertEqual([record["cert_number"] for record in filtered], ["1", "2", "3"])

    def test_inventory_filter_applies_date_min_and_max(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records

        dummy = InventoryDummy()
        dummy.inventory_person_var = FieldVar("")
        dummy.inventory_sport_var = FieldVar("")
        dummy.inventory_search_var = FieldVar("")
        dummy.inventory_min_var = FieldVar("")
        dummy.inventory_max_var = FieldVar("")
        dummy.inventory_date_min_var = FieldVar("2026-06-01")
        dummy.inventory_date_max_var = FieldVar("2026-06-30")
        rows = [
            {"status": "Active", "cert_number": "111", "card_title": "May Card", "date_added": "2026-05-31", "inventory_value": 50},
            {"status": "Active", "cert_number": "222", "card_title": "June Card", "date_added": "2026-06-15", "inventory_value": 50},
            {"status": "Active", "cert_number": "333", "card_title": "July Card", "date_added": "2026-07-01", "inventory_value": 50},
        ]

        self.assertEqual([row["cert_number"] for row in dummy._filtered_inventory_records(rows)], ["222"])

    def test_inventory_filter_applies_grader(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records

        dummy = InventoryDummy()
        dummy.inventory_person_var = FieldVar("")
        dummy.inventory_sport_var = FieldVar("")
        dummy.inventory_grader_var = FieldVar("psa")
        dummy.inventory_search_var = FieldVar("")
        dummy.inventory_min_var = FieldVar("")
        dummy.inventory_max_var = FieldVar("")
        rows = [
            {"status": "Active", "cert_number": "111", "card_title": "PSA Card", "grader": "PSA", "inventory_value": 50},
            {"status": "Active", "cert_number": "222", "card_title": "CGC Card", "grader": "CGC", "inventory_value": 50},
            {"status": "Active", "cert_number": "333", "card_title": "No Grader Card", "grader": "", "inventory_value": 50},
        ]

        self.assertEqual([row["cert_number"] for row in dummy._filtered_inventory_records(rows)], ["111"])

    def test_inventory_grader_filter_accepts_multiple_checked_graders(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records

        dummy = InventoryDummy()
        dummy.inventory_person_var = FieldVar("")
        dummy.inventory_sport_var = FieldVar("")
        dummy.inventory_grader_var = FieldVar("PSA, CGC")
        dummy.inventory_search_var = FieldVar("")
        dummy.inventory_min_var = FieldVar("")
        dummy.inventory_max_var = FieldVar("")
        rows = [
            {"status": "Active", "cert_number": "111", "card_title": "PSA Card", "grader": "PSA", "inventory_value": 50},
            {"status": "Active", "cert_number": "222", "card_title": "BGS Card", "grader": "BGS", "inventory_value": 50},
            {"status": "Active", "cert_number": "333", "card_title": "CGC Card", "grader": "CGC", "inventory_value": 50},
            {"status": "Active", "cert_number": "444", "card_title": "No Grader Card", "grader": "", "inventory_value": 50},
        ]

        self.assertEqual([row["cert_number"] for row in dummy._filtered_inventory_records(rows)], ["111", "333"])

    def test_inventory_filter_applies_card_year_from_title(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records
            _inventory_record_card_years = app.CardPipelineApp._inventory_record_card_years

        dummy = InventoryDummy()
        dummy.inventory_person_var = FieldVar("")
        dummy.inventory_sport_var = FieldVar("")
        dummy.inventory_grader_var = FieldVar("")
        dummy.inventory_year_var = FieldVar("2023")
        dummy.inventory_search_var = FieldVar("")
        dummy.inventory_min_var = FieldVar("")
        dummy.inventory_max_var = FieldVar("")
        rows = [
            {"status": "Active", "cert_number": "111", "card_title": "2022 Topps Chrome Test Card PSA 10", "inventory_value": 50},
            {"status": "Active", "cert_number": "222", "card_title": "2023 Panini Prizm Test Card PSA 10", "inventory_value": 50},
            {"status": "Active", "cert_number": "333", "year": "2023", "card_title": "No Year In Title", "inventory_value": 50},
            {"status": "Sold", "cert_number": "444", "card_title": "2023 Sold Card", "inventory_value": 50},
        ]

        self.assertEqual([row["cert_number"] for row in dummy._filtered_inventory_records(rows)], ["222"])

    def test_table_sorting_handles_money_and_blank_values(self) -> None:
        class SortDummy:
            _money_value = app.CardPipelineApp._money_value
            _expense_related_label = app.CardPipelineApp._expense_related_label
            _profit_added_sort_value = app.CardPipelineApp._profit_added_sort_value
            _record_sort_value = app.CardPipelineApp._record_sort_value
            _sorted_records = app.CardPipelineApp._sorted_records

        dummy = SortDummy()
        inventory_rows = [
            {"cert_number": "1", "purchase_price": "$9.00"},
            {"cert_number": "2", "purchase_price": "$100.00"},
            {"cert_number": "3", "purchase_price": ""},
        ]

        self.assertEqual([row["cert_number"] for row in dummy._sorted_records(inventory_rows, "purchase", False, "inventory")], ["1", "2", "3"])
        self.assertEqual([row["cert_number"] for row in dummy._sorted_records(inventory_rows, "purchase", True, "inventory")], ["2", "1", "3"])

        profit_rows = [
            {"card_title": "A", "profit": "$5.00"},
            {"card_title": "B", "profit": "$15.00"},
            {"card_title": "C", "profit": ""},
        ]
        self.assertEqual([row["card_title"] for row in dummy._sorted_records(profit_rows, "profit", True, "profit")], ["B", "A", "C"])

    def test_profit_date_sort_uses_added_timestamp_within_same_day(self) -> None:
        class SortDummy:
            _money_value = app.CardPipelineApp._money_value
            _expense_related_label = app.CardPipelineApp._expense_related_label
            _profit_added_sort_value = app.CardPipelineApp._profit_added_sort_value
            _record_sort_value = app.CardPipelineApp._record_sort_value
            _sorted_records = app.CardPipelineApp._sorted_records

        dummy = SortDummy()
        rows = [
            {"card_title": "Older card", "date_added": "2026-07-13", "ledger_added_at": "2026-07-13T10:00:00"},
            {"card_title": "Newer card", "date_added": "2026-07-13", "ledger_added_at": "2026-07-13T10:05:00"},
            {"card_title": "Shipping", "record_type": "expense", "date_added": "2026-07-13", "expense_id": "20260713100600000000"},
            {"card_title": "Yesterday", "date_added": "2026-07-12", "ledger_added_at": "2026-07-12T23:59:00"},
        ]

        sorted_rows = dummy._sorted_records(rows, "date", True, "profit", "Sold Cards")

        self.assertEqual([row["card_title"] for row in sorted_rows], ["Shipping", "Newer card", "Older card", "Yesterday"])

    def test_assignment_explanation_shows_payout_rate_and_category(self) -> None:
        class ExplainDummy:
            _assignment_decision_detail = app.CardPipelineApp._assignment_decision_detail

        dummy = ExplainDummy()
        decision = assignment_engine.AssignmentDecision(
            "Arena Club",
            True,
            2622.0,
            "accepted and payout tier matched",
            2760.0,
            0.95,
            "Vintage",
            100,
            None,
        )

        detail = dummy._assignment_decision_detail(decision)

        self.assertIn("95%", detail)
        self.assertIn("category: Vintage", detail)
        self.assertIn("$100", detail)
        self.assertNotIn("accepted and payout tier matched", detail)

    def test_inventory_refresh_purges_non_active_rows_from_ledger(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records
            refresh_inventory_tab = app.CardPipelineApp.refresh_inventory_tab

        with TemporaryDirectory() as tmp:
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = InventoryDummy()
            try:
                dummy._save_inventory_ledger([
                    dummy._normalize_inventory_record({"assigned_person": "Kevin Hambone", "cert_number": "111", "source_sheet": "A.xlsx", "status": "Active"}),
                    dummy._normalize_inventory_record({"assigned_person": "Kevin Hambone", "cert_number": "222", "source_sheet": "B.xlsx", "status": "Sold"}),
                    dummy._normalize_inventory_record({"assigned_person": "Kevin Hambone", "cert_number": "333", "source_sheet": "C.xlsx", "status": "Company Sheet"}),
                ])

                dummy.refresh_inventory_tab()

                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual([row["cert_number"] for row in ledger], ["111"])
            finally:
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_inventory_table_values_can_be_copied_without_editing(self) -> None:
        class FakeTree:
            def item(self, _row_id, _option):
                return ("2026-06-18", "Kevin Hambone", "Basketball", "12345678", "PSA", "Test Card")

        class FakeStatus:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class InventoryDummy:
            _tree_cell_text = app.CardPipelineApp._tree_cell_text
            _inventory_tree_cell_text = app.CardPipelineApp._inventory_tree_cell_text
            _tree_row_text = app.CardPipelineApp._tree_row_text
            _inventory_tree_row_text = app.CardPipelineApp._inventory_tree_row_text
            _copy_inventory_text = app.CardPipelineApp._copy_inventory_text
            copy_inventory_cell_value = app.CardPipelineApp.copy_inventory_cell_value
            copy_inventory_row_values = app.CardPipelineApp.copy_inventory_row_values
            copy_tree_cell_value = app.CardPipelineApp.copy_tree_cell_value
            copy_tree_row_values = app.CardPipelineApp.copy_tree_row_values

            def __init__(self):
                self.inventory_tree = FakeTree()
                self.status_var = FakeStatus()
                self.clipboard = ""

            def clipboard_clear(self):
                self.clipboard = ""

            def clipboard_append(self, text):
                self.clipboard = text

            def clipboard_get(self):
                return self.clipboard

            def update(self):
                pass

        dummy = InventoryDummy()
        dummy.copy_inventory_cell_value("row-1", "#4")
        self.assertEqual(dummy.clipboard, "12345678")
        self.assertEqual(dummy.status_var.value, "Copied inventory cell.")

        dummy.copy_inventory_row_values("row-1")
        self.assertEqual(dummy.clipboard, "2026-06-18\tKevin Hambone\tBasketball\t12345678\tPSA\tTest Card")
        self.assertEqual(dummy.status_var.value, "Copied inventory row.")

        dummy.copy_tree_cell_value(dummy.inventory_tree, "row-1", "#5", "sheet cell")
        self.assertEqual(dummy.clipboard, "PSA")
        self.assertEqual(dummy.status_var.value, "Copied sheet cell.")

        dummy.copy_tree_row_values(dummy.inventory_tree, "row-1", "sheet row")
        self.assertEqual(dummy.clipboard, "2026-06-18\tKevin Hambone\tBasketball\t12345678\tPSA\tTest Card")
        self.assertEqual(dummy.status_var.value, "Copied sheet row.")

    def test_copy_replaces_clipboard_when_first_clear_does_not_settle(self) -> None:
        class ClipboardDummy:
            _copy_inventory_text = app.CardPipelineApp._copy_inventory_text

            def __init__(self):
                self.clipboard = "12345678"
                self.clear_calls = 0

            def clipboard_clear(self):
                self.clear_calls += 1
                if self.clear_calls > 1:
                    self.clipboard = ""

            def clipboard_append(self, text):
                self.clipboard += text

            def clipboard_get(self):
                return self.clipboard

            def update(self):
                pass

        dummy = ClipboardDummy()
        dummy._copy_inventory_text("12345678")

        self.assertEqual(dummy.clipboard, "12345678")
        self.assertEqual(dummy.clear_calls, 2)

    def test_inventory_table_shows_source_values_separately(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        class FakeTree:
            def __init__(self):
                self.rows = []

            def get_children(self):
                return []

            def delete(self, *_args):
                return None

            def insert(self, *_args, **kwargs):
                self.rows.append(kwargs["values"])
                return "row-1"

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records
            refresh_inventory_tab = app.CardPipelineApp.refresh_inventory_tab

            def __init__(self):
                self.inventory_person_var = FieldVar("")
                self.inventory_sport_var = FieldVar("")
                self.inventory_search_var = FieldVar("")
                self.inventory_min_var = FieldVar("")
                self.inventory_max_var = FieldVar("")
                self.inventory_tree = FakeTree()
                self.inventory_metric_var = FieldVar("")
                self.inventory_status_var = FieldVar("")

            def _load_inventory_ledger(self):
                return [
                    self._normalize_inventory_record(
                        {
                            "date_added": "2026-06-18",
                            "assigned_person": "Kevin Hambone",
                            "sport": "football",
                            "cert_number": "141119049",
                            "grader": "PSA",
                            "card_title": "2025 Panini Mosaic Elevate Travis Hunter PSA 10",
                            "purchase_price": 40,
                            "inventory_value": 41,
                            "card_ladder_value": 41,
                            "card_ladder_comps_average": 38.87,
                            "cy_value": "",
                            "cy_confidence": "",
                            "paid_with": "Cash",
                            "best_company": "FANATICS",
                            "estimated_payout": 38.95,
                            "source_sheet": "SCOTSBORO_HAMBONE_6_16_26.xlsx",
                            "status": "Active",
                        }
                    )
                ]

            def _save_inventory_ledger(self, _rows):
                raise AssertionError("active-only refresh should not rewrite this fixture")

            def _refresh_person_combo_values(self):
                return None

        self.assertNotIn("value", app.INVENTORY_TABLE_COLUMNS)
        self.assertIn("card_ladder", app.INVENTORY_TABLE_COLUMNS)
        self.assertIn("paid_with", app.INVENTORY_TABLE_COLUMNS)
        self.assertEqual(app.INVENTORY_TABLE_COLUMNS.index("paid_with"), app.INVENTORY_TABLE_COLUMNS.index("payout") + 1)
        self.assertIn("comps", app.INVENTORY_TABLE_COLUMNS)
        self.assertIn("cy_estimate", app.INVENTORY_TABLE_COLUMNS)
        self.assertIn("cy_confidence", app.INVENTORY_TABLE_COLUMNS)

        dummy = InventoryDummy()
        dummy.refresh_inventory_tab()

        row = dummy.inventory_tree.rows[0]
        columns = app.INVENTORY_TABLE_COLUMNS
        self.assertEqual(row[columns.index("card_ladder")], "$41.00")
        self.assertEqual(row[columns.index("paid_with")], "Cash")
        self.assertEqual(row[columns.index("comps")], "$38.87")
        self.assertEqual(row[columns.index("cy_estimate")], "")
        self.assertEqual(row[columns.index("company")], "FANATICS")
        self.assertEqual(row[columns.index("payout")], "$38.95")
        self.assertIn("Source Value: $41.00", dummy.inventory_metric_var.value)

    def test_filtered_inventory_refresh_only_enriches_visible_rows(self) -> None:
        class FieldVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        class FakeTree:
            def get_children(self):
                return []

            def delete(self, *_args):
                return None

            def insert(self, *_args, **_kwargs):
                return "row"

        class FakeAssignment:
            def recommend(self, row, person=""):
                return types.SimpleNamespace(company=f"{person} Club", payout=88)

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _enrich_inventory_record_assignment = app.CardPipelineApp._enrich_inventory_record_assignment
            _inventory_sport_filter_values = app.CardPipelineApp._inventory_sport_filter_values
            _filtered_inventory_records = app.CardPipelineApp._filtered_inventory_records
            refresh_inventory_tab = app.CardPipelineApp.refresh_inventory_tab
            update_inventory_payouts = app.CardPipelineApp.update_inventory_payouts
            _refresh_person_combo_values = lambda self: None

        with TemporaryDirectory() as tmp:
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = InventoryDummy()
            dummy.assignment_engine = FakeAssignment()
            dummy.inventory_person_var = FieldVar("Kevin")
            dummy.inventory_sport_var = FieldVar("")
            dummy.inventory_search_var = FieldVar("")
            dummy.inventory_min_var = FieldVar("")
            dummy.inventory_max_var = FieldVar("")
            dummy.inventory_tree = FakeTree()
            dummy.inventory_metric_var = FieldVar("")
            dummy.inventory_status_var = FieldVar("")
            try:
                dummy._save_inventory_ledger([
                    dummy._normalize_inventory_record({"assigned_person": "Kevin Hambone", "cert_number": "1", "source_sheet": "A.xlsx", "card_title": "Kevin Card", "best_company": "Old", "estimated_payout": 1}),
                    dummy._normalize_inventory_record({"assigned_person": "Lucas", "cert_number": "2", "source_sheet": "B.xlsx", "card_title": "Lucas Card", "best_company": "Old", "estimated_payout": 1}),
                ])

                dummy.update_inventory_payouts()

                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                by_person = {record["assigned_person"]: record for record in ledger}
                self.assertEqual(by_person["Kevin Hambone"]["best_company"], "Kevin Hambone Club")
                self.assertEqual(by_person["Kevin Hambone"]["estimated_payout"], 88)
                self.assertEqual(by_person["Lucas"]["best_company"], "Old")
                self.assertEqual(by_person["Lucas"]["estimated_payout"], 1)
                self.assertEqual(dummy.inventory_status_var.value, "Updated payouts for 1 visible inventory card(s); changed 1.")
            finally:
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_inventory_record_assignment_enrichment_adds_company_and_payout(self) -> None:
        class FakeAssignment:
            def recommend(self, row, person=""):
                self.last_row = row
                return types.SimpleNamespace(company="Arena Club", payout=95, source_value=150)

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _enrich_inventory_record_assignment = app.CardPipelineApp._enrich_inventory_record_assignment

        dummy = InventoryDummy()
        dummy.assignment_engine = FakeAssignment()
        record = dummy._enrich_inventory_record_assignment(
            {
                "assigned_person": "Hambone",
                "cert_number": "123",
                "card_title": "Test Card",
                "source_sheet": "Lot.xlsx",
                "purchase_price": 40,
                "inventory_value": 100,
                "card_ladder_value": 150,
                "card_ladder_comps_average": 100,
                "cy_value": 90,
            }
        )
        self.assertEqual(record["best_company"], "Arena Club")
        self.assertEqual(record["estimated_payout"], 95)
        self.assertEqual(record["inventory_value"], 150)
        self.assertEqual(record["card_ladder_value"], 150)
        self.assertEqual(record["card_ladder_comps_average"], 100)
        self.assertEqual(dummy.assignment_engine.last_row.card_ladder_value, 150)
        self.assertEqual(dummy.assignment_engine.last_row.card_ladder_comps_average, 100)

    def test_inventory_assignment_recomputes_nobody_takes_with_stale_payout(self) -> None:
        class FakeAssignment:
            def recommend(self, row, person=""):
                self.last_person = person
                return types.SimpleNamespace(company="Fanatics", payout=476.16, source_value=512)

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _enrich_inventory_record_assignment = app.CardPipelineApp._enrich_inventory_record_assignment

        dummy = InventoryDummy()
        dummy.assignment_engine = FakeAssignment()
        record = dummy._enrich_inventory_record_assignment(
            {
                "assigned_person": "Mikey",
                "cert_number": "0018849545",
                "grader": "BGS",
                "sport": "basketball",
                "card_title": "2024-25 Panini National Treasures Emerald #111 Cam Spencer JSY AU BGS 8.5",
                "source_sheet": "smalltown.sportscards_7_7_26.xlsx",
                "purchase_price": 440,
                "card_ladder_value": 512,
                "card_ladder_comps_average": 480,
                "best_company": app.NO_COMPANY_TAKES_LABEL,
                "estimated_payout": 476.16,
            }
        )

        self.assertEqual(record["best_company"], "Fanatics")
        self.assertEqual(record["estimated_payout"], 476.16)
        self.assertEqual(record["inventory_value"], 512)
        self.assertEqual(dummy.assignment_engine.last_person, "Mikey")

    def test_inventory_assignment_uses_saved_sport_for_company_rules(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _enrich_inventory_record_assignment = app.CardPipelineApp._enrich_inventory_record_assignment

        dummy = InventoryDummy()
        dummy.assignment_engine = assignment_engine.AssignmentEngine([
            assignment_engine.AssignmentCompany(
                "Arena Club",
                assignment_engine.CompanyRules(
                    accept_all=True,
                    blocks=[assignment_engine.AssignmentRule("Football Anthony Richardson", block=True)],
                ),
                [assignment_engine.PayoutTier(0, 500, 1.05)],
            ),
            assignment_engine.AssignmentCompany(
                "Fanatics",
                assignment_engine.CompanyRules(accept_all=True),
                [assignment_engine.PayoutTier(0, 500, 0.95)],
            ),
        ])

        record = dummy._enrich_inventory_record_assignment(
            {
                "assigned_person": "Tyler Hamlin",
                "cert_number": "98514082",
                "sport": "football",
                "grader": "PSA",
                "card_title": "2023 Panini Select 117 Anthony Richardson PSA 10",
                "source_sheet": "TYLER_NASHVILLE_INVENTORY.xlsx",
                "purchase_price": 20,
                "card_ladder_value": 31,
                "card_ladder_comps_average": 31,
            },
            force=True,
        )

        self.assertEqual(record["best_company"], "Fanatics")
        self.assertEqual(record["estimated_payout"], 29.45)

    def test_inventory_refresh_hydrates_source_values_for_card_ladder_company(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _inventory_source_sheet_path = app.CardPipelineApp._inventory_source_sheet_path
            _hydrate_inventory_record_source_values = app.CardPipelineApp._hydrate_inventory_record_source_values
            _enrich_inventory_record_assignment = app.CardPipelineApp._enrich_inventory_record_assignment

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            received_dir = root / "RECEIVED SHEETS"
            received_dir.mkdir()
            source = received_dir / "Lot.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Certification Number", "Grader", "Card Description", "Purchase Price", "Card Ladder Value", "Comps"])
            sheet.append(["87378266", "PSA", "2023 Panini Phoenix Playing With Fire Bryce Young PSA 10", 20, 44, 32])
            workbook.save(source)

            old_received = app.RECEIVED_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            app.RECEIVED_SHEETS_DIR = received_dir
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            app.INCOMING_SHEETS_DIR = root / "INCOMING SHEETS"
            dummy = InventoryDummy()
            dummy.assignment_engine = assignment_engine.AssignmentEngine([
                assignment_engine.AssignmentCompany(
                    "Arena Club",
                    assignment_engine.CompanyRules(accept_all=True),
                    [assignment_engine.PayoutTier(0, 500, 1.0)],
                    value_source="comps",
                ),
                assignment_engine.AssignmentCompany(
                    "Fanatics",
                    assignment_engine.CompanyRules(accept_all=True),
                    [assignment_engine.PayoutTier(0, 500, 0.95)],
                    value_source="card_ladder",
                ),
            ])
            try:
                record = dummy._enrich_inventory_record_assignment(
                    {
                        "assigned_person": "Kevin Hambone",
                        "cert_number": "87378266",
                        "grader": "PSA",
                        "card_title": "2023 Panini Phoenix Playing With Fire Bryce Young PSA 10",
                        "purchase_price": 20,
                        "inventory_value": 32,
                        "source_sheet": "Lot.xlsx",
                        "status": "Active",
                    },
                    force=True,
                )

                self.assertEqual(record["best_company"], "Fanatics")
                self.assertEqual(record["estimated_payout"], 41.8)
                self.assertEqual(record["inventory_value"], 44)
                self.assertEqual(record["card_ladder_value"], 44)
                self.assertEqual(record["card_ladder_comps_average"], 32)
            finally:
                app.RECEIVED_SHEETS_DIR = old_received
                app.WORKING_SHEETS_DIR = old_working
                app.INCOMING_SHEETS_DIR = old_incoming

    def test_add_inventory_records_recalculates_stale_nobody_takes_assignment(self) -> None:
        class FakeAssignment:
            def recommend(self, row, person=""):
                self.last_person = person
                return types.SimpleNamespace(company="Arena Club", payout=95)

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _enrich_inventory_record_assignment = app.CardPipelineApp._enrich_inventory_record_assignment
            add_inventory_records = app.CardPipelineApp.add_inventory_records
            refresh_inventory_tab = lambda self: None

        with TemporaryDirectory() as tmp:
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = InventoryDummy()
            dummy.assignment_engine = FakeAssignment()
            try:
                dummy.add_inventory_records([
                    {
                        "assigned_person": "Kevin Hambone",
                        "cert_number": "123",
                        "card_title": "Test Card",
                        "source_sheet": "Hambone.xlsx",
                        "purchase_price": 40,
                        "inventory_value": 100,
                        "best_company": app.NO_COMPANY_TAKES_LABEL,
                    }
                ])

                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(ledger[0]["best_company"], "Arena Club")
                self.assertEqual(ledger[0]["estimated_payout"], 95)
                self.assertEqual(dummy.assignment_engine.last_person, "Kevin Hambone")
            finally:
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_retarget_inventory_rows_for_source_changes_owner_without_duplicates(self) -> None:
        class FakeAssignment:
            def recommend(self, row, person=""):
                self.last_person = person
                return types.SimpleNamespace(company=f"{person} Club", payout=88)

        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _enrich_inventory_record_assignment = app.CardPipelineApp._enrich_inventory_record_assignment
            _retarget_inventory_rows_for_source = app.CardPipelineApp._retarget_inventory_rows_for_source

        with TemporaryDirectory() as tmp:
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = InventoryDummy()
            dummy.assignment_engine = FakeAssignment()
            try:
                dummy._save_inventory_ledger([
                    dummy._normalize_inventory_record({"assigned_person": "Lucas", "cert_number": "123", "source_sheet": "Lot A.xlsx", "card_title": "Test Card", "inventory_value": 100, "best_company": "Arena Club", "estimated_payout": 95}),
                    dummy._normalize_inventory_record({"assigned_person": "Mikey", "cert_number": "123", "source_sheet": "Lot A.xlsx", "card_title": "Test Card", "inventory_value": 100, "best_company": "Arena Club", "estimated_payout": 90}),
                ])

                self.assertEqual(dummy._retarget_inventory_rows_for_source("Lot A.xlsx", "Mikey"), 1)

                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(len(ledger), 1)
                self.assertEqual(ledger[0]["assigned_person"], "Mikey")
                self.assertEqual(ledger[0]["inventory_key"], "123|lot a.xlsx|mikey")
                self.assertEqual(ledger[0]["best_company"], "Mikey Club")
                self.assertEqual(ledger[0]["estimated_payout"], 88)
            finally:
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_received_inventory_reconcile_skips_company_sheet_rows(self) -> None:
        class ReconcileDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _company_sheet_source_cert_keys = app.CardPipelineApp._company_sheet_source_cert_keys
            _received_inventory_accounted_source_cert_keys = app.CardPipelineApp._received_inventory_accounted_source_cert_keys
            _received_certs_in_workbook = app.CardPipelineApp._received_certs_in_workbook
            _received_inventory_candidate_records_for_sheet = app.CardPipelineApp._received_inventory_candidate_records_for_sheet
            _received_inventory_candidate_records = app.CardPipelineApp._received_inventory_candidate_records
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _load_inventory_ledger = lambda self: []
            _load_profit_ledger = lambda self: []
            _inventory_deleted_source_cert_keys = lambda self: set()
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _is_personal_lucas = lambda self: False
            _personal_default_person = app.CardPipelineApp._personal_default_person

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            received_dir = root / "RECEIVED SHEETS"
            received_dir.mkdir()
            received_path = received_dir / "Hambone Lot.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Certification Number", "Grader", "Card Description", "Purchase Price", "Comps"])
            sheet.append(["111", "PSA", "Hambone Inventory Card", 40, 100])
            sheet.append(["222", "PSA", "Company Card", 50, 120])
            workbook.save(received_path)

            company_dir = root / "COMPANY SHEETS"
            company_path = company_dir / "Arena Club" / "Arena Club.xlsx"
            company_path.parent.mkdir(parents=True)
            company_workbook = Workbook()
            company_sheet = company_workbook.active
            company_sheet.title = "Week of 2026-06-15"
            company_sheet.append(["Date Added", "Source Sheet", "Source", "Certification Number", "Grader", "Card Description", "Purchase Price", "Card Ladder Value", "Comps", "CY Estimate", "CY Confidence", "Best Company", "Estimated Payout", "Status", "Notes"])
            company_sheet.append(["2026-06-17", "Hambone Lot.xlsx", "", "222", "PSA", "Company Card", 50, "", 120, "", "", "Arena Club", 90, "Received", ""])
            company_workbook.save(company_path)

            old_received = app.RECEIVED_SHEETS_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_company = app.COMPANY_SHEETS_DIR
            app.RECEIVED_SHEETS_DIR = received_dir
            app.INCOMING_SHEETS_DIR = root / "INCOMING SHEETS"
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            app.COMPANY_SHEETS_DIR = company_dir
            dummy = ReconcileDummy()
            dummy.home_sheet_markers = {"Received|Hambone Lot.xlsx": {"assigned_person": "Hambone"}}
            try:
                records = dummy._received_inventory_candidate_records()
                self.assertEqual(len(records), 1)
                self.assertEqual(records[0]["cert_number"], "111")
                self.assertEqual(records[0]["assigned_person"], "Hambone")
            finally:
                app.RECEIVED_SHEETS_DIR = old_received
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working
                app.COMPANY_SHEETS_DIR = old_company

    def test_received_inventory_reconcile_defaults_unassigned_sheet_markers(self) -> None:
        class ReconcileDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _company_sheet_source_cert_keys = app.CardPipelineApp._company_sheet_source_cert_keys
            _received_inventory_accounted_source_cert_keys = app.CardPipelineApp._received_inventory_accounted_source_cert_keys
            _received_certs_in_workbook = app.CardPipelineApp._received_certs_in_workbook
            _received_inventory_candidate_records_for_sheet = app.CardPipelineApp._received_inventory_candidate_records_for_sheet
            _received_inventory_candidate_records = app.CardPipelineApp._received_inventory_candidate_records
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _load_inventory_ledger = lambda self: []
            _load_profit_ledger = lambda self: []
            _inventory_deleted_source_cert_keys = lambda self: set()
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _is_personal_lucas = lambda self: False
            _personal_default_person = app.CardPipelineApp._personal_default_person

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            received_dir = root / "RECEIVED SHEETS"
            received_dir.mkdir()
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Certification Number", "Grader", "Card Description", "Purchase Price", "Comps"])
            sheet.append(["151740304", "PSA", "Deleted Test Card", 40, 100])
            workbook.save(received_dir / "HAMBONE_MIKEY_6_16_TEST.xlsx")

            old_received = app.RECEIVED_SHEETS_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_company = app.COMPANY_SHEETS_DIR
            app.RECEIVED_SHEETS_DIR = received_dir
            app.INCOMING_SHEETS_DIR = root / "INCOMING SHEETS"
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            app.COMPANY_SHEETS_DIR = root / "COMPANY SHEETS"
            dummy = ReconcileDummy()
            dummy.home_sheet_markers = {}
            try:
                records = dummy._received_inventory_candidate_records()
                self.assertEqual(len(records), 1)
                self.assertEqual(records[0]["cert_number"], "151740304")
                self.assertEqual(records[0]["assigned_person"], "Unassigned")
            finally:
                app.RECEIVED_SHEETS_DIR = old_received
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working
                app.COMPANY_SHEETS_DIR = old_company

    def test_personal_received_inventory_reconcile_defaults_blank_marker_to_mikey(self) -> None:
        class ReconcileDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _company_sheet_source_cert_keys = lambda self: set()
            _received_inventory_accounted_source_cert_keys = app.CardPipelineApp._received_inventory_accounted_source_cert_keys
            _received_certs_in_workbook = app.CardPipelineApp._received_certs_in_workbook
            _received_inventory_candidate_records_for_sheet = app.CardPipelineApp._received_inventory_candidate_records_for_sheet
            _received_inventory_candidate_records = app.CardPipelineApp._received_inventory_candidate_records
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _load_inventory_ledger = lambda self: []
            _load_profit_ledger = lambda self: []
            _inventory_deleted_source_cert_keys = lambda self: set()
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _is_personal_lucas = lambda self: True
            _personal_default_person = app.CardPipelineApp._personal_default_person

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            received_dir = root / "RECEIVED SHEETS"
            received_dir.mkdir()
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Certification Number", "Grader", "Sport", "Card Description", "Purchase Price", "Comps"])
            sheet.append(["57355243", "PSA", "football", "2020 Panini Mosaic 213 Tee Higgins Mosaic PSA 10", 30, 40])
            workbook.save(received_dir / "cardking10x_tees.xlsx")

            old_received = app.RECEIVED_SHEETS_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_company = app.COMPANY_SHEETS_DIR
            app.RECEIVED_SHEETS_DIR = received_dir
            app.INCOMING_SHEETS_DIR = root / "INCOMING SHEETS"
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            app.COMPANY_SHEETS_DIR = root / "COMPANY SHEETS"
            dummy = ReconcileDummy()
            dummy.home_sheet_markers = {"Received|cardking10x_tees.xlsx": {"all_received": True}}
            try:
                records = dummy._received_inventory_candidate_records()
                self.assertEqual(len(records), 1)
                self.assertEqual(records[0]["cert_number"], "57355243")
                self.assertEqual(records[0]["assigned_person"], "Mikey")
            finally:
                app.RECEIVED_SHEETS_DIR = old_received
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working
                app.COMPANY_SHEETS_DIR = old_company

    def test_received_inventory_reconcile_skips_already_sold_profit_rows(self) -> None:
        class ReconcileDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _company_sheet_source_cert_keys = lambda self: set()
            _received_inventory_accounted_source_cert_keys = app.CardPipelineApp._received_inventory_accounted_source_cert_keys
            _received_certs_in_workbook = app.CardPipelineApp._received_certs_in_workbook
            _received_inventory_candidate_records_for_sheet = app.CardPipelineApp._received_inventory_candidate_records_for_sheet
            _received_inventory_candidate_records = app.CardPipelineApp._received_inventory_candidate_records
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _inventory_deleted_source_cert_keys = lambda self: set()
            _is_personal_lucas = lambda self: False
            _personal_default_person = app.CardPipelineApp._personal_default_person

            def _load_inventory_ledger(self):
                return []

            def _load_profit_ledger(self):
                return [
                    {
                        "record_type": "sale",
                        "source_sheet": "Lot A.xlsx",
                        "original_source_sheet": "Lot A.xlsx",
                        "cert_number": "222",
                        "card_title": "Already Sold Card",
                        "purchase_price": 50,
                        "sale_price": 75,
                    }
                ]

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            received_dir = root / "RECEIVED SHEETS"
            received_dir.mkdir()
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Certification Number", "Grader", "Card Description", "Purchase Price", "Comps"])
            sheet.append(["111", "PSA", "Still Active Candidate", 40, 100])
            sheet.append(["222", "PSA", "Already Sold Card", 50, 120])
            workbook.save(received_dir / "Lot A.xlsx")

            old_received = app.RECEIVED_SHEETS_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            app.RECEIVED_SHEETS_DIR = received_dir
            app.INCOMING_SHEETS_DIR = root / "INCOMING SHEETS"
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            dummy = ReconcileDummy()
            dummy.home_sheet_markers = {"Received|Lot A.xlsx": {"assigned_person": "Kevin Hambone"}}
            try:
                records = dummy._received_inventory_candidate_records()
                self.assertEqual([record["cert_number"] for record in records], ["111"])
            finally:
                app.RECEIVED_SHEETS_DIR = old_received
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working

    def test_received_inventory_reconcile_skips_deleted_tombstone_rows(self) -> None:
        class ReconcileDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _company_sheet_source_cert_keys = lambda self: set()
            _received_inventory_accounted_source_cert_keys = app.CardPipelineApp._received_inventory_accounted_source_cert_keys
            _received_certs_in_workbook = app.CardPipelineApp._received_certs_in_workbook
            _received_inventory_candidate_records_for_sheet = app.CardPipelineApp._received_inventory_candidate_records_for_sheet
            _received_inventory_candidate_records = app.CardPipelineApp._received_inventory_candidate_records
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _is_personal_lucas = lambda self: False
            _personal_default_person = app.CardPipelineApp._personal_default_person

            def _load_inventory_ledger(self):
                return []

            def _load_profit_ledger(self):
                return []

            def _inventory_deleted_source_cert_keys(self):
                return {("lot a.xlsx", "222")}

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            received_dir = root / "RECEIVED SHEETS"
            received_dir.mkdir()
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Certification Number", "Grader", "Card Description", "Purchase Price", "Comps"])
            sheet.append(["111", "PSA", "Still Active Candidate", 40, 100])
            sheet.append(["222", "PSA", "Deleted Inventory Card", 50, 120])
            workbook.save(received_dir / "Lot A.xlsx")

            old_received = app.RECEIVED_SHEETS_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            app.RECEIVED_SHEETS_DIR = received_dir
            app.INCOMING_SHEETS_DIR = root / "INCOMING SHEETS"
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            dummy = ReconcileDummy()
            dummy.home_sheet_markers = {"Received|Lot A.xlsx": {"assigned_person": "Kevin Hambone"}}
            try:
                records = dummy._received_inventory_candidate_records()
                self.assertEqual([record["cert_number"] for record in records], ["111"])
            finally:
                app.RECEIVED_SHEETS_DIR = old_received
                app.INCOMING_SHEETS_DIR = old_incoming
                app.WORKING_SHEETS_DIR = old_working

    def test_inventory_refresh_enrich_preserves_non_active_ledger_rows(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            refresh_inventory_tab = app.CardPipelineApp.refresh_inventory_tab

            def __init__(self):
                self.inventory_filter_after_id = None
                self.saved_rows = None

            def _load_inventory_ledger(self):
                return [
                    {"cert_number": "111", "source_sheet": "Lot A.xlsx", "assigned_person": "Kevin", "status": "Active", "card_title": "Active Card"},
                    {"cert_number": "222", "source_sheet": "Lot A.xlsx", "assigned_person": "Kevin", "status": "Sold", "card_title": "Sold Card"},
                ]

            def _save_inventory_ledger(self, rows):
                self.saved_rows = rows

            def _enrich_inventory_record_assignment(self, record, force=False):
                updated = dict(record)
                updated["best_company"] = "Arena Club"
                return updated

            def _filtered_inventory_records(self, rows):
                return rows

        dummy = InventoryDummy()
        dummy.refresh_inventory_tab(enrich=True)

        self.assertIsNotNone(dummy.saved_rows)
        self.assertEqual([row["cert_number"] for row in dummy.saved_rows], ["222", "111"])
        self.assertEqual(dummy.saved_rows[0]["status"], "Sold")
        self.assertEqual(dummy.saved_rows[1]["best_company"], "Arena Club")

    def test_received_inventory_candidates_preserve_source_sheet_sport(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_sport_from_value = app.CardPipelineApp._inventory_sport_from_value
            _received_inventory_candidate_records_for_sheet = app.CardPipelineApp._received_inventory_candidate_records_for_sheet

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "manual-baseball.xlsx"
            write_working_sheet(
                path,
                [
                    WorkbookRow(
                        excel_row=2,
                        cert_number="0010355805",
                        grader="BGS",
                        card_title="2017 Bowman Chrome Prospect Autographs Gold Shimmer Refractors #CPACF Clint Frazier BGS 9.5",
                        category="baseball",
                        card_ladder_comps_average=432,
                    )
                ],
            )

            records = InventoryDummy()._received_inventory_candidate_records_for_sheet("Received", path, "James Copeland", set())

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["sport"], "baseball")

    def test_create_raw_rows_get_item_ids_before_working_sheet_save(self) -> None:
        class CreateDummy:
            _next_raw_item_id = app.CardPipelineApp._next_raw_item_id
            _raw_item_id_namespace = lambda self: "TEAM"
            _ensure_raw_item_ids_for_rows = app.CardPipelineApp._ensure_raw_item_ids_for_rows

            def _load_inventory_ledger(self):
                return [{"item_id": f"RAW-TEAM-{datetime.now().strftime('%Y%m%d')}-0003"}]

        rows = [
            WorkbookRow(excel_row=2, cert_number="", grader="", card_title="Raw Michael Jordan Insert", category="basketball"),
            WorkbookRow(excel_row=3, cert_number="", grader="", card_title="Raw Kobe Bryant Insert", category="basketball"),
            WorkbookRow(excel_row=4, cert_number="12345678", grader="PSA", card_title="Graded Card", category="football"),
        ]
        added = CreateDummy()._ensure_raw_item_ids_for_rows(rows)

        self.assertEqual(added, 2)
        self.assertEqual(rows[0].item_id, f"RAW-TEAM-{datetime.now().strftime('%Y%m%d')}-0004")
        self.assertEqual(rows[1].item_id, f"RAW-TEAM-{datetime.now().strftime('%Y%m%d')}-0005")
        self.assertEqual(rows[2].item_id, "")

    def test_raw_item_ids_are_namespaced_by_profile(self) -> None:
        class RawIdDummy:
            _next_raw_item_id = app.CardPipelineApp._next_raw_item_id

            def __init__(self, namespace: str):
                self.namespace = namespace

            def _raw_item_id_namespace(self):
                return self.namespace

        today = datetime.now().strftime("%Y%m%d")
        team_id = RawIdDummy("TEAM")._next_raw_item_id([])
        personal_id = RawIdDummy("MIKEY")._next_raw_item_id([])

        self.assertEqual(team_id, f"RAW-TEAM-{today}-0001")
        self.assertEqual(personal_id, f"RAW-MIKEY-{today}-0001")
        self.assertNotEqual(team_id, personal_id)

    def test_create_raw_ids_skip_live_incoming_sheet_ids(self) -> None:
        class RawIdDummy:
            _next_raw_item_id = app.CardPipelineApp._next_raw_item_id
            _ensure_raw_item_ids_for_rows = app.CardPipelineApp._ensure_raw_item_ids_for_rows
            _raw_item_id_namespace = lambda self: "MIKEY"

            def _load_inventory_ledger(self):
                return []

            def _live_sheet_raw_item_records(self):
                today = datetime.now().strftime("%Y%m%d")
                return [{"item_id": f"RAW-MIKEY-{today}-0001"}]

        today = datetime.now().strftime("%Y%m%d")
        row = WorkbookRow(excel_row=2, cert_number="", grader="", card_title="Raw New Card", category="baseball")

        RawIdDummy()._ensure_raw_item_ids_for_rows([row])

        self.assertEqual(row.item_id, f"RAW-MIKEY-{today}-0002")

    def test_stage_sheet_raw_id_backfill_writes_missing_item_ids(self) -> None:
        class RawIdDummy:
            _next_raw_item_id = app.CardPipelineApp._next_raw_item_id
            _raw_item_id_namespace = lambda self: "MIKEY"
            _workbook_header_lookup = app.CardPipelineApp._workbook_header_lookup
            _ensure_workbook_item_id_column = app.CardPipelineApp._ensure_workbook_item_id_column
            _ensure_raw_item_ids_in_sheet_paths = app.CardPipelineApp._ensure_raw_item_ids_in_sheet_paths

            def _load_inventory_ledger(self):
                return []

            def _live_sheet_raw_item_records(self):
                today = datetime.now().strftime("%Y%m%d")
                return [{"item_id": f"RAW-MIKEY-{today}-0001"}]

        today = datetime.now().strftime("%Y%m%d")
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "raw-lot.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Cards"
            sheet.append(["Item ID", "Certification Number", "Company", "Sport", "Card Description", "Purchase Price"])
            sheet.append(["", "", "RAW", "baseball", "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1", 400])
            sheet.append([f"RAW-MIKEY-{today}-0099", "", "", "", "", ""])
            workbook.save(path)

            result = RawIdDummy()._ensure_raw_item_ids_in_sheet_paths([path])

            self.assertEqual(result["ids_added"], 1)
            saved = load_workbook(path, read_only=True, data_only=True)
            try:
                self.assertEqual(saved["Cards"].cell(2, 1).value, f"RAW-MIKEY-{today}-0002")
                self.assertIsNone(saved["Cards"].cell(3, 1).value)
            finally:
                saved.close()

    def test_stage_sheet_raw_id_backfill_rewrites_duplicate_item_ids(self) -> None:
        class RawIdDummy:
            _next_raw_item_id = app.CardPipelineApp._next_raw_item_id
            _raw_item_id_namespace = lambda self: "MIKEY"
            _workbook_header_lookup = app.CardPipelineApp._workbook_header_lookup
            _ensure_workbook_item_id_column = app.CardPipelineApp._ensure_workbook_item_id_column
            _ensure_raw_item_ids_in_sheet_paths = app.CardPipelineApp._ensure_raw_item_ids_in_sheet_paths

            def _load_inventory_ledger(self):
                return []

            def _live_sheet_raw_item_records(self):
                return []

        today = datetime.now().strftime("%Y%m%d")
        duplicate_id = f"RAW-MIKEY-{today}-0001"
        with TemporaryDirectory() as tmp:
            first_path = Path(tmp) / "raw-a.xlsx"
            second_path = Path(tmp) / "raw-b.xlsx"
            for path, title in (
                (first_path, "2024 Topps Dynasty Gunnar Henderson Rookie Nike Patch Auto 1/1"),
                (second_path, "2015 Panini Immaculate Kobe Bryant Auto /60"),
            ):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Cards"
                sheet.append(["Item ID", "Certification Number", "Company", "Sport", "Card Description", "Purchase Price"])
                sheet.append([duplicate_id, "", "RAW", "baseball", title, 400])
                workbook.save(path)

            result = RawIdDummy()._ensure_raw_item_ids_in_sheet_paths([first_path, second_path])

            self.assertEqual(result["ids_added"], 1)
            first = load_workbook(first_path, read_only=True, data_only=True)
            second = load_workbook(second_path, read_only=True, data_only=True)
            try:
                self.assertEqual(first["Cards"].cell(2, 1).value, duplicate_id)
                self.assertEqual(second["Cards"].cell(2, 1).value, f"RAW-MIKEY-{today}-0002")
            finally:
                first.close()
                second.close()

    def test_working_sheet_writer_persists_raw_item_id_column(self) -> None:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "raw-create.xlsx"
            row = WorkbookRow(
                excel_row=2,
                cert_number="",
                grader="",
                card_title="Raw Michael Jordan Insert",
                category="basketball",
                item_id="RAW-20260709-0001",
            )
            write_working_sheet(path, [row], {2: "manual"})

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                self.assertEqual(sheet.cell(1, 1).value, "Item ID")
                self.assertEqual(sheet.cell(2, 1).value, "RAW-20260709-0001")
                self.assertEqual(sheet.cell(1, 2).value, "Certification Number")
                self.assertEqual(sheet.cell(2, 5).value, "Raw Michael Jordan Insert")
            finally:
                workbook.close()

    def test_home_all_received_marker_adds_received_sheet_inventory(self) -> None:
        class Status:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class MarkerDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _received_certs_in_workbook = app.CardPipelineApp._received_certs_in_workbook
            _company_sheet_source_cert_keys = lambda self: set()
            _received_inventory_candidate_records_for_sheet = app.CardPipelineApp._received_inventory_candidate_records_for_sheet
            _sync_received_sheet_inventory_to_ledger = app.CardPipelineApp._sync_received_sheet_inventory_to_ledger
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _sheet_path_for_stage = app.CardPipelineApp._sheet_path_for_stage
            _delete_sheet_marker = app.CardPipelineApp._delete_sheet_marker
            _move_home_sheet_to_stage = app.CardPipelineApp._move_home_sheet_to_stage
            _move_sheet_to_received = app.CardPipelineApp._move_sheet_to_received
            _move_received_sheet_to_incoming = app.CardPipelineApp._move_received_sheet_to_incoming
            _move_working_sheet_to_incoming = app.CardPipelineApp._move_working_sheet_to_incoming
            _marker_for_stage = app.CardPipelineApp._marker_for_stage
            _retarget_inventory_rows_for_source = lambda self, source, person: 0
            _is_personal_lucas = lambda self: False
            _personal_default_person = app.CardPipelineApp._personal_default_person
            _enrich_inventory_record_assignment = lambda self, record, force=False: record
            add_inventory_records = app.CardPipelineApp.add_inventory_records
            save_home_sheet_markers = app.CardPipelineApp.save_home_sheet_markers
            _load_sheet_markers = app.CardPipelineApp._load_sheet_markers
            _save_sheet_markers = app.CardPipelineApp._save_sheet_markers
            refresh_working_sheets = lambda self: None
            refresh_received_sheets = lambda self: None
            refresh_incoming_index = lambda self: None
            refresh_home = lambda self: None
            refresh_inventory_tab = lambda self, enrich=False: None

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            incoming_dir = root / "INCOMING SHEETS"
            received_dir = root / "RECEIVED SHEETS"
            incoming_dir.mkdir()
            received_dir.mkdir()
            source_path = incoming_dir / "Hambone Lot.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Certification Number", "Grader", "Card Description", "Purchase Price", "Comps"])
            sheet.append(["111", "PSA", "Hambone Inventory Card", 40, 100])
            workbook.save(source_path)

            old_pipeline = app.CARD_PIPELINE_DIR
            old_incoming = app.INCOMING_SHEETS_DIR
            old_received = app.RECEIVED_SHEETS_DIR
            old_working = app.WORKING_SHEETS_DIR
            old_company = app.COMPANY_SHEETS_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_markers = app.SHEET_MARKERS_PATH
            app.CARD_PIPELINE_DIR = root
            app.INCOMING_SHEETS_DIR = incoming_dir
            app.RECEIVED_SHEETS_DIR = received_dir
            app.WORKING_SHEETS_DIR = root / "WORKING SHEETS"
            app.COMPANY_SHEETS_DIR = root / "COMPANY SHEETS"
            app.INVENTORY_LEDGER_PATH = root / "inventory_ledger.json"
            app.SHEET_MARKERS_PATH = root / "sheet_markers.json"
            dummy = MarkerDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.home_selected_sheet_key = "Incoming|Hambone Lot.xlsx"
            dummy.home_sheet_markers = {"Incoming|Hambone Lot.xlsx": {"assigned_person": "Kevin Hambone"}}
            dummy.home_sheet_paths = {"Incoming": {"Hambone Lot.xlsx": source_path}, "Working": {}, "Received": {}}
            dummy.received_sheet_paths = {}
            dummy.deleted_sheet_marker_keys = set()
            dummy.status_var = Status()
            try:
                dummy.save_home_sheet_markers(
                    {
                        "incoming_proper": False,
                        "tracking_number": "",
                        "all_received": True,
                        "assigned_person": "Kevin Hambone",
                    }
                )

                inventory = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertTrue((received_dir / "Hambone Lot.xlsx").exists())
                self.assertEqual(len(inventory), 1)
                self.assertEqual(inventory[0]["cert_number"], "111")
                self.assertEqual(inventory[0]["assigned_person"], "Kevin Hambone")
                self.assertIn("Added 1 inventory row", dummy.status_var.value)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INCOMING_SHEETS_DIR = old_incoming
                app.RECEIVED_SHEETS_DIR = old_received
                app.WORKING_SHEETS_DIR = old_working
                app.COMPANY_SHEETS_DIR = old_company
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.SHEET_MARKERS_PATH = old_markers

    def test_inventory_rows_can_be_removed_for_deleted_source_sheet(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _remove_inventory_rows_for_source = app.CardPipelineApp._remove_inventory_rows_for_source

        with TemporaryDirectory() as tmp:
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = InventoryDummy()
            try:
                dummy._save_inventory_ledger([
                    dummy._normalize_inventory_record({"assigned_person": "Unassigned", "cert_number": "151740304", "source_sheet": "HAMBONE_MIKEY_6_16_TEST.xlsx"}),
                    dummy._normalize_inventory_record({"assigned_person": "Hambone", "cert_number": "222", "source_sheet": "SCOTSBORO_HAMBONE_6_16_26.xlsx"}),
                ])

                self.assertEqual(dummy._remove_inventory_rows_for_source("HAMBONE_MIKEY_6_16_TEST.xlsx"), 1)

                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(len(ledger), 1)
                self.assertEqual(ledger[0]["source_sheet"], "SCOTSBORO_HAMBONE_6_16_26.xlsx")
            finally:
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_inventory_rows_can_be_deleted_by_key(self) -> None:
        class InventoryDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_inventory_deleted_tombstones = app.CardPipelineApp._load_inventory_deleted_tombstones
            _save_inventory_deleted_tombstones = app.CardPipelineApp._save_inventory_deleted_tombstones
            _record_inventory_deleted_tombstones = app.CardPipelineApp._record_inventory_deleted_tombstones
            _delete_inventory_records_by_keys = app.CardPipelineApp._delete_inventory_records_by_keys
            _delete_inventory_photo_files_for_removed_records = lambda self, removed, kept: 0

        with TemporaryDirectory() as tmp:
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_tombstones = app.INVENTORY_DELETED_TOMBSTONES_PATH
            old_pipeline = app.CARD_PIPELINE_DIR
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_DELETED_TOMBSTONES_PATH = Path(tmp) / "inventory_deleted_tombstones.json"
            app.CARD_PIPELINE_DIR = Path(tmp)
            dummy = InventoryDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                first = dummy._normalize_inventory_record({"assigned_person": "Kevin Hambone", "cert_number": "111", "source_sheet": "A.xlsx"})
                second = dummy._normalize_inventory_record({"assigned_person": "James Copeland", "cert_number": "222", "source_sheet": "B.xlsx"})
                third = dummy._normalize_inventory_record({"assigned_person": "Kevin Hambone", "cert_number": "333", "source_sheet": "C.xlsx"})
                dummy._save_inventory_ledger([first, second, third])

                deleted = dummy._delete_inventory_records_by_keys({first["inventory_key"], third["inventory_key"]})

                self.assertEqual(deleted, 2)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(len(ledger), 1)
                self.assertEqual(ledger[0]["inventory_key"], second["inventory_key"])
                self.assertEqual(ledger[0]["cert_number"], "222")
                tombstones = json.loads(app.INVENTORY_DELETED_TOMBSTONES_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual({item["cert_number"] for item in tombstones}, {"111", "333"})
                self.assertEqual({item["source_sheet"] for item in tombstones}, {"A.xlsx", "C.xlsx"})
            finally:
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_DELETED_TOMBSTONES_PATH = old_tombstones
                app.CARD_PIPELINE_DIR = old_pipeline

    def test_delete_person_records_unassigns_markers_inventory_and_profit(self) -> None:
        class DeletePersonDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            delete_person_records = app.CardPipelineApp.delete_person_records

            def _save_sheet_markers(self):
                self.saved_markers = True

        with TemporaryDirectory() as tmp:
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_profit = app.PROFIT_LEDGER_PATH
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            dummy = DeletePersonDummy()
            dummy.home_sheet_markers = {
                "Incoming|A.xlsx": {"assigned_person": "Lucas"},
                "Incoming|B.xlsx": {"assigned_person": "Mikey"},
            }
            dummy.saved_markers = False
            try:
                dummy._save_inventory_ledger([
                    dummy._normalize_inventory_record({"assigned_person": "Lucas", "cert_number": "1", "source_sheet": "A.xlsx"}),
                    dummy._normalize_inventory_record({"assigned_person": "Mikey", "cert_number": "2", "source_sheet": "B.xlsx"}),
                ])
                dummy._save_profit_ledger([
                    dummy._normalize_profit_record({"assigned_person": "Lucas", "cert_number": "1", "source_sheet": "A.xlsx", "company": "Arena", "date_added": "2026-06-17"}),
                    dummy._normalize_profit_record({"assigned_person": "Mikey", "cert_number": "2", "source_sheet": "B.xlsx", "company": "Arena", "date_added": "2026-06-17"}),
                ])

                counts = dummy.delete_person_records("Lucas")

                self.assertEqual(counts, {"markers": 1, "inventory": 1, "profit": 1})
                self.assertTrue(dummy.saved_markers)
                self.assertEqual(dummy.home_sheet_markers["Incoming|A.xlsx"]["assigned_person"], "")
                inventory = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                profit = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                self.assertEqual(inventory[0]["assigned_person"], "Unassigned")
                self.assertEqual(inventory[1]["assigned_person"], "Mikey")
                self.assertEqual(profit[0]["assigned_person"], "")
                self.assertEqual(profit[1]["assigned_person"], "Mikey")
            finally:
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.PROFIT_LEDGER_PATH = old_profit

    def test_home_person_filter_limits_sheet_names(self) -> None:
        class Var:
            def __init__(self, value: str) -> None:
                self.value = value

            def get(self) -> str:
                return self.value

        class HomeDummy:
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _home_person_filter = app.CardPipelineApp._home_person_filter
            _home_sheet_matches_person_filter = app.CardPipelineApp._home_sheet_matches_person_filter
            _filtered_home_sheet_names = app.CardPipelineApp._filtered_home_sheet_names

        dummy = HomeDummy()
        dummy.home_person_var = Var("Kevin")
        dummy.home_sheet_paths = {
            "Incoming": {
                "kevin.xlsx": Path("kevin.xlsx"),
                "james.xlsx": Path("james.xlsx"),
                "blank.xlsx": Path("blank.xlsx"),
            }
        }
        dummy.home_sheet_markers = {
            "Incoming|kevin.xlsx": {"assigned_person": "Kevin Hambone"},
            "Incoming|james.xlsx": {"assigned_person": "James Copeland"},
            "Incoming|blank.xlsx": {},
        }

        self.assertEqual(dummy._filtered_home_sheet_names("Incoming"), ["kevin.xlsx"])

    def test_create_seller_terms_apply_and_restore_purchase_prices(self) -> None:
        class Var:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        class SellerTermsDummy:
            _money_value = app.CardPipelineApp._money_value
            _seller_terms_rate = app.CardPipelineApp._seller_terms_rate
            _load_seller_terms = app.CardPipelineApp._load_seller_terms
            _seller_terms_match = app.CardPipelineApp._seller_terms_match
            _seller_terms_company_decision = app.CardPipelineApp._seller_terms_company_decision
            _seller_terms_company_price = app.CardPipelineApp._seller_terms_company_price
            _restore_create_seller_term_prices = app.CardPipelineApp._restore_create_seller_term_prices
            _network_mode_enabled = app.CardPipelineApp._network_mode_enabled
            apply_create_seller_terms = app.CardPipelineApp.apply_create_seller_terms

            def __init__(self):
                self.create_network_mode_var = Var(True)
                self.seller_terms_seller_var = Var("John")
                self.seller_terms_sheet_type_var = Var("Arena Club")
                self.status_var = Var("")
                self.refreshed = 0
                self.assignment_engine = types.SimpleNamespace(
                    evaluate=lambda row: (
                        [types.SimpleNamespace(company="Arena Club", accepted=True, payout=95.0, source_value=100.0)]
                        if row.cert_number == "1"
                        else []
                    )
                )
                self.intake_rows = [
                    WorkbookRow(excel_row=2, cert_number="1", grader="PSA", card_title="Test", existing_value=10, cy_value=100),
                    WorkbookRow(excel_row=3, cert_number="2", grader="PSA", card_title="No Arena Match", existing_value=20),
                ]

            def _refresh_table(self):
                self.refreshed += 1

        with TemporaryDirectory() as tmp:
            old_terms = app.SELLER_TERMS_PATH
            app.SELLER_TERMS_PATH = Path(tmp) / "seller_terms.csv"
            app.SELLER_TERMS_PATH.write_text(
                "Seller,Sheet Type,Seller Rate,Deduction\n"
                "John,Arena Club,80%,\n"
                "Mary,Arena Club,,5%\n",
                encoding="utf-8",
            )
            dummy = SellerTermsDummy()
            try:
                self.assertEqual(dummy.apply_create_seller_terms(), 1)
                self.assertEqual(dummy.intake_rows[0].existing_value, 80)
                self.assertEqual(dummy.intake_rows[1].existing_value, 20)
                self.assertIn("Arena Club rule value", dummy.status_var.value)

                dummy.seller_terms_seller_var.set("Mary")
                dummy.seller_terms_sheet_type_var.set("Arena Club")
                self.assertEqual(dummy.apply_create_seller_terms(), 1)
                self.assertEqual(dummy.intake_rows[0].existing_value, 90)
                self.assertEqual(dummy.intake_rows[1].existing_value, 20)
                self.assertIn("payout minus 5%", dummy.status_var.value)

                dummy.seller_terms_sheet_type_var.set("")
                self.assertEqual(dummy.apply_create_seller_terms(), 0)
                self.assertEqual(dummy.intake_rows[0].existing_value, 10)
                self.assertEqual(dummy.intake_rows[1].existing_value, 20)

                dummy.seller_terms_sheet_type_var.set("Arena Club")
                dummy.create_network_mode_var.set(False)
                self.assertEqual(dummy.apply_create_seller_terms(), 0)
                self.assertEqual(dummy.intake_rows[0].existing_value, 10)
                self.assertEqual(dummy.intake_rows[1].existing_value, 20)
            finally:
                app.SELLER_TERMS_PATH = old_terms

    def test_inventory_recomp_empty_scope_only_queues_missing_selected_fields(self) -> None:
        class RecompScopeDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_recomp_record_matches_scope = app.CardPipelineApp._inventory_recomp_record_matches_scope

        dummy = RecompScopeDummy()
        full_record = {"card_ladder_value": 50, "card_ladder_comps_average": 40, "cy_value": 30, "cy_confidence": 4}
        missing_comps = {"card_ladder_value": 50, "card_ladder_comps_average": None, "cy_value": 30, "cy_confidence": 4}
        missing_cy = {"card_ladder_value": 50, "card_ladder_comps_average": 40, "cy_value": None, "cy_confidence": ""}
        missing_cy_confidence = {"card_ladder_value": 50, "card_ladder_comps_average": 40, "cy_value": 30, "cy_confidence": ""}

        self.assertTrue(dummy._inventory_recomp_record_matches_scope(full_record, {"scope": app.COMP_SCOPE_ALL, "card_ladder_comps": True}))
        self.assertFalse(dummy._inventory_recomp_record_matches_scope(full_record, {"scope": app.COMP_SCOPE_EMPTY, "card_ladder_value": True, "card_ladder_comps": True}))
        self.assertTrue(dummy._inventory_recomp_record_matches_scope(missing_comps, {"scope": app.COMP_SCOPE_EMPTY, "card_ladder_comps": True}))
        self.assertFalse(dummy._inventory_recomp_record_matches_scope(missing_cy, {"scope": app.COMP_SCOPE_EMPTY, "card_ladder_comps": True}))
        self.assertTrue(dummy._inventory_recomp_record_matches_scope(missing_cy, {"scope": app.COMP_SCOPE_EMPTY, "cy": True}))
        self.assertTrue(dummy._inventory_recomp_record_matches_scope(missing_cy_confidence, {"scope": app.COMP_SCOPE_EMPTY, "cy": True}))

    def test_inventory_recomp_sync_preserves_cy_confidence(self) -> None:
        class RecompSyncDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _sync_inventory_recomp_results = app.CardPipelineApp._sync_inventory_recomp_results

            def _inventory_sport_from_value(self, sport, card_title):
                return sport

            def _enrich_inventory_record_assignment(self, record, force=False):
                return self._normalize_inventory_record(record)

        with TemporaryDirectory() as tmp:
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = RecompSyncDummy()
            original = dummy._normalize_inventory_record(
                {
                    "assigned_person": "Kevin Hambone",
                    "cert_number": "123",
                    "card_title": "Test Card",
                    "source_sheet": "INVENTORY.xlsx",
                    "status": "Active",
                    "cy_value": None,
                    "cy_confidence": "",
                }
            )
            dummy._save_inventory_ledger([original])
            dummy.state = types.SimpleNamespace(
                lock=threading.Lock(),
                rows=[
                    WorkbookRow(
                        excel_row=2,
                        cert_number="123",
                        card_title="Test Card",
                        grader="PSA",
                        cy_value=88.0,
                        cy_confidence=4,
                        status="Ready",
                    )
                ],
            )
            dummy.inventory_recomp_context = {
                "keys_by_excel_row": {2: original["inventory_key"]},
                "features": {"cy": True},
            }
            try:
                self.assertEqual(dummy._sync_inventory_recomp_results(), 1)
                rows = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(rows[0]["cy_value"], 88.0)
                self.assertEqual(rows[0]["cy_confidence"], 4)
            finally:
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_received_inventory_candidate_preserves_cy_confidence(self) -> None:
        class InventoryCandidateDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_sport_from_value = app.CardPipelineApp._inventory_sport_from_value
            _received_inventory_candidate_records_for_sheet = app.CardPipelineApp._received_inventory_candidate_records_for_sheet

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / "POKE_TEST.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Certification Number", "Company", "Sport", "Card Description", "Purchase Price", "Card Ladder Value", "Comps", "CY Estimate", "CY Confidence", "Source", "RECEIVED"])
            sheet.append(["12345678", "PSA", "pokemon", "Pokemon Test PSA 10", 50, None, None, 86, 6, "CY import", "X"])
            workbook.save(path)
            workbook.close()

            records = InventoryCandidateDummy()._received_inventory_candidate_records_for_sheet("Received", path, "Kevin Hambone", company_keys=set())

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["cy_value"], 86.0)
        self.assertEqual(records[0]["cy_confidence"], 6)

    def test_edit_inventory_row_updates_visible_fields_and_rebuilds_key(self) -> None:
        class FakeTree:
            def selection(self):
                return ["row-1"]

        class FakeStatus:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class InventoryEditDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _update_inventory_record_by_key = app.CardPipelineApp._update_inventory_record_by_key
            edit_selected_inventory_row = app.CardPipelineApp.edit_selected_inventory_row
            refresh_inventory_tab = lambda self: None

            def _inventory_edit_row_dialog(self, _record):
                return {
                    "date_added": "2026-06-20",
                    "assigned_person": "James Copeland",
                    "sport": "football",
                    "cert_number": "83861755",
                    "grader": "PSA",
                    "card_title": "2010 Donruss Rated Rookies 95 Tim Tebow PSA 10",
                    "purchase_price": 100,
                    "card_ladder_value": 116,
                    "card_ladder_comps_average": 99.97,
                    "cy_value": None,
                    "cy_confidence": "",
                    "best_company": "FANATICS",
                    "estimated_payout": 107.88,
                    "source_sheet": "JAMES_NASHVILLE_PICKUPS.xlsx",
                }

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = InventoryEditDummy()
            dummy.inventory_tree = FakeTree()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.status_var = FakeStatus()
            record = dummy._normalize_inventory_record({"assigned_person": "Lucas", "cert_number": "1", "source_sheet": "Old.xlsx", "status": "Active"})
            dummy._save_inventory_ledger([record])
            dummy.inventory_tree_records = {"row-1": record}
            try:
                dummy.edit_selected_inventory_row()

                rows = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(len(rows), 1)
                edited = rows[0]
                self.assertEqual(edited["assigned_person"], "James Copeland")
                self.assertEqual(edited["cert_number"], "83861755")
                self.assertEqual(edited["source_sheet"], "JAMES_NASHVILLE_PICKUPS.xlsx")
                self.assertEqual(edited["card_ladder_value"], 116)
                self.assertEqual(edited["card_ladder_comps_average"], 99.97)
                self.assertEqual(edited["best_company"], "FANATICS")
                self.assertEqual(edited["estimated_payout"], 107.88)
                self.assertEqual(edited["inventory_key"], "83861755|james_nashville_pickups.xlsx|james copeland")
                self.assertEqual(dummy.status_var.value, "Edited 1 inventory row(s).")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_inventory_records_can_move_to_company_sheets(self) -> None:
        class FakeTree:
            def selection(self):
                return ["row-1"]

        class FakeAssignment:
            def recommend(self, row, person=""):
                return types.SimpleNamespace(company="Arena Club", payout=90)

        class FakeStatus:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class InventoryMoveDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _mark_inventory_records_moved_to_company = app.CardPipelineApp._mark_inventory_records_moved_to_company
            _inventory_record_can_move_to_company_sheet = app.CardPipelineApp._inventory_record_can_move_to_company_sheet
            _company_sheet_reset_schedules = app.CardPipelineApp._company_sheet_reset_schedules
            _company_sheet_schedule_for_company = app.CardPipelineApp._company_sheet_schedule_for_company
            _company_sheet_week_start_for_company = app.CardPipelineApp._company_sheet_week_start_for_company
            _company_sheet_name_lookup_for_rows = app.CardPipelineApp._company_sheet_name_lookup_for_rows
            move_selected_inventory_to_company_sheets = app.CardPipelineApp.move_selected_inventory_to_company_sheets
            _move_inventory_records_to_company_sheets = app.CardPipelineApp._move_inventory_records_to_company_sheets
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            record_profit_sales = app.CardPipelineApp.record_profit_sales
            refresh_inventory_tab = lambda self: None
            refresh_profit_tab = lambda self: None
            _append_activity = lambda self, action, summary, details=None: None

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_company = app.COMPANY_SHEETS_DIR
            old_profit = app.PROFIT_LEDGER_PATH
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.CARD_PIPELINE_DIR = root
            app.COMPANY_SHEETS_DIR = root / "COMPANY SHEETS"
            app.PROFIT_LEDGER_PATH = root / "profit_ledger.json"
            app.INVENTORY_LEDGER_PATH = root / "inventory_ledger.json"
            dummy = InventoryMoveDummy()
            dummy.assignment_engine = FakeAssignment()
            dummy.inventory_tree = FakeTree()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.status_var = FakeStatus()
            record = dummy._normalize_inventory_record(
                {
                    "assigned_person": "Lucas",
                    "cert_number": "123",
                    "grader": "PSA",
                    "card_title": "Test Card",
                    "source_sheet": "Quick Load.xlsx",
                    "purchase_price": 40,
                    "inventory_value": 100,
                    "best_company": "Arena Club",
                    "estimated_payout": 90,
                    "status": "Active",
                }
            )
            dummy._save_inventory_ledger([record])
            dummy.inventory_tree_records = {"row-1": record}
            try:
                with patch("app.messagebox.askyesno", return_value=True):
                    dummy.move_selected_inventory_to_company_sheets()

                company_rows = read_company_profit_records(app.COMPANY_SHEETS_DIR)
                profit_rows = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                inventory = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(len(company_rows), 1)
                self.assertEqual(company_rows[0]["company"], "Arena Club")
                self.assertEqual(len(profit_rows), 1)
                self.assertEqual(profit_rows[0]["assigned_person"], "Lucas")
                self.assertEqual(inventory, [])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.COMPANY_SHEETS_DIR = old_company
                app.PROFIT_LEDGER_PATH = old_profit
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_inventory_move_uses_stored_best_company_when_recommendation_goes_stale(self) -> None:
        class FakeAssignment:
            def recommend(self, row, person=""):
                return types.SimpleNamespace(company=app.NO_COMPANY_TAKES_LABEL, payout=None)

        class FakeStatus:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        class InventoryMoveDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _mark_inventory_records_moved_to_company = app.CardPipelineApp._mark_inventory_records_moved_to_company
            _company_sheet_reset_schedules = app.CardPipelineApp._company_sheet_reset_schedules
            _company_sheet_schedule_for_company = app.CardPipelineApp._company_sheet_schedule_for_company
            _company_sheet_week_start_for_company = app.CardPipelineApp._company_sheet_week_start_for_company
            _company_sheet_name_lookup_for_rows = app.CardPipelineApp._company_sheet_name_lookup_for_rows
            _move_inventory_records_to_company_sheets = app.CardPipelineApp._move_inventory_records_to_company_sheets
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            record_profit_sales = app.CardPipelineApp.record_profit_sales
            refresh_inventory_tab = lambda self: None
            refresh_profit_tab = lambda self: None
            _append_activity = lambda self, action, summary, details=None: None

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_company = app.COMPANY_SHEETS_DIR
            old_profit = app.PROFIT_LEDGER_PATH
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.CARD_PIPELINE_DIR = root
            app.COMPANY_SHEETS_DIR = root / "COMPANY SHEETS"
            app.PROFIT_LEDGER_PATH = root / "profit_ledger.json"
            app.INVENTORY_LEDGER_PATH = root / "inventory_ledger.json"
            dummy = InventoryMoveDummy()
            dummy.assignment_engine = FakeAssignment()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.status_var = FakeStatus()
            record = dummy._normalize_inventory_record(
                {
                    "assigned_person": "Lucas",
                    "cert_number": "141119049",
                    "grader": "PSA",
                    "card_title": "2025 Panini Prizm 325 Justin Herbert PSA 8",
                    "source_sheet": "SCOTSBORO_HAMBONE_6_16_26.xlsx",
                    "purchase_price": 10,
                    "card_ladder_value": 14,
                    "card_ladder_comps_average": 14,
                    "best_company": "FANATICS",
                    "estimated_payout": 13.30,
                    "status": "Active",
                }
            )
            dummy._save_inventory_ledger([record])
            try:
                dummy._move_inventory_records_to_company_sheets([record])

                company_rows = read_company_profit_records(app.COMPANY_SHEETS_DIR)
                profit_rows = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                inventory = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(len(company_rows), 1)
                self.assertEqual(company_rows[0]["company"], "FANATICS")
                self.assertEqual(company_rows[0]["sale_price"], 13.30)
                self.assertEqual(len(profit_rows), 1)
                self.assertEqual(inventory, [])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.COMPANY_SHEETS_DIR = old_company
                app.PROFIT_LEDGER_PATH = old_profit
                app.INVENTORY_LEDGER_PATH = old_inventory


    def test_inventory_company_move_preserves_photo_paths_for_refund(self) -> None:
        class InventoryMoveDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _company_sheet_name_lookup_for_rows = lambda self, rows: {}
            _mark_inventory_records_moved_to_company = app.CardPipelineApp._mark_inventory_records_moved_to_company
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            record_profit_sales = app.CardPipelineApp.record_profit_sales
            _append_profit_records = app.CardPipelineApp._append_profit_records
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _delete_inventory_photo_files_for_removed_records = lambda self, removed, kept=None: 0
            _move_inventory_records_to_company_sheets = app.CardPipelineApp._move_inventory_records_to_company_sheets
            _append_activity = lambda self, action, summary, details=None: None
            refresh_inventory_tab = lambda self: None
            refresh_profit_tab = lambda self: None

            def __init__(self):
                self.lucas_identity = {"display_name": "Tester", "machine": "Test"}
                self.status_var = types.SimpleNamespace(set=lambda _value: None)
                self.assignment_engine = types.SimpleNamespace()

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            old_pipeline = app.CARD_PIPELINE_DIR
            old_company = app.COMPANY_SHEETS_DIR
            old_profit = app.PROFIT_LEDGER_PATH
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.CARD_PIPELINE_DIR = root
            app.COMPANY_SHEETS_DIR = root / "COMPANY SHEETS"
            app.PROFIT_LEDGER_PATH = root / "profit_ledger.json"
            app.INVENTORY_LEDGER_PATH = root / "inventory_ledger.json"
            dummy = InventoryMoveDummy()
            record = dummy._normalize_inventory_record(
                {
                    "assigned_person": "James Copeland",
                    "cert_number": "4151253025",
                    "grader": "CGC",
                    "card_title": "1997 Pokemon Jungle Japanese Mr. Mime Holo CGC 7.5",
                    "source_sheet": "Kevin_6_16_2026.xlsx",
                    "purchase_price": 12,
                    "card_ladder_value": 18,
                    "card_ladder_comps_average": 16,
                    "best_company": "FANATICS",
                    "estimated_payout": 17.1,
                    "status": "Active",
                    "photo_paths": ["mr-mime-front.jpg"],
                }
            )
            dummy._save_inventory_ledger([record])
            try:
                dummy._move_inventory_records_to_company_sheets([record])
                profit = [dummy._normalize_profit_record(item) for item in dummy._load_profit_ledger()]
                self.assertEqual(profit[0]["photo_paths"], ["mr-mime-front.jpg"])
                self.assertEqual(profit[0]["photo_count"], 1)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.COMPANY_SHEETS_DIR = old_company
                app.PROFIT_LEDGER_PATH = old_profit
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_nobody_takes_inventory_record_cannot_move_to_company_sheet(self) -> None:
        class InventoryDummy:
            _inventory_record_can_move_to_company_sheet = app.CardPipelineApp._inventory_record_can_move_to_company_sheet

        dummy = InventoryDummy()
        self.assertFalse(
            dummy._inventory_record_can_move_to_company_sheet(
                {"status": "Active", "best_company": app.NO_COMPANY_TAKES_LABEL, "estimated_payout": None}
            )
        )
        self.assertTrue(
            dummy._inventory_record_can_move_to_company_sheet(
                {"status": "Active", "best_company": "Arena Club", "estimated_payout": 95}
            )
        )
        self.assertFalse(
            dummy._inventory_record_can_move_to_company_sheet(
                {"item_type": "Raw", "status": "Active", "best_company": "Arena Club", "estimated_payout": 95}
            )
        )

    def test_mark_inventory_record_sold_writes_profit_and_removes_inventory(self) -> None:
        class SoldDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _safe_inventory_photo_path = app.CardPipelineApp._safe_inventory_photo_path
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _delete_inventory_photo_files_for_removed_records = app.CardPipelineApp._delete_inventory_photo_files_for_removed_records
            _mark_inventory_record_sold = app.CardPipelineApp._mark_inventory_record_sold
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _inventory_sale_profit_record = app.CardPipelineApp._inventory_sale_profit_record
            _inventory_sale_expense_record = app.CardPipelineApp._inventory_sale_expense_record
            _general_sold_sheet_name = app.CardPipelineApp._general_sold_sheet_name
            mark_inventory_record_sold = app.CardPipelineApp.mark_inventory_record_sold
            record_profit_sales = app.CardPipelineApp.record_profit_sales
            refresh_profit_tab = lambda self: None
            _append_activity = lambda self, action, summary, details=None: None

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_profit = app.PROFIT_LEDGER_PATH
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            dummy = SoldDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            record = dummy._normalize_inventory_record(
                {
                    "assigned_person": "Hambone",
                    "cert_number": "123",
                    "grader": "PSA",
                    "card_title": "Test Card",
                    "source_sheet": "Lot.xlsx",
                    "purchase_price": 40,
                    "inventory_value": 100,
                    "best_company": "Arena Club",
                    "estimated_payout": 90,
                    "status": "Active",
                }
            )
            dummy._save_inventory_ledger([record])
            try:
                self.assertTrue(dummy.mark_inventory_record_sold(record, "Arena Club", 95, expense_type="Shipping", expense_amount=12.5, expense_notes="label"))
                inventory = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                profit = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                self.assertEqual(inventory, [])
                self.assertEqual(len(profit), 2)
                sale_row = next(record for record in profit if record.get("record_type") != "expense")
                expense_row = next(record for record in profit if record.get("record_type") == "expense")
                self.assertEqual(sale_row["company"], "Arena Club")
                self.assertEqual(sale_row["assigned_person"], "Hambone")
                self.assertEqual(sale_row["sale_price"], 95.0)
                self.assertEqual(sale_row["profit"], 55.0)
                self.assertEqual(expense_row["company"], "Expense: Shipping")
                self.assertEqual(expense_row["assigned_person"], "Hambone")
                self.assertEqual(expense_row["related_type"], "Card")
                self.assertEqual(expense_row["source_sheet"], "Lot.xlsx")
                self.assertEqual(expense_row["cert_number"], "123")
                self.assertEqual(expense_row["expense_amount"], 12.5)
                self.assertEqual(expense_row["profit"], -12.5)
                self.assertEqual(expense_row["notes"], "label")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.PROFIT_LEDGER_PATH = old_profit
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_photo_scan_links_matching_cert(self) -> None:
        class PhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_files = app.CardPipelineApp._inventory_photo_files
            _inventory_photo_certs_from_cards = app.CardPipelineApp._inventory_photo_certs_from_cards
            _active_inventory_keys_by_cert = app.CardPipelineApp._active_inventory_keys_by_cert
            _link_inventory_photo_to_keys = app.CardPipelineApp._link_inventory_photo_to_keys
            _inventory_photo_match_keys = app.CardPipelineApp._inventory_photo_match_keys
            _inventory_photo_best_title_match = app.CardPipelineApp._inventory_photo_best_title_match
            _inventory_photo_card_match_text = app.CardPipelineApp._inventory_photo_card_match_text
            _match_text_tokens = app.CardPipelineApp._match_text_tokens
            _inventory_photo_filename_hint = app.CardPipelineApp._inventory_photo_filename_hint
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_capture_group_key = app.CardPipelineApp._inventory_photo_capture_group_key
            _inventory_photo_base64 = lambda self, path: "stub"
            _inventory_photo_scan_group_nearby_unmatched = app.CardPipelineApp._inventory_photo_scan_group_nearby_unmatched
            _inventory_photo_scan_worker = app.CardPipelineApp._inventory_photo_scan_worker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            photo = app.INVENTORY_PHOTOS_DIR / "card.jpg"
            photo.write_bytes(b"fake image")
            dummy = PhotoDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            dummy.inventory_photo_client = object()
            dummy.events = __import__("queue").Queue()
            record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "12345678", "card_title": "Test", "status": "Active"})
            dummy._save_inventory_ledger([record])
            try:
                with patch.object(app, "identify_cards_sync", return_value=[{"cert_number": "12345678"}]):
                    dummy._inventory_photo_scan_worker(app.INVENTORY_PHOTOS_DIR)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(ledger[0]["photo_paths"], [str(photo)])
                state = json.loads(app.INVENTORY_PHOTO_STATE_PATH.read_text(encoding="utf-8"))
                state_record = next(iter(state["photos"].values()))
                self.assertEqual(state_record["certs"], ["12345678"])
                self.assertEqual(state_record["status"], "linked")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_photo_scan_does_not_group_by_time_only(self) -> None:
        class PhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_files = app.CardPipelineApp._inventory_photo_files
            _inventory_photo_certs_from_cards = app.CardPipelineApp._inventory_photo_certs_from_cards
            _active_inventory_keys_by_cert = app.CardPipelineApp._active_inventory_keys_by_cert
            _link_inventory_photo_to_keys = app.CardPipelineApp._link_inventory_photo_to_keys
            _inventory_photo_match_keys = app.CardPipelineApp._inventory_photo_match_keys
            _inventory_photo_best_title_match = app.CardPipelineApp._inventory_photo_best_title_match
            _inventory_photo_card_match_text = app.CardPipelineApp._inventory_photo_card_match_text
            _match_text_tokens = app.CardPipelineApp._match_text_tokens
            _inventory_photo_filename_hint = app.CardPipelineApp._inventory_photo_filename_hint
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_capture_group_key = app.CardPipelineApp._inventory_photo_capture_group_key
            _inventory_photo_base64 = lambda self, path: "stub"
            _inventory_photo_scan_group_nearby_unmatched = app.CardPipelineApp._inventory_photo_scan_group_nearby_unmatched
            _inventory_photo_scan_worker = app.CardPipelineApp._inventory_photo_scan_worker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            front = app.INVENTORY_PHOTOS_DIR / "1-front.jpg"
            back = app.INVENTORY_PHOTOS_DIR / "2-back.jpg"
            front.write_bytes(b"front image")
            back.write_bytes(b"back image")
            now = time.time()
            os.utime(front, (now, now))
            os.utime(back, (now + 10, now + 10))
            dummy = PhotoDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            dummy.inventory_photo_client = object()
            dummy.events = queue.Queue()
            record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "12345678", "card_title": "Test", "status": "Active"})
            dummy._save_inventory_ledger([record])
            try:
                with patch.object(app, "identify_cards_sync", side_effect=[[{"cert_number": "12345678"}], [{}]]):
                    dummy._inventory_photo_scan_worker(app.INVENTORY_PHOTOS_DIR)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(ledger[0]["photo_paths"], [str(front)])
                state = json.loads(app.INVENTORY_PHOTO_STATE_PATH.read_text(encoding="utf-8"))
                grouped = [photo for photo in state["photos"].values() if photo.get("auto_grouped_from")]
                self.assertEqual(grouped, [])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_photo_scan_groups_matching_filename_prefix_even_when_not_nearby(self) -> None:
        class PhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_files = app.CardPipelineApp._inventory_photo_files
            _inventory_photo_certs_from_cards = app.CardPipelineApp._inventory_photo_certs_from_cards
            _active_inventory_keys_by_cert = app.CardPipelineApp._active_inventory_keys_by_cert
            _link_inventory_photo_to_keys = app.CardPipelineApp._link_inventory_photo_to_keys
            _inventory_photo_match_keys = app.CardPipelineApp._inventory_photo_match_keys
            _inventory_photo_best_title_match = app.CardPipelineApp._inventory_photo_best_title_match
            _inventory_photo_card_match_text = app.CardPipelineApp._inventory_photo_card_match_text
            _match_text_tokens = app.CardPipelineApp._match_text_tokens
            _inventory_photo_filename_hint = app.CardPipelineApp._inventory_photo_filename_hint
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_capture_group_key = app.CardPipelineApp._inventory_photo_capture_group_key
            _inventory_photo_base64 = lambda self, path: "stub"
            _inventory_photo_scan_group_nearby_unmatched = app.CardPipelineApp._inventory_photo_scan_group_nearby_unmatched
            _inventory_photo_scan_worker = app.CardPipelineApp._inventory_photo_scan_worker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            front = app.INVENTORY_PHOTOS_DIR / "20260706-Card001-01.jpg"
            back = app.INVENTORY_PHOTOS_DIR / "20260706-Card001-02.jpg"
            other = app.INVENTORY_PHOTOS_DIR / "20260706-Card002-01.jpg"
            front.write_bytes(b"front image")
            back.write_bytes(b"back image")
            other.write_bytes(b"other image")
            now = time.time()
            os.utime(front, (now, now))
            os.utime(back, (now + 600, now + 600))
            os.utime(other, (now + 601, now + 601))
            dummy = PhotoDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            dummy.inventory_photo_client = object()
            dummy.events = queue.Queue()
            record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "12345678", "card_title": "Test", "status": "Active"})
            dummy._save_inventory_ledger([record])
            try:
                with patch.object(app, "identify_cards_sync", side_effect=[[{"cert_number": "12345678"}], [{}], [{}]]):
                    dummy._inventory_photo_scan_worker(app.INVENTORY_PHOTOS_DIR)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(ledger[0]["photo_paths"], [str(front), str(back)])
                self.assertNotIn(str(other), ledger[0]["photo_paths"])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_photo_scan_groups_shortcuts_bracketed_filename_prefix(self) -> None:
        class PhotoDummy:
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_capture_group_key = app.CardPipelineApp._inventory_photo_capture_group_key

        dummy = PhotoDummy()
        front = {"filename": "[20260706-1226]-Card[1]-[1].jpg"}
        back = {"filename": "[20260706-1226]-Card[1]-[2].jpg"}
        next_card = {"filename": "[20260706-1226]-Card[2]-[1].jpg"}
        titled_front = {"filename": "[20260706-1226]-Card[1]-[1]-Kobe Purple Wave.jpg"}
        titled_back = {"filename": "[20260706-1226]-Card[1]-[2]-Kobe Purple Wave.jpg"}

        self.assertEqual(
            dummy._inventory_photo_capture_group_key(front),
            dummy._inventory_photo_capture_group_key(back),
        )
        self.assertEqual(
            dummy._inventory_photo_capture_group_key(titled_front),
            dummy._inventory_photo_capture_group_key(titled_back),
        )
        self.assertEqual(
            dummy._inventory_photo_capture_group_key(front),
            dummy._inventory_photo_capture_group_key(titled_front),
        )
        self.assertNotEqual(
            dummy._inventory_photo_capture_group_key(front),
            dummy._inventory_photo_capture_group_key(next_card),
        )

    def test_inventory_photo_scan_saves_progress_after_each_photo(self) -> None:
        class PhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_files = app.CardPipelineApp._inventory_photo_files
            _inventory_photo_certs_from_cards = app.CardPipelineApp._inventory_photo_certs_from_cards
            _active_inventory_keys_by_cert = app.CardPipelineApp._active_inventory_keys_by_cert
            _link_inventory_photo_to_keys = app.CardPipelineApp._link_inventory_photo_to_keys
            _inventory_photo_match_keys = app.CardPipelineApp._inventory_photo_match_keys
            _inventory_photo_best_title_match = app.CardPipelineApp._inventory_photo_best_title_match
            _inventory_photo_card_match_text = app.CardPipelineApp._inventory_photo_card_match_text
            _match_text_tokens = app.CardPipelineApp._match_text_tokens
            _inventory_photo_filename_hint = app.CardPipelineApp._inventory_photo_filename_hint
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_capture_group_key = app.CardPipelineApp._inventory_photo_capture_group_key
            _inventory_photo_base64 = lambda self, path: "stub"
            _inventory_photo_scan_group_nearby_unmatched = app.CardPipelineApp._inventory_photo_scan_group_nearby_unmatched
            _inventory_photo_scan_worker = app.CardPipelineApp._inventory_photo_scan_worker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            first = app.INVENTORY_PHOTOS_DIR / "a.jpg"
            second = app.INVENTORY_PHOTOS_DIR / "b.jpg"
            first.write_bytes(b"fake image one")
            second.write_bytes(b"fake image two")
            dummy = PhotoDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            dummy.inventory_photo_client = object()
            dummy.events = __import__("queue").Queue()
            dummy._save_inventory_ledger([])
            save_counts: list[int] = []
            original_save = dummy._save_inventory_photo_state

            def tracking_save(state: dict[str, object]) -> None:
                photos = state.get("photos")
                save_counts.append(len(photos) if isinstance(photos, dict) else 0)
                original_save(state)

            dummy._save_inventory_photo_state = tracking_save
            try:
                with patch.object(app, "identify_cards_sync", return_value=[{"cert_number": "12345678"}]):
                    dummy._inventory_photo_scan_worker(app.INVENTORY_PHOTOS_DIR)
                self.assertIn(1, save_counts)
                self.assertGreaterEqual(save_counts.count(2), 1)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_photo_scan_retries_previous_no_match_but_skips_active_link(self) -> None:
        class PhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_files = app.CardPipelineApp._inventory_photo_files
            _inventory_photo_certs_from_cards = app.CardPipelineApp._inventory_photo_certs_from_cards
            _active_inventory_keys_by_cert = app.CardPipelineApp._active_inventory_keys_by_cert
            _link_inventory_photo_to_keys = app.CardPipelineApp._link_inventory_photo_to_keys
            _inventory_photo_match_keys = app.CardPipelineApp._inventory_photo_match_keys
            _inventory_photo_best_title_match = app.CardPipelineApp._inventory_photo_best_title_match
            _inventory_photo_card_match_text = app.CardPipelineApp._inventory_photo_card_match_text
            _match_text_tokens = app.CardPipelineApp._match_text_tokens
            _inventory_photo_filename_hint = app.CardPipelineApp._inventory_photo_filename_hint
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_capture_group_key = app.CardPipelineApp._inventory_photo_capture_group_key
            _inventory_photo_shared_folder = app.CardPipelineApp._inventory_photo_shared_folder
            _inventory_photo_relative_path = app.CardPipelineApp._inventory_photo_relative_path
            _inventory_photo_storage_value = app.CardPipelineApp._inventory_photo_storage_value
            _inventory_photo_path_candidates = app.CardPipelineApp._inventory_photo_path_candidates
            _inventory_photo_used_path_keys = app.CardPipelineApp._inventory_photo_used_path_keys
            _inventory_photo_used_hashes = app.CardPipelineApp._inventory_photo_used_hashes
            _inventory_record_references_photo = app.CardPipelineApp._inventory_record_references_photo
            _inventory_photo_scan_can_skip = app.CardPipelineApp._inventory_photo_scan_can_skip
            _sold_inventory_cert_numbers = app.CardPipelineApp._sold_inventory_cert_numbers
            _sold_inventory_photo_used_keys = app.CardPipelineApp._sold_inventory_photo_used_keys
            _inventory_photo_image_matches_sold_photo = app.CardPipelineApp._inventory_photo_image_matches_sold_photo
            _inventory_photo_state_matches_sold_cert = app.CardPipelineApp._inventory_photo_state_matches_sold_cert
            _inventory_photo_base64 = lambda self, path: "stub"
            _inventory_photo_scan_group_nearby_unmatched = app.CardPipelineApp._inventory_photo_scan_group_nearby_unmatched
            _inventory_photo_scan_worker = app.CardPipelineApp._inventory_photo_scan_worker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_profit = app.PROFIT_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            retry_photo = app.INVENTORY_PHOTOS_DIR / "retry.jpg"
            linked_photo = app.INVENTORY_PHOTOS_DIR / "linked.jpg"
            stale_linked_photo = app.INVENTORY_PHOTOS_DIR / "stale-linked.jpg"
            sold_photo = app.INVENTORY_PHOTOS_DIR / "sold.jpg"
            sold_path_photo = app.INVENTORY_PHOTOS_DIR / "sold-path.jpg"
            attached_stale_photo = app.INVENTORY_PHOTOS_DIR / "attached-stale.jpg"
            retry_photo.write_bytes(b"retry image")
            linked_photo.write_bytes(b"linked image")
            stale_linked_photo.write_bytes(b"stale linked image")
            sold_photo.write_bytes(b"sold image")
            sold_path_photo.write_bytes(b"sold path image")
            attached_stale_photo.write_bytes(b"attached stale image")
            dummy = PhotoDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            dummy.inventory_photo_client = object()
            dummy.events = __import__("queue").Queue()
            retry_record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "12345678", "card_title": "Retry Card", "status": "Active"})
            linked_record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "87654321", "card_title": "Linked Card", "status": "Active", "photo_paths": [str(linked_photo)]})
            attached_stale_record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "22222222", "card_title": "Already Attached", "status": "Active", "photo_paths": [str(attached_stale_photo)]})
            dummy._save_inventory_ledger([retry_record, linked_record, attached_stale_record])
            dummy._save_profit_ledger([
                {"cert_number": "55555555", "card_title": "Sold Card", "sale_price": 20},
                {"cert_number": "44444444", "card_title": "Sold Path Card", "sale_price": 25, "photo_paths": [str(sold_path_photo)]},
            ])
            retry_sha = dummy._inventory_photo_file_hash(retry_photo)
            linked_sha = dummy._inventory_photo_file_hash(linked_photo)
            stale_linked_sha = dummy._inventory_photo_file_hash(stale_linked_photo)
            sold_sha = dummy._inventory_photo_file_hash(sold_photo)
            sold_path_sha = dummy._inventory_photo_file_hash(sold_path_photo)
            attached_stale_sha = dummy._inventory_photo_file_hash(attached_stale_photo)
            dummy._save_inventory_photo_state(
                {
                    "version": 1,
                    "photos": {
                        retry_sha: {"filename": retry_photo.name, "relative_path": retry_photo.name, "cards": [{"cert_number": "99999999"}], "certs": ["99999999"], "linked_keys": [], "status": "no_matching_inventory"},
                        linked_sha: {"filename": linked_photo.name, "relative_path": linked_photo.name, "cards": [{"cert_number": "87654321"}], "certs": ["87654321"], "linked_keys": [linked_record["inventory_key"]], "status": "linked"},
                        stale_linked_sha: {"filename": stale_linked_photo.name, "relative_path": stale_linked_photo.name, "cards": [{"cert_number": "11111111"}], "certs": ["11111111"], "linked_keys": ["old-key"], "status": "missing_from_album"},
                        sold_sha: {"filename": sold_photo.name, "relative_path": sold_photo.name, "cards": [{"cert_number": "55555555"}], "certs": ["55555555"], "linked_keys": [], "status": "no_matching_inventory"},
                        attached_stale_sha: {"filename": attached_stale_photo.name, "relative_path": attached_stale_photo.name, "cards": [{"cert_number": "22222222"}], "certs": ["22222222"], "linked_keys": [], "status": "no_matching_inventory"},
                    },
                }
            )
            try:
                with patch.object(app, "identify_cards_sync", return_value=[{"cert_number": "12345678"}]) as identify:
                    dummy._inventory_photo_scan_worker(app.INVENTORY_PHOTOS_DIR)
                self.assertEqual(identify.call_count, 1)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                by_cert = {row["cert_number"]: row for row in ledger}
                self.assertEqual(by_cert["12345678"]["photo_paths"], [retry_photo.name])
                self.assertEqual(by_cert["87654321"]["photo_paths"], [str(linked_photo)])
                state = json.loads(app.INVENTORY_PHOTO_STATE_PATH.read_text(encoding="utf-8"))
                self.assertEqual(state["photos"][retry_sha]["status"], "linked")
                self.assertEqual(state["photos"][retry_sha]["certs"], ["12345678"])
                self.assertEqual(state["photos"][stale_linked_sha]["status"], "missing_from_album")
                self.assertIn("last_seen", state["photos"][stale_linked_sha])
                self.assertEqual(state["photos"][sold_sha]["status"], "sold_inventory")
                self.assertIn("last_seen", state["photos"][sold_sha])
                self.assertEqual(state["photos"][sold_path_sha]["status"], "sold_inventory")
                self.assertIn("last_seen", state["photos"][sold_path_sha])
                self.assertEqual(state["photos"][attached_stale_sha]["status"], "linked")
                self.assertEqual(state["photos"][attached_stale_sha]["linked_keys"], [attached_stale_record["inventory_key"]])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.PROFIT_LEDGER_PATH = old_profit
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_photo_scan_links_by_card_details_when_cert_missing(self) -> None:
        class PhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_files = app.CardPipelineApp._inventory_photo_files
            _inventory_photo_certs_from_cards = app.CardPipelineApp._inventory_photo_certs_from_cards
            _active_inventory_keys_by_cert = app.CardPipelineApp._active_inventory_keys_by_cert
            _link_inventory_photo_to_keys = app.CardPipelineApp._link_inventory_photo_to_keys
            _inventory_photo_match_keys = app.CardPipelineApp._inventory_photo_match_keys
            _inventory_photo_best_title_match = app.CardPipelineApp._inventory_photo_best_title_match
            _inventory_photo_card_match_text = app.CardPipelineApp._inventory_photo_card_match_text
            _match_text_tokens = app.CardPipelineApp._match_text_tokens
            _inventory_photo_filename_hint = app.CardPipelineApp._inventory_photo_filename_hint
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_capture_group_key = app.CardPipelineApp._inventory_photo_capture_group_key
            _inventory_photo_base64 = lambda self, path: "stub"
            _inventory_photo_scan_group_nearby_unmatched = app.CardPipelineApp._inventory_photo_scan_group_nearby_unmatched
            _inventory_photo_scan_worker = app.CardPipelineApp._inventory_photo_scan_worker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            photo = app.INVENTORY_PHOTOS_DIR / "card.jpg"
            photo.write_bytes(b"fake image")
            dummy = PhotoDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            dummy.inventory_photo_client = object()
            dummy.events = __import__("queue").Queue()
            record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "11111111", "card_title": "2024 Panini Prizm Test Player Silver PSA 10", "status": "Active"})
            dummy._save_inventory_ledger([record])
            ocr_card = {
                "grading_company": "PSA",
                "player": "Test Player",
                "year": "2024",
                "set": "Panini Prizm",
                "parallel": "Silver",
                "grade": "10",
            }
            try:
                with patch.object(app, "identify_cards_sync", return_value=[ocr_card]):
                    dummy._inventory_photo_scan_worker(app.INVENTORY_PHOTOS_DIR)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(ledger[0]["photo_paths"], [str(photo)])
                state = json.loads(app.INVENTORY_PHOTO_STATE_PATH.read_text(encoding="utf-8"))
                state_record = next(iter(state["photos"].values()))
                self.assertEqual(state_record["cards"], [ocr_card])
                self.assertEqual(state_record["status"], "linked")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_photo_scan_links_raw_card_by_filename_hint(self) -> None:
        class PhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_files = app.CardPipelineApp._inventory_photo_files
            _inventory_photo_certs_from_cards = app.CardPipelineApp._inventory_photo_certs_from_cards
            _active_inventory_keys_by_cert = app.CardPipelineApp._active_inventory_keys_by_cert
            _link_inventory_photo_to_keys = app.CardPipelineApp._link_inventory_photo_to_keys
            _inventory_photo_match_keys = app.CardPipelineApp._inventory_photo_match_keys
            _inventory_photo_best_title_match = app.CardPipelineApp._inventory_photo_best_title_match
            _inventory_photo_card_match_text = app.CardPipelineApp._inventory_photo_card_match_text
            _match_text_tokens = app.CardPipelineApp._match_text_tokens
            _inventory_photo_filename_hint = app.CardPipelineApp._inventory_photo_filename_hint
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_capture_group_key = app.CardPipelineApp._inventory_photo_capture_group_key
            _inventory_photo_base64 = lambda self, path: "stub"
            _inventory_photo_scan_group_nearby_unmatched = app.CardPipelineApp._inventory_photo_scan_group_nearby_unmatched
            _inventory_photo_scan_worker = app.CardPipelineApp._inventory_photo_scan_worker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            photo = app.INVENTORY_PHOTOS_DIR / "[20260706-1226]-Card[1]-[1]-stockton immaculate laundry tag 3.jpg"
            photo.write_bytes(b"fake image")
            dummy = PhotoDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            dummy.inventory_photo_client = object()
            dummy.events = __import__("queue").Queue()
            record = dummy._normalize_inventory_record(
                {
                    "assigned_person": "Kevin",
                    "item_id": "RAW-STOCKTON-1",
                    "cert_number": "",
                    "card_title": "John Stockton Panini Immaculate Laundry Tag /3",
                    "status": "Active",
                }
            )
            dummy._save_inventory_ledger([record])
            try:
                with patch.object(app, "identify_cards_sync", return_value=[]):
                    dummy._inventory_photo_scan_worker(app.INVENTORY_PHOTOS_DIR)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(ledger[0]["photo_paths"], [str(photo)])
                state = json.loads(app.INVENTORY_PHOTO_STATE_PATH.read_text(encoding="utf-8"))
                state_record = next(iter(state["photos"].values()))
                self.assertEqual(state_record["filename_hint"], "stockton immaculate laundry tag 3")
                self.assertEqual(state_record["linked_keys"], [record["inventory_key"]])
                self.assertEqual(state_record["status"], "linked")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_photo_scan_does_not_title_match_when_cert_mismatches(self) -> None:
        class PhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_files = app.CardPipelineApp._inventory_photo_files
            _inventory_photo_certs_from_cards = app.CardPipelineApp._inventory_photo_certs_from_cards
            _active_inventory_keys_by_cert = app.CardPipelineApp._active_inventory_keys_by_cert
            _link_inventory_photo_to_keys = app.CardPipelineApp._link_inventory_photo_to_keys
            _inventory_photo_match_keys = app.CardPipelineApp._inventory_photo_match_keys
            _inventory_photo_best_title_match = app.CardPipelineApp._inventory_photo_best_title_match
            _inventory_photo_card_match_text = app.CardPipelineApp._inventory_photo_card_match_text
            _match_text_tokens = app.CardPipelineApp._match_text_tokens
            _inventory_photo_filename_hint = app.CardPipelineApp._inventory_photo_filename_hint
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_base64 = lambda self, path: "stub"
            _inventory_photo_scan_group_nearby_unmatched = app.CardPipelineApp._inventory_photo_scan_group_nearby_unmatched
            _inventory_photo_scan_worker = app.CardPipelineApp._inventory_photo_scan_worker

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            photo = app.INVENTORY_PHOTOS_DIR / "card.jpg"
            photo.write_bytes(b"fake image")
            dummy = PhotoDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            dummy.inventory_photo_client = object()
            dummy.events = __import__("queue").Queue()
            record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "161034883", "card_title": "1986 Fleer Michael Jordan PSA 4.5", "status": "Active"})
            dummy._save_inventory_ledger([record])
            ocr_card = {
                "cert_number": "06052050",
                "grading_company": "PSA",
                "player": "Michael Jordan",
                "year": "1986",
                "set": "Fleer",
                "card_number": "#57",
                "grade": "8",
            }
            try:
                with patch.object(app, "identify_cards_sync", return_value=[ocr_card]):
                    dummy._inventory_photo_scan_worker(app.INVENTORY_PHOTOS_DIR)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(ledger[0]["photo_paths"], [])
                state = json.loads(app.INVENTORY_PHOTO_STATE_PATH.read_text(encoding="utf-8"))
                state_record = next(iter(state["photos"].values()))
                self.assertEqual(state_record["cards"], [ocr_card])
                self.assertEqual(state_record["certs"], ["06052050"])
                self.assertEqual(state_record["linked_keys"], [])
                self.assertEqual(state_record["status"], "no_matching_inventory")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_photo_scan_rescues_single_bgs_cert_from_full_photo(self) -> None:
        class PhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_shared_folder = app.CardPipelineApp._inventory_photo_shared_folder
            _inventory_photo_relative_path = app.CardPipelineApp._inventory_photo_relative_path
            _inventory_photo_storage_value = app.CardPipelineApp._inventory_photo_storage_value
            _inventory_photo_path_candidates = app.CardPipelineApp._inventory_photo_path_candidates
            _inventory_record_references_photo = app.CardPipelineApp._inventory_record_references_photo
            _inventory_photo_scan_can_skip = app.CardPipelineApp._inventory_photo_scan_can_skip
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_files = app.CardPipelineApp._inventory_photo_files
            _inventory_photo_certs_from_cards = app.CardPipelineApp._inventory_photo_certs_from_cards
            _inventory_photo_rescue_single_bgs_cert = app.CardPipelineApp._inventory_photo_rescue_single_bgs_cert
            _active_inventory_keys_by_cert = app.CardPipelineApp._active_inventory_keys_by_cert
            _link_inventory_photo_to_keys = app.CardPipelineApp._link_inventory_photo_to_keys
            _inventory_photo_match_keys = app.CardPipelineApp._inventory_photo_match_keys
            _inventory_photo_best_title_match = app.CardPipelineApp._inventory_photo_best_title_match
            _inventory_photo_card_match_text = app.CardPipelineApp._inventory_photo_card_match_text
            _match_text_tokens = app.CardPipelineApp._match_text_tokens
            _inventory_photo_filename_hint = app.CardPipelineApp._inventory_photo_filename_hint
            _compact_match_text = app.CardPipelineApp._compact_match_text
            _inventory_photo_capture_group_key = app.CardPipelineApp._inventory_photo_capture_group_key
            _sold_inventory_cert_numbers = lambda self: set()
            _inventory_photo_state_matches_sold_cert = app.CardPipelineApp._inventory_photo_state_matches_sold_cert
            _inventory_photo_base64 = lambda self, path: "stub-b64"
            _inventory_photo_scan_group_nearby_unmatched = app.CardPipelineApp._inventory_photo_scan_group_nearby_unmatched
            _inventory_photo_scan_worker = app.CardPipelineApp._inventory_photo_scan_worker

        class Verification:
            def get(self, key, default=None):
                return {"cert_number": "0010133787", "grading_company": "BGS", "label_text": "CERT 0010133787"}.get(key, default)

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            old_verify = app._verify_cert_only_sync
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            photo = app.INVENTORY_PHOTOS_DIR / "kobe-bgs.jpg"
            photo.write_bytes(b"fake image")
            dummy = PhotoDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            dummy.inventory_photo_client = object()
            dummy.events = __import__("queue").Queue()
            record = dummy._normalize_inventory_record({"assigned_person": "Mikey", "cert_number": "0010133787", "card_title": "Kobe Bryant BGS", "status": "Active"})
            dummy._save_inventory_ledger([record])
            try:
                app._verify_cert_only_sync = lambda _client, _image_b64: Verification()
                with patch.object(app, "identify_cards_sync", return_value=[{"grading_company": "BGS", "cert_number": "", "label_text": "BECKETT 10 AUTOGRAPH"}]):
                    dummy._inventory_photo_scan_worker(app.INVENTORY_PHOTOS_DIR)
                ledger = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(ledger[0]["photo_paths"], [photo.name])
                state = json.loads(app.INVENTORY_PHOTO_STATE_PATH.read_text(encoding="utf-8"))
                state_record = next(iter(state["photos"].values()))
                self.assertEqual(state_record["certs"], ["0010133787"])
                self.assertEqual(state_record["linked_keys"], [record["inventory_key"]])
            finally:
                app._verify_cert_only_sync = old_verify
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_create_photo_ocr_rescues_single_bgs_cert_from_full_photo(self) -> None:
        class PhotoCreateDummy:
            _inventory_photo_rescue_single_bgs_cert = app.CardPipelineApp._inventory_photo_rescue_single_bgs_cert
            _photo_card_to_row = app.CardPipelineApp._photo_card_to_row
            _photo_card_has_inventory = app.CardPipelineApp._photo_card_has_inventory
            _photo_scan_worker = app.CardPipelineApp._photo_scan_worker

        class Verification:
            def get(self, key, default=None):
                return {"cert_number": "0010133787", "grading_company": "BGS", "label_text": "CERT 0010133787"}.get(key, default)

        with TemporaryDirectory() as tmp:
            old_verify = app._verify_cert_only_sync
            photo = Path(tmp) / "kobe-bgs.jpg"
            photo.write_bytes(b"fake image")
            dummy = PhotoCreateDummy()
            dummy.photo_paths = [photo]
            dummy.photo_client = object()
            dummy.events = __import__("queue").Queue()
            try:
                app._verify_cert_only_sync = lambda _client, _image_b64: Verification()
                with patch.object(app, "identify_cards_sync", return_value=[{"grading_company": "BGS", "cert_number": "", "label_text": "BECKETT 10 AUTOGRAPH", "grade": "10"}]):
                    dummy._photo_scan_worker()
                rows = []
                while not dummy.events.empty():
                    event, payload = dummy.events.get()
                    if event == "photo_rows":
                        rows.extend(payload)
                self.assertEqual(len(rows), 1)
                self.assertEqual(rows[0]["cert_number"], "0010133787")
                self.assertEqual(rows[0]["grader"], "BGS")
            finally:
                app._verify_cert_only_sync = old_verify

    def test_unattached_photo_picker_hides_stale_linked_state_records(self) -> None:
        class PhotoPickerDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _inventory_photo_file_hash = app.CardPipelineApp._inventory_photo_file_hash
            _inventory_photo_paths = app.CardPipelineApp._inventory_photo_paths
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_shared_folder = app.CardPipelineApp._inventory_photo_shared_folder
            _inventory_photo_relative_path = app.CardPipelineApp._inventory_photo_relative_path
            _inventory_photo_storage_value = app.CardPipelineApp._inventory_photo_storage_value
            _inventory_photo_path_candidates = app.CardPipelineApp._inventory_photo_path_candidates
            _inventory_photo_used_path_keys = app.CardPipelineApp._inventory_photo_used_path_keys
            _inventory_photo_used_hashes = app.CardPipelineApp._inventory_photo_used_hashes
            _inventory_photo_state_used_keys = app.CardPipelineApp._inventory_photo_state_used_keys
            _sold_inventory_cert_numbers = app.CardPipelineApp._sold_inventory_cert_numbers
            _sold_inventory_photo_used_keys = app.CardPipelineApp._sold_inventory_photo_used_keys
            _inventory_photo_state_matches_sold_cert = app.CardPipelineApp._inventory_photo_state_matches_sold_cert
            _inventory_unattached_photo_paths = app.CardPipelineApp._inventory_unattached_photo_paths

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_profit = app.PROFIT_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            photo = app.INVENTORY_PHOTOS_DIR / "[20260708-0246]-Card[10]-[1]-[].jpg"
            sold_path_photo = app.INVENTORY_PHOTOS_DIR / "sold-path.jpg"
            photo.write_bytes(b"fake image")
            sold_path_photo.write_bytes(b"sold path image")
            dummy = PhotoPickerDummy()
            dummy.app_settings = {}
            dummy._save_inventory_ledger([])
            dummy._save_profit_ledger([
                {"cert_number": "65774395", "card_title": "Sold Card", "sale_price": 20},
                {"cert_number": "44444444", "card_title": "Sold Path Card", "sale_price": 25, "photo_paths": [str(sold_path_photo)]},
            ])
            sha = dummy._inventory_photo_file_hash(photo)
            app.INVENTORY_PHOTO_STATE_PATH.write_text(
                json.dumps(
                    {
                        "version": 1,
                        "photos": {
                            sha: {
                                "path": str(Path(tmp) / "missing-phone-album" / photo.name),
                                "filename": photo.name,
                                "certs": ["65774395"],
                                "linked_keys": [],
                                "status": "no_matching_inventory",
                            }
                        },
                    }
                ),
                encoding="utf-8",
            )
            try:
                self.assertEqual(dummy._inventory_unattached_photo_paths(), [])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.PROFIT_LEDGER_PATH = old_profit
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state

    def test_inventory_sold_archives_unshared_photo_file_for_two_weeks(self) -> None:
        class PhotoSoldDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _safe_inventory_photo_path = app.CardPipelineApp._safe_inventory_photo_path
            _deleted_archive_metadata_path = app.CardPipelineApp._deleted_archive_metadata_path
            _unique_deleted_archive_path = app.CardPipelineApp._unique_deleted_archive_path
            _archive_deleted_file = app.CardPipelineApp._archive_deleted_file
            _purge_expired_deleted_archive = app.CardPipelineApp._purge_expired_deleted_archive
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _delete_inventory_photo_files_for_removed_records = app.CardPipelineApp._delete_inventory_photo_files_for_removed_records
            _mark_inventory_record_sold = app.CardPipelineApp._mark_inventory_record_sold
            _append_activity = lambda self, action, summary, details=None: None

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            old_deleted_archive = app.DELETED_ARCHIVE_DIR
            old_deleted_photos = app.DELETED_INVENTORY_PHOTOS_DIR
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.DELETED_ARCHIVE_DIR = Path(tmp) / "DELETED ARCHIVE"
            app.DELETED_INVENTORY_PHOTOS_DIR = app.DELETED_ARCHIVE_DIR / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            photo = app.INVENTORY_PHOTOS_DIR / "card.jpg"
            photo.write_bytes(b"fake image")
            dummy = PhotoSoldDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {}
            record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "123", "card_title": "Test", "status": "Active", "photo_paths": [str(photo)]})
            dummy._save_inventory_ledger([record])
            try:
                self.assertEqual(dummy._mark_inventory_record_sold(str(record["inventory_key"]), "Arena Club", 10), 1)
                self.assertFalse(photo.exists())
                archived = list(app.DELETED_INVENTORY_PHOTOS_DIR.rglob("card.jpg"))
                self.assertEqual(len(archived), 1)
                metadata_path = archived[0].with_name("card.jpg.archive.json")
                self.assertTrue(metadata_path.exists())
                metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
                self.assertEqual(metadata["original_path"], str(photo))
                self.assertEqual(metadata["reason"], "inventory_photo_removed")
                self.assertEqual(metadata["retention_days"], 14)
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state
                app.DELETED_ARCHIVE_DIR = old_deleted_archive
                app.DELETED_INVENTORY_PHOTOS_DIR = old_deleted_photos

    def test_inventory_photo_export_to_desktop_copies_each_photo_without_overwrite(self) -> None:
        class PhotoExportDummy:
            _desktop_photo_export_destination = app.CardPipelineApp._desktop_photo_export_destination
            _export_inventory_photos_to_desktop = app.CardPipelineApp._export_inventory_photos_to_desktop

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            desktop = root / "Desktop"
            source_dir = root / "photos"
            source_dir.mkdir(parents=True)
            first = source_dir / "front.jpg"
            second = source_dir / "back.jpg"
            first.write_bytes(b"front")
            second.write_bytes(b"back")
            record = {"card_title": "2001 Pokemon Neo Discovery - 1st Ed. 46/75 Scyther CGC 10"}
            desktop.mkdir(parents=True)
            preexisting = desktop / "2001-Pokemon-Neo-Discovery-1st-Ed.-46-75-Scyther-CGC-10-photo-1-front.jpg"
            preexisting.write_bytes(b"existing")
            dummy = PhotoExportDummy()
            with patch.object(app.Path, "home", return_value=root):
                exported = dummy._export_inventory_photos_to_desktop(record, [first, second])
            self.assertEqual(len(exported), 2)
            self.assertEqual(exported[0].name, "2001-Pokemon-Neo-Discovery-1st-Ed.-46-75-Scyther-CGC-10-photo-1-front-2.jpg")
            self.assertEqual(exported[1].name, "2001-Pokemon-Neo-Discovery-1st-Ed.-46-75-Scyther-CGC-10-photo-2-back.jpg")
            self.assertEqual(preexisting.read_bytes(), b"existing")
            self.assertEqual(exported[0].read_bytes(), b"front")
            self.assertEqual(exported[1].read_bytes(), b"back")

    def test_inventory_sold_archives_matching_source_and_shared_photo_files(self) -> None:
        class PhotoSoldDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_shared_folder = app.CardPipelineApp._inventory_photo_shared_folder
            _inventory_photo_relative_path = app.CardPipelineApp._inventory_photo_relative_path
            _inventory_photo_path_candidates = app.CardPipelineApp._inventory_photo_path_candidates
            _inventory_photo_safe_candidates = app.CardPipelineApp._inventory_photo_safe_candidates
            _safe_inventory_photo_path = app.CardPipelineApp._safe_inventory_photo_path
            _deleted_archive_metadata_path = app.CardPipelineApp._deleted_archive_metadata_path
            _unique_deleted_archive_path = app.CardPipelineApp._unique_deleted_archive_path
            _archive_deleted_file = app.CardPipelineApp._archive_deleted_file
            _purge_expired_deleted_archive = app.CardPipelineApp._purge_expired_deleted_archive
            _load_inventory_photo_state = app.CardPipelineApp._load_inventory_photo_state
            _save_inventory_photo_state = app.CardPipelineApp._save_inventory_photo_state
            _delete_inventory_photo_files_for_removed_records = app.CardPipelineApp._delete_inventory_photo_files_for_removed_records
            _mark_inventory_record_sold = app.CardPipelineApp._mark_inventory_record_sold
            _append_activity = lambda self, action, summary, details=None: None

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_photo_state = app.INVENTORY_PHOTO_STATE_PATH
            old_deleted_archive = app.DELETED_ARCHIVE_DIR
            old_deleted_photos = app.DELETED_INVENTORY_PHOTOS_DIR
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "LUCAS_PERSONAL" / "INVENTORY PHOTOS"
            app.INVENTORY_PHOTO_STATE_PATH = Path(tmp) / "inventory_photo_state.json"
            app.DELETED_ARCHIVE_DIR = Path(tmp) / "DELETED ARCHIVE"
            app.DELETED_INVENTORY_PHOTOS_DIR = app.DELETED_ARCHIVE_DIR / "INVENTORY PHOTOS"
            source_dir = Path(tmp) / "iCloud Source"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            source_dir.mkdir(parents=True)
            shared_photo = app.INVENTORY_PHOTOS_DIR / "card.jpg"
            source_photo = source_dir / "card.jpg"
            shared_photo.write_bytes(b"shared image")
            source_photo.write_bytes(b"source image")
            dummy = PhotoSoldDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            dummy.app_settings = {"inventory_photo_folder": str(source_dir)}
            record = dummy._normalize_inventory_record({"assigned_person": "Kevin", "cert_number": "123", "card_title": "Test", "status": "Active", "photo_paths": [str(shared_photo)]})
            dummy._save_inventory_ledger([record])
            try:
                self.assertEqual(dummy._mark_inventory_record_sold(str(record["inventory_key"]), "Arena Club", 10), 1)
                self.assertFalse(shared_photo.exists())
                self.assertFalse(source_photo.exists())
                archived = sorted(path.name for path in app.DELETED_INVENTORY_PHOTOS_DIR.rglob("card*.jpg"))
                self.assertEqual(archived, ["card-2.jpg", "card.jpg"])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.INVENTORY_PHOTO_STATE_PATH = old_photo_state
                app.DELETED_ARCHIVE_DIR = old_deleted_archive
                app.DELETED_INVENTORY_PHOTOS_DIR = old_deleted_photos

    def test_refund_restore_brings_archived_inventory_photo_back(self) -> None:
        class PhotoRestoreDummy:
            _inventory_photo_relative_path = app.CardPipelineApp._inventory_photo_relative_path
            _inventory_photo_shared_folder = app.CardPipelineApp._inventory_photo_shared_folder
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_path_candidates = app.CardPipelineApp._inventory_photo_path_candidates
            _deleted_archive_metadata_path = app.CardPipelineApp._deleted_archive_metadata_path
            _restore_inventory_photo_files_for_records = app.CardPipelineApp._restore_inventory_photo_files_for_records
            _append_activity = lambda self, action, summary, details=None: None

        with TemporaryDirectory() as tmp:
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_deleted_archive = app.DELETED_ARCHIVE_DIR
            old_deleted_photos = app.DELETED_INVENTORY_PHOTOS_DIR
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.DELETED_ARCHIVE_DIR = Path(tmp) / "DELETED ARCHIVE"
            app.DELETED_INVENTORY_PHOTOS_DIR = app.DELETED_ARCHIVE_DIR / "INVENTORY PHOTOS"
            archive_dir = app.DELETED_INVENTORY_PHOTOS_DIR / "2026-07-10"
            archive_dir.mkdir(parents=True)
            original = app.INVENTORY_PHOTOS_DIR / "card.jpg"
            archived = archive_dir / "card.jpg"
            archived.write_bytes(b"fake image")
            archived.with_name("card.jpg.archive.json").write_text(
                json.dumps(
                    {
                        "original_path": str(original),
                        "archive_path": str(archived),
                        "reason": "inventory_photo_removed",
                    }
                ),
                encoding="utf-8",
            )
            try:
                dummy = PhotoRestoreDummy()
                dummy.app_settings = {}
                restored = dummy._restore_inventory_photo_files_for_records([{"photo_paths": [str(original)]}])

                self.assertEqual(restored, 1)
                self.assertTrue(original.exists())
                self.assertFalse(archived.exists())
                self.assertFalse(archived.with_name("card.jpg.archive.json").exists())
            finally:
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.DELETED_ARCHIVE_DIR = old_deleted_archive
                app.DELETED_INVENTORY_PHOTOS_DIR = old_deleted_photos


    def test_refund_restore_recovers_photo_paths_from_archive_metadata(self) -> None:
        class PhotoRestoreDummy:
            _inventory_photo_relative_path = app.CardPipelineApp._inventory_photo_relative_path
            _inventory_photo_shared_folder = app.CardPipelineApp._inventory_photo_shared_folder
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_path_candidates = app.CardPipelineApp._inventory_photo_path_candidates
            _deleted_archive_metadata_path = app.CardPipelineApp._deleted_archive_metadata_path
            _restore_inventory_photo_files_for_records = app.CardPipelineApp._restore_inventory_photo_files_for_records
            _append_activity = lambda self, action, summary, details=None: None

        with TemporaryDirectory() as tmp:
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            old_deleted_archive = app.DELETED_ARCHIVE_DIR
            old_deleted_photos = app.DELETED_INVENTORY_PHOTOS_DIR
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "INVENTORY PHOTOS"
            app.DELETED_ARCHIVE_DIR = Path(tmp) / "DELETED ARCHIVE"
            app.DELETED_INVENTORY_PHOTOS_DIR = app.DELETED_ARCHIVE_DIR / "INVENTORY PHOTOS"
            archive_dir = app.DELETED_INVENTORY_PHOTOS_DIR / "2026-07-10"
            archive_dir.mkdir(parents=True)
            original = app.INVENTORY_PHOTOS_DIR / "mr-mime-front.jpg"
            archived = archive_dir / "mr-mime-front.jpg"
            archived.write_bytes(b"fake image")
            archived.with_name("mr-mime-front.jpg.archive.json").write_text(
                json.dumps(
                    {
                        "original_path": str(original),
                        "archive_path": str(archived),
                        "reason": "inventory_photo_removed",
                        "details": {
                            "cert_number": "4151253025",
                            "card_title": "1997 Pokemon Jungle Japanese Mr. Mime Holo CGC 7.5",
                            "source_sheet": "Kevin_6_16_2026.xlsx",
                        },
                    }
                ),
                encoding="utf-8",
            )
            try:
                dummy = PhotoRestoreDummy()
                dummy.app_settings = {}
                record = {"cert_number": "4151253025", "card_title": "1997 Pokemon Jungle Japanese Mr. Mime Holo CGC 7.5", "source_sheet": "Kevin_6_16_2026.xlsx", "photo_paths": []}
                restored = dummy._restore_inventory_photo_files_for_records([record])

                self.assertEqual(restored, 1)
                self.assertEqual(record["photo_paths"], [str(original)])
                self.assertTrue(original.exists())
            finally:
                app.INVENTORY_PHOTOS_DIR = old_photo_dir
                app.DELETED_ARCHIVE_DIR = old_deleted_archive
                app.DELETED_INVENTORY_PHOTOS_DIR = old_deleted_photos

    def test_manual_inventory_add_accepts_cert_without_grader(self) -> None:
        class ManualAddDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_sport_from_value = app.CardPipelineApp._inventory_sport_from_value
            _next_raw_item_id = app.CardPipelineApp._next_raw_item_id
            _raw_item_id_namespace = lambda self: "TEAM"
            add_raw_inventory_card = app.CardPipelineApp.add_raw_inventory_card
            refresh_inventory_tab = lambda self: None
            _append_activity = lambda self, action, summary, details=None: None

            def _raw_inventory_card_dialog(self):
                return {
                    "assigned_person": "Kevin Hambone",
                    "cert_number": "12345678",
                    "grader": "",
                    "card_title": "2024 Panini Prizm Test Card",
                    "purchase_price": 25,
                }

            def _enrich_inventory_record_assignment(self, record, force=False):
                return self._normalize_inventory_record(record)

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = ManualAddDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            class Status:
                def __init__(self):
                    self.value = ""
                def set(self, value):
                    self.value = value
            dummy.status_var = Status()
            dummy.inventory_status_var = Status()
            try:
                dummy.add_raw_inventory_card()
                inventory = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(len(inventory), 1)
                self.assertEqual(inventory[0]["cert_number"], "12345678")
                self.assertEqual(inventory[0]["grader"], "")
                self.assertEqual(inventory[0]["item_type"], "Graded")
                self.assertEqual(inventory[0]["item_id"], "")
                self.assertEqual(inventory[0]["source_sheet"], "Manual Inventory")
                self.assertEqual(dummy.status_var.value, "Added inventory card 12345678.")
                self.assertEqual(dummy.inventory_status_var.value, "Added inventory card 12345678.")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_raw_inventory_record_uses_item_id_for_sold_profit_and_expense(self) -> None:
        class SoldDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _mark_inventory_record_sold = app.CardPipelineApp._mark_inventory_record_sold
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _inventory_sale_profit_record = app.CardPipelineApp._inventory_sale_profit_record
            _inventory_sale_expense_record = app.CardPipelineApp._inventory_sale_expense_record
            _general_sold_sheet_name = app.CardPipelineApp._general_sold_sheet_name
            _next_raw_item_id = app.CardPipelineApp._next_raw_item_id
            _raw_item_id_namespace = lambda self: "TEAM"
            mark_inventory_record_sold = app.CardPipelineApp.mark_inventory_record_sold
            record_profit_sales = app.CardPipelineApp.record_profit_sales
            refresh_profit_tab = lambda self: None
            _append_activity = lambda self, action, summary, details=None: None

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_profit = app.PROFIT_LEDGER_PATH
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            dummy = SoldDummy()
            dummy.lucas_identity = {"display_name": "Tester", "machine": "Test"}
            try:
                item_id = dummy._next_raw_item_id([])
                record = dummy._normalize_inventory_record(
                    {
                        "item_type": "Raw",
                        "item_id": item_id,
                        "assigned_person": "Hambone",
                        "card_title": "2024 Panini Prizm Test Raw Card",
                        "purchase_price": 40,
                        "card_ladder_value": 100,
                        "best_company": "Arena Club",
                        "estimated_payout": 90,
                        "source_sheet": "Raw Inventory",
                        "photo_paths": ["front.jpg", "back.jpg"],
                        "status": "Active",
                    }
                )
                self.assertEqual(record["inventory_key"], item_id.lower())
                dummy._save_inventory_ledger([record])

                self.assertTrue(dummy.mark_inventory_record_sold(record, "Arena Club", 95, expense_type="Shipping", expense_amount=5))

                inventory = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                profit = json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8"))
                self.assertEqual(inventory, [])
                sale_row = next(record for record in profit if record.get("record_type") != "expense")
                expense_row = next(record for record in profit if record.get("record_type") == "expense")
                self.assertEqual(sale_row["item_type"], "Raw")
                self.assertEqual(sale_row["item_id"], item_id)
                self.assertEqual(sale_row["cert_number"], "")
                self.assertEqual(sale_row["photo_paths"], ["front.jpg", "back.jpg"])
                self.assertEqual(sale_row["photo_count"], 2)
                self.assertEqual(expense_row["item_id"], item_id)
                self.assertEqual(expense_row["cert_number"], "")
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.PROFIT_LEDGER_PATH = old_profit
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_blank_inventory_sale_company_uses_person_general_sold_sheet(self) -> None:
        class SoldDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _general_sold_sheet_name = app.CardPipelineApp._general_sold_sheet_name
            _inventory_sale_profit_record = app.CardPipelineApp._inventory_sale_profit_record

        dummy = SoldDummy()
        record = dummy._inventory_sale_profit_record(
            {
                "assigned_person": "Kevin Hambone",
                "cert_number": "123",
                "card_title": "Test Card",
                "source_sheet": "Original.xlsx",
                "purchase_price": 40,
            },
            "",
            95,
        )
        self.assertEqual(record["company"], "General Sold")
        self.assertEqual(record["source_sheet"], "Kevin Hambone General Sold")
        self.assertEqual(record["original_source_sheet"], "Original.xlsx")
        self.assertEqual(record["assigned_person"], "Kevin Hambone")

    def test_refund_profit_record_returns_card_to_inventory_and_removes_company_row(self) -> None:
        class RefundDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            add_inventory_records = app.CardPipelineApp.add_inventory_records
            _enrich_inventory_record_assignment = lambda self, record: record
            refresh_inventory_tab = lambda self: None

        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            company_dir = root / "COMPANY SHEETS"
            company_path = company_dir / "Arena Club" / "Arena Club.xlsx"
            company_path.parent.mkdir(parents=True)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Week of 2026-06-15"
            sheet.append(["Date Added", "Source Sheet", "Source", "Certification Number", "Grader", "Card Description", "Purchase Price", "Card Ladder Value", "Comps", "CY Estimate", "CY Confidence", "Best Company", "Estimated Payout", "Status", "Notes"])
            sheet.append(["2026-06-17", "Lot A.xlsx", "", "123", "PSA", "Test Card", 40, "", 100, "", "", "Arena Club", 90, "Received", ""])
            workbook.save(company_path)

            old_company = app.COMPANY_SHEETS_DIR
            old_profit = app.PROFIT_LEDGER_PATH
            old_inventory = app.INVENTORY_LEDGER_PATH
            app.COMPANY_SHEETS_DIR = company_dir
            app.PROFIT_LEDGER_PATH = root / "profit_ledger.json"
            app.INVENTORY_LEDGER_PATH = root / "inventory_ledger.json"
            record = {
                "date_added": "2026-06-17",
                "company": "Arena Club",
                "weekly_sheet_name": "Arena Club.xlsx:Week of 2026-06-15",
                "source_sheet": "Lot A.xlsx",
                "cert_number": "123",
                "grader": "PSA",
                "card_title": "Test Card",
                "purchase_price": 40,
                "sale_price": 90,
                "assigned_person": "Lucas",
                "photo_paths": ["front.jpg"],
            }
            dummy = RefundDummy()
            normalized = dummy._normalize_profit_record(record)
            app.PROFIT_LEDGER_PATH.write_text(json.dumps([normalized]), encoding="utf-8")
            try:
                ledger = [dummy._normalize_profit_record(item) for item in dummy._load_profit_ledger()]
                kept = [item for item in ledger if item["ledger_key"] != normalized["ledger_key"]]
                dummy._save_profit_ledger(kept)
                remove_company_sheet_rows_for_source = app.remove_company_sheet_rows_for_source
                remove_company_sheet_rows_for_source(app.COMPANY_SHEETS_DIR, "Lot A.xlsx", {"123"})
                dummy.add_inventory_records([
                    dummy._normalize_inventory_record(
                        {
                            "assigned_person": "Lucas",
                            "cert_number": "123",
                            "grader": "PSA",
                            "card_title": "Test Card",
                            "purchase_price": 40,
                            "inventory_value": 90,
                            "source_sheet": "Lot A.xlsx",
                            "photo_paths": normalized.get("photo_paths") or [],
                            "status": "Active",
                            "notes": "Refunded from sold cards",
                        }
                    )
                ])

                self.assertEqual(json.loads(app.PROFIT_LEDGER_PATH.read_text(encoding="utf-8")), [])
                inventory = json.loads(app.INVENTORY_LEDGER_PATH.read_text(encoding="utf-8"))["items"]
                self.assertEqual(inventory[0]["photo_paths"], ["front.jpg"])
                self.assertEqual(inventory[0]["photo_count"], 1)
                self.assertEqual(read_company_profit_records(company_dir), [])
                self.assertEqual(len(inventory), 1)
                self.assertEqual(inventory[0]["status"], "Active")
                self.assertEqual(inventory[0]["assigned_person"], "Lucas")
            finally:
                app.COMPANY_SHEETS_DIR = old_company
                app.PROFIT_LEDGER_PATH = old_profit
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_profit_records_are_enriched_with_assigned_person_from_sheet_marker(self) -> None:
        class ProfitDummy:
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _money_value = app.CardPipelineApp._money_value
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _home_sheet_key = app.CardPipelineApp._home_sheet_key
            _split_home_sheet_key = app.CardPipelineApp._split_home_sheet_key
            _person_for_profit_record = app.CardPipelineApp._person_for_profit_record
            _enrich_profit_records_with_people = app.CardPipelineApp._enrich_profit_records_with_people
            _filtered_profit_records = app.CardPipelineApp._filtered_profit_records
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _profit_today = lambda self: datetime(2026, 6, 17).date()
            _profit_period_bounds = app.CardPipelineApp._profit_period_bounds

        dummy = ProfitDummy()
        dummy.home_sheet_markers = {"Received|Lot A.xlsx": {"assigned_person": "Lucas"}}
        dummy.profit_person_var = types.SimpleNamespace(get=lambda: "luc")
        dummy.profit_period_var = types.SimpleNamespace(get=lambda: "Total")
        dummy.profit_search_var = types.SimpleNamespace(get=lambda: "")

        rows = dummy._enrich_profit_records_with_people([
            {
                "date_added": "2026-06-11",
                "company": "Arena Club",
                "source_sheet": "Lot A.xlsx",
                "assigned_person": "Unassigned",
                "cert_number": "123",
                "card_title": "Test Card",
                "purchase_price": 40,
                "sale_price": 90,
            }
        ])

        self.assertEqual(rows[0]["assigned_person"], "Lucas")
        self.assertEqual(rows[0]["profit"], 50)
        self.assertEqual(dummy._filtered_profit_records(rows), rows)

        dummy.profit_search_var = types.SimpleNamespace(get=lambda: "123 test")
        self.assertEqual(dummy._filtered_profit_records(rows), rows)

        dummy.profit_search_var = types.SimpleNamespace(get=lambda: "fanatics")
        self.assertEqual(dummy._filtered_profit_records(rows), [])

    def test_profit_period_filter_and_chart_series_support_daily_and_overall_views(self) -> None:
        class ProfitDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _profit_today = lambda self: datetime(2026, 6, 17).date()
            _profit_period_bounds = app.CardPipelineApp._profit_period_bounds
            _canonical_profit_period = app.CardPipelineApp._canonical_profit_period
            _profit_period_label = app.CardPipelineApp._profit_period_label
            _profit_graph_label = app.CardPipelineApp._profit_graph_label
            _profit_plot_label = app.CardPipelineApp._profit_plot_label
            _profit_chart_title = app.CardPipelineApp._profit_chart_title
            _filtered_profit_records = app.CardPipelineApp._filtered_profit_records
            _profit_chart_series = app.CardPipelineApp._profit_chart_series
            _profit_chart_lines = app.CardPipelineApp._profit_chart_lines
            _profit_chart_tooltip_value = app.CardPipelineApp._profit_chart_tooltip_value
            _profit_sport_label = app.CardPipelineApp._profit_sport_label
            _profit_company_label = app.CardPipelineApp._profit_company_label
            _profit_company_chart_series = app.CardPipelineApp._profit_company_chart_series
            _profit_chart_bucket_label = app.CardPipelineApp._profit_chart_bucket_label
            _profit_chart_bucket_display = app.CardPipelineApp._profit_chart_bucket_display
            _profit_chart_bucket_range = app.CardPipelineApp._profit_chart_bucket_range

        rows = [
            {"assigned_person": "Lucas", "date_added": "2026-06-17", "profit": 30},
            {"assigned_person": "Lucas", "date_added": "2026-06-13", "profit": 20},
            {"assigned_person": "Lucas", "date_added": "2026-06-10", "profit": 100},
            {"assigned_person": "Mikey", "date_added": "2026-06-17", "profit": 7},
            {"assigned_person": "Lucas", "date_added": "not-a-date", "profit": 50},
        ]

        dummy = ProfitDummy()
        self.assertEqual(dummy._profit_period_label(), "Year")
        self.assertEqual(dummy._profit_graph_label(), "Overall Profit")
        self.assertEqual(dummy._profit_chart_title(), "Overall Profit (Year)")
        self.assertEqual(dummy._profit_chart_tooltip_value(123.45, False), "$123.45")
        self.assertEqual(dummy._profit_chart_tooltip_value(0.1234, True), "12.34%")

        dummy.profit_person_var = types.SimpleNamespace(get=lambda: "luc")
        dummy.profit_period_var = types.SimpleNamespace(get=lambda: "5 Days")
        dummy.profit_graph_var = types.SimpleNamespace(get=lambda: "Daily Trend")
        dummy.profit_search_var = types.SimpleNamespace(get=lambda: "")
        self.assertEqual(dummy._profit_chart_title(), "Daily Trend (5 Days)")

        filtered = dummy._filtered_profit_records(rows)
        days, daily_values = dummy._profit_chart_series(filtered)
        dummy.profit_graph_var = types.SimpleNamespace(get=lambda: "Overall Profit")
        self.assertEqual(dummy._profit_chart_title(), "Overall Profit (5 Days)")
        overall_days, overall_values = dummy._profit_chart_series(filtered)

        self.assertEqual([record["date_added"] for record in filtered], ["2026-06-17", "2026-06-13"])
        self.assertEqual(days, ["2026-06-13", "2026-06-14", "2026-06-15", "2026-06-16", "2026-06-17"])
        self.assertEqual(daily_values, [20, 0.0, 0.0, 0.0, 30])
        self.assertEqual(overall_days, days)
        self.assertEqual(overall_values, [20, 20.0, 20.0, 20.0, 50.0])

        ytd_rows = [
            {"assigned_person": "Lucas", "date_added": "2026-01-05", "profit": 40},
            {"assigned_person": "Lucas", "date_added": "2026-06-17", "profit": 30},
        ]
        dummy.profit_period_var = types.SimpleNamespace(get=lambda: "YTD")
        dummy.profit_graph_var = types.SimpleNamespace(get=lambda: "Daily Trend")
        ytd_filtered = dummy._filtered_profit_records(ytd_rows)
        ytd_days, ytd_values = dummy._profit_chart_series(ytd_filtered)

        self.assertEqual(ytd_days[0], "2026-01-01")
        self.assertEqual(ytd_days[-1], "2026-06-17")
        self.assertEqual(len(ytd_days), 168)
        self.assertEqual(ytd_values[ytd_days.index("2026-01-05")], 40)
        self.assertEqual(ytd_values[ytd_days.index("2026-06-17")], 30)

        sport_rows = [
            {"assigned_person": "Lucas", "date_added": "2026-01-05", "sport": "football", "card_title": "2024 Panini Prizm Jayden Daniels", "profit": 20, "sale_price": 100},
            {"assigned_person": "Lucas", "date_added": "2026-02-01", "sport": "football", "card_title": "2024 Panini Prizm Jayden Daniels Silver", "profit": 15, "sale_price": 75},
            {"assigned_person": "Lucas", "date_added": "2026-01-09", "sport": "baseball", "card_title": "2024 Topps Chrome Shohei Ohtani", "profit": 30, "sale_price": 150},
            {"assigned_person": "Lucas", "date_added": "2026-02-02", "sport": "basketball", "card_title": "Victor Wembanyama", "profit": 25, "sale_price": 100},
        ]
        dummy.profit_period_var = types.SimpleNamespace(get=lambda: "Year")
        dummy.profit_graph_var = types.SimpleNamespace(get=lambda: "Overall Profit")
        dummy.profit_plot_var = types.SimpleNamespace(get=lambda: "By Sport")
        labels, lines, percent_mode = dummy._profit_chart_lines(sport_rows)
        self.assertFalse(percent_mode)
        self.assertEqual(labels[:2], ["2026-01", "2026-02"])
        line_lookup = {line["label"]: line["values"] for line in lines}
        self.assertEqual(line_lookup["Football"][:3], [20.0, 35.0, 35.0])
        self.assertEqual(line_lookup["Baseball"][:3], [30.0, 30.0, 30.0])
        self.assertEqual(line_lookup["Basketball"][:3], [0.0, 25.0, 25.0])

        dummy.profit_period_var = types.SimpleNamespace(get=lambda: "Year")
        dummy.profit_graph_var = types.SimpleNamespace(get=lambda: "Profit to Sales Ratio")
        dummy.profit_plot_var = types.SimpleNamespace(get=lambda: "By Sport")
        labels, lines, percent_mode = dummy._profit_chart_lines(sport_rows)
        self.assertTrue(percent_mode)
        self.assertEqual(labels[:2], ["2026-01", "2026-02"])
        line_lookup = {line["label"]: line["values"] for line in lines}
        self.assertAlmostEqual(line_lookup["Football"][0], 0.20)
        self.assertAlmostEqual(line_lookup["Football"][1], 0.20)
        self.assertAlmostEqual(line_lookup["Baseball"][0], 0.20)
        self.assertAlmostEqual(line_lookup["Basketball"][1], 0.25)

        company_rows = [
            {"assigned_person": "Lucas", "date_added": "2026-06-17", "profit": 30, "company": "Fanatics"},
            {"assigned_person": "Lucas", "date_added": "2026-06-16", "profit": -40, "company": "Arena Club"},
            {"assigned_person": "Lucas", "date_added": "2026-06-15", "profit": 10, "company": "Fanatics"},
            {"assigned_person": "Lucas", "date_added": "2026-06-14", "profit": 5, "company": ""},
            {"assigned_person": "Lucas", "date_added": "2026-06-17", "profit": -100, "record_type": "expense", "company": "Fees"},
        ]
        dummy.profit_period_var = types.SimpleNamespace(get=lambda: "5 Days")
        dummy.profit_graph_var = types.SimpleNamespace(get=lambda: "Profit by Company")
        dummy.profit_plot_var = types.SimpleNamespace(get=lambda: "By Sport")
        company_labels, company_lines = dummy._profit_company_chart_series(company_rows)
        self.assertEqual(company_labels, ["2026-06-13", "2026-06-14", "2026-06-15", "2026-06-16", "2026-06-17"])
        company_values = {line["label"]: line["values"] for line in company_lines}
        self.assertEqual(company_values["Fanatics"], [0.0, 0.0, 10.0, 0.0, 30.0])
        self.assertEqual(company_values["Arena Club"], [0.0, 0.0, 0.0, -40.0, 0.0])
        self.assertEqual(company_values["General Sold"], [0.0, 5.0, 0.0, 0.0, 0.0])
        company_chart_labels, company_chart_lines, company_percent_mode = dummy._profit_chart_lines(company_rows)
        self.assertFalse(company_percent_mode)
        self.assertEqual(company_chart_labels, company_labels)
        self.assertEqual([line["label"] for line in company_chart_lines], ["Fanatics", "Arena Club", "General Sold"])
        self.assertEqual(dummy._profit_chart_title(), "Profit by Company (5 Days)")

    def test_profit_periods_include_calendar_month_and_last_thirty_days(self) -> None:
        class ProfitDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _profit_today = lambda self: datetime(2026, 7, 5).date()
            _profit_period_bounds = app.CardPipelineApp._profit_period_bounds
            _canonical_profit_period = app.CardPipelineApp._canonical_profit_period
            _filtered_profit_records = app.CardPipelineApp._filtered_profit_records

        rows = [
            {"assigned_person": "Lucas", "date_added": "2026-06-05", "profit": 10},
            {"assigned_person": "Lucas", "date_added": "2026-06-06", "profit": 20},
            {"assigned_person": "Lucas", "date_added": "2026-07-05", "profit": 30},
        ]

        dummy = ProfitDummy()
        dummy.profit_person_var = types.SimpleNamespace(get=lambda: "")
        dummy.profit_search_var = types.SimpleNamespace(get=lambda: "")

        period_start, period_end = dummy._profit_period_bounds("Last 30 Days")
        dummy.profit_period_var = types.SimpleNamespace(get=lambda: "Last 30 Days")
        filtered = dummy._filtered_profit_records(rows)

        self.assertEqual(period_start.isoformat(), "2026-06-06")
        self.assertEqual(period_end.isoformat(), "2026-07-05")
        self.assertEqual([record["date_added"] for record in filtered], ["2026-06-06", "2026-07-05"])

        period_start, period_end = dummy._profit_period_bounds("Calendar Month")
        dummy.profit_period_var = types.SimpleNamespace(get=lambda: "Calendar Month")
        filtered = dummy._filtered_profit_records(rows)

        self.assertEqual(period_start.isoformat(), "2026-07-01")
        self.assertEqual(period_end.isoformat(), "2026-07-05")
        self.assertEqual([record["date_added"] for record in filtered], ["2026-07-05"])
        self.assertEqual(dummy._canonical_profit_period("Month"), "Calendar Month")

    def test_personal_instagram_inventory_plan_separates_post_remove_and_missing_photo(self) -> None:
        class InstagramDummy:
            _instagram_inventory_plan = app.CardPipelineApp._instagram_inventory_plan
            _instagram_inventory_photo_url = app.CardPipelineApp._instagram_inventory_photo_url
            _instagram_inventory_photo_id = app.CardPipelineApp._instagram_inventory_photo_id
            _instagram_post_photo_id = app.CardPipelineApp._instagram_post_photo_id
            _instagram_cover_photo_path = app.CardPipelineApp._instagram_cover_photo_path
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def _load_instagram_inventory_state(self):
                return {
                    "version": 1,
                    "posts": {
                        "old-key": {
                            "status": "posted",
                            "media_id": "179000",
                            "caption": "Sold Card",
                        }
                    },
                }

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token", "public_photo_base_url": "https://example.test/photos"}

            def _instagram_inventory_active_records(self):
                return [
                    {
                        "inventory_key": "new-key",
                        "status": "Active",
                        "card_title": "New Card",
                        "cert_number": "111",
                        "photo_paths": [
                            "[20260708-0923]-Card[10]-[2]-[new card].jpg",
                            "[20260708-0923]-Card[10]-[1]-[new card].jpg",
                        ],
                    },
                    {"inventory_key": "missing-key", "status": "Active", "card_title": "No Photo", "cert_number": "222", "photo_paths": []},
                ]

            def _inventory_photo_paths_for_record(self, record):
                return [Path("/tmp") / value for value in record.get("photo_paths") or []]

            def _inventory_photo_relative_path(self, path):
                return Path(path.name)

        plan = InstagramDummy()._instagram_inventory_plan()

        self.assertEqual(plan["active_count"], 2)
        self.assertEqual(len(plan["to_post"]), 1)
        self.assertEqual(plan["to_post"][0]["caption"], "New Card")
        self.assertEqual(plan["to_post"][0]["photo_url"], "https://example.test/photos/%5B20260708-0923%5D-Card%5B10%5D-%5B1%5D-%5Bnew%20card%5D.jpg")
        self.assertEqual(Path(plan["to_post"][0]["photo_path"]).name, "[20260708-0923]-Card[10]-[1]-[new card].jpg")
        self.assertEqual([Path(path).name for path in plan["to_post"][0]["photo_paths"]], [
            "[20260708-0923]-Card[10]-[1]-[new card].jpg",
            "[20260708-0923]-Card[10]-[2]-[new card].jpg",
        ])
        self.assertEqual(plan["to_post"][0]["photo_urls"], [
            "https://example.test/photos/%5B20260708-0923%5D-Card%5B10%5D-%5B1%5D-%5Bnew%20card%5D.jpg",
            "https://example.test/photos/%5B20260708-0923%5D-Card%5B10%5D-%5B2%5D-%5Bnew%20card%5D.jpg",
        ])
        self.assertEqual(len(plan["to_remove"]), 1)
        self.assertEqual(plan["to_remove"][0]["media_id"], "179000")
        self.assertEqual(len(plan["missing_photos"]), 1)

    def test_personal_instagram_inventory_plan_skips_repost_when_identity_already_posted(self) -> None:
        class InstagramDummy:
            _instagram_inventory_plan = app.CardPipelineApp._instagram_inventory_plan
            _inventory_photo_encoded_id = app.CardPipelineApp._inventory_photo_encoded_id
            _instagram_inventory_photo_url = app.CardPipelineApp._instagram_inventory_photo_url
            _instagram_inventory_photo_id = app.CardPipelineApp._instagram_inventory_photo_id
            _instagram_post_photo_id = app.CardPipelineApp._instagram_post_photo_id
            _instagram_cover_photo_path = app.CardPipelineApp._instagram_cover_photo_path
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def _load_instagram_inventory_state(self):
                return {
                    "version": 1,
                    "posts": {
                        "old-import-key": {
                            "status": "posted",
                            "media_id": "179-old",
                            "caption": "2023 Panini Prizm Victor Wembanyama Silver PSA 10",
                            "inventory_identity": "cert:0019267453",
                        }
                    },
                }

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token", "public_photo_base_url": "https://example.test/photos"}

            def _instagram_inventory_active_records(self):
                return [
                    {
                        "inventory_key": "rebuilt-key",
                        "status": "Active",
                        "card_title": "2023 Panini Prizm Victor Wembanyama Silver PSA 10",
                        "cert_number": "0019267453",
                        "photo_paths": ["front.jpg"],
                    }
                ]

            def _inventory_photo_paths_for_record(self, record):
                return [Path("/tmp") / value for value in record.get("photo_paths") or []]

            def _inventory_photo_relative_path(self, path):
                return Path(path.name)

        plan = InstagramDummy()._instagram_inventory_plan()

        self.assertEqual(plan["posted_count"], 1)
        self.assertEqual(plan["to_post"], [])

    def test_personal_instagram_inventory_plan_does_not_remove_active_identity_under_old_key(self) -> None:
        class InstagramDummy:
            _instagram_inventory_plan = app.CardPipelineApp._instagram_inventory_plan
            _instagram_inventory_photo_url = app.CardPipelineApp._instagram_inventory_photo_url
            _instagram_cover_photo_path = app.CardPipelineApp._instagram_cover_photo_path
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def _load_instagram_inventory_state(self):
                return {
                    "version": 1,
                    "posts": {
                        "old-key": {
                            "status": "posted",
                            "media_id": "179-active",
                            "caption": "2023 Panini Prizm Victor Wembanyama Silver PSA 10",
                            "inventory_identity": "cert:0019267453",
                        }
                    },
                }

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token", "public_photo_base_url": "https://example.test/photos"}

            def _instagram_inventory_active_records(self):
                return [
                    {
                        "inventory_key": "new-key-after-rebuild",
                        "status": "Active",
                        "card_title": "2023 Panini Prizm Victor Wembanyama Silver PSA 10",
                        "cert_number": "0019267453",
                        "photo_paths": ["front.jpg"],
                    }
                ]

            def _inventory_photo_paths_for_record(self, record):
                return [Path("/tmp") / value for value in record.get("photo_paths") or []]

            def _inventory_photo_relative_path(self, path):
                return Path(path.name)

        plan = InstagramDummy()._instagram_inventory_plan()

        self.assertEqual(plan["posted_count"], 1)
        self.assertEqual(plan["to_post"], [])
        self.assertEqual(plan["to_remove"], [])

    def test_personal_instagram_inventory_plan_queues_active_post_when_cover_photo_changes(self) -> None:
        class InstagramDummy:
            _instagram_inventory_plan = app.CardPipelineApp._instagram_inventory_plan
            _inventory_photo_encoded_id = app.CardPipelineApp._inventory_photo_encoded_id
            _instagram_inventory_photo_url = app.CardPipelineApp._instagram_inventory_photo_url
            _instagram_inventory_photo_id = app.CardPipelineApp._instagram_inventory_photo_id
            _instagram_post_photo_id = app.CardPipelineApp._instagram_post_photo_id
            _instagram_cover_photo_path = app.CardPipelineApp._instagram_cover_photo_path
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def _load_instagram_inventory_state(self):
                return {
                    "version": 1,
                    "posts": {
                        "raw-key": {
                            "status": "posted",
                            "media_id": "179-back",
                            "caption": "1957 Topps Johnny Unitas Rookie SGC 3",
                            "inventory_identity": "item:rawkey",
                            "photo_url": f"https://example.test/instagram/media/token/{self._instagram_inventory_photo_id(Path('/tmp/[20260708-0923]-Card[10]-[2]-[unitas].jpg'))}/back.jpg",
                        }
                    },
                }

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token", "public_photo_base_url": ""}

            def _instagram_inventory_active_records(self):
                return [
                    {
                        "inventory_key": "raw-key",
                        "item_id": "RAW-KEY",
                        "status": "Active",
                        "card_title": "1957 Topps Johnny Unitas Rookie SGC 3",
                        "cert_number": "",
                        "photo_paths": [
                            "[20260708-0923]-Card[10]-[2]-[unitas].jpg",
                            "[20260708-0923]-Card[10]-[1]-[unitas].jpg",
                        ],
                    }
                ]

            def _inventory_photo_paths_for_record(self, record):
                return [Path("/tmp") / value for value in record.get("photo_paths") or []]

            def _inventory_photo_relative_path(self, path):
                return Path(path.name)

        plan = InstagramDummy()._instagram_inventory_plan()

        self.assertEqual(plan["posted_count"], 1)
        self.assertEqual(plan["to_post"], [])
        self.assertEqual(len(plan["to_remove"]), 1)
        self.assertEqual(plan["to_remove"][0]["media_id"], "179-back")
        self.assertEqual(plan["to_remove"][0]["reason"], "cover_photo_changed")
        self.assertEqual(plan["to_remove"][0]["expected_photo_id"], InstagramDummy()._instagram_inventory_photo_id(Path("/tmp/[20260708-0923]-Card[10]-[1]-[unitas].jpg")))

    def test_personal_instagram_inventory_plan_ignores_changed_bridge_token_for_same_photo(self) -> None:
        class InstagramDummy:
            _instagram_inventory_plan = app.CardPipelineApp._instagram_inventory_plan
            _inventory_photo_encoded_id = app.CardPipelineApp._inventory_photo_encoded_id
            _instagram_inventory_photo_url = app.CardPipelineApp._instagram_inventory_photo_url
            _instagram_inventory_photo_id = app.CardPipelineApp._instagram_inventory_photo_id
            _instagram_post_photo_id = app.CardPipelineApp._instagram_post_photo_id
            _instagram_cover_photo_path = app.CardPipelineApp._instagram_cover_photo_path
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def _load_instagram_inventory_state(self):
                photo_id = self._instagram_inventory_photo_id(Path("/tmp/[20260708-1009]-Card[6]-[1]-[jordan].jpg"))
                return {
                    "version": 1,
                    "posts": {
                        "raw-key": {
                            "status": "posted",
                            "media_id": "179-front",
                            "caption": "1998 Fleer Ultra Michael Jordan Star Power",
                            "inventory_identity": "item:rawkey",
                            "photo_url": f"https://old-tunnel.trycloudflare.com/instagram/media/old-token/{photo_id}/front.jpg",
                        }
                    },
                }

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token", "public_photo_base_url": ""}

            def _instagram_inventory_active_records(self):
                return [
                    {
                        "inventory_key": "raw-key",
                        "item_id": "RAW-KEY",
                        "status": "Active",
                        "card_title": "1998 Fleer Ultra Michael Jordan Star Power",
                        "cert_number": "",
                        "photo_paths": ["[20260708-1009]-Card[6]-[1]-[jordan].jpg"],
                    }
                ]

            def _inventory_photo_paths_for_record(self, record):
                return [Path("/tmp") / value for value in record.get("photo_paths") or []]

            def _inventory_photo_relative_path(self, path):
                return Path(path.name)

        plan = InstagramDummy()._instagram_inventory_plan()

        self.assertEqual(plan["posted_count"], 1)
        self.assertEqual(plan["to_post"], [])
        self.assertEqual(plan["to_remove"], [])

    def test_company_sheet_week_start_uses_configured_reset_day_and_time(self) -> None:
        before_reset = datetime(2026, 7, 8, 11, 30)
        after_reset = datetime(2026, 7, 8, 12, 30)

        self.assertEqual(
            app.company_sheet_week_start_for_schedule(before_reset, "Wednesday", "12:00").isoformat(),
            "2026-07-01",
        )
        self.assertEqual(
            app.company_sheet_week_start_for_schedule(after_reset, "Wednesday", "12:00").isoformat(),
            "2026-07-08",
        )
        self.assertEqual(
            app.company_sheet_week_start_for_schedule(datetime(2026, 7, 6, 19, 59), "Monday", "8:00 PM").isoformat(),
            "2026-06-29",
        )

    def test_company_sheet_reset_schedules_prefer_assignment_rules(self) -> None:
        class Dummy:
            _company_sheet_reset_schedules = app.CardPipelineApp._company_sheet_reset_schedules

        dummy = Dummy()
        dummy.app_settings = {
            "company_sheet_reset_schedules": {
                "Arena Club": {"weekday": "Monday", "time": "20:00"},
            }
        }
        dummy.assignment_engine = assignment_engine.AssignmentEngine(
            [
                assignment_engine.AssignmentCompany(
                    "Arena Club",
                    assignment_engine.CompanyRules(accept_all=True),
                    [assignment_engine.PayoutTier(rate=0.8)],
                    reset_weekday="Wednesday",
                    reset_time="12:30",
                )
            ]
        )

        schedules = dummy._company_sheet_reset_schedules()

        self.assertEqual(schedules["Arena Club"], {"weekday": "Wednesday", "time": "12:30"})

    def test_append_company_sheet_rows_uses_company_sheet_name_lookup(self) -> None:
        row = WorkbookRow(
            excel_row=2,
            cert_number="1234567890",
            card_title="Test Card",
            grader="PSA",
            existing_value=10,
            best_company="Fanatics",
            estimated_payout=15,
        )
        with TemporaryDirectory() as tmpdir:
            result = append_company_sheet_rows(
                Path(tmpdir),
                [row],
                sheet_name_lookup={"Fanatics": "Week of 2026-07-08"},
            )
            self.assertEqual(result["rows_added"], 1)
            self.assertIn("Week of 2026-07-08", result["added_records"][0]["weekly_sheet_name"])

    def test_personal_instagram_inventory_photo_url_can_use_bridge_media_route(self) -> None:
        class InstagramDummy:
            _instagram_inventory_photo_url = app.CardPipelineApp._instagram_inventory_photo_url
            _instagram_inventory_photo_id = app.CardPipelineApp._instagram_inventory_photo_id
            _inventory_photo_encoded_id = app.CardPipelineApp._inventory_photo_encoded_id

            def __init__(self):
                self.state = app.BridgeState()

            def _inventory_photo_storage_value(self, path):
                return path.name

        dummy = InstagramDummy()
        url = dummy._instagram_inventory_photo_url(
            Path("/tmp/front photo.jpg"),
            {"public_bridge_url": "https://lucas-example.trycloudflare.com", "public_photo_base_url": ""},
        )

        self.assertTrue(url.startswith(f"https://lucas-example.trycloudflare.com/instagram/media/{dummy.state.instagram_media_token}/"))
        self.assertTrue(url.endswith("/front-photo.jpg"))

    def test_instagram_env_config_prefers_background_tunnel_url(self) -> None:
        class InstagramDummy:
            _instagram_env_config = app.CardPipelineApp._instagram_env_config

            def _instagram_background_tunnel_enabled(self):
                return True

            def _ensure_instagram_background_tunnel(self):
                return "https://hidden-background.trycloudflare.com"

        dummy = InstagramDummy()
        with patch.dict(
            os.environ,
            {
                "LUCAS_INSTAGRAM_USER_ID": "178",
                "LUCAS_INSTAGRAM_ACCESS_TOKEN": "token",
                "LUCAS_INSTAGRAM_PUBLIC_BRIDGE_URL": "https://manual-url.trycloudflare.com",
            },
            clear=False,
        ):
            config = dummy._instagram_env_config()

        self.assertEqual(config["public_bridge_url"], "https://hidden-background.trycloudflare.com")

    def test_instagram_bridge_media_requires_token(self) -> None:
        state = app.BridgeState()
        state.instagram_media_resolver = lambda photo_id: (b"jpg-bytes", "image/jpeg") if photo_id == "abc" else None

        self.assertIsNone(state.get_instagram_media("bad-token", "abc"))
        self.assertEqual(state.get_instagram_media(state.instagram_media_token, "abc"), (b"jpg-bytes", "image/jpeg"))

    def test_mobile_inventory_photo_payload_and_media_response(self) -> None:
        class MobilePhotoDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _mobile_inventory_json_record = app.CardPipelineApp._mobile_inventory_json_record
            _mobile_inventory_photo_items = app.CardPipelineApp._mobile_inventory_photo_items
            _inventory_photo_source_folder = app.CardPipelineApp._inventory_photo_source_folder
            _inventory_photo_shared_folder = app.CardPipelineApp._inventory_photo_shared_folder
            _inventory_photo_relative_path = app.CardPipelineApp._inventory_photo_relative_path
            _inventory_photo_storage_value = app.CardPipelineApp._inventory_photo_storage_value
            _inventory_photo_encoded_id = app.CardPipelineApp._inventory_photo_encoded_id
            _inventory_photo_path_candidates = app.CardPipelineApp._inventory_photo_path_candidates
            _inventory_photo_safe_candidates = app.CardPipelineApp._inventory_photo_safe_candidates
            _safe_inventory_photo_path = app.CardPipelineApp._safe_inventory_photo_path
            mobile_inventory_photo_response = app.CardPipelineApp.mobile_inventory_photo_response
            _inventory_photo_media_response = app.CardPipelineApp._inventory_photo_media_response

        with TemporaryDirectory() as tmp:
            old_photo_dir = app.INVENTORY_PHOTOS_DIR
            app.INVENTORY_PHOTOS_DIR = Path(tmp) / "source"
            app.INVENTORY_PHOTOS_DIR.mkdir(parents=True)
            photo = app.INVENTORY_PHOTOS_DIR / "front photo.jpg"
            photo.write_bytes(b"jpg-bytes")
            try:
                dummy = MobilePhotoDummy()
                dummy.state = app.BridgeState()
                dummy.mobile_pin = "123456"
                record = dummy._mobile_inventory_json_record(
                    {
                        "assigned_person": "Kevin",
                        "cert_number": "123",
                        "card_title": "Test Card",
                        "status": "Active",
                        "photo_paths": [str(photo)],
                    }
                )

                self.assertEqual(record["photo_count"], 1)
                self.assertIn("/mobile/api/inventory/photo/", record["photos"][0]["url"])
                self.assertIn("pin=123456", record["photos"][0]["url"])
                self.assertEqual(dummy.mobile_inventory_photo_response(record["photos"][0]["id"]), (b"jpg-bytes", "image/jpeg"))
            finally:
                app.INVENTORY_PHOTOS_DIR = old_photo_dir

    def test_mobile_inventory_photo_requires_pin(self) -> None:
        state = app.BridgeState()
        state.mobile_pin_provider = lambda: "123456"
        state.mobile_inventory_photo_resolver = lambda photo_id: (b"jpg-bytes", "image/jpeg") if photo_id == "abc" else None

        self.assertIsNone(state.get_mobile_inventory_photo(None, {"pin": ["bad"]}, "abc"))
        self.assertEqual(state.get_mobile_inventory_photo(None, {"pin": ["123456"]}, "abc"), (b"jpg-bytes", "image/jpeg"))

    def test_instagram_publish_retries_until_media_ready(self) -> None:
        class InstagramDummy:
            _instagram_publish_media_with_retry = app.CardPipelineApp._instagram_publish_media_with_retry

            def __init__(self):
                self.calls = 0
                self.events = queue.Queue()

            def _instagram_api_json(self, endpoint, params=None, method="GET"):
                self.calls += 1
                if self.calls == 1:
                    raise RuntimeError("Media ID is not available: media is not ready for publishing")
                return {"id": "179123"}

        dummy = InstagramDummy()
        with patch.object(app.time, "sleep", lambda _seconds: None):
            result = dummy._instagram_publish_media_with_retry("178", "creation", "Test Card")

        self.assertEqual(result["id"], "179123")
        self.assertEqual(dummy.calls, 2)

    def test_instagram_inventory_sync_queues_inactive_posts_for_manual_delete(self) -> None:
        class InstagramDummy:
            _instagram_inventory_sync_worker = app.CardPipelineApp._instagram_inventory_sync_worker
            _record_instagram_removed_post = app.CardPipelineApp._record_instagram_removed_post
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def __init__(self):
                self.state = {
                    "version": 1,
                    "posts": {
                        "sold-key": {
                            "status": "posted",
                            "media_id": "179-sold",
                            "caption": "Sold Card",
                            "permalink": "https://instagram.test/p/sold",
                        }
                    },
                }
                self.events = queue.Queue()
                self.activities = []
                self.deleted = []

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_api_json(self, endpoint, params=None, method="GET"):
                self.deleted.append((endpoint, method))
                return {"success": True}

            def _instagram_inventory_active_records(self):
                return []

            def _append_activity(self, action, summary, details):
                self.activities.append((action, summary, details))

        dummy = InstagramDummy()
        dummy._instagram_inventory_sync_worker(
            {
                "config": {"user_id": "178"},
                "to_post": [],
                "to_remove": [
                    {
                        "inventory_key": "sold-key",
                        "media_id": "179-sold",
                        "caption": "Sold Card",
                        "permalink": "https://instagram.test/p/sold",
                    }
                ],
            }
        )

        self.assertEqual(dummy.deleted, [])
        self.assertEqual(dummy.state["posts"]["sold-key"]["status"], "delete_review_needed")
        self.assertIn("delete_queued_at", dummy.state["posts"]["sold-key"])
        self.assertNotIn("removed_posts", dummy.state)
        self.assertIn("queued 1", dummy.activities[-1][1])

    def test_instagram_inventory_sync_does_not_call_delete_api_for_review_rows(self) -> None:
        class InstagramDummy:
            _instagram_inventory_sync_worker = app.CardPipelineApp._instagram_inventory_sync_worker
            _record_instagram_removed_post = app.CardPipelineApp._record_instagram_removed_post
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def __init__(self):
                self.state = {
                    "version": 1,
                    "posts": {
                        "old-key": {
                            "status": "posted",
                            "media_id": "179-old",
                            "caption": "Old Card",
                        }
                    },
                }
                self.events = queue.Queue()
                self.activities = []
                self.deleted = []

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_api_json(self, endpoint, params=None, method="GET"):
                self.deleted.append((endpoint, method))
                raise RuntimeError("Unsupported delete request")

            def _instagram_inventory_active_records(self):
                return []

            def _append_activity(self, action, summary, details):
                self.activities.append((action, summary, details))

        dummy = InstagramDummy()
        dummy._instagram_inventory_sync_worker(
            {
                "config": {"user_id": "178"},
                "to_post": [],
                "to_remove": [{"inventory_key": "old-key", "media_id": "179-old", "caption": "Old Card"}],
            }
        )

        entry = dummy.state["posts"]["old-key"]
        self.assertEqual(dummy.deleted, [])
        self.assertEqual(entry["status"], "delete_review_needed")
        self.assertNotIn("delete_error", entry)
        self.assertIn("queued 1", dummy.activities[-1][1])

    def test_instagram_manual_delete_clears_tracked_post(self) -> None:
        class InstagramDummy:
            _mark_instagram_posts_manually_deleted = app.CardPipelineApp._mark_instagram_posts_manually_deleted
            _record_instagram_removed_post = app.CardPipelineApp._record_instagram_removed_post

            def __init__(self):
                self.state = {
                    "version": 1,
                    "posts": {
                        "old-key": {
                            "status": "delete_review_needed",
                            "media_id": "179-old",
                            "caption": "Old Card",
                            "permalink": "https://instagram.test/p/old",
                        }
                    },
                    "duplicate_posts": [{"inventory_key": "old-key", "media_id": "179-old"}],
                }
                self.activities = []

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _append_activity(self, action, summary, details=None):
                self.activities.append((action, summary, details))

        dummy = InstagramDummy()
        marked = dummy._mark_instagram_posts_manually_deleted([
            {"inventory_key": "old-key", "media_id": "179-old", "caption": "Old Card"}
        ])

        self.assertEqual(marked, 1)
        self.assertEqual(dummy.state["posts"], {})
        self.assertEqual(dummy.state["duplicate_posts"], [])
        self.assertEqual(dummy.state["removed_posts"][0]["media_id"], "179-old")
        self.assertEqual(dummy.state["removed_posts"][0]["reason"], "manual_delete_confirmed")
        self.assertEqual(dummy.activities[-1][0], "Instagram Manual Delete")

    def test_instagram_inventory_sync_skips_post_when_preview_item_is_no_longer_active(self) -> None:
        class InstagramDummy:
            _instagram_inventory_sync_worker = app.CardPipelineApp._instagram_inventory_sync_worker
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def __init__(self):
                self.state = {"version": 1, "posts": {}}
                self.events = queue.Queue()
                self.activities = []
                self.api_calls = []

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_inventory_active_records(self):
                return []

            def _instagram_api_json(self, endpoint, params=None, method="GET"):
                self.api_calls.append((endpoint, method, params))
                return {"id": "should-not-happen"}

            def _instagram_publish_media_with_retry(self, user_id, creation_id, caption):
                self.api_calls.append(("publish", "POST", {"creation_id": creation_id}))
                return {"id": "should-not-happen"}

            def _append_activity(self, action, summary, details):
                self.activities.append((action, summary, details))

        dummy = InstagramDummy()
        dummy._instagram_inventory_sync_worker(
            {
                "config": {"user_id": "178"},
                "to_post": [
                    {
                        "inventory_key": "sold-before-sync",
                        "record": {"inventory_key": "sold-before-sync", "card_title": "Sold Card", "cert_number": "12345"},
                        "caption": "Sold Card",
                        "photo_url": "https://example.test/sold.jpg",
                    }
                ],
                "to_remove": [],
            }
        )

        self.assertEqual(dummy.api_calls, [])
        self.assertEqual(dummy.state["posts"], {})
        self.assertIn("posted 0", dummy.activities[-1][1])

    def test_instagram_inventory_sync_stores_permalink_after_post(self) -> None:
        class InstagramDummy:
            _instagram_inventory_sync_worker = app.CardPipelineApp._instagram_inventory_sync_worker
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def __init__(self):
                self.state = {"version": 1, "posts": {}}
                self.events = queue.Queue()
                self.activities = []

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_inventory_active_records(self):
                return [{"inventory_key": "ready-key", "status": "Active", "card_title": "Ready Card", "item_id": "RAW-1"}]

            def _instagram_api_json(self, endpoint, params=None, method="GET"):
                if method == "POST":
                    return {"id": "creation-1"}
                if endpoint == "media-1":
                    return {"permalink": "https://instagram.test/p/ready"}
                return {}

            def _instagram_publish_media_with_retry(self, user_id, creation_id, caption):
                return {"id": "media-1"}

            def _instagram_inventory_photo_id(self, path):
                return "photo-1"

            def _append_activity(self, action, summary, details):
                self.activities.append((action, summary, details))

        dummy = InstagramDummy()
        dummy._instagram_inventory_sync_worker(
            {
                "config": {"user_id": "178"},
                "to_post": [
                    {
                        "inventory_key": "ready-key",
                        "record": {"inventory_key": "ready-key", "status": "Active", "card_title": "Ready Card", "item_id": "RAW-1"},
                        "caption": "Ready Card",
                        "photo_url": "https://example.test/ready.jpg",
                    }
                ],
                "to_remove": [],
            }
        )

        self.assertEqual(dummy.state["posts"]["ready-key"]["media_id"], "media-1")
        self.assertEqual(dummy.state["posts"]["ready-key"]["permalink"], "https://instagram.test/p/ready")

    def test_instagram_inventory_sync_posts_multiple_photos_as_carousel(self) -> None:
        class InstagramDummy:
            _instagram_inventory_sync_worker = app.CardPipelineApp._instagram_inventory_sync_worker
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def __init__(self):
                self.state = {"version": 1, "posts": {}}
                self.events = queue.Queue()
                self.activities = []
                self.api_calls = []

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_inventory_active_records(self):
                return [{"inventory_key": "raw-key", "status": "Active", "card_title": "Front Back Card", "item_id": "RAW-1"}]

            def _instagram_api_json(self, endpoint, params=None, method="GET"):
                self.api_calls.append((endpoint, method, dict(params or {})))
                if method == "POST" and params and params.get("is_carousel_item"):
                    return {"id": f"child-{len([call for call in self.api_calls if call[2].get('is_carousel_item')])}"}
                if method == "POST" and params and params.get("media_type") == "CAROUSEL":
                    return {"id": "carousel-parent"}
                if endpoint == "media-carousel":
                    return {"permalink": "https://instagram.test/p/carousel"}
                return {}

            def _instagram_publish_media_with_retry(self, user_id, creation_id, caption):
                self.api_calls.append(("publish", "POST", {"creation_id": creation_id}))
                return {"id": "media-carousel"}

            def _instagram_inventory_photo_id(self, path):
                return Path(str(path)).stem

            def _append_activity(self, action, summary, details):
                self.activities.append((action, summary, details))

        dummy = InstagramDummy()
        dummy._instagram_inventory_sync_worker(
            {
                "config": {"user_id": "178"},
                "to_post": [
                    {
                        "inventory_key": "raw-key",
                        "record": {"inventory_key": "raw-key", "status": "Active", "card_title": "Front Back Card", "item_id": "RAW-1"},
                        "caption": "Front Back Card",
                        "photo_path": Path("/tmp/front.jpg"),
                        "photo_paths": [Path("/tmp/front.jpg"), Path("/tmp/back.jpg")],
                        "photo_url": "https://example.test/front.jpg",
                        "photo_urls": ["https://example.test/front.jpg", "https://example.test/back.jpg"],
                    }
                ],
                "to_remove": [],
            }
        )

        post_calls = [call for call in dummy.api_calls if call[0] == "178/media"]
        self.assertEqual(post_calls[0][2], {"image_url": "https://example.test/front.jpg", "is_carousel_item": "true"})
        self.assertEqual(post_calls[1][2], {"image_url": "https://example.test/back.jpg", "is_carousel_item": "true"})
        self.assertEqual(post_calls[2][2], {"media_type": "CAROUSEL", "children": "child-1,child-2", "caption": "Front Back Card"})
        self.assertIn(("publish", "POST", {"creation_id": "carousel-parent"}), dummy.api_calls)
        posted = dummy.state["posts"]["raw-key"]
        self.assertEqual(posted["media_type"], "CAROUSEL")
        self.assertEqual(posted["photo_urls"], ["https://example.test/front.jpg", "https://example.test/back.jpg"])
        self.assertEqual(posted["child_creation_ids"], ["child-1", "child-2"])
        self.assertEqual(posted["permalink"], "https://instagram.test/p/carousel")

    def test_instagram_ready_photo_urls_drops_blank_and_duplicate_urls(self) -> None:
        urls = app.instagram_ready_photo_urls(
            {
                "photo_url": "https://example.test/front.jpg",
                "photo_urls": [
                    "",
                    "https://example.test/front.jpg",
                    " https://example.test/front.jpg ",
                    "https://example.test/back.jpg",
                ],
            }
        )

        self.assertEqual(urls, ["https://example.test/front.jpg", "https://example.test/back.jpg"])

    def test_instagram_inventory_sync_continues_after_single_post_error(self) -> None:
        class InstagramDummy:
            _instagram_inventory_sync_worker = app.CardPipelineApp._instagram_inventory_sync_worker
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def __init__(self):
                self.state = {"version": 1, "posts": {}}
                self.events = queue.Queue()
                self.activities = []

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_inventory_active_records(self):
                return [
                    {"inventory_key": "bad-key", "status": "Active", "card_title": "Bad Card", "item_id": "RAW-BAD"},
                    {"inventory_key": "good-key", "status": "Active", "card_title": "Good Card", "item_id": "RAW-GOOD"},
                ]

            def _instagram_api_json(self, endpoint, params=None, method="GET"):
                if method == "POST" and params and params.get("caption") == "Bad Card":
                    raise RuntimeError("image fetch failed")
                if method == "POST":
                    return {"id": "creation-good"}
                if endpoint == "media-good":
                    return {"permalink": "https://instagram.test/p/good"}
                return {}

            def _instagram_publish_media_with_retry(self, user_id, creation_id, caption):
                return {"id": "media-good"}

            def _instagram_inventory_photo_id(self, path):
                return "photo-good"

            def _append_activity(self, action, summary, details):
                self.activities.append((action, summary, details))

        dummy = InstagramDummy()
        dummy._instagram_inventory_sync_worker(
            {
                "config": {"user_id": "178"},
                "to_post": [
                    {
                        "inventory_key": "bad-key",
                        "record": {"inventory_key": "bad-key", "status": "Active", "card_title": "Bad Card", "item_id": "RAW-BAD"},
                        "caption": "Bad Card",
                        "photo_url": "https://example.test/bad.jpg",
                    },
                    {
                        "inventory_key": "good-key",
                        "record": {"inventory_key": "good-key", "status": "Active", "card_title": "Good Card", "item_id": "RAW-GOOD"},
                        "caption": "Good Card",
                        "photo_url": "https://example.test/good.jpg",
                    },
                ],
                "to_remove": [],
            }
        )

        self.assertEqual(dummy.state["posts"]["bad-key"]["status"], "post_error")
        self.assertIn("image fetch failed", dummy.state["posts"]["bad-key"]["post_error"])
        self.assertEqual(dummy.state["posts"]["good-key"]["media_id"], "media-good")
        self.assertIn("posted 1", dummy.activities[-1][1])
        self.assertIn("errors 1", dummy.activities[-1][1])

    def test_instagram_inventory_sync_does_not_delete_active_identity_from_stale_remove_plan(self) -> None:
        class InstagramDummy:
            _instagram_inventory_sync_worker = app.CardPipelineApp._instagram_inventory_sync_worker
            _instagram_delete_media_post = app.CardPipelineApp._instagram_delete_media_post
            _record_instagram_removed_post = app.CardPipelineApp._record_instagram_removed_post
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_post_entry_identity = app.CardPipelineApp._instagram_post_entry_identity
            _instagram_active_identity_map = app.CardPipelineApp._instagram_active_identity_map

            def __init__(self):
                self.state = {
                    "version": 1,
                    "posts": {
                        "old-key": {
                            "status": "posted",
                            "media_id": "179-active",
                            "caption": "Active Card",
                            "inventory_identity": "cert:12345",
                        }
                    },
                }
                self.events = queue.Queue()
                self.activities = []
                self.deleted = []

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_inventory_active_records(self):
                return [{"inventory_key": "new-key", "status": "Active", "card_title": "Active Card", "cert_number": "12345"}]

            def _instagram_api_json(self, endpoint, params=None, method="GET"):
                self.deleted.append((endpoint, method))
                return {"success": True}

            def _append_activity(self, action, summary, details):
                self.activities.append((action, summary, details))

        dummy = InstagramDummy()
        dummy._instagram_inventory_sync_worker(
            {
                "config": {"user_id": "178"},
                "to_post": [],
                "to_remove": [
                    {
                        "inventory_key": "old-key",
                        "media_id": "179-active",
                        "caption": "Active Card",
                        "inventory_identity": "cert:12345",
                    }
                ],
            }
        )

        self.assertEqual(dummy.deleted, [])
        self.assertEqual(dummy.state["posts"]["old-key"]["media_id"], "179-active")
        self.assertIn("removed 0", dummy.activities[-1][1])

    def test_instagram_auto_sync_runs_ready_posts_and_removals_once_daily(self) -> None:
        class ImmediateThread:
            def __init__(self, target, args=(), daemon=None):
                self.target = target
                self.args = args
                self.daemon = daemon

            def start(self):
                self.target(*self.args)

        class InstagramDummy:
            _instagram_auto_sync_due = app.CardPipelineApp._instagram_auto_sync_due
            _mark_instagram_auto_sync_completed = app.CardPipelineApp._mark_instagram_auto_sync_completed
            _run_instagram_auto_sync_if_due = app.CardPipelineApp._run_instagram_auto_sync_if_due
            _instagram_auto_sync_worker = app.CardPipelineApp._instagram_auto_sync_worker
            _instagram_daily_post_limit = app.CardPipelineApp._instagram_daily_post_limit
            _instagram_content_publishing_limit = lambda self, config=None: {"limit": 100, "quota_usage": 0, "remaining": 100, "error": ""}

            def __init__(self):
                self.state = {"version": 1, "posts": {}}
                self.instagram_auto_sync_running = False
                self.events = queue.Queue()
                self.synced_plan = None

            def _personal_instagram_sync_enabled(self):
                return True

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token"}

            def _instagram_inventory_plan(self):
                return {
                    "config": {"user_id": "178", "access_token": "token"},
                    "to_post": [
                        {"inventory_key": "ready", "caption": "Ready Card", "photo_url": "https://example.test/ready.jpg"},
                        {"inventory_key": "waiting", "caption": "Waiting Card", "photo_url": ""},
                    ],
                    "to_remove": [{"inventory_key": "sold", "media_id": "179-sold", "caption": "Sold Card"}],
                    "missing_public_urls": [],
                }

            def _instagram_inventory_sync_worker(self, plan):
                self.synced_plan = plan

        dummy = InstagramDummy()
        with patch.object(app.threading, "Thread", ImmediateThread):
            self.assertTrue(dummy._run_instagram_auto_sync_if_due())

        self.assertFalse(dummy.instagram_auto_sync_running)
        self.assertEqual([item["inventory_key"] for item in dummy.synced_plan["to_post"]], ["ready"])
        self.assertEqual([item["inventory_key"] for item in dummy.synced_plan["to_remove"]], ["sold"])
        self.assertEqual(dummy.state["last_auto_sync_date"], datetime.now().date().isoformat())
        self.assertIn("posted=1", dummy.state["last_auto_sync_summary"])

    def test_instagram_auto_sync_caps_daily_posts_at_configured_limit(self) -> None:
        class ImmediateThread:
            def __init__(self, target, args=(), daemon=None):
                self.target = target
                self.args = args
                self.daemon = daemon

            def start(self):
                self.target(*self.args)

        class InstagramDummy:
            _instagram_auto_sync_due = app.CardPipelineApp._instagram_auto_sync_due
            _mark_instagram_auto_sync_completed = app.CardPipelineApp._mark_instagram_auto_sync_completed
            _run_instagram_auto_sync_if_due = app.CardPipelineApp._run_instagram_auto_sync_if_due
            _instagram_auto_sync_worker = app.CardPipelineApp._instagram_auto_sync_worker
            _instagram_daily_post_limit = app.CardPipelineApp._instagram_daily_post_limit
            _instagram_content_publishing_limit = lambda self, config=None: {"limit": 100, "quota_usage": 0, "remaining": 100, "error": ""}

            def __init__(self):
                self.state = {"version": 1, "posts": {}}
                self.instagram_auto_sync_running = False
                self.events = queue.Queue()
                self.synced_plan = None

            def _personal_instagram_sync_enabled(self):
                return True

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token", "daily_post_limit": "75"}

            def _instagram_inventory_plan(self):
                return {
                    "config": {"user_id": "178", "access_token": "token", "daily_post_limit": "75"},
                    "to_post": [
                        {"inventory_key": f"ready-{index}", "caption": f"Ready Card {index}", "photo_url": f"https://example.test/{index}.jpg"}
                        for index in range(80)
                    ],
                    "to_remove": [],
                    "missing_public_urls": [],
                }

            def _instagram_inventory_sync_worker(self, plan):
                self.synced_plan = plan

        dummy = InstagramDummy()
        with patch.object(app.threading, "Thread", ImmediateThread):
            self.assertTrue(dummy._run_instagram_auto_sync_if_due())

        self.assertEqual(len(dummy.synced_plan["to_post"]), 75)
        self.assertEqual(dummy.synced_plan["to_post"][0]["inventory_key"], "ready-0")
        self.assertEqual(dummy.synced_plan["to_post"][-1]["inventory_key"], "ready-74")
        self.assertIn("daily_limit=75", dummy.state["last_auto_sync_summary"])

    def test_instagram_auto_sync_caps_daily_posts_at_meta_remaining_quota(self) -> None:
        class ImmediateThread:
            def __init__(self, target, args=(), daemon=None):
                self.target = target
                self.args = args
                self.daemon = daemon

            def start(self):
                self.target(*self.args)

        class InstagramDummy:
            _instagram_auto_sync_due = app.CardPipelineApp._instagram_auto_sync_due
            _mark_instagram_auto_sync_completed = app.CardPipelineApp._mark_instagram_auto_sync_completed
            _run_instagram_auto_sync_if_due = app.CardPipelineApp._run_instagram_auto_sync_if_due
            _instagram_auto_sync_worker = app.CardPipelineApp._instagram_auto_sync_worker
            _instagram_daily_post_limit = app.CardPipelineApp._instagram_daily_post_limit
            _instagram_content_publishing_limit = lambda self, config=None: {"limit": 100, "quota_usage": 96, "remaining": 4, "error": ""}

            def __init__(self):
                self.state = {"version": 1, "posts": {}}
                self.instagram_auto_sync_running = False
                self.events = queue.Queue()
                self.synced_plan = None

            def _personal_instagram_sync_enabled(self):
                return True

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token", "daily_post_limit": "75"}

            def _instagram_inventory_plan(self):
                return {
                    "config": {"user_id": "178", "access_token": "token", "daily_post_limit": "75"},
                    "to_post": [
                        {"inventory_key": f"ready-{index}", "caption": f"Ready Card {index}", "photo_url": f"https://example.test/{index}.jpg"}
                        for index in range(20)
                    ],
                    "to_remove": [],
                    "missing_public_urls": [],
                }

            def _instagram_inventory_sync_worker(self, plan):
                self.synced_plan = plan

        dummy = InstagramDummy()
        with patch.object(app.threading, "Thread", ImmediateThread):
            self.assertTrue(dummy._run_instagram_auto_sync_if_due())

        self.assertEqual(len(dummy.synced_plan["to_post"]), 4)
        self.assertEqual(dummy.synced_plan["to_post"][-1]["inventory_key"], "ready-3")

    def test_instagram_manual_sync_plan_caps_ready_posts_at_configured_limit(self) -> None:
        class InstagramDummy:
            _instagram_daily_post_limit = app.CardPipelineApp._instagram_daily_post_limit
            _instagram_limited_manual_sync_plan = app.CardPipelineApp._instagram_limited_manual_sync_plan

        plan = {
            "config": {"daily_post_limit": "3"},
            "to_post": [
                {"inventory_key": f"ready-{index}", "caption": f"Ready Card {index}", "photo_url": f"https://example.test/{index}.jpg"}
                for index in range(5)
            ]
            + [{"inventory_key": "missing-url", "caption": "Missing URL", "photo_url": ""}],
            "to_remove": [{"inventory_key": "sold", "media_id": "179-sold", "caption": "Sold Card"}],
            "missing_public_urls": [{"inventory_key": "missing-url"}],
        }

        limited_plan, ready_total, limit = InstagramDummy()._instagram_limited_manual_sync_plan(plan)

        self.assertEqual(limit, 3)
        self.assertEqual(ready_total, 5)
        self.assertEqual([item["inventory_key"] for item in limited_plan["to_post"]], ["ready-0", "ready-1", "ready-2"])
        self.assertEqual([item["inventory_key"] for item in limited_plan["to_remove"]], ["sold"])
        self.assertEqual(limited_plan["missing_public_urls"], [])

    def test_instagram_meta_publish_limit_defaults_to_stricter_graph_limit(self) -> None:
        class InstagramDummy:
            _instagram_meta_publish_limit = app.CardPipelineApp._instagram_meta_publish_limit

        with patch.dict(app.os.environ, {}, clear=True):
            self.assertEqual(InstagramDummy()._instagram_meta_publish_limit(), 50)
        with patch.dict(app.os.environ, {"LUCAS_INSTAGRAM_META_POST_LIMIT": "100"}, clear=True):
            self.assertEqual(InstagramDummy()._instagram_meta_publish_limit(), 100)

    def test_instagram_manual_sync_plan_caps_ready_posts_at_meta_remaining_quota(self) -> None:
        class InstagramDummy:
            _instagram_daily_post_limit = app.CardPipelineApp._instagram_daily_post_limit
            _instagram_limited_manual_sync_plan = app.CardPipelineApp._instagram_limited_manual_sync_plan

        plan = {
            "config": {"daily_post_limit": "75"},
            "to_post": [
                {"inventory_key": f"ready-{index}", "caption": f"Ready Card {index}", "photo_url": f"https://example.test/{index}.jpg"}
                for index in range(20)
            ],
            "to_remove": [{"inventory_key": "sold", "media_id": "179-sold", "caption": "Sold Card"}],
            "missing_public_urls": [],
        }

        limited_plan, ready_total, limit = InstagramDummy()._instagram_limited_manual_sync_plan(plan, meta_remaining=4)

        self.assertEqual(limit, 4)
        self.assertEqual(ready_total, 20)
        self.assertEqual([item["inventory_key"] for item in limited_plan["to_post"]], ["ready-0", "ready-1", "ready-2", "ready-3"])
        self.assertEqual([item["inventory_key"] for item in limited_plan["to_remove"]], ["sold"])

    def test_instagram_auto_sync_due_skips_after_daily_completion(self) -> None:
        class InstagramDummy:
            _instagram_auto_sync_due = app.CardPipelineApp._instagram_auto_sync_due

            def __init__(self):
                self.instagram_auto_sync_running = False
                self.state = {"last_auto_sync_date": "2026-07-07"}

            def _personal_instagram_sync_enabled(self):
                return True

            def _load_instagram_inventory_state(self):
                return self.state

        dummy = InstagramDummy()
        self.assertFalse(dummy._instagram_auto_sync_due(datetime(2026, 7, 7, 9, 30)))
        self.assertTrue(dummy._instagram_auto_sync_due(datetime(2026, 7, 8, 9, 30)))

    def test_instagram_import_existing_posts_matches_cert_and_leaves_unclear_unmatched(self) -> None:
        class InstagramDummy:
            _instagram_import_existing_posts = app.CardPipelineApp._instagram_import_existing_posts
            _instagram_find_inventory_match_for_post = app.CardPipelineApp._instagram_find_inventory_match_for_post
            _instagram_match_text_tokens = app.CardPipelineApp._instagram_match_text_tokens
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_record_duplicate_post = app.CardPipelineApp._instagram_record_duplicate_post
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_record_duplicate_post = app.CardPipelineApp._instagram_record_duplicate_post
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_record_duplicate_post = app.CardPipelineApp._instagram_record_duplicate_post

            def __init__(self):
                self.state = {"version": 1, "posts": {}}
                self.activities = []

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token"}

            def _instagram_inventory_active_records(self):
                return [
                    {
                        "inventory_key": "cert-key",
                        "status": "Active",
                        "card_title": "2023 Panini Prizm Victor Wembanyama Silver PSA 10",
                        "cert_number": "0019267453",
                    },
                    {
                        "inventory_key": "raw-key",
                        "status": "Active",
                        "card_title": "Kobe Bryant Purple Wave",
                        "item_id": "RAW-20260706-0001",
                    },
                ]

            def _instagram_existing_media_posts(self, config, limit=500):
                return [
                    {
                        "id": "179-cert",
                        "caption": "2023 Panini Prizm Victor Wembanyama Silver PSA 10 cert 0019267453",
                        "permalink": "https://instagram.test/p/cert",
                        "timestamp": "2026-07-01T00:00:00+0000",
                    },
                    {
                        "id": "179-unclear",
                        "caption": "fresh raw card available",
                        "permalink": "https://instagram.test/p/raw",
                    },
                ]

            def _append_activity(self, action, summary, details):
                self.activities.append((action, summary, details))

        dummy = InstagramDummy()
        result = dummy._instagram_import_existing_posts(use_ocr=False)

        self.assertEqual(result["imported"], 1)
        self.assertEqual(len(result["unmatched"]), 1)
        self.assertIn("cert-key", dummy.state["posts"])
        self.assertEqual(dummy.state["posts"]["cert-key"]["media_id"], "179-cert")
        self.assertEqual(dummy.state["posts"]["cert-key"]["matched_by"], "cert_number")
        self.assertNotIn("raw-key", dummy.state["posts"])

    def test_instagram_import_existing_posts_queues_duplicate_media_for_removal(self) -> None:
        class InstagramDummy:
            _instagram_import_existing_posts = app.CardPipelineApp._instagram_import_existing_posts
            _instagram_find_inventory_match_for_post = app.CardPipelineApp._instagram_find_inventory_match_for_post
            _instagram_match_text_tokens = app.CardPipelineApp._instagram_match_text_tokens
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_record_duplicate_post = app.CardPipelineApp._instagram_record_duplicate_post
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_record_duplicate_post = app.CardPipelineApp._instagram_record_duplicate_post
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_record_duplicate_post = app.CardPipelineApp._instagram_record_duplicate_post

            def __init__(self):
                self.state = {"version": 1, "posts": {}}

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token"}

            def _instagram_inventory_active_records(self):
                return [
                    {
                        "inventory_key": "cert-key",
                        "status": "Active",
                        "card_title": "2023 Panini Prizm Victor Wembanyama Silver PSA 10",
                        "cert_number": "0019267453",
                    }
                ]

            def _instagram_existing_media_posts(self, config, limit=500):
                return [
                    {
                        "id": "179-keep",
                        "caption": "2023 Panini Prizm Victor Wembanyama Silver PSA 10 cert 0019267453",
                        "permalink": "https://instagram.test/p/keep",
                    },
                    {
                        "id": "179-duplicate",
                        "caption": "duplicate 2023 Panini Prizm Victor Wembanyama Silver PSA 10 cert 0019267453",
                        "permalink": "https://instagram.test/p/dupe",
                    },
                ]

            def _append_activity(self, action, summary, details):
                pass

        dummy = InstagramDummy()
        result = dummy._instagram_import_existing_posts(use_ocr=False)

        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["duplicates_found"], 1)
        self.assertEqual(dummy.state["posts"]["cert-key"]["media_id"], "179-keep")
        self.assertEqual(dummy.state["duplicate_posts"][0]["media_id"], "179-duplicate")

    def test_instagram_import_existing_posts_can_match_strong_title_overlap(self) -> None:
        class InstagramDummy:
            _instagram_import_existing_posts = app.CardPipelineApp._instagram_import_existing_posts
            _instagram_find_inventory_match_for_post = app.CardPipelineApp._instagram_find_inventory_match_for_post
            _instagram_match_text_tokens = app.CardPipelineApp._instagram_match_text_tokens
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_record_duplicate_post = app.CardPipelineApp._instagram_record_duplicate_post

            def __init__(self):
                self.state = {"version": 1, "posts": {}}

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token"}

            def _instagram_inventory_active_records(self):
                return [
                    {
                        "inventory_key": "title-key",
                        "status": "Active",
                        "card_title": "2024 Topps Chrome Caitlin Clark Blue Refractor PSA 9",
                        "cert_number": "",
                    }
                ]

            def _instagram_existing_media_posts(self, config, limit=500):
                return [
                    {
                        "id": "179-title",
                        "caption": "2024 Topps Chrome Caitlin Clark Blue Refractor PSA 9",
                        "permalink": "https://instagram.test/p/title",
                    }
                ]

            def _append_activity(self, action, summary, details):
                pass

        dummy = InstagramDummy()
        result = dummy._instagram_import_existing_posts()

        self.assertEqual(result["imported"], 1)
        self.assertEqual(dummy.state["posts"]["title-key"]["matched_by"], "title_tokens")

    def test_instagram_import_existing_posts_can_match_ocr_cert_text(self) -> None:
        class InstagramDummy:
            _instagram_import_existing_posts = app.CardPipelineApp._instagram_import_existing_posts
            _instagram_find_inventory_match_for_post = app.CardPipelineApp._instagram_find_inventory_match_for_post
            _instagram_match_text_tokens = app.CardPipelineApp._instagram_match_text_tokens
            _instagram_inventory_identity = app.CardPipelineApp._instagram_inventory_identity
            _instagram_record_duplicate_post = app.CardPipelineApp._instagram_record_duplicate_post

            def __init__(self):
                self.state = {"version": 1, "posts": {}}
                self.events = queue.Queue()

            def _load_instagram_inventory_state(self):
                return self.state

            def _save_instagram_inventory_state(self, state):
                self.state = state

            def _instagram_env_config(self):
                return {"user_id": "178", "access_token": "token"}

            def _instagram_inventory_active_records(self):
                return [
                    {
                        "inventory_key": "ocr-key",
                        "status": "Active",
                        "card_title": "2018 Panini Prizm Luka Doncic Silver PSA 10",
                        "cert_number": "1234567890",
                    }
                ]

            def _instagram_existing_media_posts(self, config, limit=500):
                return [
                    {
                        "id": "179-ocr",
                        "caption": "",
                        "media_type": "IMAGE",
                        "media_url": "https://instagram.test/media.jpg",
                        "permalink": "https://instagram.test/p/ocr",
                    }
                ]

            def _instagram_ocr_post_text(self, post, client=None):
                return "PSA Cert 1234567890 Luka Doncic Silver"

            def _append_activity(self, action, summary, details):
                pass

        dummy = InstagramDummy()
        with patch.object(app, "genai", object()), patch.object(app, "genai_types", object()), patch.dict(os.environ, {"GOOGLE_API_KEY": "key"}), patch.object(app, "make_photo_ocr_client", lambda _api_key: object()):
            result = dummy._instagram_import_existing_posts(use_ocr=True)

        self.assertEqual(result["imported"], 1)
        self.assertEqual(result["ocr_attempted"], 1)
        self.assertEqual(result["ocr_matched"], 1)
        self.assertEqual(dummy.state["posts"]["ocr-key"]["matched_by"], "ocr_cert_number")

    def test_year_end_report_data_rolls_up_overall_quarterly_and_people(self) -> None:
        class ReportDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _quarter_label_for_date = app.CardPipelineApp._quarter_label_for_date
            _empty_year_end_bucket = app.CardPipelineApp._empty_year_end_bucket
            _year_end_report_data = app.CardPipelineApp._year_end_report_data
            _build_year_end_report_workbook = app.CardPipelineApp._build_year_end_report_workbook
            _write_year_end_sheet = app.CardPipelineApp._write_year_end_sheet
            _style_year_end_sheet = app.CardPipelineApp._style_year_end_sheet
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _inventory_sport_from_value = app.CardPipelineApp._inventory_sport_from_value
            _expense_related_label = app.CardPipelineApp._expense_related_label
            _is_manual_company_profit_backfill = lambda self, _record: False
            _dedupe_profit_records = lambda self, rows: (rows, 0)
            _enrich_profit_records_with_people = lambda self, rows: rows

            def __init__(self):
                self.profit = [
                    {"assigned_person": "Kevin", "date_added": "2026-02-15", "purchase_price": 100, "sale_price": 150, "card_title": "Card A", "cert_number": "111"},
                    {"assigned_person": "Kevin", "record_type": "expense", "date_added": "2026-03-01", "expense_type": "Shipping", "expense_amount": 25, "related_type": "General", "notes": "Supplies"},
                    {"assigned_person": "Mikey", "date_added": "2026-04-02", "purchase_price": 120, "sale_price": 200, "card_title": "Card B", "cert_number": "222"},
                    {"assigned_person": "Kevin", "date_added": "2025-12-31", "purchase_price": 10, "sale_price": 20, "card_title": "Prior Year"},
                ]
                self.inventory = [
                    {"assigned_person": "Kevin", "status": "Active", "card_title": "Inventory A", "purchase_price": 100, "inventory_value": 300, "cert_number": "333"},
                    {"assigned_person": "Mikey", "status": "Sold", "card_title": "Inventory Sold", "purchase_price": 10, "inventory_value": 50, "cert_number": "444"},
                ]

            def _load_profit_ledger(self):
                return self.profit

            def _load_inventory_ledger(self):
                return self.inventory

        dummy = ReportDummy()
        data = dummy._year_end_report_data(2026)

        self.assertEqual(data["overall"]["sales"], 350)
        self.assertEqual(data["overall"]["purchase"], 220)
        self.assertEqual(data["overall"]["gross_profit"], 130)
        self.assertEqual(data["overall"]["expenses"], 25)
        self.assertEqual(data["overall"]["net_profit"], 105)
        self.assertEqual(data["overall"]["inventory_value"], 300)
        self.assertEqual(data["overall"]["quarters"]["Q1"]["sales"], 150)
        self.assertEqual(data["overall"]["quarters"]["Q1"]["expenses"], 25)
        self.assertEqual(data["overall"]["quarters"]["Q2"]["sales"], 200)
        self.assertEqual(data["people"]["Kevin"]["net_profit"], 25)
        self.assertEqual(data["people"]["Kevin"]["inventory_value"], 300)
        self.assertEqual(data["people"]["Mikey"]["gross_profit"], 80)
        workbook = dummy._build_year_end_report_workbook(data)
        self.assertIn("Expenses by Type", workbook.sheetnames)
        self.assertIn("Per Person", workbook.sheetnames)

    def test_profit_sheet_rows_group_by_person_and_source_sheet(self) -> None:
        class ProfitDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_sheet_rows = app.CardPipelineApp._profit_sheet_rows

        dummy = ProfitDummy()
        rows = [
            {"assigned_person": "Lucas", "source_sheet": "Lot A.xlsx", "company": "Arena Club", "purchase_price": 40, "sale_price": 90, "profit": 50, "date_added": "2026-06-11"},
            {"assigned_person": "Lucas", "source_sheet": "Lot A.xlsx", "company": "Fanatics", "purchase_price": 20, "sale_price": 30, "profit": 10, "date_added": "2026-06-12"},
            {"assigned_person": "Mikey", "source_sheet": "Lot B.xlsx", "company": "Arena Club", "purchase_price": 100, "sale_price": 80, "profit": -20, "date_added": "2026-06-12"},
        ]

        grouped = dummy._profit_sheet_rows(rows)
        lot_a = next(item for item in grouped if item["sheet"] == "Lot A.xlsx")

        self.assertEqual(lot_a["person"], "Lucas")
        self.assertEqual(lot_a["cards"], 2)
        self.assertEqual(lot_a["purchase"], 60)
        self.assertEqual(lot_a["sale"], 120)
        self.assertEqual(lot_a["profit"], 60)
        self.assertEqual(lot_a["companies"], "Arena Club, Fanatics")


class PhotoOcrSpeedTests(unittest.TestCase):
    def test_detect_regions_skips_extra_label_sweeps_when_dense_target_is_met(self) -> None:
        calls: list[str] = []

        def fake_detect(_client, _bytes, _mime, prompt):
            calls.append(prompt)
            if prompt == multi_card_extraction.DETECTION_PROMPT:
                return [
                    {"card_index": index + 1, "position": f"slot {index + 1}", "bbox": [index * 50, 0, index * 50 + 40, 400], "detection_confidence": "high"}
                    for index in range(multi_card_extraction.PHOTO_OCR_REGION_TARGET)
                ]
            return []

        with patch.object(multi_card_extraction, "_detect_regions_for_prompt", side_effect=fake_detect), \
                patch.object(multi_card_extraction, "_detect_best_row_regions", return_value=[]), \
                patch.object(multi_card_extraction, "_add_uncovered_edge_regions", side_effect=lambda regions: regions):
            regions = multi_card_extraction._detect_regions_sync(object(), b"image", "image/jpeg")

        self.assertEqual(len(regions), multi_card_extraction.PHOTO_OCR_REGION_TARGET)
        self.assertIn(multi_card_extraction.DETECTION_PROMPT, calls)
        self.assertNotIn(multi_card_extraction.LABEL_DETECTION_PROMPT, calls)
        self.assertNotIn(multi_card_extraction.LABEL_SWEEP_PROMPT, calls)

    def test_detect_regions_skips_extra_sweeps_for_good_small_batch(self) -> None:
        calls: list[str] = []

        def fake_detect(_client, _bytes, _mime, prompt):
            calls.append(prompt)
            if prompt == multi_card_extraction.DETECTION_PROMPT:
                return [
                    {"card_index": 1, "position": "left", "bbox": [0, 0, 200, 400], "detection_confidence": "high"},
                    {"card_index": 2, "position": "middle", "bbox": [220, 0, 420, 400], "detection_confidence": "high"},
                    {"card_index": 3, "position": "right", "bbox": [440, 0, 640, 400], "detection_confidence": "high"},
                ]
            if prompt == multi_card_extraction.LABEL_DETECTION_PROMPT:
                return [
                    {"card_index": index + 1, "position": f"label {index + 1}", "bbox": [index * 50, 0, index * 50 + 40, 100], "detection_confidence": "high"}
                    for index in range(6)
                ]
            return []

        with patch.object(multi_card_extraction, "_detect_regions_for_prompt", side_effect=fake_detect), \
                patch.object(multi_card_extraction, "_detect_best_row_regions", return_value=[]), \
                patch.object(multi_card_extraction, "_detect_best_prompt_regions", return_value=[]), \
                patch.object(multi_card_extraction, "_add_uncovered_edge_regions", side_effect=lambda regions: regions):
            regions = multi_card_extraction._detect_regions_sync(object(), b"image", "image/jpeg")

        self.assertTrue(regions)
        self.assertEqual(len(regions), 3)
        self.assertIn(multi_card_extraction.DETECTION_PROMPT, calls)
        self.assertNotIn(multi_card_extraction.LABEL_DETECTION_PROMPT, calls)
        self.assertNotIn(multi_card_extraction.LABEL_SWEEP_PROMPT, calls)

    def test_detect_regions_uses_label_sweeps_for_low_confidence_small_batch(self) -> None:
        calls: list[str] = []

        def fake_detect(_client, _bytes, _mime, prompt):
            calls.append(prompt)
            if prompt == multi_card_extraction.DETECTION_PROMPT:
                return [
                    {"card_index": 1, "position": "left", "bbox": [0, 0, 200, 400], "detection_confidence": "low"},
                    {"card_index": 2, "position": "middle", "bbox": [220, 0, 420, 400], "detection_confidence": "high"},
                    {"card_index": 3, "position": "right", "bbox": [440, 0, 640, 400], "detection_confidence": "high"},
                ]
            if prompt == multi_card_extraction.LABEL_DETECTION_PROMPT:
                return [
                    {"card_index": index + 1, "position": f"label {index + 1}", "bbox": [index * 50, 0, index * 50 + 40, 100], "detection_confidence": "high"}
                    for index in range(6)
                ]
            return []

        with patch.object(multi_card_extraction, "_detect_regions_for_prompt", side_effect=fake_detect), \
                patch.object(multi_card_extraction, "_detect_best_row_regions", return_value=[]), \
                patch.object(multi_card_extraction, "_detect_best_prompt_regions", return_value=[]), \
                patch.object(multi_card_extraction, "_add_uncovered_edge_regions", side_effect=lambda regions: regions):
            regions = multi_card_extraction._detect_regions_sync(object(), b"image", "image/jpeg")

        self.assertTrue(regions)
        self.assertIn(multi_card_extraction.LABEL_DETECTION_PROMPT, calls)

    def test_identify_cards_reports_crop_progress_and_preserves_order(self) -> None:
        callbacks: list[str] = []
        regions = [
            {"card_index": 1, "position": "left", "bbox": [0, 0, 200, 400], "detection_confidence": "high"},
            {"card_index": 2, "position": "right", "bbox": [220, 0, 420, 400], "detection_confidence": "medium"},
        ]

        def fake_identify(_client, crop_b64):
            return {
                "is_graded_slab": True,
                "grading_company": "PSA",
                "cert_number": "111" if crop_b64 == "crop-1" else "222",
                "player": "Player",
                "year": "2020",
                "set": "Test",
                "card_number": "",
                "parallel": "",
                "subset": "",
                "grade": "10",
                "category": "baseball",
                "confidence": "high",
                "label_text": "label",
            }

        with patch.object(multi_card_extraction, "_prepare_image", return_value=(b"image", "image/jpeg")), \
                patch.object(multi_card_extraction, "_detect_regions_sync", return_value=regions), \
                patch.object(multi_card_extraction, "_decode_image", return_value=object()), \
                patch.object(multi_card_extraction, "_crop_region_to_base64", side_effect=["crop-1", "crop-2"]), \
                patch.object(multi_card_extraction, "_identify_crop_sync", side_effect=fake_identify):
            cards = multi_card_extraction.identify_cards_sync(object(), "fake-b64", progress_callback=callbacks.append)

        self.assertEqual([card["cert_number"] for card in cards], ["111", "222"])
        self.assertTrue(any("Detected 2 card(s)" in message for message in callbacks))
        self.assertTrue(any("Read 2/2" in message for message in callbacks))

    def test_identify_cards_keeps_detected_slab_when_crop_ocr_fails(self) -> None:
        regions = [
            {"card_index": 1, "position": "left", "bbox": [0, 0, 200, 400], "detection_confidence": "high"},
            {"card_index": 2, "position": "right", "bbox": [220, 0, 420, 400], "detection_confidence": "medium"},
        ]

        def fake_identify(_client, crop_b64):
            if crop_b64 == "crop-2":
                raise RuntimeError("label unreadable")
            return {
                "is_graded_slab": True,
                "grading_company": "PSA",
                "cert_number": "111",
                "player": "Player",
                "year": "2020",
                "set": "Test",
                "card_number": "",
                "parallel": "",
                "subset": "",
                "grade": "10",
                "category": "baseball",
                "confidence": "high",
                "label_text": "label",
            }

        with patch.object(multi_card_extraction, "_prepare_image", return_value=(b"image", "image/jpeg")), \
                patch.object(multi_card_extraction, "_detect_regions_sync", return_value=regions), \
                patch.object(multi_card_extraction, "_decode_image", return_value=object()), \
                patch.object(multi_card_extraction, "_crop_region_to_base64", side_effect=["crop-1", "crop-2"]), \
                patch.object(multi_card_extraction, "_identify_crop_sync", side_effect=fake_identify):
            cards = multi_card_extraction.identify_cards_sync(object(), "fake-b64")

        self.assertEqual(len(cards), 2)
        self.assertEqual(cards[0]["cert_number"], "111")
        self.assertEqual(cards[1]["card_index"], 2)
        self.assertTrue(cards[1]["is_graded_slab"])
        self.assertIn("label unreadable", cards[1]["error"])

    def test_bgs_blank_cert_runs_cert_only_fallback(self) -> None:
        responses = [
            '{"mode":"crop","is_graded_slab":true,"grading_company":"BGS","cert_number":"","player":"","year":"","set":"","card_number":"","parallel":"","subset":"","attributes":"AUTOGRAPH","grade":"10","category":"","confidence":"high","label_text":"BECKETT 10 AUTOGRAPH"}',
            '{"mode":"cert_verify","grading_company":"BGS","cert_number":"0010133787","confidence":"medium","label_text":"CERT 0010133787"}',
        ]

        class Response:
            def __init__(self, text):
                self.text = text

        with patch.object(multi_card_extraction, "_prepare_image", return_value=(b"image", "image/jpeg")), \
                patch.object(multi_card_extraction, "_generate_with_retry", side_effect=[Response(text) for text in responses]):
            card = multi_card_extraction._identify_crop_sync(object(), "fake-b64")

        self.assertEqual(card["grading_company"], "BGS")
        self.assertEqual(card["cert_number"], "0010133787")
        self.assertEqual(card["cert_verified"], "YES")

    def test_photo_table_accepts_detected_slab_without_readable_inventory(self) -> None:
        card = {
            "card_index": 2,
            "position": "right",
            "is_graded_slab": True,
            "detection_confidence": "medium",
            "error": "label unreadable",
        }

        self.assertTrue(app.CardPipelineApp._photo_card_has_inventory(object(), card))
        row = app.CardPipelineApp._photo_card_to_row(object(), Path("dense.jpg"), card)
        self.assertEqual(row["source"], "Photo: dense.jpg")
        self.assertIn("right", row["notes"])
        self.assertIn("OCR review needed", row["notes"])

    def test_mobile_inventory_search_and_duplicate_add_flow(self) -> None:
        class MobileDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _inventory_workbook_row = app.CardPipelineApp._inventory_workbook_row
            _inventory_source_sheet_path = app.CardPipelineApp._inventory_source_sheet_path
            _hydrate_inventory_record_source_values = app.CardPipelineApp._hydrate_inventory_record_source_values
            _enrich_inventory_record_assignment = app.CardPipelineApp._enrich_inventory_record_assignment
            _mobile_inventory_payload_record = app.CardPipelineApp._mobile_inventory_payload_record
            _mobile_inventory_json_record = app.CardPipelineApp._mobile_inventory_json_record
            _mobile_inventory_sport_filters = app.CardPipelineApp._mobile_inventory_sport_filters
            _canonical_person_choice = app.CardPipelineApp._canonical_person_choice
            _next_raw_item_id = app.CardPipelineApp._next_raw_item_id
            _raw_item_id_namespace = lambda self: "TEAM"
            mobile_inventory_search = app.CardPipelineApp.mobile_inventory_search
            mobile_inventory_add = app.CardPipelineApp.mobile_inventory_add
            _append_activity = lambda self, action, summary, details=None: None

            def __init__(self, root: Path) -> None:
                self.events = queue.Queue()
                self.lucas_identity = {"display_name": "test", "machine": "test"}

            def _is_personal_lucas(self):
                return False

            def _known_people(self):
                return sorted(
                    {
                        str(record.get("assigned_person") or "").strip()
                        for record in self._load_inventory_ledger()
                        if str(record.get("assigned_person") or "").strip()
                    },
                    key=str.lower,
                )

        old_root = app.CARD_PIPELINE_DIR
        old_inventory = app.INVENTORY_LEDGER_PATH
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            try:
                app.CARD_PIPELINE_DIR = root
                app.INVENTORY_LEDGER_PATH = root / "inventory_ledger.json"
                dummy = MobileDummy(root)
                dummy._save_inventory_ledger([
                    dummy._normalize_inventory_record(
                        {
                            "assigned_person": "Kevin Hambone",
                            "cert_number": "123456",
                            "grader": "PSA",
                            "card_title": "2024 Test Player Silver PSA 10",
                            "sport": "Basketball",
                            "purchase_price": 42,
                            "inventory_value": 88,
                            "source_sheet": "Existing.xlsx",
                            "source": "Barcode",
                        }
                    ),
                    dummy._normalize_inventory_record(
                        {
                            "assigned_person": "Kevin Hambone",
                            "cert_number": "654321",
                            "grader": "BGS",
                            "card_title": "2023 Baseball Test Prospect BGS 9.5",
                            "sport": "Baseball",
                            "purchase_price": 15,
                            "inventory_value": 35,
                            "source_sheet": "Existing.xlsx",
                            "source": "Barcode",
                        }
                    )
                ])

                search = dummy.mobile_inventory_search({"query": "123456"})
                self.assertTrue(search["ok"])
                self.assertEqual(search["count"], 1)
                self.assertEqual(search["items"][0]["purchase_price_display"], "$42.00")
                self.assertEqual(search["people"], ["Kevin Hambone"])
                self.assertEqual(dummy.mobile_inventory_search({"person": "Kevin"})["count"], 2)
                self.assertEqual(dummy.mobile_inventory_search({"person": "Mike"})["count"], 0)
                self.assertEqual(dummy.mobile_inventory_search({"sport": "baseball"})["items"][0]["cert_number"], "654321")
                multi_category = dummy.mobile_inventory_search({"sport": ["baseball", "basketball"]})
                self.assertEqual(multi_category["count"], 2)
                self.assertEqual({item["cert_number"] for item in multi_category["items"]}, {"123456", "654321"})
                self.assertEqual(dummy.mobile_inventory_search({"query": "Test Player Silver"})["items"][0]["cert_number"], "123456")
                self.assertEqual(dummy.mobile_inventory_search({"query": "Barcode"})["count"], 0)
                self.assertEqual(dummy.mobile_inventory_search({"query": "Kevin"})["count"], 0)
                self.assertEqual(dummy.mobile_inventory_search({"query": "Existing"})["count"], 0)
                limited = dummy.mobile_inventory_search({"limit": 1})
                self.assertEqual(limited["count"], 1)

                duplicate = dummy.mobile_inventory_add({"assigned_person": "Kevin Hambone", "cert_number": "123456", "purchase_price": "50"})
                self.assertFalse(duplicate["ok"])
                self.assertTrue(duplicate["duplicate"])

                update = dummy.mobile_inventory_add({"assigned_person": "Kevin Hambone", "cert_number": "123456", "purchase_price": "50", "update_existing": True})
                self.assertTrue(update["ok"])
                self.assertEqual(update["action"], "updated")
                self.assertEqual(update["record"]["purchase_price"], 50.0)

                rejected_person = dummy.mobile_inventory_add({"assigned_person": "New Person", "cert_number": "777888", "purchase_price": "12.50"})
                self.assertFalse(rejected_person["ok"])
                self.assertIn("People Rules", rejected_person["error"])

                added = dummy.mobile_inventory_add({"assigned_person": "Kevin Hambone", "cert_number": "777888", "grader": "SGC", "card_title": "Mobile Added Card", "purchase_price": "12.50", "source": "Show"})
                self.assertTrue(added["ok"])
                self.assertEqual(added["action"], "added")
                self.assertEqual(added["record"]["source"], "Show")

                raw_added = dummy.mobile_inventory_add({"assigned_person": "Kevin Hambone", "card_title": "2024 Panini Prizm Test Raw Card", "purchase_price": "8"})
                self.assertTrue(raw_added["ok"])
                self.assertEqual(raw_added["record"]["item_type"], "Raw")
                self.assertTrue(raw_added["record"]["item_id"].startswith("RAW-"))
                self.assertEqual(raw_added["record"]["cert_number"], "")
                raw_search = dummy.mobile_inventory_search({"query": raw_added["record"]["item_id"]})
                self.assertEqual(raw_search["count"], 0)
                raw_search = dummy.mobile_inventory_search({"query": "Panini Prizm Test Raw"})
                self.assertEqual(raw_search["count"], 1)
                self.assertEqual(raw_search["items"][0]["item_id"], raw_added["record"]["item_id"])
            finally:
                app.CARD_PIPELINE_DIR = old_root
                app.INVENTORY_LEDGER_PATH = old_inventory

    def test_mobile_expenses_profit_summary_and_payouts(self) -> None:
        class MobileFinanceDummy:
            _money_value = app.CardPipelineApp._money_value
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _person_for_profit_record = app.CardPipelineApp._person_for_profit_record
            _enrich_profit_records_with_people = app.CardPipelineApp._enrich_profit_records_with_people
            _append_profit_records = app.CardPipelineApp._append_profit_records
            _canonical_person_choice = app.CardPipelineApp._canonical_person_choice
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _profit_today = lambda self: datetime(2026, 6, 19).date()
            _profit_added_sort_value = app.CardPipelineApp._profit_added_sort_value
            _profit_period_bounds = app.CardPipelineApp._profit_period_bounds
            _canonical_profit_period = app.CardPipelineApp._canonical_profit_period
            _mobile_profit_rows = app.CardPipelineApp._mobile_profit_rows
            _mobile_profit_chart_series = app.CardPipelineApp._mobile_profit_chart_series
            mobile_profit_summary = app.CardPipelineApp.mobile_profit_summary
            mobile_expense_add = app.CardPipelineApp.mobile_expense_add
            mobile_payouts = app.CardPipelineApp.mobile_payouts

            def __init__(self):
                self.events = queue.Queue()
                self.lucas_identity = {"display_name": "Tester", "machine": "Test"}
                self.home_sheet_markers = {}
                self.home_sheet_paths = {"Incoming": {}, "Received": {}}
                self.home_sheet_summaries = {}

            def _known_people(self):
                return ["Kevin Hambone", "Mike Seller"]

            def _payout_sheet_items(self):
                return [
                    {
                        "name": "Lot A.xlsx",
                        "stage": "Received",
                        "person": "Mike Seller",
                        "paid": False,
                        "row_count": 3,
                        "received_count": 3,
                        "payout_balance": 120.0,
                        "status": "Ready",
                    },
                    {
                        "name": "Lot B.xlsx",
                        "stage": "Received",
                        "person": "Kevin Hambone",
                        "paid": True,
                        "row_count": 2,
                        "received_count": 2,
                        "payout_balance": 0.0,
                        "status": "Paid",
                    },
                ]

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_ledger = app.PROFIT_LEDGER_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            try:
                dummy = MobileFinanceDummy()
                dummy._save_profit_ledger(
                    [
                        {
                            "assigned_person": "Kevin Hambone",
                            "cert_number": "123",
                            "source_sheet": "Lot A.xlsx",
                            "company": "Arena",
                            "purchase_price": 40,
                            "sale_price": 100,
                            "date_added": "2026-06-18",
                        }
                    ]
                )

                added = dummy.mobile_expense_add(
                    {
                        "person": "Kevin Hambone",
                        "date": "2026-06-18",
                        "expense_type": "Travel Meal",
                        "amount": "15",
                        "related_type": "Card",
                        "cert_number": "123",
                        "notes": "Dinner",
                    }
                )
                self.assertTrue(added["ok"])
                summary = dummy.mobile_profit_summary({"person": "Kevin", "period": "YTD", "graph": "Overall Profit"})
                self.assertTrue(summary["ok"])
                self.assertEqual(summary["totals"]["gross_profit"], 60.0)
                self.assertEqual(summary["totals"]["expenses"], 15.0)
                self.assertEqual(summary["totals"]["net_profit"], 45.0)
                self.assertEqual(summary["chart"]["values"][-1], 45.0)

                payouts = dummy.mobile_payouts({"person": "Mike"})
                self.assertTrue(payouts["ok"])
                self.assertEqual(payouts["totals"]["balance"], 120.0)
                self.assertEqual(payouts["summary"][0]["person"], "Mike Seller")
                self.assertEqual(payouts["details"][0]["status"], "Ready")

                dummy._save_profit_ledger(
                    [
                        {
                            "assigned_person": "Kevin Hambone",
                            "company": "Z Older",
                            "card_title": "Older Timestamp",
                            "purchase_price": 10,
                            "sale_price": 15,
                            "date_added": "2026-06-18",
                            "ledger_added_at": "2026-06-18T08:00:00",
                        },
                        {
                            "assigned_person": "Kevin Hambone",
                            "company": "A Newer",
                            "card_title": "Newer Timestamp",
                            "purchase_price": 10,
                            "sale_price": 20,
                            "date_added": "2026-06-18",
                            "ledger_added_at": "2026-06-18T09:00:00",
                        },
                    ]
                )
                recent = dummy.mobile_profit_summary({"person": "Kevin", "period": "YTD"})["recent"]
                self.assertEqual([row["title"] for row in recent[:2]], ["Newer Timestamp", "Older Timestamp"])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.PROFIT_LEDGER_PATH = old_ledger

    def test_mobile_inventory_mark_sold_removes_inventory_and_records_method(self) -> None:
        class MobileSaleDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _mobile_inventory_json_record = app.CardPipelineApp._mobile_inventory_json_record
            _mark_inventory_record_sold = app.CardPipelineApp._mark_inventory_record_sold
            _general_sold_sheet_name = app.CardPipelineApp._general_sold_sheet_name
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _person_for_profit_record = app.CardPipelineApp._person_for_profit_record
            _append_profit_records = app.CardPipelineApp._append_profit_records
            _profit_record_date = app.CardPipelineApp._profit_record_date
            _inventory_sale_profit_record = app.CardPipelineApp._inventory_sale_profit_record
            _mobile_inventory_sale_match = app.CardPipelineApp._mobile_inventory_sale_match
            _mobile_inventory_title_key = app.CardPipelineApp._mobile_inventory_title_key
            mobile_inventory_mark_sold = app.CardPipelineApp.mobile_inventory_mark_sold
            _append_activity = lambda self, action, summary, details=None: None

            def __init__(self):
                self.events = queue.Queue()
                self.lucas_identity = {"display_name": "Tester", "machine": "Test"}
                self.home_sheet_markers = {}

            def _known_people(self):
                return ["Kevin Hambone"]

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_profit = app.PROFIT_LEDGER_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            try:
                dummy = MobileSaleDummy()
                record = dummy._normalize_inventory_record(
                    {
                        "assigned_person": "Kevin Hambone",
                        "cert_number": "999111",
                        "grader": "PSA",
                        "card_title": "Mobile Sold Card",
                        "purchase_price": 40,
                        "inventory_value": 90,
                        "source_sheet": "Lot A.xlsx",
                    }
                )
                dummy._save_inventory_ledger([record])

                result = dummy.mobile_inventory_mark_sold(
                    {
                        "inventory_key": record["inventory_key"],
                        "sale_price": "125",
                        "sale_date": "2026-06-20",
                        "sale_method": "Venmo",
                        "company": "Cash Buyer",
                    }
                )
                self.assertTrue(result["ok"])
                self.assertEqual(dummy._load_inventory_ledger(), [])
                profit = [dummy._normalize_profit_record(item) for item in dummy._load_profit_ledger()]
                self.assertEqual(len(profit), 1)
                self.assertEqual(profit[0]["date_added"], "2026-06-20")
                self.assertEqual(profit[0]["company"], "Cash Buyer")
                self.assertEqual(profit[0]["sale_price"], 125.0)
                self.assertEqual(profit[0]["profit"], 85.0)
                self.assertEqual(profit[0]["sale_method"], "Venmo")
                self.assertIn("Sale method: Venmo", profit[0]["notes"])

                stale_key_record = dummy._normalize_inventory_record(
                    {
                        "assigned_person": "Kevin Hambone",
                        "cert_number": "555222",
                        "grader": "PSA",
                        "card_title": "Cached Mobile Sale Card",
                        "purchase_price": 25,
                        "source_sheet": "Lot B.xlsx",
                    }
                )
                dummy._save_inventory_ledger([stale_key_record])
                stale_result = dummy.mobile_inventory_mark_sold(
                    {
                        "inventory_key": "old-cached-key",
                        "cert_number": "555222",
                        "card_title": "Cached Mobile Sale Card",
                        "sale_price": "60",
                        "sale_date": "2026-06-21",
                    }
                )
                self.assertTrue(stale_result["ok"])
                self.assertEqual(dummy._load_inventory_ledger(), [])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.PROFIT_LEDGER_PATH = old_profit

    def test_mobile_queue_sync_applies_actions_once(self) -> None:
        class MobileQueueDummy:
            _money_value = app.CardPipelineApp._money_value
            _inventory_record_key = app.CardPipelineApp._inventory_record_key
            _normalize_inventory_record = app.CardPipelineApp._normalize_inventory_record
            _load_inventory_ledger = app.CardPipelineApp._load_inventory_ledger
            _save_inventory_ledger = app.CardPipelineApp._save_inventory_ledger
            _mobile_inventory_payload_record = app.CardPipelineApp._mobile_inventory_payload_record
            _mobile_inventory_json_record = app.CardPipelineApp._mobile_inventory_json_record
            _inventory_title_with_grader = app.CardPipelineApp._inventory_title_with_grader
            _next_raw_item_id = app.CardPipelineApp._next_raw_item_id
            _raw_item_id_namespace = lambda self: "TEAM"
            mobile_inventory_add = app.CardPipelineApp.mobile_inventory_add
            _profit_record_key = app.CardPipelineApp._profit_record_key
            _load_profit_ledger = app.CardPipelineApp._load_profit_ledger
            _save_profit_ledger = app.CardPipelineApp._save_profit_ledger
            _normalize_profit_record = app.CardPipelineApp._normalize_profit_record
            _person_for_profit_record = app.CardPipelineApp._person_for_profit_record
            _append_profit_records = app.CardPipelineApp._append_profit_records
            _profit_record_date = app.CardPipelineApp._profit_record_date
            mobile_expense_add = app.CardPipelineApp.mobile_expense_add
            _load_mobile_action_log = app.CardPipelineApp._load_mobile_action_log
            _save_mobile_action_log = app.CardPipelineApp._save_mobile_action_log
            _apply_mobile_queue_action = app.CardPipelineApp._apply_mobile_queue_action
            mobile_queue_sync = app.CardPipelineApp.mobile_queue_sync
            _append_activity = lambda self, action, summary, details=None: None

            def __init__(self):
                self.events = queue.Queue()
                self.lucas_identity = {"display_name": "Tester", "machine": "Test"}
                self.home_sheet_markers = {}

            def _known_people(self):
                return ["Kevin Hambone"]

            def _canonical_person_choice(self, value):
                text = str(value or "").strip()
                return text if text == "Kevin Hambone" else None

            def _enrich_inventory_record_assignment(self, record, force=False):
                return self._normalize_inventory_record(record)

        with TemporaryDirectory() as tmp:
            old_pipeline = app.CARD_PIPELINE_DIR
            old_inventory = app.INVENTORY_LEDGER_PATH
            old_profit = app.PROFIT_LEDGER_PATH
            old_mobile_log = app.MOBILE_ACTION_LOG_PATH
            app.CARD_PIPELINE_DIR = Path(tmp)
            app.INVENTORY_LEDGER_PATH = Path(tmp) / "inventory_ledger.json"
            app.PROFIT_LEDGER_PATH = Path(tmp) / "profit_ledger.json"
            app.MOBILE_ACTION_LOG_PATH = Path(tmp) / "mobile_action_log.json"
            try:
                dummy = MobileQueueDummy()
                payload = {
                    "client_id": "phone-1",
                    "actions": [
                        {
                            "id": "phone-1-add-1",
                            "type": "inventory.add",
                            "payload": {
                                "cert_number": "777888",
                                "grader": "PSA",
                                "card_title": "Queued Mobile Card",
                                "purchase_price": "12.50",
                                "assigned_person": "Kevin Hambone",
                            },
                        },
                        {
                            "id": "phone-1-expense-1",
                            "type": "expense.add",
                            "payload": {
                                "person": "Kevin Hambone",
                                "date": "2026-06-22",
                                "expense_type": "Shipping",
                                "amount": "5",
                            },
                        },
                    ],
                }
                first = dummy.mobile_queue_sync(payload)
                self.assertTrue(first["ok"])
                self.assertEqual(first["applied"], 2)
                self.assertEqual(first["skipped"], 0)
                self.assertEqual(len(dummy._load_inventory_ledger()), 1)
                self.assertEqual(len(dummy._load_profit_ledger()), 1)

                second = dummy.mobile_queue_sync(payload)
                self.assertTrue(second["ok"])
                self.assertEqual(second["applied"], 0)
                self.assertEqual(second["skipped"], 2)
                self.assertEqual(len(dummy._load_inventory_ledger()), 1)
                self.assertEqual(len(dummy._load_profit_ledger()), 1)

                title_result = dummy.mobile_inventory_add(
                    {
                        "cert_number": "123321",
                        "grader": "PSA",
                        "card_title": "2025 DONRUSS JAXSON DART #14 DOWNTOWN! DOWNTOWN 10",
                        "purchase_price": "100",
                        "assigned_person": "Kevin Hambone",
                    }
                )
                self.assertTrue(title_result["ok"])
                self.assertIn("PSA 10", title_result["record"]["card_title"])

                raw_result = dummy.mobile_inventory_add(
                    {
                        "cert_number": "4710408",
                        "card_title": "1996 Skybox Rising Stars Kobe Bryant Blank Rookie",
                        "purchase_price": "425",
                        "assigned_person": "Kevin Hambone",
                    }
                )
                self.assertTrue(raw_result["ok"])
                self.assertEqual(raw_result["record"]["item_type"], "Raw")
                self.assertTrue(str(raw_result["record"]["item_id"]).startswith("RAW-TEAM-"))
                self.assertEqual(raw_result["record"]["cert_number"], "")
                self.assertIn("Mobile entered cert/item: 4710408", raw_result["record"]["notes"])
            finally:
                app.CARD_PIPELINE_DIR = old_pipeline
                app.INVENTORY_LEDGER_PATH = old_inventory
                app.PROFIT_LEDGER_PATH = old_profit
                app.MOBILE_ACTION_LOG_PATH = old_mobile_log

    def test_mobile_card_identify_returns_search_query_from_photo_ocr(self) -> None:
        class MobilePhotoDummy:
            _photo_card_to_row = app.CardPipelineApp._photo_card_to_row
            _photo_card_has_inventory = app.CardPipelineApp._photo_card_has_inventory
            _load_photo_env = lambda self: None
            _mobile_image_parts = app.CardPipelineApp._mobile_image_parts
            _parse_mobile_quick_card_response = app.CardPipelineApp._parse_mobile_quick_card_response
            _mobile_quick_card_to_row = app.CardPipelineApp._mobile_quick_card_to_row
            _mobile_single_card_quick_read = app.CardPipelineApp._mobile_single_card_quick_read

        dummy = MobilePhotoDummy()
        quick_response = json.dumps(
            {
                "grading_company": "PSA",
                "cert_number": "123456789",
                "player": "Test Player",
                "year": "2024",
                "set": "Prizm",
                "grade": "10",
                "confidence": "high",
            }
        )

        class FakeModels:
            def generate_content(self, **_kwargs):
                return types.SimpleNamespace(text=quick_response)

        class FakeClient:
            models = FakeModels()

        fake_genai = types.SimpleNamespace(Client=lambda api_key: FakeClient())
        fake_genai_types = types.SimpleNamespace(
            Part=types.SimpleNamespace(from_bytes=lambda **kwargs: kwargs),
            GenerateContentConfig=lambda **kwargs: kwargs,
            ThinkingConfig=lambda **kwargs: kwargs,
        )
        with patch.object(app, "genai", fake_genai), \
                patch.object(app, "genai_types", fake_genai_types), \
                patch.object(app, "identify_cards_sync") as fallback_ocr, \
                patch.dict(app.os.environ, {"GOOGLE_API_KEY": "test-key"}):
            result = app.CardPipelineApp.mobile_card_identify(dummy, {"image": "data:image/jpeg;base64,ZmFrZQ=="})

        self.assertTrue(result["ok"])
        self.assertEqual(result["mode"], "quick")
        self.assertEqual(result["query"], "123456789")
        self.assertEqual(result["card"]["grader"], "PSA")
        self.assertIn("Test Player", result["card"]["card_title"])
        fallback_ocr.assert_not_called()

    def test_mobile_card_identify_rejects_oversized_photos(self) -> None:
        class MobilePhotoDummy:
            _load_photo_env = lambda self: None
            _mobile_image_parts = app.CardPipelineApp._mobile_image_parts

        oversized = base64.b64encode(b"x" * (8 * 1024 * 1024 + 1)).decode("ascii")
        fake_genai = types.SimpleNamespace(Client=lambda api_key: object())
        with patch.object(app, "genai", fake_genai), \
                patch.object(app, "identify_cards_sync") as fallback_ocr, \
                patch.dict(app.os.environ, {"GOOGLE_API_KEY": "test-key"}):
            result = app.CardPipelineApp.mobile_card_identify(MobilePhotoDummy(), {"image": f"data:image/jpeg;base64,{oversized}"})

        self.assertFalse(result["ok"])
        self.assertIn("too large", result["error"])
        fallback_ocr.assert_not_called()


if __name__ == "__main__":
    unittest.main()
from openpyxl import Workbook, load_workbook
